"""
excel-tools — MCP server for round-trip Excel editing.
"""
import sys
from pathlib import Path

_SERVER_DIRECTORY = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(_SERVER_DIRECTORY))

import copy
import html
import importlib.metadata
import json
import os
import platform
import re
import subprocess
import tempfile
import threading
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timezone

from mcp.server.fastmcp import FastMCP
from mcp.types import TextContent

from cancellable import run_named_operation, run_worker_cli
from process_lifecycle import (
    ManagedProcessCancelled,
    remove_path_with_retries,
    run_cancellable_in_thread,
    run_managed_process,
)
from package_tools import package_edit_verifier_patterns, register_package_tools
from preservation import (
    _MemorySampler,
    cleanup_excel_backups,
    find_latest_excel_backup,
    inspect_workbook_pair,
    inspection_pair_performance,
    package_signature_report,
    start_excel_backup_cleanup,
    verify_xlsx_preservation,
)

from save_transaction import (
    SaveTransactionError,
    commit_staging_file,
    create_staging_path,
    execute_save_stage,
    file_state,
    remove_staging_path,
    require_preservation_success,
    verification_reference,
)

from core import (
    _cell_baseline,
    _close_openpyxl_workbook,
    _expanded_implicit_cell,
    _validate_image_creation_payload,
    build_shape_rich_text_xml,
    diff_xlsx_package,
    inspect_xlsx_package,
    reconstruct_excel,
    serialize_excel,
    uri_to_path,
)

# Session cache lives at module level so it persists across tool calls
_sessions: dict[str, dict] = {}
_SESSION_LOCK = threading.RLock()
_BUSY_SESSIONS: dict[str, dict] = {}
start_excel_backup_cleanup()


class ExcelOperationError(RuntimeError):
    def __init__(self, details: dict):
        self.details = details
        super().__init__(json.dumps(details, ensure_ascii=False, sort_keys=True, default=str))


def _session_busy_error(session_key: str) -> ExcelOperationError:
    busy = dict(_BUSY_SESSIONS.get(session_key) or {})
    return ExcelOperationError({
        "code": "EXCEL_SESSION_BUSY",
        "session_key": session_key,
        "operation": busy.get("operation"),
        "started_at": busy.get("started_at"),
        "message": "The Excel session is busy with a save transaction.",
    })

# Style keys accepted by excel_set_style
_STYLE_KEYS = frozenset({"fill", "fcolor", "bold", "italic", "strike", "underline", "uline", "size", "font", "wrap", "halign", "valign", "numfmt"})

# Blank cell template used when inserting empty cells
_EMPTY_CELL: dict = {
    "v": None, "fill": None, "bold": False, "italic": False,
    "size": None, "font": None, "fcolor": None, "wrap": False,
    "halign": None, "valign": None, "numfmt": "General", "merge": {}, "border": {},
}


def _normalize_underline(value):
    if value is True:
        return "single"
    if value in (False, "", None):
        return None
    return value


def _apply_style(cell: dict, style: dict) -> None:
    for key, value in style.items():
        if key == "underline":
            key = "uline"
        if key in _STYLE_KEYS:
            if key == "uline":
                value = _normalize_underline(value)
            cell[key] = value
            if key == "fill":
                cell.pop("_fill_raw", None)
            if key in {"fcolor", "bold", "italic", "strike", "underline", "uline", "size", "font"}:
                cell.pop("_font_raw", None)


def _drop_raw_fills(sheet: dict) -> None:
    for row in sheet.get("rows", []):
        for cell in row.get("cells", []):
            cell.pop("_fill_raw", None)
            cell.pop("_font_raw", None)
            for side in (cell.get("border") or {}).values():
                if isinstance(side, dict):
                    side.pop("_color_raw", None)


_LEGACY_OR_BINARY_EXTS = {".xls", ".xlsb"}


def _check_supported(path) -> None:
    ext = Path(str(path)).suffix.lower()
    if ext in _LEGACY_OR_BINARY_EXTS:
        raise ValueError(
            f"'{ext}' files are not supported by the edit engine. Use convert_to_markdown "
            "for read-only extraction, or convert to an OOXML workbook (.xlsx/.xlsm/.xltx/.xltm).")


def _resolve_session_key(session_key: str) -> str:
    """Return the canonical key for a session, tolerating path-case differences."""
    with _SESSION_LOCK:
        if session_key in _sessions:
            return session_key
        try:
            alt = str(Path(session_key).resolve())
        except (OSError, ValueError):
            alt = None
        if alt and alt in _sessions:
            return alt
    raise ValueError(f"Session '{session_key}' not found. Call excel_load first.")


def _get_session(session_key: str) -> dict:
    with _SESSION_LOCK:
        resolved = _resolve_session_key(session_key)
        if resolved in _BUSY_SESSIONS:
            raise _session_busy_error(resolved)
        return _sessions[resolved]


def _begin_session_operation(session_key: str, operation: str) -> tuple[str, dict]:
    with _SESSION_LOCK:
        resolved = _resolve_session_key(session_key)
        if resolved in _BUSY_SESSIONS:
            raise _session_busy_error(resolved)
        _BUSY_SESSIONS[resolved] = {
            "operation": operation,
            "started_at": datetime.now(timezone.utc).isoformat(),
        }
        return resolved, _sessions[resolved]


def _end_session_operation(session_key: str) -> None:
    with _SESSION_LOCK:
        _BUSY_SESSIONS.pop(session_key, None)


def _store_session(session_key: str, data: dict) -> None:
    with _SESSION_LOCK:
        if session_key in _BUSY_SESSIONS:
            raise _session_busy_error(session_key)
        _sessions[session_key] = data


def _find_sheet(data: dict, name: str) -> dict:
    sheet = next((s for s in data["sheets"] if s["name"] == name), None)
    if sheet is None:
        available = [s["name"] for s in data["sheets"]]
        raise ValueError(f"Sheet '{name}' not found. Available: {available}")
    return sheet



def _excel_family_ext(path: str | Path) -> str:
    return Path(path).suffix.lower()


def _check_save_extension_compatible(data: dict, dest: str) -> None:
    source = data["source"]
    src_ext = _excel_family_ext(source)
    dst_ext = _excel_family_ext(dest)
    macro_exts = {".xlsm", ".xltm"}
    non_macro_exts = {".xlsx", ".xltx"}
    if src_ext in macro_exts and dst_ext in non_macro_exts:
        raise ValueError(f"Refusing to save macro-enabled {src_ext} workbook as {dst_ext}; choose a macro-enabled extension.")
    if src_ext in non_macro_exts and dst_ext in macro_exts:
        edits = data.get("_package_edits") or {}
        upserted = edits.get("upsert") or {}
        deleted = set(edits.get("delete") or [])
        content_type = (
            (edits.get("content_types") or {}).get("overrides") or {}
        ).get("xl/workbook.xml")
        expected_content_type = {
            ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            ".xltm": "application/vnd.ms-excel.template.macroEnabled.main+xml",
        }[dst_ext]
        explicit_macro_promotion = (
            "xl/vbaProject.bin" in upserted
            and "xl/vbaProject.bin" not in deleted
            and content_type == expected_content_type
        )
        if explicit_macro_promotion:
            return
        raise ValueError(f"Refusing to save non-macro {src_ext} workbook as {dst_ext}; choose a non-macro extension.")

def _excel_range_to_indices(range_ref: str) -> tuple[int, int, int, int]:
    """Convert an Excel A1 range to 0-based inclusive row/column bounds."""
    import openpyxl.utils

    normalized = range_ref.split("!", 1)[-1].replace("$", "")
    if ":" not in normalized:
        normalized = f"{normalized}:{normalized}"
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(normalized)
    return min_row - 1, max_row - 1, min_col - 1, max_col - 1

def _range_from_args(
    range_ref: str | None,
    start_row: int | None,
    end_row: int | None,
    start_col: int | None,
    end_col: int | None,
    max_row: int,
    max_col: int,
) -> tuple[int, int, int, int]:
    if range_ref:
        return _excel_range_to_indices(range_ref)
    r1 = 0 if start_row is None else start_row
    r2 = max_row - 1 if end_row is None else end_row - 1
    c1 = 0 if start_col is None else start_col
    c2 = max_col - 1 if end_col is None else end_col - 1
    if min(r1, r2, c1, c2) < 0 or r1 > r2 or c1 > c2:
        raise ValueError("Invalid range bounds. Use 0-based start and exclusive end indexes.")
    return r1, r2, c1, c2

def _limit_workbook_data(data: dict, max_rows: int | None = None, max_cols: int | None = None) -> dict:
    """Return a copy of serialized workbook data trimmed for Markdown export."""
    if max_rows is None and max_cols is None:
        return data
    limited = copy.deepcopy(data)
    for sheet in limited.get("sheets", []):
        rows = sheet.get("rows", [])
        if max_rows is not None:
            sheet["rows"] = rows[:max_rows]
        if max_cols is not None:
            for row in sheet.get("rows", []):
                row["cells"] = row.get("cells", [])[:max_cols]
    return limited


# ── Structural-shift machinery ────────────────────────────────────────────────
# When rows/columns are inserted or deleted, every piece of coordinate-anchored
# metadata must follow: merges, hyperlinks, comments, validations, conditional
# formatting, tables, auto filter, freeze panes, print titles, drawing anchors.

def _shift_col_dims(sheet: dict, at_col: int, delta: int) -> None:
    """Shift column width/hidden/outline maps after a column insert (+N) / delete (-1)."""
    import openpyxl.utils

    def shift(d: dict) -> dict:
        result = {}
        for letter, value in (d or {}).items():
            idx = openpyxl.utils.column_index_from_string(letter) - 1  # 0-based
            if delta > 0:
                new_idx = idx + delta if idx >= at_col else idx
            else:
                if idx == at_col:
                    continue  # drop the deleted column's entry
                new_idx = idx + delta if idx > at_col else idx
            result[openpyxl.utils.get_column_letter(new_idx + 1)] = value
        return result

    sheet["cw"] = shift(sheet.get("cw") or {})
    if sheet.get("ch"):
        sheet["ch"] = shift(sheet["ch"])
    if sheet.get("co"):
        sheet["co"] = shift(sheet["co"])


def _insert_maps(pos: int, count: int):
    """(strict, clamp) index maps for inserting `count` slots at `pos`."""
    def shift(i):
        return i + count if i >= pos else i
    return shift, shift


def _delete_maps(deleted: set):
    """(strict, clamp) index maps for deleting an index set.

    strict: deleted index → None; clamp: deleted index → position of the
    element that takes its place (for anchors that must survive).
    """
    import bisect
    sd = sorted(deleted)

    def strict(i):
        if i in deleted:
            return None
        return i - bisect.bisect_left(sd, i)

    def clamp(i):
        return i - bisect.bisect_left(sd, i)

    return strict, clamp


_COORD_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?(\d+)$")
_RANGE_RE = re.compile(
    r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)(?::(\$?)([A-Za-z]{1,3})(\$?)(\d+))?$")
_ROWS_RE = re.compile(r"^(\$?)(\d+):(\$?)(\d+)$")
_COLS_RE = re.compile(r"^(\$?)([A-Za-z]{1,3}):(\$?)([A-Za-z]{1,3})$")


def _coord_to_rc(coord: str):
    import openpyxl.utils
    m = _COORD_RE.match(str(coord))
    if not m:
        return None
    return int(m.group(2)) - 1, openpyxl.utils.column_index_from_string(m.group(1).upper()) - 1


def _rc_to_coord(r: int, c: int) -> str:
    import openpyxl.utils
    return f"{openpyxl.utils.get_column_letter(c + 1)}{r + 1}"


def _shift_span(lo: int, hi: int, smap, cmap):
    """Map an inclusive index span; None if nothing survives a deletion."""
    a, b = smap(lo), smap(hi)
    if a is not None and b is not None:
        return a, b
    if hi - lo <= 50000:
        vals = [v for v in (smap(i) for i in range(lo, hi + 1)) if v is not None]
        if not vals:
            return None
        return min(vals), max(vals)
    a, b = cmap(lo), cmap(hi)
    return (a, b) if a <= b else None


def _shift_ref(ref: str, row_maps=None, col_maps=None):
    """Shift one A1-style ref/range ($ preserved). None = entirely removed."""
    import openpyxl.utils as U
    ref = str(ref)
    m = _RANGE_RE.match(ref)
    if m:
        d1, col1, d2, row1, d3, col2, d4, row2 = m.groups()
        c1 = U.column_index_from_string(col1.upper()) - 1
        r1 = int(row1) - 1
        c2 = U.column_index_from_string((col2 or col1).upper()) - 1
        r2 = int(row2 or row1) - 1
        if row_maps:
            span = _shift_span(min(r1, r2), max(r1, r2), *row_maps)
            if span is None:
                return None
            r1, r2 = span
        if col_maps:
            span = _shift_span(min(c1, c2), max(c1, c2), *col_maps)
            if span is None:
                return None
            c1, c2 = span
        first = f"{d1}{U.get_column_letter(c1 + 1)}{d2}{r1 + 1}"
        if m.group(6) is None:
            return first
        second = f"{d3}{U.get_column_letter(c2 + 1)}{d4}{r2 + 1}"
        return f"{first}:{second}"
    m = _ROWS_RE.match(ref)
    if m:
        if not row_maps:
            return ref
        d1, row1, d2, row2 = m.groups()
        span = _shift_span(int(row1) - 1, int(row2) - 1, *row_maps)
        if span is None:
            return None
        return f"{d1}{span[0] + 1}:{d2}{span[1] + 1}"
    m = _COLS_RE.match(ref)
    if m:
        if not col_maps:
            return ref
        d1, col1, d2, col2 = m.groups()
        span = _shift_span(U.column_index_from_string(col1.upper()) - 1,
                           U.column_index_from_string(col2.upper()) - 1, *col_maps)
        if span is None:
            return None
        return f"{d1}{U.get_column_letter(span[0] + 1)}:{d2}{U.get_column_letter(span[1] + 1)}"
    return ref


def _shift_sqref(sqref: str, row_maps=None, col_maps=None):
    parts = []
    for token in str(sqref).split():
        s = _shift_ref(token, row_maps, col_maps)
        if s:
            parts.append(s)
    return " ".join(parts) or None


def _shift_formula_value(value, current_sheet, target_sheet: str, row_maps=None, col_maps=None):
    if not isinstance(value, str) or not value:
        return value
    hash_prefix = "#" if value.startswith("#") else ""
    body = value[1:] if hash_prefix else value
    had_equals = body.startswith("=")
    formula = body if had_equals else "=" + body
    shifted = _shift_formula_str(formula, current_sheet, target_sheet, row_maps, col_maps)
    shifted = shifted if had_equals else shifted[1:]
    return hash_prefix + shifted


def _shift_formula_model(value, current_sheet, target_sheet: str, row_maps=None, col_maps=None):
    if isinstance(value, str):
        return _shift_formula_value(value, current_sheet, target_sheet, row_maps, col_maps)
    if isinstance(value, dict):
        for key in ("text", "attr_text"):
            if isinstance(value.get(key), str):
                value[key] = _shift_formula_value(value[key], current_sheet, target_sheet, row_maps, col_maps)
                break
    return value


def _shift_xml_formula_nodes(xml: str, tags, current_sheet, target_sheet: str,
                             row_maps=None, col_maps=None) -> str:
    for tag in tags:
        pattern = rf"(<(?:[A-Za-z_][\w.-]*:)?{re.escape(tag)}\b[^>]*>)(.*?)(</(?:[A-Za-z_][\w.-]*:)?{re.escape(tag)}>)"

        def repl(match):
            value = html.unescape(match.group(2))
            shifted = _shift_formula_value(value, current_sheet, target_sheet, row_maps, col_maps)
            return match.group(1) + html.escape(shifted, quote=False) + match.group(3)

        xml = re.sub(pattern, repl, xml, flags=re.DOTALL)
    return xml


def _shift_range_value(value, current_sheet, target_sheet: str, row_maps=None, col_maps=None):
    shifted = _shift_formula_value(value, current_sheet, target_sheet, row_maps, col_maps)
    return None if isinstance(shifted, str) and "#REF!" in shifted else shifted


def _split_reference_union(value: str) -> list[str]:
    parts = []
    current = []
    quoted = False
    index = 0
    while index < len(value):
        char = value[index]
        if char == "'":
            current.append(char)
            if quoted and index + 1 < len(value) and value[index + 1] == "'":
                current.append("'")
                index += 2
                continue
            quoted = not quoted
        elif char == "," and not quoted:
            parts.append("".join(current))
            current = []
        else:
            current.append(char)
        index += 1
    parts.append("".join(current))
    return parts


def _shift_reference_union(value, current_sheet, target_sheet: str, row_maps=None, col_maps=None):
    if not isinstance(value, str) or not value:
        return value
    shifted = []
    for part in _split_reference_union(value):
        updated = _shift_range_value(part.strip(), current_sheet, target_sheet, row_maps, col_maps)
        if updated:
            shifted.append(updated)
    return ",".join(shifted) or None


def _shift_sort_state_ranges(sort_state: dict | None, row_maps=None, col_maps=None):
    if not sort_state:
        return sort_state
    if sort_state.get("ref"):
        sort_state["ref"] = _shift_ref(sort_state["ref"], row_maps, col_maps)
        if not sort_state["ref"]:
            return None
    kept = []
    for condition in sort_state.get("conditions", sort_state.get("sortCondition", [])) or []:
        if condition.get("ref"):
            condition["ref"] = _shift_ref(condition["ref"], row_maps, col_maps)
            if not condition["ref"]:
                continue
        kept.append(condition)
    if "conditions" in sort_state:
        sort_state["conditions"] = kept
    elif "sortCondition" in sort_state:
        sort_state["sortCondition"] = kept
    return sort_state


def _shift_auto_filter_ranges(model: dict | None, row_maps=None, col_maps=None):
    if not model:
        return model
    old_ref = model.get("ref")
    old_min_col = None
    if old_ref and col_maps:
        try:
            from openpyxl.utils.cell import range_boundaries
            old_min_col = range_boundaries(str(old_ref))[0] - 1
        except Exception:
            old_min_col = None
    if old_ref:
        model["ref"] = _shift_ref(old_ref, row_maps, col_maps)
        if not model["ref"]:
            return None
    if col_maps and old_min_col is not None and model.get("ref"):
        try:
            from openpyxl.utils.cell import range_boundaries
            new_min_col = range_boundaries(str(model["ref"]))[0] - 1
        except Exception:
            new_min_col = None
        if new_min_col is not None:
            kept_columns = []
            for column in model.get("filter_columns") or []:
                absolute = old_min_col + int(column.get("colId", 0))
                mapped = col_maps[0](absolute)
                if mapped is None:
                    continue
                column["colId"] = mapped - new_min_col
                kept_columns.append(column)
            if "filter_columns" in model:
                model["filter_columns"] = kept_columns
    model["sort_state"] = _shift_sort_state_ranges(model.get("sort_state"), row_maps, col_maps)
    return model


def _shift_anchor_value(value, row_maps=None, col_maps=None):
    rc = _coord_to_rc(str(value)) if value else None
    if rc is None:
        return value
    row, column = rc
    if row_maps:
        row = row_maps[1](row)
    if col_maps:
        column = col_maps[1](column)
    return _rc_to_coord(row, column)


def _shift_coord_map(d, row_maps=None, col_maps=None):
    if not d:
        return d
    out = {}
    for coord, value in d.items():
        rc = _coord_to_rc(coord)
        if rc is None:
            out[coord] = value
            continue
        r, c = rc
        if row_maps:
            r = row_maps[0](r)
        if col_maps:
            c = col_maps[0](c) if c is not None else None
        if r is None or c is None:
            continue
        out[_rc_to_coord(r, c)] = value
    return out or None


def _shift_dv_xml(xml: str, row_maps=None, col_maps=None,
                  current_sheet=None, target_sheet=None):
    def repl(match):
        block = match.group(0)
        sqref = re.search(r'sqref="([^"]+)"', block)
        if sqref:
            shifted_sqref = _shift_sqref(sqref.group(1), row_maps, col_maps)
            if shifted_sqref is None:
                return ""
            block = block.replace(sqref.group(0), f'sqref="{shifted_sqref}"', 1)
        if target_sheet is not None:
            block = _shift_xml_formula_nodes(
                block, ("formula1", "formula2"), current_sheet, target_sheet, row_maps, col_maps,
            )
        return block

    xml = re.sub(r"<dataValidation\b[^>]*/>|<dataValidation\b[^>]*>.*?</dataValidation>",
                 repl, xml, flags=re.DOTALL)
    count = len(re.findall(r"<dataValidation\b", xml)) - len(re.findall(r"<dataValidations\b", xml))
    if count <= 0:
        return None
    return re.sub(r'(<dataValidations\b[^>]*?count=")\d+(")', rf"\g<1>{count}\g<2>", xml, count=1)


def _shift_cf_blocks(blocks: list, row_maps=None, col_maps=None,
                     current_sheet=None, target_sheet=None):
    out = []
    for block in blocks:
        sqref = re.search(r'sqref="([^"]+)"', block)
        if sqref:
            shifted_sqref = _shift_sqref(sqref.group(1), row_maps, col_maps)
            if shifted_sqref is None:
                continue
            block = block.replace(sqref.group(0), f'sqref="{shifted_sqref}"', 1)
        if target_sheet is not None:
            block = _shift_xml_formula_nodes(
                block, ("formula",), current_sheet, target_sheet, row_maps, col_maps,
            )
        out.append(block)
    return out or None


def _shift_drawing_anchors(xml: str, row_maps=None, col_maps=None) -> str:
    if row_maps:
        clamp = row_maps[1]
        xml = re.sub(r"(<(?:\w+:)?row>)(\d+)(</(?:\w+:)?row>)",
                     lambda m: f"{m.group(1)}{clamp(int(m.group(2)))}{m.group(3)}", xml)
    if col_maps:
        clamp = col_maps[1]
        xml = re.sub(r"(<(?:\w+:)?col>)(\d+)(</(?:\w+:)?col>)",
                     lambda m: f"{m.group(1)}{clamp(int(m.group(2)))}{m.group(3)}", xml)
    return xml


def _shift_sheet_meta(sheet: dict, row_maps=None, col_maps=None) -> None:
    """Shift all coordinate-anchored sheet metadata after a structural change."""
    sheet_name = sheet["name"]
    if sheet.get("hyperlinks"):
        sheet["hyperlinks"] = _shift_coord_map(sheet["hyperlinks"], row_maps, col_maps)
    if sheet.get("comments"):
        sheet["comments"] = _shift_coord_map(sheet["comments"], row_maps, col_maps)

    kept = []
    for validation in sheet.get("validations") or []:
        shifted_sqref = _shift_sqref(validation.get("sqref", ""), row_maps, col_maps)
        if shifted_sqref is None:
            continue
        validation["sqref"] = shifted_sqref
        kept.append(validation)
    if sheet.get("validations") is not None:
        sheet["validations"] = kept

    if sheet.get("data_validations_xml"):
        shifted_xml = _shift_dv_xml(
            sheet["data_validations_xml"], row_maps, col_maps,
            current_sheet=sheet_name, target_sheet=sheet_name,
        )
        if shifted_xml is None:
            sheet.pop("data_validations_xml", None)
        else:
            sheet["data_validations_xml"] = shifted_xml

    if sheet.get("cf_xml"):
        shifted_blocks = _shift_cf_blocks(
            sheet["cf_xml"], row_maps, col_maps,
            current_sheet=sheet_name, target_sheet=sheet_name,
        )
        if shifted_blocks is None:
            sheet.pop("cf_xml", None)
        else:
            sheet["cf_xml"] = shifted_blocks

    if sheet.get("auto_filter"):
        sheet["auto_filter"] = _shift_ref(sheet["auto_filter"], row_maps, col_maps)

    if sheet.get("tables"):
        kept_tables = []
        for table in sheet["tables"]:
            shifted_ref = _shift_ref(table.get("ref", ""), row_maps, col_maps)
            if shifted_ref is None:
                continue
            table["ref"] = shifted_ref
            table["auto_filter"] = _shift_auto_filter_ranges(table.get("auto_filter"), row_maps, col_maps)
            table["sort_state"] = _shift_sort_state_ranges(table.get("sort_state"), row_maps, col_maps)
            kept_tables.append(table)
        sheet["tables"] = kept_tables or None

    if sheet.get("freeze"):
        rc = _coord_to_rc(str(sheet["freeze"]))
        if rc:
            row, column = rc
            if row_maps:
                row = row_maps[1](row)
            if col_maps:
                column = col_maps[1](column)
            sheet["freeze"] = _rc_to_coord(row, column) if (row > 0 or column > 0) else None

    print_titles = sheet.get("print_titles")
    if print_titles:
        if print_titles.get("rows") and row_maps:
            print_titles["rows"] = _shift_range_value(
                print_titles["rows"], sheet_name, sheet_name, row_maps, None,
            )
        if print_titles.get("cols") and col_maps:
            print_titles["cols"] = _shift_range_value(
                print_titles["cols"], sheet_name, sheet_name, None, col_maps,
            )
        if not print_titles.get("rows") and not print_titles.get("cols"):
            sheet["print_titles"] = None

    if sheet.get("print_area"):
        sheet["print_area"] = _shift_reference_union(
            sheet["print_area"], sheet_name, sheet_name, row_maps, col_maps,
        )

    drawing_data = sheet.get("drawing_data")
    if drawing_data and drawing_data.get("drawing_xml"):
        drawing_data["drawing_xml"] = _shift_drawing_anchors(
            drawing_data["drawing_xml"], row_maps, col_maps,
        )
    for creation in sheet.get("drawing_creations") or []:
        if creation.get("anchor"):
            creation["anchor"] = _shift_anchor_value(creation["anchor"], row_maps, col_maps)

    _shift_advanced_sheet_meta(sheet, row_maps, col_maps)


def _capture_merge_regions(rows: list) -> list:
    """Merge regions [(r1,c1,r2,c2)] from actual grid positions + spans."""
    regions = []
    for r, row in enumerate(rows):
        for c, cd in enumerate(row.get("cells") or []):
            mi = cd.get("merge")
            if isinstance(mi, dict) and (mi.get("rowspan", 1) > 1 or mi.get("colspan", 1) > 1):
                regions.append([r, c,
                                r + mi.get("rowspan", 1) - 1,
                                c + mi.get("colspan", 1) - 1])
    return regions


def _restamp_merges(rows: list, regions: list, implicit_defaults: dict | None = None) -> None:
    """Clear all merge markers and re-stamp them from a region list."""
    for row in rows:
        for cd in row.get("cells") or []:
            cd["merge"] = {}
    n_rows = len(rows)
    for r1, c1, r2, c2 in regions:
        if not (0 <= r1 < n_rows):
            continue
        r2 = min(r2, n_rows - 1)
        if r2 == r1 and c2 == c1:
            continue  # collapsed to a single cell — no merge left
        cells = rows[r1].get("cells") or []
        if c1 >= len(cells):
            continue
        origin = _promote_implicit_cell(cells[c1], implicit_defaults)
        origin["merge"] = {"r1": r1, "c1": c1, "r2": r2, "c2": c2,
                           "rowspan": r2 - r1 + 1, "colspan": c2 - c1 + 1}
        for r in range(r1, r2 + 1):
            row_cells = rows[r].get("cells") or []
            for c in range(c1, min(c2 + 1, len(row_cells))):
                if r == r1 and c == c1:
                    continue
                target = _promote_implicit_cell(row_cells[c], implicit_defaults)
                target["merge"] = "slave"


def _regions_after_insert(regions: list, pos: int, count: int, axis: str) -> list:
    out = []
    for r1, c1, r2, c2 in regions:
        lo, hi = (r1, r2) if axis == "row" else (c1, c2)
        if pos <= lo:
            lo, hi = lo + count, hi + count
        elif pos <= hi:
            hi += count  # inserted inside the merge → region extends
        out.append([lo, c1, hi, c2] if axis == "row" else [r1, lo, r2, hi])
    return out


def _regions_after_delete(regions: list, maps, axis: str) -> list:
    smap, cmap = maps
    out = []
    for r1, c1, r2, c2 in regions:
        lo, hi = (r1, r2) if axis == "row" else (c1, c2)
        span = _shift_span(lo, hi, smap, cmap)
        if span is None:
            continue
        lo, hi = span
        out.append([lo, c1, hi, c2] if axis == "row" else [r1, lo, r2, hi])
    return out


def _block_merge_regions(block_rows: list, row_offset: int) -> list:
    """Self-contained merges inside an inserted block, clamped to the block."""
    out = []
    n = len(block_rows)
    for r1, c1, r2, c2 in _capture_merge_regions(block_rows):
        out.append([r1 + row_offset, c1, min(r2, n - 1) + row_offset, c2])
    return out


def _apply_row_insert(data: dict, sheet: dict, pos: int, new_rows: list) -> None:
    regions = _capture_merge_regions(sheet["rows"])
    block_regions = _block_merge_regions(new_rows, pos)
    sheet["rows"] = sheet["rows"][:pos] + new_rows + sheet["rows"][pos:]
    regions = _regions_after_insert(regions, pos, len(new_rows), "row") + block_regions
    _restamp_merges(sheet["rows"], regions, sheet.get("_implicit_cell_defaults"))
    maps = _insert_maps(pos, len(new_rows))
    _shift_sheet_meta(sheet, row_maps=maps)
    _shift_formulas_workbook(data, sheet["name"], row_maps=maps)
    _mark_dirty(data, "structure", f"sheets/{sheet['name']}/rows")


def _apply_row_delete(data: dict, sheet: dict, to_delete: set) -> None:
    regions = _capture_merge_regions(sheet["rows"])
    maps = _delete_maps(set(to_delete))
    sheet["rows"] = [row for i, row in enumerate(sheet["rows"]) if i not in to_delete]
    _restamp_merges(
        sheet["rows"],
        _regions_after_delete(regions, maps, "row"),
        sheet.get("_implicit_cell_defaults"),
    )
    _shift_sheet_meta(sheet, row_maps=maps)
    _shift_formulas_workbook(data, sheet["name"], row_maps=maps)
    _mark_dirty(data, "structure", f"sheets/{sheet['name']}/rows")


def _finish_col_insert(data: dict, sheet: dict, regions: list, pos: int, count: int = 1) -> None:
    _restamp_merges(
        sheet["rows"],
        _regions_after_insert(regions, pos, count, "col"),
        sheet.get("_implicit_cell_defaults"),
    )
    _shift_col_dims(sheet, pos, +count)
    maps = _insert_maps(pos, count)
    _shift_sheet_meta(sheet, col_maps=maps)
    _shift_formulas_workbook(data, sheet["name"], col_maps=maps)
    _mark_dirty(data, "structure", f"sheets/{sheet['name']}/columns")


def _finish_col_delete(data: dict, sheet: dict, regions: list, col: int) -> None:
    maps = _delete_maps({col})
    _restamp_merges(
        sheet["rows"],
        _regions_after_delete(regions, maps, "col"),
        sheet.get("_implicit_cell_defaults"),
    )
    _shift_col_dims(sheet, col, -1)
    _shift_sheet_meta(sheet, col_maps=maps)
    _shift_formulas_workbook(data, sheet["name"], col_maps=maps)
    _mark_dirty(data, "structure", f"sheets/{sheet['name']}/columns")


def _strip_private(obj):
    """Drop internal keys (_fill_raw, _font_raw, …) from tool output JSON."""
    if isinstance(obj, dict):
        return {k: _strip_private(v) for k, v in obj.items()
                if not (isinstance(k, str) and k.startswith("_"))}
    if isinstance(obj, list):
        return [_strip_private(x) for x in obj]
    return obj


_MISSING = object()
_COLOR_TYPES = {"rgb", "theme", "indexed", "auto"}
_RICH_FONT_KEYS = {
    "name", "size", "bold", "italic", "underline", "strike", "color",
    "vertAlign", "charset", "family", "scheme", "outline", "shadow",
    "condense", "extend",
}


def _mark_dirty(data: dict, feature: str, path: str | None = None) -> None:
    """Record feature-level dirty state without persisting the workbook."""
    features = data.setdefault("_dirty_features", [])
    if feature not in features:
        features.append(feature)
    if path:
        paths = data.setdefault("_dirty_paths", [])
        if path not in paths:
            paths.append(path)


def _mutation_result(data: dict, feature: str, before, after, path: str | None = None, **extra) -> str:
    _mark_dirty(data, feature, path)
    payload = {
        "feature": feature,
        "path": path,
        "before": _strip_private(before),
        "after": _strip_private(after),
        "dirty_features": list(data.get("_dirty_features") or []),
        "dirty_paths": list(data.get("_dirty_paths") or []),
    }
    payload.update(extra)
    return json.dumps(payload, default=str, ensure_ascii=False)


def _new_empty_cell(*, present: bool = False) -> dict:
    cell = copy.deepcopy(_EMPTY_CELL)
    cell["present"] = bool(present)
    return cell


def _promote_implicit_cell(
    cell: dict,
    implicit_defaults: dict | None = None,
    *,
    present: bool = False,
) -> dict:
    if not cell.get("_implicit"):
        return cell
    expanded = _expanded_implicit_cell(cell, implicit_defaults)
    cell.clear()
    cell.update(expanded)
    if present:
        cell["present"] = True
    return cell


def _cell_model_for_read(cell: dict, implicit_defaults: dict | None = None) -> dict:
    if not cell.get("_implicit"):
        return cell
    return _expanded_implicit_cell(cell, implicit_defaults)


def _ensure_cell(
    sheet: dict,
    row_index: int,
    col_index: int,
    *,
    capture_baseline: bool = False,
) -> dict:
    if row_index < 0 or col_index < 0:
        raise ValueError("Cell coordinates must be non-negative.")
    rows = sheet.setdefault("rows", [])
    while row_index >= len(rows):
        rows.append({"h": None, "hidden": False, "outline": 0, "cells": []})
    row = rows[row_index]
    cells = row.setdefault("cells", [])
    while col_index >= len(cells):
        cells.append(_new_empty_cell())
    cell = cells[col_index]
    if capture_baseline:
        _promote_implicit_cell(cell, sheet.get("_implicit_cell_defaults"))
        _cell_baseline(cell)
    return cell


def _cell_from_a1(data: dict, sheet_name: str, cell: str, *, create: bool = False):
    sheet = _find_sheet(data, sheet_name)
    r1, r2, c1, c2 = _excel_range_to_indices(cell)
    if r1 != r2 or c1 != c2:
        raise ValueError(f"Expected a single A1 cell reference, got {cell!r}.")
    if create:
        return sheet, r1, c1, _ensure_cell(sheet, r1, c1, capture_baseline=True)
    rows = sheet.get("rows", [])
    if r1 >= len(rows) or c1 >= len(rows[r1].get("cells", [])):
        return sheet, r1, c1, None
    return sheet, r1, c1, rows[r1]["cells"][c1]


def _cell_coord(row_index: int, col_index: int) -> str:
    return _rc_to_coord(row_index, col_index)


def _normalize_color_object(color):
    if color is None:
        return None
    if isinstance(color, str):
        return {"type": "rgb", "rgb": color}
    if not isinstance(color, dict):
        raise ValueError("Color must be an ARGB string, color object, or null.")
    result = copy.deepcopy(color)
    color_type = result.get("type")
    if color_type is None:
        for candidate in ("rgb", "theme", "indexed", "auto"):
            if candidate in result:
                color_type = candidate
                result["type"] = candidate
                break
    if color_type not in _COLOR_TYPES:
        raise ValueError("Color type must be rgb, theme, indexed, or auto.")
    required = "auto" if color_type == "auto" else color_type
    if required not in result:
        result[required] = True if color_type == "auto" else None
    return result


def _resolved_rgb(color) -> str | None:
    color = _normalize_color_object(color)
    return color.get("rgb") if color and color.get("type") == "rgb" else None


def _whole_cell_font(cell: dict) -> dict:
    raw = copy.deepcopy(cell.get("_font_raw") or {})
    mapping = {
        "name": cell.get("font"),
        "size": cell.get("size"),
        "bold": cell.get("bold"),
        "italic": cell.get("italic"),
        "underline": cell.get("uline"),
        "strike": cell.get("strike"),
        "vertAlign": cell.get("vAlign"),
    }
    for key, value in mapping.items():
        if key not in raw and value is not None:
            raw[key] = value
    if "color" not in raw and cell.get("fcolor"):
        raw["color"] = {"type": "rgb", "rgb": cell["fcolor"]}
    return raw


def _normalize_rich_font(style: dict | None) -> dict:
    style = copy.deepcopy(style or {})
    if "font" in style and isinstance(style["font"], dict):
        nested = style.pop("font")
        nested.update(style)
        style = nested
    result = {key: value for key, value in style.items() if key in _RICH_FONT_KEYS}
    if "color" in result:
        result["color"] = _normalize_color_object(result["color"])
    if "underline" in result:
        result["underline"] = _normalize_underline(result["underline"])
    return result


def _normalize_rich_run(run: dict) -> dict:
    text = str(run.get("text", ""))
    font = _normalize_rich_font(run.get("font") or run.get("style") or {
        key: run[key] for key in _RICH_FONT_KEYS if key in run
    })
    xml_space = run.get("xml_space") or run.get("xml:space")
    if xml_space is None and (text[:1].isspace() or text[-1:].isspace()):
        xml_space = "preserve"
    return {"text": text, "font": font, "xml_space": xml_space}


def _reindex_rich_runs(runs: list[dict]) -> tuple[list[dict], str]:
    normalized = []
    offset = 0
    for item in runs:
        run = _normalize_rich_run(item)
        if not run["text"]:
            continue
        run["start"] = offset
        offset += len(run["text"])
        run["end"] = offset
        normalized.append(run)
    return normalized, "".join(run["text"] for run in normalized)


def _rich_text_model(cell: dict | None) -> dict:
    cell = cell or _new_empty_cell()
    existing = copy.deepcopy(cell.get("rich_text") or {})
    raw_runs = existing.get("runs")
    is_rich = bool(raw_runs is not None or cell.get("rich_text") is not None)
    if raw_runs is None:
        value = cell.get("value", cell.get("v"))
        text = "" if value is None or _is_formula_value(cell) else str(value)
        raw_runs = [{"text": text, "font": _whole_cell_font(cell)}] if text else []
    runs, text = _reindex_rich_runs(raw_runs)
    return {
        "is_rich_text": is_rich,
        "text": text,
        "plain_text": text,
        "runs": runs,
        "phonetic_runs": copy.deepcopy(existing.get("phonetic_runs") or existing.get("rPh") or []),
        "phonetic_properties": copy.deepcopy(existing.get("phonetic_properties") or existing.get("phoneticPr")),
    }


def _chars_from_rich(model: dict) -> list[tuple[str, dict]]:
    chars = []
    for run in model.get("runs") or []:
        font = copy.deepcopy(run.get("font") or {})
        chars.extend((char, copy.deepcopy(font)) for char in run.get("text", ""))
    return chars


def _runs_from_chars(chars: list[tuple[str, dict]]) -> list[dict]:
    runs = []
    for char, font in chars:
        font = _normalize_rich_font(font)
        if runs and runs[-1]["font"] == font:
            runs[-1]["text"] += char
        else:
            runs.append({"text": char, "font": font})
    return _reindex_rich_runs(runs)[0]


def _set_rich_text_cell(cell: dict, model: dict) -> None:
    runs, text = _reindex_rich_runs(model.get("runs") or [])
    cell["rich_text"] = {
        "text": text,
        "plain_text": text,
        "runs": runs,
        "phonetic_runs": copy.deepcopy(model.get("phonetic_runs") or []),
        "phonetic_properties": copy.deepcopy(model.get("phonetic_properties")),
    }
    cell["v"] = text
    cell["value"] = text
    cell["dt"] = "s"
    cell["data_type"] = "inlineStr"
    cell["present"] = True
    cell.pop("formula", None)
    cell.pop("cached_value", None)
    cell.pop("cached_value_state", None)


def _formula_model(cell: dict) -> dict | None:
    formula = cell.get("formula")
    if isinstance(formula, str):
        formula = {"text": formula}
    elif isinstance(formula, dict):
        formula = copy.deepcopy(formula)
    elif _is_formula_value(cell):
        formula = {"text": cell.get("v"), "type": "normal", "attributes": {}}
    else:
        return None
    formula.setdefault("type", "normal")
    formula.setdefault("attributes", copy.deepcopy(cell.get("formula_attributes") or {}))
    # A formula reloaded from disk carries core.py's raw XML-parsed key name
    # "cache_state" (see _extract_ooxml_semantics), not "cached_value_state"
    # (the public-tool/_set_formula_cell convention) -- check both, else a
    # freshly reloaded cell always reported "missing" here even when a real
    # cached value (or an explicit empty one) was actually present.
    state = formula.get("cached_value_state")
    if state is None:
        state = formula.get("cache_state", cell.get("cached_value_state", "missing"))
    formula["cached_value_state"] = state
    if "cached_value" not in formula and "cached_value" in cell:
        formula["cached_value"] = copy.deepcopy(cell.get("cached_value"))
    return formula


def _set_formula_cell(
    cell: dict,
    formula,
    *,
    formula_type: str = "normal",
    attributes: dict | None = None,
    cached_value=_MISSING,
    cache_policy: str = "clear",
) -> None:
    if isinstance(formula, dict):
        payload = copy.deepcopy(formula)
        text = payload.get("text") or payload.get("formula")
        formula_type = payload.get("type", formula_type)
        attributes = payload.get("attributes", payload.get("formula_attributes", attributes))
        if "cached_value" in payload:
            cached_value = payload["cached_value"]
        cache_policy = payload.get("cache_policy", cache_policy)
        explicit_state = payload.get("cached_value_state")
    else:
        text = formula
        explicit_state = None
    if not isinstance(text, str) or not text:
        raise ValueError("formula text is required.")
    if not text.startswith("="):
        text = "=" + text
    model = {"text": text, "type": formula_type or "normal", "attributes": copy.deepcopy(attributes or {})}
    if cached_value is not _MISSING:
        model["cached_value"] = copy.deepcopy(cached_value)
        model["cached_value_state"] = explicit_state or ("empty" if cached_value is None else "value")
        cell["cached_value"] = copy.deepcopy(cached_value)
        cell["cached_value_state"] = model["cached_value_state"]
    elif cache_policy == "preserve" and "cached_value" in cell:
        model["cached_value"] = copy.deepcopy(cell.get("cached_value"))
        model["cached_value_state"] = cell.get("cached_value_state", "value")
    else:
        model["cached_value_state"] = explicit_state or "missing"
        cell.pop("cached_value", None)
        cell["cached_value_state"] = model["cached_value_state"]
    cell["formula"] = model
    cell["formula_attributes"] = copy.deepcopy(model["attributes"])
    cell["v"] = text
    cell["value"] = text
    cell["dt"] = "f"
    cell["data_type"] = "f"
    cell["present"] = True
    cell.pop("rich_text", None)


def _apply_typed_cell_payload(cell: dict, payload: dict) -> None:
    clear = payload.get("clear", False)
    if clear:
        policy = "all" if clear is True else str(clear)
        merge = copy.deepcopy(cell.get("merge", {}))
        if policy in {"all", "cell"}:
            cell.clear()
            cell.update(_new_empty_cell(present=True))
            cell["merge"] = merge
        elif policy in {"content", "value"}:
            for key in (
                "v", "value", "dt", "data_type", "formula", "formula_attributes",
                "cached_value", "cached_value_state", "rich_text", "qp",
            ):
                cell.pop(key, None)
            cell["v"] = None
            cell["present"] = True
        else:
            raise ValueError("clear must be false, true, 'content', 'value', or 'all'.")

    if "formula" in payload and payload["formula"] is not None:
        _set_formula_cell(
            cell,
            payload["formula"],
            formula_type=payload.get("formula_type", "normal"),
            attributes=payload.get("formula_attributes"),
            cached_value=payload["cached_value"] if "cached_value" in payload else _MISSING,
            cache_policy=payload.get("cache_policy", "clear"),
        )
        return

    if "rich_text" in payload and payload["rich_text"] is not None:
        rich = payload["rich_text"]
        if isinstance(rich, list):
            rich = {"runs": rich}
        _set_rich_text_cell(cell, rich)
        return

    if "value" in payload:
        value = payload["value"]
        if cell.get("rich_text"):
            current = _rich_text_model(cell)["text"]
            policy = payload.get("rich_text_policy")
            if policy == "preserve_runs_if_text_equal" and str(value or "") == current:
                cell["v"] = value
                cell["value"] = value
                cell["present"] = True
                return
            if policy != "replace_all":
                raise ValueError(
                    "Editing a rich-text cell requires rich_text_policy='replace_all' "
                    "or 'preserve_runs_if_text_equal'."
                )
            cell.pop("rich_text", None)
        cell["v"] = copy.deepcopy(value)
        cell["value"] = copy.deepcopy(value)
        data_type = payload.get("data_type")
        if data_type is None and isinstance(value, str):
            data_type = "s"
        if data_type is not None:
            cell["dt"] = data_type
            cell["data_type"] = data_type
        else:
            cell.pop("dt", None)
            cell.pop("data_type", None)
        cell.pop("formula", None)
        cell.pop("formula_attributes", None)
        cell.pop("cached_value", None)
        cell.pop("cached_value_state", None)
        cell["present"] = True

    if "data_type" in payload and "value" not in payload:
        value = payload["data_type"]
        if value is None:
            cell.pop("dt", None)
            cell.pop("data_type", None)
        else:
            cell["dt"] = value
            cell["data_type"] = value
    if "cached_value" in payload and "formula" not in payload:
        cell["cached_value"] = copy.deepcopy(payload["cached_value"])
        cell["cached_value_state"] = "empty" if payload["cached_value"] is None else "value"
    if "present" in payload:
        cell["present"] = bool(payload["present"])


def _cell_has_semantics(cell: dict, sheet: dict | None = None, coord: str | None = None) -> bool:
    if cell.get("present") is not None:
        return bool(cell.get("present"))
    if cell.get("_present") is not None:
        return bool(cell.get("_present"))
    if cell.get("v") is not None or cell.get("rich_text") or cell.get("formula"):
        return True
    if any(cell.get(key) not in (None, False, {}, "General") for key in (
        "fill", "bold", "italic", "strike", "size", "font", "fcolor", "uline",
        "wrap", "halign", "valign", "rot", "indent", "shrink", "border", "xf",
        "named_style", "qp",
    )):
        return True
    if sheet is not None and coord:
        if coord in (sheet.get("hyperlinks") or {}) or coord in (sheet.get("comments") or {}):
            return True
        for validation in sheet.get("validations") or []:
            if coord in str(validation.get("sqref", "")).split():
                return True
    return False


def _cell_public_view(
    cell: dict | None,
    *,
    sheet: dict | None = None,
    coord: str | None = None,
    include_rich_text: bool = False,
    include_formula_cache: bool = False,
    include_semantics: bool = False,
):
    if cell is None:
        return {"present": False, "coordinate": coord} if include_semantics else None
    cell = _cell_model_for_read(cell, sheet.get("_implicit_cell_defaults") if sheet else None)
    result = _strip_private(cell)
    # The raw cell dict carries internal "rich_text"/"cached_value"/
    # "cached_value_state" bookkeeping keys. Strip them unconditionally so the
    # include_rich_text/include_formula_cache flags actually gate response
    # size/shape instead of being bypassed by this leak; the curated
    # "rich_text" and "formula.cached_value*" fields below are the sole public
    # source of truth for this data.
    result.pop("rich_text", None)
    result.pop("cached_value", None)
    result.pop("cached_value_state", None)
    result["value"] = cell.get("value", cell.get("v"))
    result["data_type"] = cell.get("data_type", cell.get("dt"))
    if coord:
        result["coordinate"] = coord
    formula = _formula_model(cell)
    if formula:
        if not include_formula_cache:
            formula.pop("cached_value", None)
            formula.pop("cached_value_state", None)
        result["formula"] = formula
    if include_rich_text:
        result["rich_text"] = _rich_text_model(cell)
    if include_semantics:
        font = _whole_cell_font(cell)
        fill_raw = copy.deepcopy(cell.get("_fill_raw") or {})
        fill = {
            "type": "gradient" if fill_raw.get("is_gradient") else "pattern",
            "pattern_type": fill_raw.get("patternType"),
            "foreground": copy.deepcopy(fill_raw.get("fgColor")) or (
                {"type": "rgb", "rgb": cell.get("fill")} if cell.get("fill") else None
            ),
            "background": copy.deepcopy(fill_raw.get("bgColor")),
        }
        border = copy.deepcopy(cell.get("border") or {})
        for side in border.values():
            if isinstance(side, dict) and side.get("_color_raw"):
                side["color_object"] = copy.deepcopy(side["_color_raw"])
        result["present"] = _cell_has_semantics(cell, sheet, coord)
        result["semantics"] = {
            "font": font,
            "fill": fill,
            "alignment": copy.deepcopy(cell.get("alignment") or {
                "horizontal": cell.get("halign"),
                "vertical": cell.get("valign"),
                "wrapText": cell.get("wrap"),
                "textRotation": cell.get("rot"),
                "indent": cell.get("indent"),
                "shrinkToFit": cell.get("shrink"),
            }),
            "border": border,
            "protection": copy.deepcopy(cell.get("protection") or {
                "locked": cell.get("locked", True),
                "hidden": cell.get("hidden_cell", False),
            }),
            "xf": copy.deepcopy(cell.get("xf") or {}),
            "named_style": cell.get("named_style"),
            "quote_prefix": bool(cell.get("qp")),
        }
    return result


def _patch_mapping(target: dict, updates: dict, *, clear_none: bool = False) -> dict:
    before = copy.deepcopy(target)
    for key, value in updates.items():
        if value is None and clear_none:
            target.pop(key, None)
        else:
            target[key] = copy.deepcopy(value)
    return before


def _sheet_index(data: dict, sheet_name: str | None) -> int | None:
    if sheet_name is None:
        return None
    for index, sheet in enumerate(data.get("sheets") or []):
        if sheet.get("name") == sheet_name:
            return index
    raise ValueError(f"Sheet {sheet_name!r} not found.")


def _shift_advanced_sheet_meta(sheet: dict, row_maps=None, col_maps=None) -> None:
    if sheet.get("auto_filter_model"):
        sheet["auto_filter_model"] = _shift_auto_filter_ranges(
            sheet["auto_filter_model"], row_maps, col_maps,
        )
    for key in ("ignored_errors", "protected_ranges"):
        kept = []
        for item in sheet.get(key) or []:
            current = copy.deepcopy(item)
            if current.get("sqref"):
                current["sqref"] = _shift_sqref(current["sqref"], row_maps, col_maps)
                if not current["sqref"]:
                    continue
            kept.append(current)
        if key in sheet:
            sheet[key] = kept
    for view in sheet.get("sheet_views") or []:
        if view.get("topLeftCell"):
            view["topLeftCell"] = _shift_ref(view["topLeftCell"], row_maps, col_maps)
        pane = view.get("pane") or {}
        if pane.get("topLeftCell"):
            pane["topLeftCell"] = _shift_ref(pane["topLeftCell"], row_maps, col_maps)
        for selection in view.get("selections") or []:
            if selection.get("activeCell"):
                selection["activeCell"] = _shift_ref(selection["activeCell"], row_maps, col_maps)
            if selection.get("sqref"):
                selection["sqref"] = _shift_sqref(selection["sqref"], row_maps, col_maps)
    breaks = sheet.get("page_breaks") or {}
    for axis, maps in (("rows", row_maps), ("columns", col_maps)):
        if maps:
            updated = []
            for item in breaks.get(axis) or []:
                value = maps[0](int(item.get("id", 1)) - 1)
                if value is None:
                    continue
                copied = copy.deepcopy(item)
                copied["id"] = value + 1
                updated.append(copied)
            if axis in breaks:
                breaks[axis] = updated
        if axis in breaks or f"{axis}_count" in breaks or f"{axis}_manualBreakCount" in breaks:
            entries = breaks.get(axis) or []
            breaks[f"{axis}_count"] = len(entries)
            breaks[f"{axis}_manualBreakCount"] = sum(
                1 for item in entries if item.get("man", True)
            )


_SHEETNAME_SIMPLE_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")


def _formula_sheet_prefix(name: str) -> str:
    """Sheet prefix for formulas — quoted when Excel requires quoting
    (spaces/specials, or names that look like cell references such as D2)."""
    looks_like_ref = bool(re.match(r"^[A-Za-z]{1,3}\d+$", name)) or \
        bool(re.match(r"^[Rr]\d|^[Cc]\d", name))
    if _SHEETNAME_SIMPLE_RE.match(name) and not looks_like_ref:
        return f"{name}!"
    return "'" + name.replace("'", "''") + "'!"


def _rename_sheet_in_formula(text, old: str, new: str):
    """Rewrite 'Old Name'! / OldName! references in a formula/defined-name."""
    if not text or not isinstance(text, str):
        return text
    new_ref = _formula_sheet_prefix(new)
    text = text.replace("'" + old.replace("'", "''") + "'!", new_ref)
    if _SHEETNAME_SIMPLE_RE.match(old):
        text = re.sub(rf"(?<![A-Za-z0-9_.!'\"]){re.escape(old)}!",
                      lambda m: new_ref, text)
    return text


def _rewrite_sheet_reference_value(value, old: str, new: str | None):
    if not isinstance(value, str) or not value:
        return value
    from openpyxl.formula import Tokenizer
    hash_prefix = "#" if value.startswith("#") else ""
    body = value[1:] if hash_prefix else value
    had_equals = body.startswith("=")
    formula = body if had_equals else "=" + body
    try:
        tokenizer = Tokenizer(formula)
        changed = False
        replacement = _formula_sheet_prefix(new) if new is not None else "#REF!"
        for token in tokenizer.items:
            if token.type != "OPERAND" or token.subtype != "RANGE":
                continue
            sheet_name, reference = _split_sheet_prefix(token.value)
            if sheet_name == old:
                token.value = replacement + reference
                changed = True
        rendered = tokenizer.render() if changed else formula
        rendered = rendered if had_equals else rendered[1:]
        return hash_prefix + rendered
    except Exception:
        if new is None:
            old_prefixes = (_formula_sheet_prefix(old), "'" + old.replace("'", "''") + "'!")
            for prefix in old_prefixes:
                body = body.replace(prefix, "#REF!")
            return hash_prefix + body
        return _rename_sheet_in_formula(value, old, new)


def _value_targets_sheet(value, sheet_name: str) -> bool:
    if not isinstance(value, str) or not value:
        return False
    from openpyxl.formula import Tokenizer
    body = value[1:] if value.startswith("#") else value
    formula = body if body.startswith("=") else "=" + body
    try:
        return any(
            token.type == "OPERAND"
            and token.subtype == "RANGE"
            and _split_sheet_prefix(token.value)[0] == sheet_name
            for token in Tokenizer(formula).items
        )
    except Exception:
        return _formula_sheet_prefix(sheet_name) in value or (
            "'" + sheet_name.replace("'", "''") + "'!"
        ) in value


def _rewrite_xml_formula_references(xml: str, tags, old: str, new: str | None) -> str:
    for tag in tags:
        pattern = rf"(<(?:[A-Za-z_][\w.-]*:)?{re.escape(tag)}\b[^>]*>)(.*?)(</(?:[A-Za-z_][\w.-]*:)?{re.escape(tag)}>)"

        def repl(match):
            value = html.unescape(match.group(2))
            rewritten = _rewrite_sheet_reference_value(value, old, new)
            return match.group(1) + html.escape(rewritten, quote=False) + match.group(3)

        xml = re.sub(pattern, repl, xml, flags=re.DOTALL)
    return xml


def _rewrite_sheet_cell_formulas(sheet: dict, old: str, new: str | None) -> None:
    for row in sheet.get("rows") or []:
        for cell in row.get("cells") or []:
            if not _is_formula_value(cell):
                continue
            rewritten = _rewrite_sheet_reference_value(cell["v"], old, new)
            if rewritten == cell["v"]:
                continue
            cell["v"] = rewritten
            cell["value"] = rewritten
            if isinstance(cell.get("formula"), dict):
                cell["formula"]["text"] = rewritten
                cell["formula"]["cached_value_state"] = "missing"
                cell["formula"].pop("cached_value", None)
            cell.pop("cached_value", None)
            cell["cached_value_state"] = "missing"


def _rewrite_drawing_formula_files(sheet: dict, old: str, new: str | None) -> None:
    import base64
    drawing_data = sheet.get("drawing_data") or {}
    for path, payload in list((drawing_data.get("files") or {}).items()):
        if "/charts/" not in path or not path.lower().endswith(".xml"):
            continue
        try:
            xml = base64.b64decode(payload).decode("utf-8")
            rewritten = _rewrite_xml_formula_references(xml, ("f",), old, new)
            drawing_data["files"][path] = base64.b64encode(rewritten.encode("utf-8")).decode("ascii")
        except Exception:
            continue


def _rewrite_sheet_metadata_references(sheet: dict, old: str, new: str | None) -> None:
    _rewrite_sheet_cell_formulas(sheet, old, new)
    for validation in sheet.get("validations") or []:
        for key in ("formula1", "formula2"):
            if isinstance(validation.get(key), str):
                validation[key] = _rewrite_sheet_reference_value(validation[key], old, new)
    if sheet.get("data_validations_xml"):
        sheet["data_validations_xml"] = _rewrite_xml_formula_references(
            sheet["data_validations_xml"], ("formula1", "formula2"), old, new,
        )
    if sheet.get("cf_xml"):
        sheet["cf_xml"] = [
            _rewrite_xml_formula_references(block, ("formula",), old, new)
            for block in sheet["cf_xml"]
        ]

    rewritten_links = {}
    for coordinate, hyperlink in (sheet.get("hyperlinks") or {}).items():
        item = copy.deepcopy(hyperlink)
        location = item.get("location")
        if new is None and _value_targets_sheet(location, old):
            item.pop("location", None)
            if not item.get("target"):
                continue
        elif isinstance(location, str):
            item["location"] = _rewrite_sheet_reference_value(location, old, new)
        rewritten_links[coordinate] = item
    if "hyperlinks" in sheet:
        sheet["hyperlinks"] = rewritten_links

    for table in sheet.get("tables") or []:
        for column in table.get("columns") or []:
            for key in ("calculatedColumnFormula", "totalsRowFormula"):
                value = column.get(key)
                if isinstance(value, str):
                    column[key] = _rewrite_sheet_reference_value(value, old, new)
                elif isinstance(value, dict):
                    for text_key in ("text", "attr_text"):
                        if isinstance(value.get(text_key), str):
                            value[text_key] = _rewrite_sheet_reference_value(value[text_key], old, new)
                            break

    if sheet.get("print_area"):
        sheet["print_area"] = _rewrite_sheet_reference_value(sheet["print_area"], old, new)
    for key in ("rows", "cols"):
        if isinstance((sheet.get("print_titles") or {}).get(key), str):
            sheet["print_titles"][key] = _rewrite_sheet_reference_value(
                sheet["print_titles"][key], old, new,
            )
    for creation in sheet.get("drawing_creations") or []:
        if creation.get("type") == "chart" and isinstance(creation.get("source_range"), str):
            creation["source_range"] = _rewrite_sheet_reference_value(
                creation["source_range"], old, new,
            )
    _rewrite_drawing_formula_files(sheet, old, new)


def _rewrite_workbook_sheet_references(data: dict, old: str, new: str | None) -> None:
    for sheet in data.get("sheets") or []:
        _rewrite_sheet_metadata_references(sheet, old, new)
    for defined_name in data.get("named_ranges") or []:
        if isinstance(defined_name.get("value"), str):
            defined_name["value"] = _rewrite_sheet_reference_value(
                defined_name["value"], old, new,
            )


def _remap_workbook_view_sheet_ids(data: dict, mapper, fallback_index: int | None = None) -> None:
    sheet_count = len(data.get("sheets") or [])
    fallback = 0 if not sheet_count else max(0, min(fallback_index or 0, sheet_count - 1))
    views = list(data.get("workbook_views") or [])
    if data.get("wb_view") is not None:
        views.append(data["wb_view"])
    for view in views:
        for key in ("activeTab", "firstSheet"):
            if view.get(key) is None:
                continue
            try:
                mapped = mapper(int(view[key]))
            except Exception:
                continue
            view[key] = fallback if mapped is None else mapped


def _copy_scoped_defined_names(data: dict, source_index: int) -> list[dict]:
    return [
        copy.deepcopy(item)
        for item in data.get("named_ranges") or []
        if item.get("sheet_id") is not None and int(item["sheet_id"]) == source_index
    ]


def _append_copied_defined_names(data: dict, names: list[dict], new_index: int,
                                 old_sheet: str, new_sheet: str) -> None:
    for item in names:
        item["sheet_id"] = new_index
        if isinstance(item.get("value"), str):
            item["value"] = _rewrite_sheet_reference_value(item["value"], old_sheet, new_sheet)
        data.setdefault("named_ranges", []).append(item)


def _ensure_unique_sheet_codename(data: dict, sheet: dict) -> None:
    properties = sheet.get("sheet_properties") or {}
    code_name = properties.get("codeName")
    if not code_name:
        return
    existing = {
        (other.get("sheet_properties") or {}).get("codeName")
        for other in data.get("sheets") or []
        if other is not sheet
    }
    if code_name not in existing:
        return
    suffix = 2
    candidate = f"{code_name}_{suffix}"
    while candidate in existing:
        suffix += 1
        candidate = f"{code_name}_{suffix}"
    properties["codeName"] = candidate
    sheet["sheet_properties"] = properties


def _remap_named_range_sheet_ids(data: dict, mapper) -> None:
    """Remap localSheetId indices after sheet add/copy/move/delete; None = drop."""
    kept = []
    for nr in data.get("named_ranges") or []:
        sid = nr.get("sheet_id")
        if sid is None:
            kept.append(nr)
            continue
        new = mapper(int(sid))
        if new is None:
            continue
        nr["sheet_id"] = new
        kept.append(nr)
    data["named_ranges"] = kept


def _dxf_blocks(xml: str | None) -> list[str]:
    if not xml:
        return []
    return re.findall(
        r"<(?:[A-Za-z_][\w.-]*:)?dxf\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?dxf>)",
        xml,
        flags=re.DOTALL,
    )


def _render_dxfs_xml(existing_xml: str | None, blocks: list[str]) -> str | None:
    if not blocks:
        return None
    if existing_xml:
        opening = re.search(r"<(?:[A-Za-z_][\w.-]*:)?dxfs\b[^>]*>", existing_xml)
        closing = re.search(r"</(?:[A-Za-z_][\w.-]*:)?dxfs>", existing_xml)
        if opening and closing:
            tag = opening.group(0)
            if re.search(r'\bcount="\d+"', tag):
                tag = re.sub(r'\bcount="\d+"', f'count="{len(blocks)}"', tag, count=1)
            else:
                tag = tag[:-1] + f' count="{len(blocks)}">'
            return tag + "".join(blocks) + closing.group(0)
    return f'<dxfs count="{len(blocks)}">' + "".join(blocks) + "</dxfs>"


def _copy_dxf_dependencies(src_data: dict, dst_data: dict, sheet: dict) -> None:
    source_blocks = _dxf_blocks(src_data.get("dxfs_xml"))
    if not source_blocks:
        return
    destination_blocks = _dxf_blocks(dst_data.get("dxfs_xml"))
    mapping = {}
    for source_index, block in enumerate(source_blocks):
        try:
            destination_index = destination_blocks.index(block)
        except ValueError:
            destination_index = len(destination_blocks)
            destination_blocks.append(block)
        mapping[source_index] = destination_index
    dst_data["dxfs_xml"] = _render_dxfs_xml(dst_data.get("dxfs_xml"), destination_blocks)

    def remap_xml(xml: str) -> str:
        return re.sub(
            r'\bdxfId="(\d+)"',
            lambda match: f'dxfId="{mapping.get(int(match.group(1)), int(match.group(1)))}"',
            xml,
        )

    if sheet.get("cf_xml"):
        sheet["cf_xml"] = [remap_xml(block) for block in sheet["cf_xml"]]
    table_dxf_keys = (
        "headerRowDxfId", "dataDxfId", "totalsRowDxfId", "headerRowBorderDxfId",
        "tableBorderDxfId", "totalsRowBorderDxfId",
    )
    column_dxf_keys = ("headerRowDxfId", "dataDxfId", "totalsRowDxfId")
    for table in sheet.get("tables") or []:
        for key in table_dxf_keys:
            if table.get(key) is not None:
                table[key] = mapping.get(int(table[key]), int(table[key]))
        for column in table.get("columns") or []:
            for key in column_dxf_keys:
                if column.get(key) is not None:
                    column[key] = mapping.get(int(column[key]), int(column[key]))


def _drawing_part_paths(data: dict) -> set[str]:
    paths = set()
    for sheet in data.get("sheets") or []:
        drawing_data = sheet.get("drawing_data") or {}
        if drawing_data.get("drawing_file"):
            paths.add(drawing_data["drawing_file"])
        paths.update((drawing_data.get("files") or {}).keys())
    return paths


def _unique_drawing_part_path(path: str, used: set[str]) -> str:
    import posixpath
    if path not in used:
        used.add(path)
        return path
    directory, filename = posixpath.split(path)
    match = re.match(r"^(.*?)(\d+)?(\.[^.]+)$", filename)
    stem, suffix = (match.group(1), match.group(3)) if match else (filename, "")
    index = 1
    candidate = posixpath.join(directory, f"{stem}{index}{suffix}")
    while candidate in used:
        index += 1
        candidate = posixpath.join(directory, f"{stem}{index}{suffix}")
    used.add(candidate)
    return candidate


def _passthrough_part_paths(data: dict) -> set[str]:
    paths = set(((data.get("_lossless") or {}).get("package_graph") or {}).get("parts") or {})
    for sheet in data.get("sheets") or []:
        passthrough = sheet.get("passthrough_relationships") or {}
        paths.update((passthrough.get("parts") or {}).keys())
    package_edits = data.get("_package_edits") or {}
    upserts = package_edits.get("upsert") or {}
    if isinstance(upserts, dict):
        paths.update(str(path).lstrip("/") for path in upserts)
    elif isinstance(upserts, list):
        paths.update(
            str(record.get("part") or record.get("name") or record.get("part_name")).lstrip("/")
            for record in upserts
            if record.get("part") or record.get("name") or record.get("part_name")
        )
    return paths


def _rebase_copied_passthrough_parts(dst_data: dict, sheet: dict) -> None:
    passthrough = sheet.get("passthrough_relationships") or {}
    parts = passthrough.get("parts") or {}
    if not parts:
        return
    used = _passthrough_part_paths(dst_data)
    mapping = {
        path: _unique_drawing_part_path(path, used)
        for path in parts
    }
    passthrough["parts"] = {
        mapping[path]: payload
        for path, payload in parts.items()
    }
    for relationship in passthrough.get("relationships") or []:
        target_part = relationship.get("target_part")
        if target_part in mapping:
            relationship["target_part"] = mapping[target_part]


def _rebase_copied_drawing_parts(dst_data: dict, sheet: dict) -> None:
    import posixpath
    drawing_data = sheet.get("drawing_data") or {}
    old_drawing = drawing_data.get("drawing_file")
    if not old_drawing:
        return
    used = _drawing_part_paths(dst_data)
    part_mapping = {old_drawing: _unique_drawing_part_path(old_drawing, used)}
    for path in (drawing_data.get("files") or {}):
        part_mapping[path] = _unique_drawing_part_path(path, used)

    new_drawing = part_mapping[old_drawing]
    drawing_data["drawing_file"] = new_drawing
    drawing_data["files"] = {
        part_mapping[path]: payload
        for path, payload in (drawing_data.get("files") or {}).items()
    }
    relationships = drawing_data.get("drawing_rels")
    if relationships:
        old_directory = posixpath.dirname(old_drawing)
        new_directory = posixpath.dirname(new_drawing)

        def replace_target(match):
            target = match.group(1)
            absolute = target.lstrip("/") if target.startswith("/") else posixpath.normpath(
                posixpath.join(old_directory, target)
            )
            mapped = part_mapping.get(absolute)
            if not mapped:
                return match.group(0)
            relative = posixpath.relpath(mapped, new_directory)
            return f'Target="{relative}"'

        drawing_data["drawing_rels"] = re.sub(
            r'Target="([^"]+)"', replace_target, relationships,
        )


def _dedupe_table_names(data: dict, sheet: dict) -> None:
    """Table names, displayNames, and ids must be unique workbook-wide."""
    existing_names = {
        value
        for other in data["sheets"] if other is not sheet
        for table in (other.get("tables") or [])
        for value in (table.get("name"), table.get("displayName"))
        if value
    }
    existing_ids = {
        int(table["id"])
        for other in data["sheets"] if other is not sheet
        for table in (other.get("tables") or [])
        if table.get("id") is not None
    }
    next_id = max(existing_ids, default=0) + 1
    for table in sheet.get("tables") or []:
        base_name = table.get("displayName") or table.get("name") or "Table"
        candidate = base_name
        suffix = 2
        while candidate in existing_names:
            candidate = f"{base_name}_{suffix}"
            suffix += 1
        table["name"] = candidate
        table["displayName"] = candidate
        existing_names.add(candidate)
        table_id = int(table.get("id") or 0)
        if not table_id or table_id in existing_ids:
            table_id = next_id
            next_id += 1
        table["id"] = table_id
        existing_ids.add(table_id)


# ── Formula handling ──────────────────────────────────────────────────────────

def _is_formula_value(cd: dict) -> bool:
    v = cd.get("v")
    return isinstance(v, str) and v.startswith("=") and cd.get("dt") != "s"


def _normalize_input_value(value):
    """Apply the Excel-like input contract for written cell values.

    Returns (value, force_text). A leading apostrophe forces the rest to be
    stored as literal text (the apostrophe itself is NOT stored — it becomes
    the quotePrefix style flag, exactly like typing 'text in Excel).
    A leading "=" that does not even tokenize as a formula is stored as text.
    """
    if isinstance(value, str):
        if value.startswith("'"):
            return value[1:], True
        if value.startswith("="):
            from openpyxl.formula import Tokenizer
            try:
                Tokenizer(value)
            except Exception:
                return value, True
    return value, False


def _store_cell_value(cell: dict, value, implicit_defaults: dict | None = None) -> None:
    _promote_implicit_cell(cell, implicit_defaults)
    _cell_baseline(cell)
    value, force_text = _normalize_input_value(value)
    cell["v"] = value
    if force_text:
        cell["dt"] = "s"
        cell["qp"] = True
    else:
        cell.pop("dt", None)
        cell.pop("qp", None)


def _split_sheet_prefix(ref: str):
    """'My Sheet'!A5 → ("My Sheet", "A5"); Data!C3 → ("Data", "C3"); A5 → (None, "A5")."""
    m = re.match(r"^'((?:[^']|'')+)'!(.*)$", ref)
    if m:
        return m.group(1).replace("''", "'"), m.group(2)
    m = re.match(r"^([A-Za-z_][A-Za-z0-9_.]*)!(.*)$", ref)
    if m:
        return m.group(1), m.group(2)
    return None, ref


def _shift_formula_str(formula: str, current_sheet, target_sheet: str,
                       row_maps=None, col_maps=None) -> str:
    """Shift cell/range references in one formula that target target_sheet.

    Only RANGE operand tokens are touched (string literals, names and
    structured table references pass through untouched). References whose
    whole area was deleted become #REF!. Any parse problem leaves the
    formula unchanged — never worse than not shifting.
    """
    from openpyxl.formula import Tokenizer
    try:
        tok = Tokenizer(formula)
        changed = False
        for t in tok.items:
            if t.type != "OPERAND" or t.subtype != "RANGE":
                continue
            sheet, rest = _split_sheet_prefix(t.value)
            if sheet is None:
                if current_sheet != target_sheet:
                    continue
            elif sheet != target_sheet:
                continue
            new_rest = _shift_ref(rest, row_maps, col_maps)
            if new_rest is None:
                new_rest = "#REF!"
            if new_rest != rest:
                t.value = t.value[: len(t.value) - len(rest)] + new_rest
                changed = True
        return tok.render() if changed else formula
    except Exception:
        return formula


def _shift_formulas_workbook(data: dict, target_sheet: str,
                             row_maps=None, col_maps=None) -> None:
    """Rewrite formula/name references after a structural worksheet edit."""
    for sheet in data["sheets"]:
        current_sheet = sheet["name"]
        for row in sheet["rows"]:
            for cell in row.get("cells") or []:
                formula = _formula_model(cell)
                if not formula:
                    continue
                old_text = cell.get("v")
                if not isinstance(old_text, str) or not old_text.startswith("="):
                    continue
                new_text = _shift_formula_str(
                    old_text, current_sheet, target_sheet, row_maps, col_maps,
                )
                if new_text != old_text:
                    formula["text"] = new_text
                    formula["cached_value_state"] = "missing"
                    formula.pop("cached_value", None)
                    cell["formula"] = formula
                    cell["v"] = new_text
                    cell["value"] = new_text
                    cell.pop("cached_value", None)
                    cell["cached_value_state"] = "missing"

        for validation in sheet.get("validations") or []:
            for key in ("formula1", "formula2"):
                if isinstance(validation.get(key), str):
                    validation[key] = _shift_formula_value(
                        validation[key], current_sheet, target_sheet, row_maps, col_maps,
                    )

        if current_sheet != target_sheet:
            if sheet.get("data_validations_xml"):
                sheet["data_validations_xml"] = _shift_xml_formula_nodes(
                    sheet["data_validations_xml"], ("formula1", "formula2"),
                    current_sheet, target_sheet, row_maps, col_maps,
                )
            if sheet.get("cf_xml"):
                sheet["cf_xml"] = [
                    _shift_xml_formula_nodes(
                        block, ("formula",), current_sheet, target_sheet, row_maps, col_maps,
                    )
                    for block in sheet["cf_xml"]
                ]

        for hyperlink in (sheet.get("hyperlinks") or {}).values():
            if isinstance(hyperlink.get("location"), str):
                hyperlink["location"] = _shift_formula_value(
                    hyperlink["location"], current_sheet, target_sheet, row_maps, col_maps,
                )

        for table in sheet.get("tables") or []:
            for column in table.get("columns") or []:
                for key in ("calculatedColumnFormula", "totalsRowFormula"):
                    if column.get(key) is not None:
                        column[key] = _shift_formula_model(
                            column[key], current_sheet, target_sheet, row_maps, col_maps,
                        )

        shifted_creations = []
        for creation in sheet.get("drawing_creations") or []:
            if creation.get("type") == "chart" and creation.get("source_range"):
                creation["source_range"] = _shift_range_value(
                    creation["source_range"], current_sheet, target_sheet, row_maps, col_maps,
                )
                if not creation["source_range"]:
                    continue
            shifted_creations.append(creation)
        if "drawing_creations" in sheet:
            sheet["drawing_creations"] = shifted_creations

    for defined_name in data.get("named_ranges") or []:
        value = defined_name.get("value")
        if not isinstance(value, str) or not value:
            continue
        current_sheet = None
        scope = defined_name.get("sheet_id")
        if isinstance(scope, int) and 0 <= scope < len(data["sheets"]):
            current_sheet = data["sheets"][scope]["name"]
        defined_name["value"] = _shift_formula_value(
            value, current_sheet, target_sheet, row_maps, col_maps,
        )


def _rename_sheet_in_cell_formulas(data: dict, old: str, new: str) -> None:
    from openpyxl.formula import Tokenizer
    new_prefix = _formula_sheet_prefix(new)
    for sd in data["sheets"]:
        for row in sd["rows"]:
            for cd in row.get("cells") or []:
                if not _is_formula_value(cd):
                    continue
                try:
                    tok = Tokenizer(cd["v"])
                    changed = False
                    for t in tok.items:
                        if t.type == "OPERAND" and t.subtype == "RANGE":
                            sheet, rest = _split_sheet_prefix(t.value)
                            if sheet == old:
                                t.value = new_prefix + rest
                                changed = True
                    if changed:
                        rendered = tok.render()
                        cd["v"] = rendered
                        cd["value"] = rendered
                        if isinstance(cd.get("formula"), dict):
                            cd["formula"]["text"] = rendered
                            cd["formula"]["cached_value_state"] = "missing"
                            cd["formula"].pop("cached_value", None)
                        cd.pop("cached_value", None)
                        cd["cached_value_state"] = "missing"
                except Exception:
                    pass


_INSTRUCTIONS = """\
excel-tools MCP — Excel round-trip editing.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXCEL EDITING WORKFLOW
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Standard flow:
  1. excel_get_info     — lightweight: sheet names + dimensions, no session needed
  2. excel_load         — load file into server session → returns session_key
  3. excel_to_markdown  — read full content; rows annotated with 0-based row_index,
                          columns annotated as col_0, col_1, … so you know exact
                          indices to pass to edit/delete/insert tools
  4. <edit tools>       — excel_edit_cells / excel_insert_rows / excel_delete_rows / …
  5. excel_save         — write session back to .xlsx (omit output_path to overwrite
                          the original file; pass a new path to save elsewhere)

IMPORTANT — index conventions:
  • All row_index and col_index values are 0-based.
  • excel_to_markdown shows them explicitly — use those values directly.
  • Merge info (r1 c1 r2 c2) returned by excel_get_cell is also 0-based.
  • Merged cells: origin shows [M rowspan×colspan]; slave cells are blank.
    Only edit/delete the ORIGIN cell of a merged range.

IMPORTANT — multi-insert pattern (avoid index drift):
  1. Clone ALL template rows first (excel_clone_rows) before any inserts.
  2. Insert from BOTTOM to TOP (largest after_index first).
  This ensures earlier rows are not shifted before later inserts.

IMPORTANT — automatic reference shifting:
  • Inserting/deleting rows & columns automatically shifts merges, hyperlinks,
    comments, data validations, conditional-formatting ranges, tables,
    auto-filter, freeze panes, print titles, image/chart anchors, AND cell
    formulas / defined names referencing the edited sheet (from every loaded
    sheet). References whose entire area was deleted become #REF!.
  • Caveat: with a sheet_name-filtered session, formulas on UNLOADED sheets
    are not rewritten — load the full workbook before structural edits if
    other sheets reference the edited one.
  • Editing a slave (merged) cell raises an error — edit the origin cell.

IMPORTANT — writing values that start with = + - :
  • A value starting with "=" is stored as a FORMULA.
  • To store literal text that starts with "=" (or any text), prefix it with
    a single apostrophe: "'=not a formula" → cell text «=not a formula».
    The apostrophe is not stored; it becomes Excel's quote-prefix flag,
    exactly like typing 'text in Excel.
  • Values starting with "+" or "-" are already stored as text automatically
    when they are not numbers — no prefix needed.
  • Merging over an existing merged region raises an error — unmerge first.
  • convert_to_markdown is read-only and can accept Excel-family files readable
    by the converter.
  • Session/edit/save tools support OOXML Excel packages: .xlsx, .xlsm,
    .xltx, and .xltm. Macro/template parts are preserved best-effort.
    Legacy/binary .xls and .xlsb need read-only conversion or conversion to
    OOXML before editing.
  • excel_load with sheet_name loads ONLY that sheet, but excel_save merges
    the other sheets back from disk automatically — nothing is lost.
  • excel_save validates the generated .xlsx before replacing the destination.
    If validation fails, the existing destination file is left untouched.
  • Advanced DrawingML/charts/images/unknown OOXML parts are preserved
    best-effort. Use excel_validate_workbook and excel_verify_preservation for risky
    files before trusting a save workflow.

IMPORTANT — session_key:
  • session_key is the absolute file path returned by excel_load.
  • It persists server-side across all tool calls in this conversation.
  • You do NOT need to reload the file between edits — just reuse the same key.
  • Call excel_reload to discard in-memory changes and re-read from disk.
  • Call excel_close when done to free server memory.

Quick reference — one-shot conversion:
  convert_to_markdown  read-only Markdown export; supports sheet/range/max limits
  excel_get_workbook_summary compact file summary without excel_load/session
  excel_get_sheet_preview compact top-left sheet preview without session

Quick reference — session lifecycle:
  excel_get_info       sheet names + dimensions (no session needed)
  excel_load           load file → session_key
  excel_save           write session back with a two-day pre-save backup
  excel_save_as_copy   save to a different .xlsx path without overwriting source
  excel_validate_workbook validate .xlsx ZIP/XML structure + feature summary
  excel_diff_package   compare package manifests and semantic part hashes
  excel_verify_preservation compare semantic workbook features; auto-uses backup
  excel_reload         reload from disk, discard unsaved changes
  excel_close          remove session from cache, free memory

Quick reference — sheet management:
  excel_add_sheet      add a new empty sheet
  excel_delete_sheet   delete a sheet (cannot delete the only sheet)
  excel_rename_sheet   rename a sheet
  excel_copy_sheet     duplicate a sheet within the same workbook
  excel_move_sheet     reorder a sheet to a new position

Quick reference — reading:
  excel_to_markdown    annotated Markdown view; supports max_rows/max_cols
  excel_to_markdown_range Markdown table for one A1/0-based range
  excel_list_tables    list Excel table objects captured in session
  excel_list_defined_names list workbook defined names/named ranges
  excel_get_rows       row range as JSON; values_only=True for compact output
  excel_read_range     exact A1 or 0-based rectangular range; token-efficient
  excel_get_cell       single cell with full style metadata
  excel_get_column     all cells in a column
  excel_find_cells     find literal/regex values or formulas across workbook
  excel_get_shapes     list captured DrawingML shape/image/chart metadata

Quick reference — editing rows:
  excel_edit_cells     edit cell values across one or more rows
  excel_insert_rows    insert rows at one or more positions in one call
  excel_clone_rows     deep-clone rows → JSON for modification before insert
  excel_copy_row       clone a row AND insert immediately (one step)
  excel_delete_rows    delete rows by index list OR start_row+end_row range
  excel_clear_range    clear values and/or styles from a rectangular range

Quick reference — editing columns:
  excel_insert_column  insert a new empty column at a position
  excel_copy_column    copy a column to a new position
  excel_delete_column  delete a column from all rows

Quick reference — merge:
  excel_merge_cells    merge a range (unmerge=False) or unmerge origin (unmerge=True)

Quick reference — formatting:
  excel_set_style      set style on a cell/range (fill/fcolor/font/strike/align/numfmt)
  excel_set_font_color set font color on a cell or range
  excel_set_strike     enable/disable strikethrough on a cell or range
  excel_set_borders    set/remove borders on a cell range
  excel_set_dimension  set row height (axis="row") or column width (axis="col")
                       • axis="row", index=3, size=20   → set row 3 height to 20pt
                       • axis="col", index=1, size=22   → set col 1 (B) width to 22
                       • size=null resets to auto
  excel_autofit_cols   auto-fit column widths to content (heuristic)
  excel_freeze_panes   freeze header rows and/or columns
  excel_set_data_validation  add dropdown list validation to a cell range
  excel_update_shape_text    replace DrawingML shape text with plain or rich runs
  excel_set_shape_style      set simple DrawingML shape fill/outline/text color

Quick reference — search & fill:
  excel_find_rows      find rows matching a value or regex in a column
  excel_fill_column    fill a column range with a constant or sequence
  excel_fill_rows      clone a template row N times and insert (stamp pattern)
"""

mcp = FastMCP("excel-tools", instructions=_INSTRUCTIONS)


def _install_deferred_heavy_cancellation() -> None:
    import anyio
    from mcp.shared.session import RequestResponder
    from mcp.types import CallToolRequest

    if getattr(RequestResponder, "_docloupe_heavy_cancel_patch", False):
        return

    heavy_tool_names = frozenset({
        "excel_load",
        "excel_save",
        "excel_save_as_copy",
        "excel_verify_preservation",
    })
    original_cancel = RequestResponder.cancel
    original_respond = RequestResponder.respond

    async def cancel(responder) -> None:
        request = getattr(responder.request, "root", None)
        if isinstance(request, CallToolRequest) and request.params.name in heavy_tool_names:
            responder._cancel_scope.cancel()
            responder._docloupe_deferred_cancel = True
            return
        await original_cancel(responder)

    async def respond(responder, response) -> None:
        if getattr(responder, "_docloupe_deferred_cancel", False):
            if not responder._entered:
                raise RuntimeError("RequestResponder must be used as a context manager")
            assert not responder._completed, "Request already responded to"
            responder._completed = True
            with anyio.CancelScope(shield=True):
                await responder._session._send_response(
                    request_id=responder.request_id,
                    response=response,
                )
            return
        await original_respond(responder, response)

    RequestResponder.cancel = cancel
    RequestResponder.respond = respond
    RequestResponder._docloupe_heavy_cancel_patch = True


_install_deferred_heavy_cancellation()
_package_tools = register_package_tools(mcp, _get_session)


# ── 1. Info ───────────────────────────────────────────────────────────────────

@mcp.tool()
def convert_to_markdown(file_path: str, sheet_name: str | None = None, range_ref: str | None = None, max_rows: int | None = None, max_cols: int | None = None, include_styles: bool = False) -> TextContent:
    """
    Convert an Excel-family file to Markdown in one call without creating a session.

    This read-only conversion is for quick inspection/export workflows. Use
    excel_load + excel_to_markdown when you need Markdown generated from an
    editable .xlsx in-memory session, including unsaved changes.

    Args:
        file_path: Path or file:// URI to an Excel-family file readable by the converter.
        sheet_name: Sheet to export; omit to export all sheets.
        range_ref: Optional A1 range to export from sheet_name.
        max_rows: Optional maximum rows per exported sheet/range.
        max_cols: Optional maximum columns per exported sheet/range.
        include_styles: Reserved for clients; current Markdown export remains content-focused.

    Returns:
        Markdown content (text/markdown) representing workbook sheets.
    """
    from excel_converter import convert_excel_to_markdown

    path = uri_to_path(file_path)
    data = serialize_excel(str(path), sheet_name)
    if range_ref:
        if not sheet_name:
            raise ValueError("sheet_name is required when range_ref is provided")
        r1, r2, c1, c2 = _excel_range_to_indices(range_ref)
        for sheet in data.get("sheets", []):
            sheet["rows"] = [
                {**row, "cells": row.get("cells", [])[c1:c2 + 1]}
                for row in sheet.get("rows", [])[r1:r2 + 1]
            ]
    data = _limit_workbook_data(data, max_rows=max_rows, max_cols=max_cols)
    markdown = convert_excel_to_markdown(data)
    return TextContent(type="text", text=markdown, mimeType="text/markdown")

@mcp.tool()
def excel_get_info(uri: str) -> str:
    """
    Return summary info about an Excel file: sheet names, row and column counts.

    Use this first to understand the file structure before loading.

    Args:
        uri: Local file path or file:// URI to the .xlsx file

    Returns:
        JSON: {source, sheets: [{name, max_row, max_column}]}
    """
    import openpyxl
    path = uri_to_path(uri)
    wb = openpyxl.load_workbook(
        str(path),
        read_only=False,
        data_only=False,
        keep_vba=False,
    )
    try:
        info = {
            "source": str(path),
            "sheets": [],
        }
        for name in wb.sheetnames:
            ws = wb[name]
            info["sheets"].append({
                "name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "state": ws.sheet_state,
                "hidden": ws.sheet_state != "visible",
                "freeze_panes": ws.freeze_panes,
                "merged_ranges": len(ws.merged_cells.ranges),
                "table_count": len(getattr(ws, "tables", {}) or {}),
            })
    finally:
        _close_openpyxl_workbook(wb)
    return json.dumps(info, ensure_ascii=False)


# ── 2. Load ───────────────────────────────────────────────────────────────────


def _load_worker_operation(uri: str, sheet_name: str | None = None) -> dict:
    path = uri_to_path(uri)
    _check_supported(path)
    return serialize_excel(str(path), sheet_name)


def _publish_loaded_session(
    data: dict,
    sheet_name: str | None,
    load_metrics: dict | None = None,
    tool_started_at: float | None = None,
) -> str:
    import_started_at = time.perf_counter()
    if not isinstance(data, dict):
        raise ExcelOperationError({
            "code": "EXCEL_LOAD_RESULT_INVALID",
            "message": "The load worker did not return a workbook dictionary.",
        })
    source = data.get("source")
    sheets = data.get("sheets")
    if not isinstance(source, str) or not source:
        raise ExcelOperationError({
            "code": "EXCEL_LOAD_RESULT_INVALID",
            "message": "The load worker result is missing a source path.",
        })
    if not isinstance(sheets, list) or any(
        not isinstance(sheet, dict)
        or not isinstance(sheet.get("name"), str)
        or not isinstance(sheet.get("rows"), list)
        for sheet in sheets
    ):
        raise ExcelOperationError({
            "code": "EXCEL_LOAD_RESULT_INVALID",
            "message": "The load worker result contains an invalid sheet model.",
        })

    data["_sheet_filter"] = sheet_name
    data["_loaded_disk_names"] = [sheet["name"] for sheet in sheets]
    session_key = str(Path(source).resolve())
    data["source"] = session_key
    metrics_record = copy.deepcopy(load_metrics) if load_metrics is not None else None
    if metrics_record is not None:
        data["_load_metrics"] = metrics_record
    _store_session(session_key, data)

    sheet_names = [sheet["name"] for sheet in sheets]
    total_rows = sum(len(sheet["rows"]) for sheet in sheets)
    result = f"Loaded: session_key={session_key!r} | sheets={sheet_names} | total_rows={total_rows}"
    if metrics_record is not None:
        metrics_record["parent_session_import_seconds"] = round(
            time.perf_counter() - import_started_at,
            6,
        )
        if tool_started_at is not None:
            metrics_record["total_tool_seconds"] = round(
                time.perf_counter() - tool_started_at,
                6,
            )
        result += " | load_metrics=" + json.dumps(
            metrics_record,
            ensure_ascii=False,
            separators=(",", ":"),
        )
    return result


def excel_load(uri: str, sheet_name: str | None = None) -> str:
    """Synchronously load a workbook for direct Python callers."""
    return _publish_loaded_session(_load_worker_operation(uri, sheet_name), sheet_name)


@mcp.tool(name="excel_load")
async def _excel_load_tool(
    uri: str,
    sheet_name: str | None = None,
    timeout_seconds: float | None = None,
) -> str:
    """
    Load an OOXML workbook through a cancellable subprocess worker.

    The parent validates the bounded JSON artifact completely and only then
    publishes the decoded workbook model into the session cache. Timeout,
    cancellation, worker failure, or invalid artifacts create no new session.
    """
    import anyio

    tool_started_at = time.perf_counter()
    path = uri_to_path(uri)
    _check_supported(path)
    outcome = await run_named_operation(
        "load",
        {"uri": str(path), "sheet_name": sheet_name},
        timeout_seconds=timeout_seconds,
        operation_label="excel_load",
        return_metadata=True,
    )
    if (
        not isinstance(outcome, dict)
        or not isinstance(outcome.get("metrics"), dict)
        or "result" not in outcome
    ):
        raise ExcelOperationError({
            "code": "EXCEL_LOAD_RESULT_INVALID",
            "message": "The load supervisor returned an invalid result envelope.",
        })
    worker_metrics = outcome["metrics"].get("worker")
    parent_metrics = outcome["metrics"].get("parent")
    if not isinstance(worker_metrics, dict):
        worker_metrics = {}
    if not isinstance(parent_metrics, dict):
        parent_metrics = {}
    load_metrics = {
        "worker_serialization_seconds": worker_metrics.get("operation_seconds"),
        "worker_json_encode_seconds": worker_metrics.get("json_encode_seconds"),
        "worker_artifact_write_seconds": worker_metrics.get("artifact_write_seconds"),
        "artifact_bytes": parent_metrics.get(
            "artifact_bytes",
            worker_metrics.get("artifact_bytes"),
        ),
        "parent_artifact_read_seconds": parent_metrics.get("artifact_read_seconds"),
        "parent_artifact_validate_seconds": parent_metrics.get("artifact_validate_seconds"),
        "parent_json_decode_seconds": parent_metrics.get("json_decode_seconds"),
        "supervisor_seconds": outcome["metrics"].get("supervisor_seconds"),
    }
    return await anyio.to_thread.run_sync(
        lambda: _publish_loaded_session(
            outcome["result"],
            sheet_name,
            load_metrics,
            tool_started_at,
        )
    )


@mcp.tool()
def excel_get_session_status(session_key: str) -> str:
    """Return lightweight session status without reading or mutating workbook data."""
    with _SESSION_LOCK:
        resolved = _resolve_session_key(session_key)
        busy = copy.deepcopy(_BUSY_SESSIONS.get(resolved))
        data = _sessions[resolved]
        return json.dumps({
            "session_key": resolved,
            "busy": busy is not None,
            "operation": busy.get("operation") if busy else None,
            "started_at": busy.get("started_at") if busy else None,
            "sheet_filter": data.get("_sheet_filter"),
            "dirty_feature_count": len(data.get("_dirty_features") or []),
            "dirty_path_count": len(data.get("_dirty_paths") or []),
        }, ensure_ascii=False, indent=2)


# ── 3. Save ───────────────────────────────────────────────────────────────────

def _save_verifier_patterns(dirty_paths: list[str], dirty_features: list[str]) -> list[str]:
    """Translate session mutation paths into the verifier's semantic namespace."""
    patterns: list[str] = []

    def add(*values: str) -> None:
        for value in values:
            if value and value not in patterns:
                patterns.append(value)

    printing_parts = {"page_setup", "print_options", "header_footer", "page_breaks"}
    worksheet_parts = {
        "properties": "sheet_properties",
        "views": "sheet_views",
        "auto_filter": "filtering",
        "data_validations": "validation",
        "conditional_formatting": "validation",
        "ignored_errors": "ignored_errors",
        "hyperlinks": "hyperlinks",
        "tables": "tables",
        "comments": "comments",
        "protected_ranges": "protection",
    }
    for raw_path in dirty_paths:
        path = str(raw_path).strip("/")
        if not path:
            continue
        if path == "workbook":
            add("workbook/*", "document_properties/*")
            continue
        if path.startswith("workbook/calculation_properties"):
            add("workbook/calculation")
            continue
        if path.startswith("workbook/properties"):
            add("workbook/workbook_properties")
            continue
        if path.startswith("workbook/protection"):
            add("workbook/workbook_protection")
            continue
        if path.startswith("workbook/views"):
            add("workbook/views")
            continue
        if path.startswith("workbook/defined_names"):
            add("workbook/defined_names")
            continue
        if path.startswith("workbook/named_styles"):
            add("styles/named_styles")
            continue
        if path.startswith("document_properties"):
            add("document_properties/*")
            continue
        if not path.startswith("sheets/"):
            continue

        parts = path.split("/")
        if len(parts) < 2:
            continue
        sheet_name = parts[1]
        worksheet = f"worksheets/{sheet_name}"
        tail = parts[2:]
        if not tail:
            add("workbook/sheets", worksheet, worksheet + "/*")
            continue
        head = tail[0]
        if head == "state":
            add("workbook/sheets")
        elif head == "cells" and len(tail) >= 2:
            add(f"{worksheet}/cells/{tail[1]}*")
        elif head == "range" and len(tail) >= 3 and tail[-1] == "xf":
            try:
                first_row, last_row, first_col, last_col = _excel_range_to_indices(tail[1])
                cell_count = (last_row - first_row + 1) * (last_col - first_col + 1)
                if cell_count <= 4096:
                    for row_index in range(first_row, last_row + 1):
                        for col_index in range(first_col, last_col + 1):
                            add(f"{worksheet}/cells/{_cell_coord(row_index, col_index)}/style")
                else:
                    add(f"{worksheet}/cells/*/style")
            except Exception:
                add(f"{worksheet}/cells/*/style")
        elif head == "rows":
            if len(tail) >= 2 and tail[1].isdigit():
                add(f"{worksheet}/rows/{int(tail[1]) + 1}")
            else:
                add(worksheet + "/*")
        elif head == "columns":
            add(worksheet + "/*")
        elif head in printing_parts:
            add(f"{worksheet}/printing")
        elif head in {"print_area", "print_titles"}:
            add("workbook/defined_names")
        elif head in {"drawing_creations", "drawing_shapes"}:
            add(
                f"{worksheet}/relationships",
                "package/xl/drawings/*",
                "package/xl/charts/*",
                "package/xl/media/*",
                "package/relationships/*",
                "package/content_types/*",
            )
        elif head in worksheet_parts:
            add(f"{worksheet}/{worksheet_parts[head]}")

    features = set(dirty_features)
    if "structure" in features:
        add(
            "workbook/defined_names",
            "workbook/sheets",
            "worksheets/*/cells/*/formula",
            "package/xl/drawings/*",
            "package/xl/charts/*",
            "package/relationships/*",
        )
    if "drawings" in features:
        add(
            "worksheets/*/relationships",
            "package/xl/drawings/*",
            "package/xl/charts/*",
            "package/xl/media/*",
            "package/relationships/*",
            "package/content_types/*",
        )
    return patterns


def _save_stage_operation(
    session_data: dict,
    *,
    staging_path: str,
    verification_reference_path: str | None,
    verify_preservation: bool,
    max_differences: int,
    requested_paths: list[str],
    intentional_edit: bool,
) -> dict:
    return execute_save_stage(
        session_data,
        staging_path=staging_path,
        verification_reference_path=verification_reference_path,
        verify_preservation=verify_preservation,
        max_differences=max_differences,
        requested_paths=requested_paths,
        intentional_edit=intentional_edit,
        reconstruct=reconstruct_excel,
        verify=verify_xlsx_preservation,
        signature_report=package_signature_report,
    )


def _save_report_text(report: dict) -> str:
    backup = report.get("backup")
    backup_line = (
        f"Backup: {backup['backup_path']}\nBackup expires: {backup['expires_at']}"
        if backup else "Backup: not created (new target)"
    )
    lines = [
        f"Saved: {report['saved_path']} ({report['file_size'] // 1024} KB)",
        f"Sheets: {report['sheet_count']}, Total rows: {report['total_rows']}",
        backup_line,
        f"Verify: {report['verification_call']}",
    ]
    package_signatures = report["package_signatures"]
    lines.append(
        "Package signatures: "
        f"{package_signatures['status']} "
        f"(before={len(package_signatures['parts_before'])}, "
        f"after={len(package_signatures['parts_after'])}, "
        f"parts_preserved={package_signatures['parts_preserved']})"
    )
    requested = report.get("requested_semantic_paths") or []
    if requested:
        lines.append("Requested semantic paths: " + ", ".join(requested))
    verification = report["verification"]
    if verification["status"] == "completed":
        lines.append(
            "Verification: "
            f"preservation_ok={verification['preservation_ok']}, "
            f"changes={verification['change_count']}, "
            f"unapproved={verification['unapproved_difference_count']}"
        )
        changed = verification.get("changed_semantic_paths") or []
        if changed:
            preview = changed[:20]
            suffix = f" (+{len(changed) - len(preview)} more)" if len(changed) > len(preview) else ""
            lines.append("Changed semantic paths: " + ", ".join(preview) + suffix)
    elif verification["status"] == "error":
        lines.append("Verification: error — " + verification["error"])
    elif verification["status"] == "skipped":
        lines.append("Verification: skipped — " + verification["reason"])
    if report.get("sheet_filter_merged"):
        lines.append("Note: sheet-filtered session — unloaded sheets were merged back from disk.")
    if report.get("warnings"):
        lines.append("WARNINGS: " + "; ".join(report["warnings"]))
    return "\n".join(lines)


def _prepare_save_transaction(
    session_key: str,
    output_path: str | None,
    report_format: str,
    verify_preservation: bool,
    max_differences: int,
    operation: str,
) -> dict:
    normalized_format = str(report_format).strip().lower()
    if normalized_format not in {"text", "json"}:
        raise ValueError("report_format must be 'text' or 'json'.")
    if verify_preservation and not 1 <= int(max_differences) <= 5000:
        raise ValueError("max_differences must be between 1 and 5000.")

    resolved_session, data = _begin_session_operation(session_key, operation)
    staging: Path | None = None
    try:
        if output_path:
            _check_supported(output_path)
        dest = output_path or data.get("_default_output_path") or data["source"]
        _check_save_extension_compatible(data, dest)
        destination = Path(dest).expanduser().resolve()
        reference = verification_reference(data, destination)
        dirty_features = list(data.get("_dirty_features") or [])
        dirty_paths = list(data.get("_dirty_paths") or [])
        package_paths = package_edit_verifier_patterns(data, reference)
        if package_paths and "package" not in dirty_features:
            dirty_features.append("package")
        requested_paths = _save_verifier_patterns(dirty_paths, dirty_features)
        for package_path in package_paths:
            if package_path not in requested_paths:
                requested_paths.append(package_path)
        requested_semantic_paths = list(dirty_paths)
        for package_path in package_paths:
            if package_path not in requested_semantic_paths:
                requested_semantic_paths.append(package_path)
        staging = create_staging_path(destination)
        return {
            "session_key": resolved_session,
            "data": data,
            "destination": destination,
            "staging": staging,
            "destination_state": file_state(destination),
            "verification_reference": reference,
            "verification_reference_state": file_state(reference) if reference else None,
            "report_format": normalized_format,
            "verify_preservation": bool(verify_preservation),
            "max_differences": int(max_differences),
            "dirty_features": dirty_features,
            "dirty_paths": dirty_paths,
            "requested_semantic_paths": requested_semantic_paths,
            "requested_paths": requested_paths,
            "intentional_edit": bool(dirty_features or dirty_paths or package_paths),
        }
    except BaseException:
        remove_staging_path(staging)
        _end_session_operation(resolved_session)
        raise


def _save_stage_payload(transaction: dict) -> dict:
    return {
        "staging_path": str(transaction["staging"]),
        "verification_reference_path": transaction["verification_reference"],
        "verify_preservation": transaction["verify_preservation"],
        "max_differences": transaction["max_differences"],
        "requested_paths": transaction["requested_paths"],
        "intentional_edit": transaction["intentional_edit"],
    }


def _commit_save_transaction(transaction: dict, stage_result: dict) -> str:
    require_preservation_success(stage_result)
    destination = transaction["destination"]
    reference = transaction["verification_reference"]
    backup = commit_staging_file(
        staging_path=transaction["staging"],
        destination=destination,
        expected_destination_state=transaction["destination_state"],
        verification_reference_path=reference,
        expected_reference_state=transaction["verification_reference_state"],
    )

    verification = copy.deepcopy(stage_result["verification"])
    effective_reference = reference
    if backup and reference and Path(reference).resolve() == destination:
        effective_reference = backup["backup_path"]
        verification["reference_path"] = effective_reference

    data = transaction["data"]
    with _SESSION_LOCK:
        current = _sessions.get(transaction["session_key"])
        if current is not data:
            raise ExcelOperationError({
                "code": "EXCEL_SAVE_SESSION_REPLACED",
                "session_key": transaction["session_key"],
                "message": "The session changed before the save transaction committed.",
            })
        data["_verification_baseline_path"] = str(destination)
        data["_dirty_features"] = []
        data["_dirty_paths"] = []
        data["_new_workbook"] = False

    verification_call = f"excel_verify_preservation(after_path={str(destination)!r}"
    if effective_reference:
        verification_call += f", before_path={effective_reference!r}"
    if transaction["requested_paths"]:
        verification_call += f", requested_paths={transaction['requested_paths']!r}"
    verification_call += ")"

    report = {
        "schema_version": 1,
        "saved_path": str(destination),
        "source_path": data.get("source"),
        "file_size": destination.stat().st_size,
        "sheet_count": stage_result["sheet_count"],
        "total_rows": stage_result["total_rows"],
        "backup": backup,
        "package_signatures": stage_result["package_signatures"],
        "performance": stage_result.get("inspection_performance"),
        "warnings": stage_result["warnings"],
        "dirty_features": transaction["dirty_features"],
        "requested_semantic_paths": transaction["requested_semantic_paths"],
        "changed_semantic_paths": verification.get("changed_semantic_paths") or [],
        "verification": verification,
        "verification_call": verification_call,
        "sheet_filter_merged": stage_result["sheet_filter_merged"],
    }
    if transaction["report_format"] == "json":
        return json.dumps(report, default=str, ensure_ascii=False, indent=2)
    return _save_report_text(report)


def _finish_save_transaction(transaction: dict | None) -> bool:
    if transaction is None:
        return True
    removed = remove_staging_path(transaction.get("staging"))
    _end_session_operation(transaction["session_key"])
    return removed


def _excel_save_sync_impl(
    session_key: str,
    output_path: str | None = None,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
    operation: str = "excel_save",
) -> str:
    transaction = _prepare_save_transaction(
        session_key,
        output_path,
        report_format,
        verify_preservation,
        max_differences,
        operation,
    )
    try:
        stage_result = _save_stage_operation(
            transaction["data"],
            **_save_stage_payload(transaction),
        )
        return _commit_save_transaction(transaction, stage_result)
    finally:
        _finish_save_transaction(transaction)


async def _excel_save_async_impl(
    session_key: str,
    output_path: str | None = None,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
    timeout_seconds: float | None = None,
    operation: str = "excel_save",
) -> str:
    import anyio

    transaction = await anyio.to_thread.run_sync(
        lambda: _prepare_save_transaction(
            session_key,
            output_path,
            report_format,
            verify_preservation,
            max_differences,
            operation,
        )
    )
    try:
        stage_result = await run_named_operation(
            "save_stage",
            _save_stage_payload(transaction),
            timeout_seconds=timeout_seconds,
            operation_label=operation,
            input_data=transaction["data"],
            failure_cleanup_paths=(transaction["staging"],),
        )
        with anyio.CancelScope(shield=True):
            return await anyio.to_thread.run_sync(
                _commit_save_transaction,
                transaction,
                stage_result,
            )
    finally:
        with anyio.CancelScope(shield=True):
            removed = await anyio.to_thread.run_sync(
                _finish_save_transaction,
                transaction,
            )
            if not removed:
                raise ExcelOperationError({
                    "code": "EXCEL_SAVE_STAGING_CLEANUP_FAILED",
                    "staging_path": str(transaction["staging"]),
                    "message": "The save staging file could not be removed.",
                })


def excel_save(
    session_key: str,
    output_path: str | None = None,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
) -> str:
    """Synchronously execute the transactional save path for direct Python callers."""
    return _excel_save_sync_impl(
        session_key,
        output_path,
        report_format,
        verify_preservation,
        max_differences,
    )


def excel_save_as_copy(
    session_key: str,
    output_path: str,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
) -> str:
    """Save to another path with the same structured report and verifier options as excel_save."""
    if not output_path:
        raise ValueError("output_path is required for excel_save_as_copy.")
    source = Path(_get_session(session_key)["source"]).resolve()
    dest = Path(output_path).resolve()
    if source == dest:
        raise ValueError("excel_save_as_copy output_path must differ from the source. Use excel_save to overwrite.")
    return _excel_save_sync_impl(
        session_key,
        str(dest),
        operation="excel_save_as_copy",
        report_format=report_format,
        verify_preservation=verify_preservation,
        max_differences=max_differences,
    )


@mcp.tool(name="excel_save")
async def _excel_save_tool(
    session_key: str,
    output_path: str | None = None,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
    timeout_seconds: float | None = None,
) -> str:
    """Transactionally save a session through a cancellable worker.

    Reconstruction and optional preservation verification finish against a
    same-directory staging workbook before the destination is backed up and
    atomically replaced. Cancellation, timeout, worker failure, or failed
    verification leaves the destination unchanged.
    """
    return await _excel_save_async_impl(
        session_key,
        output_path,
        report_format,
        verify_preservation,
        max_differences,
        timeout_seconds,
        "excel_save",
    )


@mcp.tool(name="excel_save_as_copy")
async def _excel_save_as_copy_tool(
    session_key: str,
    output_path: str,
    report_format: str = "text",
    verify_preservation: bool = False,
    max_differences: int = 200,
    timeout_seconds: float | None = None,
) -> str:
    """Transactionally save a session to another path through a cancellable worker."""
    if not output_path:
        raise ValueError("output_path is required for excel_save_as_copy.")
    source = Path(_get_session(session_key)["source"]).resolve()
    destination = Path(output_path).resolve()
    if source == destination:
        raise ValueError(
            "excel_save_as_copy output_path must differ from the source. "
            "Use excel_save to overwrite."
        )
    return await _excel_save_async_impl(
        session_key,
        str(destination),
        report_format,
        verify_preservation,
        max_differences,
        timeout_seconds,
        "excel_save_as_copy",
    )


_VERIFICATION_CLASSIFICATIONS = (
    "REQUESTED",
    "APPROVED_NORMALIZATION",
    "UNAPPROVED_LOSS",
    "FIXTURE_GAP",
    "VERIFIER_GAP",
    "PACKAGE_INVALID",
)
_BLOCKING_VERIFICATION_CLASSIFICATIONS = (
    "UNAPPROVED_LOSS",
    "FIXTURE_GAP",
    "VERIFIER_GAP",
    "PACKAGE_INVALID",
)


def _verification_runtime_metadata() -> dict:
    repository = Path(__file__).resolve().parents[2]
    runtime_roots = []
    bundled_root = getattr(sys, "_MEIPASS", None)
    if bundled_root:
        runtime_roots.append(Path(bundled_root))
    runtime_roots.extend((repository, Path(__file__).resolve().parent))

    build_metadata = {}
    for root in runtime_roots:
        try:
            candidate = json.loads(
                (root / "excel-build-metadata.json").read_text(encoding="utf-8")
            )
        except Exception:
            continue
        if isinstance(candidate, dict):
            build_metadata = candidate
            break

    server_version = os.environ.get("DOCLOUPE_SERVER_VERSION") or build_metadata.get("server_version")
    if not server_version:
        for root in runtime_roots:
            try:
                server_version = json.loads(
                    (root / "server.json").read_text(encoding="utf-8")
                ).get("version")
            except Exception:
                continue
            if server_version:
                break

    commit_sha = os.environ.get("DOCLOUPE_COMMIT_SHA") or build_metadata.get("commit_sha")
    working_tree_dirty = build_metadata.get("working_tree_dirty")
    try:
        if not commit_sha:
            commit = subprocess.run(
                ["git", "rev-parse", "HEAD"],
                cwd=repository,
                capture_output=True,
                text=True,
                timeout=3,
                check=False,
            )
            if commit.returncode == 0:
                commit_sha = commit.stdout.strip() or None
        if working_tree_dirty is None:
            status = subprocess.run(
                ["git", "status", "--porcelain"],
                cwd=repository,
                capture_output=True,
                text=True,
                timeout=3,
                check=False,
            )
            if status.returncode == 0:
                working_tree_dirty = bool(status.stdout.strip())
    except Exception:
        pass

    embedded_libraries = build_metadata.get("library_versions") or {}
    libraries = {}
    for distribution in ("openpyxl", "mcp"):
        try:
            libraries[distribution] = importlib.metadata.version(distribution)
        except importlib.metadata.PackageNotFoundError:
            libraries[distribution] = embedded_libraries.get(distribution)
    return {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "server": "excel-tools",
        "server_version": server_version,
        "commit_sha": commit_sha,
        "working_tree_dirty": working_tree_dirty,
        "python_version": platform.python_version(),
        "platform": platform.platform(),
        "library_versions": libraries,
    }


def _package_invalid_verification_report(
    before_file: dict,
    after_file: dict,
    requested_paths: list[str] | None,
    approved_normalizations: list[str | dict] | None,
    fixture_gap_paths: list[str] | None,
    verifier_gap_paths: list[str] | None,
) -> dict:
    changes = []
    for role, metadata in (("before", before_file), ("after", after_file)):
        if metadata["package_valid"] and metadata["loadable"]:
            continue
        changes.append({
            "path": f"files/{role}",
            "category": "package_validity",
            "severity": "critical",
            "classification": "PACKAGE_INVALID",
            "before": None if role == "after" else metadata,
            "after": metadata if role == "after" else None,
        })
    classification_counts = {name: 0 for name in _VERIFICATION_CLASSIFICATIONS}
    classification_counts["PACKAGE_INVALID"] = len(changes)
    return {
        "schema_version": 2,
        "before_path": before_file["path"],
        "after_path": after_file["path"],
        "before_sha256": before_file["sha256"],
        "after_sha256": after_file["sha256"],
        "before_size": before_file["size"],
        "after_size": after_file["size"],
        "equivalent": False,
        "preservation_ok": False,
        "recommendation": "Repair the invalid or unloadable package before using it as preservation evidence.",
        "severity_counts": {"critical": len(changes), "high": 0, "medium": 0, "info": 0},
        "classification_counts": classification_counts,
        "requested_paths": list(requested_paths or []),
        "approved_normalizations": [
            item if isinstance(item, str) else item.get("path") or item.get("pattern")
            for item in (approved_normalizations or [])
        ],
        "approved_normalization_rules": copy.deepcopy(approved_normalizations or []),
        "normalization_evidence_complete": all(
            isinstance(item, dict)
            and bool(str(item.get("rationale") or "").strip())
            and item.get("bidirectional") is True
            and isinstance(item.get("evidence"), dict)
            and "before" in item["evidence"]
            and "after" in item["evidence"]
            for item in (approved_normalizations or [])
        ),
        "fixture_gap_paths": list(fixture_gap_paths or []),
        "verifier_gap_paths": list(verifier_gap_paths or []),
        "unapproved_difference_count": 0,
        "blocking_issue_count": len(changes),
        "part_diff": None,
        "change_count": len(changes),
        "changes": changes,
        "truncated": False,
    }


@mcp.tool()
def excel_validate_workbook(path: str) -> str:
    """Validate an .xlsx package and report advanced features found."""
    return json.dumps(inspect_xlsx_package(str(uri_to_path(path))), ensure_ascii=False, indent=2)


@mcp.tool()
def excel_diff_package(before_path: str, after_path: str) -> str:
    """Compare two .xlsx ZIP package manifests for save diagnostics."""
    return json.dumps(
        diff_xlsx_package(str(uri_to_path(before_path)), str(uri_to_path(after_path))),
        ensure_ascii=False,
        indent=2,
    )


def excel_verify_preservation(
    after_path: str,
    before_path: str | None = None,
    max_differences: int = 200,
    requested_paths: list[str] | None = None,
    approved_normalizations: list[str | dict] | None = None,
    fixture_gap_paths: list[str] | None = None,
    verifier_gap_paths: list[str] | None = None,
    fixture_id: str | None = None,
) -> str:
    """
    Compare a saved OOXML workbook with a reference and report semantic losses.

    If before_path is omitted, the newest unexpired pre-save backup associated
    with after_path is used automatically. The verifier checks values, formulas,
    rich-text runs, cell style semantics, sheet/workbook settings, print/filter/
    protection metadata, tables, document properties, and advanced OOXML parts.

    Args:
        after_path: Workbook produced by excel_save or another save workflow.
        before_path: Optional explicit reference workbook.
        max_differences: Maximum detailed changes returned (1-5000).
        approved_normalizations: Two-way normalization rules with path, rationale,
            bidirectional=true, and evidence.before/evidence.after. Legacy path-only
            entries are reported as VERIFIER_GAP when they match a change.
        fixture_gap_paths: Known differences caused by missing or invalid fixture data.
        verifier_gap_paths: Known differences the verifier cannot model correctly yet.
        fixture_id: Stable report identifier; defaults to the reference file stem.
    """
    tool_started_at = time.perf_counter()
    memory_sampler = _MemorySampler()
    memory_sampler.start()
    left_inspection = None
    right_inspection = None
    metadata_seconds = 0.0
    semantic_verification_seconds = 0.0
    try:
        cleanup_excel_backups()
        resolved_after = str(uri_to_path(after_path))
        backup = None
        if before_path:
            resolved_before = str(uri_to_path(before_path))
        else:
            backup = find_latest_excel_backup(resolved_after)
            if not backup:
                raise ValueError(
                    "No unexpired pre-save backup was found for after_path. "
                    "Pass before_path explicitly or save the workbook with excel_save first."
                )
            resolved_before = backup["backup_path"]

        metadata_started_at = time.perf_counter()
        left_inspection, right_inspection = inspect_workbook_pair(
            resolved_before,
            resolved_after,
            check_loadable=True,
        )
        metadata_seconds = time.perf_counter() - metadata_started_at
        before_file = left_inspection.file_metadata()
        after_file = right_inspection.file_metadata()
        if all(
            item["package_valid"] and item["loadable"]
            for item in (before_file, after_file)
        ):
            semantic_started_at = time.perf_counter()
            report = verify_xlsx_preservation(
                resolved_before,
                resolved_after,
                max_differences,
                requested_paths=requested_paths,
                approved_normalizations=approved_normalizations,
                fixture_gap_paths=fixture_gap_paths,
                verifier_gap_paths=verifier_gap_paths,
                before_inspection=left_inspection,
                after_inspection=right_inspection,
            )
            semantic_verification_seconds = time.perf_counter() - semantic_started_at
        else:
            report = _package_invalid_verification_report(
                before_file,
                after_file,
                requested_paths,
                approved_normalizations,
                fixture_gap_paths,
                verifier_gap_paths,
            )
        report["fixture_id"] = (
            str(fixture_id).strip()
            if fixture_id and str(fixture_id).strip()
            else Path(resolved_after).stem
        )
        report["fixture_id_source"] = (
            "provided" if fixture_id and str(fixture_id).strip() else "after_filename"
        )
        report["files"] = {"before": before_file, "after": after_file}
        report["runtime"] = _verification_runtime_metadata()
        report["backup"] = (
            {
                key: backup.get(key)
                for key in (
                    "backup_path",
                    "reference_path",
                    "saved_path",
                    "created_at",
                    "expires_at",
                    "retention_days",
                )
            }
            if backup
            else None
        )
    finally:
        if left_inspection is not None:
            left_inspection.release_raw_parts()
        if right_inspection is not None:
            right_inspection.release_raw_parts()
        memory_metrics = memory_sampler.finish()

    report["performance"] = inspection_pair_performance(
        left_inspection,
        right_inspection,
        metadata_seconds=metadata_seconds,
        semantic_verification_seconds=semantic_verification_seconds,
        total_tool_seconds=time.perf_counter() - tool_started_at,
        memory=memory_metrics,
    )
    return json.dumps(report, ensure_ascii=False, indent=2)


@mcp.tool(name="excel_verify_preservation")
async def _excel_verify_preservation_tool(
    after_path: str,
    before_path: str | None = None,
    max_differences: int = 200,
    requested_paths: list[str] | None = None,
    approved_normalizations: list[str | dict] | None = None,
    fixture_gap_paths: list[str] | None = None,
    verifier_gap_paths: list[str] | None = None,
    fixture_id: str | None = None,
    timeout_seconds: float | None = None,
) -> str:
    """
    Compare two workbooks in a spawned, cancellable worker process.

    The synchronous `excel_verify_preservation` function remains available for
    direct Python tests. MCP calls use this wrapper so timeout or host-task
    cancellation stops and joins the worker before returning an error.

    Args:
        after_path: Workbook produced by excel_save or another save workflow.
        before_path: Optional explicit reference workbook; otherwise use the
            newest unexpired pre-save backup for after_path.
        max_differences: Maximum detailed changes returned (1-5000).
        requested_paths: Semantic paths intentionally changed by the caller.
        approved_normalizations: Fully evidenced two-way normalization rules.
        fixture_gap_paths: Known differences caused by fixture defects.
        verifier_gap_paths: Known differences the verifier cannot model yet.
        fixture_id: Stable report identifier.
        timeout_seconds: Optional per-call timeout; otherwise use the configured
            verify or global heavy-operation timeout.
    """
    return await run_named_operation(
        "verify_preservation",
        {
            "after_path": after_path,
            "before_path": before_path,
            "max_differences": max_differences,
            "requested_paths": requested_paths,
            "approved_normalizations": approved_normalizations,
            "fixture_gap_paths": fixture_gap_paths,
            "verifier_gap_paths": verifier_gap_paths,
            "fixture_id": fixture_id,
        },
        timeout_seconds=timeout_seconds,
        operation_label="excel_verify_preservation",
    )


def _summary_count(record: dict, keys: tuple[str, ...], label: str) -> int | None:
    value = next((record[key] for key in keys if record.get(key) is not None), None)
    if value is None:
        return None
    if not isinstance(value, int) or isinstance(value, bool) or value < 0:
        raise ValueError(f"{label} must be a non-negative integer.")
    return value


def _summary_item_id(item: object, index: int) -> str:
    if isinstance(item, dict):
        value = item.get("key") or item.get("name") or item.get("fixture_id")
        if value is not None and str(value).strip():
            return str(value).strip()
    return f"item-{index + 1}"


def _coverage_state(checked: int | None, passed: int | None) -> dict:
    if checked is None or passed is None:
        return {
            "checked_count": 0,
            "passed_count": 0,
            "failed_count": 0,
            "complete": None,
        }
    if passed > checked:
        raise ValueError("Evidence-state passed_count cannot exceed checked_count.")
    return {
        "checked_count": checked,
        "passed_count": passed,
        "failed_count": checked - passed,
        "complete": passed == checked,
    }


def _normalize_coverage_evidence(record: dict, index: int) -> dict:
    if not isinstance(record, dict):
        raise ValueError(f"coverage_reports[{index}] must be an object.")
    coverage_id = str(
        record.get("coverage_id")
        or record.get("id")
        or record.get("name")
        or f"coverage-{index + 1}"
    ).strip()
    if not coverage_id:
        raise ValueError(f"coverage_reports[{index}] requires a non-empty coverage_id.")
    results = record.get("results")
    fixtures = record.get("fixtures")
    if results is not None and not isinstance(results, list):
        raise ValueError(f"coverage_reports[{index}].results must be a list.")
    if fixtures is not None and not isinstance(fixtures, list):
        raise ValueError(f"coverage_reports[{index}].fixtures must be a list.")
    explicit_mode = str(record.get("coverage_mode") or record.get("mode") or "").strip().upper()
    if explicit_mode and explicit_mode not in {"REPRODUCTION", "POST_FIX_PRESERVATION", "GENERAL"}:
        raise ValueError(
            f"coverage_reports[{index}].coverage_mode must be REPRODUCTION, "
            "POST_FIX_PRESERVATION, or GENERAL."
        )
    coverage_mode = explicit_mode or ("POST_FIX_PRESERVATION" if fixtures is not None else "REPRODUCTION")
    checked = _summary_count(
        record,
        (
            "checked_count",
            "checked_key_count",
            "checked_case_count",
            "tested_key_count",
            "case_count",
        ),
        f"coverage_reports[{index}].checked_count",
    )
    passed = _summary_count(
        record,
        (
            "passed_count",
            "preserved_key_count",
            "preserved_case_count",
            "fully_reproduced_count",
            "reproduced_count",
        ),
        f"coverage_reports[{index}].passed_count",
    )
    failed_items = copy.deepcopy(
        record.get("failed_items")
        or record.get("failed_keys")
        or record.get("failed_cases")
        or record.get("failed")
        or []
    )
    if not isinstance(failed_items, list):
        raise ValueError(f"coverage_reports[{index}].failed_items must be a list.")
    failed = _summary_count(
        record,
        ("failed_count", "failed_key_count", "failed_case_count"),
        f"coverage_reports[{index}].failed_count",
    )
    if checked is None:
        checked = len(fixtures if fixtures is not None else results or []) or None
    if passed is None and coverage_mode == "POST_FIX_PRESERVATION" and fixtures is not None:
        passed = sum(item.get("preservation_ok") is True for item in fixtures if isinstance(item, dict))
    if passed is None and results is not None:
        passed = sum(item.get("reproduced") is True for item in results if isinstance(item, dict))
    if not failed_items and checked is not None and passed is not None and passed < checked:
        source_items = fixtures if fixtures is not None else results or []
        failed_items = [
            _summary_item_id(item, item_index)
            for item_index, item in enumerate(source_items)
            if not isinstance(item, dict)
            or not (
                item.get("preservation_ok") is True
                if coverage_mode == "POST_FIX_PRESERVATION"
                else item.get("reproduced") is True
            )
        ]
    if checked is None:
        raise ValueError(f"coverage_reports[{index}] requires a checked count.")
    if passed is None:
        if failed is None:
            raise ValueError(f"coverage_reports[{index}] requires a passed or failed count.")
        passed = checked - failed
    if failed is None:
        failed = checked - passed
    if passed < 0 or failed < 0 or passed + failed != checked:
        raise ValueError(
            f"coverage_reports[{index}] counts must satisfy passed + failed == checked."
        )
    if len(failed_items) > failed:
        raise ValueError(
            f"coverage_reports[{index}] has more failed items than failed_count."
        )
    source_present = _summary_count(
        record,
        ("source_semantic_present_count", "source_before_ok_count"),
        f"coverage_reports[{index}].source_semantic_present_count",
    )
    if source_present is None and coverage_mode == "REPRODUCTION":
        source_present = (
            sum(item.get("source_ok") is True for item in results if isinstance(item, dict))
            if results is not None
            else checked
        )
    legacy_reproduced = _summary_count(
        record,
        ("legacy_bug_reproduced_count", "fully_reproduced_count", "reproduced_count"),
        f"coverage_reports[{index}].legacy_bug_reproduced_count",
    )
    if legacy_reproduced is None and coverage_mode == "REPRODUCTION":
        legacy_reproduced = passed
    fixed_preserved = _summary_count(
        record,
        ("fixed_output_preserved_count",),
        f"coverage_reports[{index}].fixed_output_preserved_count",
    )
    if fixed_preserved is None and coverage_mode == "POST_FIX_PRESERVATION":
        fixed_preserved = passed

    fixture_graph_checked = 0
    fixture_graph_valid = 0
    explicit_graph_valid = record.get("fixture_graph_valid")
    if explicit_graph_valid is not None:
        if not isinstance(explicit_graph_valid, bool):
            raise ValueError(f"coverage_reports[{index}].fixture_graph_valid must be true or false.")
        fixture_graph_checked = checked
        fixture_graph_valid = checked if explicit_graph_valid else 0
    else:
        for item in fixtures or results or []:
            if not isinstance(item, dict):
                continue
            value = item.get("fixture_graph_valid")
            if value is None:
                value = item.get("package_valid")
            if isinstance(value, bool):
                fixture_graph_checked += 1
                fixture_graph_valid += value

    fixture_gap_items = copy.deepcopy(record.get("fixture_gap_items") or [])
    verifier_gap_items = copy.deepcopy(record.get("verifier_gap_items") or [])
    if not isinstance(fixture_gap_items, list) or not isinstance(verifier_gap_items, list):
        raise ValueError(
            f"coverage_reports[{index}].fixture_gap_items and verifier_gap_items must be lists."
        )
    for item_index, item in enumerate(results or []):
        if not isinstance(item, dict):
            continue
        item_id = _summary_item_id(item, item_index)
        classification = str(item.get("classification") or "").upper()
        if item.get("source_ok") is False or classification == "FIXTURE_GAP":
            if item_id not in fixture_gap_items:
                fixture_gap_items.append(item_id)
        if item.get("verifier_gap") is True or classification == "VERIFIER_GAP":
            if item_id not in verifier_gap_items:
                verifier_gap_items.append(item_id)
    unclassified_failures = [
        item for item in failed_items
        if item not in fixture_gap_items and item not in verifier_gap_items
    ]
    evidence_states = {
        "source_semantic_present": _coverage_state(
            checked if source_present is not None else None,
            source_present,
        ),
        "legacy_bug_reproduced": _coverage_state(
            checked if legacy_reproduced is not None else None,
            legacy_reproduced,
        ),
        "fixed_output_preserved": _coverage_state(
            checked if fixed_preserved is not None else None,
            fixed_preserved,
        ),
        "fixture_graph_valid": _coverage_state(
            fixture_graph_checked if fixture_graph_checked else None,
            fixture_graph_valid if fixture_graph_checked else None,
        ),
    }
    required_states = {
        "REPRODUCTION": ("source_semantic_present", "legacy_bug_reproduced", "fixture_graph_valid"),
        "POST_FIX_PRESERVATION": ("fixed_output_preserved", "fixture_graph_valid"),
        "GENERAL": (),
    }[coverage_mode]
    complete = (
        failed == 0
        and passed == checked
        and all(evidence_states[name]["complete"] is True for name in required_states)
    )

    return {
        "coverage_id": coverage_id,
        "path": record.get("path"),
        "coverage_mode": coverage_mode,
        "checked_count": checked,
        "passed_count": passed,
        "failed_count": failed,
        "failed_items": failed_items,
        "failure_classifications": {
            "FIXTURE_GAP": fixture_gap_items,
            "VERIFIER_GAP": verifier_gap_items,
            "NOT_REPRODUCED": unclassified_failures,
        },
        "evidence_states": evidence_states,
        "complete": complete,
    }


def _verification_fixture_graph_valid(record: dict, index: int) -> bool | None:
    explicit = record.get("fixture_graph_valid")
    if explicit is not None:
        if not isinstance(explicit, bool):
            raise ValueError(
                f"verification_reports[{index}].fixture_graph_valid must be true or false."
            )
        return explicit
    files = record.get("files")
    if not isinstance(files, dict):
        return None
    statuses = []
    for role in ("before", "after"):
        metadata = files.get(role)
        if not isinstance(metadata, dict):
            return None
        for key in ("package_valid", "loadable"):
            value = metadata.get(key)
            if not isinstance(value, bool):
                return None
            statuses.append(value)
    return all(statuses)


def _normalize_verification_evidence(record: dict, index: int, max_changes: int) -> dict:
    if not isinstance(record, dict):
        raise ValueError(f"verification_reports[{index}] must be an object.")
    counts = record.get("classification_counts") or {}
    if not isinstance(counts, dict):
        raise ValueError(
            f"verification_reports[{index}].classification_counts must be an object."
        )
    normalized_counts = {}
    for classification in _VERIFICATION_CLASSIFICATIONS:
        value = counts.get(classification, 0)
        if not isinstance(value, int) or isinstance(value, bool) or value < 0:
            raise ValueError(
                f"verification_reports[{index}].classification_counts.{classification} "
                "must be a non-negative integer."
            )
        normalized_counts[classification] = value
    changes = copy.deepcopy(record.get("changes") or [])
    if not isinstance(changes, list):
        raise ValueError(f"verification_reports[{index}].changes must be a list.")
    blocking_issue_count = _summary_count(
        record,
        ("blocking_issue_count",),
        f"verification_reports[{index}].blocking_issue_count",
    )
    if blocking_issue_count is None:
        blocking_issue_count = sum(
            normalized_counts[classification]
            for classification in _BLOCKING_VERIFICATION_CLASSIFICATIONS
        )
    change_count = _summary_count(
        record,
        ("change_count",),
        f"verification_reports[{index}].change_count",
    )
    if change_count is None:
        change_count = len(changes)
    if change_count < len(changes):
        raise ValueError(
            f"verification_reports[{index}].change_count cannot be smaller than changes length."
        )
    returned_changes = changes[:max_changes]
    fixture_id = str(
        record.get("fixture_id") or record.get("after_path") or f"verification-{index + 1}"
    ).strip()
    if not fixture_id:
        fixture_id = f"verification-{index + 1}"
    fixture_graph_valid = _verification_fixture_graph_valid(record, index)
    has_blocking_classification = any(
        normalized_counts[classification] > 0
        for classification in _BLOCKING_VERIFICATION_CLASSIFICATIONS
    )
    fixed_output_preserved = (
        record.get("preservation_ok") is True
        and blocking_issue_count == 0
        and not has_blocking_classification
    )
    return {
        "fixture_id": fixture_id,
        "before_path": record.get("before_path"),
        "after_path": record.get("after_path"),
        "files": copy.deepcopy(record.get("files")),
        "backup": copy.deepcopy(record.get("backup")),
        "preservation_ok": record.get("preservation_ok") is True,
        "fixed_output_preserved": fixed_output_preserved,
        "fixture_graph_valid": fixture_graph_valid,
        "classification_counts": normalized_counts,
        "unapproved_difference_count": normalized_counts["UNAPPROVED_LOSS"],
        "invalid_package_count": normalized_counts["PACKAGE_INVALID"],
        "blocking_issue_count": blocking_issue_count,
        "change_count": change_count,
        "changes": returned_changes,
        "changes_truncated": len(changes) > len(returned_changes),
        "part_diff": copy.deepcopy(record.get("part_diff")),
    }


def _normalize_backup_evidence(record: dict, index: int) -> dict:
    if not isinstance(record, dict):
        raise ValueError(f"backup_checks[{index}] must be an object.")
    name = str(record.get("check_id") or record.get("name") or f"backup-check-{index + 1}").strip()
    if not name:
        raise ValueError(f"backup_checks[{index}] requires a non-empty name.")
    if not isinstance(record.get("passed"), bool):
        raise ValueError(f"backup_checks[{index}].passed must be true or false.")
    return {
        "check_id": name,
        "passed": record["passed"],
        "details": copy.deepcopy(record.get("details") or {}),
    }


def _aggregate_evidence_states(states: list[dict]) -> dict:
    checked = sum(state["checked_count"] for state in states)
    passed = sum(state["passed_count"] for state in states)
    return {
        "checked_count": checked,
        "passed_count": passed,
        "failed_count": checked - passed,
        "complete": checked > 0 and passed == checked,
    }


@mcp.tool()
def excel_build_preservation_summary(
    coverage_reports: list[dict],
    verification_reports: list[dict],
    backup_checks: list[dict],
    evidence_id: str = "excel-preservation-final",
    output_path: str | None = None,
    overwrite: bool = False,
    max_changes_per_report: int = 20,
) -> str:
    """Aggregate coverage, file-verification, and backup evidence into one JSON report."""
    if not isinstance(max_changes_per_report, int) or isinstance(max_changes_per_report, bool):
        raise ValueError("max_changes_per_report must be an integer from 0 to 200.")
    if not 0 <= max_changes_per_report <= 200:
        raise ValueError("max_changes_per_report must be an integer from 0 to 200.")
    if not coverage_reports:
        raise ValueError("coverage_reports must contain at least one report.")
    if not verification_reports:
        raise ValueError("verification_reports must contain at least one report.")
    if not backup_checks:
        raise ValueError("backup_checks must contain at least one check.")

    coverage = [
        _normalize_coverage_evidence(record, index)
        for index, record in enumerate(coverage_reports)
    ]
    verifications = [
        _normalize_verification_evidence(record, index, max_changes_per_report)
        for index, record in enumerate(verification_reports)
    ]
    backups = [
        _normalize_backup_evidence(record, index)
        for index, record in enumerate(backup_checks)
    ]

    classification_counts = {name: 0 for name in _VERIFICATION_CLASSIFICATIONS}
    for report in verifications:
        for name, value in report["classification_counts"].items():
            classification_counts[name] += value
    source_states = [
        item["evidence_states"]["source_semantic_present"]
        for item in coverage
        if item["evidence_states"]["source_semantic_present"]["checked_count"]
    ]
    reproduction_states = [
        item["evidence_states"]["legacy_bug_reproduced"]
        for item in coverage
        if item["evidence_states"]["legacy_bug_reproduced"]["checked_count"]
    ]
    reproduction_graph_states = [
        item["evidence_states"]["fixture_graph_valid"]
        for item in coverage
        if item["coverage_mode"] == "REPRODUCTION"
        and item["evidence_states"]["fixture_graph_valid"]["checked_count"]
    ]
    fixed_output_state = {
        "checked_count": len(verifications),
        "passed_count": sum(item["fixed_output_preserved"] for item in verifications),
        "failed_count": sum(not item["fixed_output_preserved"] for item in verifications),
        "complete": all(item["fixed_output_preserved"] for item in verifications),
    }
    verification_graph_state = {
        "checked_count": len(verifications),
        "passed_count": sum(item["fixture_graph_valid"] is True for item in verifications),
        "failed_count": sum(item["fixture_graph_valid"] is not True for item in verifications),
        "unknown_count": sum(item["fixture_graph_valid"] is None for item in verifications),
        "complete": all(item["fixture_graph_valid"] is True for item in verifications),
    }
    fixture_graph_state = _aggregate_evidence_states(
        reproduction_graph_states + [verification_graph_state]
    )
    fixture_graph_state["unknown_count"] = verification_graph_state["unknown_count"]
    evidence_states = {
        "source_semantic_present": _aggregate_evidence_states(source_states),
        "legacy_bug_reproduced": _aggregate_evidence_states(reproduction_states),
        "fixed_output_preserved": fixed_output_state,
        "fixture_graph_valid": fixture_graph_state,
    }
    failed_coverage = [item["coverage_id"] for item in coverage if not item["complete"]]
    failed_verifications = [
        item["fixture_id"]
        for item in verifications
        if not item["fixed_output_preserved"]
        or item["fixture_graph_valid"] is not True
        or item["blocking_issue_count"] > 0
        or any(
            item["classification_counts"][classification] > 0
            for classification in _BLOCKING_VERIFICATION_CLASSIFICATIONS
        )
    ]
    failed_backups = [item["check_id"] for item in backups if not item["passed"]]
    final_gate_passed = not (failed_coverage or failed_verifications or failed_backups)
    resolved_output = None
    if output_path:
        resolved_output = Path(uri_to_path(output_path)).expanduser().resolve()
        if resolved_output.exists() and not overwrite:
            raise FileExistsError(
                f"Summary output already exists: {resolved_output}. Pass overwrite=true to replace it."
            )

    summary = {
        "schema_version": 2,
        "evidence_id": str(evidence_id).strip() or "excel-preservation-final",
        "runtime": _verification_runtime_metadata(),
        "output_path": str(resolved_output) if resolved_output else None,
        "evidence_states": evidence_states,
        "coverage": {
            "reports": coverage,
            "report_count": len(coverage),
            "checked_count": sum(item["checked_count"] for item in coverage),
            "passed_count": sum(item["passed_count"] for item in coverage),
            "failed_count": sum(item["failed_count"] for item in coverage),
            "failed_coverage_ids": failed_coverage,
        },
        "verification": {
            "reports": verifications,
            "report_count": len(verifications),
            "preservation_ok_count": sum(item["preservation_ok"] for item in verifications),
            "fixed_output_preserved_count": fixed_output_state["passed_count"],
            "fixture_graph_valid_count": verification_graph_state["passed_count"],
            "fixture_graph_unknown_count": verification_graph_state["unknown_count"],
            "classification_counts": classification_counts,
            "unapproved_difference_count": classification_counts["UNAPPROVED_LOSS"],
            "invalid_package_count": classification_counts["PACKAGE_INVALID"],
            "fixture_gap_count": classification_counts["FIXTURE_GAP"],
            "verifier_gap_count": classification_counts["VERIFIER_GAP"],
            "failed_fixture_ids": failed_verifications,
        },
        "backup_retention": {
            "checks": backups,
            "check_count": len(backups),
            "passed_check_count": sum(item["passed"] for item in backups),
            "failed_check_ids": failed_backups,
        },
        "final_gate_passed": final_gate_passed,
        "failed_gate_reasons": {
            "coverage": failed_coverage,
            "verification": failed_verifications,
            "backup_retention": failed_backups,
        },
    }
    payload = json.dumps(summary, ensure_ascii=False, indent=2)
    if resolved_output:
        resolved_output.parent.mkdir(parents=True, exist_ok=True)
        resolved_output.write_text(payload + "\n", encoding="utf-8")
    return payload


# ── 4. Reload / Close ─────────────────────────────────────────────────────────

@mcp.tool()
def excel_reload(session_key: str) -> str:
    """
    Reload session data from disk, discarding any unsaved in-memory changes.

    Args:
        session_key: Key returned by excel_load (must still point to a valid file)

    Returns:
        Same summary as excel_load
    """
    session_key = _resolve_session_key(session_key)
    with _SESSION_LOCK:
        if session_key in _BUSY_SESSIONS:
            raise _session_busy_error(session_key)
    sheet_filter = _sessions[session_key].get("_sheet_filter")
    data = serialize_excel(session_key, sheet_filter)
    data["_sheet_filter"] = sheet_filter
    data["_loaded_disk_names"] = [s["name"] for s in data["sheets"]]
    data["source"] = session_key
    _store_session(session_key, data)
    sheet_names = [s["name"] for s in data["sheets"]]
    total_rows = sum(len(s["rows"]) for s in data["sheets"])
    return f"Reloaded: session_key={session_key!r} | sheets={sheet_names} | total_rows={total_rows}"


@mcp.tool()
def excel_close(session_key: str) -> str:
    """
    Remove a session from the server cache to free memory.

    Args:
        session_key: Key returned by excel_load

    Returns:
        Confirmation
    """
    session_key = _resolve_session_key(session_key)
    with _SESSION_LOCK:
        if session_key in _BUSY_SESSIONS:
            raise _session_busy_error(session_key)
        del _sessions[session_key]
    return f"Closed session '{session_key}'."


# ── 5. To Markdown ────────────────────────────────────────────────────────────

@mcp.tool()
def excel_to_markdown(session_key: str, sheet_name: str | None = None, max_rows: int | None = None, max_cols: int | None = None) -> TextContent:
    """
    Export session data as Markdown tables annotated with 0-based row/column indices.

    Column headers show col_N (header text) using row 0 as the header row.
    When a sheet contains merged cells a "merge" column is inserted showing RxC
    for origin cells. Slave cells are shown as blank.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Sheet to export; omit to export all sheets
        max_rows: Optional maximum rows per exported sheet
        max_cols: Optional maximum columns per exported sheet

    Returns:
        Markdown content (text/markdown) — one table per sheet
    """
    data = _limit_workbook_data(_get_session(session_key), max_rows=max_rows, max_cols=max_cols)
    from excel_converter import convert_excel_to_markdown
    markdown = convert_excel_to_markdown(data, sheet_name=sheet_name)
    return TextContent(type="text", text=markdown, mimeType="text/markdown")


# ── LibreOffice capture ───────────────────────────────────────────────────────

def _find_soffice(hint: str | None = None) -> str:
    import shutil
    if hint:
        return hint
    found = shutil.which("soffice")
    if found:
        return found
    from pathlib import Path
    for candidate in (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if Path(candidate).exists():
            return candidate
    raise FileNotFoundError(
        "LibreOffice (soffice) not found. Install LibreOffice or pass soffice_path explicitly."
    )


def _excel_capture_impl(
    session_key: str,
    sheet_name: str,
    output_path: str,
    soffice_path: str | None,
    timeout_seconds: float,
    cancel_event: threading.Event | None,
) -> str:
    import shutil

    lo = _find_soffice(soffice_path)
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    command_label = (lo, "--headless", "--convert-to", "png")
    if cancel_event is not None and cancel_event.is_set():
        raise ManagedProcessCancelled(command_label, process_tree_stopped=True)

    tmp_dir = Path(tempfile.mkdtemp(prefix="docloupe-excel-capture-"))
    output_stage: Path | None = None
    try:
        tmp_xlsx = tmp_dir / "capture.xlsx"
        reconstruct_excel({"source": "", "sheets": [sheet]}, str(tmp_xlsx))
        profile = tmp_dir / "lo_profile"
        command = [
            lo,
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            f"-env:UserInstallation=file:///{profile.as_posix()}",
            "--convert-to",
            "png",
            "--outdir",
            str(tmp_dir),
            str(tmp_xlsx),
        ]
        result = run_managed_process(
            command,
            timeout_seconds=timeout_seconds,
            cancel_event=cancel_event,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if cancel_event is not None and cancel_event.is_set():
            raise ManagedProcessCancelled(
                command,
                process_tree_stopped=result.process_tree_stopped,
                stdout=result.stdout,
                stderr=result.stderr,
            )

        pngs = sorted(tmp_dir.glob("*.png"))
        if result.returncode != 0 or not pngs:
            detail = f"{result.stdout or ''}{result.stderr or ''}".strip()
            raise RuntimeError(
                f"LibreOffice produced no PNG (exit_code={result.returncode}). {detail}"
            )

        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        descriptor, stage_name = tempfile.mkstemp(
            prefix=f".{output.name}.docloupe-",
            suffix=".tmp",
            dir=output.parent,
        )
        os.close(descriptor)
        output_stage = Path(stage_name)
        shutil.copyfile(pngs[0], output_stage)

        with output_stage.open("rb") as stream:
            header = stream.read(24)
        if len(header) < 24 or header[:8] != b"\x89PNG\r\n\x1a\n":
            raise RuntimeError("LibreOffice produced an invalid PNG file.")
        width = int.from_bytes(header[16:20], "big")
        height = int.from_bytes(header[20:24], "big")
        if cancel_event is not None and cancel_event.is_set():
            raise ManagedProcessCancelled(command, process_tree_stopped=True)

        os.replace(output_stage, output)
        output_stage = None
        return f"Saved {width}×{height}px → {output}"
    finally:
        if output_stage is not None:
            output_stage.unlink(missing_ok=True)
        if not remove_path_with_retries(tmp_dir):
            raise RuntimeError(f"Could not remove LibreOffice workspace: {tmp_dir}")


def excel_capture(
    session_key: str,
    sheet_name: str,
    output_path: str,
    soffice_path: str | None = None,
    timeout_seconds: float = 120.0,
) -> str:
    """Render a sheet as a PNG through a bounded LibreOffice process tree."""
    return _excel_capture_impl(
        session_key,
        sheet_name,
        output_path,
        soffice_path,
        timeout_seconds,
        None,
    )


@mcp.tool(name="excel_capture")
async def _excel_capture_tool(
    session_key: str,
    sheet_name: str,
    output_path: str,
    soffice_path: str | None = None,
    timeout_seconds: float = 120.0,
) -> str:
    """Render a sheet as PNG; cancellation stops LibreOffice and its descendants."""
    return await run_cancellable_in_thread(
        lambda cancel_event: _excel_capture_impl(
            session_key,
            sheet_name,
            output_path,
            soffice_path,
            timeout_seconds,
            cancel_event,
        )
    )


@mcp.tool()
def excel_extract_images(
    session_key: str,
    sheet_name: str,
    output_dir: str,
) -> str:
    """
    Extract all embedded images from a sheet and save them to a directory.

    Images are read directly from the source file on disk (not from in-memory
    edits). Call excel_save first if you want the saved state to be reflected.

    Args:
        session_key: Key returned by excel_load (used as the source file path)
        sheet_name: Name of the sheet to extract images from
        output_dir: Directory to save images into (created if it does not exist)

    Returns:
        List of saved image paths with their anchor cell references
    """
    import openpyxl
    import openpyxl.utils
    from pathlib import Path

    def _ext(data: bytes) -> str:
        if data[:8] == b'\x89PNG\r\n\x1a\n':
            return 'png'
        if data[:2] == b'\xff\xd8':
            return 'jpg'
        if data[:4] == b'GIF8':
            return 'gif'
        if data[:4] == b'RIFF' and data[8:12] == b'WEBP':
            return 'webp'
        return 'bin'

    def _anchor_cell(img) -> str:
        a = img.anchor
        if isinstance(a, str):
            return a
        try:
            fr = a._from
            cell = f"{openpyxl.utils.get_column_letter(fr.col + 1)}{fr.row + 1}"
            if hasattr(a, 'to'):
                t = a.to
                cell += f":{openpyxl.utils.get_column_letter(t.col + 1)}{t.row + 1}"
            return cell
        except Exception:
            return str(a)

    _get_session(session_key)  # validate session exists
    wb = openpyxl.load_workbook(session_key)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    images = ws._images
    if not images:
        return f"No images found in sheet '{sheet_name}'."

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    saved = []
    for i, img in enumerate(images, 1):
        data = img._data()
        ext = _ext(data)
        name = f"image_{i:02d}.{ext}"
        path = out / name
        path.write_bytes(data)
        saved.append(f"{path}  [{_anchor_cell(img)}]")

    lines = [f"Extracted {len(saved)} image(s) from sheet '{sheet_name}':"] + saved
    return "\n".join(lines)


# ── 6. Get rows ───────────────────────────────────────────────────────────────

@mcp.tool()
def excel_get_rows(
    session_key: str,
    sheet_name: str,
    start_row: int = 0,
    end_row: int | None = None,
    values_only: bool = False,
    include_rich_text: bool = False,
    include_formula_cache: bool = False,
    include_semantics: bool = False,
) -> str:
    """Read session rows with optional lossless cell semantics."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"][start_row:end_row]
    semantic_mode = include_rich_text or include_formula_cache or include_semantics
    if values_only and not semantic_mode:
        result = [
            [cd.get("v") if cd.get("merge") != "slave" else None for cd in row.get("cells", [])]
            for row in rows
        ]
    else:
        result = []
        for offset, row in enumerate(rows):
            row_index = start_row + offset
            public_row = _strip_private({key: value for key, value in row.items() if key != "cells"})
            public_row["row_index"] = row_index
            public_row["cells"] = [
                _cell_public_view(
                    cell,
                    sheet=sheet,
                    coord=_cell_coord(row_index, col_index),
                    include_rich_text=include_rich_text,
                    include_formula_cache=include_formula_cache,
                    include_semantics=include_semantics,
                )
                for col_index, cell in enumerate(row.get("cells", []))
            ]
            result.append(public_row)
    return json.dumps(result, default=str, ensure_ascii=False)


@mcp.tool()
def excel_read_range(
    session_key: str,
    sheet_name: str,
    range_ref: str | None = None,
    start_row: int | None = None,
    end_row: int | None = None,
    start_col: int | None = None,
    end_col: int | None = None,
    values_only: bool = True,
    include_rich_text: bool = False,
    include_formula_cache: bool = False,
    include_semantics: bool = False,
) -> str:
    """Read an exact rectangle with optional rich-text/formula/style semantics."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet.get("rows", [])
    max_row = len(rows)
    max_col = max((len(row.get("cells", [])) for row in rows), default=0)
    if range_ref:
        requested = _excel_range_to_indices(range_ref)
        max_row = max(max_row, requested[1] + 1)
        max_col = max(max_col, requested[3] + 1)
    if max_row == 0 or max_col == 0:
        return json.dumps({"sheet_name": sheet_name, "range": None, "values": []}, ensure_ascii=False)
    r1, r2, c1, c2 = _range_from_args(range_ref, start_row, end_row, start_col, end_col, max_row, max_col)
    semantic_mode = include_rich_text or include_formula_cache or include_semantics
    values = []
    for row_index in range(r1, r2 + 1):
        cells = rows[row_index].get("cells", []) if row_index < len(rows) else []
        row_values = []
        for col_index in range(c1, c2 + 1):
            cell = cells[col_index] if col_index < len(cells) else None
            coord = _cell_coord(row_index, col_index)
            if values_only and not semantic_mode:
                row_values.append(None if not cell or cell.get("merge") == "slave" else cell.get("v"))
            else:
                row_values.append(_cell_public_view(
                    cell,
                    sheet=sheet,
                    coord=coord,
                    include_rich_text=include_rich_text,
                    include_formula_cache=include_formula_cache,
                    include_semantics=include_semantics,
                ))
        values.append(row_values)
    result = {
        "sheet_name": sheet_name,
        "range": {"start_row": r1, "end_row": r2, "start_col": c1, "end_col": c2},
        "values_only": values_only and not semantic_mode,
        "values": values,
    }
    return json.dumps(result, default=str, ensure_ascii=False)

@mcp.tool()
def excel_find_cells(
    session_key: str,
    query: str,
    sheet_name: str | None = None,
    regex: bool = False,
    case_sensitive: bool = False,
    match_in: str = "value",
    max_results: int = 100,
) -> str:
    """Find cells by literal text or regex across one sheet or the whole workbook."""
    data = _get_session(session_key)
    flags = 0 if case_sensitive else re.IGNORECASE
    pattern = re.compile(query if regex else re.escape(query), flags)
    sheets = [_find_sheet(data, sheet_name)] if sheet_name else data.get("sheets", [])
    results = []
    for sheet in sheets:
        for row_index, row in enumerate(sheet.get("rows", [])):
            for col_index, cell in enumerate(row.get("cells", [])):
                if cell.get("merge") == "slave":
                    continue
                haystack = cell.get("v")
                if match_in == "formula":
                    if not isinstance(haystack, str) or not haystack.startswith("="):
                        continue
                elif match_in != "value":
                    raise ValueError("match_in must be 'value' or 'formula'")
                text = "" if haystack is None else str(haystack)
                if pattern.search(text):
                    results.append({"sheet_name": sheet["name"], "row_index": row_index, "col_index": col_index, "value": haystack})
                    if len(results) >= max_results:
                        return json.dumps({"query": query, "truncated": True, "count": len(results), "matches": results}, default=str, ensure_ascii=False)
    return json.dumps({"query": query, "truncated": False, "count": len(results), "matches": results}, default=str, ensure_ascii=False)

@mcp.tool()
def excel_get_workbook_summary(file_path: str) -> str:
    """Return a compact read-only workbook summary without creating a session."""
    path = uri_to_path(file_path)
    data = serialize_excel(str(path))
    sheets = []
    for sheet in data.get("sheets", []):
        rows = sheet.get("rows", [])
        max_cols = max((len(row.get("cells", [])) for row in rows), default=0)
        formula_count = 0
        merge_origins = 0
        non_empty = 0
        for row in rows:
            for cell in row.get("cells", []):
                value = cell.get("v")
                if value not in (None, ""):
                    non_empty += 1
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                merge = cell.get("merge")
                if isinstance(merge, dict) and merge:
                    merge_origins += 1
        sheets.append({
            "name": sheet.get("name"),
            "rows": len(rows),
            "columns": max_cols,
            "non_empty_cells": non_empty,
            "formula_count": formula_count,
            "merged_ranges": merge_origins,
            "freeze": sheet.get("freeze"),
            "validations": len(sheet.get("validations") or []),
        })
    return json.dumps({"source": str(path), "sheet_count": len(sheets), "sheets": sheets}, default=str, ensure_ascii=False)
@mcp.tool()
def excel_to_markdown_range(session_key: str, sheet_name: str, range_ref: str | None = None, start_row: int | None = None, end_row: int | None = None, start_col: int | None = None, end_col: int | None = None) -> TextContent:
    """Export one worksheet range as a compact Markdown table."""
    data = json.loads(excel_read_range(session_key, sheet_name, range_ref, start_row, end_row, start_col, end_col, True))
    values = data["values"]
    if not values:
        return TextContent(type="text", text="", mimeType="text/markdown")
    headers = ["" if value is None else str(value) for value in values[0]]
    if not headers:
        return TextContent(type="text", text="", mimeType="text/markdown")
    rows = values[1:] if len(values) > 1 else []
    def cell(value):
        return "" if value is None else str(value).replace("|", "\\|").replace("\n", "<br>")
    lines = ["| " + " | ".join(cell(value) for value in headers) + " |"]
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    lines.extend("| " + " | ".join(cell((row + [None] * len(headers))[index]) for index in range(len(headers))) + " |" for row in rows)
    return TextContent(type="text", text="\n".join(lines), mimeType="text/markdown")

@mcp.tool()
def excel_list_tables(session_key: str, sheet_name: str | None = None) -> str:
    """List Excel table objects captured in the loaded workbook session."""
    data = _get_session(session_key)
    sheets = [_find_sheet(data, sheet_name)] if sheet_name else data.get("sheets", [])
    tables = []
    for sheet in sheets:
        for table in sheet.get("tables") or []:
            tables.append({"sheet_name": sheet["name"], **table})
    return json.dumps({"count": len(tables), "tables": tables}, default=str, ensure_ascii=False)

@mcp.tool()
def excel_list_defined_names(session_key: str) -> str:
    """List workbook defined names and named ranges from the loaded session."""
    data = _get_session(session_key)
    names = data.get("named_ranges") or []
    return json.dumps({"count": len(names), "defined_names": _strip_private(names)}, default=str, ensure_ascii=False)

@mcp.tool()
def excel_get_sheet_preview(file_path: str, max_rows: int = 20, max_cols: int = 10, sheet_name: str | None = None) -> str:
    """Return compact top-left previews for one sheet or all sheets without creating a session."""
    path = uri_to_path(file_path)
    data = serialize_excel(str(path), sheet_name)
    previews = []
    for sheet in data.get("sheets", []):
        rows = []
        for row in sheet.get("rows", [])[:max_rows]:
            values = []
            for cell in row.get("cells", [])[:max_cols]:
                values.append(None if cell.get("merge") == "slave" else cell.get("v"))
            rows.append(values)
        previews.append({"sheet_name": sheet["name"], "rows": rows, "truncated_rows": len(sheet.get("rows", [])) > max_rows})
    return json.dumps({"source": str(path), "max_rows": max_rows, "max_cols": max_cols, "sheets": previews}, default=str, ensure_ascii=False)
# ── 7. Get cell ───────────────────────────────────────────────────────────────

@mcp.tool()
def excel_get_cell(
    session_key: str,
    sheet_name: str,
    row_index: int,
    col_index: int,
    include_rich_text: bool = False,
    include_formula_cache: bool = False,
    include_semantics: bool = False,
) -> str:
    """Read one 0-based cell with optional lossless semantic fields."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet.get("rows", [])
    cell = None
    if 0 <= row_index < len(rows):
        cells = rows[row_index].get("cells", [])
        if 0 <= col_index < len(cells):
            cell = cells[col_index]
    if cell is None and not include_semantics:
        row_max = len(rows) - 1
        raise ValueError(f"Cell [{row_index},{col_index}] is outside the materialized session grid (last row {row_max}).")
    result = _cell_public_view(
        cell,
        sheet=sheet,
        coord=_cell_coord(row_index, col_index),
        include_rich_text=include_rich_text,
        include_formula_cache=include_formula_cache,
        include_semantics=include_semantics,
    )
    return json.dumps(result, default=str, ensure_ascii=False)


# ── 8. Get column ─────────────────────────────────────────────────────────────

@mcp.tool()
def excel_get_column(
    session_key: str,
    sheet_name: str,
    col_index: int,
    start_row: int = 0,
    end_row: int | None = None,
) -> str:
    """
    Get all cells in a column as JSON.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_index: 0-based column index
        start_row: 0-based start row (inclusive), default 0
        end_row: 0-based end row (exclusive); omit for all rows

    Returns:
        JSON array of {row_index, cell} objects
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"][start_row:end_row]
    result = [
        {
            "row_index": start_row + i,
            "cell": (
                _strip_private(_cell_model_for_read(
                    row["cells"][col_index],
                    sheet.get("_implicit_cell_defaults"),
                ))
                if col_index < len(row["cells"])
                else None
            ),
        }
        for i, row in enumerate(rows)
    ]
    return json.dumps(result, default=str, ensure_ascii=False)


# ── 9. Sheet management ───────────────────────────────────────────────────────

@mcp.tool()
def excel_add_sheet(
    session_key: str,
    sheet_name: str,
    position: int | None = None,
) -> str:
    """
    Add a new empty sheet to the workbook session.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name for the new sheet (must be unique)
        position: 0-based position to insert at; omit to append at the end

    Returns:
        Confirmation with updated sheet list
    """
    data = _get_session(session_key)
    if any(sheet["name"] == sheet_name for sheet in data["sheets"]):
        raise ValueError(f"Sheet '{sheet_name}' already exists.")
    sheet_count = len(data["sheets"])
    actual = sheet_count if position is None else int(position)
    if not 0 <= actual <= sheet_count:
        raise ValueError(f"Position {actual} out of range (0–{sheet_count}).")
    new_sheet = _empty_sheet_model(sheet_name)
    data["sheets"].insert(actual, new_sheet)
    if actual < sheet_count:
        mapper = lambda index: index + 1 if index >= actual else index
        _remap_named_range_sheet_ids(data, mapper)
        _remap_workbook_view_sheet_ids(data, mapper)
    _mark_dirty(data, "workbook", f"sheets/{sheet_name}")
    return f"Added sheet '{sheet_name}' at position {actual}. Sheets: {[s['name'] for s in data['sheets']]}"


@mcp.tool()
def excel_delete_sheet(session_key: str, sheet_name: str) -> str:
    """
    Delete a sheet from the workbook session. Cannot delete the only sheet.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet to delete

    Returns:
        Confirmation with updated sheet list
    """
    data = _get_session(session_key)
    idx = next((i for i, s in enumerate(data["sheets"]) if s["name"] == sheet_name), None)
    if idx is None:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {[s['name'] for s in data['sheets']]}")
    if len(data["sheets"]) == 1:
        raise ValueError("Cannot delete the only sheet in a workbook.")
    data["sheets"].pop(idx)
    mapper = lambda index: None if index == idx else (index - 1 if index > idx else index)
    _remap_named_range_sheet_ids(data, mapper)
    _remap_workbook_view_sheet_ids(data, mapper, fallback_index=idx)
    _rewrite_workbook_sheet_references(data, sheet_name, None)
    _mark_dirty(data, "workbook", f"sheets/{sheet_name}")
    return f"Deleted '{sheet_name}'. Remaining: {[s['name'] for s in data['sheets']]}"


@mcp.tool()
def excel_rename_sheet(session_key: str, sheet_name: str, new_name: str) -> str:
    """
    Rename a sheet in the workbook session.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Current name of the sheet
        new_name: New name (must be unique)

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    if any(s["name"] == new_name for s in data["sheets"]):
        raise ValueError(f"Sheet '{new_name}' already exists.")
    _find_sheet(data, sheet_name)["name"] = new_name
    _rewrite_workbook_sheet_references(data, sheet_name, new_name)
    _mark_dirty(data, "workbook", f"sheets/{new_name}")
    return f"Renamed '{sheet_name}' → '{new_name}'."


@mcp.tool()
def excel_copy_sheet(
    session_key: str,
    source_sheet: str,
    new_name: str,
    position: int | None = None,
) -> str:
    """
    Duplicate a sheet within the same workbook session.

    All rows, styles, column widths, freeze panes, and validations are copied.

    Args:
        session_key: Key returned by excel_load
        source_sheet: Name of the sheet to copy
        new_name: Name for the copy (must be unique)
        position: 0-based insertion position; omit to append

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    if any(sheet["name"] == new_name for sheet in data["sheets"]):
        raise ValueError(f"Sheet '{new_name}' already exists.")
    source_index = _sheet_index(data, source_sheet)
    scoped_names = _copy_scoped_defined_names(data, source_index)
    new_sheet = copy.deepcopy(_find_sheet(data, source_sheet))
    new_sheet["name"] = new_name
    _rewrite_sheet_metadata_references(new_sheet, source_sheet, new_name)
    _rebase_copied_drawing_parts(data, new_sheet)
    _rebase_copied_passthrough_parts(data, new_sheet)

    sheet_count = len(data["sheets"])
    actual = sheet_count if position is None else int(position)
    if not 0 <= actual <= sheet_count:
        raise ValueError(f"Position {actual} out of range (0–{sheet_count}).")
    data["sheets"].insert(actual, new_sheet)
    if actual < sheet_count:
        mapper = lambda index: index + 1 if index >= actual else index
        _remap_named_range_sheet_ids(data, mapper)
        _remap_workbook_view_sheet_ids(data, mapper)
    _append_copied_defined_names(data, scoped_names, actual, source_sheet, new_name)
    _ensure_unique_sheet_codename(data, new_sheet)
    _dedupe_table_names(data, new_sheet)
    _mark_dirty(data, "workbook", f"sheets/{new_name}")
    return f"Copied '{source_sheet}' → '{new_name}' at position {actual}."


@mcp.tool()
def excel_copy_sheet_to(
    src_session_key: str,
    src_sheet_name: str,
    dst_session_key: str,
    new_name: str | None = None,
    position: int | None = None,
) -> str:
    """
    Copy a sheet from one loaded workbook session into another.

    Both source and destination files must be loaded with excel_load first.
    After copying, call excel_save on the destination session_key to persist.

    Args:
        src_session_key: session_key of the source file
        src_sheet_name: Name of the sheet to copy from the source
        dst_session_key: session_key of the destination file
        new_name: Name for the sheet in the destination; defaults to src_sheet_name
        position: 0-based insertion position in the destination; omit to append

    Returns:
        Confirmation
    """
    src_data = _get_session(src_session_key)
    dst_data = _get_session(dst_session_key)
    source_index = _sheet_index(src_data, src_sheet_name)
    scoped_names = _copy_scoped_defined_names(src_data, source_index)
    sheet_copy = copy.deepcopy(_find_sheet(src_data, src_sheet_name))
    target_name = new_name or src_sheet_name

    if any(sheet["name"] == target_name for sheet in dst_data["sheets"]):
        raise ValueError(f"Sheet '{target_name}' already exists in destination. Specify a different new_name.")
    sheet_count = len(dst_data["sheets"])
    actual = sheet_count if position is None else int(position)
    if not 0 <= actual <= sheet_count:
        raise ValueError(f"Position {actual} out of range (0–{sheet_count}).")

    if src_data.get("theme_xml") != dst_data.get("theme_xml"):
        _drop_raw_fills(sheet_copy)
    sheet_copy["name"] = target_name
    _rewrite_sheet_metadata_references(sheet_copy, src_sheet_name, target_name)
    _copy_dxf_dependencies(src_data, dst_data, sheet_copy)
    _rebase_copied_drawing_parts(dst_data, sheet_copy)
    _rebase_copied_passthrough_parts(dst_data, sheet_copy)

    dst_data["sheets"].insert(actual, sheet_copy)
    if actual < sheet_count:
        mapper = lambda index: index + 1 if index >= actual else index
        _remap_named_range_sheet_ids(dst_data, mapper)
        _remap_workbook_view_sheet_ids(dst_data, mapper)
    _append_copied_defined_names(dst_data, scoped_names, actual, src_sheet_name, target_name)
    _ensure_unique_sheet_codename(dst_data, sheet_copy)
    _dedupe_table_names(dst_data, sheet_copy)
    _mark_dirty(dst_data, "workbook", f"sheets/{target_name}")

    return (
        f"Copied sheet '{src_sheet_name}' from '{src_session_key}' "
        f"→ '{target_name}' in '{dst_session_key}' at position {actual}. "
        f"Call excel_save on the destination to persist."
    )


@mcp.tool()
def excel_move_sheet(
    session_key: str,
    sheet_name: str,
    position: int,
) -> str:
    """
    Move a sheet to a new position within the workbook.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet to move
        position: New 0-based position

    Returns:
        Confirmation with updated sheet order
    """
    data = _get_session(session_key)
    idx = next((i for i, s in enumerate(data["sheets"]) if s["name"] == sheet_name), None)
    if idx is None:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {[s['name'] for s in data['sheets']]}")
    n = len(data["sheets"])
    if not (0 <= position < n):
        raise ValueError(f"Position {position} out of range (0–{n-1}).")
    sheet = data["sheets"].pop(idx)
    data["sheets"].insert(position, sheet)
    # Remap localSheetId of defined names to the new sheet order
    order = list(range(n))
    moved = order.pop(idx)
    order.insert(position, moved)
    mapping = {old: new for new, old in enumerate(order)}
    mapper = lambda index: mapping.get(index, index)
    _remap_named_range_sheet_ids(data, mapper)
    _remap_workbook_view_sheet_ids(data, mapper)
    _mark_dirty(data, "workbook", f"sheets/{sheet_name}")
    return f"Moved '{sheet_name}' to position {position}. Order: {[s['name'] for s in data['sheets']]}"


# ── 10. Clone rows ────────────────────────────────────────────────────────────

@mcp.tool()
def excel_clone_rows(
    session_key: str,
    sheet_name: str,
    start_row: int,
    end_row: int | None = None,
) -> str:
    """
    Deep-clone one or more rows and return them as a JSON array WITHOUT inserting.

    Use with excel_insert_rows to insert the cloned block at one or more positions.
    Clone BEFORE any inserts to avoid index drift.

    For a single row: pass start_row only (end_row defaults to start_row).
    For a range:      pass start_row and end_row (both inclusive).

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        start_row: 0-based index of the first row to clone (inclusive)
        end_row: 0-based index of the last row to clone (inclusive);
                 defaults to start_row for single-row clone

    Returns:
        JSON array of cloned row objects
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    n = len(sheet["rows"])
    if end_row is None:
        end_row = start_row
    if not (0 <= start_row <= end_row < n):
        raise ValueError(f"Row range [{start_row}, {end_row}] out of bounds (0–{n-1})")
    cloned_rows = copy.deepcopy(sheet["rows"][start_row:end_row + 1])
    implicit_defaults = sheet.get("_implicit_cell_defaults")
    for row in cloned_rows:
        row["cells"] = [
            _cell_model_for_read(cell, implicit_defaults)
            for cell in row.get("cells", [])
        ]
    return json.dumps(cloned_rows, default=str, ensure_ascii=False)


# ── 11. Copy row ──────────────────────────────────────────────────────────────

@mcp.tool()
def excel_copy_row(
    session_key: str,
    sheet_name: str,
    row_index: int,
    after_index: int,
) -> str:
    """
    Clone a row and insert the copy immediately at a new position (one step).

    Contrast with excel_clone_rows, which only returns the row as JSON so you
    can modify it before inserting.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        row_index: 0-based index of the row to copy
        after_index: 0-based index to insert AFTER; use -1 to prepend

    Returns:
        Confirmation with new row count
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    n = len(sheet["rows"])
    if not (0 <= row_index < n):
        raise ValueError(f"row_index {row_index} out of range (0–{n-1})")
    cloned = copy.deepcopy(sheet["rows"][row_index])
    pos = after_index + 1
    _apply_row_insert(data, sheet, pos, [cloned])
    return f"Copied row {row_index} → inserted at position {pos}. Sheet '{sheet_name}' now has {len(sheet['rows'])} rows."


# ── 12. Insert rows ───────────────────────────────────────────────────────────

@mcp.tool()
def excel_insert_rows(
    session_key: str,
    sheet_name: str,
    inserts: list[dict],
) -> str:
    """
    Insert rows at one or more positions in a single call.

    Positions are automatically sorted bottom-to-top, so each after_index
    refers to the original row positions (before any insertions).

    For a single insert: pass a list with one item.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the target sheet
        inserts: List of {"after_index": int, "rows_json": list | str} objects.
                 after_index — 0-based original index to insert AFTER; -1 to prepend.
                 rows_json   — a row object, a list of row objects, or a JSON string
                               (as returned by excel_clone_rows).

    Returns:
        Summary: total rows inserted, new total row count
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)

    parsed: list[tuple[int, list]] = []
    for entry in inserts:
        rows_json = entry["rows_json"]
        if isinstance(rows_json, str):
            rows_json = json.loads(rows_json)
        if isinstance(rows_json, dict):
            rows_json = [rows_json]
        parsed.append((entry["after_index"], rows_json))

    parsed.sort(key=lambda x: x[0], reverse=True)

    total_inserted = 0
    for after_index, new_rows in parsed:
        pos = after_index + 1
        _apply_row_insert(data, sheet, pos, new_rows)
        total_inserted += len(new_rows)

    return (
        f"Inserted {total_inserted} row(s) across {len(parsed)} position(s). "
        f"Sheet '{sheet_name}' now has {len(sheet['rows'])} rows."
    )


# ── 13. Insert column ─────────────────────────────────────────────────────────

@mcp.tool()
def excel_insert_column(
    session_key: str,
    sheet_name: str,
    after_col_index: int,
) -> str:
    """
    Insert a new empty column after the given column index.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        after_col_index: 0-based column index to insert AFTER; use -1 to prepend

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    pos = after_col_index + 1
    regions = _capture_merge_regions(sheet["rows"])
    for row in sheet["rows"]:
        new_cell = copy.deepcopy(_EMPTY_CELL)
        row["cells"] = row["cells"][:pos] + [new_cell] + row["cells"][pos:]
    _finish_col_insert(data, sheet, regions, pos)
    return f"Inserted empty column at position {pos} in sheet '{sheet_name}'."


# ── 14. Edit cells ────────────────────────────────────────────────────────────

@mcp.tool()
def excel_edit_cells(
    session_key: str,
    sheet_name: str,
    edits: list[dict],
) -> str:
    """Edit scalar values or typed partial cell payloads without saving."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    typed_keys = {
        "value", "data_type", "formula", "formula_type", "formula_attributes",
        "cached_value", "cache_policy", "rich_text", "rich_text_policy", "clear", "present",
    }
    changes = []
    edited_rows = set()
    for entry in edits:
        row_index = int(entry["row_index"])
        if row_index < 0:
            raise ValueError(f"row_index must be >= 0, got {row_index}")
        for col_key, value in entry["edits"].items():
            col_index = int(col_key)
            cell = _ensure_cell(sheet, row_index, col_index, capture_baseline=True)
            if cell.get("merge") == "slave":
                raise ValueError(
                    f"Cell [{row_index},{col_index}] is a slave cell of a merged range; edit its origin."
                )
            before = _cell_public_view(cell, sheet=sheet, coord=_cell_coord(row_index, col_index),
                                       include_rich_text=True, include_formula_cache=True,
                                       include_semantics=True)
            if isinstance(value, dict) and typed_keys.intersection(value):
                _apply_typed_cell_payload(cell, value)
            else:
                if cell.get("rich_text"):
                    # A bare scalar value has no way to carry rich_text_policy,
                    # so silently clearing the existing runs here would be
                    # exactly the "silently guessing" behavior the typed
                    # payload path (_apply_typed_cell_payload) already refuses
                    # to do. Require the caller to use the typed payload form
                    # ({"value": ..., "rich_text_policy": "replace_all" |
                    # "preserve_runs_if_text_equal"}) instead.
                    raise ValueError(
                        f"Cell [{row_index},{col_index}] has rich text; editing its plain value "
                        "requires the typed payload form with an explicit rich_text_policy "
                        "('replace_all' or 'preserve_runs_if_text_equal')."
                    )
                _store_cell_value(cell, value, sheet.get("_implicit_cell_defaults"))
                cell["value"] = copy.deepcopy(cell.get("v"))
                cell["data_type"] = cell.get("dt")
                cell["present"] = True
                cell.pop("rich_text", None)
                cell.pop("formula", None)
                cell.pop("cached_value", None)
                cell.pop("cached_value_state", None)
            path = f"sheets/{sheet_name}/cells/{_cell_coord(row_index, col_index)}"
            _mark_dirty(data, "cells", path)
            after = _cell_public_view(cell, sheet=sheet, coord=_cell_coord(row_index, col_index),
                                      include_rich_text=True, include_formula_cache=True,
                                      include_semantics=True)
            changes.append({"path": path, "before": before, "after": after})
            edited_rows.add(row_index)
    return json.dumps({
        "rows_edited": len(edited_rows),
        "cells_updated": len(changes),
        "changes": changes,
        "dirty_features": data.get("_dirty_features", []),
        "dirty_paths": data.get("_dirty_paths", []),
    }, default=str, ensure_ascii=False)


# ── 15. Delete rows ───────────────────────────────────────────────────────────

@mcp.tool()
def excel_delete_rows(
    session_key: str,
    sheet_name: str,
    row_indices: list[int] | None = None,
    start_row: int | None = None,
    end_row: int | None = None,
) -> str:
    """
    Delete one or more rows by index list or by a contiguous range.

    Provide row_indices OR start_row+end_row (or both — they are merged).
    end_row is EXCLUSIVE (Python convention): to delete rows 14–18 inclusive
    pass start_row=14, end_row=19.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        row_indices: List of 0-based row indices to delete
        start_row: Start of a contiguous range (0-based, inclusive)
        end_row: End of a contiguous range (0-based, EXCLUSIVE)

    Returns:
        Confirmation with remaining row count
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    n = len(sheet["rows"])

    to_delete: set[int] = set(row_indices or [])
    if start_row is not None and end_row is not None:
        to_delete.update(range(start_row, end_row))
    elif start_row is not None or end_row is not None:
        raise ValueError("Provide both start_row and end_row together.")
    if not to_delete:
        raise ValueError("Provide row_indices or start_row+end_row.")

    invalid = [i for i in to_delete if not (0 <= i < n)]
    if invalid:
        raise ValueError(f"Row indices out of range (0–{n-1}): {sorted(invalid)}")
    _apply_row_delete(data, sheet, to_delete)
    return f"Deleted {len(to_delete)} row(s). Sheet '{sheet_name}' now has {len(sheet['rows'])} rows."


# ── 16. Clear range ───────────────────────────────────────────────────────────

@mcp.tool()
def excel_clear_range(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    r2: int,
    c2: int,
    clear_values: bool = True,
    clear_styles: bool = False,
) -> str:
    """
    Clear values and/or styles from a rectangular cell range.

    Slave cells of merged regions are skipped (their content belongs to the
    origin). To also remove merge structure, call excel_merge_cells with
    unmerge=True first.

    All coordinates are 0-based, inclusive on both ends.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        r1: 0-based top row (inclusive)
        c1: 0-based left column (inclusive)
        r2: 0-based bottom row (inclusive)
        c2: 0-based right column (inclusive)
        clear_values: If True (default), set cell values to null
        clear_styles: If True, reset fill/font/alignment/border to defaults

    Returns:
        Summary: cells cleared
    """
    _STYLE_DEFAULTS = {
        "fill": None, "bold": False, "italic": False, "size": None,
        "fcolor": None, "wrap": False, "halign": None, "valign": None,
        "numfmt": "General", "border": {},
    }

    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]

    cells_cleared = 0
    for r in range(r1, r2 + 1):
        if r >= len(rows):
            continue
        row_cells = rows[r]["cells"]
        for c in range(c1, c2 + 1):
            if c >= len(row_cells):
                continue
            cell = row_cells[c]
            if cell.get("_implicit") or cell.get("merge") == "slave":
                continue
            if clear_values or clear_styles:
                _cell_baseline(cell)
            if clear_values:
                cell["v"] = None
                cell.pop("dt", None)
                cell.pop("qp", None)
            if clear_styles:
                for k, v in _STYLE_DEFAULTS.items():
                    cell[k] = v
                cell.pop("_fill_raw", None)
                cell.pop("_font_raw", None)
                cell.pop("qp", None)
            path = f"sheets/{sheet_name}/cells/{_cell_coord(r, c)}"
            if clear_values:
                _mark_dirty(data, "cells", path)
            if clear_styles:
                _mark_dirty(data, "cell_style", path + "/style")
            cells_cleared += 1

    actions = []
    if clear_values:
        actions.append("values")
    if clear_styles:
        actions.append("styles")
    return (
        f"Cleared {' + '.join(actions)} in {cells_cleared} cell(s) "
        f"[{r1},{c1}]–[{r2},{c2}] in sheet '{sheet_name}'."
    )


# ── 17. Copy / Delete column ──────────────────────────────────────────────────

@mcp.tool()
def excel_copy_column(
    session_key: str,
    sheet_name: str,
    col_index: int,
    after_col_index: int,
) -> str:
    """
    Copy a column and insert it after a given column index, preserving all styles.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_index: 0-based index of the column to copy
        after_col_index: 0-based index to insert AFTER; use -1 to prepend

    Returns:
        Confirmation
    """
    import openpyxl.utils
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    pos = after_col_index + 1
    # Capture source width before shifting
    src_letter = openpyxl.utils.get_column_letter(col_index + 1)
    src_width = sheet["cw"].get(src_letter)
    regions = _capture_merge_regions(sheet["rows"])
    for row in sheet["rows"]:
        src = copy.deepcopy(row["cells"][col_index]) if col_index < len(row["cells"]) else copy.deepcopy(_EMPTY_CELL)
        if src.get("merge"):
            src["merge"] = {}  # copied cells never carry the source's merge state
        row["cells"] = row["cells"][:pos] + [src] + row["cells"][pos:]
    _finish_col_insert(data, sheet, regions, pos)
    if src_width:
        sheet["cw"][openpyxl.utils.get_column_letter(pos + 1)] = src_width
    return f"Copied column {col_index} → inserted at position {pos} in sheet '{sheet_name}'."


@mcp.tool()
def excel_delete_column(session_key: str, sheet_name: str, col_index: int) -> str:
    """
    Delete a column from all rows in a sheet.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_index: 0-based column index to delete

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    regions = _capture_merge_regions(sheet["rows"])
    removed = 0
    for row in sheet["rows"]:
        if col_index < len(row["cells"]):
            row["cells"].pop(col_index)
            removed += 1
    _finish_col_delete(data, sheet, regions, col_index)
    return f"Deleted column {col_index} from {removed} row(s) in sheet '{sheet_name}'."


# ── 18. Merge cells ───────────────────────────────────────────────────────────

@mcp.tool()
def excel_merge_cells(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    r2: int | None = None,
    c2: int | None = None,
    unmerge: bool = False,
) -> str:
    """
    Merge a rectangular range of cells, or unmerge a merged region.

    Merge:   pass r1, c1, r2, c2 (all required when unmerge=False).
             The top-left cell (r1, c1) becomes the origin and keeps its value.
             All other cells in the range are marked as slave cells.

    Unmerge: pass r1, c1 of the origin cell and unmerge=True.
             r2, c2 are ignored — the full merge region is found automatically.
             All cells in the region revert to independent cells.

    All coordinates are 0-based.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        r1: Top row (0-based)
        c1: Left column (0-based)
        r2: Bottom row (0-based, inclusive) — required for merge, ignored for unmerge
        c2: Right column (0-based, inclusive) — required for merge, ignored for unmerge
        unmerge: If True, unmerge the region whose origin is at (r1, c1)

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]
    n_rows = len(rows)

    if unmerge:
        if not (0 <= r1 < n_rows):
            raise ValueError(f"r1={r1} out of range (0–{n_rows-1})")
        cell = rows[r1]["cells"][c1]
        if cell["merge"] == "slave":
            raise ValueError(f"Cell [{r1},{c1}] is a slave cell. Pass the origin (top-left) of the merge.")
        mi = cell["merge"]
        if not isinstance(mi, dict) or (mi.get("rowspan", 1) <= 1 and mi.get("colspan", 1) <= 1):
            raise ValueError(f"Cell [{r1},{c1}] is not a merge origin.")
        er1, ec1 = mi.get("r1", r1), mi.get("c1", c1)
        er2, ec2 = mi.get("r2", r1), mi.get("c2", c1)
        for r in range(er1, er2 + 1):
            for c in range(ec1, ec2 + 1):
                target = _promote_implicit_cell(
                    rows[r]["cells"][c],
                    sheet.get("_implicit_cell_defaults"),
                )
                _cell_baseline(target)
                target["merge"] = {}
        _mark_dirty(data, "structure", f"sheets/{sheet_name}/merges")
        return f"Unmerged [{er1},{ec1}]–[{er2},{ec2}] in sheet '{sheet_name}'."

    # Merge
    if r2 is None or c2 is None:
        raise ValueError("r2 and c2 are required for merge. Pass unmerge=True to unmerge.")
    if not (0 <= r1 <= r2 < n_rows):
        raise ValueError(f"Row range [{r1}, {r2}] out of bounds (0–{n_rows-1})")
    for r in range(r1, r2 + 1):
        n_cols = len(rows[r]["cells"])
        if not (0 <= c1 <= c2 < n_cols):
            raise ValueError(f"Col range [{c1}, {c2}] out of bounds for row {r} (0–{n_cols-1})")

    # Reject overlap with any existing merged region (would corrupt the file)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            mi = rows[r]["cells"][c].get("merge")
            if mi == "slave" or (isinstance(mi, dict)
                                 and (mi.get("rowspan", 1) > 1 or mi.get("colspan", 1) > 1)):
                raise ValueError(
                    f"Range overlaps an existing merged region at [{r},{c}]. "
                    "Unmerge it first (excel_merge_cells with unmerge=True).")

    origin = _promote_implicit_cell(
        rows[r1]["cells"][c1],
        sheet.get("_implicit_cell_defaults"),
    )
    _cell_baseline(origin)
    origin["merge"] = {
        "r1": r1, "c1": c1, "r2": r2, "c2": c2,
        "rowspan": r2 - r1 + 1, "colspan": c2 - c1 + 1,
    }
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            if not (r == r1 and c == c1):
                target = _promote_implicit_cell(
                    rows[r]["cells"][c],
                    sheet.get("_implicit_cell_defaults"),
                )
                _cell_baseline(target)
                target["merge"] = "slave"
    _mark_dirty(data, "structure", f"sheets/{sheet_name}/merges")
    return (
        f"Merged [{r1},{c1}]–[{r2},{c2}] "
        f"({r2-r1+1}×{c2-c1+1}) in sheet '{sheet_name}'."
    )


# ── 19. Style ─────────────────────────────────────────────────────────────────

@mcp.tool()
def excel_set_style(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    r2: int | None = None,
    c2: int | None = None,
    style: dict = {},
) -> str:
    """Patch whole-cell font/fill/alignment/protection/XF semantics; rich runs stay intact."""
    r2 = r1 if r2 is None else r2
    c2 = c1 if c2 is None else c2
    if r1 < 0 or c1 < 0 or r2 < r1 or c2 < c1:
        raise ValueError("Invalid style range.")
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    changes = []
    for row_index in range(r1, r2 + 1):
        for col_index in range(c1, c2 + 1):
            cell = _ensure_cell(sheet, row_index, col_index, capture_baseline=True)
            if cell.get("merge") == "slave":
                continue
            before = _cell_public_view(cell, include_semantics=True, include_rich_text=True)
            _apply_structured_style(cell, style)
            after = _cell_public_view(cell, include_semantics=True, include_rich_text=True)
            path = f"sheets/{sheet_name}/cells/{_cell_coord(row_index, col_index)}/style"
            _mark_dirty(data, "cell_style", path)
            changes.append({"path": path, "before": before, "after": after})
    return json.dumps({"cells_styled": len(changes), "changes": changes,
                       "dirty_features": data.get("_dirty_features", [])},
                      default=str, ensure_ascii=False)


@mcp.tool()
def excel_set_font_color(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    color: str | dict | None,
    r2: int | None = None,
    c2: int | None = None,
) -> str:
    """Set font color on a cell or range. Color is ARGB hex, or null for auto."""
    style = {"font": {"color": color}} if isinstance(color, dict) else {"fcolor": color}
    return excel_set_style(session_key, sheet_name, r1, c1, r2, c2, style)


@mcp.tool()
def excel_set_strike(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    enabled: bool = True,
    r2: int | None = None,
    c2: int | None = None,
) -> str:
    """Enable or disable strikethrough on a cell or range."""
    return excel_set_style(session_key, sheet_name, r1, c1, r2, c2, {"strike": enabled})



def _shape_anchors(drawing_xml: str) -> list[re.Match]:
    return list(re.finditer(
        r"<(?:(?:xdr:)?)(twoCellAnchor|oneCellAnchor|absoluteAnchor)\b.*?</(?:(?:xdr:)?)(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)>",
        drawing_xml,
        re.DOTALL,
    ))


def _rgb_hex(color: str | None) -> str | None:
    if color is None:
        return None
    value = color.strip().lstrip("#").upper()
    if len(value) == 8:
        value = value[2:]
    if not re.fullmatch(r"[0-9A-F]{6}", value):
        raise ValueError("Color must be RGB or ARGB hex, e.g. 'FF0000' or 'FFFF0000'.")
    return value


_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
ET.register_namespace("a", _A_NS)
ET.register_namespace("xdr", _XDR_NS)
_SHAPE_CLEAR = "__DOCLOUPE_CLEAR__"
_SHAPE_KEEP = "__DOCLOUPE_KEEP__"


def _solid_fill_xml(rgb: str) -> str:
    return f'<a:solidFill><a:srgbClr val="{rgb}"/></a:solidFill>'


def _et_local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _et_tostring(elem: ET.Element) -> str:
    return ET.tostring(elem, encoding="unicode", short_empty_elements=True)


def _parse_sp_pr(sp_pr: str) -> ET.Element:
    wrapper = ET.fromstring(
        f'<docloupe:root xmlns:docloupe="urn:docloupe" xmlns:xdr="{_XDR_NS}" xmlns:a="{_A_NS}">{sp_pr}</docloupe:root>'
    )
    return list(wrapper)[0]


def _first_child_index(elem: ET.Element, names: set[str]) -> int | None:
    for idx, child in enumerate(list(elem)):
        if _et_local(child.tag) in names:
            return idx
    return None


def _new_solid_fill(rgb: str) -> ET.Element:
    fill = ET.Element(f"{{{_A_NS}}}solidFill")
    ET.SubElement(fill, f"{{{_A_NS}}}srgbClr", {"val": rgb})
    return fill


def _new_no_fill() -> ET.Element:
    return ET.Element(f"{{{_A_NS}}}noFill")


def _replace_direct_fill(elem: ET.Element, replacement: ET.Element) -> None:
    fill_names = {"noFill", "solidFill", "gradFill", "blipFill", "pattFill", "grpFill"}
    for idx, child in enumerate(list(elem)):
        if _et_local(child.tag) in fill_names:
            elem.remove(child)
            elem.insert(idx, replacement)
            return
    geom_idx = _first_child_index(elem, {"prstGeom", "custGeom"})
    insert_at = geom_idx + 1 if geom_idx is not None else 0
    elem.insert(insert_at, replacement)


def _set_shape_fill(sp_pr: str, rgb: str | None) -> str:
    root = _parse_sp_pr(sp_pr)
    _replace_direct_fill(root, _new_no_fill() if rgb is None else _new_solid_fill(rgb))
    return _et_tostring(root)


def _set_shape_line(sp_pr: str, rgb_marker, width_pt: float | None) -> str:
    root = _parse_sp_pr(sp_pr)
    if width_pt is not None and width_pt < 0:
        raise ValueError("outline_width_pt must be >= 0.")
    line = next((c for c in list(root) if _et_local(c.tag) == "ln"), None)
    if line is None:
        line = ET.Element(f"{{{_A_NS}}}ln")
        root.append(line)
    if width_pt is not None:
        line.set("w", str(int(round(width_pt * 12700))))
    if rgb_marker is _SHAPE_CLEAR:
        _replace_direct_fill(line, _new_no_fill())
    elif rgb_marker is not _SHAPE_KEEP:
        _replace_direct_fill(line, _new_solid_fill(rgb_marker))
    return _et_tostring(root)


def _set_shape_text_color(anchor_xml: str, rgb: str) -> str:
    def patch_rpr(match: re.Match) -> str:
        tag = match.group(0)
        if tag.endswith("/>"):
            tag = tag[:-2] + ">" + _solid_fill_xml(rgb) + "</a:rPr>"
        elif re.search(r"<a:solidFill\b.*?</a:solidFill>|<a:noFill\s*/>", tag, re.DOTALL):
            tag = re.sub(r"<a:solidFill\b.*?</a:solidFill>|<a:noFill\s*/>", _solid_fill_xml(rgb), tag, count=1, flags=re.DOTALL)
        else:
            tag = tag.replace("</a:rPr>", _solid_fill_xml(rgb) + "</a:rPr>", 1)
        return tag
    updated, count = re.subn(r"<a:rPr\b.*?</a:rPr>|<a:rPr\b[^/]*/>", patch_rpr, anchor_xml, flags=re.DOTALL)
    return updated if count else anchor_xml

@mcp.tool()
def excel_get_shapes(session_key: str, sheet_name: str | None = None) -> str:
    """List DrawingML shapes/images/charts captured from loaded sheets."""
    data = _get_session(session_key)
    result = {}
    for sheet in data["sheets"]:
        if sheet_name and sheet["name"] != sheet_name:
            continue
        result[sheet["name"]] = sheet.get("shapes") or []
    if sheet_name and sheet_name not in result:
        _find_sheet(data, sheet_name)
    return json.dumps(result, ensure_ascii=False, indent=2)


@mcp.tool()
def excel_update_shape_text(
    session_key: str,
    sheet_name: str,
    shape_index: int,
    text: str | None = None,
    rich_text: dict | None = None,
) -> str:
    """Replace shape text with a plain string or the common rich-text run model."""
    if rich_text is not None:
        source_runs = rich_text.get("runs") or []
        if not source_runs and rich_text.get("text") is not None:
            source_runs = [{"text": str(rich_text.get("text") or "")}]
        runs, plain = _reindex_rich_runs(source_runs)
        model = {**copy.deepcopy(rich_text), "runs": runs, "text": plain}
    elif text is not None:
        runs, plain = _reindex_rich_runs([{"text": str(text)}])
        model = {"runs": runs, "text": plain}
    else:
        raise ValueError("Provide text or rich_text.")

    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    drawing = sheet.get("drawing_data")
    if not drawing or not drawing.get("drawing_xml"):
        raise ValueError(f"Sheet '{sheet_name}' has no captured DrawingML shapes.")
    if shape_index < 1:
        raise ValueError("shape_index is 1-based and must be >= 1.")

    anchors = list(re.finditer(
        r"<(?:(?:xdr:)?)(twoCellAnchor|oneCellAnchor|absoluteAnchor)\b.*?</(?:(?:xdr:)?)(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)>",
        drawing["drawing_xml"],
        re.DOTALL,
    ))
    if shape_index > len(anchors):
        raise ValueError(f"shape_index {shape_index} out of bounds; found {len(anchors)} shape(s).")

    match = anchors[shape_index - 1]
    anchor_xml = match.group(0)
    paragraphs = list(re.finditer(r"<a:p\b[^>]*>.*?</a:p>", anchor_xml, re.DOTALL))
    if not paragraphs:
        raise ValueError(f"Shape {shape_index} has no editable text paragraphs.")

    first_paragraph = paragraphs[0].group(0)
    last_paragraph = paragraphs[-1].group(0)
    paragraph_properties = re.search(
        r"<a:pPr\b.*?</a:pPr>|<a:pPr\b[^>]*/>",
        first_paragraph,
        re.DOTALL,
    )
    end_properties = re.search(
        r"<a:endParaRPr\b.*?</a:endParaRPr>|<a:endParaRPr\b[^>]*/>",
        last_paragraph,
        re.DOTALL,
    )
    runs_xml = build_shape_rich_text_xml(model)
    new_paragraph = (
        "<a:p>"
        + (paragraph_properties.group(0) if paragraph_properties else "")
        + runs_xml
        + (end_properties.group(0) if end_properties else "<a:endParaRPr/>")
        + "</a:p>"
    )
    new_anchor = (
        anchor_xml[:paragraphs[0].start()]
        + new_paragraph
        + anchor_xml[paragraphs[-1].end():]
    )
    drawing["drawing_xml"] = drawing["drawing_xml"][:match.start()] + new_anchor + drawing["drawing_xml"][match.end():]
    shapes = sheet.get("shapes") or []
    if shape_index <= len(shapes):
        shapes[shape_index - 1]["text"] = plain
        shapes[shape_index - 1]["rich_text"] = model
    _mark_dirty(data, "drawings", f"sheets/{sheet_name}/drawing_shapes/{shape_index}/text")
    return f"Updated text for shape {shape_index} on sheet '{sheet_name}'."



@mcp.tool()
def excel_set_shape_style(
    session_key: str,
    sheet_name: str,
    shape_index: int,
    fill_color: str | None = None,
    outline_color: str | None = None,
    outline_width_pt: float | None = None,
    text_color: str | None = None,
    clear_fill: bool = False,
    clear_outline: bool = False,
) -> str:
    """
    Set simple DrawingML shape style by 1-based shape index.

    Colors accept RGB or ARGB hex. Use clear_fill=True or clear_outline=True to remove fill/outline.
    Supports simple DrawingML shapes/textboxes captured by excel_load.
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    drawing = sheet.get("drawing_data")
    if not drawing or not drawing.get("drawing_xml"):
        raise ValueError(f"Sheet '{sheet_name}' has no captured DrawingML shapes.")
    if shape_index < 1:
        raise ValueError("shape_index is 1-based and must be >= 1.")

    anchors = _shape_anchors(drawing["drawing_xml"])
    if shape_index > len(anchors):
        raise ValueError(f"shape_index {shape_index} out of bounds; found {len(anchors)} shape(s).")

    if clear_fill and fill_color is not None:
        raise ValueError("Use either fill_color or clear_fill, not both.")
    if clear_outline and outline_color is not None:
        raise ValueError("Use either outline_color or clear_outline, not both.")
    fill_rgb = _rgb_hex(fill_color) if fill_color is not None else None
    outline_rgb = _rgb_hex(outline_color) if outline_color is not None else None
    text_rgb = _rgb_hex(text_color) if text_color is not None else None

    match = anchors[shape_index - 1]
    anchor_xml = match.group(0)
    sp_match = re.search(r"<(?:(?:xdr:)?sp)\b.*?</(?:(?:xdr:)?sp)>", anchor_xml, re.DOTALL)
    if not sp_match:
        raise ValueError(f"Shape {shape_index} is not a simple editable DrawingML shape.")
    shape_xml = sp_match.group(0)
    sp_pr_match = re.search(r"<(?:(?:xdr:)?spPr)\b.*?</(?:(?:xdr:)?spPr)>", shape_xml, re.DOTALL)
    if not sp_pr_match:
        raise ValueError(f"Shape {shape_index} has no editable shape properties.")

    sp_pr = sp_pr_match.group(0)
    if fill_color is not None or clear_fill:
        sp_pr = _set_shape_fill(sp_pr, None if clear_fill else fill_rgb)
    if outline_color is not None or outline_width_pt is not None or clear_outline:
        outline_marker = _SHAPE_CLEAR if clear_outline else (outline_rgb if outline_color is not None else _SHAPE_KEEP)
        sp_pr = _set_shape_line(sp_pr, outline_marker, outline_width_pt)
    shape_xml = shape_xml[:sp_pr_match.start()] + sp_pr + shape_xml[sp_pr_match.end():]
    anchor_xml = anchor_xml[:sp_match.start()] + shape_xml + anchor_xml[sp_match.end():]
    if text_rgb is not None:
        anchor_xml = _set_shape_text_color(anchor_xml, text_rgb)

    drawing["drawing_xml"] = drawing["drawing_xml"][:match.start()] + anchor_xml + drawing["drawing_xml"][match.end():]
    shapes = sheet.get("shapes") or []
    if shape_index <= len(shapes):
        if fill_color is not None or clear_fill:
            shapes[shape_index - 1]["fill_color"] = None if clear_fill else fill_rgb
        if outline_color is not None or clear_outline:
            shapes[shape_index - 1]["outline_color"] = None if clear_outline else outline_rgb
        if outline_width_pt is not None:
            shapes[shape_index - 1]["outline_width_pt"] = outline_width_pt
        if text_color is not None:
            shapes[shape_index - 1]["text_color"] = text_rgb
    _mark_dirty(data, "drawings", f"sheets/{sheet_name}/drawing_shapes/{shape_index}/style")
    return f"Updated style for shape {shape_index} on sheet '{sheet_name}'."
# ── 20. Borders ───────────────────────────────────────────────────────────────

@mcp.tool()
def excel_set_borders(
    session_key: str,
    sheet_name: str,
    r1: int,
    c1: int,
    r2: int,
    c2: int,
    style: str | None = None,
    sides: list[str] | None = None,
    color: str | dict | None = None,
    border: dict | None = None,
) -> str:
    """Patch border sides/flags while preserving unspecified sibling sides."""
    valid_sides = {"top", "bottom", "left", "right", "start", "end", "vertical", "horizontal", "diagonal"}
    updates = copy.deepcopy(border or {})
    if border is None:
        selected = sides or ["top", "bottom", "left", "right"]
        bad = set(selected) - valid_sides
        if bad:
            raise ValueError(f"Invalid border sides: {sorted(bad)}")
        color_object = _normalize_color_object(color) if color is not None else None
        side = None if style in {None, "none"} else {"style": style}
        if side is not None and color_object is not None:
            side["color"] = _resolved_rgb(color_object)
            side["_color_raw"] = color_object
        updates.update({name: copy.deepcopy(side) for name in selected})
    unknown = set(updates) - valid_sides - {"outline", "diagonalUp", "diagonalDown"}
    if unknown:
        raise ValueError(f"Unknown border fields: {sorted(unknown)}")

    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    changed = 0
    for row_index in range(r1, r2 + 1):
        for col_index in range(c1, c2 + 1):
            cell = _ensure_cell(sheet, row_index, col_index, capture_baseline=True)
            if cell.get("merge") == "slave":
                continue
            current = copy.deepcopy(cell.get("border") or {})
            for key, value in updates.items():
                if key in {"outline", "diagonalUp", "diagonalDown"}:
                    current[key] = bool(value)
                elif value is None or value == {"style": "none"}:
                    current.pop(key, None)
                else:
                    side_value = copy.deepcopy(value)
                    if side_value.get("color") is not None:
                        color_object = _normalize_color_object(side_value["color"])
                        side_value["_color_raw"] = color_object
                        side_value["color"] = _resolved_rgb(color_object)
                    current[key] = side_value
            cell["border"] = current
            cell["present"] = True
            _mark_dirty(data, "cell_style", f"sheets/{sheet_name}/cells/{_cell_coord(row_index, col_index)}/border")
            changed += 1
    return json.dumps({"cells_updated": changed, "border_patch": _strip_private(updates),
                       "dirty_features": data.get("_dirty_features", [])},
                      default=str, ensure_ascii=False)


# ── 21. Dimensions ────────────────────────────────────────────────────────────

@mcp.tool()
def excel_set_dimension(
    session_key: str,
    sheet_name: str,
    axis: str,
    index: int,
    size: float | None,
) -> str:
    """
    Set the height of a row or the width of a column.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        axis: "row" to set row height, "col" to set column width
        index: 0-based row or column index
        size: Height in points (rows) or width in character units (cols).
              Pass null to reset to auto.

    Returns:
        Confirmation
    """
    import openpyxl.utils
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)

    if axis == "row":
        rows = sheet["rows"]
        if not (0 <= index < len(rows)):
            raise ValueError(f"row index {index} out of range (0–{len(rows)-1})")
        rows[index]["h"] = size
        _mark_dirty(data, "row_properties", f"sheets/{sheet_name}/rows/{index}")
        return f"Set row {index} height to {size!r} in sheet '{sheet_name}'."
    elif axis == "col":
        col_letter = openpyxl.utils.get_column_letter(index + 1)
        if size is None:
            sheet["cw"].pop(col_letter, None)
        else:
            sheet["cw"][col_letter] = size
        _mark_dirty(data, "column_properties", f"sheets/{sheet_name}/columns")
        return f"Set column {index} ({col_letter}) width to {size!r} in sheet '{sheet_name}'."
    else:
        raise ValueError(f"axis must be 'row' or 'col', got {axis!r}")


@mcp.tool()
def excel_set_row_height(
    session_key: str,
    sheet_name: str,
    row_heights: dict,
) -> str:
    """
    Set height for one or more rows in a single call.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        row_heights: Map of {row_index: height}. Height in points.
                     Pass null to reset a row to auto height.
                     Example: {"0": 30, "1": 20, "5": null}

    Returns:
        Confirmation
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]
    updated = 0
    for idx_str, height in row_heights.items():
        idx = int(idx_str)
        if not (0 <= idx < len(rows)):
            raise ValueError(f"row index {idx} out of range (0–{len(rows)-1})")
        rows[idx]["h"] = height
        _mark_dirty(data, "row_properties", f"sheets/{sheet_name}/rows/{idx}")
        updated += 1
    return f"Set height for {updated} row(s) in sheet '{sheet_name}'."


@mcp.tool()
def excel_set_column_width(
    session_key: str,
    sheet_name: str,
    col_widths: dict,
) -> str:
    """
    Set width for one or more columns in a single call.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_widths: Map of {col: width}. Width in character units.
                    Column key accepts letter ("A", "B") or 0-based integer ("0", "1").
                    Pass null to remove an explicit width (resets to default).
                    Example: {"A": 20, "B": 15, "2": 30}

    Returns:
        Confirmation
    """
    import openpyxl.utils
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    updated = 0
    for key, width in col_widths.items():
        try:
            letter = openpyxl.utils.get_column_letter(int(key) + 1)
        except ValueError:
            letter = key.upper()
        if width is None:
            sheet["cw"].pop(letter, None)
        else:
            sheet["cw"][letter] = float(width)
        updated += 1
    if updated:
        _mark_dirty(data, "column_properties", f"sheets/{sheet_name}/columns")
    return f"Set width for {updated} column(s) in sheet '{sheet_name}'."


@mcp.tool()
def excel_autofit_cols(
    session_key: str,
    sheet_name: str,
    col_indices: list[int] | None = None,
    min_width: float = 8.0,
    max_width: float = 60.0,
) -> str:
    """
    Estimate and set column widths based on content length (heuristic approximation).

    openpyxl cannot measure rendered text, so widths are estimated from string
    length, font size, and bold flag. Results are usually close but may need
    manual adjustment.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_indices: 0-based column indices to fit; omit for all columns
        min_width: Minimum column width (default 8.0)
        max_width: Maximum column width cap (default 60.0)

    Returns:
        JSON: {columns_fitted, widths: {col_index: width}}
    """
    import openpyxl.utils
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]

    n_cols = max((len(r["cells"]) for r in rows), default=0)
    targets = col_indices if col_indices is not None else list(range(n_cols))

    updated: dict[int, float] = {}
    for c in targets:
        max_len = 0.0
        for row in rows:
            if c < len(row["cells"]):
                cell = row["cells"][c]
                if cell.get("merge") != "slave" and cell["v"] is not None:
                    text_len = len(str(cell["v"]))
                    size = cell.get("size") or 11
                    factor = 1.2 if cell.get("bold") else 1.0
                    est = text_len * (size / 11) * factor
                    if est > max_len:
                        max_len = est
        width = round(max(min_width, min(max_width, max_len * 1.1 + 2)), 1)
        sheet["cw"][openpyxl.utils.get_column_letter(c + 1)] = width
        updated[c] = width

    if updated:
        _mark_dirty(data, "column_properties", f"sheets/{sheet_name}/columns")
    return json.dumps({"sheet": sheet_name, "columns_fitted": len(updated),
                       "widths": {str(k): v for k, v in updated.items()}}, ensure_ascii=False)


# ── 22. Freeze panes ──────────────────────────────────────────────────────────

@mcp.tool()
def excel_freeze_panes(
    session_key: str,
    sheet_name: str,
    row: int,
    col: int,
) -> str:
    """
    Freeze rows above `row` and/or columns to the left of `col`.

    row=1, col=0 → freeze first row only (most common for headers)
    row=0, col=1 → freeze first column only
    row=1, col=1 → freeze both header row and first column
    row=0, col=0 → unfreeze

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        row: First unfrozen row (0-based); 0 = no row freeze
        col: First unfrozen column (0-based); 0 = no column freeze

    Returns:
        Confirmation with freeze reference cell
    """
    import openpyxl.utils
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    views = copy.deepcopy(
        sheet.get("sheet_views")
        or ([sheet.get("sheet_view")] if sheet.get("sheet_view") else [{}])
    )
    first_view = views[0]
    if row == 0 and col == 0:
        sheet["freeze"] = None
        first_view.pop("pane", None)
        sheet["sheet_views"] = views
        sheet["sheet_view"] = copy.deepcopy(first_view)
        _mark_dirty(data, "sheet_views", f"sheets/{sheet_name}/views")
        return f"Unfrozen panes in sheet '{sheet_name}'."
    col_letter = openpyxl.utils.get_column_letter(col + 1)
    ref = f"{col_letter}{row + 1}"
    sheet["freeze"] = ref
    pane = {
        "topLeftCell": ref,
        "state": "frozen",
    }
    if row > 0:
        pane["ySplit"] = row
    if col > 0:
        pane["xSplit"] = col
    if row > 0 and col > 0:
        pane["activePane"] = "bottomRight"
    elif row > 0:
        pane["activePane"] = "bottomLeft"
    else:
        pane["activePane"] = "topRight"
    first_view["pane"] = pane
    sheet["sheet_views"] = views
    sheet["sheet_view"] = copy.deepcopy(first_view)
    _mark_dirty(data, "sheet_views", f"sheets/{sheet_name}/views")
    frozen = []
    if row > 0:
        frozen.append(f"rows 0–{row-1}")
    if col > 0:
        frozen.append(f"cols 0–{col-1}")
    return f"Freeze → {ref!r} in sheet '{sheet_name}' ({', '.join(frozen)} frozen)."


# ── 23. Data validation ───────────────────────────────────────────────────────

@mcp.tool()
def excel_set_data_validation(
    session_key: str,
    sheet_name: str,
    start_row: int | None = None,
    start_col: int | None = None,
    end_row: int | None = None,
    end_col: int | None = None,
    options: list[str] | None = None,
    allow_blank: bool | None = True,
    validation: dict | None = None,
    sqref: str | list[str] | None = None,
    mode: str = "append",
) -> str:
    """Add, replace, or patch a general Excel data-validation rule."""
    if sqref is None:
        if None in {start_row, start_col, end_row, end_col}:
            raise ValueError("Provide sqref or all four 0-based range bounds.")
        sqref = f"{_cell_coord(start_row, start_col)}:{_cell_coord(end_row, end_col)}"
    if isinstance(sqref, list):
        sqref = " ".join(sqref)
    item = copy.deepcopy(validation or {})
    if options is not None:
        item.setdefault("type", "list")
        item["formula1"] = '"' + ",".join(str(option).replace('"', "") for option in options) + '"'
    if allow_blank is not None:
        item["allowBlank"] = allow_blank
        item["allow_blank"] = allow_blank
    item["sqref"] = str(sqref)
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    validations = sheet.setdefault("validations", [])
    before = copy.deepcopy(validations)
    if mode == "replace":
        validations[:] = [item]
    elif mode == "patch":
        match_index = next((index for index, existing in enumerate(validations)
                            if existing.get("sqref") == item["sqref"]), None)
        if match_index is None:
            validations.append(item)
        else:
            validations[match_index].update(item)
    elif mode == "append":
        validations.append(item)
    else:
        raise ValueError("mode must be append, patch, or replace.")
    # reconstruct_excel prefers a byte-for-byte raw-XML passthrough of the
    # sheet's original <dataValidations> block (captured at load time) over
    # whatever this call just built, for the sake of preserving unknown/
    # unsupported extension attributes on validations this tool doesn't
    # model. That passthrough is now stale: without invalidating it here, a
    # brand-new validation added through this very call would be silently
    # discarded on save (the caller sees a success response, but the file
    # never gains the rule). Drop it so the openpyxl-object path (driven by
    # the up-to-date `validations` list above) is what actually gets written.
    sheet.pop("data_validations_xml", None)
    return _mutation_result(data, "data_validation", before, validations,
                            f"sheets/{sheet_name}/data_validations")


_CF_RULE_RE = re.compile(
    r"<(?:[A-Za-z_][\w.-]*:)?cfRule\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?cfRule>)",
    re.DOTALL,
)


def _cf_elements(xml: str, tag_name: str) -> list[str]:
    pattern = re.compile(
        rf"<(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}\b[^>]*"
        rf"(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}>)",
        re.DOTALL,
    )
    return [match.group(0) for match in pattern.finditer(xml)]


def _cf_xml_attributes(xml: str) -> dict[str, str]:
    opening = xml[:xml.find(">") + 1] if ">" in xml else xml
    return {
        name: html.unescape(value)
        for name, value in re.findall(r'([A-Za-z_][\w:.-]*)\s*=\s*"([^"]*)"', opening)
    }


def _cf_set_xml_attribute(xml: str, name: str, value) -> str:
    end = xml.find(">")
    if end < 0:
        raise ValueError("Malformed conditional-formatting XML.")
    opening = xml[:end + 1]
    tail = xml[end + 1:]
    attr_pattern = re.compile(rf'(\s{re.escape(name)}\s*=\s*)"[^"]*"')
    if value is None:
        opening = attr_pattern.sub("", opening, count=1)
    else:
        if isinstance(value, bool):
            value = "1" if value else "0"
        encoded = html.escape(str(value), quote=True)
        if attr_pattern.search(opening):
            opening = attr_pattern.sub(lambda match: f'{match.group(1)}"{encoded}"', opening, count=1)
        else:
            insert_at = -2 if opening.endswith("/>") else -1
            opening = opening[:insert_at] + f' {name}="{encoded}"' + opening[insert_at:]
    return opening + tail


def _cf_replace_elements(xml: str, tag_name: str, replacements: list[str]) -> str:
    pattern = re.compile(
        rf"<(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}\b[^>]*"
        rf"(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}>)",
        re.DOTALL,
    )
    matches = list(pattern.finditer(xml))
    if matches:
        start = matches[0].start()
        end = matches[-1].end()
        return xml[:start] + "".join(replacements) + xml[end:]
    closing = re.search(r"</(?:[A-Za-z_][\w.-]*:)?cfRule>", xml)
    if not closing:
        if xml.rstrip().endswith("/>"):
            base = xml.rstrip()
            return base[:-2] + ">" + "".join(replacements) + "</cfRule>"
        raise ValueError("Malformed conditional-formatting rule XML.")
    return xml[:closing.start()] + "".join(replacements) + xml[closing.start():]


def _normalize_cf_sqref(sqref: str | list[str]) -> str:
    from openpyxl.utils.cell import range_boundaries

    value = " ".join(str(item) for item in sqref) if isinstance(sqref, list) else str(sqref)
    tokens = value.split()
    if not tokens:
        raise ValueError("Conditional-formatting sqref cannot be empty.")
    for token in tokens:
        if "!" in token or "," in token:
            raise ValueError(f"Conditional-formatting sqref must be worksheet-local, got {token!r}.")
        try:
            bounds = range_boundaries(token.replace("$", ""))
        except Exception as exc:
            raise ValueError(f"Invalid conditional-formatting range {token!r}.") from exc
        if None in bounds:
            raise ValueError(f"Invalid conditional-formatting range {token!r}.")
    return " ".join(tokens)


def _cf_block_sqref(block: str) -> str:
    return _cf_xml_attributes(block).get("sqref", "")


def _cf_rule_records(blocks: list[str]) -> list[dict]:
    records = []
    for block_index, block in enumerate(blocks):
        sqref = _cf_block_sqref(block)
        for rule_index, match in enumerate(_CF_RULE_RE.finditer(block)):
            raw = match.group(0)
            attrs = _cf_xml_attributes(raw)
            records.append({
                "rule_id": f"b{block_index}:r{rule_index}",
                "block_index": block_index,
                "rule_index": rule_index,
                "sqref": sqref,
                "raw": raw,
                "attrs": attrs,
            })
    return records


def _cf_rule_record(blocks: list[str], rule_id: str) -> dict:
    match = re.fullmatch(r"b(\d+):r(\d+)", str(rule_id))
    if not match:
        raise ValueError(f"Invalid conditional-formatting rule_id {rule_id!r}.")
    block_index, rule_index = int(match.group(1)), int(match.group(2))
    for record in _cf_rule_records(blocks):
        if record["block_index"] == block_index and record["rule_index"] == rule_index:
            return record
    raise ValueError(f"Conditional-formatting rule {rule_id!r} not found.")


def _cf_replace_rule(blocks: list[str], block_index: int, rule_index: int, replacement: str | None) -> None:
    block = blocks[block_index]
    matches = list(_CF_RULE_RE.finditer(block))
    if rule_index < 0 or rule_index >= len(matches):
        raise ValueError("Conditional-formatting rule index is out of range.")
    target = matches[rule_index]
    updated = block[:target.start()] + (replacement or "") + block[target.end():]
    if _CF_RULE_RE.search(updated):
        blocks[block_index] = updated
    else:
        blocks.pop(block_index)


def _cf_append_rule(blocks: list[str], sqref: str, rule_xml: str) -> tuple[int, int]:
    for block_index, block in enumerate(blocks):
        if _cf_block_sqref(block) != sqref:
            continue
        closing = re.search(r"</(?:[A-Za-z_][\w.-]*:)?conditionalFormatting>", block)
        if not closing:
            raise ValueError("Malformed conditional-formatting block XML.")
        rule_index = len(list(_CF_RULE_RE.finditer(block)))
        blocks[block_index] = block[:closing.start()] + rule_xml + block[closing.start():]
        return block_index, rule_index
    blocks.append(f'<conditionalFormatting sqref="{html.escape(sqref, quote=True)}">{rule_xml}</conditionalFormatting>')
    return len(blocks) - 1, 0


def _cf_adjust_priorities(blocks: list[str], transform) -> None:
    for block_index, block in enumerate(list(blocks)):
        rule_index = 0

        def replace(match):
            nonlocal rule_index
            raw = match.group(0)
            attrs = _cf_xml_attributes(raw)
            try:
                priority = int(attrs.get("priority", 0))
            except ValueError:
                priority = 0
            updated = transform(block_index, rule_index, priority)
            rule_index += 1
            return _cf_set_xml_attribute(raw, "priority", updated) if updated is not None else raw

        blocks[block_index] = _CF_RULE_RE.sub(replace, block)


def _cf_compact_priorities(blocks: list[str]) -> None:
    records = _cf_rule_records(blocks)
    ordered = sorted(
        records,
        key=lambda item: (
            int(item["attrs"].get("priority", 0)) if str(item["attrs"].get("priority", "")).isdigit() and int(item["attrs"]["priority"]) > 0 else 10**9,
            item["block_index"],
            item["rule_index"],
        ),
    )
    assignments = {
        (record["block_index"], record["rule_index"]): priority
        for priority, record in enumerate(ordered, 1)
    }
    _cf_adjust_priorities(blocks, lambda block_index, rule_index, _priority: assignments[(block_index, rule_index)])


def _cf_color_xml(color, tag_name: str = "color") -> str:
    if isinstance(color, str):
        color = {"rgb": color}
    if not isinstance(color, dict):
        raise ValueError("Conditional-formatting color must be a string or object.")
    attrs = []
    for key in ("rgb", "indexed", "theme", "tint", "auto"):
        if color.get(key) is None:
            continue
        value = color[key]
        if isinstance(value, bool):
            value = "1" if value else "0"
        attrs.append(f'{key}="{html.escape(str(value), quote=True)}"')
    if not attrs:
        raise ValueError("Conditional-formatting color requires rgb, indexed, theme, tint, or auto.")
    return f"<{tag_name} {' '.join(attrs)}/>"


def _cf_cfvo_xml(value) -> str:
    if isinstance(value, (int, float, str)):
        value = {"type": "num", "val": value}
    if not isinstance(value, dict) or not value.get("type"):
        raise ValueError("Conditional-formatting cfvo requires a type.")
    attrs = [f'type="{html.escape(str(value["type"]), quote=True)}"']
    if value.get("val") is not None:
        attrs.append(f'val="{html.escape(str(value["val"]), quote=True)}"')
    if value.get("gte") is not None:
        attrs.append(f'gte="{1 if bool(value["gte"]) else 0}"')
    return f"<cfvo {' '.join(attrs)}/>"


def _cf_complex_rule_xml(rule_type: str, model: dict) -> str:
    if rule_type == "colorScale":
        values = model.get("cfvo") or model.get("thresholds") or []
        colors = model.get("colors") or []
        if len(values) not in {2, 3} or len(colors) != len(values):
            raise ValueError("colorScale requires two or three matching cfvo and color entries.")
        return "<colorScale>" + "".join(_cf_cfvo_xml(item) for item in values) + "".join(_cf_color_xml(item) for item in colors) + "</colorScale>"
    if rule_type == "dataBar":
        values = model.get("cfvo") or model.get("thresholds") or []
        if len(values) != 2 or model.get("color") is None:
            raise ValueError("dataBar requires exactly two cfvo entries and one color.")
        attrs = []
        for key in ("minLength", "maxLength", "showValue", "gradient", "border", "direction", "axisPosition"):
            if model.get(key) is None:
                continue
            value = model[key]
            if isinstance(value, bool):
                value = "1" if value else "0"
            attrs.append(f'{key}="{html.escape(str(value), quote=True)}"')
        opening = "<dataBar" + (" " + " ".join(attrs) if attrs else "") + ">"
        return opening + "".join(_cf_cfvo_xml(item) for item in values) + _cf_color_xml(model["color"]) + "</dataBar>"
    if rule_type == "iconSet":
        values = model.get("cfvo") or model.get("thresholds") or []
        if len(values) < 3:
            raise ValueError("iconSet requires at least three cfvo entries.")
        attrs = []
        for key in ("iconSet", "showValue", "percent", "reverse"):
            if model.get(key) is None:
                continue
            value = model[key]
            if isinstance(value, bool):
                value = "1" if value else "0"
            attrs.append(f'{key}="{html.escape(str(value), quote=True)}"')
        opening = "<iconSet" + (" " + " ".join(attrs) if attrs else "") + ">"
        return opening + "".join(_cf_cfvo_xml(item) for item in values) + "</iconSet>"
    raise ValueError(f"Unsupported conditional-formatting type {rule_type!r}.")


def _cf_dxf_xml(model) -> str:
    if isinstance(model, str):
        if not re.search(r"<(?:[A-Za-z_][\w.-]*:)?dxf\b", model):
            raise ValueError("Raw differential style must contain a dxf element.")
        return model
    if not isinstance(model, dict):
        raise ValueError("Differential style must be an object or raw dxf XML.")
    children = []
    font = model.get("font")
    if isinstance(font, dict):
        font_parts = []
        for key, tag in (("bold", "b"), ("italic", "i"), ("strike", "strike"), ("outline", "outline"), ("shadow", "shadow"), ("condense", "condense"), ("extend", "extend")):
            if font.get(key) is True:
                font_parts.append(f"<{tag}/>")
            elif font.get(key) is False:
                font_parts.append(f'<{tag} val="0"/>')
        underline = font.get("underline", font.get("u"))
        if underline:
            font_parts.append("<u/>" if underline is True or underline == "single" else f'<u val="{html.escape(str(underline), quote=True)}"/>')
        if font.get("name") is not None:
            font_parts.append(f'<name val="{html.escape(str(font["name"]), quote=True)}"/>')
        if font.get("size", font.get("sz")) is not None:
            font_parts.append(f'<sz val="{html.escape(str(font.get("size", font.get("sz"))), quote=True)}"/>')
        if font.get("color") is not None:
            font_parts.append(_cf_color_xml(font["color"]))
        for key, tag in (("vertAlign", "vertAlign"), ("charset", "charset"), ("family", "family"), ("scheme", "scheme")):
            if font.get(key) is not None:
                font_parts.append(f'<{tag} val="{html.escape(str(font[key]), quote=True)}"/>')
        children.append("<font>" + "".join(font_parts) + "</font>")
    number_format = model.get("numFmt", model.get("number_format"))
    if number_format is not None:
        if isinstance(number_format, str):
            number_format = {"formatCode": number_format, "numFmtId": 0}
        children.append(
            f'<numFmt numFmtId="{int(number_format.get("numFmtId", 0))}" '
            f'formatCode="{html.escape(str(number_format["formatCode"]), quote=True)}"/>'
        )
    fill = model.get("fill")
    if fill is not None:
        if isinstance(fill, str):
            fill = {"fgColor": fill, "patternType": "solid"}
        pattern = fill.get("patternType", fill.get("pattern_type", "solid"))
        foreground = fill.get("fgColor", fill.get("foreground", fill.get("color")))
        background = fill.get("bgColor", fill.get("background"))
        fill_parts = []
        if foreground is not None:
            fill_parts.append(_cf_color_xml(foreground, "fgColor"))
        if background is not None:
            fill_parts.append(_cf_color_xml(background, "bgColor"))
        children.append(f'<fill><patternFill patternType="{html.escape(str(pattern), quote=True)}">' + "".join(fill_parts) + "</patternFill></fill>")
    alignment = model.get("alignment")
    if isinstance(alignment, dict):
        attrs = []
        for key, value in alignment.items():
            if value is None:
                continue
            if isinstance(value, bool):
                value = "1" if value else "0"
            attrs.append(f'{key}="{html.escape(str(value), quote=True)}"')
        children.append("<alignment" + (" " + " ".join(attrs) if attrs else "") + "/>")
    border = model.get("border")
    if isinstance(border, dict):
        border_parts = []
        for side_name in ("left", "right", "top", "bottom", "diagonal", "vertical", "horizontal", "start", "end"):
            side = border.get(side_name)
            if side is None:
                continue
            if isinstance(side, str):
                side = {"style": side}
            style_attr = f' style="{html.escape(str(side["style"]), quote=True)}"' if side.get("style") else ""
            color_xml = _cf_color_xml(side["color"]) if side.get("color") is not None else ""
            border_parts.append(f"<{side_name}{style_attr}>{color_xml}</{side_name}>")
        children.append("<border>" + "".join(border_parts) + "</border>")
    protection = model.get("protection")
    if isinstance(protection, dict):
        attrs = []
        for key, value in protection.items():
            if value is None:
                continue
            if isinstance(value, bool):
                value = "1" if value else "0"
            attrs.append(f'{key}="{html.escape(str(value), quote=True)}"')
        children.append("<protection" + (" " + " ".join(attrs) if attrs else "") + "/>")
    return "<dxf>" + "".join(children) + "</dxf>"


def _cf_typed_attributes(xml: str) -> dict:
    boolean_keys = {
        "auto", "wrapText", "shrinkToFit", "justifyLastLine", "locked", "hidden",
        "diagonalUp", "diagonalDown", "outline", "shadow", "condense", "extend",
        "showValue", "gradient", "border", "percent", "reverse",
    }
    result = {}
    for key, value in _cf_xml_attributes(xml).items():
        lowered = value.lower()
        if key in boolean_keys and lowered in {"true", "false", "1", "0"}:
            result[key] = lowered in {"true", "1"}
        else:
            try:
                result[key] = int(value)
            except ValueError:
                try:
                    result[key] = float(value)
                except ValueError:
                    result[key] = value
    return result


def _cf_parse_dxf_xml(raw: str) -> dict:
    model = {}
    fonts = _cf_elements(raw, "font")
    if fonts:
        font_xml = fonts[0]
        font = {}
        for key, tag in (("bold", "b"), ("italic", "i"), ("strike", "strike"), ("outline", "outline"), ("shadow", "shadow"), ("condense", "condense"), ("extend", "extend")):
            elements = _cf_elements(font_xml, tag)
            if elements:
                value = _cf_xml_attributes(elements[0]).get("val", "1")
                font[key] = value.lower() not in {"0", "false"}
        underlines = _cf_elements(font_xml, "u")
        if underlines:
            font["underline"] = _cf_xml_attributes(underlines[0]).get("val", "single")
        for key, tag in (("name", "name"), ("size", "sz"), ("vertAlign", "vertAlign"), ("charset", "charset"), ("family", "family"), ("scheme", "scheme")):
            elements = _cf_elements(font_xml, tag)
            if elements:
                value = _cf_xml_attributes(elements[0]).get("val")
                if value is not None:
                    if key == "size":
                        try:
                            value = float(value)
                        except ValueError:
                            pass
                    font[key] = value
        colors = _cf_elements(font_xml, "color")
        if colors:
            font["color"] = _cf_typed_attributes(colors[0])
        model["font"] = font
    number_formats = _cf_elements(raw, "numFmt")
    if number_formats:
        model["numFmt"] = _cf_typed_attributes(number_formats[0])
    fills = _cf_elements(raw, "fill")
    if fills:
        pattern_fills = _cf_elements(fills[0], "patternFill")
        if pattern_fills:
            fill = _cf_typed_attributes(pattern_fills[0])
            foreground = _cf_elements(pattern_fills[0], "fgColor")
            background = _cf_elements(pattern_fills[0], "bgColor")
            if foreground:
                fill["fgColor"] = _cf_typed_attributes(foreground[0])
            if background:
                fill["bgColor"] = _cf_typed_attributes(background[0])
            model["fill"] = fill
    alignments = _cf_elements(raw, "alignment")
    if alignments:
        model["alignment"] = _cf_typed_attributes(alignments[0])
    borders = _cf_elements(raw, "border")
    if borders:
        border = {}
        for side_name in ("left", "right", "top", "bottom", "diagonal", "vertical", "horizontal", "start", "end"):
            sides = _cf_elements(borders[0], side_name)
            if not sides:
                continue
            side = _cf_typed_attributes(sides[0])
            colors = _cf_elements(sides[0], "color")
            if colors:
                side["color"] = _cf_typed_attributes(colors[0])
            border[side_name] = side
        model["border"] = border
    protections = _cf_elements(raw, "protection")
    if protections:
        model["protection"] = _cf_typed_attributes(protections[0])
    return model


def _cf_resolve_dxf(data: dict, rule: dict, existing_id: int | None = None, *, clear_none: bool = False) -> tuple[int | None, str | None]:
    blocks = _dxf_blocks(data.get("dxfs_xml"))
    dxf_id = existing_id
    if "dxf" in rule:
        if rule["dxf"] is None:
            if clear_none:
                dxf_id = None
        else:
            dxf_model = rule["dxf"]
            if isinstance(dxf_model, dict) and existing_id is not None and 0 <= existing_id < len(blocks) and rule.get("dxf_mode", "patch") == "patch":
                merged = _cf_parse_dxf_xml(blocks[existing_id])
                _deep_patch_mapping(merged, dxf_model, clear_none=clear_none)
                dxf_model = merged
            raw = _cf_dxf_xml(dxf_model)
            try:
                dxf_id = blocks.index(raw)
            except ValueError:
                dxf_id = len(blocks)
                blocks.append(raw)
    elif "dxfId" in rule:
        if rule["dxfId"] is None and clear_none:
            dxf_id = None
        elif rule["dxfId"] is not None:
            dxf_id = int(rule["dxfId"])
            if dxf_id < 0 or dxf_id >= len(blocks):
                raise ValueError(f"dxfId {dxf_id} does not exist.")
    return dxf_id, _render_dxfs_xml(data.get("dxfs_xml"), blocks)


def _cf_rule_type(value) -> str:
    aliases = {
        "formula": "expression",
        "cell_is": "cellIs",
        "color_scale": "colorScale",
        "data_bar": "dataBar",
        "icon_set": "iconSet",
    }
    normalized = aliases.get(str(value), str(value))
    if normalized not in {"cellIs", "expression", "colorScale", "dataBar", "iconSet"}:
        raise ValueError(f"Unsupported conditional-formatting type {value!r}.")
    return normalized


def _cf_normalize_rule_input(rule: dict) -> dict:
    if not isinstance(rule, dict):
        raise ValueError("Conditional-formatting rule must be an object.")
    result = copy.deepcopy(rule)
    aliases = {
        "stop_if_true": "stopIfTrue",
        "dxf_id": "dxfId",
        "color_scale": "colorScale",
        "data_bar": "dataBar",
        "icon_set": "iconSet",
        "formulas": "formula",
    }
    for source, target in aliases.items():
        if source in result:
            result[target] = result.pop(source)
    if "type" in result:
        result["type"] = _cf_rule_type(result["type"])
    return result


def _cf_formula_values(rule: dict) -> list[str]:
    value = rule.get("formula")
    if isinstance(value, str):
        return [value]
    if isinstance(value, list):
        return [str(item) for item in value]
    return []


def _cf_build_rule_xml(rule: dict, priority: int, dxf_id: int | None) -> str:
    rule_type = _cf_rule_type(rule.get("type"))
    attrs = {"type": rule_type, "priority": int(priority)}
    for key in ("stopIfTrue", "aboveAverage", "percent", "bottom", "operator", "text", "timePeriod", "rank", "stdDev", "equalAverage"):
        if rule.get(key) is not None:
            attrs[key] = rule[key]
    if dxf_id is not None:
        attrs["dxfId"] = int(dxf_id)
    attr_text = []
    for key, value in attrs.items():
        if isinstance(value, bool):
            value = "1" if value else "0"
        attr_text.append(f'{key}="{html.escape(str(value), quote=True)}"')
    children = []
    if rule_type in {"cellIs", "expression"}:
        formulas = _cf_formula_values(rule)
        if not formulas:
            raise ValueError(f"{rule_type} conditional formatting requires formula values.")
        if rule_type == "cellIs":
            operator = rule.get("operator")
            if not operator:
                raise ValueError("cellIs conditional formatting requires an operator.")
            expected = 2 if operator in {"between", "notBetween"} else 1
            if len(formulas) != expected:
                raise ValueError(f"cellIs operator {operator!r} requires {expected} formula value(s).")
        children.extend(f"<formula>{html.escape(value)}</formula>" for value in formulas)
    else:
        children.append(_cf_complex_rule_xml(rule_type, rule.get(rule_type) or {}))
    return "<cfRule " + " ".join(attr_text) + ">" + "".join(children) + "</cfRule>"


def _cf_parse_complex_rule(raw: str, rule_type: str) -> dict | None:
    elements = _cf_elements(raw, rule_type)
    if not elements:
        return None
    element = elements[0]
    result = _cf_typed_attributes(element)
    result["cfvo"] = [_cf_xml_attributes(item) for item in _cf_elements(element, "cfvo")]
    if rule_type in {"colorScale", "dataBar"}:
        colors = [_cf_xml_attributes(item) for item in _cf_elements(element, "color")]
        if rule_type == "colorScale":
            result["colors"] = colors
        elif colors:
            result["color"] = colors[0]
    return result


def _cf_rule_model(record: dict, dxf_blocks: list[str], *, include_raw_xml: bool = False) -> dict:
    attrs = record["attrs"]
    model = {
        "rule_id": record["rule_id"],
        "sqref": record["sqref"],
        "type": attrs.get("type"),
    }
    for key in ("priority", "dxfId", "rank", "stdDev"):
        if attrs.get(key) is not None:
            try:
                model[key] = int(attrs[key])
            except ValueError:
                model[key] = attrs[key]
    for key in ("stopIfTrue", "aboveAverage", "percent", "bottom", "equalAverage"):
        if attrs.get(key) is not None:
            model[key] = attrs[key].lower() in {"1", "true"}
    for key in ("operator", "text", "timePeriod"):
        if attrs.get(key) is not None:
            model[key] = attrs[key]
    formulas = [
        html.unescape(re.sub(r"<[^>]+>", "", item))
        for item in _cf_elements(record["raw"], "formula")
    ]
    if formulas:
        model["formula"] = formulas
    if model.get("type") in {"colorScale", "dataBar", "iconSet"}:
        complex_model = _cf_parse_complex_rule(record["raw"], model["type"])
        if complex_model is not None:
            model[model["type"]] = complex_model
    dxf_id = model.get("dxfId")
    if isinstance(dxf_id, int) and 0 <= dxf_id < len(dxf_blocks):
        model["dxf"] = _cf_parse_dxf_xml(dxf_blocks[dxf_id])
        model["dxf_xml"] = dxf_blocks[dxf_id]
    if include_raw_xml:
        model["raw_xml"] = record["raw"]
    return model


def _cf_patch_rule_xml(raw: str, candidate: dict, updates: dict, priority: int, dxf_id: int | None, *, clear_none: bool) -> str:
    current_type = _cf_xml_attributes(raw).get("type")
    if candidate.get("type") != current_type:
        return _cf_build_rule_xml(candidate, priority, dxf_id)
    updated = _cf_set_xml_attribute(raw, "priority", priority)
    for key in ("stopIfTrue", "aboveAverage", "percent", "bottom", "operator", "text", "timePeriod", "rank", "stdDev", "equalAverage"):
        if key in updates:
            value = updates[key]
            updated = _cf_set_xml_attribute(updated, key, None if value is None and clear_none else value)
    if "dxf" in updates or "dxfId" in updates:
        updated = _cf_set_xml_attribute(updated, "dxfId", dxf_id)
    if "formula" in updates:
        formulas = _cf_formula_values(candidate)
        updated = _cf_replace_elements(
            updated,
            "formula",
            [f"<formula>{html.escape(value)}</formula>" for value in formulas],
        )
    rule_type = candidate["type"]
    if rule_type in {"colorScale", "dataBar", "iconSet"} and rule_type in updates:
        updated = _cf_replace_elements(updated, rule_type, [_cf_complex_rule_xml(rule_type, candidate[rule_type])])
    _cf_build_rule_xml(candidate, priority, dxf_id)
    return updated


@mcp.tool()
def excel_get_conditional_formats(
    session_key: str,
    sheet_name: str,
    sqref: str | list[str] | None = None,
    include_raw_xml: bool = False,
) -> str:
    """List conditional-formatting rules with stable per-session rule IDs."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    expected_sqref = _normalize_cf_sqref(sqref) if sqref is not None else None
    records = _cf_rule_records(sheet.get("cf_xml") or [])
    dxf_blocks = _dxf_blocks(data.get("dxfs_xml"))
    rules = [
        _cf_rule_model(record, dxf_blocks, include_raw_xml=include_raw_xml)
        for record in records
        if expected_sqref is None or record["sqref"] == expected_sqref
    ]
    return json.dumps({"count": len(rules), "rules": rules}, default=str, ensure_ascii=False)


@mcp.tool()
def excel_add_conditional_format(
    session_key: str,
    sheet_name: str,
    sqref: str | list[str],
    rule: dict,
) -> str:
    """Add a cellIs, expression, colorScale, dataBar, or iconSet rule."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    normalized_sqref = _normalize_cf_sqref(sqref)
    normalized_rule = _cf_normalize_rule_input(rule)
    blocks = copy.deepcopy(sheet.get("cf_xml") or [])
    _cf_compact_priorities(blocks)
    count = len(_cf_rule_records(blocks))
    requested_priority = normalized_rule.get("priority", count + 1)
    priority = int(requested_priority)
    if priority < 1 or priority > count + 1:
        raise ValueError(f"priority must be between 1 and {count + 1}.")
    _cf_adjust_priorities(
        blocks,
        lambda _block, _rule, current: current + 1 if current >= priority else current,
    )
    dxf_id, dxfs_xml = _cf_resolve_dxf(data, normalized_rule)
    rule_xml = _cf_build_rule_xml(normalized_rule, priority, dxf_id)
    before = [
        _cf_rule_model(record, _dxf_blocks(data.get("dxfs_xml")))
        for record in _cf_rule_records(sheet.get("cf_xml") or [])
    ]
    block_index, rule_index = _cf_append_rule(blocks, normalized_sqref, rule_xml)
    sheet["cf_xml"] = blocks
    data["dxfs_xml"] = dxfs_xml
    after_records = _cf_rule_records(blocks)
    dxf_blocks = _dxf_blocks(dxfs_xml)
    after = [_cf_rule_model(record, dxf_blocks) for record in after_records]
    rule_id = f"b{block_index}:r{rule_index}"
    return _mutation_result(
        data,
        "conditional_formatting",
        before,
        after,
        f"sheets/{sheet_name}/conditional_formatting/{rule_id}",
        rule_id=rule_id,
    )


@mcp.tool()
def excel_update_conditional_format(
    session_key: str,
    sheet_name: str,
    rule_id: str,
    updates: dict,
) -> str:
    """Patch one conditional-formatting rule without replacing omitted XML siblings."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    normalized_updates = _cf_normalize_rule_input(updates)
    clear_none = bool(normalized_updates.pop("clear_nulls", False))
    blocks = copy.deepcopy(sheet.get("cf_xml") or [])
    _cf_compact_priorities(blocks)
    record = _cf_rule_record(blocks, rule_id)
    old_priority = int(record["attrs"].get("priority", 1))
    count = len(_cf_rule_records(blocks))
    new_priority = int(normalized_updates.get("priority", old_priority))
    if new_priority < 1 or new_priority > count:
        raise ValueError(f"priority must be between 1 and {count}.")
    existing_dxf = int(record["attrs"]["dxfId"]) if str(record["attrs"].get("dxfId", "")).isdigit() else None
    dxf_id, dxfs_xml = _cf_resolve_dxf(data, normalized_updates, existing_dxf, clear_none=clear_none)
    dxf_blocks_before = _dxf_blocks(data.get("dxfs_xml"))
    before = [_cf_rule_model(item, dxf_blocks_before) for item in _cf_rule_records(sheet.get("cf_xml") or [])]
    current_model = _cf_rule_model(record, dxf_blocks_before)
    candidate = copy.deepcopy(current_model)
    patch = {
        key: value for key, value in normalized_updates.items()
        if key not in {"priority", "sqref", "dxf", "dxfId", "raw_xml", "dxf_xml"}
    }
    _deep_patch_mapping(candidate, patch, clear_none=clear_none)
    candidate["priority"] = new_priority
    candidate["dxfId"] = dxf_id
    candidate.pop("rule_id", None)
    candidate.pop("sqref", None)
    candidate.pop("raw_xml", None)
    candidate.pop("dxf_xml", None)
    _cf_build_rule_xml(candidate, new_priority, dxf_id)

    if new_priority < old_priority:
        _cf_adjust_priorities(
            blocks,
            lambda block_index, index, current: current + 1
            if (block_index, index) != (record["block_index"], record["rule_index"]) and new_priority <= current < old_priority
            else current,
        )
    elif new_priority > old_priority:
        _cf_adjust_priorities(
            blocks,
            lambda block_index, index, current: current - 1
            if (block_index, index) != (record["block_index"], record["rule_index"]) and old_priority < current <= new_priority
            else current,
        )
    record = _cf_rule_record(blocks, rule_id)
    updated_raw = _cf_patch_rule_xml(
        record["raw"],
        candidate,
        normalized_updates,
        new_priority,
        dxf_id,
        clear_none=clear_none,
    )
    target_sqref = _normalize_cf_sqref(normalized_updates["sqref"]) if "sqref" in normalized_updates else record["sqref"]
    if target_sqref == record["sqref"]:
        _cf_replace_rule(blocks, record["block_index"], record["rule_index"], updated_raw)
        new_rule_id = rule_id
    else:
        _cf_replace_rule(blocks, record["block_index"], record["rule_index"], None)
        block_index, rule_index = _cf_append_rule(blocks, target_sqref, updated_raw)
        new_rule_id = f"b{block_index}:r{rule_index}"
    sheet["cf_xml"] = blocks
    data["dxfs_xml"] = dxfs_xml
    dxf_blocks_after = _dxf_blocks(dxfs_xml)
    after = [_cf_rule_model(item, dxf_blocks_after) for item in _cf_rule_records(blocks)]
    return _mutation_result(
        data,
        "conditional_formatting",
        before,
        after,
        f"sheets/{sheet_name}/conditional_formatting/{new_rule_id}",
        rule_id=new_rule_id,
    )


@mcp.tool()
def excel_delete_conditional_format(session_key: str, sheet_name: str, rule_id: str) -> str:
    """Delete one conditional-formatting rule and compact remaining priorities."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    blocks = copy.deepcopy(sheet.get("cf_xml") or [])
    _cf_compact_priorities(blocks)
    record = _cf_rule_record(blocks, rule_id)
    deleted_priority = int(record["attrs"].get("priority", 1))
    dxf_blocks = _dxf_blocks(data.get("dxfs_xml"))
    before = [_cf_rule_model(item, dxf_blocks) for item in _cf_rule_records(sheet.get("cf_xml") or [])]
    _cf_replace_rule(blocks, record["block_index"], record["rule_index"], None)
    _cf_adjust_priorities(
        blocks,
        lambda _block, _rule, current: current - 1 if current > deleted_priority else current,
    )
    if blocks:
        sheet["cf_xml"] = blocks
    else:
        sheet.pop("cf_xml", None)
    after = [_cf_rule_model(item, dxf_blocks) for item in _cf_rule_records(blocks)]
    return _mutation_result(
        data,
        "conditional_formatting",
        before,
        after,
        f"sheets/{sheet_name}/conditional_formatting/{rule_id}",
    )


# ── 24. Find rows / Fill column ───────────────────────────────────────────────

@mcp.tool()
def excel_find_rows(
    session_key: str,
    sheet_name: str,
    col_index: int,
    value: str | None = None,
    pattern: str | None = None,
    start_row: int = 0,
    end_row: int | None = None,
) -> str:
    """
    Find all rows where a column cell matches a value or regex pattern.

    value and pattern are mutually exclusive — provide exactly one.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_index: 0-based column index to search in
        value: Exact match (compared as string)
        pattern: Python regex pattern (re.search)
        start_row: 0-based start row (inclusive), default 0
        end_row: 0-based end row (exclusive); omit for all rows

    Returns:
        JSON array of {row_index, values} for each matching row
    """
    import re
    if value is None and pattern is None:
        raise ValueError("Provide either value or pattern.")
    if value is not None and pattern is not None:
        raise ValueError("value and pattern are mutually exclusive.")

    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]

    results = []
    for r_idx, row in enumerate(rows[start_row:end_row], start=start_row):
        cells = row["cells"]
        if col_index >= len(cells):
            continue
        cell_val = cells[col_index]["v"]
        cell_str = str(cell_val) if cell_val is not None else ""
        matched = (cell_str == str(value)) if value is not None else bool(re.search(pattern, cell_str))
        if matched:
            results.append({
                "row_index": r_idx,
                "values": [cd["v"] if cd.get("merge") != "slave" else None for cd in cells],
            })
    return json.dumps(results, default=str, ensure_ascii=False)


@mcp.tool()
def excel_fill_column(
    session_key: str,
    sheet_name: str,
    col_index: int,
    start_row: int,
    end_row: int,
    value: str | int | float | None = None,
    sequence_start: int | None = None,
    step: int = 1,
) -> str:
    """
    Fill a column range with a constant value or an auto-incrementing sequence.

    Constant fill: pass value — every cell gets the same value.
    Sequence fill: pass sequence_start — cells get sequence_start,
                   sequence_start+step, sequence_start+2×step, …

    start_row and end_row are both inclusive (0-based).
    Slave cells of merged regions are skipped.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        col_index: 0-based column index to fill
        start_row: 0-based start row (inclusive)
        end_row: 0-based end row (inclusive)
        value: Constant fill value (mutually exclusive with sequence_start)
        sequence_start: First integer of the sequence
        step: Increment between sequence values (default 1)

    Returns:
        Summary: cells filled
    """
    if value is not None and sequence_start is not None:
        raise ValueError("value and sequence_start are mutually exclusive.")

    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    rows = sheet["rows"]
    n = len(rows)
    if not (0 <= start_row <= end_row < n):
        raise ValueError(f"Row range [{start_row}, {end_row}] out of bounds (0–{n-1})")

    cells_filled = 0
    seq = sequence_start
    for r in range(start_row, end_row + 1):
        row_cells = rows[r]["cells"]
        if col_index >= len(row_cells):
            continue
        cell = row_cells[col_index]
        if cell.get("merge") == "slave":
            continue
        if sequence_start is not None:
            _store_cell_value(cell, seq, sheet.get("_implicit_cell_defaults"))
            seq += step
        else:
            _store_cell_value(cell, value, sheet.get("_implicit_cell_defaults"))
        _mark_dirty(data, "cells", f"sheets/{sheet_name}/cells/{_cell_coord(r, col_index)}")
        cells_filled += 1

    return f"Filled {cells_filled} cell(s) in col {col_index}, rows {start_row}–{end_row} in sheet '{sheet_name}'."


# ── 25. Fill rows (stamp pattern) ─────────────────────────────────────────────

@mcp.tool()
def excel_fill_rows(
    session_key: str,
    sheet_name: str,
    template_row: int,
    after_index: int,
    count: int,
) -> str:
    """
    Clone a template row N times and insert all copies in one call.

    More efficient than clone_rows + insert_rows when inserting many rows with
    the same format (e.g. stamping a formatted template row for a data table).
    All cell values, styles, and borders from the template are preserved in every copy.

    Args:
        session_key: Key returned by excel_load
        sheet_name: Name of the sheet
        template_row: 0-based index of the row to clone as the template
        after_index: Insert the block AFTER this 0-based row index; use -1 to prepend
        count: Number of copies to insert (must be > 0)

    Returns:
        Confirmation with new row count
    """
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    n = len(sheet["rows"])
    if not (0 <= template_row < n):
        raise ValueError(f"template_row {template_row} out of range (0–{n - 1})")
    if count <= 0:
        raise ValueError(f"count must be > 0, got {count}")

    template = sheet["rows"][template_row]
    new_rows = [copy.deepcopy(template) for _ in range(count)]
    pos = after_index + 1
    _apply_row_insert(data, sheet, pos, new_rows)
    return (
        f"Inserted {count} copy/copies of row {template_row} after index {after_index}. "
        f"Sheet '{sheet_name}' now has {len(sheet['rows'])} rows."
    )


def _apply_font_payload(cell: dict, payload: dict) -> None:
    raw = copy.deepcopy(cell.get("_font_raw") or _whole_cell_font(cell))
    aliases = {
        "name": "font", "size": "size", "bold": "bold", "italic": "italic",
        "underline": "uline", "strike": "strike", "vertAlign": "vAlign",
    }
    for key, value in payload.items():
        if key == "color":
            color = _normalize_color_object(value)
            if color is None:
                raw.pop("color", None)
                cell["fcolor"] = None
            else:
                raw["color"] = color
                cell["fcolor"] = _resolved_rgb(color)
        elif key in aliases:
            if key == "underline":
                value = _normalize_underline(value)
            raw[key] = copy.deepcopy(value)
            cell[aliases[key]] = copy.deepcopy(value)
        elif key in _RICH_FONT_KEYS:
            raw[key] = copy.deepcopy(value)
    cell["_font_raw"] = raw


def _apply_fill_payload(cell: dict, payload) -> None:
    if payload is None or isinstance(payload, str):
        cell["fill"] = payload
        cell.pop("_fill_raw", None)
        cell["fill_model"] = None if payload is None else {
            "type": "pattern", "pattern_type": "solid",
            "foreground": {"type": "rgb", "rgb": payload},
        }
        return
    model = copy.deepcopy(payload)
    foreground = model.get("foreground", model.get("fgColor", model.get("color")))
    background = model.get("background", model.get("bgColor"))
    foreground = _normalize_color_object(foreground) if foreground is not None else None
    background = _normalize_color_object(background) if background is not None else None
    pattern_type = model.get("pattern_type", model.get("patternType", "solid"))
    fill_type = model.get("type", "pattern")
    cell["fill_model"] = {
        **model,
        "type": fill_type,
        "pattern_type": pattern_type,
        "foreground": foreground,
        "background": background,
    }
    cell["fill"] = _resolved_rgb(foreground)
    cell["_fill_raw"] = {
        "rgb": cell["fill"],
        "is_gradient": fill_type == "gradient",
        "patternType": pattern_type,
        "fgColor": foreground,
        "bgColor": background,
        **({"xml": model["xml"]} if model.get("xml") else {}),
    }


def _apply_alignment_payload(cell: dict, payload: dict) -> None:
    alignment = copy.deepcopy(cell.get("alignment") or {})
    alignment.update(copy.deepcopy(payload))
    cell["alignment"] = alignment
    aliases = {
        "horizontal": "halign", "vertical": "valign", "wrapText": "wrap",
        "wrap_text": "wrap", "textRotation": "rot", "text_rotation": "rot",
        "indent": "indent", "shrinkToFit": "shrink", "shrink_to_fit": "shrink",
    }
    for key, value in payload.items():
        if key in aliases:
            cell[aliases[key]] = copy.deepcopy(value)


def _apply_protection_payload(cell: dict, payload: dict) -> None:
    protection = copy.deepcopy(cell.get("protection") or {})
    protection.update(copy.deepcopy(payload))
    cell["protection"] = protection
    if "locked" in payload:
        cell["locked"] = bool(payload["locked"])
    if "hidden" in payload:
        cell["hidden_cell"] = bool(payload["hidden"])


def _apply_structured_style(cell: dict, style: dict) -> None:
    flat = {key: value for key, value in style.items() if key in _STYLE_KEYS and not isinstance(value, dict)}
    _apply_style(cell, flat)
    font_payload = style.get("font") if isinstance(style.get("font"), dict) else None
    if font_payload:
        _apply_font_payload(cell, font_payload)
    if "fill" in style and isinstance(style.get("fill"), dict):
        _apply_fill_payload(cell, style["fill"])
    if isinstance(style.get("alignment"), dict):
        _apply_alignment_payload(cell, style["alignment"])
    if isinstance(style.get("protection"), dict):
        _apply_protection_payload(cell, style["protection"])
    if isinstance(style.get("xf"), dict):
        xf = copy.deepcopy(cell.get("xf") or {})
        xf.update(copy.deepcopy(style["xf"]))
        cell["xf"] = xf
        if "quotePrefix" in style["xf"]:
            cell["qp"] = bool(style["xf"]["quotePrefix"])
    if "named_style" in style:
        cell["named_style"] = style["named_style"]
    cell["present"] = True


# ── Public preservation-grade semantic tools ──────────────────────────────────


def _empty_sheet_model(name: str) -> dict:
    return {
        "name": name,
        "state": "visible",
        "cw": {},
        "ch": {},
        "co": None,
        "rows": [],
        "freeze": None,
        "validations": [],
        "sheet_view": {},
        "sheet_views": [{}],
        "page_setup": None,
        "page_margins": None,
        "header_footer": None,
        "tables": [],
        "hyperlinks": {},
        "comments": {},
    }


@mcp.tool()
def excel_create_workbook(
    format: str = "xlsx",
    sheet_names: list[str] | None = None,
    active_sheet: str | int | None = None,
    template_path: str | None = None,
    macro_template_path: str | None = None,
    document_properties: dict | None = None,
    target_path: str | None = None,
) -> str:
    """Create a new in-memory workbook session; persistence still requires excel_save."""
    import uuid

    extension = format.lower().lstrip(".")
    if extension not in {"xlsx", "xltx", "xlsm", "xltm"}:
        raise ValueError("format must be xlsx, xltx, xlsm, or xltm.")
    if extension in {"xlsm", "xltm"} and not macro_template_path:
        raise ValueError("Macro-enabled workbooks require macro_template_path; VBA is never synthesized.")
    selected_template = macro_template_path or template_path
    if selected_template:
        template = uri_to_path(selected_template)
        _check_supported(template)
        if not template.is_file():
            raise FileNotFoundError(template)
        data = serialize_excel(str(template))
        source = str(template.resolve())
    else:
        names = sheet_names or ["Sheet1"]
        if not names or len(set(names)) != len(names):
            raise ValueError("sheet_names must contain at least one unique name.")
        data = {
            "source": "",
            "sheets": [_empty_sheet_model(name) for name in names],
            "named_ranges": [],
            "named_styles": [{"name": "Normal", "builtinId": 0}],
            "doc_props": None,
            "wb_view": {"activeTab": 0},
            "workbook_views": [{"activeTab": 0}],
        }
        source = ""

    if selected_template and sheet_names:
        names = list(sheet_names)
        if not names or len(set(names)) != len(names):
            raise ValueError("sheet_names must contain at least one unique name.")
        existing = data.get("sheets", [])
        for index, name in enumerate(names):
            if index < len(existing):
                existing[index]["name"] = name
            else:
                existing.append(_empty_sheet_model(name))
        data["sheets"] = existing[:len(names)]

    token = uuid.uuid4().hex
    if target_path:
        default_output = str(Path(target_path).expanduser().resolve())
        if Path(default_output).suffix.lower() != f".{extension}":
            raise ValueError(f"target_path must use .{extension} for format={extension}.")
    else:
        default_output = str((Path.cwd() / f"untitled-{token[:12]}.{extension}").resolve())
    data["source"] = source or default_output
    data["_default_output_path"] = default_output
    data["_new_workbook"] = True
    data["_sheet_filter"] = None
    data["_loaded_disk_names"] = [sheet["name"] for sheet in data.get("sheets", [])]
    if document_properties:
        data["doc_props"] = copy.deepcopy(document_properties)

    if isinstance(active_sheet, str):
        active_index = _sheet_index(data, active_sheet)
    elif active_sheet is None:
        active_index = 0
    else:
        active_index = int(active_sheet)
    if not 0 <= active_index < len(data["sheets"]):
        raise ValueError("active_sheet is outside the sheet list.")
    data["wb_view"] = {**(data.get("wb_view") or {}), "activeTab": active_index}
    views = data.get("workbook_views") or [copy.deepcopy(data["wb_view"])]
    views[0] = {**views[0], "activeTab": active_index}
    data["workbook_views"] = views
    session_key = f"new:{token}:{extension}"
    _store_session(session_key, data)
    _mark_dirty(data, "workbook", "workbook")
    return json.dumps({
        "session_key": session_key,
        "default_output_path": default_output,
        "format": extension,
        "sheets": [sheet["name"] for sheet in data["sheets"]],
        "active_sheet": data["sheets"][active_index]["name"],
        "capabilities": {
            "macros": extension in {"xlsm", "xltm"},
            "template_based": bool(selected_template),
            "requires_explicit_save": True,
        },
    }, ensure_ascii=False)


@mcp.tool()
def excel_get_rich_text(session_key: str, sheet_name: str, cell: str) -> str:
    """Read rich-text runs, Unicode offsets, whitespace, and phonetic metadata."""
    data = _get_session(session_key)
    _, _, _, cell_data = _cell_from_a1(data, sheet_name, cell, create=False)
    model = _rich_text_model(cell_data)
    model["cell"] = cell
    return json.dumps(model, default=str, ensure_ascii=False)


@mcp.tool()
def excel_edit_rich_text(
    session_key: str,
    sheet_name: str,
    cell: str,
    operations: list[dict],
    expected_text: str | None = None,
) -> str:
    """Patch rich text by runs or Unicode character ranges without saving."""
    data = _get_session(session_key)
    sheet, row_index, col_index, cell_data = _cell_from_a1(data, sheet_name, cell, create=True)
    before = _rich_text_model(cell_data)
    if expected_text is not None and before["text"] != expected_text:
        raise ValueError(f"expected_text mismatch: expected {expected_text!r}, found {before['text']!r}.")
    model = copy.deepcopy(before)

    for operation in operations:
        op = operation.get("op")
        if op == "replace_runs":
            model["runs"], model["text"] = _reindex_rich_runs(operation.get("runs") or [])
        elif op == "style_run":
            run_index = int(operation["run_index"])
            runs = copy.deepcopy(model.get("runs") or [])
            if not 0 <= run_index < len(runs):
                raise ValueError(f"run_index {run_index} is outside 0..{len(runs)-1}.")
            runs[run_index]["font"] = {
                **(runs[run_index].get("font") or {}),
                **_normalize_rich_font(operation.get("style") or {}),
            }
            model["runs"], model["text"] = _reindex_rich_runs(runs)
        elif op in {"style_range", "insert_text", "delete_range", "replace_range"}:
            chars = _chars_from_rich(model)
            start = int(operation.get("start", 0))
            end = int(operation.get("end", start))
            if not 0 <= start <= end <= len(chars):
                raise ValueError(f"Invalid Unicode range [{start}, {end}) for text length {len(chars)}.")
            if op == "style_range":
                patch = _normalize_rich_font(operation.get("style") or {})
                chars[start:end] = [(char, {**font, **patch}) for char, font in chars[start:end]]
            elif op == "delete_range":
                del chars[start:end]
            else:
                text = str(operation.get("text", ""))
                style = operation.get("style")
                if style is None:
                    if start > 0:
                        style = chars[start - 1][1]
                    elif start < len(chars):
                        style = chars[start][1]
                    else:
                        style = {}
                insertion = [(char, _normalize_rich_font(style)) for char in text]
                if op == "replace_range":
                    chars[start:end] = insertion
                else:
                    chars[start:start] = insertion
            model["runs"] = _runs_from_chars(chars)
            model["text"] = "".join(char for char, _ in chars)
        elif op == "set_phonetic":
            model["phonetic_runs"] = copy.deepcopy(operation.get("runs") or operation.get("phonetic_runs") or [])
            if "properties" in operation or "phonetic_properties" in operation:
                properties = copy.deepcopy(
                    operation.get("properties", operation.get("phonetic_properties"))
                )
                if properties is not None and not isinstance(properties, dict):
                    raise ValueError("Phonetic properties must be an object or null.")
                if properties is not None:
                    properties.setdefault("fontId", 0)
                model["phonetic_properties"] = properties
        else:
            raise ValueError(
                "Unsupported rich-text operation. Use replace_runs, style_range, style_run, "
                "insert_text, delete_range, replace_range, or set_phonetic."
            )

    _set_rich_text_cell(cell_data, model)
    after = _rich_text_model(cell_data)
    return _mutation_result(
        data,
        "rich_text",
        before,
        after,
        f"sheets/{sheet_name}/cells/{_cell_coord(row_index, col_index)}/rich_text",
    )


@mcp.tool()
def excel_set_formula(
    session_key: str,
    sheet_name: str,
    cell: str,
    formula: str | dict,
    formula_type: str = "normal",
    formula_attributes: dict | None = None,
    cached_value=None,
    cached_value_present: bool = False,
    cache_policy: str = "clear",
) -> str:
    """Set a normal/shared/array/data-table formula and explicit cache state."""
    data = _get_session(session_key)
    _, row_index, col_index, cell_data = _cell_from_a1(data, sheet_name, cell, create=True)
    before = _cell_public_view(cell_data, include_formula_cache=True, include_semantics=True)
    cached = cached_value if cached_value_present else _MISSING
    _set_formula_cell(
        cell_data,
        formula,
        formula_type=formula_type,
        attributes=formula_attributes,
        cached_value=cached,
        cache_policy=cache_policy,
    )
    after = _cell_public_view(cell_data, include_formula_cache=True, include_semantics=True)
    return _mutation_result(
        data,
        "formula",
        before,
        after,
        f"sheets/{sheet_name}/cells/{_cell_coord(row_index, col_index)}/formula",
    )


def _defined_name_scope(data: dict, sheet_name: str | None, local_sheet_id: int | None) -> int | None:
    if local_sheet_id is not None and sheet_name is not None:
        raise ValueError("Provide sheet_name or local_sheet_id, not both.")
    return int(local_sheet_id) if local_sheet_id is not None else _sheet_index(data, sheet_name)


def _find_defined_name(data: dict, name: str, scope: int | None):
    return next((item for item in data.get("named_ranges") or []
                 if item.get("name") == name and item.get("sheet_id") == scope), None)


# These three built-ins are persisted through dedicated worksheet properties
# (ws.print_area / ws.print_title_rows+cols / ws.auto_filter.ref), not through
# a raw <definedName> written from the generic named_ranges list -- see
# reconstruct_excel's "Named ranges" section. Routing them through the
# generic add/update/delete-defined-name tools would silently be dropped on
# save (add/update) or leave a stale print-area/print-titles field behind
# (delete), so they are classified as built-ins and rejected here with a
# pointer to the tool that actually owns them.
_RESERVED_BUILTIN_DEFINED_NAMES = {"_xlnm.Print_Area", "_xlnm.Print_Titles", "_xlnm._FilterDatabase"}
_BUILTIN_DEFINED_NAME_TOOL = {
    "_xlnm.Print_Area": "excel_set_print_area",
    "_xlnm.Print_Titles": "excel_set_print_titles",
    "_xlnm._FilterDatabase": "excel_set_auto_filter",
}


def _reject_builtin_defined_name(name: str) -> None:
    if name in _RESERVED_BUILTIN_DEFINED_NAMES:
        raise ValueError(
            f"{name!r} is a reserved built-in name, not an arbitrary user-defined "
            f"name; manage it via {_BUILTIN_DEFINED_NAME_TOOL[name]} instead."
        )


@mcp.tool()
def excel_add_defined_name(
    session_key: str,
    name: str,
    value: str,
    sheet_name: str | None = None,
    local_sheet_id: int | None = None,
    metadata: dict | None = None,
) -> str:
    """Add a workbook- or worksheet-scoped defined name with full metadata."""
    _reject_builtin_defined_name(name)
    data = _get_session(session_key)
    scope = _defined_name_scope(data, sheet_name, local_sheet_id)
    if _find_defined_name(data, name, scope):
        raise ValueError(f"Defined name {name!r} already exists in scope {scope!r}.")
    item = {"name": name, "value": value, "sheet_id": scope, **copy.deepcopy(metadata or {})}
    data.setdefault("named_ranges", []).append(item)
    return _mutation_result(data, "defined_names", None, item, f"workbook/defined_names/{name}:{scope}")


@mcp.tool()
def excel_update_defined_name(
    session_key: str,
    name: str,
    updates: dict,
    sheet_name: str | None = None,
    local_sheet_id: int | None = None,
) -> str:
    """Partially update a defined name without clearing omitted metadata."""
    _reject_builtin_defined_name(name)
    data = _get_session(session_key)
    scope = _defined_name_scope(data, sheet_name, local_sheet_id)
    item = _find_defined_name(data, name, scope)
    if item is None:
        raise ValueError(f"Defined name {name!r} was not found in scope {scope!r}.")
    before = copy.deepcopy(item)
    for key, value in updates.items():
        mapped = "sheet_id" if key in {"localSheetId", "local_sheet_id"} else key
        if value is None and updates.get("clear_nulls"):
            item.pop(mapped, None)
        elif key != "clear_nulls":
            item[mapped] = copy.deepcopy(value)
    return _mutation_result(data, "defined_names", before, item, f"workbook/defined_names/{name}:{scope}")


@mcp.tool()
def excel_delete_defined_name(
    session_key: str,
    name: str,
    sheet_name: str | None = None,
    local_sheet_id: int | None = None,
) -> str:
    """Delete one defined name from its exact scope."""
    _reject_builtin_defined_name(name)
    data = _get_session(session_key)
    scope = _defined_name_scope(data, sheet_name, local_sheet_id)
    names = data.setdefault("named_ranges", [])
    item = _find_defined_name(data, name, scope)
    if item is None:
        raise ValueError(f"Defined name {name!r} was not found in scope {scope!r}.")
    names.remove(item)
    return _mutation_result(data, "defined_names", item, None, f"workbook/defined_names/{name}:{scope}")


@mcp.tool()
def excel_set_auto_filter(
    session_key: str,
    sheet_name: str,
    ref: str | None = None,
    filter_columns: list[dict] | None = None,
    sort_state: dict | None = None,
    mode: str = "patch",
) -> str:
    """Set or patch filter criteria and sort state without deleting siblings."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("auto_filter_model") or {
        "ref": sheet.get("auto_filter"), "filter_columns": [], "sort_state": None,
    })
    model = {} if mode == "replace" else copy.deepcopy(before)
    if ref is not None:
        model["ref"] = ref
        sheet["auto_filter"] = ref
    if filter_columns is not None:
        if mode == "patch":
            by_id = {str(item.get("colId")): copy.deepcopy(item)
                     for item in model.get("filter_columns") or []}
            for item in filter_columns:
                key = str(item.get("colId"))
                by_id[key] = {**by_id.get(key, {}), **copy.deepcopy(item)}
            model["filter_columns"] = list(by_id.values())
        else:
            model["filter_columns"] = copy.deepcopy(filter_columns)
    if sort_state is not None:
        model["sort_state"] = copy.deepcopy(sort_state)
    if mode not in {"patch", "replace"}:
        raise ValueError("mode must be patch or replace.")
    sheet["auto_filter_model"] = model
    return _mutation_result(data, "auto_filter", before, model, f"sheets/{sheet_name}/auto_filter")


def _table_by_name(data: dict, name: str):
    for sheet in data.get("sheets") or []:
        for table in sheet.get("tables") or []:
            if table.get("name") == name or table.get("displayName") == name:
                return sheet, table
    return None, None


def _deep_patch_mapping(target: dict, updates: dict, *, clear_none: bool = False) -> None:
    for key, value in updates.items():
        if value is None and clear_none:
            target.pop(key, None)
        elif isinstance(value, dict) and isinstance(target.get(key), dict):
            _deep_patch_mapping(target[key], value, clear_none=clear_none)
        else:
            target[key] = copy.deepcopy(value)


def _table_ref_bounds(ref: str) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries

    normalized = str(ref or "").strip()
    if not normalized or "!" in normalized or "," in normalized or " " in normalized:
        raise ValueError(f"Table ref must be one worksheet-local A1 range, got {ref!r}.")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(normalized.replace("$", ""))
    except Exception as exc:
        raise ValueError(f"Invalid table ref {ref!r}.") from exc
    if None in (min_col, min_row, max_col, max_row) or min_col > max_col or min_row > max_row:
        raise ValueError(f"Invalid table ref {ref!r}.")
    return min_col - 1, min_row - 1, max_col - 1, max_row - 1


def _table_header_name(sheet: dict, row_index: int, col_index: int, ordinal: int) -> str:
    rows = sheet.get("rows") or []
    if 0 <= row_index < len(rows):
        cells = rows[row_index].get("cells") or []
        if 0 <= col_index < len(cells):
            value = cells[col_index].get("value", cells[col_index].get("v"))
            if value is not None and str(value).strip():
                return str(value)
    return f"Column{ordinal}"


def _reconcile_table_columns(
    sheet: dict,
    existing: list[dict] | None,
    column_updates: list[dict] | None,
    bounds: tuple[int, int, int, int],
    *,
    clear_none: bool = False,
) -> list[dict]:
    min_col, min_row, max_col, _ = bounds
    width = max_col - min_col + 1
    if column_updates is not None and not isinstance(column_updates, list):
        raise ValueError("Table columns update must be a list.")
    if column_updates is not None and len(column_updates) > width:
        raise ValueError(f"Table columns update has {len(column_updates)} entries for a {width}-column ref.")

    columns = copy.deepcopy(existing or [])[:width]
    while len(columns) < width:
        columns.append({})
    for index, patch in enumerate(column_updates or []):
        if patch is None:
            if clear_none:
                columns[index] = {}
            continue
        if not isinstance(patch, dict):
            raise ValueError(f"Table column update at index {index} must be an object.")
        _deep_patch_mapping(columns[index], patch, clear_none=clear_none)

    used_ids: set[int] = set()
    next_id = max(
        (int(item.get("id")) for item in existing or [] if str(item.get("id", "")).isdigit()),
        default=0,
    ) + 1
    for index, column in enumerate(columns):
        try:
            column_id = int(column.get("id"))
        except (TypeError, ValueError):
            column_id = 0
        if column_id <= 0 or column_id in used_ids:
            while next_id in used_ids:
                next_id += 1
            column_id = next_id
            next_id += 1
        column["id"] = column_id
        used_ids.add(column_id)
        if not column.get("name"):
            column["name"] = _table_header_name(sheet, min_row, min_col + index, index + 1)
    return columns


def _rebase_table_ref_value(
    ref: str | None,
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
):
    if not ref:
        return ref
    import openpyxl.utils as U

    value = str(ref)
    prefix = ""
    local_ref = value
    if "!" in value:
        sheet_prefix, local_ref = value.rsplit("!", 1)
        prefix = f"{sheet_prefix}!"

    def map_axis(
        position: int,
        old_min: int,
        old_max: int,
        new_min: int,
        new_max: int,
        *,
        resize_edges: bool,
    ) -> int:
        if not resize_edges:
            return position + new_min - old_min
        if position == old_min:
            return new_min
        if position == old_max:
            return new_max
        if old_min < position < old_max:
            return min(max(new_min + position - old_min, new_min), new_max)
        return position + new_min - old_min

    old_min_col, old_min_row, old_max_col, old_max_row = old_bounds
    new_min_col, new_min_row, new_max_col, new_max_row = new_bounds
    match = _RANGE_RE.match(local_ref)
    if match:
        abs_col1, col1, abs_row1, row1, abs_col2, col2, abs_row2, row2 = match.groups()
        col_index1 = U.column_index_from_string(col1.upper()) - 1
        row_index1 = int(row1) - 1
        col_index2 = U.column_index_from_string((col2 or col1).upper()) - 1
        row_index2 = int(row2 or row1) - 1
        resize_columns = col_index1 != col_index2
        resize_rows = row_index1 != row_index2
        mapped_col1 = map_axis(
            col_index1, old_min_col, old_max_col, new_min_col, new_max_col,
            resize_edges=resize_columns,
        )
        mapped_row1 = map_axis(
            row_index1, old_min_row, old_max_row, new_min_row, new_max_row,
            resize_edges=resize_rows,
        )
        first = f"{abs_col1}{U.get_column_letter(mapped_col1 + 1)}{abs_row1}{mapped_row1 + 1}"
        if col2 is None:
            return f"{prefix}{first}"
        mapped_col2 = map_axis(
            col_index2, old_min_col, old_max_col, new_min_col, new_max_col,
            resize_edges=resize_columns,
        )
        mapped_row2 = map_axis(
            row_index2, old_min_row, old_max_row, new_min_row, new_max_row,
            resize_edges=resize_rows,
        )
        second = f"{abs_col2}{U.get_column_letter(mapped_col2 + 1)}{abs_row2}{mapped_row2 + 1}"
        return f"{prefix}{first}:{second}"
    return value


def _rebase_table_sort_state(
    sort_state: dict | None,
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
) -> None:
    if not isinstance(sort_state, dict):
        return
    if sort_state.get("ref"):
        sort_state["ref"] = _rebase_table_ref_value(sort_state["ref"], old_bounds, new_bounds)
    for condition in sort_state.get("conditions", sort_state.get("sortCondition", [])) or []:
        if isinstance(condition, dict) and condition.get("ref"):
            condition["ref"] = _rebase_table_ref_value(condition["ref"], old_bounds, new_bounds)


def _rebase_table_nested_refs(
    table: dict,
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
) -> None:
    auto_filter = table.get("auto_filter")
    if isinstance(auto_filter, dict):
        if auto_filter.get("ref"):
            auto_filter["ref"] = _rebase_table_ref_value(auto_filter["ref"], old_bounds, new_bounds)
        _rebase_table_sort_state(auto_filter.get("sort_state"), old_bounds, new_bounds)
    _rebase_table_sort_state(table.get("sort_state"), old_bounds, new_bounds)


def _normalize_table_nested_models(table: dict, ref: str, width: int) -> None:
    auto_filter = table.get("auto_filter")
    if auto_filter is not None and not isinstance(auto_filter, dict):
        raise ValueError("Table auto_filter must be an object or null.")
    if isinstance(auto_filter, dict):
        auto_filter["ref"] = ref
        kept_columns = []
        for column in auto_filter.get("filter_columns") or []:
            try:
                column_id = int(column.get("colId", 0))
            except (AttributeError, TypeError, ValueError):
                continue
            if 0 <= column_id < width:
                kept_columns.append(column)
        if "filter_columns" in auto_filter:
            auto_filter["filter_columns"] = kept_columns
        nested_sort = auto_filter.get("sort_state")
        if isinstance(nested_sort, dict):
            nested_sort.setdefault("ref", ref)
    sort_state = table.get("sort_state")
    if sort_state is not None and not isinstance(sort_state, dict):
        raise ValueError("Table sort_state must be an object or null.")
    if isinstance(sort_state, dict):
        sort_state.setdefault("ref", ref)


@mcp.tool()
def excel_add_table(
    session_key: str,
    sheet_name: str,
    name: str,
    ref: str,
    table: dict | None = None,
    style: dict | None = None,
) -> str:
    """Add a table model including columns, formulas, filters, totals, and style flags."""
    data = _get_session(session_key)
    if _table_by_name(data, name)[1] is not None:
        raise ValueError(f"Table name {name!r} must be unique workbook-wide.")
    sheet = _find_sheet(data, sheet_name)
    item = {"name": name, "displayName": name, "ref": ref, **copy.deepcopy(table or {})}
    if style is not None:
        item["style"] = copy.deepcopy(style)
    sheet["tables"] = sheet.get("tables") or []
    sheet["tables"].append(item)
    return _mutation_result(data, "tables", None, item, f"sheets/{sheet_name}/tables/{name}")


@mcp.tool()
def excel_update_table(session_key: str, name: str, updates: dict) -> str:
    """Partially update a table and reconcile its range-bound metadata."""
    if not isinstance(updates, dict):
        raise ValueError("Table updates must be an object.")
    data = _get_session(session_key)
    sheet, table = _table_by_name(data, name)
    if table is None:
        raise ValueError(f"Table {name!r} not found.")

    requested_name = updates.get("name") if "name" in updates else None
    requested_display_name = updates.get("displayName") if "displayName" in updates else None
    if "name" in updates and "displayName" in updates and requested_name != requested_display_name:
        raise ValueError("Table name and displayName must match when both are provided.")
    if "name" in updates or "displayName" in updates:
        final_name = requested_name if "name" in updates else requested_display_name
        if not isinstance(final_name, str) or not final_name.strip():
            raise ValueError("Table name must be a non-empty string.")
        final_name = final_name.strip()
    else:
        final_name = str(table.get("name") or table.get("displayName") or name)
    for other_sheet in data.get("sheets") or []:
        for other_table in other_sheet.get("tables") or []:
            if other_table is table:
                continue
            if other_table.get("name") == final_name or other_table.get("displayName") == final_name:
                raise ValueError(f"Table name {final_name!r} must be unique workbook-wide.")

    old_ref = table.get("ref")
    final_ref = updates.get("ref", old_ref)
    old_bounds = _table_ref_bounds(old_ref)
    new_bounds = _table_ref_bounds(final_ref)
    clear_none = bool(updates.get("clear_nulls"))
    column_updates = updates.get("columns") if "columns" in updates else None

    before = copy.deepcopy(table)
    candidate = copy.deepcopy(table)
    if old_bounds != new_bounds:
        _rebase_table_nested_refs(candidate, old_bounds, new_bounds)
    patch = {
        key: value for key, value in updates.items()
        if key not in {"clear_nulls", "columns", "name", "displayName", "ref"}
    }
    _deep_patch_mapping(candidate, patch, clear_none=clear_none)
    candidate["name"] = final_name
    candidate["displayName"] = final_name
    candidate["ref"] = str(final_ref)
    candidate["columns"] = _reconcile_table_columns(
        sheet,
        table.get("columns"),
        column_updates,
        new_bounds,
        clear_none=clear_none,
    )
    _normalize_table_nested_models(candidate, candidate["ref"], new_bounds[2] - new_bounds[0] + 1)

    table.clear()
    table.update(candidate)
    return _mutation_result(
        data,
        "tables",
        before,
        table,
        f"sheets/{sheet['name']}/tables/{final_name}",
    )


@mcp.tool()
def excel_delete_table(session_key: str, name: str) -> str:
    """Delete one workbook-unique table by name."""
    data = _get_session(session_key)
    sheet, table = _table_by_name(data, name)
    if table is None:
        raise ValueError(f"Table {name!r} not found.")
    sheet["tables"].remove(table)
    return _mutation_result(data, "tables", table, None, f"sheets/{sheet['name']}/tables/{name}")


@mcp.tool()
def excel_set_hyperlink(
    session_key: str,
    sheet_name: str,
    cell: str,
    target: str | None = None,
    location: str | None = None,
    display: str | None = None,
    tooltip: str | None = None,
    relationship: dict | None = None,
) -> str:
    """Partially set external/internal hyperlink metadata on one cell."""
    data = _get_session(session_key)
    sheet, row_index, col_index, cell_data = _cell_from_a1(data, sheet_name, cell, create=True)
    links = sheet.setdefault("hyperlinks", {})
    before = copy.deepcopy(links.get(cell))
    item = copy.deepcopy(before or {})
    for key, value in {
        "target": target, "location": location, "display": display,
        "tooltip": tooltip, "relationship": relationship,
    }.items():
        if value is not None:
            item[key] = copy.deepcopy(value)
    if not item.get("target") and not item.get("location"):
        raise ValueError("A hyperlink requires target or location.")
    links[cell] = item
    if display is not None:
        cell_data["v"] = display
        cell_data["value"] = display
        cell_data["dt"] = "s"
    cell_data["present"] = True
    return _mutation_result(data, "hyperlinks", before, item, f"sheets/{sheet_name}/hyperlinks/{cell}")


@mcp.tool()
def excel_remove_hyperlink(session_key: str, sheet_name: str, cell: str) -> str:
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = (sheet.get("hyperlinks") or {}).pop(cell, None)
    return _mutation_result(data, "hyperlinks", before, None, f"sheets/{sheet_name}/hyperlinks/{cell}")


@mcp.tool()
def excel_set_comment(
    session_key: str,
    sheet_name: str,
    cell: str,
    text: str,
    author: str = "",
    comment_type: str = "legacy",
    metadata: dict | None = None,
) -> str:
    """Set a legacy comment; threaded comments must use package-level support."""
    if comment_type != "legacy":
        raise ValueError("Threaded comments are not converted to legacy comments; core/package support is required.")
    data = _get_session(session_key)
    sheet, _, _, cell_data = _cell_from_a1(data, sheet_name, cell, create=True)
    comments = sheet.setdefault("comments", {})
    before = copy.deepcopy(comments.get(cell))
    item = {"text": text, "author": author, **copy.deepcopy(metadata or {})}
    comments[cell] = item
    cell_data["present"] = True
    return _mutation_result(data, "comments", before, item, f"sheets/{sheet_name}/comments/{cell}")


@mcp.tool()
def excel_remove_comment(session_key: str, sheet_name: str, cell: str) -> str:
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = (sheet.get("comments") or {}).pop(cell, None)
    return _mutation_result(data, "comments", before, None, f"sheets/{sheet_name}/comments/{cell}")


@mcp.tool()
def excel_set_ignored_errors(
    session_key: str,
    sheet_name: str,
    rules: list[dict],
    mode: str = "patch",
) -> str:
    """Patch ignored-error rules with multi-range sqref and all OOXML flags."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("ignored_errors") or [])
    if mode == "replace":
        after = copy.deepcopy(rules)
    elif mode == "patch":
        by_sqref = {item.get("sqref"): copy.deepcopy(item) for item in before}
        for rule in rules:
            sqref = rule.get("sqref")
            if isinstance(sqref, list):
                sqref = " ".join(sqref)
            if not sqref:
                raise ValueError("Each ignored-error rule requires sqref.")
            by_sqref[sqref] = {**by_sqref.get(sqref, {}), **copy.deepcopy(rule), "sqref": sqref}
        after = list(by_sqref.values())
    else:
        raise ValueError("mode must be patch or replace.")
    sheet["ignored_errors"] = after
    return _mutation_result(data, "ignored_errors", before, after, f"sheets/{sheet_name}/ignored_errors")


def _named_style(data: dict, name: str):
    return next((style for style in data.get("named_styles") or [] if style.get("name") == name), None)


@mcp.tool()
def excel_add_named_style(session_key: str, name: str, style: dict, metadata: dict | None = None) -> str:
    """Add a named style without renumbering existing style records."""
    data = _get_session(session_key)
    if _named_style(data, name):
        raise ValueError(f"Named style {name!r} already exists.")
    item = {"name": name, "style": copy.deepcopy(style), **copy.deepcopy(metadata or {})}
    data.setdefault("named_styles", [{"name": "Normal", "builtinId": 0}]).append(item)
    return _mutation_result(data, "named_styles", None, item, f"workbook/named_styles/{name}")


@mcp.tool()
def excel_update_named_style(session_key: str, name: str, updates: dict) -> str:
    """Partially update a named style while preserving its identity."""
    data = _get_session(session_key)
    item = _named_style(data, name)
    if item is None:
        raise ValueError(f"Named style {name!r} not found.")
    before = copy.deepcopy(item)
    _patch_mapping(item, updates, clear_none=bool(updates.get("clear_nulls")))
    item.pop("clear_nulls", None)
    return _mutation_result(data, "named_styles", before, item, f"workbook/named_styles/{name}")


@mcp.tool()
def excel_delete_named_style(session_key: str, name: str) -> str:
    """Delete a non-Normal named style only when no cell still references it."""
    if name == "Normal":
        raise ValueError("The built-in Normal style cannot be deleted.")
    data = _get_session(session_key)
    item = _named_style(data, name)
    if item is None:
        raise ValueError(f"Named style {name!r} not found.")
    referenced = []
    for sheet in data.get("sheets") or []:
        for row_index, row in enumerate(sheet.get("rows") or []):
            for col_index, cell in enumerate(row.get("cells") or []):
                if cell.get("named_style") == name:
                    referenced.append(f"{sheet['name']}!{_cell_coord(row_index, col_index)}")
    if referenced:
        raise ValueError(f"Named style {name!r} is still used by {referenced[:20]}.")
    data["named_styles"].remove(item)
    return _mutation_result(data, "named_styles", item, None, f"workbook/named_styles/{name}")


@mcp.tool()
def excel_set_cell_style_semantics(
    session_key: str,
    sheet_name: str,
    range_ref: str,
    xf: dict | None = None,
    named_style: str | None = None,
    exact_default_policy: str = "preserve",
    dry_run: bool = False,
) -> str:
    """Set expert XF flags/base-style association on an A1 range."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    if named_style is not None and _named_style(data, named_style) is None:
        raise ValueError(f"Named style {named_style!r} not found.")
    r1, r2, c1, c2 = _excel_range_to_indices(range_ref)
    changes = []
    for row_index in range(r1, r2 + 1):
        for col_index in range(c1, c2 + 1):
            if dry_run:
                rows = sheet.get("rows", [])
                if row_index < len(rows) and col_index < len(rows[row_index].get("cells", [])):
                    cell = _cell_model_for_read(
                        rows[row_index]["cells"][col_index],
                        sheet.get("_implicit_cell_defaults"),
                    )
                else:
                    cell = _new_empty_cell()
            else:
                cell = _ensure_cell(sheet, row_index, col_index, capture_baseline=True)
            before = {"xf": copy.deepcopy(cell.get("xf") or {}), "named_style": cell.get("named_style")}
            after = copy.deepcopy(before)
            if xf is not None:
                after["xf"].update(copy.deepcopy(xf))
                after["xf"]["exact_default_policy"] = exact_default_policy
            if named_style is not None:
                after["named_style"] = named_style
            changes.append({"cell": _cell_coord(row_index, col_index), "before": before, "after": after})
            if not dry_run:
                cell["xf"] = after["xf"]
                cell["named_style"] = after["named_style"]
                cell["present"] = True
    if not dry_run:
        _mark_dirty(data, "cell_style", f"sheets/{sheet_name}/range/{range_ref}/xf")
    return json.dumps({"dry_run": dry_run, "changes": changes,
                       "dirty_features": data.get("_dirty_features", [])},
                      default=str, ensure_ascii=False)


@mcp.tool()
def excel_get_workbook_semantics(session_key: str) -> str:
    """Read workbook calculation/properties/protection/views/document semantics."""
    data = _get_session(session_key)
    return json.dumps({
        "calculation_properties": _strip_private(data.get("calculation_properties") or {}),
        "workbook_properties": _strip_private(data.get("workbook_properties") or {}),
        "document_properties": _strip_private(data.get("doc_props") or {}),
        "app_properties": _strip_private(data.get("app_props") or {}),
        "custom_properties": _strip_private(data.get("custom_doc_props") or []),
        "workbook_protection": _strip_private(data.get("workbook_protection") or {}),
        "workbook_views": _strip_private(data.get("workbook_views") or ([data.get("wb_view")] if data.get("wb_view") else [])),
        "dirty_features": list(data.get("_dirty_features") or []),
    }, default=str, ensure_ascii=False)


@mcp.tool()
def excel_set_calculation_properties(session_key: str, properties: dict) -> str:
    """Partially set calculation properties while preserving explicit false and zero."""
    data = _get_session(session_key)
    target = data.setdefault("calculation_properties", {})
    before = _patch_mapping(target, properties, clear_none=bool(properties.get("clear_nulls")))
    target.pop("clear_nulls", None)
    return _mutation_result(data, "calculation_properties", before, target, "workbook/calculation_properties")


@mcp.tool()
def excel_set_workbook_properties(
    session_key: str,
    properties: dict,
    date_system_policy: str | None = None,
) -> str:
    """Patch workbookPr fields; changing date1904 requires an explicit date policy."""
    data = _get_session(session_key)
    target = data.setdefault("workbook_properties", {})
    # A never-loaded/never-set baseline is the OOXML default (date1904=False),
    # not "unknown" -- otherwise setting the already-effective False value on
    # a brand-new workbook would incorrectly demand a date_system_policy.
    if "date1904" in properties and bool(properties["date1904"]) != bool(target.get("date1904", False)):
        if date_system_policy not in {"preserve_displayed_dates", "preserve_serial_values"}:
            raise ValueError(
                "Changing date1904 requires date_system_policy='preserve_displayed_dates' "
                "or 'preserve_serial_values'."
            )
    before = _patch_mapping(target, properties, clear_none=bool(properties.get("clear_nulls")))
    target.pop("clear_nulls", None)
    if "date1904" in properties:
        target["date_system_policy"] = date_system_policy
    return _mutation_result(data, "workbook_properties", before, target, "workbook/properties")


@mcp.tool()
def excel_set_document_properties(
    session_key: str,
    core: dict | None = None,
    app: dict | None = None,
    custom: list[dict] | None = None,
    modified_policy: str = "preserve",
) -> str:
    """Patch core/app/custom document properties with typed custom values."""
    if modified_policy not in {"preserve", "set_explicit", "update_on_save"}:
        raise ValueError("modified_policy must be preserve, set_explicit, or update_on_save.")
    data = _get_session(session_key)
    before = {
        "core": copy.deepcopy(data.get("doc_props") or {}),
        "app": copy.deepcopy(data.get("app_props") or {}),
        "custom": copy.deepcopy(data.get("custom_doc_props") or []),
        "modified_policy": data.get("modified_policy", "preserve"),
    }
    if core is not None:
        # data["doc_props"] can exist and be explicitly None (e.g. a fresh
        # excel_create_workbook session with no template) -- setdefault does
        # NOT replace an existing None value, so build the dict ourselves.
        target = data.get("doc_props") or {}
        data["doc_props"] = target
        _patch_mapping(target, core, clear_none=bool(core.get("clear_nulls")))
        target.pop("clear_nulls", None)
    if app is not None:
        target = data.get("app_props") or {}
        data["app_props"] = target
        _patch_mapping(target, app, clear_none=bool(app.get("clear_nulls")))
        target.pop("clear_nulls", None)
    if custom is not None:
        data["custom_doc_props"] = copy.deepcopy(custom)
    data["modified_policy"] = modified_policy
    after = {
        "core": data.get("doc_props") or {}, "app": data.get("app_props") or {},
        "custom": data.get("custom_doc_props") or [], "modified_policy": modified_policy,
    }
    return _mutation_result(data, "document_properties", before, after, "document_properties")


@mcp.tool()
def excel_set_workbook_protection(
    session_key: str,
    properties: dict,
    already_hashed: bool = True,
) -> str:
    """Patch workbook protection without rehashing caller-supplied hashes."""
    data = _get_session(session_key)
    target = data.setdefault("workbook_protection", {})
    before = copy.deepcopy(target)
    target.update(copy.deepcopy(properties))
    target["already_hashed"] = bool(already_hashed)
    return _mutation_result(data, "workbook_protection", before, target, "workbook/protection")


@mcp.tool()
def excel_get_workbook_views(session_key: str) -> str:
    data = _get_session(session_key)
    views = data.get("workbook_views") or ([data.get("wb_view")] if data.get("wb_view") else [])
    return json.dumps({"views": _strip_private(views)}, default=str, ensure_ascii=False)


@mcp.tool()
def excel_set_workbook_views(session_key: str, views: list[dict], mode: str = "replace") -> str:
    """Replace or index-patch the complete ordered workbook-view list."""
    data = _get_session(session_key)
    before = copy.deepcopy(data.get("workbook_views") or ([data.get("wb_view")] if data.get("wb_view") else []))
    if mode == "replace":
        after = copy.deepcopy(views)
    elif mode == "patch":
        after = copy.deepcopy(before)
        for view in views:
            index = int(view.get("index", 0))
            patch = {key: copy.deepcopy(value) for key, value in view.items() if key != "index"}
            while index >= len(after):
                after.append({})
            after[index].update(patch)
    else:
        raise ValueError("mode must be replace or patch.")
    if not after:
        raise ValueError("A workbook must keep at least one workbook view.")
    data["workbook_views"] = after
    data["wb_view"] = copy.deepcopy(after[0])
    return _mutation_result(data, "workbook_views", before, after, "workbook/views")


@mcp.tool()
def excel_get_sheet_semantics(session_key: str, sheet_name: str) -> str:
    """Read state, properties, views, printing, breaks, protection, and errors."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    keys = (
        "state", "sheet_properties", "sheet_view", "sheet_views", "print_options",
        "page_setup", "page_margins", "header_footer", "page_breaks", "print_area",
        "print_titles", "protected_ranges", "ignored_errors", "protection",
    )
    return json.dumps({key: _strip_private(sheet.get(key)) for key in keys}, default=str, ensure_ascii=False)


@mcp.tool()
def excel_set_sheet_state(session_key: str, sheet_name: str, state: str) -> str:
    """Set visible/hidden/veryHidden while retaining at least one visible sheet."""
    if state not in {"visible", "hidden", "veryHidden"}:
        raise ValueError("state must be visible, hidden, or veryHidden.")
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = sheet.get("state", "visible")
    if state != "visible":
        others_visible = any(other is not sheet and other.get("state", "visible") == "visible"
                             for other in data.get("sheets") or [])
        if not others_visible:
            raise ValueError("At least one worksheet must remain visible.")
    sheet["state"] = state
    return _mutation_result(data, "sheet_state", before, state, f"sheets/{sheet_name}/state")


@mcp.tool()
def excel_set_sheet_properties(session_key: str, sheet_name: str, properties: dict) -> str:
    """Partially set CodeName/filter/published/sync/outline/pageSetup properties."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    target = sheet.setdefault("sheet_properties", {})
    before = _patch_mapping(target, properties, clear_none=bool(properties.get("clear_nulls")))
    target.pop("clear_nulls", None)
    return _mutation_result(data, "sheet_properties", before, target, f"sheets/{sheet_name}/properties")


@mcp.tool()
def excel_get_sheet_views(session_key: str, sheet_name: str) -> str:
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    views = sheet.get("sheet_views") or ([sheet.get("sheet_view")] if sheet.get("sheet_view") else [])
    return json.dumps({"sheet_name": sheet_name, "views": _strip_private(views)}, default=str, ensure_ascii=False)


@mcp.tool()
def excel_set_sheet_views(session_key: str, sheet_name: str, views: list[dict], mode: str = "replace") -> str:
    """Replace or index-patch all sheet views, panes, and selections."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("sheet_views") or ([sheet.get("sheet_view")] if sheet.get("sheet_view") else []))
    if mode == "replace":
        after = copy.deepcopy(views)
    elif mode == "patch":
        after = copy.deepcopy(before or [{}])
        for view in views:
            index = int(view.get("index", 0))
            patch = {key: copy.deepcopy(value) for key, value in view.items() if key != "index"}
            while index >= len(after):
                after.append({})
            after[index].update(patch)
    else:
        raise ValueError("mode must be replace or patch.")
    sheet["sheet_views"] = after
    sheet["sheet_view"] = copy.deepcopy(after[0] if after else {})
    return _mutation_result(data, "sheet_views", before, after, f"sheets/{sheet_name}/views")


@mcp.tool()
def excel_set_row_properties(session_key: str, sheet_name: str, row_index: int, properties: dict) -> str:
    """Patch one row's height/hidden/outline/collapse/style/phonetic flags."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    _ensure_cell(sheet, row_index, 0)
    row = sheet["rows"][row_index]
    before = copy.deepcopy(row)
    aliases = {"height": "h", "outlineLevel": "outline"}
    for key, value in properties.items():
        row[aliases.get(key, key)] = copy.deepcopy(value)
    return _mutation_result(data, "row_properties", before, row, f"sheets/{sheet_name}/rows/{row_index}/properties")


@mcp.tool()
def excel_set_page_setup(
    session_key: str,
    sheet_name: str,
    properties: dict,
    present: bool | None = None,
    exact: bool = False,
) -> str:
    """Patch full pageSetup attributes and preserve explicit-empty presence."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("page_setup"))
    target = {} if exact else copy.deepcopy(sheet.get("page_setup") or {})
    target.update(copy.deepcopy(properties))
    if present is not None:
        target["present"] = bool(present)
    sheet["page_setup"] = target if present is not False else None
    return _mutation_result(data, "page_setup", before, sheet.get("page_setup"), f"sheets/{sheet_name}/page_setup")


@mcp.tool()
def excel_set_print_options(
    session_key: str,
    sheet_name: str,
    properties: dict,
    present: bool | None = None,
    exact: bool = False,
) -> str:
    """Patch printOptions and preserve explicit false/empty attributes."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("print_options"))
    target = {} if exact else copy.deepcopy(sheet.get("print_options") or {})
    target.update(copy.deepcopy(properties))
    if present is not None:
        target["present"] = bool(present)
    sheet["print_options"] = target if present is not False else None
    return _mutation_result(data, "print_options", before, sheet.get("print_options"), f"sheets/{sheet_name}/print_options")


def _flatten_header_footer(model: dict) -> dict:
    result = copy.deepcopy(model)
    aliases = {
        ("odd_header", "left"): "hl", ("odd_header", "center"): "hc", ("odd_header", "right"): "hr",
        ("odd_footer", "left"): "fl", ("odd_footer", "center"): "fc", ("odd_footer", "right"): "fr",
    }
    for (section, position), alias in aliases.items():
        if isinstance(model.get(section), dict) and position in model[section]:
            result[alias] = model[section][position]
    return result


@mcp.tool()
def excel_set_header_footer(session_key: str, sheet_name: str, sections: dict, properties: dict | None = None) -> str:
    """Patch odd/even/first headers and footers without clearing sibling sections."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("header_footer") or {})
    model = copy.deepcopy(before)
    for section, value in sections.items():
        if isinstance(value, dict) and isinstance(model.get(section), dict):
            model[section].update(copy.deepcopy(value))
        else:
            model[section] = copy.deepcopy(value)
    if properties:
        model.update(copy.deepcopy(properties))
    sheet["header_footer"] = _flatten_header_footer(model)
    return _mutation_result(data, "header_footer", before, sheet["header_footer"], f"sheets/{sheet_name}/header_footer")


@mcp.tool()
def excel_set_page_breaks(
    session_key: str,
    sheet_name: str,
    row_breaks: list[dict] | None = None,
    column_breaks: list[dict] | None = None,
    mode: str = "replace",
    exact_attributes: dict | None = None,
) -> str:
    """Set row/column breaks and synchronize count/manualBreakCount metadata."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("page_breaks") or {})
    target = {} if mode == "replace" else copy.deepcopy(before)
    if row_breaks is not None:
        target["rows"] = copy.deepcopy(row_breaks)
    if column_breaks is not None:
        target["columns"] = copy.deepcopy(column_breaks)
    for axis in ("rows", "columns"):
        entries = target.get(axis) or []
        target[f"{axis}_count"] = len(entries)
        target[f"{axis}_manualBreakCount"] = sum(1 for item in entries if item.get("man", True))
    if exact_attributes:
        target["exact_attributes"] = copy.deepcopy(exact_attributes)
    sheet["page_breaks"] = target
    return _mutation_result(data, "page_breaks", before, target, f"sheets/{sheet_name}/page_breaks")


@mcp.tool()
def excel_set_print_area(session_key: str, sheet_name: str, areas: str | list[str] | None) -> str:
    """Set/clear multi-area print ranges through the common defined-name model."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    scope = _sheet_index(data, sheet_name)
    before = copy.deepcopy(sheet.get("print_area"))
    value = None if areas is None else (",".join(areas) if isinstance(areas, list) else areas)
    sheet["print_area"] = value
    existing = _find_defined_name(data, "_xlnm.Print_Area", scope)
    if value is None:
        if existing:
            data["named_ranges"].remove(existing)
    elif existing:
        existing["value"] = value
    else:
        data.setdefault("named_ranges", []).append({"name": "_xlnm.Print_Area", "value": value, "sheet_id": scope, "builtin": True})
    return _mutation_result(data, "print_area", before, value, f"sheets/{sheet_name}/print_area")


@mcp.tool()
def excel_set_print_titles(
    session_key: str,
    sheet_name: str,
    repeated_rows: str | None = None,
    repeated_columns: str | None = None,
) -> str:
    """Set repeated print rows/columns and synchronized built-in defined name."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("print_titles"))
    value = {"rows": repeated_rows, "cols": repeated_columns}
    sheet["print_titles"] = value if repeated_rows or repeated_columns else None
    scope = _sheet_index(data, sheet_name)
    existing = _find_defined_name(data, "_xlnm.Print_Titles", scope)
    pieces = [part for part in (repeated_columns, repeated_rows) if part]
    text = ",".join(pieces) if pieces else None
    if text is None and existing:
        data["named_ranges"].remove(existing)
    elif existing:
        existing["value"] = text
    elif text:
        data.setdefault("named_ranges", []).append({"name": "_xlnm.Print_Titles", "value": text, "sheet_id": scope, "builtin": True})
    return _mutation_result(data, "print_titles", before, sheet.get("print_titles"), f"sheets/{sheet_name}/print_titles")


@mcp.tool()
def excel_set_sheet_protection(
    session_key: str,
    sheet_name: str,
    properties: dict | None = None,
    enabled: bool | None = None,
    already_hashed: bool = True,
) -> str:
    """Patch or disable worksheet protection, including legacy and modern hashes."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("protection"))
    updates = copy.deepcopy(properties or {})
    clear_nulls = bool(updates.pop("clear_nulls", False))
    if "sheet" in updates:
        property_enabled = bool(updates.pop("sheet"))
        if enabled is not None and bool(enabled) != property_enabled:
            raise ValueError("enabled conflicts with properties['sheet'].")
        enabled = property_enabled
    allowed = {
        "password", "selectLockedCells", "selectUnlockedCells", "algorithmName",
        "objects", "insertRows", "insertHyperlinks", "autoFilter", "scenarios",
        "formatColumns", "deleteColumns", "insertColumns", "pivotTables", "deleteRows",
        "formatCells", "saltValue", "formatRows", "sort", "spinCount", "hashValue",
    }
    unknown = set(updates) - allowed
    if unknown:
        raise ValueError(f"Unknown worksheet protection fields: {sorted(unknown)}")
    if enabled is False:
        if updates:
            raise ValueError("Cannot patch worksheet protection properties while disabling it.")
        sheet["protection"] = None
        return _mutation_result(
            data,
            "sheet_protection",
            before,
            None,
            f"sheets/{sheet_name}/protection",
        )

    target = copy.deepcopy(before or {})
    _patch_mapping(target, updates, clear_none=clear_nulls)
    if target.get("password") is not None:
        target["password_is_hashed"] = bool(already_hashed)
    elif clear_nulls and "password" in updates:
        target.pop("password_is_hashed", None)
    if target.get("password") is not None and target.get("hashValue") is not None:
        raise ValueError("Use either legacy password or modern hashValue, not both.")
    if enabled is True or before is not None or updates:
        sheet["protection"] = target
    after = copy.deepcopy(sheet.get("protection"))
    return _mutation_result(
        data,
        "sheet_protection",
        before,
        after,
        f"sheets/{sheet_name}/protection",
    )


@mcp.tool()
def excel_set_protected_ranges(
    session_key: str,
    sheet_name: str,
    ranges: list[dict],
    mode: str = "replace",
) -> str:
    """Add/update/delete editable protected ranges without toggling sheet protection."""
    data = _get_session(session_key)
    sheet = _find_sheet(data, sheet_name)
    before = copy.deepcopy(sheet.get("protected_ranges") or [])
    if mode == "replace":
        after = copy.deepcopy(ranges)
    elif mode == "patch":
        by_name = {item.get("name"): copy.deepcopy(item) for item in before}
        for item in ranges:
            name = item.get("name")
            if not name:
                raise ValueError("Each protected range requires name.")
            if item.get("delete"):
                by_name.pop(name, None)
            else:
                by_name[name] = {**by_name.get(name, {}), **copy.deepcopy(item)}
        after = list(by_name.values())
    else:
        raise ValueError("mode must be replace or patch.")
    sheet["protected_ranges"] = after
    return _mutation_result(data, "protected_ranges", before, after, f"sheets/{sheet_name}/protected_ranges")


def _append_drawing_creation(data: dict, sheet_name: str, drawing_type: str, payload: dict) -> str:
    import uuid

    sheet = _find_sheet(data, sheet_name)
    item = {"id": uuid.uuid4().hex, "type": drawing_type, **copy.deepcopy(payload)}
    sheet.setdefault("drawing_creations", []).append(item)
    return _mutation_result(
        data,
        "drawings",
        None,
        item,
        f"sheets/{sheet_name}/drawing_creations/{item['id']}",
    )


@mcp.tool()
def excel_add_image(
    session_key: str,
    sheet_name: str,
    anchor: str,
    source_path: str | None = None,
    base64_data: str | None = None,
    mime_type: str | None = None,
    width: float | None = None,
    height: float | None = None,
    name: str | None = None,
) -> str:
    """Queue an image creation with bytes, anchor, dimensions, and relationship metadata."""
    import base64
    import mimetypes

    if bool(source_path) == bool(base64_data):
        raise ValueError("Provide exactly one of source_path or base64_data.")
    if source_path:
        path = Path(source_path).expanduser().resolve()
        raw = path.read_bytes()
        base64_data = base64.b64encode(raw).decode("ascii")
        mime_type = mime_type or mimetypes.guess_type(path.name)[0]
        name = name or path.name
    _, image_format = _validate_image_creation_payload(base64_data)
    if mime_type is None:
        mime_type = "image/jpeg" if image_format in {"jpeg", "jpg"} else f"image/{image_format}"
    data = _get_session(session_key)
    return _append_drawing_creation(data, sheet_name, "image", {
        "anchor": anchor, "base64": base64_data, "mime_type": mime_type,
        "width": width, "height": height, "name": name,
    })


@mcp.tool()
def excel_add_chart(
    session_key: str,
    sheet_name: str,
    chart_type: str,
    source_range: str,
    anchor: str,
    title: str | None = None,
    width: float | None = None,
    height: float | None = None,
    options: dict | None = None,
) -> str:
    """Queue a relationship-aware chart creation from a worksheet range."""
    data = _get_session(session_key)
    return _append_drawing_creation(data, sheet_name, "chart", {
        "chart_type": chart_type, "source_range": source_range, "anchor": anchor,
        "title": title, "width": width, "height": height,
        "options": copy.deepcopy(options or {}),
    })


@mcp.tool()
def excel_add_shape(
    session_key: str,
    sheet_name: str,
    shape_type: str,
    anchor: str,
    text: str | None = None,
    rich_text: dict | None = None,
    width: float | None = None,
    height: float | None = None,
    style: dict | None = None,
    name: str | None = None,
) -> str:
    """Queue a DrawingML shape using the common rich-text style model."""
    data = _get_session(session_key)
    if rich_text is not None:
        runs, plain = _reindex_rich_runs(rich_text.get("runs") or [])
        rich_text = {**copy.deepcopy(rich_text), "runs": runs, "text": plain}
        text = plain
    return _append_drawing_creation(data, sheet_name, "shape", {
        "shape_type": shape_type, "anchor": anchor, "text": text,
        "rich_text": rich_text, "width": width, "height": height,
        "style": copy.deepcopy(style or {}), "name": name,
    })


async def _run_frozen_stdio_server() -> None:
    import anyio
    from io import TextIOWrapper
    from mcp.server.stdio import stdio_server

    stdin_text = TextIOWrapper(sys.stdin.buffer, encoding="utf-8", errors="replace")
    stdout_text = TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    try:
        async with stdio_server(
            stdin=anyio.wrap_file(stdin_text),
            stdout=anyio.wrap_file(stdout_text),
        ) as (read_stream, write_stream):
            await mcp._mcp_server.run(
                read_stream,
                write_stream,
                mcp._mcp_server.create_initialization_options(),
            )
    finally:
        for stream in (stdin_text, stdout_text):
            try:
                stream.detach()
            except (OSError, ValueError):
                pass


if __name__ == "__main__":
    if len(sys.argv) >= 2 and sys.argv[1] == "--excel-worker":
        if len(sys.argv) != 3:
            raise SystemExit(2)
        raise SystemExit(run_worker_cli(sys.argv[2]))
    if getattr(sys, "frozen", False):
        import anyio
        anyio.run(_run_frozen_stdio_server)
    else:
        mcp.run(transport="stdio")
