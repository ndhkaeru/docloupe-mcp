"""
Full-metadata Excel serializer / reconstructor.

Serialize: Excel → dict with per-cell fill, font, merge, alignment, numfmt,
           column widths, row heights, borders.
Reconstruct: dict → Excel (.xlsx) preserving all of the above.
"""

import base64
import copy
import hashlib
import json
import re
import shutil
import zipfile
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape, quoteattr
import os
from pathlib import Path
from urllib.parse import urlparse, unquote

# lxml can hang inside the PyInstaller onefile build on this toolchain.
# Force openpyxl to use the stdlib/et_xmlfile XML path before any openpyxl import.
os.environ.setdefault("OPENPYXL_LXML", "False")


# ── URI → Path ────────────────────────────────────────────────────────────────

def uri_to_path(uri: str) -> Path:
    if uri.startswith("file://"):
        path = unquote(urlparse(uri).path)
        # Windows: /D:/foo/bar.xlsx → D:/foo/bar.xlsx
        if path.startswith("/") and len(path) > 2 and path[2] == ":":
            path = path[1:]
        return Path(path)
    return Path(uri)


def _close_openpyxl_workbook(workbook) -> None:
    """Close every archive owned by an openpyxl workbook."""
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        try:
            vba_archive.close()
        except Exception:
            pass
    try:
        workbook.close()
    except Exception:
        pass


# ── Color helpers ─────────────────────────────────────────────────────────────

# Default Office theme base colors (indices 0-11: dk1,lt1,dk2,lt2,accent1-6,hlink,folHlink)
_OFFICE_THEME_COLORS: list[str] = [
    "000000", "FFFFFF", "44546A", "E7E6E6",
    "4472C4", "ED7D31", "A5A5A5", "FFC000",
    "5B9BD5", "70AD47", "0563C1", "954F72",
]


def _apply_tint(hex6: str, tint: float) -> str:
    """Apply Excel luminance tint to a 6-char hex RGB and return 8-char ARGB."""
    import colorsys
    r, g, b = int(hex6[0:2], 16) / 255, int(hex6[2:4], 16) / 255, int(hex6[4:6], 16) / 255
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l = l * (1 + tint) if tint < 0 else l + (1 - l) * tint
    l = max(0.0, min(1.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return f"FF{int(r * 255):02X}{int(g * 255):02X}{int(b * 255):02X}"


def _wb_theme_colors(wb) -> list[str]:
    """
    Extract the 12 base theme RGB values from the workbook's theme XML.
    Falls back to Office 2016 defaults if the theme is absent or unparseable.
    """
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return list(_OFFICE_THEME_COLORS)
    try:
        from xml.etree import ElementTree as ET
        ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        root = ET.fromstring(raw if isinstance(raw, (bytes, str)) else raw)
        scheme = root.find(f".//{{{ns}}}clrScheme")
        if scheme is None:
            return list(_OFFICE_THEME_COLORS)
        result = list(_OFFICE_THEME_COLORS)
        for i, child in enumerate(scheme):
            if i >= 12:
                break
            srgb = child.find(f"{{{ns}}}srgbClr")
            if srgb is not None:
                val = srgb.get("val", "")
                if len(val) == 6:
                    result[i] = val.upper()
                continue
            sys_clr = child.find(f"{{{ns}}}sysClr")
            if sys_clr is not None:
                last = sys_clr.get("lastClr", "")
                if len(last) == 6:
                    result[i] = last.upper()
        return result
    except Exception:
        return list(_OFFICE_THEME_COLORS)


def _resolve_color(fc, theme_colors: list[str]) -> str | None:
    """
    Resolve an openpyxl Color to an 8-char ARGB hex string, or None if transparent/absent.
    Handles rgb, indexed, and theme types.
    Only excludes "00000000" (fully transparent) — all real colors including black and white
    are preserved so fill round-trips correctly.
    """
    if fc is None:
        return None
    try:
        if fc.type == "rgb":
            rgb = fc.rgb
            if isinstance(rgb, str) and rgb != "00000000":
                return rgb
        elif fc.type == "indexed":
            from openpyxl.styles.colors import COLOR_INDEX
            idx = fc.indexed
            if idx is not None and 0 <= idx < len(COLOR_INDEX):
                argb = COLOR_INDEX[idx]
                if argb != "00000000":
                    return argb
        elif fc.type == "theme":
            idx = fc.theme
            if idx is not None and 0 <= idx < len(theme_colors):
                base = theme_colors[idx]
                tint = fc.tint or 0.0
                return _apply_tint(base, tint) if tint else f"FF{base}"
    except Exception:
        pass
    return None


def _parse_xml_attrs(attr_text: str) -> dict[str, str]:
    """Parse a simple XML attribute string into {name: value}."""
    import re
    return {m.group(1): m.group(2) for m in re.finditer(r'([\w:.-]+)="([^"]*)"', attr_text)}


def _extract_xml_children(parent_xml: str, child_tag: str) -> list[str]:
    """Return raw child XML elements for a simple OOXML container."""
    import re
    return [m.group(0) for m in re.finditer(
        rf"<{child_tag}\b[^>]*/>|<{child_tag}\b[^>]*>.*?</{child_tag}>",
        parent_xml,
        re.DOTALL,
    )]


def _color_ref_from_openpyxl(fc) -> dict | None:
    """Serialize an openpyxl color reference for later reconstruction."""
    if fc is None:
        return None
    ctype = getattr(fc, "type", None)
    try:
        if ctype == "rgb" and isinstance(fc.rgb, str) and fc.rgb != "00000000":
            ref = {"type": "rgb", "rgb": fc.rgb}
        elif ctype == "theme" and fc.theme is not None:
            ref = {"type": "theme", "theme": int(fc.theme)}
        elif ctype == "indexed" and fc.indexed is not None:
            ref = {"type": "indexed", "indexed": int(fc.indexed)}
        elif ctype == "auto" and fc.auto is not None:
            ref = {"type": "auto", "auto": bool(fc.auto)}
        else:
            return None
        tint = getattr(fc, "tint", None)
        if tint:
            ref["tint"] = float(tint)
        return ref
    except Exception:
        return None


def _make_color_from_ref(ref: dict):
    from openpyxl.styles.colors import Color

    ctype = ref.get("type")
    kw = {}
    if ctype == "theme":
        kw["theme"] = int(ref["theme"])
    elif ctype == "indexed":
        kw["indexed"] = int(ref["indexed"])
    elif ctype == "auto":
        kw["auto"] = bool(ref.get("auto", True))
    elif ctype == "rgb":
        kw["rgb"] = ref["rgb"]
    else:
        return None
    if ref.get("tint") is not None:
        kw["tint"] = float(ref["tint"])
    return Color(**kw)


def _usable_raw_fill(cd: dict) -> dict | None:
    raw = cd.get("_fill_raw")
    if not raw:
        return None
    # Keep the raw fill only while the public resolved fill has not changed.
    if cd.get("fill") != raw.get("rgb"):
        return None
    return raw


def _usable_raw_font(cd: dict) -> dict | None:
    raw = cd.get("_font_raw")
    if not raw:
        return None
    checks = (
        ("font", "name"),
        ("size", "size"),
        ("bold", "bold"),
        ("italic", "italic"),
        ("uline", "underline"),
        ("strike", "strike"),
        ("vAlign", "vertAlign"),
    )
    for public_key, raw_key in checks:
        if raw.get(raw_key) != cd.get(public_key):
            return None
    return raw


def _usable_raw_color_ref(ref: dict | None, current_rgb: str | None) -> dict | None:
    if not ref:
        return None
    raw_rgb = ref.get("rgb")
    if current_rgb == raw_rgb:
        return ref
    if current_rgb is None and raw_rgb in (None, "FF000000", "00000000"):
        return ref
    return None


def _font_raw_from_openpyxl(font, fcolor: str | None) -> dict | None:
    if font is None:
        return None
    raw = {
        "name":      font.name,
        "size":      font.size,
        "bold":      bool(font.bold),
        "italic":    bool(font.italic),
        "underline": font.underline,
        "strike":    bool(font.strike),
        "vertAlign": font.vertAlign,
    }
    color_ref = _color_ref_from_openpyxl(font.color)
    if color_ref:
        if color_ref.get("type") != "rgb":
            color_ref["rgb"] = fcolor
        raw["color"] = color_ref
    for attr in ("charset", "family", "scheme", "outline", "shadow", "condense", "extend"):
        value = getattr(font, attr, None)
        if value is not None:
            raw[attr] = value
    return raw if len(raw) > 7 else None


def _apply_raw_font_kwargs(fk: dict, cd: dict) -> None:
    raw = _usable_raw_font(cd)
    if not raw:
        if cd.get("fcolor"):
            fk["color"] = cd["fcolor"]
        return

    color_ref = _usable_raw_color_ref(raw.get("color"), cd.get("fcolor"))
    if color_ref:
        color = _make_color_from_ref(color_ref)
        if color is not None:
            fk["color"] = color
    elif cd.get("fcolor"):
        fk["color"] = cd["fcolor"]

    for attr in ("charset", "family", "scheme", "outline", "shadow", "condense", "extend"):
        if attr in raw:
            fk[attr] = raw[attr]


def _make_pattern_fill_from_raw(raw: dict):
    from openpyxl.styles import PatternFill

    if raw.get("is_gradient"):
        import hashlib
        # Placeholder only. _inject_raw_fills patches the saved fill XML back to
        # the original gradient fill.
        color = "FF" + hashlib.sha1((raw.get("xml") or "").encode("utf-8")).hexdigest()[:6].upper()
        return PatternFill("solid", fgColor=color)

    pattern_type = raw.get("patternType") or "solid"
    fg = _make_color_from_ref(raw.get("fgColor") or {})
    bg = _make_color_from_ref(raw.get("bgColor") or {})
    kwargs = {"fill_type": pattern_type}
    if fg is not None:
        kwargs["fgColor"] = fg
    if bg is not None:
        kwargs["bgColor"] = bg
    return PatternFill(**kwargs)


def _fill_xml_has_color_reference(fill_xml: str) -> bool:
    import re
    return bool(re.search(r"<(?:fgColor|bgColor)\b[^>]*(?:\btheme=|\bindexed=|\bauto=|\btint=)", fill_xml))


def _fill_xml_should_preserve(fill_xml: str) -> bool:
    import re

    if _fill_xml_has_color_reference(fill_xml):
        return True
    if re.search(r"<gradientFill\b", fill_xml):
        return True
    pattern_m = re.search(r"<patternFill\b([^>]*)", fill_xml)
    if not pattern_m:
        return False
    pattern_type = _parse_xml_attrs(pattern_m.group(1)).get("patternType")
    if not pattern_type or pattern_type in ("none", "solid"):
        return False
    if pattern_type == "gray125":
        # Bare gray125 is openpyxl's built-in default fill slot; gray125 with
        # explicit colors is a real user fill and must be preserved.
        return "<fgColor" in fill_xml or "<bgColor" in fill_xml
    return True


def _extract_raw_fill_data(xlsx_path, sheet_file_map: dict) -> dict:
    """
    Extract raw fill XML per cell for fills that use theme/indexed/auto/tint
    references. The public API still exposes resolved RGB; this metadata is used
    only to keep the original OOXML color reference on save.
    """
    import zipfile, re

    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            if "xl/styles.xml" not in zf.namelist():
                return result
            styles_xml = zf.read("xl/styles.xml").decode("utf-8")
            fills_m = re.search(r"<fills\b[^>]*>.*?</fills>", styles_xml, re.DOTALL)
            xfs_m = re.search(r"<cellXfs\b[^>]*>.*?</cellXfs>", styles_xml, re.DOTALL)
            if not fills_m or not xfs_m:
                return result

            fills = _extract_xml_children(fills_m.group(0), "fill")
            style_to_fill: dict[int, int] = {}
            for idx, xf_m in enumerate(re.finditer(r"<xf\b([^>]*)/?>", xfs_m.group(0))):
                attrs = _parse_xml_attrs(xf_m.group(1))
                fill_id = attrs.get("fillId")
                if fill_id is not None:
                    style_to_fill[idx] = int(fill_id)

            fill_raw: dict[int, str] = {
                idx: xml for idx, xml in enumerate(fills)
                if _fill_xml_should_preserve(xml)
            }
            if not fill_raw:
                return result

            for sname, sheet_file in sheet_file_map.items():
                if sheet_file not in zf.namelist():
                    continue
                sheet_xml = zf.read(sheet_file).decode("utf-8")
                cells = {}
                for cell_m in re.finditer(r"<c\b([^>]*)", sheet_xml):
                    attrs = _parse_xml_attrs(cell_m.group(1))
                    coord = attrs.get("r")
                    style_idx = attrs.get("s")
                    if coord is None or style_idx is None:
                        continue
                    fill_id = style_to_fill.get(int(style_idx))
                    if fill_id in fill_raw:
                        cells[coord] = {"xml": fill_raw[fill_id]}
                if cells:
                    result[sname] = cells
    except Exception:
        pass
    return result


def _extract_sheet_view_attrs(xlsx_path, sheet_file_map: dict) -> dict:
    """Extract raw sheetView opening-tag attributes per sheet."""
    import zipfile, re

    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            for sname, sheet_file in sheet_file_map.items():
                if sheet_file not in zf.namelist():
                    continue
                sheet_xml = zf.read(sheet_file).decode("utf-8")
                m = re.search(r"<sheetView\b([^>]*)>", sheet_xml)
                if not m:
                    continue
                raw_attrs = m.group(1).strip()
                if raw_attrs.endswith("/"):
                    raw_attrs = raw_attrs[:-1].rstrip()
                result[sname] = {
                    "raw": raw_attrs,
                    "attrs": _parse_xml_attrs(raw_attrs),
                }
    except Exception:
        pass
    return result


def _extract_row_attrs(xlsx_path, sheet_file_map: dict) -> dict:
    """Extract exact non-coordinate row attributes per worksheet."""
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as archive:
            for sheet_name, sheet_file in sheet_file_map.items():
                if sheet_file not in archive.namelist():
                    continue
                root = ET.fromstring(archive.read(sheet_file))
                sheet_data = root.find(_qname("sheetData"))
                rows = {}
                for row in (list(sheet_data) if sheet_data is not None else []):
                    row_number = row.get("r")
                    if not row_number:
                        continue
                    attrs = {
                        _local_name(key): value
                        for key, value in row.attrib.items()
                        if _local_name(key) != "r"
                    }
                    if attrs:
                        rows[row_number] = attrs
                if rows:
                    result[sheet_name] = rows
    except Exception:
        pass
    return result


def _extract_sheet_format_data(xlsx_path, sheet_file_map: dict) -> dict:
    """Extract raw worksheet root attrs, sheetFormatPr XML, and cols XML per sheet."""
    import zipfile, re

    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            for sname, sheet_file in sheet_file_map.items():
                if sheet_file not in zf.namelist():
                    continue
                sheet_xml = zf.read(sheet_file).decode("utf-8")
                root_m = re.search(r"<worksheet\b([^>]*)>", sheet_xml)
                sf_m = re.search(
                    r"<sheetFormatPr\b[^>]*/>|<sheetFormatPr\b[^>]*>.*?</sheetFormatPr>",
                    sheet_xml,
                    re.DOTALL,
                )
                cols_m = re.search(r"<cols\b[^>]*>.*?</cols>", sheet_xml, re.DOTALL)
                if root_m or sf_m or cols_m:
                    result[sname] = {
                        "root_attrs": root_m.group(1).strip() if root_m else "",
                        "sheetFormatPr": sf_m.group(0) if sf_m else None,
                        "cols": cols_m.group(0) if cols_m else None,
                    }
    except Exception:
        pass
    return result


def _serialize_sheet_view(sv, raw_view: dict | None = None) -> dict:
    attrs = {}
    for attr in getattr(type(sv), "__attrs__", ()):
        try:
            value = getattr(sv, attr)
        except Exception:
            continue
        if value is not None:
            attrs[attr] = value

    # Backwards-compatible alias used by older session data.
    if "zoomScale" in attrs:
        attrs["zoom"] = attrs["zoomScale"]
    if raw_view:
        attrs["_raw_attrs"] = raw_view.get("raw")
        attrs["_raw_attr_values"] = raw_view.get("attrs")
    return attrs


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ser_border_side(side, theme_colors: list) -> dict | None:
    """Return compact border side dict, or None if no border style set."""
    if side is None or not side.border_style:
        return None
    d = {"style": side.border_style}
    color = _resolve_color(side.color, theme_colors)
    color_ref = _color_ref_from_openpyxl(side.color)
    if color_ref:
        color_ref["rgb"] = color
        d["_color_raw"] = color_ref
    # Suppress default black — it is the implicit border color
    if color and color not in ("FF000000", "00000000"):
        d["color"] = color
    return d


def _ser_validations(ws) -> list:
    result = []
    for dv in ws.data_validations.dataValidation:
        item = {"formula1": dv.formula1, "formula2": dv.formula2}
        for attr in getattr(type(dv), "__attrs__", ()):
            value = getattr(dv, attr, None)
            if attr == "sqref":
                value = str(value)
            if value is not None:
                item[attr] = value
        result.append(item)
    return result


def _ser_column_dimensions(ws) -> tuple[dict, dict, dict]:
    import openpyxl.utils

    widths = {}
    hidden = {}
    outline = {}
    for key, cd in ws.column_dimensions.items():
        try:
            start = cd.min or openpyxl.utils.column_index_from_string(key)
            end = cd.max or start
        except Exception:
            start = end = openpyxl.utils.column_index_from_string(key)
        for idx in range(start, end + 1):
            letter = openpyxl.utils.get_column_letter(idx)
            widths[letter] = cd.width
            if cd.hidden:
                hidden[letter] = True
            if cd.outlineLevel:
                outline[letter] = cd.outlineLevel
    return widths, hidden, outline


def _dimension_state(cw: dict | None, ch: dict | None, co: dict | None) -> dict:
    def _widths(values):
        result = {}
        for key, value in (values or {}).items():
            if value is None:
                result[str(key)] = None
            else:
                try:
                    result[str(key)] = round(float(value), 10)
                except Exception:
                    result[str(key)] = value
        return result

    return {
        "cw": _widths(cw),
        "ch": {str(k): bool(v) for k, v in (ch or {}).items() if v},
        "co": {str(k): int(v) for k, v in (co or {}).items() if v},
    }


def _normalize_dimension_state(state: dict | None) -> dict:
    state = state or {}
    return _dimension_state(state.get("cw"), state.get("ch"), state.get("co"))


def _make_border_side(sd):
    from openpyxl.styles import Side
    if not sd:
        return Side()
    from openpyxl.styles.colors import Color
    kw = {"border_style": sd["style"]}
    raw_color = _usable_raw_color_ref(sd.get("_color_raw"), sd.get("color"))
    if raw_color:
        color = _make_color_from_ref(raw_color)
        if color is not None:
            kw["color"] = color
    elif sd.get("color"):
        kw["color"] = Color(rgb=sd["color"])
    return Side(**kw)


def _xlsx_sheet_file_map(wb_xml: str, rels_xml: str) -> dict:
    """Return {sheet_name: zip_path} from workbook XML and its .rels XML."""
    import re
    from html import unescape
    rel_map = {}
    for m in re.finditer(r'<(?:\w+:)?Relationship\b([^>]+)/>', rels_xml):
        attrs = m.group(1)
        id_m  = re.search(r'\bId="([^"]+)"', attrs)
        tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
        if id_m and tgt_m:
            rel_map[id_m.group(1)] = tgt_m.group(1)
    result = {}
    for m in re.finditer(r'<sheet\b[^>]+\bname="([^"]+)"[^>]+\br:id="([^"]+)"', wb_xml):
        raw_name, rid = m.group(1), m.group(2)
        sname = unescape(raw_name)
        if rid not in rel_map:
            continue
        t = rel_map[rid]
        # Normalize: /xl/worksheets/sheet1.xml or ../worksheets/sheet1.xml
        if t.startswith("/"):
            t = t.lstrip("/")           # /xl/... → xl/...
        elif t.startswith("../"):
            t = "xl/" + t[3:]           # ../worksheets/... → xl/worksheets/...
        elif not t.startswith("xl/"):
            t = "xl/" + t
        result[sname] = t
    return result


def _normalize_rel_target(target: str, prefix: str = "xl/") -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    if target.startswith("../"):
        # Both worksheet and drawing rels live one directory below xl/
        # (xl/worksheets/, xl/drawings/), so ../foo always resolves to xl/foo.
        return "xl/" + target[3:]
    if not target.startswith(prefix):
        return prefix + target
    return target


def _xlsx_parts(path: str) -> set[str]:
    import zipfile
    with zipfile.ZipFile(str(path), "r") as zf:
        return set(zf.namelist())


def inspect_xlsx_package(path: str) -> dict:
    """Return a compact, validation-oriented summary of an .xlsx package."""
    import re
    import xml.etree.ElementTree as ET
    import zipfile

    required = {"[Content_Types].xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
    errors: list[str] = []
    warnings: list[str] = []
    features: dict[str, int] = {}
    parts: list[str] = []

    try:
        with zipfile.ZipFile(str(path), "r") as zf:
            bad = zf.testzip()
            if bad:
                errors.append(f"corrupt zip member: {bad}")
            parts = sorted(zf.namelist())
            missing = sorted(required - set(parts))
            if missing:
                errors.append(f"missing required parts: {missing}")

            xml_parts = [p for p in parts if p.endswith((".xml", ".rels"))]
            for part in xml_parts:
                try:
                    ET.fromstring(zf.read(part))
                except Exception as exc:
                    errors.append(f"invalid XML in {part}: {exc}")

            prefixes = {
                "drawings": "xl/drawings/",
                "charts": "xl/charts/",
                "media": "xl/media/",
                "vml_drawings": "xl/drawings/vmlDrawing",
                "pivot_tables": "xl/pivotTables/",
                "slicers": "xl/slicers/",
                "external_links": "xl/externalLinks/",
                "custom_xml": "customXml/",
            }
            for name, prefix in prefixes.items():
                features[name] = sum(1 for p in parts if p.startswith(prefix))
            features["vba_project"] = int("xl/vbaProject.bin" in parts)
            features["unknown_parts"] = sum(
                1 for p in parts
                if not p.startswith(("_rels/", "docProps/", "xl/", "customXml/"))
                and p != "[Content_Types].xml"
            )

            referenced_parts: set[str] = set()
            for rel_part in [p for p in parts if p.endswith(".rels")]:
                rel_base = ""
                if rel_part == "_rels/.rels":
                    rel_base = ""
                elif "/_rels/" in rel_part:
                    folder, rel_name = rel_part.split("/_rels/", 1)
                    rel_base = folder.rsplit("/", 1)[0] + "/" if "/" in folder else ""
                    rel_base += rel_name[:-5]
                try:
                    rel_xml = zf.read(rel_part).decode("utf-8")
                    for match in re.finditer(r'\bTarget="([^"]+)"', rel_xml):
                        target = match.group(1)
                        if target.startswith(("http://", "https://", "mailto:")):
                            continue
                        if target.startswith("/"):
                            referenced_parts.add(target.lstrip("/"))
                        else:
                            import posixpath
                            referenced_parts.add(posixpath.normpath(posixpath.join(posixpath.dirname(rel_base), target)))
                except Exception:
                    pass
            orphan_advanced = [
                p for p in parts
                if p.startswith(("xl/pivotTables/", "xl/externalLinks/"))
                and p not in referenced_parts
            ]
            if orphan_advanced:
                warnings.append(f"advanced parts are present but not referenced by relationships: {orphan_advanced[:8]}")
            if any(features.get(k, 0) for k in ("vml_drawings", "pivot_tables", "slicers", "external_links", "vba_project")):
                warnings.append("workbook contains advanced parts that are preserved best-effort only")
    except Exception as exc:
        errors.append(f"cannot read xlsx package: {exc}")

    return {
        "path": str(path),
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "part_count": len(parts),
        "features": features,
    }


def validate_xlsx(path: str) -> list[str]:
    """Raise ValueError if the saved workbook package is structurally broken."""
    report = inspect_xlsx_package(path)
    if not report["valid"]:
        raise ValueError("Saved workbook failed validation: " + "; ".join(report["errors"]))
    return report["warnings"]


def diff_xlsx_package(before: str, after: str) -> dict:
    """Compare ZIP package manifests and semantic part content hashes after a save."""
    from preservation import package_content_diff

    return package_content_diff(before, after)




def _restore_missing_package_parts(source_path: str | None, xlsx_path: str) -> str | None:
    """Best-effort copy of advanced OOXML parts that openpyxl dropped."""
    if not source_path or str(source_path) == str(xlsx_path):
        return None
    import os
    import zipfile

    if not os.path.exists(str(source_path)):
        return None

    tmp = str(xlsx_path) + ".~parts.tmp"
    def merge_relationships(current_xml: str, source_xml: str, restored_parts: list[str]) -> str:
        import posixpath
        import re

        restored = set(restored_parts)
        existing_ids = set(re.findall(r'\bId="([^"]+)"', current_xml))
        existing_targets = set(re.findall(r'\bTarget="([^"]+)"', current_xml))
        additions = []
        for match in re.finditer(r'<Relationship\b([^>]*)/>', source_xml):
            rel = match.group(0)
            attrs = match.group(1)
            target_m = re.search(r'\bTarget="([^"]+)"', attrs)
            id_m = re.search(r'\bId="([^"]+)"', attrs)
            if not target_m:
                continue
            target = target_m.group(1)
            norm = target.lstrip("/") if target.startswith("/") else posixpath.normpath("xl/" + target)
            if norm not in restored and "vbaProject.bin" not in norm:
                continue
            if target in existing_targets:
                continue
            if id_m and id_m.group(1) in existing_ids:
                next_id = 1
                while f"rId{next_id}" in existing_ids:
                    next_id += 1
                new_id = f"rId{next_id}"
                existing_ids.add(new_id)
                rel = re.sub(r'\bId="[^"]+"', f'Id="{new_id}"', rel, count=1)
            elif id_m:
                existing_ids.add(id_m.group(1))
            existing_targets.add(target)
            additions.append(rel)
        if additions:
            current_xml = current_xml.replace("</Relationships>", "".join(additions) + "</Relationships>", 1)
        return current_xml

    def merge_content_types(current_xml: str, source_xml: str, restored_parts: list[str]) -> str:
        import re
        workbook_override = re.search(r'<Override\b[^>]*\bPartName="/xl/workbook.xml"[^>]*/>', source_xml)
        if workbook_override and (
            "macroEnabled" in workbook_override.group(0)
            or "template" in workbook_override.group(0)
        ):
            current_xml = re.sub(
                r'<Override\b[^>]*\bPartName="/xl/workbook.xml"[^>]*/>',
                workbook_override.group(0),
                current_xml,
                count=1,
            )
        needed_exts = {p.rsplit(".", 1)[-1].lower() for p in restored_parts if "." in p}
        for match in re.finditer(r'<Default\b[^>]*\bExtension="([^"]+)"[^>]*/>', source_xml):
            ext = match.group(1).lower()
            if ext in needed_exts and f'Extension="{ext}"' not in current_xml:
                current_xml = current_xml.replace("</Types>", match.group(0) + "</Types>", 1)
        for part in restored_parts:
            part_name = "/" + part
            if f'PartName="{part_name}"' in current_xml:
                continue
            match = re.search(rf'<Override\b[^>]*\bPartName="{re.escape(part_name)}"[^>]*/>', source_xml)
            if match:
                current_xml = current_xml.replace("</Types>", match.group(0) + "</Types>", 1)
        return current_xml

    try:
        with zipfile.ZipFile(str(source_path), "r") as src, zipfile.ZipFile(str(xlsx_path), "r") as cur:
            current = set(cur.namelist())
            missing = [p for p in src.namelist() if p not in current]
            source_names = set(src.namelist())
            if not missing and "xl/vbaProject.bin" not in source_names:
                return None
            unsafe_prefixes = (
                "xl/worksheets/",      # regenerated sheets own their rel ids
                "xl/drawings/",        # handled by _inject_drawing_data
                "xl/media/",           # handled through drawing relationships
                "xl/charts/",          # handled through drawing relationships
                "xl/printerSettings/", # old sheet rels/pageSetup ids are not stable
            )
            restored = [
                p for p in missing
                if p not in {"xl/workbook.xml", "xl/styles.xml", "xl/sharedStrings.xml", "[Content_Types].xml", "xl/calcChain.xml"}
                and not p.startswith(unsafe_prefixes)
            ]
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in cur.infolist():
                    raw = cur.read(item.filename)
                    if item.filename == "[Content_Types].xml":
                        raw = merge_content_types(
                            raw.decode("utf-8"),
                            src.read("[Content_Types].xml").decode("utf-8"),
                            restored,
                        ).encode("utf-8")
                    elif item.filename == "xl/_rels/workbook.xml.rels" and "xl/_rels/workbook.xml.rels" in source_names:
                        raw = merge_relationships(
                            raw.decode("utf-8"),
                            src.read("xl/_rels/workbook.xml.rels").decode("utf-8"),
                            restored,
                        ).encode("utf-8")
                    zout.writestr(item, raw)
                for part in restored:
                    zout.writestr(part, src.read(part))
        os.replace(tmp, str(xlsx_path))
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"advanced package part passthrough failed: {exc}"
    return None


def _build_chart_creation(ws, creation: dict):
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils.cell import range_boundaries

    chart_type = str(creation.get("chart_type") or "").strip().lower()
    if chart_type in {"bar", "column"}:
        chart = BarChart()
        chart.type = "bar" if chart_type == "bar" else "col"
    elif chart_type == "line":
        chart = LineChart()
    else:
        raise ValueError(f"Unsupported chart_type: {chart_type!r}")

    source_range = str(creation.get("source_range") or "").strip()
    try:
        min_col, min_row, max_col, max_row = range_boundaries(source_range)
    except Exception as exc:
        raise ValueError(f"Invalid chart source_range: {source_range!r}") from exc
    if max_col - min_col < 1 or max_row - min_row < 1:
        raise ValueError("Chart source_range must contain headers and at least two columns.")

    chart.add_data(
        Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row))
    if creation.get("title") is not None:
        chart.title = str(creation["title"])
    if creation.get("width") is not None:
        chart.width = float(creation["width"])
    if creation.get("height") is not None:
        chart.height = float(creation["height"])
    ws.add_chart(chart, str(creation.get("anchor") or "A1"))
    return chart


def _validate_image_creation_payload(encoded: str) -> tuple[bytes, str]:
    from io import BytesIO

    from PIL import Image as PillowImage

    if not isinstance(encoded, str) or not encoded:
        raise ValueError("Image creation requires non-empty base64 data.")
    try:
        raw = base64.b64decode(encoded, validate=True)
    except Exception as exc:
        raise ValueError("Image creation contains invalid base64 data.") from exc

    try:
        with PillowImage.open(BytesIO(raw)) as probe:
            image_format = (probe.format or "").lower()
            probe.verify()
    except Exception as exc:
        raise ValueError("Image creation contains unsupported or invalid image bytes.") from exc
    if image_format not in {"png", "jpeg", "jpg", "gif", "bmp", "tiff"}:
        raise ValueError(f"Unsupported image format: {image_format or 'unknown'}")
    return raw, image_format


def _build_image_creation(ws, creation: dict) -> dict:
    from io import BytesIO

    from openpyxl.drawing.image import Image as OpenpyxlImage

    raw, image_format = _validate_image_creation_payload(creation.get("base64"))

    image = OpenpyxlImage(BytesIO(raw))
    if creation.get("width") is not None:
        image.width = float(creation["width"])
    if creation.get("height") is not None:
        image.height = float(creation["height"])
    ws.add_image(image, str(creation.get("anchor") or "A1"))
    return {**creation, "_raw_bytes": raw, "_image_format": image_format}


def _stage_drawing_creations(wb, data: dict) -> dict:
    plan: dict[str, dict[str, list[dict]]] = {}
    for sheet_data in data.get("sheets") or []:
        creations = sheet_data.get("drawing_creations") or []
        if not creations:
            continue
        sheet_name = sheet_data["name"]
        ws = wb[sheet_name]
        sheet_plan = {"charts": [], "images": [], "shapes": []}
        for creation in creations:
            creation_type = creation.get("type")
            if creation_type == "chart":
                _build_chart_creation(ws, creation)
                sheet_plan["charts"].append(copy.deepcopy(creation))
            elif creation_type == "image":
                sheet_plan["images"].append(_build_image_creation(ws, copy.deepcopy(creation)))
            elif creation_type == "shape":
                sheet_plan["shapes"].append(copy.deepcopy(creation))
            else:
                raise ValueError(f"Unsupported drawing creation type: {creation_type!r}")
        plan[sheet_name] = sheet_plan
    return plan


def _extract_drawing_data(xlsx_path, sheet_file_map: dict) -> dict:
    """
    Extract drawing/chart/image files per sheet name.
    Returns {sname: {"drawing_xml": str, "drawing_rels": str, "files": {path: b64str}}}.
    """
    import zipfile, re, base64
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            namelist = set(zf.namelist())
            for sname, sheet_file in sheet_file_map.items():
                # xl/worksheets/sheetN.xml → xl/worksheets/_rels/sheetN.xml.rels
                parts = sheet_file.rsplit("/", 1)
                sheet_rels_file = parts[0] + "/_rels/" + parts[1] + ".rels"
                if sheet_rels_file not in namelist:
                    continue
                rels_content = zf.read(sheet_rels_file).decode("utf-8")
                drawing_file = None
                for rm in re.finditer(r'<Relationship\b([^>]+)/>', rels_content):
                    attrs = rm.group(1)
                    type_m = re.search(r'\bType="([^"]+)"', attrs)
                    # Only DrawingML drawings — NOT vmlDrawing (comment shapes),
                    # which openpyxl regenerates itself on save.
                    if not type_m or not type_m.group(1).rstrip("/").endswith("/drawing"):
                        continue
                    tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
                    if tgt_m:
                        drawing_file = _normalize_rel_target(tgt_m.group(1), "xl/drawings/")
                        break
                if not drawing_file or drawing_file not in namelist:
                    continue

                sd = {
                    "drawing_file": drawing_file,
                    "drawing_xml": zf.read(drawing_file).decode("utf-8"),
                    "drawing_rels": None,
                    "files": {},
                }

                dr_rels_path = drawing_file.rsplit("/", 1)
                dr_rels_file = dr_rels_path[0] + "/_rels/" + dr_rels_path[1] + ".rels"
                if dr_rels_file in namelist:
                    dr_rels = zf.read(dr_rels_file).decode("utf-8")
                    sd["drawing_rels"] = dr_rels
                    for rm in re.finditer(r'<Relationship\b([^>]+)/>', dr_rels):
                        attrs = rm.group(1)
                        tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
                        if tgt_m:
                            tgt = _normalize_rel_target(tgt_m.group(1), "xl/")
                            if tgt in namelist:
                                sd["files"][tgt] = base64.b64encode(zf.read(tgt)).decode()

                result[sname] = sd
    except Exception:
        pass
    return result


def _shape_texts_from_drawing_xml(xml: str) -> list[str]:
    import re
    from html import unescape
    texts = []
    for block in re.findall(r"<a:t(?:\s[^>]*)?>(.*?)</a:t>", xml, flags=re.DOTALL):
        texts.append(unescape(re.sub(r"<[^>]+>", "", block)))
    return texts


def _drawingml_attrs(fragment: str) -> dict[str, str]:
    from html import unescape

    return {
        name.split(":")[-1]: unescape(value)
        for name, _quote, value in re.findall(
            r"([A-Za-z_][\w:.-]*)\s*=\s*(['\"])(.*?)\2",
            fragment,
            flags=re.DOTALL,
        )
    }


def _shape_run_font_from_xml(run_xml: str) -> dict:
    rpr_match = re.search(
        r"<a:rPr\b([^>]*)>(.*?)</a:rPr>|<a:rPr\b([^>]*)/>",
        run_xml,
        flags=re.DOTALL,
    )
    if not rpr_match:
        return {}
    attrs = _drawingml_attrs(rpr_match.group(1) or rpr_match.group(3) or "")
    body = rpr_match.group(2) or ""
    font = {}
    for source, target in (("b", "bold"), ("i", "italic")):
        if source in attrs:
            font[target] = attrs[source].lower() in {"1", "true", "on"}
    if "strike" in attrs:
        font["strike"] = attrs["strike"] not in {"noStrike", "none"}
    if "u" in attrs and attrs["u"] not in {"none", "0", "false"}:
        font["underline"] = {
            "sng": "single",
            "dbl": "double",
        }.get(attrs["u"], attrs["u"])
    if "sz" in attrs:
        try:
            font["size"] = int(attrs["sz"]) / 100
        except (TypeError, ValueError):
            pass
    color_match = re.search(r"<a:srgbClr\b[^>]*\bval=\"([0-9A-Fa-f]{6,8})\"", body)
    if color_match:
        rgb = color_match.group(1).upper()
        font["color"] = {"type": "rgb", "rgb": rgb if len(rgb) == 8 else "FF" + rgb}
    latin_match = re.search(r"<a:latin\b[^>]*\btypeface=\"([^\"]*)\"", body)
    if latin_match:
        from html import unescape
        font["name"] = unescape(latin_match.group(1))
    return font


def _shape_rich_text_from_drawing_xml(xml: str) -> dict | None:
    from html import unescape

    paragraphs = re.findall(r"<a:p\b[^>]*>(.*?)</a:p>", xml, flags=re.DOTALL)
    if not paragraphs:
        return None
    runs = []
    offset = 0

    def append(text: str, font: dict | None = None) -> None:
        nonlocal offset
        if not text:
            return
        run = {"text": text, "start": offset, "end": offset + len(text)}
        if font:
            run["font"] = font
        runs.append(run)
        offset = run["end"]

    for paragraph_index, paragraph in enumerate(paragraphs):
        if paragraph_index:
            append("\n")
        tokens = re.findall(
            r"<a:r\b[^>]*>.*?</a:r>|<a:fld\b[^>]*>.*?</a:fld>|<a:br\b[^>]*/>",
            paragraph,
            flags=re.DOTALL,
        )
        if not tokens:
            tokens = re.findall(r"<a:t(?:\s[^>]*)?>.*?</a:t>", paragraph, flags=re.DOTALL)
        for token in tokens:
            if re.match(r"<a:br\b", token):
                append("\n")
                continue
            text_match = re.search(r"<a:t(?:\s[^>]*)?>(.*?)</a:t>", token, flags=re.DOTALL)
            if not text_match:
                continue
            text = unescape(re.sub(r"<[^>]+>", "", text_match.group(1)))
            append(text, _shape_run_font_from_xml(token))
    text = "".join(run["text"] for run in runs)
    return {"text": text, "runs": runs}


def _extract_shape_inventory(drawing_data: dict) -> dict:
    """Build lightweight shape metadata from preserved DrawingML XML."""
    import re
    from html import unescape
    result: dict[str, list[dict]] = {}
    for sname, sd in drawing_data.items():
        xml = sd.get("drawing_xml") or ""
        shapes = []
        anchor_re = r"<(?:(?:xdr:)?)(twoCellAnchor|oneCellAnchor|absoluteAnchor)\b.*?</(?:(?:xdr:)?)(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)>"
        for idx, match in enumerate(re.finditer(anchor_re, xml, re.DOTALL), 1):
            anchor_xml = match.group(0)
            name_m = re.search(r'<(?:xdr:)?cNvPr\b[^>]*\bname="([^"]*)"', anchor_xml)
            id_m = re.search(r'<(?:xdr:)?cNvPr\b[^>]*\bid="([^"]*)"', anchor_xml)
            rel_m = re.search(r'\br:embed="([^"]+)"|\br:link="([^"]+)"', anchor_xml)
            kind = "shape"
            if re.search(r"<(?:xdr:)?pic\b", anchor_xml):
                kind = "picture"
            elif "graphicData" in anchor_xml and "/chart" in anchor_xml:
                kind = "chart"
            elif re.search(r"<(?:xdr:)?sp\b", anchor_xml):
                kind = "shape"
            rich_text = _shape_rich_text_from_drawing_xml(anchor_xml) if kind == "shape" else None
            shapes.append({
                "index": idx,
                "id": id_m.group(1) if id_m else None,
                "name": unescape(name_m.group(1)) if name_m else None,
                "type": kind,
                "text": (rich_text or {}).get("text") or "".join(_shape_texts_from_drawing_xml(anchor_xml)) or None,
                "rich_text": rich_text,
                "relationship_id": next((g for g in (rel_m.groups() if rel_m else ()) if g), None),
            })
        if shapes:
            result[sname] = shapes
    return result


def _inject_drawing_data(xlsx_path: str, drawing_data: dict,
                         sheet_name_to_new_file: dict) -> str | None:
    """
    Inject preserved DrawingML drawings, charts, and media into a saved xlsx file.
    sheet_name_to_new_file: {sheet_name: "xl/worksheets/sheetN.xml"} in the NEW file.
    """
    import zipfile, re, os, base64
    if not drawing_data:
        return None
    tmp = str(xlsx_path) + ".~draw.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            existing = set(zin.namelist())
            extra_files: dict[str, bytes] = {}
            sheet_xml_patches: dict[str, dict] = {}

            for sname, sd in drawing_data.items():
                new_sheet_file = sheet_name_to_new_file.get(sname)
                old_drawing = sd.get("drawing_file")
                drawing_xml = sd.get("drawing_xml")
                if not new_sheet_file or not old_drawing or not drawing_xml:
                    continue

                # Keep original DrawingML part names/relationships. Excel desktop
                # is stricter than XML validators; remapping complex drawing
                # packages can break hidden cross-part references.
                extra_files[old_drawing] = drawing_xml.encode("utf-8")

                old_rels_content = sd.get("drawing_rels")
                if old_rels_content:
                    dr_rels_path = old_drawing.rsplit("/", 1)
                    extra_files[dr_rels_path[0] + "/_rels/" + dr_rels_path[1] + ".rels"] = old_rels_content.encode("utf-8")

                for fp, payload in (sd.get("files") or {}).items():
                    extra_files[fp] = base64.b64decode(payload)

                parts = new_sheet_file.rsplit("/", 1)
                new_sheet_rels = parts[0] + "/_rels/" + parts[1] + ".rels"
                rid = "rId1"
                if new_sheet_rels in existing:
                    rels_now = zin.read(new_sheet_rels).decode("utf-8")
                    used = [int(m) for m in re.findall(r'\bId="rId(\d+)"', rels_now)]
                    rid = f"rId{max(used, default=0) + 1}"

                sheet_xml_patches[new_sheet_file] = {
                    "drawing_rId": rid,
                    "sheet_rels_file": new_sheet_rels,
                    "rel_target": "../drawings/" + old_drawing.rsplit("/", 1)[1],
                }

            if not sheet_xml_patches:
                return None

            ct_additions = set()
            for fp in extra_files:
                ext = fp.rsplit(".", 1)[-1].lower()
                if ext == "xml":
                    if "/charts/" in fp:
                        ct_additions.add(("Override", fp, "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"))
                    elif "/drawings/drawing" in fp:
                        ct_additions.add(("Override", fp, "application/vnd.openxmlformats-officedocument.drawing+xml"))
                elif ext == "png":
                    ct_additions.add(("Default", "png", "image/png"))
                elif ext in {"jpg", "jpeg"}:
                    ct_additions.add(("Default", ext, "image/jpeg"))
                elif ext == "gif":
                    ct_additions.add(("Default", "gif", "image/gif"))
                elif ext == "emf":
                    ct_additions.add(("Default", "emf", "image/x-emf"))
                elif ext == "wmf":
                    ct_additions.add(("Default", "wmf", "image/x-wmf"))

            draw_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
            rels_patches = {
                patch["sheet_rels_file"]: (
                    f'<Relationship Id="{patch["drawing_rId"]}" '
                    f'Type="{draw_type}" Target="{patch["rel_target"]}"/>'
                )
                for patch in sheet_xml_patches.values()
            }

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)

                    if item.filename in sheet_xml_patches:
                        patch = sheet_xml_patches[item.filename]
                        content = raw.decode("utf-8")
                        if "<drawing" not in content:
                            rns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                            tag = f'<drawing xmlns:r="{rns}" r:id="{patch["drawing_rId"]}"/>'
                            anchor = re.search(
                                r"<(?:legacyDrawing|legacyDrawingHF|picture|oleObjects"
                                r"|controls|webPublishItems|tableParts|extLst)\b",
                                content,
                            )
                            pos = anchor.start() if anchor else content.rfind("</worksheet>")
                            content = content[:pos] + tag + content[pos:]
                        raw = content.encode("utf-8")

                    elif item.filename in rels_patches:
                        content = raw.decode("utf-8")
                        if rels_patches[item.filename] not in content:
                            content = content.replace("</Relationships>", rels_patches[item.filename] + "</Relationships>", 1)
                        raw = content.encode("utf-8")

                    elif item.filename == "[Content_Types].xml" and ct_additions:
                        content = raw.decode("utf-8")
                        for kind, part_or_ext, content_type in ct_additions:
                            if kind == "Default":
                                if f'Extension="{part_or_ext}"' not in content:
                                    content = content.replace(
                                        "</Types>",
                                        f'<Default Extension="{part_or_ext}" ContentType="{content_type}"/></Types>', 1)
                            elif f'PartName="/{part_or_ext}"' not in content:
                                content = content.replace(
                                    "</Types>",
                                    f'<Override PartName="/{part_or_ext}" ContentType="{content_type}"/></Types>', 1)
                        raw = content.encode("utf-8")

                    if item.filename in extra_files:
                        raw = extra_files[item.filename]
                    zout.writestr(item, raw)

                for fp, fb in extra_files.items():
                    if fp not in existing:
                        zout.writestr(fp, fb)

                ns = "http://schemas.openxmlformats.org/package/2006/relationships"
                for rels_file, rel_entry in rels_patches.items():
                    if rels_file not in existing:
                        zout.writestr(rels_file, f'<Relationships xmlns="{ns}">{rel_entry}</Relationships>'.encode("utf-8"))

        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"drawings/charts/images passthrough failed: {e}"
    return None

def _drawing_anchor_fragments(xml: str) -> list[str]:
    pattern = re.compile(
        r"<(twoCellAnchor|oneCellAnchor|absoluteAnchor)\b.*?</\1>",
        re.DOTALL,
    )
    return [match.group(0) for match in pattern.finditer(xml or "")]


def _drawing_relationship_records(xml: str | None) -> dict[str, dict]:
    if not xml:
        return {}
    root = ET.fromstring(xml)
    return {
        child.attrib["Id"]: dict(child.attrib)
        for child in root
        if child.tag.rsplit("}", 1)[-1] == "Relationship" and child.attrib.get("Id")
    }


def _drawing_target_part(drawing_part: str, target: str) -> str:
    import posixpath

    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(drawing_part), target))


def _allocate_package_part(preferred: str, used_parts: set[str]) -> str:
    import posixpath

    preferred = preferred.lstrip("/")
    if preferred not in used_parts:
        used_parts.add(preferred)
        return preferred
    folder, filename = posixpath.split(preferred)
    match = re.match(r"^(.*?)(\d+)(\.[^.]+)$", filename)
    if match:
        stem, _, suffix = match.groups()
    else:
        stem, suffix = posixpath.splitext(filename)
    number = 1
    while True:
        candidate = posixpath.join(folder, f"{stem}{number}{suffix}")
        if candidate not in used_parts:
            used_parts.add(candidate)
            return candidate
        number += 1


def _next_relationship_id(used_ids: set[str]) -> str:
    numbers = [int(match.group(1)) for value in used_ids if (match := re.fullmatch(r"rId(\d+)", value))]
    number = max(numbers, default=0) + 1
    while f"rId{number}" in used_ids:
        number += 1
    result = f"rId{number}"
    used_ids.add(result)
    return result


def _append_before_xml_close(xml: str, local_name: str, additions: str) -> str:
    match = re.search(rf"</(?:[A-Za-z_][\w.-]*:)?{re.escape(local_name)}\s*>", xml)
    if not match:
        raise ValueError(f"Invalid XML: missing closing {local_name} element.")
    return xml[:match.start()] + additions + xml[match.start():]


def _set_tag_attribute(tag: str, name: str, value) -> str:
    pattern = re.compile(rf"(?<![\w:]){re.escape(name)}\s*=\s*([\"']).*?\1")
    replacement = f"{name}={quoteattr(str(value))}"
    if pattern.search(tag):
        return pattern.sub(replacement, tag, count=1)
    position = -2 if tag.endswith("/>") else -1
    return tag[:position] + " " + replacement + tag[position:]


def _patch_drawing_object(anchor_xml: str, object_id: int, name: str | None) -> str:
    pattern = re.compile(r"<(?:[A-Za-z_][\w.-]*:)?cNvPr\b[^>]*>")

    def replace(match):
        tag = _set_tag_attribute(match.group(0), "id", object_id)
        return _set_tag_attribute(tag, "name", name) if name is not None else tag

    patched, count = pattern.subn(replace, anchor_xml, count=1)
    if count != 1:
        raise ValueError("Generated drawing anchor has no cNvPr element.")
    anchor_pattern = re.compile(r"^<(?:[A-Za-z_][\w.-]*:)?(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)\b[^>]*>")

    def bind_namespaces(match):
        tag = match.group(0)
        namespaces = {
            "xmlns": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "xmlns:a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "xmlns:c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
            "xmlns:r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for attribute, uri in namespaces.items():
            if not re.search(rf"(?<![\w:]){re.escape(attribute)}\s*=", tag):
                tag = _set_tag_attribute(tag, attribute, uri)
        return tag

    return anchor_pattern.sub(bind_namespaces, patched, count=1)


def _replace_drawing_relationship_id(anchor_xml: str, old_id: str, new_id: str) -> str:
    for attribute in ("id", "embed", "link"):
        pattern = re.compile(
            rf"(\br:{attribute}\s*=\s*)([\"']){re.escape(old_id)}\2"
        )
        anchor_xml = pattern.sub(lambda match: match.group(1) + quoteattr(new_id), anchor_xml)
    return anchor_xml


def _drawing_rgb(value) -> str | None:
    if isinstance(value, dict):
        value = value.get("rgb") if value.get("type", "rgb") == "rgb" else None
    if value is None:
        return None
    text = str(value).strip().lstrip("#")
    if len(text) == 8:
        text = text[-6:]
    return text.upper() if re.fullmatch(r"[0-9A-Fa-f]{6}", text) else None


def _shape_run_xml(run: dict, default_color: str | None) -> str:
    text = str(run.get("text") or "")
    font = run.get("font") or run.get("style") or {}
    attrs = []
    if font.get("bold") is not None:
        attrs.append(f'b="{1 if font["bold"] else 0}"')
    if font.get("italic") is not None:
        attrs.append(f'i="{1 if font["italic"] else 0}"')
    if "strike" in font:
        attrs.append(f'strike="{"sngStrike" if font["strike"] else "noStrike"}"')
    if "underline" in font:
        underline = font.get("underline")
        underline_value = {
            "single": "sng",
            "double": "dbl",
            None: "none",
            False: "none",
        }.get(underline, underline)
        attrs.append(f'u="{escape(str(underline_value))}"')
    if font.get("size") is not None:
        attrs.append(f'sz="{max(1, int(round(float(font["size"]) * 100)))}"')
    properties = []
    color = _drawing_rgb(font.get("color")) or default_color
    if color:
        properties.append(f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>')
    if font.get("name"):
        properties.append(f'<a:latin typeface={quoteattr(str(font["name"]))}/>')
    rpr = f'<a:rPr {" ".join(attrs)}>{"".join(properties)}</a:rPr>' if attrs or properties else "<a:rPr/>"
    preserve = ' xml:space="preserve"' if text[:1].isspace() or text[-1:].isspace() else ""
    return f"<a:r>{rpr}<a:t{preserve}>{escape(text)}</a:t></a:r>"


def build_shape_rich_text_xml(rich_text: dict, default_color: str | None = None) -> str:
    result = []
    runs = rich_text.get("runs") or []
    if not runs and rich_text.get("text") is not None:
        runs = [{"text": str(rich_text.get("text") or "")}]
    for run in runs:
        parts = str(run.get("text") or "").split("\n")
        for index, part in enumerate(parts):
            if part:
                result.append(_shape_run_xml({**run, "text": part}, default_color))
            if index < len(parts) - 1:
                result.append("<a:br/>")
    return "".join(result)


def _shape_anchor_xml(creation: dict, object_id: int) -> str:
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    from openpyxl.utils.units import cm_to_EMU

    anchor = str(creation.get("anchor") or "A1").strip().upper()
    try:
        column_letter, row_number = coordinate_from_string(anchor)
        column_number = column_index_from_string(column_letter)
    except Exception as exc:
        raise ValueError(f"Invalid shape anchor: {anchor!r}") from exc
    width = float(creation.get("width") if creation.get("width") is not None else 3.0)
    height = float(creation.get("height") if creation.get("height") is not None else 1.5)
    if width <= 0 or height <= 0:
        raise ValueError("Shape width and height must be positive.")
    cx, cy = int(cm_to_EMU(width)), int(cm_to_EMU(height))
    name = creation.get("name") or f"Shape {object_id}"
    geometry = str(creation.get("shape_type") or "rect").strip() or "rect"
    if geometry == "rectangle":
        geometry = "rect"

    style = creation.get("style") or {}
    fill = _drawing_rgb(style.get("fill_color"))
    outline = _drawing_rgb(style.get("outline_color"))
    text_color = _drawing_rgb(style.get("text_color"))
    fill_xml = f'<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>' if fill else "<a:noFill/>"
    line_xml = (
        f'<a:ln><a:solidFill><a:srgbClr val="{outline}"/></a:solidFill></a:ln>'
        if outline else "<a:ln><a:noFill/></a:ln>"
    )

    rich_text = copy.deepcopy(creation.get("rich_text") or {})
    if not rich_text.get("runs") and creation.get("text") is not None:
        rich_text["runs"] = [{"text": str(creation.get("text") or "")}]
    paragraph = build_shape_rich_text_xml(rich_text, text_color)
    paragraph += "<a:endParaRPr/>"

    return (
        '<xdr:oneCellAnchor xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"<xdr:from><xdr:col>{column_number - 1}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row_number - 1}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f'<xdr:ext cx="{cx}" cy="{cy}"/>'
        "<xdr:sp><xdr:nvSpPr>"
        f'<xdr:cNvPr id="{object_id}" name={quoteattr(str(name))}/><xdr:cNvSpPr txBox="1"/>'
        "</xdr:nvSpPr><xdr:spPr>"
        f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
        f'{fill_xml}<a:prstGeom prst={quoteattr(geometry)}><a:avLst/></a:prstGeom>{line_xml}'
        f"</xdr:spPr><xdr:txBody><a:bodyPr/><a:lstStyle/><a:p>{paragraph}</a:p></xdr:txBody>"
        "</xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
    )


def _patch_sheet_drawing_reference(xml: str, relationship_id: str) -> str:
    xml = re.sub(r"<(?:[A-Za-z_][\w.-]*:)?drawing\b[^>]*/>", "", xml)
    root_match = re.search(r"<(?:[A-Za-z_][\w.-]*:)?worksheet\b", xml)
    prefix_match = re.match(r"<([A-Za-z_][\w.-]*:)?worksheet", root_match.group(0)) if root_match else None
    prefix = prefix_match.group(1) if prefix_match and prefix_match.group(1) else ""
    drawing = (
        f'<{prefix}drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        f'r:id="{relationship_id}"/>'
    )
    following = re.search(
        r"<(?:[A-Za-z_][\w.-]*:)?(?:legacyDrawing|legacyDrawingHF|picture|oleObjects|controls|webPublishItems|tableParts|extLst)\b",
        xml,
    )
    position = following.start() if following else xml.rfind(f"</{prefix}worksheet>")
    if position < 0:
        raise ValueError("Invalid worksheet XML: missing closing worksheet element.")
    return xml[:position] + drawing + xml[position:]


def _patch_sheet_drawing_relationship(xml: str | None, relationship_id: str, drawing_part: str) -> str:
    namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    if not xml:
        xml = f'<Relationships xmlns="{namespace}"></Relationships>'

    def remove_drawing(match):
        type_match = re.search(r"\bType\s*=\s*([\"'])(.*?)\1", match.group(0))
        return "" if type_match and type_match.group(2).rstrip("/").endswith("/drawing") else match.group(0)

    xml = re.sub(r"<(?:[A-Za-z_][\w.-]*:)?Relationship\b[^>]*/>", remove_drawing, xml)
    entry = (
        f'<Relationship xmlns="{namespace}" Id="{relationship_id}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        f'Target="../drawings/{drawing_part.rsplit("/", 1)[-1]}"/>'
    )
    return _append_before_xml_close(xml, "Relationships", entry)


def _patch_drawing_content_types(xml: str, removed_parts: set[str], final_parts: set[str]) -> str:
    namespace = "http://schemas.openxmlformats.org/package/2006/content-types"

    def remove_override(match):
        part_match = re.search(r"\bPartName\s*=\s*([\"'])(.*?)\1", match.group(0))
        part = part_match.group(2).lstrip("/") if part_match else None
        return "" if part in removed_parts else match.group(0)

    xml = re.sub(r"<(?:[A-Za-z_][\w.-]*:)?Override\b[^>]*/>", remove_override, xml)
    existing_overrides = {
        match.group(2).lstrip("/")
        for match in re.finditer(r"\bPartName\s*=\s*([\"'])(.*?)\1", xml)
    }
    existing_defaults = {
        match.group(2).lower()
        for match in re.finditer(r"\bExtension\s*=\s*([\"'])(.*?)\1", xml)
    }
    additions = []
    media_types = {
        "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "gif": "image/gif",
        "bmp": "image/bmp", "tif": "image/tiff", "tiff": "image/tiff", "svg": "image/svg+xml",
        "emf": "image/x-emf", "wmf": "image/x-wmf",
    }
    for part in sorted(final_parts):
        if part.startswith("xl/drawings/") and part.endswith(".xml") and "/_rels/" not in part:
            content_type = "application/vnd.openxmlformats-officedocument.drawing+xml"
        elif part.startswith("xl/charts/") and part.endswith(".xml"):
            content_type = "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
        else:
            extension = part.rsplit(".", 1)[-1].lower() if "." in part else ""
            content_type = media_types.get(extension)
            if content_type and extension not in existing_defaults:
                additions.append(
                    f'<Default xmlns="{namespace}" Extension="{extension}" ContentType="{content_type}"/>'
                )
                existing_defaults.add(extension)
            continue
        if part not in existing_overrides:
            additions.append(
                f'<Override xmlns="{namespace}" PartName="/{part}" ContentType="{content_type}"/>'
            )
            existing_overrides.add(part)
    return _append_before_xml_close(xml, "Types", "".join(additions)) if additions else xml


def _merge_drawing_packages(
    xlsx_path: str,
    drawing_data: dict,
    creation_plan: dict,
    sheet_name_to_new_file: dict,
) -> str | None:
    if not drawing_data and not creation_plan:
        return None
    tmp = str(xlsx_path) + ".~draw-merge.tmp"
    try:
        generated_data = _extract_drawing_data(xlsx_path, sheet_name_to_new_file)
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            existing_parts = set(zin.namelist())
            generated_parts: set[str] = set()
            for sheet_drawing in generated_data.values():
                drawing_part = sheet_drawing.get("drawing_file")
                if drawing_part:
                    generated_parts.add(drawing_part)
                    generated_parts.add(
                        drawing_part.rsplit("/", 1)[0] + "/_rels/" + drawing_part.rsplit("/", 1)[1] + ".rels"
                    )
                generated_parts.update((sheet_drawing.get("files") or {}).keys())

            preserved_parts: set[str] = set()
            for sheet_drawing in drawing_data.values():
                drawing_part = sheet_drawing.get("drawing_file")
                if drawing_part:
                    preserved_parts.add(drawing_part)
                    if sheet_drawing.get("drawing_rels"):
                        preserved_parts.add(
                            drawing_part.rsplit("/", 1)[0] + "/_rels/" + drawing_part.rsplit("/", 1)[1] + ".rels"
                        )
                preserved_parts.update((sheet_drawing.get("files") or {}).keys())

            used_parts = (existing_parts - generated_parts) | preserved_parts
            final_files: dict[str, bytes] = {}
            sheet_patches: dict[str, bytes] = {}
            sheet_relationship_patches: dict[str, bytes] = {}
            final_payload_parts: set[str] = set()
            affected_sheets = list(dict.fromkeys([*drawing_data.keys(), *creation_plan.keys()]))

            for sheet_name in affected_sheets:
                preserved = drawing_data.get(sheet_name) or {}
                generated = generated_data.get(sheet_name) or {}
                plan = creation_plan.get(sheet_name) or {"charts": [], "images": [], "shapes": []}
                sheet_file = sheet_name_to_new_file.get(sheet_name)
                if not sheet_file or sheet_file not in existing_parts:
                    raise ValueError(f"Cannot resolve worksheet package part for {sheet_name!r}.")

                preferred_drawing = preserved.get("drawing_file") or generated.get("drawing_file") or "xl/drawings/drawing1.xml"
                if preserved.get("drawing_file"):
                    drawing_part = preferred_drawing
                    used_parts.add(drawing_part)
                else:
                    drawing_part = _allocate_package_part(preferred_drawing, used_parts)

                drawing_xml = preserved.get("drawing_xml") or (
                    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"></xdr:wsDr>'
                )
                drawing_rels_xml = preserved.get("drawing_rels")
                for part, payload in (preserved.get("files") or {}).items():
                    final_files[part] = base64.b64decode(payload)
                    final_payload_parts.add(part)

                object_ids = {
                    int(match.group(2))
                    for match in re.finditer(
                        r"<(?:[A-Za-z_][\w.-]*:)?cNvPr\b[^>]*\bid\s*=\s*([\"'])(\d+)\1",
                        drawing_xml,
                    )
                }
                next_object_id = max(object_ids, default=0) + 1
                preserved_relationships = _drawing_relationship_records(drawing_rels_xml)
                used_relationship_ids = set(preserved_relationships)
                generated_relationships = _drawing_relationship_records(generated.get("drawing_rels"))
                relationship_map: dict[str, tuple[str, str]] = {}
                relationship_additions = []
                anchor_additions = []
                creation_indices = {"charts": 0, "images": 0}

                for anchor_xml in _drawing_anchor_fragments(generated.get("drawing_xml") or ""):
                    if re.search(r"<(?:[A-Za-z_][\w.-]*:)?pic\b", anchor_xml):
                        creation_key = "images"
                        id_match = re.search(r"\br:embed\s*=\s*([\"'])(.*?)\1", anchor_xml)
                    elif "/chart" in anchor_xml or re.search(r"<(?:[A-Za-z_][\w.-]*:)?chart\b", anchor_xml):
                        creation_key = "charts"
                        id_match = re.search(r"\br:id\s*=\s*([\"'])(.*?)\1", anchor_xml)
                    else:
                        raise ValueError("Openpyxl generated an unsupported drawing anchor.")
                    index = creation_indices[creation_key]
                    creations = plan.get(creation_key) or []
                    if index >= len(creations):
                        raise ValueError(f"Generated {creation_key[:-1]} count does not match queued creations.")
                    creation = creations[index]
                    creation_indices[creation_key] += 1
                    if not id_match:
                        raise ValueError("Generated drawing anchor has no relationship id.")
                    old_relationship_id = id_match.group(2)

                    if old_relationship_id not in relationship_map:
                        relationship = generated_relationships.get(old_relationship_id)
                        if not relationship:
                            raise ValueError(f"Missing generated drawing relationship {old_relationship_id}.")
                        old_target = _drawing_target_part(generated.get("drawing_file"), relationship["Target"])
                        payload_b64 = (generated.get("files") or {}).get(old_target)
                        if payload_b64 is None:
                            raise ValueError(f"Missing generated drawing payload {old_target}.")
                        payload = creation.get("_raw_bytes") if creation_key == "images" else base64.b64decode(payload_b64)
                        new_target = _allocate_package_part(old_target, used_parts)
                        new_relationship_id = _next_relationship_id(used_relationship_ids)
                        relationship_map[old_relationship_id] = (new_relationship_id, new_target)
                        relationship_entry = (
                            '<Relationship xmlns="http://schemas.openxmlformats.org/package/2006/relationships" '
                            f'Id="{new_relationship_id}" Type={quoteattr(relationship["Type"])} '
                            f'Target="/{new_target}"'
                            + (
                                f' TargetMode={quoteattr(relationship["TargetMode"])}'
                                if relationship.get("TargetMode") else ""
                            )
                            + "/>"
                        )
                        relationship_additions.append(relationship_entry)
                        final_files[new_target] = payload
                        final_payload_parts.add(new_target)
                    new_relationship_id, _ = relationship_map[old_relationship_id]
                    anchor_xml = _replace_drawing_relationship_id(
                        anchor_xml, old_relationship_id, new_relationship_id
                    )
                    requested_name = creation.get("name") if creation_key == "images" else None
                    anchor_additions.append(
                        _patch_drawing_object(anchor_xml, next_object_id, requested_name)
                    )
                    next_object_id += 1

                for creation_key in ("charts", "images"):
                    if creation_indices[creation_key] != len(plan.get(creation_key) or []):
                        raise ValueError(f"Queued {creation_key} were not fully materialized by openpyxl.")
                for shape in plan.get("shapes") or []:
                    anchor_additions.append(_shape_anchor_xml(shape, next_object_id))
                    next_object_id += 1

                if anchor_additions:
                    drawing_xml = _append_before_xml_close(drawing_xml, "wsDr", "".join(anchor_additions))
                if relationship_additions:
                    if not drawing_rels_xml:
                        drawing_rels_xml = (
                            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                            "</Relationships>"
                        )
                    drawing_rels_xml = _append_before_xml_close(
                        drawing_rels_xml, "Relationships", "".join(relationship_additions)
                    )

                final_files[drawing_part] = drawing_xml.encode("utf-8")
                final_payload_parts.add(drawing_part)
                drawing_rels_part = (
                    drawing_part.rsplit("/", 1)[0] + "/_rels/" + drawing_part.rsplit("/", 1)[1] + ".rels"
                )
                if drawing_rels_xml:
                    final_files[drawing_rels_part] = drawing_rels_xml.encode("utf-8")

                sheet_rels_part = (
                    sheet_file.rsplit("/", 1)[0] + "/_rels/" + sheet_file.rsplit("/", 1)[1] + ".rels"
                )
                current_sheet_rels = (
                    zin.read(sheet_rels_part).decode("utf-8") if sheet_rels_part in existing_parts else None
                )
                remaining_rels = re.sub(
                    r"<(?:[A-Za-z_][\w.-]*:)?Relationship\b[^>]*/>",
                    lambda match: "" if re.search(
                        r"\bType\s*=\s*([\"']).*?/drawing/?\1", match.group(0)
                    ) else match.group(0),
                    current_sheet_rels or "",
                )
                sheet_relationship_id = _next_relationship_id(
                    set(re.findall(r"\bId\s*=\s*[\"']([^\"']+)[\"']", remaining_rels))
                )
                sheet_patches[sheet_file] = _patch_sheet_drawing_reference(
                    zin.read(sheet_file).decode("utf-8"), sheet_relationship_id
                ).encode("utf-8")
                sheet_relationship_patches[sheet_rels_part] = _patch_sheet_drawing_relationship(
                    current_sheet_rels, sheet_relationship_id, drawing_part
                ).encode("utf-8")

            removed_content_type_parts = {
                part for part in generated_parts
                if part.startswith("xl/drawings/") or part.startswith("xl/charts/")
            }
            content_types = _patch_drawing_content_types(
                zin.read("[Content_Types].xml").decode("utf-8"),
                removed_content_type_parts,
                final_payload_parts,
            ).encode("utf-8")

            replacements = {**final_files, **sheet_patches, **sheet_relationship_patches}
            replacements["[Content_Types].xml"] = content_types
            skipped = generated_parts | set(replacements)
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename in skipped:
                        continue
                    zout.writestr(item, zin.read(item.filename))
                for part, payload in replacements.items():
                    zout.writestr(part, payload)
        os.replace(tmp, str(xlsx_path))
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"drawing creation merge failed: {exc}"
    return None


def _extract_cf_xml(xlsx_path, sheet_names: list) -> dict:
    """
    Extract raw <conditionalFormatting> XML blocks per sheet, plus the workbook's
    <dxfs> section (differential styles referenced by CF rules).
    Returns {sname: [block, ...], "__dxfs__": "<dxfs>...</dxfs>"}.
    """
    import zipfile, re
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            wb_xml   = zf.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            for sname in sheet_names:
                fp = sheet_file_map.get(sname)
                if not fp or fp not in zf.namelist():
                    continue
                content = zf.read(fp).decode("utf-8")
                blocks = re.findall(
                    r"<conditionalFormatting(?:\s[^>]*)?>.*?</conditionalFormatting>",
                    content, re.DOTALL)
                if blocks:
                    result[sname] = blocks
            # Extract dxfs section from styles.xml (needed for dxfId refs in CF rules)
            if result and "xl/styles.xml" in zf.namelist():
                styles_xml = zf.read("xl/styles.xml").decode("utf-8")
                dxfs_m = re.search(r"<dxfs\b[^>]*>.*?</dxfs>", styles_xml, re.DOTALL)
                if dxfs_m:
                    result["__dxfs__"] = dxfs_m.group(0)
    except Exception:
        pass
    return result


def _extract_data_validations_xml(xlsx_path, sheet_file_map: dict) -> dict:
    """Extract raw <dataValidations> XML per sheet."""
    import zipfile, re
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            for sname, sheet_file in sheet_file_map.items():
                if sheet_file not in zf.namelist():
                    continue
                content = zf.read(sheet_file).decode("utf-8")
                m = re.search(
                    r"<dataValidations\b[^>]*>.*?</dataValidations>",
                    content,
                    re.DOTALL,
                )
                if m:
                    result[sname] = m.group(0)
    except Exception:
        pass
    return result


def _extract_worksheet_xml_blocks(xlsx_path, sheet_file_map: dict, tag_name: str) -> dict:
    result = {}
    pattern = re.compile(
        rf"<(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?{re.escape(tag_name)}>)",
        re.DOTALL,
    )
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as archive:
            names = set(archive.namelist())
            for sheet_name, part_name in sheet_file_map.items():
                if part_name not in names:
                    continue
                match = pattern.search(archive.read(part_name).decode("utf-8"))
                if match:
                    result[sheet_name] = match.group(0)
    except Exception:
        pass
    return result


def _resolve_relationship_target(source_part: str, target: str) -> str:
    import posixpath
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


_REGENERATED_SHEET_RELATIONSHIP_SUFFIXES = (
    "/comments",
    "/drawing",
    "/hyperlink",
    "/printerSettings",
    "/table",
    "/vmlDrawing",
)


def _extract_sheet_passthrough_relationships(xlsx_path, sheet_file_map: dict) -> dict:
    """Capture worksheet relationships and target parts not rebuilt by openpyxl."""
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as archive:
            names = set(archive.namelist())
            content_types = (
                _content_type_inventory(archive.read("[Content_Types].xml"))
                if "[Content_Types].xml" in names
                else {"defaults": {}, "overrides": {}}
            )
            for sheet_name, sheet_part in sheet_file_map.items():
                rels_part = _relationship_part_for_source(sheet_part)
                if rels_part not in names:
                    continue
                relationships = []
                parts = {}
                rels_root = ET.fromstring(archive.read(rels_part))
                for node in rels_root:
                    relationship_type = str(node.get("Type") or "")
                    normalized_type = relationship_type.rstrip("/")
                    if any(
                        normalized_type.endswith(suffix)
                        for suffix in _REGENERATED_SHEET_RELATIONSHIP_SUFFIXES
                    ):
                        continue
                    target = str(node.get("Target") or "")
                    record = {
                        "Id": node.get("Id"),
                        "Type": relationship_type,
                        "Target": target,
                    }
                    target_mode = node.get("TargetMode")
                    if target_mode is not None:
                        record["TargetMode"] = target_mode
                    if target and target_mode != "External":
                        target_part = _resolve_relationship_target(sheet_part, target)
                        record["target_part"] = target_part
                        if target_part in names and target_part not in parts:
                            content_type = _content_type_for_part(target_part, content_types)
                            extension = target_part.rsplit(".", 1)[-1].lower() if "." in target_part else None
                            part_record = {
                                "data": base64.b64encode(archive.read(target_part)).decode("ascii"),
                                "content_type": content_type,
                                "content_type_source": (
                                    "override"
                                    if target_part in (content_types.get("overrides") or {})
                                    else "default"
                                ),
                                "extension": extension,
                            }
                            target_rels_part = _relationship_part_for_source(target_part)
                            if target_rels_part in names:
                                part_record["relationships_xml"] = base64.b64encode(
                                    archive.read(target_rels_part)
                                ).decode("ascii")
                            parts[target_part] = part_record
                    relationships.append(record)
                if relationships:
                    result[sheet_name] = {
                        "relationships": relationships,
                        "parts": parts,
                    }
    except Exception:
        pass
    return result


def _comment_vml_part(entries: dict[str, bytes], sheet_part: str) -> str | None:
    rels_part = _relationship_part_for_source(sheet_part)
    raw = entries.get(rels_part)
    if raw is None:
        return None
    root = ET.fromstring(raw)
    for relationship in root:
        if relationship.get("TargetMode") == "External":
            continue
        if str(relationship.get("Type", "")).endswith("/vmlDrawing"):
            return _resolve_relationship_target(sheet_part, relationship.get("Target", ""))
    return None


def _extract_comment_vml(xlsx_path, sheet_file_map: dict) -> dict:
    result = {}
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as archive:
            entries = {item.filename: archive.read(item.filename) for item in archive.infolist()}
        for sheet_name, sheet_part in sheet_file_map.items():
            vml_part = _comment_vml_part(entries, sheet_part)
            if vml_part and vml_part in entries:
                result[sheet_name] = {
                    "part_name": vml_part,
                    "xml": entries[vml_part].decode("utf-8"),
                }
    except Exception:
        pass
    return result


def _parse_ignored_errors_xml(xml: str | None) -> list[dict]:
    rules = []
    for match in re.finditer(r"<(?:[A-Za-z_][\w.-]*:)?ignoredError\b([^>]*)/?>", xml or ""):
        attrs = _parse_xml_attrs(match.group(1))
        rule = {"sqref": attrs.pop("sqref", "")}
        for key, value in attrs.items():
            rule[key] = str(value).lower() in {"1", "true"}
        if rule["sqref"]:
            rules.append(rule)
    return rules


def _inject_cf_xml(xlsx_path: str, sheet_cf: dict) -> str | None:
    """
    Patch a saved xlsx file by:
    1. Injecting stored CF XML into each sheet's XML.
    2. Replacing the <dxfs/> section in styles.xml with the original one (needed for dxfId refs).
    """
    import zipfile, re, os
    if not sheet_cf:
        return
    dxfs_xml = sheet_cf.pop("__dxfs__", None)
    sheet_cf_only = {k: v for k, v in sheet_cf.items() if v}
    if not sheet_cf_only and not dxfs_xml:
        return
    tmp = str(xlsx_path) + ".~tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            wb_xml   = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            to_patch = {sheet_file_map[sn]: blocks
                        for sn, blocks in sheet_cf_only.items()
                        if sn in sheet_file_map}
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    if item.filename in to_patch:
                        content = raw.decode("utf-8")
                        injection = "".join(to_patch[item.filename])
                        # CT_Worksheet order: conditionalFormatting must precede
                        # dataValidations/hyperlinks/pageMargins/… — insert before
                        # the first such element, not at the end of the sheet.
                        anchor = re.search(
                            r"<(?:dataValidations|hyperlinks|printOptions|pageMargins"
                            r"|pageSetup\b|headerFooter|rowBreaks|colBreaks|drawing\b"
                            r"|legacyDrawing|tableParts|extLst)\b",
                            content,
                        )
                        pos = anchor.start() if anchor else content.rfind("</worksheet>")
                        content = content[:pos] + injection + content[pos:]
                        raw = content.encode("utf-8")
                    elif item.filename == "xl/styles.xml" and dxfs_xml:
                        content = raw.decode("utf-8")
                        # Replace existing <dxfs> block, or inject at its schema
                        # position: dxfs must precede tableStyles/colors/extLst
                        # (CT_Stylesheet order) — Excel refuses the file otherwise.
                        if re.search(r"<dxfs\b", content):
                            content = re.sub(
                                r"<dxfs\b[^>]*/?>|<dxfs\b[^>]*>.*?</dxfs>",
                                dxfs_xml, content, count=1, flags=re.DOTALL)
                        else:
                            anchor = re.search(r"<(?:tableStyles|colors|extLst)\b", content)
                            pos = anchor.start() if anchor else content.rfind("</styleSheet>")
                            content = content[:pos] + dxfs_xml + content[pos:]
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"conditional formatting passthrough failed: {e}"
    return None


def _inject_data_validations_xml(xlsx_path: str, sheet_validations: dict) -> str | None:
    """Patch saved worksheet XML with original dataValidations XML."""
    import zipfile, re, os
    if not sheet_validations:
        return
    tmp = str(xlsx_path) + ".~validations.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            file_to_xml = {
                sheet_file_map[sname]: xml
                for sname, xml in sheet_validations.items()
                if sname in sheet_file_map
            }
            if not file_to_xml:
                return

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    dv_xml = file_to_xml.get(item.filename)
                    if dv_xml:
                        content = raw.decode("utf-8")
                        if re.search(r"<dataValidations\b", content):
                            content = re.sub(
                                r"<dataValidations\b[^>]*>.*?</dataValidations>",
                                dv_xml,
                                content,
                                count=1,
                                flags=re.DOTALL,
                            )
                        else:
                            content = re.sub(
                                r"(<hyperlinks\b|<pageMargins\b|</worksheet>)",
                                dv_xml + r"\1",
                                content,
                                count=1,
                            )
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"data validations passthrough failed: {e}"
    return None


def _inject_raw_fills(xlsx_path: str, data: dict) -> str | None:
    """Patch saved styles.xml fill entries back to original raw OOXML fills."""
    import zipfile, re, os
    from openpyxl.utils import get_column_letter

    sheet_targets = {}
    for sd in data.get("sheets", []):
        targets = {}
        for r_idx, row_data in enumerate(sd.get("rows", []), 1):
            for c_idx, cd in enumerate(row_data.get("cells", []), 1):
                if cd.get("merge") == "slave":
                    continue
                raw = _usable_raw_fill(cd)
                if raw and raw.get("xml"):
                    targets[f"{get_column_letter(c_idx)}{r_idx}"] = raw["xml"]
        if targets:
            sheet_targets[sd["name"]] = targets

    if not sheet_targets:
        return

    tmp = str(xlsx_path) + ".~fills.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            namelist = set(zin.namelist())
            if "xl/styles.xml" not in namelist:
                return
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)

            styles_xml = zin.read("xl/styles.xml").decode("utf-8")
            fills_m = re.search(r"<fills\b[^>]*>.*?</fills>", styles_xml, re.DOTALL)
            xfs_m = re.search(r"<cellXfs\b[^>]*>.*?</cellXfs>", styles_xml, re.DOTALL)
            if not fills_m or not xfs_m:
                return
            fills = _extract_xml_children(fills_m.group(0), "fill")
            style_to_fill: dict[int, int] = {}
            for idx, xf_m in enumerate(re.finditer(r"<xf\b([^>]*)/?>", xfs_m.group(0))):
                attrs = _parse_xml_attrs(xf_m.group(1))
                if attrs.get("fillId") is not None:
                    style_to_fill[idx] = int(attrs["fillId"])

            fill_patches: dict[int, str] = {}
            conflicts: set[int] = set()
            for sname, targets in sheet_targets.items():
                sheet_file = sheet_file_map.get(sname)
                if not sheet_file or sheet_file not in namelist:
                    continue
                sheet_xml = zin.read(sheet_file).decode("utf-8")
                coord_to_style = {}
                for cell_m in re.finditer(r"<c\b([^>]*)", sheet_xml):
                    attrs = _parse_xml_attrs(cell_m.group(1))
                    if attrs.get("r") and attrs.get("s") is not None:
                        coord_to_style[attrs["r"]] = int(attrs["s"])
                for coord, raw_xml in targets.items():
                    style_idx = coord_to_style.get(coord)
                    if style_idx is None:
                        continue
                    fill_id = style_to_fill.get(style_idx)
                    if fill_id is None or not (0 <= fill_id < len(fills)):
                        continue
                    existing = fill_patches.get(fill_id)
                    if existing is not None and existing != raw_xml:
                        conflicts.add(fill_id)
                        continue
                    fill_patches[fill_id] = raw_xml

            for fill_id in conflicts:
                fill_patches.pop(fill_id, None)
            if not fill_patches:
                return

            patched_fills = [
                fill_patches.get(idx, fill_xml)
                for idx, fill_xml in enumerate(fills)
            ]

            def _replace_fills(match):
                open_tag = re.match(r"<fills\b[^>]*>", match.group(0)).group(0)
                return open_tag + "".join(patched_fills) + "</fills>"

            patched_styles = re.sub(
                r"<fills\b[^>]*>.*?</fills>",
                _replace_fills,
                styles_xml,
                count=1,
                flags=re.DOTALL,
            )

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    if item.filename == "xl/styles.xml":
                        raw = patched_styles.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"raw fills passthrough failed: {e}"
    return None


def _inject_sheet_view_attrs(xlsx_path: str, data: dict) -> str | None:
    """Patch saved worksheet XML with original sheetView attributes."""
    import zipfile, re, os

    sheet_attrs = {
        sd["name"]: (sd.get("sheet_view") or {}).get("_raw_attrs")
        for sd in data.get("sheets", [])
        if (sd.get("sheet_view") or {}).get("_raw_attrs") is not None
    }
    if not sheet_attrs:
        return

    tmp = str(xlsx_path) + ".~sheetviews.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            file_to_attrs = {
                sheet_file_map[sname]: attrs
                for sname, attrs in sheet_attrs.items()
                if sname in sheet_file_map
            }
            if not file_to_attrs:
                return

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    attrs = file_to_attrs.get(item.filename)
                    if attrs is not None:
                        content = raw.decode("utf-8")

                        def _replace(match):
                            current_attrs = match.group(1).rstrip()
                            self_closing = current_attrs.endswith("/")
                            slash = "/" if self_closing else ""
                            sep = " " if attrs else ""
                            return f"<sheetView{sep}{attrs}{slash}>"

                        content = re.sub(
                            r"<sheetView\b([^>]*)>",
                            _replace,
                            content,
                            count=1,
                        )
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"sheetView attrs passthrough failed: {e}"
    return None


def _inject_workbook_pr_extra(xlsx_path: str, extra_attrs: dict) -> str | None:
    """Patch xl/workbook.xml's workbookPr with attributes openpyxl's own
    object model has no read/write hook for at all (filterPrivacy,
    saveExternalLinkValues, showObjects, updateLinks, ...). codeName and
    date1904 are handled natively via wb.code_name/wb.epoch before save and
    are not part of extra_attrs."""
    if not extra_attrs:
        return None
    tmp = str(xlsx_path) + ".~workbookpr.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    if item.filename == "xl/workbook.xml":
                        content = raw.decode("utf-8")

                        def _replace(match):
                            attrs = _parse_xml_attrs(match.group(1))
                            for key, value in extra_attrs.items():
                                if value is None:
                                    attrs.pop(key, None)
                                elif isinstance(value, bool):
                                    attrs[key] = "1" if value else "0"
                                else:
                                    attrs[key] = str(value)
                            return f"<workbookPr{_xml_attributes(attrs)}/>"

                        content = re.sub(r"<workbookPr\b([^>]*)/>", _replace, content, count=1)
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"workbookPr extra-attribute passthrough failed: {e}"
    return None


def _inject_doc_core_modified(xlsx_path: str, iso_value: str) -> str | None:
    """Patch docProps/core.xml's dcterms:modified to an explicit/preserved
    timestamp -- openpyxl's own save_workbook() unconditionally stamps this
    element with datetime.now() as the very last step of every save, so any
    value set on wb.properties.modified beforehand is always overwritten and
    must be fixed up afterward for the preserve/set_explicit policies."""
    from datetime import datetime as _dt
    try:
        value = _dt.fromisoformat(iso_value)
    except Exception as e:
        return f"Invalid document 'modified' timestamp {iso_value!r}: {e}"
    text = value.replace(tzinfo=None).isoformat(timespec="seconds") + "Z"
    tmp = str(xlsx_path) + ".~modified.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    if item.filename == "docProps/core.xml":
                        content = raw.decode("utf-8")
                        content = re.sub(
                            r"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                            lambda m: m.group(1) + text + m.group(2),
                            content, count=1,
                        )
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"document 'modified' policy passthrough failed: {e}"
    return None


def _inject_app_props(xlsx_path: str, app_props: dict) -> str | None:
    """Patch docProps/app.xml with extended (app) properties -- openpyxl's
    writer always emits a brand new, empty ExtendedProperties document with
    no hook to customize it, so patch the saved part directly."""
    if not app_props:
        return None
    tmp = str(xlsx_path) + ".~appprops.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    if item.filename == "docProps/app.xml":
                        content = raw.decode("utf-8")
                        for key, value in app_props.items():
                            if value is None:
                                content = re.sub(rf"<{key}\b[^>]*>.*?</{key}>", "", content, flags=re.DOTALL)
                                continue
                            escaped = escape(str(value))
                            new_child = f"<{key}>{escaped}</{key}>"
                            new_content, count = re.subn(
                                rf"<{key}\b[^>]*>.*?</{key}>", new_child, content,
                                count=1, flags=re.DOTALL,
                            )
                            content = new_content if count else content.replace(
                                "</Properties>", new_child + "</Properties>")
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"app properties passthrough failed: {e}"
    return None


def _inject_raw_cols(xlsx_path: str, data: dict) -> str | None:
    """Patch saved worksheet XML with original <cols> ranges when dimensions are unchanged."""
    import zipfile, re, os

    sheet_cols = {}
    for sd in data.get("sheets", []):
        raw = sd.get("_cols_raw") or {}
        raw_xml = raw.get("xml")
        if not raw_xml:
            continue
        current_state = _dimension_state(sd.get("cw"), sd.get("ch"), sd.get("co"))
        original_state = _normalize_dimension_state(raw.get("state"))
        if current_state == original_state:
            sheet_cols[sd["name"]] = raw_xml

    if not sheet_cols:
        return

    tmp = str(xlsx_path) + ".~cols.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            file_to_cols = {
                sheet_file_map[sname]: cols_xml
                for sname, cols_xml in sheet_cols.items()
                if sname in sheet_file_map
            }
            if not file_to_cols:
                return

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    cols_xml = file_to_cols.get(item.filename)
                    if cols_xml is not None:
                        content = raw.decode("utf-8")
                        if re.search(r"<cols\b[^>]*>.*?</cols>", content, re.DOTALL):
                            content = re.sub(
                                r"<cols\b[^>]*>.*?</cols>",
                                cols_xml,
                                content,
                                count=1,
                                flags=re.DOTALL,
                            )
                        else:
                            content = re.sub(
                                r"(<sheetData\b)",
                                cols_xml + r"\1",
                                content,
                                count=1,
                            )
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"cols passthrough failed: {e}"
    return None


def _xml_prefixed_attrs(xml: str) -> set[str]:
    import re
    return set(re.findall(r"\b([A-Za-z_][\w.-]*):[A-Za-z_][\w.-]*=", xml))


def _inject_missing_root_attrs(content: str, needed_attrs: dict[str, str]) -> str:
    import re

    root_m = re.search(r"<worksheet\b([^>]*)>", content)
    if not root_m:
        return content
    current = root_m.group(1)
    additions = []
    for key, value in needed_attrs.items():
        if re.search(rf"\b{re.escape(key)}=", current):
            continue
        additions.append(f'{key}="{value}"')
    if not additions:
        return content
    insert = " " + " ".join(additions)
    return content[:root_m.end() - 1] + insert + content[root_m.end() - 1:]


def _inject_sheet_format_pr(xlsx_path: str, data: dict) -> str | None:
    """Restore raw sheetFormatPr XML, including extension attrs like x14ac:dyDescent."""
    import zipfile, re, os

    sheet_data = {
        sd["name"]: sd.get("_sheet_format_pr")
        for sd in data.get("sheets", [])
        if sd.get("_sheet_format_pr")
    }
    if not sheet_data:
        return

    tmp = str(xlsx_path) + ".~sheetformat.tmp"
    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zin:
            wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
            rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_file_map = _xlsx_sheet_file_map(wb_xml, rels_xml)
            file_to_data = {
                sheet_file_map[sname]: sf_data
                for sname, sf_data in sheet_data.items()
                if sname in sheet_file_map
            }
            if not file_to_data:
                return

            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    raw = zin.read(item.filename)
                    sf_data = file_to_data.get(item.filename)
                    if sf_data:
                        content = raw.decode("utf-8")
                        raw_sf = sf_data.get("sheetFormatPr")
                        root_attrs = _parse_xml_attrs(sf_data.get("root_attrs") or "")
                        needed = {}
                        root_prefixes = {
                            p for p in _xml_prefixed_attrs(sf_data.get("root_attrs") or "")
                            if p not in {"xmlns", "mc"}
                        }
                        for prefix in root_prefixes:
                            ns_key = f"xmlns:{prefix}"
                            if root_attrs.get(ns_key):
                                needed[ns_key] = root_attrs[ns_key]
                            attr_key = next((k for k in root_attrs if k.startswith(prefix + ":")), None)
                            if attr_key:
                                needed[attr_key] = root_attrs[attr_key]
                        if raw_sf:
                            prefixes = _xml_prefixed_attrs(raw_sf)
                            for prefix in prefixes:
                                ns_key = f"xmlns:{prefix}"
                                if root_attrs.get(ns_key):
                                    needed[ns_key] = root_attrs[ns_key]
                            ignorable_prefixes = prefixes | root_prefixes
                            if ignorable_prefixes and root_attrs.get("xmlns:mc") and root_attrs.get("mc:Ignorable"):
                                # mc:Ignorable may only list prefixes that are
                                # actually declared in the new root — Excel
                                # refuses to open the file otherwise.
                                current_m = re.search(r"<worksheet\b([^>]*)>", content)
                                current_attrs = _parse_xml_attrs(current_m.group(1)) if current_m else {}
                                declared = {k[6:] for k in needed if k.startswith("xmlns:")}
                                declared |= {k[6:] for k in current_attrs if k.startswith("xmlns:")}
                                keep = [t for t in root_attrs["mc:Ignorable"].split()
                                        if t in declared]
                                if keep:
                                    needed["xmlns:mc"] = root_attrs["xmlns:mc"]
                                    needed["mc:Ignorable"] = " ".join(keep)
                            if re.search(r"<sheetFormatPr\b", content):
                                content = re.sub(
                                    r"<sheetFormatPr\b[^>]*/>|<sheetFormatPr\b[^>]*>.*?</sheetFormatPr>",
                                    raw_sf,
                                    content,
                                    count=1,
                                    flags=re.DOTALL,
                                )
                            else:
                                content = re.sub(
                                    r"(<sheetData\b)",
                                    raw_sf + r"\1",
                                    content,
                                    count=1,
                                )
                        if needed:
                            content = _inject_missing_root_attrs(content, needed)
                        raw = content.encode("utf-8")
                    zout.writestr(item, raw)
        os.replace(tmp, str(xlsx_path))
    except Exception as e:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"sheetFormatPr passthrough failed: {e}"
    return None


# ── Lossless package/session helpers ──────────────────────────────────────────

_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def _sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _semantic_digest(value) -> str:
    """Stable digest for the editable session model, excluding runtime markers."""
    def clean(item):
        if isinstance(item, dict):
            return {
                str(key): clean(child)
                for key, child in sorted(item.items(), key=lambda pair: str(pair[0]))
                if key not in {
                    "source", "_lossless", "_dirty", "_package_edits",
                    "_sheet_filter", "_loaded_disk_names", "_default_output_path",
                    "_new_workbook", "_load_metrics", "_dirty_features", "_dirty_paths",
                    "_verification_baseline_path",
                }
                and not str(key).startswith("_baseline")
            }
        if isinstance(item, (list, tuple)):
            return [clean(child) for child in item]
        if isinstance(item, bytes):
            return {"__bytes_sha256__": hashlib.sha256(item).hexdigest()}
        return item

    payload = json.dumps(clean(value), sort_keys=True, separators=(",", ":"), default=str, ensure_ascii=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _content_type_inventory(raw: bytes) -> dict:
    root = ET.fromstring(raw)
    defaults = {}
    overrides = {}
    for child in root:
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "Default":
            defaults[child.get("Extension", "")] = child.get("ContentType", "")
        elif tag == "Override":
            overrides[child.get("PartName", "")] = child.get("ContentType", "")
    return {"defaults": defaults, "overrides": overrides}


def _content_type_for_part(part_name: str, inventory: dict) -> str | None:
    override = inventory.get("overrides", {}).get("/" + part_name.lstrip("/"))
    if override:
        return override
    suffix = Path(part_name).suffix.lstrip(".")
    return inventory.get("defaults", {}).get(suffix)


def _relationship_inventory(raw: bytes) -> list[dict]:
    root = ET.fromstring(raw)
    return [dict(child.attrib) for child in root if child.tag.rsplit("}", 1)[-1] == "Relationship"]


def _extract_package_graph(path: str | Path) -> dict:
    """Capture the complete OOXML part/content-type/relationship graph."""
    with zipfile.ZipFile(path, "r") as archive:
        names = archive.namelist()
        content_types = (
            _content_type_inventory(archive.read("[Content_Types].xml"))
            if "[Content_Types].xml" in names
            else {"defaults": {}, "overrides": {}}
        )
        parts = {}
        relationships = {}
        for name in names:
            raw = archive.read(name)
            parts[name] = {
                "size": len(raw),
                "sha256": hashlib.sha256(raw).hexdigest(),
                "content_type": _content_type_for_part(name, content_types),
            }
            if name.endswith(".rels"):
                try:
                    relationships[name] = _relationship_inventory(raw)
                except Exception:
                    relationships[name] = []
    return {
        "parts": parts,
        "content_types": content_types,
        "relationships": relationships,
    }


def _normalise_package_part(name: str) -> str:
    normalised = str(name).replace("\\", "/").lstrip("/")
    if not normalised or normalised.startswith("../") or "/../" in normalised:
        raise ValueError(f"Invalid OOXML package part name: {name!r}")
    return normalised


def _package_edit_bytes(value) -> bytes:
    if isinstance(value, bytes):
        return value
    if isinstance(value, bytearray):
        return bytes(value)
    if isinstance(value, str):
        return value.encode("utf-8")
    if not isinstance(value, dict):
        raise ValueError("Package upsert values must be bytes, text, or an encoding object.")
    if value.get("data_base64") is not None:
        return base64.b64decode(value["data_base64"])
    if value.get("base64") is not None:
        return base64.b64decode(value["base64"])
    if value.get("bytes_base64") is not None:
        return base64.b64decode(value["bytes_base64"])
    if value.get("text") is not None:
        return str(value["text"]).encode(value.get("encoding") or "utf-8")
    if value.get("xml") is not None:
        return str(value["xml"]).encode(value.get("encoding") or "utf-8")
    if value.get("content") is not None:
        content = value["content"]
        return content if isinstance(content, bytes) else str(content).encode(value.get("encoding") or "utf-8")
    raise ValueError("Package upsert object requires base64, bytes_base64, text, xml, or content.")


def _relationship_part_for_source(source: str | None) -> str:
    if not source or str(source).strip() == "/":
        return "_rels/.rels"
    source = _normalise_package_part(source)
    parent, filename = source.rsplit("/", 1) if "/" in source else ("", source)
    prefix = f"{parent}/" if parent else ""
    return f"{prefix}_rels/{filename}.rels"


def _iter_edit_records(value, default_key: str) -> list[dict]:
    if not value:
        return []
    if isinstance(value, list):
        return [dict(record) for record in value]
    if isinstance(value, dict):
        records = []
        for key, item in value.items():
            if isinstance(item, list):
                for record in item:
                    records.append({default_key: key, **dict(record)})
            elif isinstance(item, dict):
                records.append({default_key: key, **item})
            else:
                records.append({default_key: key, "value": item})
        return records
    raise ValueError(f"Expected list or object for package {default_key} edits.")


def _apply_relationship_edits(entries: dict[str, bytes], edits) -> None:
    for record in _iter_edit_records(edits, "source"):
        rels_part = record.get("rels_part") or record.get("relationship_part")
        if rels_part:
            rels_part = _normalise_package_part(rels_part)
        else:
            rels_part = _relationship_part_for_source(record.get("source"))
        raw = entries.get(rels_part)
        root = ET.fromstring(raw) if raw else ET.Element(f"{{{_PACKAGE_REL_NS}}}Relationships")
        rel_id = record.get("Id") or record.get("id")
        op = str(record.get("op") or record.get("action") or "upsert").lower()
        existing = next((node for node in root if node.get("Id") == rel_id), None) if rel_id else None
        if op in {"delete", "remove"}:
            if existing is not None:
                root.remove(existing)
        else:
            if not rel_id:
                raise ValueError("Relationship upsert requires id/Id.")
            node = (
                existing
                if existing is not None
                else ET.SubElement(root, f"{{{_PACKAGE_REL_NS}}}Relationship")
            )
            node.set("Id", str(rel_id))
            rel_type = record.get("Type") or record.get("type")
            target = record.get("Target") or record.get("target")
            if not rel_type or target is None:
                raise ValueError("Relationship upsert requires type/Type and target/Target.")
            node.set("Type", str(rel_type))
            node.set("Target", str(target))
            target_mode = record.get("TargetMode") or record.get("target_mode")
            if target_mode is not None:
                node.set("TargetMode", str(target_mode))
            elif "TargetMode" in node.attrib:
                del node.attrib["TargetMode"]
        entries[rels_part] = ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _apply_content_type_edits(entries: dict[str, bytes], edits) -> None:
    if not edits:
        return
    raw = entries.get("[Content_Types].xml")
    root = ET.fromstring(raw) if raw else ET.Element(f"{{{_CONTENT_TYPES_NS}}}Types")
    if isinstance(edits, dict) and set(edits) <= {"defaults", "overrides"}:
        records = [
            {
                "extension": extension,
                "content_type": content_type,
                "op": "delete" if content_type is None else "upsert",
            }
            for extension, content_type in (edits.get("defaults") or {}).items()
        ]
        records.extend(
            {
                "part_name": part_name,
                "content_type": content_type,
                "op": "delete" if content_type is None else "upsert",
            }
            for part_name, content_type in (edits.get("overrides") or {}).items()
        )
    else:
        records = _iter_edit_records(edits, "part_name")
    for record in records:
        op = str(record.get("op") or record.get("action") or "upsert").lower()
        extension = record.get("Extension") or record.get("extension")
        part_name = record.get("PartName") or record.get("part_name")
        if part_name and not str(part_name).startswith("/"):
            part_name = "/" + _normalise_package_part(str(part_name))
        match = None
        for node in root:
            tag = node.tag.rsplit("}", 1)[-1]
            if extension is not None and tag == "Default" and node.get("Extension") == str(extension):
                match = node
                break
            if part_name is not None and tag == "Override" and node.get("PartName") == str(part_name):
                match = node
                break
        if op in {"delete", "remove"}:
            if match is not None:
                root.remove(match)
            continue
        content_type = record.get("ContentType") or record.get("content_type") or record.get("value")
        if not content_type:
            raise ValueError("Content-type upsert requires content_type/ContentType.")
        if extension is not None:
            node = match if match is not None else ET.SubElement(root, f"{{{_CONTENT_TYPES_NS}}}Default")
            node.set("Extension", str(extension))
        elif part_name is not None:
            node = match if match is not None else ET.SubElement(root, f"{{{_CONTENT_TYPES_NS}}}Override")
            node.set("PartName", str(part_name))
        else:
            raise ValueError("Content-type edit requires extension or part_name/PartName.")
        node.set("ContentType", str(content_type))
    entries["[Content_Types].xml"] = ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _apply_package_edits(xlsx_path: str, edits: dict | None) -> None:
    """Apply expert package edits in one ZIP rewrite so partial edits never persist."""
    if not edits:
        return
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        infos = {item.filename: item for item in archive.infolist()}
        entries = {item.filename: archive.read(item.filename) for item in archive.infolist()}

    for name in edits.get("delete") or []:
        entries.pop(_normalise_package_part(name), None)

    upserts = edits.get("upsert") or {}
    if isinstance(upserts, list):
        upserts = {
            record.get("part") or record.get("name") or record.get("part_name"): record
            for record in upserts
        }
    for name, value in upserts.items():
        if not name:
            raise ValueError("Package upsert requires a part name.")
        entries[_normalise_package_part(name)] = _package_edit_bytes(value)

    _apply_relationship_edits(entries, edits.get("relationships"))
    _apply_content_type_edits(entries, edits.get("content_types"))

    replacement = xlsx_path + ".~package-edits.tmp"
    try:
        with zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, raw in entries.items():
                info = infos.get(name)
                archive.writestr(info if info is not None else name, raw)
        os.replace(replacement, xlsx_path)
    except Exception:
        if os.path.exists(replacement):
            os.remove(replacement)
        raise


_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XML_SPACE_ATTR = "{http://www.w3.org/XML/1998/namespace}space"


def _qname(local_name: str) -> str:
    return f"{{{_SPREADSHEET_NS}}}{local_name}"


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _semantic_attrs(element) -> dict:
    result = {}
    for key, value in element.attrib.items():
        if key == _XML_SPACE_ATTR:
            key = "xml:space"
        result[key] = value
    return result


def _node_semantics(element, include_children: bool = True) -> dict:
    node = {"tag": _local_name(element.tag), "attrs": _semantic_attrs(element)}
    if element.text is not None:
        node["text"] = element.text
    if include_children:
        children = [_node_semantics(child) for child in element]
        if children:
            node["children"] = children
    return node


_XML_TRUE_VALUES = {"1", "true", "on", "yes"}


def _typed_xml_attrs(
    node: dict | None,
    *,
    booleans: set[str] | None = None,
    integers: set[str] | None = None,
    numbers: set[str] | None = None,
) -> dict:
    if not node:
        return {}
    boolean_names = booleans or set()
    integer_names = integers or set()
    number_names = numbers or set()
    result = {}
    for raw_key, raw_value in (node.get("attrs") or {}).items():
        key = _local_name(raw_key)
        if key in boolean_names:
            result[key] = str(raw_value).lower() in _XML_TRUE_VALUES
        elif key in integer_names:
            try:
                result[key] = int(raw_value)
            except (TypeError, ValueError):
                result[key] = raw_value
        elif key in number_names:
            try:
                value = float(raw_value)
                result[key] = int(value) if value.is_integer() else value
            except (TypeError, ValueError):
                result[key] = raw_value
        else:
            result[key] = raw_value
    return result


def _worksheet_node(snapshot: dict | None, tag: str) -> dict | None:
    values = ((snapshot or {}).get("nodes") or {}).get(tag) or []
    return values[-1] if values else None


def _semantic_child(node: dict | None, tag: str) -> dict | None:
    values = [child for child in (node or {}).get("children") or [] if child.get("tag") == tag]
    return values[-1] if values else None


def _split_header_footer_text(value: str | None) -> dict:
    if not value:
        return {}
    matches = list(re.finditer(r"(?<!&)&([LCR])", value))
    if not matches:
        return {"center": value}
    names = {"L": "left", "C": "center", "R": "right"}
    result = {}
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        result[names[match.group(1)]] = value[match.end():end]
    return result


def _worksheet_models_from_semantics(snapshot: dict | None) -> dict:
    sheet_property_booleans = {
        "enableFormatConditionsCalculation", "filterMode", "published", "syncHorizontal",
        "syncVertical", "transitionEvaluation", "transitionEntry",
    }
    outline_booleans = {"applyStyles", "summaryBelow", "summaryRight", "showOutlineSymbols"}
    page_setup_property_booleans = {"autoPageBreaks", "fitToPage"}
    sheet_view_booleans = {
        "windowProtection", "showFormulas", "showGridLines", "showRowColHeaders", "showZeros",
        "rightToLeft", "tabSelected", "showRuler", "showOutlineSymbols", "defaultGridColor",
        "showWhiteSpace", "zoomToFit",
    }
    sheet_view_integers = {
        "colorId", "zoomScale", "zoomScaleNormal", "zoomScaleSheetLayoutView",
        "zoomScalePageLayoutView", "workbookViewId",
    }
    page_setup_booleans = {"useFirstPageNumber", "usePrinterDefaults", "blackAndWhite", "draft"}
    page_setup_integers = {
        "scale", "fitToHeight", "fitToWidth", "firstPageNumber", "horizontalDpi",
        "verticalDpi", "copies",
    }
    print_option_booleans = {
        "horizontalCentered", "verticalCentered", "headings", "gridLines", "gridLinesSet",
    }
    header_footer_booleans = {"differentOddEven", "differentFirst", "scaleWithDoc", "alignWithMargins"}
    break_booleans = {"man", "pt"}
    break_integers = {"id", "min", "max"}

    result = {}
    sheet_properties_node = _worksheet_node(snapshot, "sheetPr")
    page_setup_properties = None
    if sheet_properties_node:
        sheet_properties = _typed_xml_attrs(sheet_properties_node, booleans=sheet_property_booleans)
        outline_node = _semantic_child(sheet_properties_node, "outlinePr")
        if outline_node:
            sheet_properties["outline"] = _typed_xml_attrs(outline_node, booleans=outline_booleans)
        page_setup_properties_node = _semantic_child(sheet_properties_node, "pageSetUpPr")
        if page_setup_properties_node:
            page_setup_properties = _typed_xml_attrs(
                page_setup_properties_node,
                booleans=page_setup_property_booleans,
            )
            sheet_properties["page_setup_properties"] = copy.deepcopy(page_setup_properties)
        result["sheet_properties"] = sheet_properties

    sheet_views_node = _worksheet_node(snapshot, "sheetViews")
    if sheet_views_node:
        views = []
        for view_node in sheet_views_node.get("children") or []:
            if view_node.get("tag") != "sheetView":
                continue
            view = _typed_xml_attrs(
                view_node,
                booleans=sheet_view_booleans,
                integers=sheet_view_integers,
            )
            pane_node = _semantic_child(view_node, "pane")
            if pane_node:
                view["pane"] = _typed_xml_attrs(pane_node, numbers={"xSplit", "ySplit"})
            selections = [
                _typed_xml_attrs(child, integers={"activeCellId"})
                for child in view_node.get("children") or []
                if child.get("tag") == "selection"
            ]
            if selections:
                view["selections"] = selections
            views.append(view)
        result["sheet_views"] = views

    page_setup_node = _worksheet_node(snapshot, "pageSetup")
    page_setup = _typed_xml_attrs(
        page_setup_node,
        booleans=page_setup_booleans,
        integers=page_setup_integers,
    ) if page_setup_node else {}
    if page_setup_node:
        page_setup["present"] = True
    if page_setup_properties and "fitToPage" in page_setup_properties:
        page_setup["fitToPage"] = page_setup_properties["fitToPage"]
    if page_setup:
        result["page_setup"] = page_setup

    print_options_node = _worksheet_node(snapshot, "printOptions")
    if print_options_node:
        print_options = _typed_xml_attrs(print_options_node, booleans=print_option_booleans)
        print_options["present"] = True
        result["print_options"] = print_options

    header_footer_node = _worksheet_node(snapshot, "headerFooter")
    if header_footer_node:
        header_footer = _typed_xml_attrs(header_footer_node, booleans=header_footer_booleans)
        section_names = {
            "oddHeader": "odd_header", "oddFooter": "odd_footer",
            "evenHeader": "even_header", "evenFooter": "even_footer",
            "firstHeader": "first_header", "firstFooter": "first_footer",
        }
        for child in header_footer_node.get("children") or []:
            section_name = section_names.get(child.get("tag"))
            if section_name:
                header_footer[section_name] = _split_header_footer_text(child.get("text"))
        odd_aliases = {
            ("odd_header", "left"): "hl", ("odd_header", "center"): "hc", ("odd_header", "right"): "hr",
            ("odd_footer", "left"): "fl", ("odd_footer", "center"): "fc", ("odd_footer", "right"): "fr",
        }
        for (section_name, position), alias in odd_aliases.items():
            if position in (header_footer.get(section_name) or {}):
                header_footer[alias] = header_footer[section_name][position]
        result["header_footer"] = header_footer

    page_breaks = {}
    for tag, axis in (("rowBreaks", "rows"), ("colBreaks", "columns")):
        container = _worksheet_node(snapshot, tag)
        if not container:
            continue
        attrs = _typed_xml_attrs(container, integers={"count", "manualBreakCount"})
        page_breaks[f"{axis}_count"] = attrs.get("count", 0)
        page_breaks[f"{axis}_manualBreakCount"] = attrs.get("manualBreakCount", 0)
        page_breaks[axis] = [
            _typed_xml_attrs(child, booleans=break_booleans, integers=break_integers)
            for child in container.get("children") or []
            if child.get("tag") == "brk"
        ]
    if page_breaks:
        result["page_breaks"] = page_breaks

    protected_ranges_node = _worksheet_node(snapshot, "protectedRanges")
    if protected_ranges_node:
        result["protected_ranges"] = [
            _typed_xml_attrs(child, integers={"spinCount"})
            for child in protected_ranges_node.get("children") or []
            if child.get("tag") == "protectedRange"
        ]
    return result


def _xml_boolean(element) -> bool:
    value = element.get("val")
    return value is None or str(value).lower() not in {"0", "false", "off", "no"}


def _color_semantics_from_xml(element) -> dict | None:
    if element is None:
        return None
    attrs = _semantic_attrs(element)
    result = {"attrs": attrs}
    if "rgb" in attrs:
        result.update({"type": "rgb", "rgb": attrs["rgb"]})
    elif "theme" in attrs:
        result.update({"type": "theme", "theme": int(attrs["theme"])})
    elif "indexed" in attrs:
        result.update({"type": "indexed", "indexed": int(attrs["indexed"])})
    elif "auto" in attrs:
        result.update({"type": "auto", "auto": str(attrs["auto"]).lower() not in {"0", "false"}})
    else:
        result["type"] = None
    if "tint" in attrs:
        result["tint"] = float(attrs["tint"])
    return result


def _font_semantics_from_rpr(rpr) -> dict | None:
    if rpr is None:
        return None
    result = {"attrs": _semantic_attrs(rpr)}
    names = {
        "rFont": "name", "sz": "size", "b": "bold", "i": "italic",
        "u": "underline", "strike": "strike", "vertAlign": "vertAlign",
        "charset": "charset", "family": "family", "scheme": "scheme",
        "outline": "outline", "shadow": "shadow", "condense": "condense",
        "extend": "extend",
    }
    boolean_tags = {"b", "i", "strike", "outline", "shadow", "condense", "extend"}
    for child in rpr:
        tag = _local_name(child.tag)
        key = names.get(tag, tag)
        if tag == "color":
            result["color"] = _color_semantics_from_xml(child)
        elif tag in boolean_tags:
            result[key] = _xml_boolean(child)
        else:
            value = child.get("val")
            if tag in {"sz"} and value is not None:
                value = float(value)
            elif tag in {"charset", "family"} and value is not None:
                try:
                    value = int(value)
                except ValueError:
                    pass
            result[key] = value
    return result


def _rich_text_semantics(container, storage: str, shared_index: int | None = None) -> dict:
    runs = []
    offset = 0
    rich_nodes = container.findall(_qname("r"))
    if rich_nodes:
        source_runs = rich_nodes
    else:
        source_runs = [container]
    for source_run in source_runs:
        text_node = source_run.find(_qname("t"))
        text = text_node.text if text_node is not None and text_node.text is not None else ""
        run = {
            "text": text,
            "start": offset,
            "end": offset + len(text),
            "font": _font_semantics_from_rpr(source_run.find(_qname("rPr"))),
            "xml_space": text_node.get(_XML_SPACE_ATTR) if text_node is not None else None,
        }
        runs.append(run)
        offset = run["end"]

    phonetic_runs = []
    for phonetic in container.findall(_qname("rPh")):
        text_node = phonetic.find(_qname("t"))
        phonetic_runs.append({
            "text": text_node.text if text_node is not None and text_node.text is not None else "",
            "start": int(phonetic.get("sb", "0")),
            "end": int(phonetic.get("eb", "0")),
            "xml_space": text_node.get(_XML_SPACE_ATTR) if text_node is not None else None,
            "attrs": _semantic_attrs(phonetic),
        })
    phonetic_pr = container.find(_qname("phoneticPr"))
    result = {
        "is_rich_text": bool(rich_nodes),
        "storage": storage,
        "runs": runs,
        "phonetic_runs": phonetic_runs,
        "phonetic_properties": _semantic_attrs(phonetic_pr) if phonetic_pr is not None else None,
        "text": "".join(run["text"] for run in runs),
    }
    if shared_index is not None:
        result["shared_string_index"] = shared_index
    return result


def _style_semantics(styles_root) -> dict:
    if styles_root is None:
        return {"cell_xfs": [], "cell_style_xfs": [], "borders": [], "named_styles": []}

    def children_of(parent_name: str, child_name: str) -> list[dict]:
        parent = styles_root.find(_qname(parent_name))
        if parent is None:
            return []
        return [_node_semantics(child) for child in parent.findall(_qname(child_name))]

    named_styles = []
    cell_styles = styles_root.find(_qname("cellStyles"))
    if cell_styles is not None:
        named_styles = [_node_semantics(node) for node in cell_styles.findall(_qname("cellStyle"))]
    return {
        "cell_xfs": children_of("cellXfs", "xf"),
        "cell_style_xfs": children_of("cellStyleXfs", "xf"),
        "fonts": children_of("fonts", "font"),
        "fills": children_of("fills", "fill"),
        "borders": children_of("borders", "border"),
        "named_styles": named_styles,
    }


_XF_BOOLEAN_ATTRS = {
    "applyNumberFormat", "applyFont", "applyFill", "applyBorder",
    "applyAlignment", "applyProtection", "pivotButton", "quotePrefix",
}
_XF_INTEGER_ATTRS = {"numFmtId", "fontId", "fillId", "borderId", "xfId"}


def _snapshot_named_style_names(wb) -> tuple[str, ...]:
    try:
        return tuple(wb._named_styles.names)
    except (AttributeError, IndexError, TypeError):
        return ()


def _safe_named_style_name(cell, named_style_names: tuple[str, ...]) -> str | None:
    try:
        style = getattr(cell, "_style")
        if style is None:
            return named_style_names[0]
        style_index = getattr(style, "xfId")
        if not isinstance(style_index, int) or isinstance(style_index, bool) or style_index < 0:
            return None
        return named_style_names[style_index]
    except (AttributeError, IndexError, TypeError):
        return None


def _cell_xf_semantics(style_id: int, definition: dict | None) -> dict:
    result = {"style_id": style_id, "definition": definition}
    attrs = (definition or {}).get("attrs") or {}
    for key, value in attrs.items():
        if key in _XF_BOOLEAN_ATTRS:
            result[key] = str(value).lower() in {"1", "true"}
        elif key in _XF_INTEGER_ATTRS:
            try:
                result[key] = int(value)
            except (TypeError, ValueError):
                result[key] = value
    return result


def _extract_ooxml_semantics(path: str | Path, sheet_file_map: dict) -> dict:
    with zipfile.ZipFile(path, "r") as archive:
        names = set(archive.namelist())
        shared_strings = []
        if "xl/sharedStrings.xml" in names:
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for index, item in enumerate(shared_root.findall(_qname("si"))):
                shared_strings.append(_rich_text_semantics(item, "shared", index))

        styles_root = ET.fromstring(archive.read("xl/styles.xml")) if "xl/styles.xml" in names else None
        styles = _style_semantics(styles_root)
        cells_by_sheet = {}
        worksheets = {}
        for sheet_name, part_name in sheet_file_map.items():
            if part_name not in names:
                continue
            root = ET.fromstring(archive.read(part_name))
            cell_map = {}
            for cell_node in root.findall(f".//{_qname('c')}"):
                coord = cell_node.get("r")
                if not coord:
                    continue
                item = {"present": True, "cell_attrs": _semantic_attrs(cell_node)}
                formula_node = cell_node.find(_qname("f"))
                if formula_node is not None:
                    cached_node = cell_node.find(_qname("v"))
                    if cached_node is None:
                        cache_state = "missing"
                        cached_value = None
                    elif cached_node.text is None:
                        cache_state = "empty"
                        cached_value = None
                    else:
                        cache_state = "value"
                        cached_value = cached_node.text
                    item["formula"] = {
                        "text": formula_node.text or "",
                        "attrs": _semantic_attrs(formula_node),
                        "cached_value": cached_value,
                        "cache_state": cache_state,
                    }
                cell_type = cell_node.get("t")
                if cell_type == "inlineStr":
                    inline = cell_node.find(_qname("is"))
                    if inline is not None:
                        rich = _rich_text_semantics(inline, "inline")
                        if rich["is_rich_text"] or rich["phonetic_runs"] or rich["phonetic_properties"]:
                            item["rich_text"] = rich
                elif cell_type == "s":
                    value_node = cell_node.find(_qname("v"))
                    try:
                        shared_index = int(value_node.text) if value_node is not None else -1
                    except (TypeError, ValueError):
                        shared_index = -1
                    if 0 <= shared_index < len(shared_strings):
                        rich = copy.deepcopy(shared_strings[shared_index])
                        if rich["is_rich_text"] or rich["phonetic_runs"] or rich["phonetic_properties"]:
                            item["rich_text"] = rich
                cell_map[coord] = item

            top_level = {}
            for child in root:
                tag = _local_name(child.tag)
                if tag == "sheetData":
                    top_level[tag] = {"attrs": _semantic_attrs(child)}
                    continue
                top_level.setdefault(tag, []).append(_node_semantics(child))
            worksheets[sheet_name] = {
                "part_name": part_name,
                "root_attrs": _semantic_attrs(root),
                "nodes": top_level,
            }
            cells_by_sheet[sheet_name] = cell_map

        workbook = {}
        if "xl/workbook.xml" in names:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))

            def _last_singleton_child(tag: str):
                # These elements are maxOccurs=1 per the schema, but some
                # producers (or patched/merged documents) can leave a stray
                # duplicate behind -- e.g. openpyxl always emits an empty
                # <workbookProtection/> even when no protection is set, and a
                # later edit can append a second, populated element rather
                # than replacing it in place. `.find()` only returns the
                # FIRST match, which would silently pick up the empty stub
                # and hide the real one. Last-one-wins mirrors how Excel
                # itself resolves duplicate singleton elements.
                matches = workbook_root.findall(_qname(tag))
                return matches[-1] if matches else None

            workbook_pr = _last_singleton_child("workbookPr")
            calc_pr = _last_singleton_child("calcPr")
            workbook_protection = _last_singleton_child("workbookProtection")
            workbook = {
                "root_attrs": _semantic_attrs(workbook_root),
                "workbook_properties": _node_semantics(workbook_pr) if workbook_pr is not None else None,
                "calculation": _node_semantics(calc_pr) if calc_pr is not None else None,
                "protection": _node_semantics(workbook_protection) if workbook_protection is not None else None,
                "views": [_node_semantics(node) for node in workbook_root.findall(f"{_qname('bookViews')}/{_qname('workbookView')}")],
                "sheets": [_node_semantics(node) for node in workbook_root.findall(f"{_qname('sheets')}/{_qname('sheet')}")],
                "defined_names": [_node_semantics(node) for node in workbook_root.findall(f"{_qname('definedNames')}/{_qname('definedName')}")],
            }

        table_parts = {}
        for name in sorted(part for part in names if part.startswith("xl/tables/") and part.endswith(".xml")):
            try:
                table_root = ET.fromstring(archive.read(name))
                table_parts[name] = _node_semantics(table_root)
            except Exception:
                pass

    return {
        "cells": cells_by_sheet,
        "worksheets": worksheets,
        "workbook": workbook,
        "styles": styles,
        "tables": table_parts,
    }


def _color_semantics_from_openpyxl(color) -> dict | None:
    ref = _color_ref_from_openpyxl(color)
    if not ref:
        return None
    return {key: value for key, value in ref.items() if key != "rgb" or ref.get("type") == "rgb"}


def _alignment_semantics(alignment) -> dict:
    result = {}
    unwrapped = copy.copy(alignment)
    for attr in getattr(type(unwrapped), "__attrs__", ()):
        value = getattr(alignment, attr, None)
        if value is not None:
            result[attr] = value
    return result


def _border_semantics(border, theme_colors: list[str]) -> dict:
    result = {}
    for attr in ("start", "end", "left", "right", "top", "bottom", "diagonal", "vertical", "horizontal"):
        side = getattr(border, attr, None)
        if side is None:
            continue
        color_ref = _color_ref_from_openpyxl(getattr(side, "color", None))
        item = {}
        if side.border_style is not None:
            item["style"] = side.border_style
        if color_ref is not None:
            item["color"] = color_ref
        resolved = _resolve_color(getattr(side, "color", None), theme_colors)
        if resolved is not None:
            item["resolved_rgb"] = resolved
        result[attr] = item
    for attr in ("diagonalUp", "diagonalDown", "outline"):
        value = getattr(border, attr, None)
        if value is not None:
            result[attr] = value
    return result


def _serialize_named_styles(wb) -> list[dict]:
    result = []
    try:
        named_styles = tuple(wb._named_styles)
    except (AttributeError, TypeError):
        return result

    for named in named_styles:
        name = getattr(named, "name", None)
        if not name:
            continue

        item = {"name": name}
        builtin_id = getattr(named, "builtinId", None)
        hidden = getattr(named, "hidden", None)
        if builtin_id is not None:
            item["builtinId"] = builtin_id
        if hidden is not None:
            item["hidden"] = hidden

        style = {}
        font = getattr(named, "font", None)
        if font is not None:
            font_spec = {}
            for source_key, target_key in (
                ("name", "name"), ("size", "size"), ("bold", "bold"),
                ("italic", "italic"), ("underline", "underline"),
                ("strike", "strike"), ("vertAlign", "vertAlign"),
                ("charset", "charset"), ("family", "family"),
                ("scheme", "scheme"), ("outline", "outline"),
                ("shadow", "shadow"), ("condense", "condense"),
                ("extend", "extend"),
            ):
                value = getattr(font, source_key, None)
                if value is None or value is False:
                    continue
                font_spec[target_key] = value
            color_ref = _color_ref_from_openpyxl(getattr(font, "color", None))
            if color_ref is not None:
                font_spec["color"] = color_ref
            if font_spec:
                style["font"] = font_spec

        fill = getattr(named, "fill", None)
        if fill is not None:
            fill_spec = {}
            pattern_type = getattr(fill, "fill_type", None)
            foreground = _color_ref_from_openpyxl(getattr(fill, "fgColor", None))
            background = _color_ref_from_openpyxl(getattr(fill, "bgColor", None))
            if pattern_type is not None:
                fill_spec["pattern_type"] = pattern_type
            if foreground is not None:
                fill_spec["foreground"] = foreground
            if background is not None:
                fill_spec["background"] = background
            if fill_spec:
                style["fill"] = fill_spec

        border = getattr(named, "border", None)
        if border is not None:
            border_spec = _border_semantics(border, [])
            side_keys = (
                "start", "end", "left", "right", "top", "bottom",
                "diagonal", "vertical", "horizontal",
            )
            has_explicit_border = any(key in border_spec for key in side_keys)
            has_explicit_border = has_explicit_border or bool(
                border_spec.get("diagonalUp") or border_spec.get("diagonalDown")
            )
            has_explicit_border = has_explicit_border or border_spec.get("outline") is False
            if has_explicit_border:
                for side_key in side_keys:
                    side = border_spec.get(side_key)
                    if isinstance(side, dict):
                        side.pop("resolved_rgb", None)
                style["border"] = border_spec

        alignment = getattr(named, "alignment", None)
        if alignment is not None:
            alignment_spec = {
                key: value
                for key, value in _alignment_semantics(alignment).items()
                if value not in (None, False, 0, 0.0)
            }
            if alignment_spec:
                style["alignment"] = alignment_spec

        protection = getattr(named, "protection", None)
        if protection is not None:
            protection_spec = {}
            locked = getattr(protection, "locked", None)
            hidden_cell = getattr(protection, "hidden", None)
            if locked is False:
                protection_spec["locked"] = False
            if hidden_cell is True:
                protection_spec["hidden"] = True
            if protection_spec:
                style["protection"] = protection_spec

        number_format = getattr(named, "number_format", None)
        if number_format not in (None, "General"):
            style["number_format"] = number_format
        item["style"] = style
        result.append(item)
    return result


def _make_border_side_semantic(item: dict | None):
    from openpyxl.styles import Side
    if item is None:
        return None
    color = _make_color_from_ref(item.get("color") or {}) if item.get("color") else None
    return Side(border_style=item.get("style"), color=color)


_IMPLICIT_CELL_DEFAULTS = {
    "v": None,
    "fill": None,
    "bold": False,
    "italic": False,
    "size": None,
    "font": None,
    "fcolor": None,
    "uline": None,
    "strike": False,
    "vAlign": None,
    "wrap": False,
    "halign": None,
    "valign": None,
    "rot": None,
    "indent": None,
    "shrink": False,
    "numfmt": "General",
    "merge": {},
    "border": {},
    "locked": True,
    "hidden_cell": False,
    "fill_color": None,
    "font_color": None,
    "present": False,
    "data_type": "n",
    "formula": None,
    "rich_text": None,
    "alignment": {},
    "border_semantics": {},
    "xf": {"style_id": 0, "definition": None},
    "named_style": None,
    "cell_attrs": {},
}


def _implicit_cell_placeholder() -> dict:
    return {"_implicit": True, "v": None, "merge": {}}


def _expanded_implicit_cell(cell_data: dict, defaults: dict | None = None) -> dict:
    expanded = copy.deepcopy(defaults or _IMPLICIT_CELL_DEFAULTS)
    expanded.update({key: value for key, value in cell_data.items() if key != "_implicit"})
    return expanded


def _style_cache_key(cell, style_id: int) -> tuple[int, int]:
    try:
        xf_id = int(getattr(getattr(cell, "_style", None), "xfId", 0) or 0)
    except (TypeError, ValueError):
        xf_id = 0
    return style_id, xf_id


def _build_cached_cell_style(
    cell,
    style_id: int,
    style_semantics: dict,
    named_style_names: tuple[str, ...],
    theme_colors: list[str],
) -> dict:
    fill = cell.fill
    fill_rgb = None
    if fill and fill.fill_type == "solid":
        fill_rgb = _resolve_color(fill.fgColor, theme_colors)

    font = cell.font
    fcolor = _resolve_color(font.color if font else None, theme_colors)
    if fcolor in ("FF000000", "00000000"):
        fcolor = None

    border = cell.border
    border_data = {}
    if border:
        for attr in (
            "start", "end", "top", "bottom", "left", "right", "diagonal", "vertical", "horizontal",
        ):
            side = _ser_border_side(getattr(border, attr), theme_colors)
            if side:
                border_data[attr] = side
        if border.diagonalUp:
            border_data["diagonalUp"] = True
        if border.diagonalDown:
            border_data["diagonalDown"] = True
        if border.outline is not None:
            border_data["outline"] = border.outline

    alignment = cell.alignment
    xf_definition = (
        style_semantics["cell_xfs"][style_id]
        if 0 <= style_id < len(style_semantics["cell_xfs"])
        else None
    )
    cell_style = {
        "fill": fill_rgb,
        "bold": bool(font.bold) if font else False,
        "italic": bool(font.italic) if font else False,
        "size": font.size if font else None,
        "font": font.name if font else None,
        "fcolor": fcolor,
        "uline": font.underline if font else None,
        "strike": bool(font.strike) if font else False,
        "vAlign": font.vertAlign if font else None,
        "wrap": bool(alignment.wrap_text) if alignment else False,
        "halign": alignment.horizontal if alignment else None,
        "valign": alignment.vertical if alignment else None,
        "rot": alignment.text_rotation if alignment else None,
        "indent": alignment.indent if alignment else None,
        "shrink": bool(alignment.shrink_to_fit) if alignment else False,
        "numfmt": cell.number_format,
        "border": border_data,
        "locked": bool(cell.protection.locked) if cell.protection else True,
        "hidden_cell": bool(cell.protection.hidden) if cell.protection else False,
        "fill_color": _color_semantics_from_openpyxl(getattr(fill, "fgColor", None)),
        "font_color": _color_semantics_from_openpyxl(getattr(font, "color", None)),
        "alignment": _alignment_semantics(alignment) if alignment else {},
        "border_semantics": _border_semantics(border, theme_colors) if border else {},
        "xf": _cell_xf_semantics(style_id, xf_definition),
        "named_style": _safe_named_style_name(cell, named_style_names),
    }
    try:
        if cell._style is not None and cell._style.quotePrefix:
            cell_style["qp"] = True
    except Exception:
        pass
    return {
        "data": cell_style,
        "font_raw": _font_raw_from_openpyxl(font, fcolor),
        "fill_preservation": {
            "patternType": fill.fill_type if fill else None,
            "fgColor": _color_ref_from_openpyxl(getattr(fill, "fgColor", None)),
            "bgColor": _color_ref_from_openpyxl(getattr(fill, "bgColor", None)),
        },
    }


def _implicit_cell_defaults_from_cached_style(cell, cached_style: dict) -> dict:
    defaults = {
        **cached_style["data"],
        "v": None,
        "merge": {},
        "data_type": cell.data_type,
        "present": False,
        "formula": None,
        "rich_text": None,
        "cell_attrs": {},
    }
    if cached_style["font_raw"]:
        defaults["_font_raw"] = cached_style["font_raw"]
    return defaults


def _cached_fill_raw(raw_cell_fill: dict | None, cached_style: dict) -> dict | None:
    if not raw_cell_fill:
        return None
    fill_raw = dict(raw_cell_fill)
    fill_raw["rgb"] = cached_style["data"]["fill"]
    fill_raw["is_gradient"] = "gradientFill" in (fill_raw.get("xml") or "")
    fill_raw["patternType"] = (
        cached_style["fill_preservation"]["patternType"]
        if not fill_raw["is_gradient"]
        else None
    )
    for key in ("fgColor", "bgColor"):
        value = cached_style["fill_preservation"].get(key)
        if value:
            fill_raw[key] = value
    return fill_raw


_CELL_CONTENT_KEYS = (
    "v", "present", "formula", "rich_text",
)
_CELL_STYLE_KEYS = (
    "fill", "fill_color", "bold", "italic", "size", "font", "fcolor", "font_color",
    "uline", "strike", "vAlign", "wrap", "halign", "valign", "rot", "indent", "shrink",
    "alignment", "numfmt", "border", "border_semantics", "locked", "hidden_cell", "qp",
    "xf", "named_style", "_fill_raw", "_font_raw",
)
_CELL_BASELINE_HASH_KEYS = (
    "_baseline_content_hash", "_baseline_style_hash", "_baseline_structure_hash",
)


def _cell_group_payloads(cell_data: dict) -> tuple[dict, dict, dict]:
    content = {key: cell_data.get(key) for key in _CELL_CONTENT_KEYS}
    content["data_type"] = cell_data.get("dt") or cell_data.get("data_type")
    return (
        content,
        {key: cell_data.get(key) for key in _CELL_STYLE_KEYS},
        {"merge": cell_data.get("merge")},
    )


def _cell_baseline(cell_data: dict) -> None:
    """Capture immutable semantic hashes once, immediately before first mutation."""
    if all(key in cell_data for key in _CELL_BASELINE_HASH_KEYS):
        return
    content, style, structure = _cell_group_payloads(cell_data)
    cell_data["_baseline_content_hash"] = _semantic_digest(content)
    cell_data["_baseline_style_hash"] = _semantic_digest(style)
    cell_data["_baseline_structure_hash"] = _semantic_digest(structure)
    cell_data.setdefault("_dirty", [])


def _xml_attributes(attrs: dict) -> str:
    return "".join(f" {key}={quoteattr(str(value))}" for key, value in attrs.items() if value is not None)


def _color_xml(color: dict | None) -> str:
    if not color:
        return ""
    attrs = dict(color.get("attrs") or {})
    if not attrs:
        color_type = color.get("type")
        if color_type and color.get(color_type) is not None:
            value = color[color_type]
            attrs[color_type] = "1" if color_type == "auto" and value is True else value
        if color.get("tint") is not None:
            attrs["tint"] = color["tint"]
    return f"<color{_xml_attributes(attrs)}/>"


def _run_properties_xml(font: dict | None) -> str:
    if not font:
        return ""
    parts = []
    scalar_tags = (
        ("name", "rFont"), ("charset", "charset"), ("family", "family"),
        ("size", "sz"), ("underline", "u"), ("vertAlign", "vertAlign"),
        ("scheme", "scheme"),
    )
    boolean_tags = (
        ("bold", "b"), ("italic", "i"), ("strike", "strike"),
        ("outline", "outline"), ("shadow", "shadow"),
        ("condense", "condense"), ("extend", "extend"),
    )
    for key, tag in scalar_tags:
        if key in font and font[key] is not None:
            parts.append(f"<{tag} val={quoteattr(str(font[key]))}/>")
    for key, tag in boolean_tags:
        if key in font:
            parts.append(f"<{tag}" + ("/>" if font[key] else ' val="0"/>'))
    if font.get("color"):
        color = _color_xml(font["color"])
        parts.append(color.replace("<color", "<color", 1))
    return "<rPr>" + "".join(parts) + "</rPr>" if parts else ""


def _text_xml(text: str, xml_space: str | None = None) -> str:
    if xml_space is None and (text[:1].isspace() or text[-1:].isspace()):
        xml_space = "preserve"
    attrs = f' xml:space={quoteattr(xml_space)}' if xml_space else ""
    return f"<t{attrs}>{escape(text)}</t>"


def _rich_text_xml(rich_text: dict) -> str:
    runs = rich_text.get("runs") or []
    if rich_text.get("is_rich_text") or any(run.get("font") for run in runs):
        body = "".join(
            "<r>" + _run_properties_xml(run.get("font"))
            + _text_xml(str(run.get("text") or ""), run.get("xml_space")) + "</r>"
            for run in runs
        )
    else:
        text = "".join(str(run.get("text") or "") for run in runs)
        body = _text_xml(text, runs[0].get("xml_space") if runs else None)

    for phonetic in rich_text.get("phonetic_runs") or []:
        attrs = dict(phonetic.get("attrs") or {})
        attrs.setdefault("sb", phonetic.get("start", 0))
        attrs.setdefault("eb", phonetic.get("end", 0))
        body += (
            f"<rPh{_xml_attributes(attrs)}>"
            + _text_xml(str(phonetic.get("text") or ""), phonetic.get("xml_space"))
            + "</rPh>"
        )
    if rich_text.get("phonetic_properties"):
        body += f"<phoneticPr{_xml_attributes(rich_text['phonetic_properties'])}/>"
    return "<is>" + body + "</is>"


_FORMULA_ERROR_TOKENS = {
    "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!",
    "#GETTING_DATA", "#SPILL!", "#CALC!", "#BLOCKED!", "#CONNECT!",
    "#EXTERNAL!", "#FIELD!", "#UNKNOWN!",
}


def _formula_cache_type_and_text(cached_value, original_type: str | None) -> tuple[str | None, str]:
    """Classify a formula's cached result so <c t=...><v> matches OOXML rules.

    Returns (t attribute or None for numeric-default, <v> body text).
    Distinguishes boolean/error/string/numeric cached values instead of
    blindly reusing a stale original cell type: previously any non-numeric
    cached value (a string, a boolean, an error token) was written with no
    t="..." attribute at all, which defaults to numeric per OOXML and produces
    a file even openpyxl itself cannot reload with data_only=True.
    """
    if isinstance(cached_value, bool):
        return "b", ("1" if cached_value else "0")
    if isinstance(cached_value, (int, float)):
        return None, str(cached_value)
    if isinstance(cached_value, str):
        if cached_value in _FORMULA_ERROR_TOKENS:
            return "e", cached_value
        if original_type == "b" and cached_value in ("0", "1"):
            return "b", cached_value
        if original_type in (None, "n"):
            try:
                float(cached_value)
                return None, cached_value
            except ValueError:
                pass
        return "str", cached_value
    return original_type, "" if cached_value is None else str(cached_value)


def _formula_xml(formula: dict, value_text: str | None = None) -> str:
    # Two conventions reach this function: core.py's own XML-parsed raw model
    # ("attrs"/"cache_state", "text" without a leading "=") and main.py's
    # public-tool model (_set_formula_cell: "attributes"/"cached_value_state",
    # "text" WITH a leading "="). Accept either so a formula set via
    # excel_set_formula writes its cache/attrs exactly like one loaded from
    # disk. <f> elements must never contain the leading "=".
    attrs_dict = dict(formula.get("attrs") or formula.get("attributes") or {})
    # main.py's public-tool model (_set_formula_cell) keeps the formula kind
    # ("shared"/"array"/"dataTable") in a separate "type" field rather than
    # folding it into "attributes" -- without this, excel_set_formula's
    # documented formula_type parameter had no effect on the actual OOXML
    # unless the caller also manually duplicated t="..." into
    # formula_attributes themselves.
    formula_type = formula.get("type")
    if formula_type and formula_type != "normal" and "t" not in attrs_dict:
        attrs_dict["t"] = formula_type
    attrs = _xml_attributes(attrs_dict)
    text = str(formula.get("text") or "").lstrip("=")
    body = f"<f{attrs}>{escape(text)}</f>"
    state = formula.get("cache_state") or formula.get("cached_value_state") or "missing"
    if state == "empty":
        body += "<v></v>"
    elif state == "value":
        if value_text is None:
            value_text = str(formula.get("cached_value") or "")
        body += f"<v>{escape(value_text)}</v>"
    return body


def _contract_is_current(cell_data: dict, key: str) -> bool:
    contract = cell_data.get(key)
    if not contract:
        return False
    if key == "formula":
        # contract["text"] may or may not already carry a leading "=" —
        # normalize before re-adding it so both conventions compare correctly
        # against cell_data["v"] (which always has the leading "=").
        logical = "=" + str(contract.get("text") or "").lstrip("=")
    else:
        logical = str(contract.get("text") or "")
    if cell_data.get("v") == logical:
        return True
    baseline = cell_data.get("_baseline_content_hash")
    current = _semantic_digest(_cell_group_payloads(cell_data)[0])
    return baseline is not None and baseline == current


def _cell_xml_fragment(
    coord: str,
    cell_data: dict,
    existing: str | None,
    *,
    include_scalar: bool = False,
) -> str | None:
    formula = cell_data.get("formula") if _contract_is_current(cell_data, "formula") else None
    rich_text = cell_data.get("rich_text") if _contract_is_current(cell_data, "rich_text") else None
    explicit_blank = bool(cell_data.get("present")) and cell_data.get("v") is None
    if not (formula or rich_text or explicit_blank or include_scalar):
        return existing

    attrs = {}
    if existing:
        opening = re.match(r"<c\b([^>]*)", existing)
        if opening:
            attrs.update(_parse_xml_attrs(opening.group(1)))
    attrs["r"] = coord
    style_id = ((cell_data.get("xf") or {}).get("style_id"))
    if style_id and "s" not in attrs:
        attrs["s"] = style_id
    if rich_text:
        attrs["t"] = "inlineStr"
        body = _rich_text_xml(rich_text)
    elif formula:
        original_type = (cell_data.get("cell_attrs") or {}).get("t")
        state = formula.get("cached_value_state") or formula.get("cache_state") or "missing"
        value_text = None
        if state == "value":
            computed_type, value_text = _formula_cache_type_and_text(formula.get("cached_value"), original_type)
            if computed_type:
                attrs["t"] = computed_type
            else:
                attrs.pop("t", None)
        elif original_type:
            attrs["t"] = original_type
        else:
            attrs.pop("t", None)
        body = _formula_xml(formula, value_text)
    elif include_scalar:
        value = cell_data.get("v")
        data_type = cell_data.get("dt") or cell_data.get("data_type")
        if isinstance(value, str) and value.startswith("=") and data_type != "s":
            attrs.pop("t", None)
            body = f"<f>{escape(value[1:])}</f>"
        elif value is None:
            attrs.pop("t", None)
            body = ""
        elif isinstance(value, bool) or data_type == "b":
            attrs["t"] = "b"
            body = f"<v>{'1' if value in (True, 1, '1') else '0'}</v>"
        elif data_type == "e":
            attrs["t"] = "e"
            body = f"<v>{escape(str(value))}</v>"
        elif data_type == "d":
            attrs["t"] = "d"
            body = f"<v>{escape(str(value))}</v>"
        elif isinstance(value, (int, float)) or data_type == "n":
            attrs.pop("t", None)
            body = f"<v>{escape(str(value))}</v>"
        else:
            attrs["t"] = "inlineStr"
            body = "<is>" + _text_xml(str(value)) + "</is>"
    else:
        body = ""
    return f"<c{_xml_attributes(attrs)}>" + body + "</c>"


def _find_cell_xml(sheet_xml: str, coord: str) -> tuple[re.Match | None, str | None]:
    coord_q = re.escape(coord)
    pattern = re.compile(
        rf'<c\b(?=[^>]*\br="{coord_q}")[^>]*?(?:/>|>.*?</c>)',
        re.DOTALL,
    )
    match = pattern.search(sheet_xml)
    return match, match.group(0) if match else None


def _insert_cell_xml(sheet_xml: str, coord: str, fragment: str) -> str:
    row_number = int(re.search(r"\d+", coord).group(0))
    row_pattern = re.compile(
        rf'(<row\b(?=[^>]*\br="{row_number}")[^>]*>)(.*?)(</row>)',
        re.DOTALL,
    )
    row_match = row_pattern.search(sheet_xml)
    if row_match:
        replacement = row_match.group(1) + row_match.group(2) + fragment + row_match.group(3)
        return sheet_xml[:row_match.start()] + replacement + sheet_xml[row_match.end():]
    row_xml = f'<row r="{row_number}">{fragment}</row>'
    return sheet_xml.replace("</sheetData>", row_xml + "</sheetData>", 1)


def _inject_cell_contracts(xlsx_path: str, data: dict) -> str | None:
    """Restore formula caches, rich-text runs, phonetics, and explicit blank cells."""
    tmp = xlsx_path + ".~cell-contracts.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            workbook_xml = source.read("xl/workbook.xml").decode("utf-8")
            rels_xml = source.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_map = _xlsx_sheet_file_map(workbook_xml, rels_xml)
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as target:
                for item in source.infolist():
                    raw = source.read(item.filename)
                    sheet_data = next(
                        (sheet for sheet in data.get("sheets", []) if sheet_map.get(sheet.get("name")) == item.filename),
                        None,
                    )
                    if sheet_data is not None:
                        xml = raw.decode("utf-8")
                        for row_index, row in enumerate(sheet_data.get("rows", []), 1):
                            for col_index, cell_data in enumerate(row.get("cells", []), 1):
                                if cell_data.get("merge") == "slave":
                                    continue
                                if not (cell_data.get("formula") or cell_data.get("rich_text") or (cell_data.get("present") and cell_data.get("v") is None)):
                                    continue
                                from openpyxl.utils import get_column_letter
                                coord = f"{get_column_letter(col_index)}{row_index}"
                                match, existing = _find_cell_xml(xml, coord)
                                fragment = _cell_xml_fragment(coord, cell_data, existing)
                                if fragment == existing:
                                    continue
                                if match:
                                    xml = xml[:match.start()] + fragment + xml[match.end():]
                                elif fragment:
                                    xml = _insert_cell_xml(xml, coord, fragment)
                        raw = xml.encode("utf-8")
                    target.writestr(item, raw)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"cell semantic contract injection failed: {exc}"
    return None


def _current_cell_group_hashes(cell_data: dict) -> tuple[str, str, str]:
    content, style, structure = _cell_group_payloads(cell_data)
    return (
        _semantic_digest(content),
        _semantic_digest(style),
        _semantic_digest(structure),
    )


def _compare_content_only_models(data: dict, baseline_data: dict) -> dict[str, list[str]] | None:
    current_workbook = _semantic_digest({key: value for key, value in data.items() if key != "sheets"})
    baseline_workbook = _semantic_digest({key: value for key, value in baseline_data.items() if key != "sheets"})
    if current_workbook != baseline_workbook:
        return None

    baseline_sheets = {sheet.get("name"): sheet for sheet in baseline_data.get("sheets", [])}
    changes = {}
    for sheet in data.get("sheets", []):
        sheet_name = sheet.get("name")
        baseline_sheet = baseline_sheets.get(sheet_name)
        if baseline_sheet is None:
            return None
        if len(sheet.get("rows", [])) != len(baseline_sheet.get("rows", [])):
            return None
        current_sheet_meta = _semantic_digest({key: value for key, value in sheet.items() if key != "rows"})
        baseline_sheet_meta = _semantic_digest({key: value for key, value in baseline_sheet.items() if key != "rows"})
        if current_sheet_meta != baseline_sheet_meta:
            return None

        changed_cells = []
        for row_index, (row, baseline_row) in enumerate(
            zip(sheet.get("rows", []), baseline_sheet.get("rows", [])),
            1,
        ):
            cells = row.get("cells", [])
            baseline_cells = baseline_row.get("cells", [])
            if len(cells) != len(baseline_cells):
                return None
            current_row_meta = _semantic_digest({key: value for key, value in row.items() if key != "cells"})
            baseline_row_meta = _semantic_digest({key: value for key, value in baseline_row.items() if key != "cells"})
            if current_row_meta != baseline_row_meta:
                return None
            for col_index, (cell_data, baseline_cell) in enumerate(zip(cells, baseline_cells), 1):
                content_hash, style_hash, structure_hash = _current_cell_group_hashes(cell_data)
                baseline_content, baseline_style, baseline_structure = _current_cell_group_hashes(baseline_cell)
                if style_hash != baseline_style or structure_hash != baseline_structure:
                    return None
                if content_hash != baseline_content:
                    from openpyxl.utils import get_column_letter
                    changed_cells.append(f"{get_column_letter(col_index)}{row_index}")
        if changed_cells:
            changes[sheet_name] = changed_cells
    return changes


def _lazy_source_content_changes(data: dict) -> dict[str, list[str]] | None:
    source = data.get("source")
    try:
        baseline_data = serialize_excel(source, data.get("_sheet_filter"))
    except Exception:
        return None
    return _compare_content_only_models(data, baseline_data)


def _content_only_changes(data: dict) -> dict[str, list[str]] | None:
    lossless = data.get("_lossless") or {}
    source = data.get("source")
    if not source or not os.path.isfile(source):
        return None
    if _sha256_file(source) != lossless.get("source_sha256"):
        return None
    current_workbook = _semantic_digest({key: value for key, value in data.items() if key != "sheets"})
    if current_workbook != lossless.get("workbook_digest"):
        return None
    dirty = data.get("_dirty") or {}
    if dirty.get("workbook"):
        return None

    changes = {}
    baseline_cells_seen = False
    for sheet in data.get("sheets", []):
        sheet_name = sheet.get("name")
        sheet_dirty = (dirty.get("sheets") or {}).get(sheet_name)
        if sheet_dirty and not (isinstance(sheet_dirty, dict) and set(sheet_dirty) <= {"cells"}):
            return None
        if len(sheet.get("rows", [])) != sheet.get("_baseline_row_count"):
            return None
        current_sheet_meta = _semantic_digest({key: value for key, value in sheet.items() if key != "rows"})
        if current_sheet_meta != sheet.get("_baseline_meta_hash"):
            return None

        changed_cells = []
        for row_index, row in enumerate(sheet.get("rows", []), 1):
            if len(row.get("cells", [])) != row.get("_baseline_cell_count"):
                return None
            current_row_meta = _semantic_digest({key: value for key, value in row.items() if key != "cells"})
            if current_row_meta != row.get("_baseline_meta_hash"):
                return None
            for col_index, cell_data in enumerate(row.get("cells", []), 1):
                present = tuple(key in cell_data for key in _CELL_BASELINE_HASH_KEYS)
                if any(present) and not all(present):
                    return None
                if not all(present):
                    continue
                baseline_cells_seen = True
                baseline_content, baseline_style, baseline_structure = (
                    cell_data[key] for key in _CELL_BASELINE_HASH_KEYS
                )
                content_hash, style_hash, structure_hash = _current_cell_group_hashes(cell_data)
                if style_hash != baseline_style or structure_hash != baseline_structure:
                    return None
                if content_hash != baseline_content:
                    from openpyxl.utils import get_column_letter
                    changed_cells.append(f"{get_column_letter(col_index)}{row_index}")
        if changed_cells:
            changes[sheet_name] = changed_cells
    if baseline_cells_seen:
        return changes
    return _lazy_source_content_changes(data)


def _cell_inner_xml(fragment: str | None) -> str:
    if not fragment or fragment.rstrip().endswith("/>"):
        return ""
    match = re.match(r"<c\b[^>]*>(.*)</c>", fragment, re.DOTALL)
    return match.group(1) if match else ""


def _merge_generated_cell_content(source_cell: str | None, generated_cell: str | None) -> str | None:
    if source_cell is None:
        if generated_cell is None:
            return None
        generated_open = re.match(r"<c\b([^>]*)", generated_cell)
        generated_attrs = _parse_xml_attrs(generated_open.group(1)) if generated_open else {}
        generated_attrs.pop("s", None)
        generated_inner = _cell_inner_xml(generated_cell)
        return f"<c{_xml_attributes(generated_attrs)}>" + generated_inner + "</c>"
    if generated_cell is None:
        return source_cell
    source_open = re.match(r"<c\b([^>]*)", source_cell)
    generated_open = re.match(r"<c\b([^>]*)", generated_cell)
    source_attrs = _parse_xml_attrs(source_open.group(1)) if source_open else {}
    generated_attrs = _parse_xml_attrs(generated_open.group(1)) if generated_open else {}
    if "t" in generated_attrs:
        source_attrs["t"] = generated_attrs["t"]
    else:
        source_attrs.pop("t", None)
    generated_inner = _cell_inner_xml(generated_cell)
    source_inner = _cell_inner_xml(source_cell)
    source_extras = re.sub(
        r"<(?:f|v|is)\b[^>]*(?:/>|>.*?</(?:f|v|is)>)",
        "",
        source_inner,
        flags=re.DOTALL,
    )
    return f"<c{_xml_attributes(source_attrs)}>" + generated_inner + source_extras + "</c>"


def _reconstruct_content_only(data: dict, output_path: str, changes: dict[str, list[str]]) -> list[str] | None:
    source_path = str(data["source"])
    temp_output = str(output_path) + ".~saving.tmp"
    try:
        sheets_by_name = {sheet.get("name"): sheet for sheet in data.get("sheets", [])}
        with zipfile.ZipFile(source_path, "r") as source:
            source_wb = source.read("xl/workbook.xml").decode("utf-8")
            source_rels = source.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            source_sheet_map = _xlsx_sheet_file_map(source_wb, source_rels)
            source_part_to_sheet = {part: name for name, part in source_sheet_map.items()}
            with zipfile.ZipFile(temp_output, "w", zipfile.ZIP_DEFLATED) as target:
                for item in source.infolist():
                    raw = source.read(item.filename)
                    sheet_name = source_part_to_sheet.get(item.filename)
                    if sheet_name in changes:
                        sheet_data = sheets_by_name.get(sheet_name)
                        if not sheet_data:
                            return None
                        source_xml = raw.decode("utf-8")
                        for coord in changes[sheet_name]:
                            from openpyxl.utils.cell import coordinate_to_tuple
                            row_number, column_number = coordinate_to_tuple(coord)
                            rows = sheet_data.get("rows", [])
                            if row_number > len(rows):
                                return None
                            cells = rows[row_number - 1].get("cells", [])
                            if column_number > len(cells):
                                return None
                            cell_data = cells[column_number - 1]
                            source_match, source_cell = _find_cell_xml(source_xml, coord)
                            generated_cell = _cell_xml_fragment(
                                coord,
                                cell_data,
                                None,
                                include_scalar=True,
                            )
                            if (
                                source_cell is None
                                and cell_data.get("v") is None
                                and not cell_data.get("present")
                                and not cell_data.get("formula")
                                and not cell_data.get("rich_text")
                            ):
                                generated_cell = None
                            merged_cell = _merge_generated_cell_content(source_cell, generated_cell)
                            if source_match and merged_cell is not None:
                                source_xml = source_xml[:source_match.start()] + merged_cell + source_xml[source_match.end():]
                            elif merged_cell is not None:
                                source_xml = _insert_cell_xml(source_xml, coord, merged_cell)
                        raw = source_xml.encode("utf-8")
                    target.writestr(item, raw)
        _apply_package_edits(temp_output, data.get("_package_edits"))
        warnings = validate_xlsx(temp_output)
        os.replace(temp_output, str(output_path))
        return warnings
    finally:
        for candidate in (temp_output,):
            if os.path.exists(candidate):
                try:
                    os.remove(candidate)
                except OSError:
                    pass


def _atomic_copy_source_package(
    data: dict,
    output_path: str,
    *,
    require_session_digest: bool = True,
) -> list[str] | None:
    """Return warnings after an exact source copy, or None when the fast path is unsafe."""
    lossless = data.get("_lossless") or {}
    source = data.get("source")
    if not source or not os.path.isfile(source):
        return None
    expected_source_hash = lossless.get("source_sha256")
    if not expected_source_hash or _sha256_file(source) != expected_source_hash:
        return None
    if require_session_digest and _semantic_digest(data) != lossless.get("session_digest"):
        return None

    out_str = str(output_path)
    tmp_out = out_str + ".~saving.tmp"
    try:
        shutil.copyfile(source, tmp_out)
        _apply_package_edits(tmp_out, data.get("_package_edits"))
        warnings = validate_xlsx(tmp_out)
        os.replace(tmp_out, out_str)
        return warnings
    except Exception:
        if os.path.exists(tmp_out):
            try:
                os.remove(tmp_out)
            except OSError:
                pass
        raise


# ── Serialize ─────────────────────────────────────────────────────────────────

def serialize_excel(uri: str, sheet_name: str | None = None) -> dict:
    """
    Serialize an Excel file to a metadata dict.

    Each cell carries: value, fill RGB, bold/italic/size/font-color,
    wrap/halign/valign, number format, merge info, border sides.
    Sheet carries: freeze_panes ref, data_validations.
    Merge origin → {rowspan, colspan, r1,c1,r2,c2}.
    Merge slave  → "slave" (skipped during reconstruct).
    """
    import openpyxl

    path = uri_to_path(uri)
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    wb = openpyxl.load_workbook(str(path), keep_vba=keep_vba, rich_text=True)
    named_style_names = _snapshot_named_style_names(wb)
    raw_theme = getattr(wb, "loaded_theme", None)
    theme_xml = None
    if raw_theme:
        import base64
        if isinstance(raw_theme, str):
            raw_theme = raw_theme.encode("utf-8")
        theme_xml = base64.b64encode(raw_theme).decode("ascii")

    # Extract theme colors once — used to resolve all theme-type cell colors
    theme_colors = _wb_theme_colors(wb)
    named_styles = _serialize_named_styles(wb)

    try:
        import zipfile as _zf
        with _zf.ZipFile(str(path), "r") as _z:
            _wb_xml = _z.read("xl/workbook.xml").decode("utf-8")
            _rels_xml = _z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        _sfm = _xlsx_sheet_file_map(_wb_xml, _rels_xml)
    except Exception:
        _sfm = {}
    raw_fill_data = _extract_raw_fill_data(path, _sfm)
    raw_sheet_views = _extract_sheet_view_attrs(path, _sfm)
    raw_row_attrs = _extract_row_attrs(path, _sfm)
    raw_sheet_formats = _extract_sheet_format_data(path, _sfm)
    raw_data_validations = _extract_data_validations_xml(path, _sfm)
    raw_sheet_extensions = _extract_worksheet_xml_blocks(path, _sfm, "extLst")
    raw_ignored_errors = _extract_worksheet_xml_blocks(path, _sfm, "ignoredErrors")
    raw_comment_vml = _extract_comment_vml(path, _sfm)
    sheet_passthrough_relationships = _extract_sheet_passthrough_relationships(path, _sfm)
    ooxml_semantics = _extract_ooxml_semantics(path, _sfm)
    raw_cell_semantics = ooxml_semantics["cells"]
    style_semantics = ooxml_semantics["styles"]

    names = [sheet_name] if sheet_name else wb.sheetnames
    sheets = []
    cell_style_cache: dict[tuple[int, int], dict] = {}

    for sname in names:
        if sname not in wb.sheetnames:
            raise ValueError(f"Sheet '{sname}' not found. Available: {wb.sheetnames}")
        ws = wb[sname]
        raw_fills_for_sheet = raw_fill_data.get(sname, {})
        raw_cells_for_sheet = raw_cell_semantics.get(sname, {})
        implicit_cell_defaults = None

        # Build merge map
        merged_map: dict = {}
        for mg in ws.merged_cells.ranges:
            merged_map[(mg.min_row, mg.min_col)] = {
                "r1": mg.min_row - 1, "c1": mg.min_col - 1,  # 0-based
                "r2": mg.max_row - 1, "c2": mg.max_col - 1,  # 0-based
                "rowspan": mg.max_row - mg.min_row + 1,
                "colspan": mg.max_col - mg.min_col + 1,
            }
            for r in range(mg.min_row, mg.max_row + 1):
                for c in range(mg.min_col, mg.max_col + 1):
                    if not (r == mg.min_row and c == mg.min_col):
                        merged_map[(r, c)] = "slave"

        rows = []
        for row in ws.iter_rows():
            rh = ws.row_dimensions[row[0].row].height
            cells = []
            for cell in row:
                mi = merged_map.get((cell.row, cell.column), {})
                raw_cell = raw_cells_for_sheet.get(cell.coordinate, {})
                if not raw_cell and not mi:
                    if implicit_cell_defaults is None:
                        style_id = int(cell.style_id or 0)
                        style_key = _style_cache_key(cell, style_id)
                        cached_style = cell_style_cache.get(style_key)
                        if cached_style is None:
                            cached_style = _build_cached_cell_style(
                                cell,
                                style_id,
                                style_semantics,
                                named_style_names,
                                theme_colors,
                            )
                            cell_style_cache[style_key] = cached_style
                        implicit_cell_defaults = _implicit_cell_defaults_from_cached_style(cell, cached_style)
                    cells.append(_implicit_cell_placeholder())
                    continue
                if raw_cell.get("rich_text"):
                    cell_value = raw_cell["rich_text"]["text"]
                elif raw_cell.get("formula"):
                    cell_value = "=" + raw_cell["formula"]["text"]
                else:
                    cell_value = cell.value

                try:
                    style_id = int(raw_cell.get("cell_attrs", {}).get("s", cell.style_id or 0))
                except (TypeError, ValueError):
                    style_id = int(cell.style_id or 0)
                style_key = _style_cache_key(cell, style_id)
                cached_style = cell_style_cache.get(style_key)
                if cached_style is None:
                    cached_style = _build_cached_cell_style(
                        cell,
                        style_id,
                        style_semantics,
                        named_style_names,
                        theme_colors,
                    )
                    cell_style_cache[style_key] = cached_style
                fill_raw = _cached_fill_raw(raw_fills_for_sheet.get(cell.coordinate), cached_style)
                cell_data = {
                    **cached_style["data"],
                    "v": cell_value,
                    "merge": mi,
                    "data_type": raw_cell.get("cell_attrs", {}).get("t") or cell.data_type,
                    "present": bool(raw_cell.get("present")),
                    "formula": copy.deepcopy(raw_cell.get("formula")),
                    "rich_text": copy.deepcopy(raw_cell.get("rich_text")),
                    "cell_attrs": copy.deepcopy(raw_cell.get("cell_attrs") or {}),
                }
                # Disambiguate literal text that LOOKS like a formula ("=…"):
                # without this marker a text cell would silently turn into a
                # broken formula on reconstruct.
                if (isinstance(cell_value, str) and cell_value.startswith("=")
                        and cell.data_type != "f" and not raw_cell.get("formula")):
                    cell_data["dt"] = "s"
                if fill_raw:
                    cell_data["_fill_raw"] = fill_raw
                if cached_style["font_raw"]:
                    cell_data["_font_raw"] = cached_style["font_raw"]
                cells.append(cell_data)
            rd = ws.row_dimensions[row[0].row]
            rows.append({
                "h":       rh,
                "hidden":  bool(rd.hidden),
                "outline": rd.outlineLevel or 0,
                "cells":   cells,
            })

        for row_number, row_data in enumerate(rows, 1):
            row_attrs = copy.deepcopy((raw_row_attrs.get(sname) or {}).get(str(row_number)) or {})
            if row_attrs:
                row_data["_row_attrs"] = row_attrs
            for raw_key, public_key in (
                ("collapsed", "collapsed"), ("thickTop", "thickTop"),
                ("thickBot", "thickBot"), ("customFormat", "customFormat"),
                ("customHeight", "customHeight"), ("ph", "phonetic"),
            ):
                if raw_key in row_attrs:
                    row_data[public_key] = str(row_attrs[raw_key]).lower() in _XML_TRUE_VALUES
            if "s" in row_attrs:
                try:
                    row_data["style"] = int(row_attrs["s"])
                except (TypeError, ValueError):
                    row_data["style"] = row_attrs["s"]

        sv = ws.sheet_view
        col_widths, col_hidden, col_outline = _ser_column_dimensions(ws)
        col_state = _dimension_state(col_widths, col_hidden, col_outline)

        # Tab color
        tc = None
        if ws.sheet_properties and ws.sheet_properties.tabColor:
            tc = _resolve_color(ws.sheet_properties.tabColor, theme_colors)

        # Print settings
        ps = ws.page_setup
        pm = ws.page_margins
        def _safe(obj, attr):
            try: return getattr(obj, attr)
            except Exception: return None
        # fitToPage lives in sheetPr/pageSetUpPr, not in pageSetup
        fit_to_page = None
        try:
            pspr = ws.sheet_properties.pageSetUpPr
            fit_to_page = pspr.fitToPage if pspr else None
        except Exception:
            pass
        page_setup = {k: v for k, v in {
            "orientation": _safe(ps, "orientation"),
            "paperSize":   _safe(ps, "paperSize"),
            "fitToPage":   fit_to_page,
            "fitToWidth":  _safe(ps, "fitToWidth"),
            "fitToHeight": _safe(ps, "fitToHeight"),
            "scale":       _safe(ps, "scale"),
        }.items() if v is not None}
        page_margins = {k: getattr(pm, k) for k in
                        ("left", "right", "top", "bottom", "header", "footer")
                        if getattr(pm, k, None) is not None}

        # Sheet protection
        prot = ws.protection
        protection = None
        if prot.sheet:
            protection = {
                key: getattr(prot, key)
                for key in getattr(type(prot), "__attrs__", ())
                if key != "sheet" and getattr(prot, key, None) is not None
            }
            if protection.get("password") is not None:
                protection["password_is_hashed"] = True

        # Print titles
        print_titles = None
        ptr = ws.print_title_rows
        ptc = ws.print_title_cols
        if ptr or ptc:
            print_titles = {"rows": ptr, "cols": ptc}

        # Print area
        print_area = str(ws.print_area) if ws.print_area else None

        # Header / footer
        hf_data = {}
        try:
            oh = ws.oddHeader
            if oh:
                if oh.left   and oh.left.text:   hf_data["hl"] = oh.left.text
                if oh.center and oh.center.text: hf_data["hc"] = oh.center.text
                if oh.right  and oh.right.text:  hf_data["hr"] = oh.right.text
            of_ = ws.oddFooter
            if of_:
                if of_.left   and of_.left.text:   hf_data["fl"] = of_.left.text
                if of_.center and of_.center.text: hf_data["fc"] = of_.center.text
                if of_.right  and of_.right.text:  hf_data["fr"] = of_.right.text
        except Exception:
            pass

        # Hyperlinks (external target and/or internal location like "Sheet2!A1")
        hyperlinks = {}
        for _row in ws.iter_rows():
            for _cell in _row:
                if _cell.hyperlink:
                    hl = _cell.hyperlink
                    target = getattr(hl, "target", None) or None
                    location = getattr(hl, "location", None) or None
                    if target or location:
                        hyperlinks[_cell.coordinate] = {
                            "target":   target,
                            "location": location,
                            "display":  getattr(hl, "display", None),
                            "tooltip":  getattr(hl, "tooltip", None),
                        }

        # Comments
        comments = {}
        for _row in ws.iter_rows():
            for _cell in _row:
                if _cell.comment:
                    comments[_cell.coordinate] = {
                        "text":   _cell.comment.text or "",
                        "author": _cell.comment.author or "",
                    }

        # Tables
        tables = []
        try:
            for _t in ws.tables.values():
                ts = _t.tableStyleInfo
                tables.append(_serialize_table(_t) or {
                    "name": _t.displayName,
                    "ref":  _t.ref,
                    "style": {
                        "name":            ts.name            if ts else None,
                        "showRowStripes":  ts.showRowStripes  if ts else None,
                        "showColStripes":  ts.showColumnStripes if ts else None,
                        "showFirstCol":    ts.showFirstColumn  if ts else None,
                        "showLastCol":     ts.showLastColumn   if ts else None,
                    } if ts else None,
                })
        except Exception:
            pass

        worksheet_semantics = copy.deepcopy(ooxml_semantics["worksheets"].get(sname))
        worksheet_models = _worksheet_models_from_semantics(worksheet_semantics)
        sheets.append({
            "name":          sname,
            "cw":            col_widths,
            "ch":            col_hidden,
            "co":            col_outline or None,
            "rows":          rows,
            "_implicit_cell_defaults": implicit_cell_defaults or copy.deepcopy(_IMPLICIT_CELL_DEFAULTS),
            "freeze":        ws.freeze_panes,
            "validations":   _ser_validations(ws),
            "sheet_view":    _serialize_sheet_view(sv, raw_sheet_views.get(sname)),
            "tab_color":     tc,
            "auto_filter":   str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
            "auto_filter_model": _serialize_auto_filter(ws.auto_filter),
            "page_setup":    page_setup or None,
            "page_margins":  page_margins or None,
            "protection":    protection,
            "print_titles":  print_titles,
            "print_area":    print_area,
            "header_footer": hf_data or None,
            "hyperlinks":    hyperlinks or None,
            "comments":      comments or None,
            "tables":        tables or None,
            "state":         ws.sheet_state,
            "worksheet_semantics": worksheet_semantics,
            "_dirty":        [],
        })
        sheets[-1].update(worksheet_models)
        if worksheet_models.get("sheet_views") is not None:
            sheets[-1]["sheet_view"] = copy.deepcopy(
                worksheet_models["sheet_views"][0] if worksheet_models["sheet_views"] else {}
            )

        raw_sheet_xml = raw_sheet_formats.get(sname) or {}
        if raw_sheet_xml.get("root_attrs") or raw_sheet_xml.get("sheetFormatPr"):
            sheets[-1]["_sheet_format_pr"] = {
                "root_attrs": raw_sheet_xml.get("root_attrs", ""),
                "sheetFormatPr": raw_sheet_xml.get("sheetFormatPr"),
            }
        if raw_sheet_xml.get("cols"):
            sheets[-1]["_cols_raw"] = {
                "xml": raw_sheet_xml["cols"],
                "state": col_state,
            }
        if raw_data_validations.get(sname):
            sheets[-1]["data_validations_xml"] = raw_data_validations[sname]
        if raw_sheet_extensions.get(sname):
            sheets[-1]["_worksheet_ext_xml"] = raw_sheet_extensions[sname]
        if raw_ignored_errors.get(sname):
            sheets[-1]["ignored_errors"] = _parse_ignored_errors_xml(raw_ignored_errors[sname])
        if raw_comment_vml.get(sname):
            sheets[-1]["_comment_vml"] = raw_comment_vml[sname]
        if sheet_passthrough_relationships.get(sname):
            sheets[-1]["passthrough_relationships"] = sheet_passthrough_relationships[sname]

    # Named ranges (workbook level + worksheet level, including built-ins).
    #
    # openpyxl's reader routes definedName elements three ways (see
    # WorkbookParser.assign_names): global-scope names land in
    # wb.defined_names; ordinary (non-reserved) worksheet-scoped names land in
    # that sheet's own ws.defined_names (NOT wb.defined_names); and the three
    # reserved built-ins (_xlnm.Print_Area / _xlnm.Print_Titles /
    # _xlnm._FilterDatabase) are consumed entirely into ws.print_area /
    # ws.print_title_rows+cols / ws.auto_filter and removed from every
    # defined-names collection outright. Only scanning wb.defined_names (as
    # before) silently dropped every worksheet-scoped name -- built-in or not.
    _DN_META_FIELDS = (
        "comment", "customMenu", "description", "help", "statusBar",
        "hidden", "function", "vbProcedure", "xlm", "functionGroupId",
        "shortcutKey", "publishToServer", "workbookParameter",
    )

    def _defined_name_record(dn, sheet_id_override=None) -> dict:
        record = {
            "name": dn.name,
            "value": dn.attr_text,
            "sheet_id": sheet_id_override if sheet_id_override is not None else dn.localSheetId,
        }
        for field in _DN_META_FIELDS:
            value = getattr(dn, field, None)
            if value is not None:
                record[field] = value
        return record

    named_ranges = []
    try:
        for name in wb.defined_names:
            named_ranges.append(_defined_name_record(wb.defined_names[name]))
        for idx, ws_scan in enumerate(wb.worksheets):
            for name in ws_scan.defined_names:
                named_ranges.append(_defined_name_record(ws_scan.defined_names[name], sheet_id_override=idx))
            # Resynthesize the reserved built-ins openpyxl consumed above so
            # they still round-trip through excel_list_defined_names /
            # excel_add_defined_name's duplicate-classification as built-ins,
            # instead of silently disappearing after a save+reload.
            if ws_scan.print_area:
                named_ranges.append({
                    "name": "_xlnm.Print_Area", "value": str(ws_scan.print_area),
                    "sheet_id": idx, "builtin": True,
                })
            if ws_scan._print_rows or ws_scan._print_cols:
                named_ranges.append({
                    "name": "_xlnm.Print_Titles", "value": str(ws_scan.print_titles),
                    "sheet_id": idx, "builtin": True,
                })
            if ws_scan.auto_filter.ref:
                from openpyxl.utils.cell import quote_sheetname
                named_ranges.append({
                    "name": "_xlnm._FilterDatabase",
                    "value": f"{quote_sheetname(ws_scan.title)}!{ws_scan.auto_filter.ref}",
                    "sheet_id": idx, "builtin": True, "hidden": True,
                })
    except Exception:
        pass

    # Extract conditional formatting as raw XML for passthrough
    cf_xml = _extract_cf_xml(path, [sd["name"] for sd in sheets])
    dxfs_xml = cf_xml.pop("__dxfs__", None)
    for sd in sheets:
        if sd["name"] in cf_xml:
            sd["cf_xml"] = cf_xml[sd["name"]]

    # Extract drawing/chart/image data for passthrough
    drawing_data = _extract_drawing_data(path, _sfm)
    shape_inventory = _extract_shape_inventory(drawing_data)
    for sd in sheets:
        if sd["name"] in drawing_data:
            sd["drawing_data"] = drawing_data[sd["name"]]
        if sd["name"] in shape_inventory:
            sd["shapes"] = shape_inventory[sd["name"]]

    # Document properties (docProps/core.xml)
    doc_props = {}
    try:
        props = wb.properties
        for attr in ("creator", "title", "subject", "description",
                     "keywords", "category", "lastModifiedBy",
                     "contentStatus", "identifier", "language",
                     "revision", "version"):
            value = getattr(props, attr, None)
            if value:
                doc_props[attr] = value
        if props.created:
            doc_props["created"] = props.created.isoformat()
        if props.modified:
            doc_props["modified"] = props.modified.isoformat()
        if props.lastPrinted:
            doc_props["lastPrinted"] = props.lastPrinted.isoformat()
    except Exception:
        pass

    # Workbook view (active tab, window geometry) -- kept for backward
    # compatibility; workbook_views below carries the FULL ordered list.
    wb_view = {}
    try:
        view = wb.views[0]
        for attr in getattr(type(view), "__attrs__", ()):
            value = getattr(view, attr, None)
            if value is not None:
                wb_view[attr] = value
    except Exception:
        pass

    # Workbook views (bookViews) -- the full ordered list, not just the first.
    workbook_views = []
    try:
        for _view_idx, _view in enumerate(wb.views):
            entry = {"index": _view_idx}
            for attr in getattr(type(_view), "__attrs__", ()):
                value = getattr(_view, attr, None)
                if value is not None:
                    entry[attr] = value
            workbook_views.append(entry)
    except Exception:
        pass

    # Calculation properties (calcPr) -- openpyxl round-trips every field
    # natively via wb.calculation.
    calculation_properties = {}
    try:
        calc = wb.calculation
        for attr in getattr(type(calc), "__attrs__", ()):
            value = getattr(calc, attr, None)
            if value is not None:
                calculation_properties[attr] = value
    except Exception:
        pass

    # Workbook protection (workbookProtection) -- openpyxl round-trips these
    # natively via wb.security, including already-hashed legacy passwords
    # and the modern hash/salt/spin-count fields.
    workbook_protection = {}
    try:
        sec = wb.security
        for attr in getattr(type(sec), "__attrs__", ()):
            value = getattr(sec, attr, None)
            if value is not None:
                workbook_protection[attr] = value
    except Exception:
        pass

    # Workbook properties (workbookPr) -- codeName/date1904 round-trip
    # natively through openpyxl's object model; everything else
    # (filterPrivacy, saveExternalLinkValues, showObjects, updateLinks, ...)
    # has no read/write hook on the Workbook object at all, so fall back to
    # the raw attrs already captured by the semantic-snapshot extractor above.
    workbook_properties = {}
    try:
        _WBPR_BOOL_ATTRS = {
            "date1904", "dateCompatibility", "showBorderUnselectedTables", "filterPrivacy",
            "promptedSolutions", "showInkAnnotation", "backupFile", "saveExternalLinkValues",
            "hidePivotFieldList", "showPivotChartFilter", "allowRefreshQuery", "publishItems",
            "checkCompatibility", "autoCompressPictures", "refreshAllConnections",
        }
        raw_wb_pr_node = (ooxml_semantics.get("workbook") or {}).get("workbook_properties") or {}
        for key, value in (raw_wb_pr_node.get("attrs") or {}).items():
            if key in _WBPR_BOOL_ATTRS:
                workbook_properties[key] = str(value).strip().lower() not in {"0", "false"}
            else:
                workbook_properties[key] = value
        if wb.code_name:
            workbook_properties["codeName"] = wb.code_name
        from openpyxl.utils.datetime import CALENDAR_MAC_1904 as _CAL_1904
        workbook_properties["date1904"] = bool(wb.excel_base_date == _CAL_1904)
    except Exception:
        pass

    # Custom document properties (docProps/custom.xml), typed.
    custom_doc_props = []
    try:
        for prop in wb.custom_doc_props:
            value = prop.value
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            custom_doc_props.append({"name": prop.name, "type": type(prop).__name__, "value": value})
    except Exception:
        pass

    # App (extended) properties (docProps/app.xml) -- openpyxl has no object
    # model for these at all; read the raw part directly.
    app_props = {}
    try:
        import zipfile as _zf_app
        with _zf_app.ZipFile(str(path), "r") as _z_app:
            if "docProps/app.xml" in _z_app.namelist():
                app_root = ET.fromstring(_z_app.read("docProps/app.xml"))
                for child in app_root:
                    tag = _local_name(child.tag)
                    if child.text and not list(child):
                        app_props[tag] = child.text
    except Exception:
        pass

    # Release the zip handle now — openpyxl workbooks have reference cycles,
    # so waiting for GC can leave the file locked on Windows.
    _close_openpyxl_workbook(wb)

    for sheet in sheets:
        for row in sheet.get("rows", []):
            row["_baseline_meta_hash"] = _semantic_digest({
                key: value for key, value in row.items() if key != "cells"
            })
            row["_baseline_cell_count"] = len(row.get("cells", []))
        sheet["_baseline_meta_hash"] = _semantic_digest({
            key: value for key, value in sheet.items() if key != "rows"
        })
        sheet["_baseline_row_count"] = len(sheet.get("rows", []))

    result = {"source": str(path), "sheets": sheets, "named_ranges": named_ranges,
              "named_styles": named_styles,
              "dxfs_xml": dxfs_xml, "theme_xml": theme_xml,
              "doc_props": doc_props or None, "wb_view": wb_view or None,
              "workbook_views": workbook_views or None,
              "calculation_properties": calculation_properties,
              "workbook_protection": workbook_protection,
              "workbook_properties": workbook_properties,
              "custom_doc_props": custom_doc_props,
              "app_props": app_props,
              "workbook_semantics": ooxml_semantics["workbook"],
              "style_semantics": style_semantics,
              "table_semantics": ooxml_semantics["tables"]}
    result["_baseline_workbook_properties"] = copy.deepcopy(workbook_properties)
    result["_lossless"] = {
        "version": 2,
        "source_sha256": _sha256_file(path),
        "session_digest": _semantic_digest(result),
        "workbook_digest": _semantic_digest({key: value for key, value in result.items() if key != "sheets"}),
        "package_graph": _extract_package_graph(path),
        "sheet_parts": _sfm,
    }
    result["_dirty"] = {"workbook": [], "sheets": {}}
    return result


# ── Reconstruct ───────────────────────────────────────────────────────────────

_TABLE_ATTRS = (
    "id", "name", "displayName", "comment", "ref", "tableType",
    "headerRowCount", "insertRow", "insertRowShift", "totalsRowCount",
    "totalsRowShown", "published", "headerRowDxfId", "dataDxfId",
    "totalsRowDxfId", "headerRowBorderDxfId", "tableBorderDxfId",
    "totalsRowBorderDxfId", "headerRowCellStyle", "dataCellStyle",
    "totalsRowCellStyle", "connectionId",
)
_TABLE_COLUMN_ATTRS = (
    "id", "uniqueName", "name", "totalsRowFunction", "totalsRowLabel",
    "queryTableFieldId", "headerRowDxfId", "dataDxfId", "totalsRowDxfId",
    "headerRowCellStyle", "dataCellStyle", "totalsRowCellStyle",
)


def _table_formula_model(value):
    if value is None:
        return None
    text = getattr(value, "attr_text", None)
    array = getattr(value, "array", None)
    if array is None:
        return text
    return {"text": text, "array": array}


def _serialize_table(table) -> dict:
    result = {
        key: getattr(table, key)
        for key in _TABLE_ATTRS
        if getattr(table, key, None) is not None
    }
    result["columns"] = []
    for column in table.tableColumns or []:
        item = {
            key: getattr(column, key)
            for key in _TABLE_COLUMN_ATTRS
            if getattr(column, key, None) is not None
        }
        calculated = _table_formula_model(column.calculatedColumnFormula)
        totals = _table_formula_model(column.totalsRowFormula)
        if calculated is not None:
            item["calculatedColumnFormula"] = calculated
        if totals is not None:
            item["totalsRowFormula"] = totals
        result["columns"].append(item)
    result["auto_filter"] = _serialize_auto_filter(table.autoFilter)
    if table.sortState is not None:
        result["sort_state"] = {
            "ref": table.sortState.ref,
            "conditions": [
                {
                    key: getattr(condition, key)
                    for key in ("ref", "descending", "sortBy", "customList", "dxfId", "iconSet", "iconId")
                    if getattr(condition, key, None) is not None
                }
                for condition in table.sortState.sortCondition or []
            ],
        }
    style = table.tableStyleInfo
    if style is not None:
        result["style"] = {
            "name": style.name,
            "showRowStripes": style.showRowStripes,
            "showColStripes": style.showColumnStripes,
            "showFirstCol": style.showFirstColumn,
            "showLastCol": style.showLastColumn,
        }
    return result


def _table_formula(value):
    if value is None:
        return None
    from openpyxl.worksheet.table import TableFormula
    if isinstance(value, dict):
        return TableFormula(array=value.get("array"), attr_text=value.get("text", value.get("attr_text")))
    return TableFormula(attr_text=str(value))


def _build_table(table_data: dict, table_id: int):
    from openpyxl.worksheet.filters import AutoFilter
    from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

    kwargs = {
        key: table_data[key]
        for key in _TABLE_ATTRS
        if key in table_data and table_data[key] is not None
    }
    kwargs["id"] = int(kwargs.get("id") or table_id)
    kwargs["displayName"] = table_data.get("displayName") or table_data.get("name")
    kwargs["name"] = table_data.get("name") or kwargs["displayName"]
    kwargs["ref"] = table_data["ref"]
    table = Table(**kwargs)

    columns = []
    for index, item in enumerate(table_data.get("columns") or [], 1):
        column_kwargs = {
            key: item[key]
            for key in _TABLE_COLUMN_ATTRS
            if key in item and item[key] is not None
        }
        column_kwargs["id"] = int(column_kwargs.get("id") or index)
        column_kwargs.setdefault("name", f"Column{index}")
        column_kwargs["calculatedColumnFormula"] = _table_formula(item.get("calculatedColumnFormula"))
        column_kwargs["totalsRowFormula"] = _table_formula(item.get("totalsRowFormula"))
        columns.append(TableColumn(**column_kwargs))
    if columns:
        table.tableColumns = columns

    filter_model = table_data.get("auto_filter")
    if filter_model:
        target = type("_FilterTarget", (), {})()
        target.auto_filter = AutoFilter()
        _apply_auto_filter_model(target, filter_model)
        table.autoFilter = target.auto_filter
    sort_model = table_data.get("sort_state")
    if sort_model:
        from openpyxl.worksheet.filters import SortCondition, SortState
        table.sortState = SortState(
            ref=sort_model.get("ref"),
            sortCondition=[SortCondition(**item) for item in sort_model.get("conditions") or []],
        )

    style = table_data.get("style")
    if style:
        def flag(short_key, long_key, default=False):
            if long_key in style:
                return bool(style.get(long_key))
            return bool(style.get(short_key, default))

        table.tableStyleInfo = TableStyleInfo(
            name=style.get("name"),
            showFirstColumn=flag("showFirstCol", "showFirstColumn"),
            showLastColumn=flag("showLastCol", "showLastColumn"),
            showRowStripes=flag("showRowStripes", "showRowStripes", True),
            showColumnStripes=flag("showColStripes", "showColumnStripes"),
        )
    return table


def _serialize_auto_filter(auto_filter) -> dict | None:
    if auto_filter is None or not auto_filter.ref:
        return None
    model = {"ref": str(auto_filter.ref), "filter_columns": [], "sort_state": None}
    for column in auto_filter.filterColumn or []:
        item = {
            key: getattr(column, key)
            for key in ("colId", "hiddenButton", "showButton")
            if getattr(column, key, None) is not None
        }
        if column.filters is not None:
            item["filters"] = list(column.filters.filter or [])
            if column.filters.blank is not None:
                item["blank"] = column.filters.blank
            if column.filters.calendarType is not None:
                item["calendarType"] = column.filters.calendarType
        if column.customFilters is not None:
            item["custom_filters"] = [
                {key: getattr(entry, key) for key in ("operator", "val") if getattr(entry, key, None) is not None}
                for entry in column.customFilters.customFilter or []
            ]
            item["and"] = column.customFilters.and_
        for source, target in (
            ("top10", "top10"), ("dynamicFilter", "dynamic_filter"),
            ("colorFilter", "color_filter"), ("iconFilter", "icon_filter"),
        ):
            value = getattr(column, source, None)
            if value is not None:
                item[target] = {
                    key: getattr(value, key)
                    for key in getattr(type(value), "__attrs__", ())
                    if getattr(value, key, None) is not None
                }
        model["filter_columns"].append(item)
    if auto_filter.sortState is not None:
        sort_state = {
            key: getattr(auto_filter.sortState, key)
            for key in ("columnSort", "caseSensitive", "sortMethod", "ref")
            if getattr(auto_filter.sortState, key, None) is not None
        }
        sort_state["conditions"] = [
            {
                key: getattr(condition, key)
                for key in ("ref", "descending", "sortBy", "customList", "dxfId", "iconSet", "iconId")
                if getattr(condition, key, None) is not None
            }
            for condition in auto_filter.sortState.sortCondition or []
        ]
        model["sort_state"] = sort_state
    return model


def _apply_auto_filter_model(ws, model: dict | None) -> None:
    if not model:
        return
    from openpyxl.worksheet.filters import (
        ColorFilter, CustomFilter, CustomFilters, DynamicFilter, FilterColumn,
        Filters, IconFilter, SortCondition, SortState, Top10,
    )

    ws.auto_filter.ref = model.get("ref")
    columns = []
    for entry in model.get("filter_columns") or []:
        kwargs = {
            key: entry[key]
            for key in ("hiddenButton", "showButton")
            if key in entry and entry[key] is not None
        }
        kwargs["colId"] = int(entry.get("colId", 0))
        filters = entry.get("filters")
        if filters is not None:
            if isinstance(filters, dict):
                values = filters.get("values", filters.get("filter", []))
                blank = filters.get("blank")
                calendar_type = filters.get("calendarType")
            else:
                values = filters
                blank = entry.get("blank")
                calendar_type = entry.get("calendarType")
            kwargs["filters"] = Filters(
                blank=blank,
                calendarType=calendar_type,
                filter=[str(value) for value in values or []],
            )
        custom_filters = entry.get("custom_filters") or entry.get("customFilters")
        if custom_filters is not None:
            kwargs["customFilters"] = CustomFilters(
                and_=entry.get("and", entry.get("and_")),
                customFilter=[CustomFilter(**item) for item in custom_filters],
            )
        for source, target, cls in (
            ("top10", "top10", Top10),
            ("dynamic_filter", "dynamicFilter", DynamicFilter),
            ("color_filter", "colorFilter", ColorFilter),
            ("icon_filter", "iconFilter", IconFilter),
        ):
            if isinstance(entry.get(source), dict):
                kwargs[target] = cls(**entry[source])
        columns.append(FilterColumn(**kwargs))
    ws.auto_filter.filterColumn = columns

    sort_model = model.get("sort_state")
    if sort_model:
        sort_kwargs = {
            key: sort_model[key]
            for key in ("columnSort", "caseSensitive", "sortMethod", "ref")
            if key in sort_model and sort_model[key] is not None
        }
        sort_kwargs["sortCondition"] = [
            SortCondition(**condition)
            for condition in sort_model.get("conditions", sort_model.get("sortCondition", []))
        ]
        ws.auto_filter.sortState = SortState(**sort_kwargs)


def _ignored_errors_xml(rules: list[dict]) -> str:
    children = []
    for rule in rules:
        sqref = rule.get("sqref")
        if isinstance(sqref, list):
            sqref = " ".join(str(value) for value in sqref)
        if not sqref:
            continue
        attrs = {"sqref": str(sqref)}
        for key, value in rule.items():
            if key == "sqref" or value is None:
                continue
            attrs[key] = "1" if bool(value) else "0"
        children.append("<ignoredError" + "".join(
            f" {key}={quoteattr(value)}" for key, value in attrs.items()
        ) + "/>")
    return "<ignoredErrors>" + "".join(children) + "</ignoredErrors>" if children else ""


def _inject_worksheet_semantics_legacy(xlsx_path: str, data: dict) -> str | None:
    requested = {}
    for sheet in data.get("sheets") or []:
        ignored_present = "ignored_errors" in sheet
        extension_xml = sheet.get("_worksheet_ext_xml")
        if ignored_present or extension_xml:
            requested[sheet.get("name")] = {
                "ignored": _ignored_errors_xml(sheet.get("ignored_errors") or []) if ignored_present else None,
                "extension": extension_xml,
            }
    if not requested:
        return None

    tmp = xlsx_path + ".~worksheet-semantics.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            workbook_xml = source.read("xl/workbook.xml").decode("utf-8")
            rels_xml = source.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_map = _xlsx_sheet_file_map(workbook_xml, rels_xml)
            file_map = {sheet_map[name]: value for name, value in requested.items() if name in sheet_map}
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as target:
                for item in source.infolist():
                    raw = source.read(item.filename)
                    patch = file_map.get(item.filename)
                    if patch is not None:
                        content = raw.decode("utf-8")
                        if patch["ignored"] is not None:
                            content = re.sub(
                                r"<(?:[A-Za-z_][\w.-]*:)?ignoredErrors\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?ignoredErrors>)",
                                "",
                                content,
                                count=1,
                                flags=re.DOTALL,
                            )
                            if patch["ignored"]:
                                anchor = re.search(
                                    r"<(?:smartTags|drawing\b|legacyDrawing|legacyDrawingHF|picture\b|oleObjects|controls|webPublishItems|tableParts|extLst)\b",
                                    content,
                                )
                                position = anchor.start() if anchor else content.rfind("</worksheet>")
                                content = content[:position] + patch["ignored"] + content[position:]
                        if patch["extension"]:
                            content = re.sub(
                                r"<(?:[A-Za-z_][\w.-]*:)?extLst\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?extLst>)",
                                "",
                                content,
                                count=1,
                                flags=re.DOTALL,
                            )
                            position = content.rfind("</worksheet>")
                            content = content[:position] + patch["extension"] + content[position:]
                        raw = content.encode("utf-8")
                    target.writestr(item, raw)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"worksheet semantic injection failed: {exc}"
    return None


def _worksheet_attr_value(value) -> str:
    if isinstance(value, bool):
        return "1" if value else "0"
    return str(value)


def _worksheet_xml_element(tag: str, model: dict | None, *, excluded: set[str] | None = None) -> str | None:
    if not isinstance(model, dict):
        return None
    skipped = {"present", *(excluded or set())}
    attrs = []
    for key, value in model.items():
        if key in skipped or key.startswith("_") or value is None or isinstance(value, (dict, list)):
            continue
        xml_key = "r:id" if key == "id" else key
        attrs.append(f" {xml_key}={quoteattr(_worksheet_attr_value(value))}")
    if not attrs and not model.get("present"):
        return None
    return f"<{tag}{''.join(attrs)}/>"


def _page_breaks_xml(model: dict, axis: str, tag: str) -> str:
    entries = model.get(axis) or []
    count = model.get(f"{axis}_count", len(entries))
    manual_count = model.get(
        f"{axis}_manualBreakCount",
        sum(1 for item in entries if item.get("man", True)),
    )
    children = []
    for item in entries:
        attrs = []
        for key in ("id", "min", "max", "man", "pt"):
            if key in item and item[key] is not None:
                attrs.append(f" {key}={quoteattr(_worksheet_attr_value(item[key]))}")
        children.append(f"<brk{''.join(attrs)}/>")
    if not children:
        return ""
    return (
        f"<{tag} count={quoteattr(str(count))} manualBreakCount={quoteattr(str(manual_count))}>"
        + "".join(children)
        + f"</{tag}>"
    )


def _protected_ranges_xml(ranges: list[dict]) -> str:
    children = []
    for item in ranges:
        if item.get("delete"):
            continue
        attrs = []
        for key, value in item.items():
            if key == "delete" or value is None:
                continue
            attrs.append(f" {key}={quoteattr(_worksheet_attr_value(value))}")
        if attrs:
            children.append(f"<protectedRange{''.join(attrs)}/>")
    return "<protectedRanges>" + "".join(children) + "</protectedRanges>" if children else ""


def _patch_row_attributes(content: str, rows: list[dict]) -> str:
    row_models = {str(index): row for index, row in enumerate(rows, 1)}
    if not row_models:
        return content

    boolean_fields = {
        "hidden": "hidden", "collapsed": "collapsed", "thickTop": "thickTop",
        "thickBot": "thickBot", "customFormat": "customFormat",
        "customHeight": "customHeight", "phonetic": "ph",
    }

    def replace(match):
        prefix, attr_text, suffix = match.groups()
        attrs = _parse_xml_attrs(attr_text)
        row_number = attrs.get("r")
        row = row_models.get(row_number)
        if row is None:
            return match.group(0)
        merged = copy.deepcopy(attrs)
        for key, value in (row.get("_row_attrs") or {}).items():
            if key != "r" and value is not None:
                merged[key] = str(value)
        if row.get("h") is None:
            merged.pop("ht", None)
        else:
            merged["ht"] = _worksheet_attr_value(row["h"])
        if row.get("outline"):
            merged["outlineLevel"] = _worksheet_attr_value(row["outline"])
        else:
            merged.pop("outlineLevel", None)
        for public_key, xml_key in boolean_fields.items():
            if public_key not in row:
                continue
            if row[public_key]:
                merged[xml_key] = "1"
            else:
                merged.pop(xml_key, None)
        if "style" in row:
            if row["style"] is None:
                merged.pop("s", None)
            else:
                merged["s"] = _worksheet_attr_value(row["style"])
        ordered = []
        if "r" in merged:
            ordered.append(("r", merged.pop("r")))
        ordered.extend(merged.items())
        rendered = "".join(f" {key}={quoteattr(str(value))}" for key, value in ordered)
        return prefix + rendered + suffix

    return re.sub(
        r"(<(?:[A-Za-z_][\w.-]*:)?row\b)([^>]*)(/?>)",
        replace,
        content,
    )


def _replace_worksheet_node(content: str, tag: str, replacement: str | None, following: tuple[str, ...]) -> str:
    if replacement is None:
        return content
    pattern = (
        rf"<(?:[A-Za-z_][\w.-]*:)?{tag}\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?{tag}>)"
    )
    content = re.sub(pattern, "", content, flags=re.DOTALL)
    if not replacement:
        return content
    following_pattern = "|".join(re.escape(value) for value in following)
    anchor = re.search(
        rf"<(?:[A-Za-z_][\w.-]*:)?(?:{following_pattern})\b",
        content,
    ) if following_pattern else None
    position = anchor.start() if anchor else content.rfind("</worksheet>")
    return content[:position] + replacement + content[position:]


def _inject_sheet_passthrough_relationships(xlsx_path: str, data: dict) -> str | None:
    requested = {
        sheet.get("name"): sheet.get("passthrough_relationships")
        for sheet in data.get("sheets") or []
        if sheet.get("passthrough_relationships")
    }
    if not requested:
        return None

    import posixpath

    replacement = xlsx_path + ".~sheet-relationships.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            infos = {item.filename: item for item in source.infolist()}
            entries = {item.filename: source.read(item.filename) for item in source.infolist()}

        sheet_file_map = _xlsx_sheet_file_map(
            entries["xl/workbook.xml"].decode("utf-8"),
            entries["xl/_rels/workbook.xml.rels"].decode("utf-8"),
        )
        content_types_root = ET.fromstring(entries["[Content_Types].xml"])
        content_types_namespace = content_types_root.tag.partition("}")[0].lstrip("{")
        override_tag = f"{{{content_types_namespace}}}Override"
        default_tag = f"{{{content_types_namespace}}}Default"

        for sheet_name, passthrough in requested.items():
            sheet_part = sheet_file_map.get(sheet_name)
            if not sheet_part:
                continue
            sheet_xml = entries[sheet_part].decode("utf-8")
            rels_part = _relationship_part_for_source(sheet_part)
            rels_root = (
                ET.fromstring(entries[rels_part])
                if rels_part in entries
                else ET.Element(f"{{{_PACKAGE_REL_NS}}}Relationships")
            )
            records = passthrough.get("relationships") or []
            reserved_ids = {str(record.get("Id")) for record in records if record.get("Id")}
            used_ids = {str(node.get("Id")) for node in rels_root if node.get("Id")}

            def next_relationship_id() -> str:
                index = 1
                while f"rId{index}" in used_ids or f"rId{index}" in reserved_ids:
                    index += 1
                relationship_id = f"rId{index}"
                used_ids.add(relationship_id)
                return relationship_id

            for record in records:
                target_part = record.get("target_part")
                target = str(record.get("Target") or "")
                if target_part and record.get("TargetMode") != "External":
                    target = posixpath.relpath(str(target_part), posixpath.dirname(sheet_part))
                desired_id = str(record.get("Id") or next_relationship_id())
                existing = next((node for node in rels_root if node.get("Id") == desired_id), None)
                same_relationship = bool(existing is not None and (
                    existing.get("Type") == record.get("Type")
                    and existing.get("Target") == target
                    and existing.get("TargetMode") == record.get("TargetMode")
                ))
                if existing is not None and not same_relationship:
                    replacement_id = next_relationship_id()
                    existing.set("Id", replacement_id)
                    sheet_xml = sheet_xml.replace(
                        f'r:id="{desired_id}"', f'r:id="{replacement_id}"'
                    )
                    sheet_xml = sheet_xml.replace(
                        f"r:id='{desired_id}'", f"r:id='{replacement_id}'"
                    )
                    existing = None
                if existing is None:
                    node = ET.SubElement(rels_root, f"{{{_PACKAGE_REL_NS}}}Relationship")
                    node.set("Id", desired_id)
                    node.set("Type", str(record.get("Type") or ""))
                    node.set("Target", target)
                    if record.get("TargetMode") is not None:
                        node.set("TargetMode", str(record["TargetMode"]))
                    used_ids.add(desired_id)

            for part_name, part_record in (passthrough.get("parts") or {}).items():
                entries[part_name] = base64.b64decode(part_record.get("data") or "")
                relationships_xml = part_record.get("relationships_xml")
                if relationships_xml:
                    entries[_relationship_part_for_source(part_name)] = base64.b64decode(
                        relationships_xml
                    )
                content_type = part_record.get("content_type")
                if not content_type:
                    continue
                if part_record.get("content_type_source") == "override":
                    part_path = f"/{part_name}"
                    override = next(
                        (node for node in content_types_root if node.get("PartName") == part_path),
                        None,
                    )
                    if override is None:
                        override = ET.SubElement(content_types_root, override_tag)
                        override.set("PartName", part_path)
                    override.set("ContentType", str(content_type))
                else:
                    extension = part_record.get("extension")
                    default = next(
                        (node for node in content_types_root if node.get("Extension") == extension),
                        None,
                    )
                    if default is None:
                        default = ET.SubElement(content_types_root, default_tag)
                        default.set("Extension", str(extension))
                        default.set("ContentType", str(content_type))
                    elif default.get("ContentType") != str(content_type):
                        part_path = f"/{part_name}"
                        override = next(
                            (node for node in content_types_root if node.get("PartName") == part_path),
                            None,
                        )
                        if override is None:
                            override = ET.SubElement(content_types_root, override_tag)
                            override.set("PartName", part_path)
                        override.set("ContentType", str(content_type))

            entries[sheet_part] = sheet_xml.encode("utf-8")
            entries[rels_part] = ET.tostring(
                rels_root, encoding="utf-8", xml_declaration=True
            )

        entries["[Content_Types].xml"] = ET.tostring(
            content_types_root, encoding="utf-8", xml_declaration=True
        )
        with zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as target:
            for name, raw in entries.items():
                target.writestr(infos.get(name, name), raw)
        os.replace(replacement, xlsx_path)
    except Exception as exc:
        if os.path.exists(replacement):
            os.remove(replacement)
        return f"worksheet relationship passthrough failed: {exc}"
    return None


def _inject_worksheet_semantics(xlsx_path: str, data: dict) -> str | None:
    requested = {}
    row_special_keys = {
        "collapsed", "thickTop", "thickBot", "customFormat", "customHeight", "style", "phonetic",
    }
    for sheet in data.get("sheets") or []:
        page_setup = sheet.get("page_setup")
        print_options = sheet.get("print_options")
        page_breaks_present = "page_breaks" in sheet
        protected_ranges_present = "protected_ranges" in sheet
        ignored_present = "ignored_errors" in sheet
        rows = sheet.get("rows") or []
        rows_present = any(
            row.get("_row_attrs") or any(key in row for key in row_special_keys)
            for row in rows
        )
        extension_xml = sheet.get("_worksheet_ext_xml")
        if not any((
            isinstance(page_setup, dict), isinstance(print_options, dict), page_breaks_present,
            protected_ranges_present, ignored_present, rows_present, bool(extension_xml),
        )):
            continue
        requested[sheet.get("name")] = {
            "page_setup": _worksheet_xml_element("pageSetup", page_setup, excluded={"fitToPage"}),
            "print_options": _worksheet_xml_element("printOptions", print_options),
            "row_breaks": _page_breaks_xml(sheet.get("page_breaks") or {}, "rows", "rowBreaks") if page_breaks_present else None,
            "column_breaks": _page_breaks_xml(sheet.get("page_breaks") or {}, "columns", "colBreaks") if page_breaks_present else None,
            "protected_ranges": _protected_ranges_xml(sheet.get("protected_ranges") or []) if protected_ranges_present else None,
            "ignored": _ignored_errors_xml(sheet.get("ignored_errors") or []) if ignored_present else None,
            "extension": extension_xml,
            "rows": rows if rows_present else None,
        }
    if not requested:
        return None

    tmp = xlsx_path + ".~worksheet-semantics.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            workbook_xml = source.read("xl/workbook.xml").decode("utf-8")
            rels_xml = source.read("xl/_rels/workbook.xml.rels").decode("utf-8")
            sheet_map = _xlsx_sheet_file_map(workbook_xml, rels_xml)
            file_map = {sheet_map[name]: value for name, value in requested.items() if name in sheet_map}
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as target:
                for item in source.infolist():
                    raw = source.read(item.filename)
                    patch = file_map.get(item.filename)
                    if patch is not None:
                        content = raw.decode("utf-8")
                        if patch["rows"] is not None:
                            content = _patch_row_attributes(content, patch["rows"])
                        content = _replace_worksheet_node(
                            content, "protectedRanges", patch["protected_ranges"],
                            ("scenarios", "autoFilter", "sortState", "dataConsolidate", "customSheetViews", "mergeCells", "phoneticPr", "conditionalFormatting", "dataValidations", "hyperlinks", "printOptions"),
                        )
                        content = _replace_worksheet_node(
                            content, "printOptions", patch["print_options"],
                            ("pageMargins", "pageSetup", "headerFooter", "rowBreaks", "colBreaks", "customProperties", "cellWatches", "ignoredErrors", "drawing", "legacyDrawing", "tableParts", "extLst"),
                        )
                        content = _replace_worksheet_node(
                            content, "pageSetup", patch["page_setup"],
                            ("headerFooter", "rowBreaks", "colBreaks", "customProperties", "cellWatches", "ignoredErrors", "drawing", "legacyDrawing", "tableParts", "extLst"),
                        )
                        content = _replace_worksheet_node(
                            content, "rowBreaks", patch["row_breaks"],
                            ("colBreaks", "customProperties", "cellWatches", "ignoredErrors", "drawing", "legacyDrawing", "tableParts", "extLst"),
                        )
                        content = _replace_worksheet_node(
                            content, "colBreaks", patch["column_breaks"],
                            ("customProperties", "cellWatches", "ignoredErrors", "drawing", "legacyDrawing", "tableParts", "extLst"),
                        )
                        content = _replace_worksheet_node(
                            content, "ignoredErrors", patch["ignored"],
                            ("smartTags", "drawing", "legacyDrawing", "legacyDrawingHF", "picture", "oleObjects", "controls", "webPublishItems", "tableParts", "extLst"),
                        )
                        content = _replace_worksheet_node(content, "extLst", patch["extension"], ())
                        raw = content.encode("utf-8")
                    target.writestr(item, raw)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"worksheet semantic injection failed: {exc}"
    return None


def _vml_comment_shapes(xml: str) -> dict[tuple[int, int], str]:
    shapes = {}
    pattern = re.compile(
        r"<(?:[A-Za-z_][\w.-]*:)?shape\b[^>]*>.*?</(?:[A-Za-z_][\w.-]*:)?shape>",
        re.DOTALL,
    )
    for match in pattern.finditer(xml):
        block = match.group(0)
        row = re.search(r"<(?:[A-Za-z_][\w.-]*:)?Row>(\d+)</(?:[A-Za-z_][\w.-]*:)?Row>", block)
        column = re.search(r"<(?:[A-Za-z_][\w.-]*:)?Column>(\d+)</(?:[A-Za-z_][\w.-]*:)?Column>", block)
        if row and column:
            shapes[(int(row.group(1)), int(column.group(1)))] = block
    return shapes


def _inject_comment_vml(xlsx_path: str, data: dict) -> str | None:
    requested = {
        sheet.get("name"): {
            "source": sheet.get("_comment_vml"),
            "comments": set((sheet.get("comments") or {}).keys()),
        }
        for sheet in data.get("sheets") or []
        if sheet.get("_comment_vml") and sheet.get("comments")
    }
    if not requested:
        return None

    tmp = xlsx_path + ".~comment-vml.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            infos = {item.filename: item for item in source.infolist()}
            entries = {item.filename: source.read(item.filename) for item in source.infolist()}
        workbook_xml = entries["xl/workbook.xml"].decode("utf-8")
        rels_xml = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
        sheet_map = _xlsx_sheet_file_map(workbook_xml, rels_xml)
        from openpyxl.utils.cell import coordinate_to_tuple

        for sheet_name, model in requested.items():
            sheet_part = sheet_map.get(sheet_name)
            if not sheet_part:
                continue
            generated_part = _comment_vml_part(entries, sheet_part)
            if not generated_part or generated_part not in entries:
                continue
            source_shapes = _vml_comment_shapes(model["source"]["xml"])
            generated_xml = entries[generated_part].decode("utf-8")
            generated_shapes = _vml_comment_shapes(generated_xml)
            coordinates = {
                (row - 1, column - 1)
                for row, column in (coordinate_to_tuple(coord) for coord in model["comments"])
            }
            for coordinate in coordinates:
                source_shape = source_shapes.get(coordinate)
                generated_shape = generated_shapes.get(coordinate)
                if source_shape and generated_shape:
                    generated_xml = generated_xml.replace(generated_shape, source_shape, 1)
            entries[generated_part] = generated_xml.encode("utf-8")

        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as target:
            for name, raw in entries.items():
                target.writestr(infos.get(name) or name, raw)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"comment VML injection failed: {exc}"
    return None


def _register_named_styles(wb, data: dict) -> None:
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Protection

    for item in data.get("named_styles") or []:
        name = item.get("name")
        if not name or name in wb.named_styles:
            continue
        style = item.get("style") or {}
        named = NamedStyle(name=name, builtinId=item.get("builtinId"), hidden=item.get("hidden"))

        font_spec = copy.deepcopy(style.get("font") or {}) if isinstance(style.get("font"), dict) else {}
        for source_key, target_key in (
            ("bold", "bold"), ("italic", "italic"), ("size", "size"),
            ("font", "name"), ("fcolor", "color"), ("uline", "underline"),
            ("strike", "strike"), ("vAlign", "vertAlign"),
        ):
            if source_key in style and target_key not in font_spec:
                value = style[source_key]
                if source_key == "font" and isinstance(value, dict):
                    continue
                font_spec[target_key] = value
        font_aliases = {
            "family": "family", "charset": "charset", "scheme": "scheme",
            "condense": "condense", "extend": "extend", "outline": "outline",
            "shadow": "shadow", "vert_align": "vertAlign", "vertAlign": "vertAlign",
        }
        font_kwargs = {}
        for key, value in font_spec.items():
            target = font_aliases.get(key, key)
            if target == "color" and isinstance(value, dict):
                value = _make_color_from_ref(value)
            if value is not None:
                font_kwargs[target] = value
        if font_kwargs:
            named.font = Font(**font_kwargs)

        fill_spec = style.get("fill")
        if isinstance(fill_spec, str):
            named.fill = PatternFill("solid", fgColor=fill_spec)
        elif isinstance(fill_spec, dict):
            pattern = fill_spec.get("pattern_type") or fill_spec.get("patternType") or fill_spec.get("fill_type")
            foreground = fill_spec.get("foreground") or fill_spec.get("fgColor")
            background = fill_spec.get("background") or fill_spec.get("bgColor")
            if isinstance(foreground, dict):
                foreground = _make_color_from_ref(foreground)
            if isinstance(background, dict):
                background = _make_color_from_ref(background)
            if pattern or foreground or background:
                fill_kwargs = {"patternType": pattern or "solid"}
                if foreground is not None:
                    fill_kwargs["fgColor"] = foreground
                if background is not None:
                    fill_kwargs["bgColor"] = background
                named.fill = PatternFill(**fill_kwargs)

        alignment_spec = style.get("alignment") if isinstance(style.get("alignment"), dict) else {}
        alignment_aliases = {
            "wrap_text": "wrapText", "text_rotation": "textRotation",
            "shrink_to_fit": "shrinkToFit", "merge_cell": "mergeCell",
        }
        alignment_kwargs = {
            alignment_aliases.get(key, key): value
            for key, value in alignment_spec.items()
            if value is not None
        }
        if alignment_kwargs:
            named.alignment = Alignment(**alignment_kwargs)

        protection_spec = style.get("protection") if isinstance(style.get("protection"), dict) else {}
        if protection_spec:
            named.protection = Protection(**{
                key: value for key, value in protection_spec.items()
                if key in {"locked", "hidden"} and value is not None
            })

        border_spec = style.get("border") if isinstance(style.get("border"), dict) else {}
        if border_spec:
            named.border = Border(
                start=_make_border_side_semantic(border_spec.get("start")),
                end=_make_border_side_semantic(border_spec.get("end")),
                top=_make_border_side_semantic(border_spec.get("top")),
                bottom=_make_border_side_semantic(border_spec.get("bottom")),
                left=_make_border_side_semantic(border_spec.get("left")),
                right=_make_border_side_semantic(border_spec.get("right")),
                diagonal=_make_border_side_semantic(border_spec.get("diagonal")),
                vertical=_make_border_side_semantic(border_spec.get("vertical")),
                horizontal=_make_border_side_semantic(border_spec.get("horizontal")),
                diagonalUp=bool(border_spec.get("diagonalUp")),
                diagonalDown=bool(border_spec.get("diagonalDown")),
                outline=border_spec.get("outline", True),
            )

        number_format = style.get("number_format", style.get("numfmt"))
        if number_format is not None:
            named.number_format = number_format
        wb.add_named_style(named)


def _inject_xf_contracts(xlsx_path: str, data: dict) -> str | None:
    requested_cells = []
    for sheet in data.get("sheets") or []:
        for row_index, row in enumerate(sheet.get("rows") or [], 1):
            for col_index, cell_data in enumerate(row.get("cells") or [], 1):
                xf = cell_data.get("xf") or {}
                overrides = {key: xf[key] for key in _XF_BOOLEAN_ATTRS | {"xfId"} if key in xf}
                if "qp" in cell_data:
                    overrides["quotePrefix"] = bool(cell_data["qp"])
                if overrides:
                    requested_cells.append((sheet.get("name"), row_index, col_index, overrides))
    if not requested_cells:
        return None

    tmp = xlsx_path + ".~xf-contracts.tmp"
    try:
        with zipfile.ZipFile(xlsx_path, "r") as source:
            infos = {item.filename: item for item in source.infolist()}
            entries = {item.filename: source.read(item.filename) for item in source.infolist()}
        workbook_xml = entries["xl/workbook.xml"].decode("utf-8")
        rels_xml = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
        sheet_map = _xlsx_sheet_file_map(workbook_xml, rels_xml)
        styles_root = ET.fromstring(entries["xl/styles.xml"])
        cell_xfs = styles_root.find(_qname("cellXfs"))
        if cell_xfs is None or not list(cell_xfs):
            raise ValueError("xl/styles.xml has no cellXfs definitions")

        sheet_roots = {}
        style_cache = {}
        from openpyxl.utils import get_column_letter
        for sheet_name, row_index, col_index, overrides in requested_cells:
            part_name = sheet_map.get(sheet_name)
            if not part_name or part_name not in entries:
                continue
            sheet_root = sheet_roots.setdefault(part_name, ET.fromstring(entries[part_name]))
            coord = f"{get_column_letter(col_index)}{row_index}"
            cell_node = sheet_root.find(f".//{_qname('c')}[@r='{coord}']")
            if cell_node is None:
                continue
            try:
                base_index = int(cell_node.get("s", "0"))
            except ValueError:
                base_index = 0
            if not 0 <= base_index < len(cell_xfs):
                base_index = 0
            xf_node = copy.deepcopy(cell_xfs[base_index])
            for key, value in overrides.items():
                if value is None:
                    xf_node.attrib.pop(key, None)
                elif key in _XF_BOOLEAN_ATTRS:
                    xf_node.set(key, "1" if bool(value) else "0")
                else:
                    xf_node.set(key, str(value))
            signature = ET.tostring(xf_node, encoding="utf-8")
            style_index = style_cache.get(signature)
            if style_index is None:
                style_index = len(cell_xfs)
                cell_xfs.append(xf_node)
                style_cache[signature] = style_index
            cell_node.set("s", str(style_index))

        cell_xfs.set("count", str(len(cell_xfs)))
        ET.register_namespace("", _SPREADSHEET_NS)
        entries["xl/styles.xml"] = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)
        for part_name, sheet_root in sheet_roots.items():
            entries[part_name] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)

        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as target:
            for name, raw in entries.items():
                target.writestr(infos.get(name) or name, raw)
        os.replace(tmp, xlsx_path)
    except Exception as exc:
        if os.path.exists(tmp):
            os.remove(tmp)
        return f"cell XF contract injection failed: {exc}"
    return None


def reconstruct_excel(data: dict, output_path: str) -> list[str]:
    """Reconstruct an Excel file from metadata dict produced by serialize_excel.

    Writes atomically: everything is assembled in a temp file that replaces
    output_path only on success, so a failure never corrupts an existing file.
    Returns warning strings for passthrough features that could not be restored.
    """
    exact_warnings = _atomic_copy_source_package(data, output_path)
    if exact_warnings is not None:
        return exact_warnings
    content_changes = _content_only_changes(data)
    if content_changes is not None:
        if not content_changes:
            unchanged_warnings = _atomic_copy_source_package(
                data,
                output_path,
                require_session_digest=False,
            )
            if unchanged_warnings is not None:
                return unchanged_warnings
        else:
            content_warnings = _reconstruct_content_only(data, output_path, content_changes)
            if content_warnings is not None:
                return content_warnings

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if data.get("theme_xml"):
        try:
            import base64
            wb.loaded_theme = base64.b64decode(data["theme_xml"])
        except Exception:
            pass

    _register_named_styles(wb, data)

    for sd in data["sheets"]:
        ws = wb.create_sheet(sd["name"])
        if sd.get("state") in {"visible", "hidden", "veryHidden"}:
            ws.sheet_state = sd["state"]

        sheet_properties = sd.get("sheet_properties") or {}
        for key, value in sheet_properties.items():
            if key in {"outline", "page_setup_properties"} or value is None:
                continue
            if hasattr(ws.sheet_properties, key):
                try:
                    setattr(ws.sheet_properties, key, value)
                except Exception:
                    pass
        outline_properties = sheet_properties.get("outline") or {}
        for key, value in outline_properties.items():
            if value is not None and hasattr(ws.sheet_properties.outlinePr, key):
                setattr(ws.sheet_properties.outlinePr, key, value)
        setup_properties = sheet_properties.get("page_setup_properties") or {}
        for key, value in setup_properties.items():
            if value is not None and hasattr(ws.sheet_properties.pageSetUpPr, key):
                setattr(ws.sheet_properties.pageSetUpPr, key, value)

        if sd.get("freeze"):
            ws.freeze_panes = sd["freeze"]

        sheet_views_data = sd.get("sheet_views")
        if sheet_views_data:
            try:
                from openpyxl.worksheet.views import Pane, Selection, SheetView
                built_views = []
                view_attrs = set(getattr(SheetView, "__attrs__", ()))
                pane_attrs = set(getattr(Pane, "__attrs__", ()))
                selection_attrs = set(getattr(Selection, "__attrs__", ()))
                for view_data in sheet_views_data:
                    attrs = {
                        key: value for key, value in view_data.items()
                        if key in view_attrs and value is not None
                    }
                    attrs.setdefault("workbookViewId", 0)
                    if isinstance(view_data.get("pane"), dict):
                        attrs["pane"] = Pane(**{
                            key: value for key, value in view_data["pane"].items()
                            if key in pane_attrs and value is not None
                        })
                    if "selections" in view_data:
                        attrs["selection"] = [
                            Selection(**{
                                key: value for key, value in selection.items()
                                if key in selection_attrs and value is not None
                            })
                            for selection in (view_data.get("selections") or [])
                        ]
                    built_views.append(SheetView(**attrs))
                ws.views.sheetView = built_views
            except Exception:
                pass

        sv_data = {} if sheet_views_data else (sd.get("sheet_view") or {})
        sv_aliases = {"zoom": "zoomScale"}
        sv_supported = set(getattr(type(ws.sheet_view), "__attrs__", ()))
        for key, value in sv_data.items():
            if key.startswith("_") or value is None:
                continue
            attr = sv_aliases.get(key, key)
            if attr not in sv_supported:
                continue
            try:
                setattr(ws.sheet_view, attr, value)
            except Exception:
                pass

        for col_letter, width in sd["cw"].items():
            if width is not None:
                ws.column_dimensions[col_letter].width = width
        for col_letter in sd.get("ch", {}):
            ws.column_dimensions[col_letter].hidden = True
        for col_letter, level in (sd.get("co") or {}).items():
            ws.column_dimensions[col_letter].outlineLevel = level

        if sd.get("tab_color"):
            from openpyxl.styles.colors import Color as _Color
            ws.sheet_properties.tabColor = _Color(rgb=sd["tab_color"])

        _apply_auto_filter_model(ws, sd.get("auto_filter_model") or (
            {"ref": sd.get("auto_filter")} if sd.get("auto_filter") else None
        ))


        ps_data = sd.get("page_setup") or {}
        if ps_data:
            for key in getattr(type(ws.page_setup), "__attrs__", ()):
                if ps_data.get(key) is not None:
                    setattr(ws.page_setup, key, ps_data[key])
            if "fitToPage" in ps_data and ps_data.get("fitToPage") is not None:
                try:
                    from openpyxl.worksheet.properties import PageSetupProperties
                    if ws.sheet_properties.pageSetUpPr is None:
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
                    ws.sheet_properties.pageSetUpPr.fitToPage = bool(ps_data["fitToPage"])
                except Exception:
                    pass

        print_options = sd.get("print_options") or {}
        for key in getattr(type(ws.print_options), "__attrs__", ()):
            if key in print_options and print_options[key] is not None:
                setattr(ws.print_options, key, print_options[key])

        pm_data = sd.get("page_margins") or {}
        if pm_data:
            for key in ("left", "right", "top", "bottom", "header", "footer"):
                if pm_data.get(key) is not None:
                    setattr(ws.page_margins, key, pm_data[key])

        prot_data = sd.get("protection")
        if prot_data is not None:
            ws.protection.sheet = True
            if prot_data.get("password") is not None:
                try:
                    ws.protection.set_password(
                        prot_data["password"],
                        already_hashed=bool(prot_data.get("password_is_hashed", True)),
                    )
                except Exception:
                    ws.protection.password = prot_data["password"]
            for key in getattr(type(ws.protection), "__attrs__", ()):
                if key not in {"sheet", "password"} and key in prot_data:
                    setattr(ws.protection, key, prot_data[key])

        # Print titles
        pt = sd.get("print_titles") or {}
        if pt.get("rows"):
            ws.print_title_rows = pt["rows"]
        if pt.get("cols"):
            ws.print_title_cols = pt["cols"]

        # Print area
        if sd.get("print_area"):
            ws.print_area = sd["print_area"]

        # Header / footer
        hf = sd.get("header_footer") or {}
        for property_name in getattr(type(ws.HeaderFooter), "__attrs__", ()):
            if property_name in hf and hf[property_name] is not None:
                setattr(ws.HeaderFooter, property_name, hf[property_name])
        section_objects = {
            "odd_header": ws.oddHeader, "odd_footer": ws.oddFooter,
            "even_header": ws.evenHeader, "even_footer": ws.evenFooter,
            "first_header": ws.firstHeader, "first_footer": ws.firstFooter,
        }
        for section_name, section_object in section_objects.items():
            section = hf.get(section_name)
            if not isinstance(section, dict):
                continue
            for position in ("left", "center", "right"):
                if position in section:
                    getattr(section_object, position).text = section[position]
        if hf.get("hl"): ws.oddHeader.left.text   = hf["hl"]
        if hf.get("hc"): ws.oddHeader.center.text = hf["hc"]
        if hf.get("hr"): ws.oddHeader.right.text  = hf["hr"]
        if hf.get("fl"): ws.oddFooter.left.text   = hf["fl"]
        if hf.get("fc"): ws.oddFooter.center.text = hf["fc"]
        if hf.get("fr"): ws.oddFooter.right.text  = hf["fr"]

        page_breaks = sd.get("page_breaks") or {}
        if page_breaks:
            try:
                from openpyxl.worksheet.pagebreak import Break
                for item in page_breaks.get("rows") or []:
                    ws.row_breaks.append(Break(**{
                        key: value for key, value in item.items()
                        if key in {"id", "min", "max", "man", "pt"} and value is not None
                    }))
                for item in page_breaks.get("columns") or []:
                    ws.col_breaks.append(Break(**{
                        key: value for key, value in item.items()
                        if key in {"id", "min", "max", "man", "pt"} and value is not None
                    }))
            except Exception:
                pass

        # Tables
        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo as TSI
            for t_data in (sd.get("tables") or []):
                t = _build_table(t_data, len(ws.tables) + 1)
                s = t_data.get("style")
                if s:
                    # Accept both the short internal aliases (showFirstCol/
                    # showLastCol/showColStripes, used by the read side below)
                    # and the canonical OOXML attribute names (showFirstColumn/
                    # showLastColumn/showColumnStripes) that a caller following
                    # the documented tableStyleInfo flag names would reasonably
                    # pass -- previously only the short aliases were honored
                    # and the canonical names were silently ignored (always
                    # False).
                    def _flag(short_key, long_key, default=False):
                        if long_key in s:
                            return bool(s.get(long_key))
                        return bool(s.get(short_key, default))

                    t.tableStyleInfo = TSI(
                        name=s.get("name"),
                        showFirstColumn=_flag("showFirstCol", "showFirstColumn"),
                        showLastColumn=_flag("showLastCol", "showLastColumn"),
                        showRowStripes=_flag("showRowStripes", "showRowStripes", True),
                        showColumnStripes=_flag("showColStripes", "showColumnStripes"),
                    )
                ws.add_table(t)
        except Exception:
            pass

        for r_idx, row_data in enumerate(sd["rows"], 1):
            if row_data.get("h") is not None:
                ws.row_dimensions[r_idx].height = row_data["h"]
            if row_data.get("hidden"):
                ws.row_dimensions[r_idx].hidden = True
            if row_data.get("outline"):
                ws.row_dimensions[r_idx].outlineLevel = row_data["outline"]
            row_dimension_keys = ("collapsed", "thickTop", "thickBot")
            if any(key in row_data for key in row_dimension_keys):
                row_dimension = ws.row_dimensions[r_idx]
                for key in row_dimension_keys:
                    if key in row_data:
                        setattr(row_dimension, key, bool(row_data[key]))

            for c_idx, cd in enumerate(row_data["cells"], 1):
                if cd.get("_implicit"):
                    public_keys = {key for key in cd if key != "_implicit"}
                    if public_keys <= {"v", "merge"} and cd.get("v") is None and not cd.get("merge"):
                        continue
                    cd = _expanded_implicit_cell(cd, sd.get("_implicit_cell_defaults"))
                if cd["merge"] == "slave":
                    continue

                style_id = (cd.get("xf") or {}).get("style_id")
                has_explicit_style = bool(
                    cd.get("fill") or cd.get("bold") or cd.get("italic") or cd.get("strike")
                    or cd.get("uline") or cd.get("fcolor") or cd.get("border")
                    or cd.get("numfmt") not in (None, "General") or style_id not in (None, 0)
                )
                if (not cd.get("present") and cd.get("v") is None
                        and not cd.get("formula") and not cd.get("rich_text")
                        and not has_explicit_style and not cd.get("merge")):
                    continue

                cell = ws.cell(row=r_idx, column=c_idx, value=cd.get("v"))
                named_style = cd.get("named_style")
                if named_style and named_style in wb.named_styles:
                    cell.style = named_style
                if cd.get("dt") == "s" and isinstance(cd["v"], str):
                    cell.data_type = "s"  # literal text, not a formula

                raw_fill = _usable_raw_fill(cd)
                if raw_fill:
                    try:
                        cell.fill = _make_pattern_fill_from_raw(raw_fill)
                    except Exception:
                        pass
                elif cd.get("fill"):
                    try:
                        cell.fill = PatternFill("solid", fgColor=cd["fill"])
                    except Exception:
                        pass

                fk: dict = {}
                if cd.get("bold"):          fk["bold"]      = True
                if cd.get("italic"):        fk["italic"]    = True
                if cd.get("size"):          fk["size"]      = cd["size"]
                if cd.get("font"):          fk["name"]      = cd["font"]
                if cd.get("uline"):
                    fk["underline"] = "single" if cd["uline"] is True else cd["uline"]
                if cd.get("strike"):        fk["strike"]    = True
                if cd.get("vAlign"):        fk["vertAlign"] = cd["vAlign"]
                _apply_raw_font_kwargs(fk, cd)
                if fk:
                    cell.font = Font(**fk)

                alignment_data = cd.get("alignment") or {}
                if alignment_data:
                    allowed_alignment = {
                        "horizontal", "vertical", "textRotation", "wrapText", "shrinkToFit",
                        "indent", "relativeIndent", "justifyLastLine", "readingOrder",
                    }
                    ak = {key: value for key, value in alignment_data.items() if key in allowed_alignment}
                else:
                    ak: dict = {}
                    if cd.get("wrap"):   ak["wrap_text"]    = True
                    if cd.get("halign"): ak["horizontal"]   = cd["halign"]
                    if cd.get("valign"): ak["vertical"]     = cd["valign"]
                    if cd.get("rot"):    ak["text_rotation"] = cd["rot"]
                    if cd.get("indent"): ak["indent"]        = cd["indent"]
                    if cd.get("shrink"): ak["shrink_to_fit"] = True
                if ak:
                    cell.alignment = Alignment(**ak)

                if cd.get("numfmt"):
                    cell.number_format = cd["numfmt"]

                border_data = cd.get("border_semantics") or {}
                bdr = cd.get("border", {})
                if border_data:
                    cell.border = Border(
                        start=_make_border_side_semantic(border_data.get("start")),
                        end=_make_border_side_semantic(border_data.get("end")),
                        top=_make_border_side_semantic(border_data.get("top")),
                        bottom=_make_border_side_semantic(border_data.get("bottom")),
                        left=_make_border_side_semantic(border_data.get("left")),
                        right=_make_border_side_semantic(border_data.get("right")),
                        diagonal=_make_border_side_semantic(border_data.get("diagonal")),
                        vertical=_make_border_side_semantic(border_data.get("vertical")),
                        horizontal=_make_border_side_semantic(border_data.get("horizontal")),
                        diagonalUp=bool(border_data.get("diagonalUp")),
                        diagonalDown=bool(border_data.get("diagonalDown")),
                        outline=border_data.get("outline", True),
                    )
                elif bdr:
                    cell.border = Border(
                        start=_make_border_side(bdr.get("start")),
                        end=_make_border_side(bdr.get("end")),
                        top=_make_border_side(bdr.get("top")),
                        bottom=_make_border_side(bdr.get("bottom")),
                        left=_make_border_side(bdr.get("left")),
                        right=_make_border_side(bdr.get("right")),
                        diagonal=_make_border_side(bdr.get("diagonal")),
                        vertical=_make_border_side(bdr.get("vertical")),
                        horizontal=_make_border_side(bdr.get("horizontal")),
                        diagonalUp=bool(bdr.get("diagonalUp")),
                        diagonalDown=bool(bdr.get("diagonalDown")),
                        outline=bdr.get("outline", True),
                    )

                locked = cd.get("locked", True)
                hidden_p = cd.get("hidden_cell", False)
                if locked is False or hidden_p:
                    cell.protection = Protection(locked=bool(locked), hidden=bool(hidden_p))

                if cd.get("qp"):
                    try:
                        from openpyxl.styles.cell_style import StyleArray
                        if cell._style is None:
                            cell._style = StyleArray()
                        cell._style.quotePrefix = 1
                    except Exception:
                        pass

                mi = cd["merge"]
                if isinstance(mi, dict) and (mi.get("rowspan", 1) > 1 or mi.get("colspan", 1) > 1):
                    ws.merge_cells(
                        start_row=r_idx, start_column=c_idx,
                        end_row=r_idx + mi["rowspan"] - 1,
                        end_column=c_idx + mi["colspan"] - 1,
                    )

        for vd in sd.get("validations", []):
            dv_kwargs = {
                key: vd.get(key)
                for key in (
                    "type", "formula1", "formula2", "showErrorMessage",
                    "showInputMessage", "showDropDown",
                    "promptTitle", "errorStyle", "error", "prompt",
                    "errorTitle", "imeMode", "operator",
                )
                if vd.get(key) is not None
            }
            if vd.get("allowBlank") is not None:
                dv_kwargs["allow_blank"] = vd["allowBlank"]
            elif vd.get("allow_blank") is not None:
                dv_kwargs["allow_blank"] = vd["allow_blank"]
            dv = DataValidation(**dv_kwargs)
            for sqref_part in vd["sqref"].split():
                dv.add(sqref_part)
            ws.add_data_validation(dv)

        # Hyperlinks
        try:
            from openpyxl.worksheet.hyperlink import Hyperlink
            for coord, hl_data in (sd.get("hyperlinks") or {}).items():
                if not (hl_data.get("target") or hl_data.get("location")):
                    continue
                cell = ws[coord]
                hl = Hyperlink(ref=coord, target=hl_data.get("target"))
                if hl_data.get("location"):
                    hl.location = hl_data["location"]
                if "display" in hl_data and hl_data.get("display") is not None:
                    hl.display = hl_data["display"]
                if hl_data.get("tooltip"):
                    hl.tooltip = hl_data["tooltip"]
                cell.hyperlink = hl
        except Exception:
            pass

        # Comments
        try:
            from openpyxl.comments import Comment
            for coord, cm in (sd.get("comments") or {}).items():
                ws[coord].comment = Comment(cm["text"], cm.get("author", ""))
        except Exception:
            pass

    # Named ranges
    #
    # The three reserved built-ins are deliberately NOT written here: Print_Area
    # and Print_Titles are already applied natively above (ws.print_area /
    # ws.print_title_rows+cols), and _FilterDatabase is auto-generated by
    # openpyxl's own writer whenever ws.auto_filter.ref is set. Writing a raw
    # duplicate definedName for any of them here corrupts the workbook (two
    # competing _xlnm.Print_Titles entries for the same sheet, one of them
    # missing the sheet-name/$ prefix that Excel requires).
    _RESERVED_BUILTIN_NAMES = {"_xlnm.Print_Area", "_xlnm.Print_Titles", "_xlnm._FilterDatabase"}
    _DN_META_FIELDS = (
        "comment", "customMenu", "description", "help", "statusBar",
        "hidden", "function", "vbProcedure", "xlm", "functionGroupId",
        "shortcutKey", "publishToServer", "workbookParameter",
    )
    try:
        from openpyxl.workbook.defined_name import DefinedName
        for nr in data.get("named_ranges") or []:
            name = nr["name"]
            if name in _RESERVED_BUILTIN_NAMES:
                continue
            kwargs = {field: nr[field] for field in _DN_META_FIELDS if nr.get(field) is not None}
            dn = DefinedName(name, attr_text=nr["value"], **kwargs)
            sheet_id = nr.get("sheet_id")
            if sheet_id is not None and 0 <= sheet_id < len(wb.worksheets):
                # Worksheet-scoped: goes in that sheet's OWN dict, keyed only
                # by name -- putting every scope into the single workbook-wide
                # wb.defined_names dict (as before) meant two different sheets
                # both naming a local range e.g. "Data" collided and one was
                # silently lost.
                dn.localSheetId = sheet_id
                wb.worksheets[sheet_id].defined_names[name] = dn
            else:
                if sheet_id is not None:
                    dn.localSheetId = sheet_id
                wb.defined_names[name] = dn
    except Exception:
        pass

    # Document properties. "modified" is special: openpyxl's own
    # save_workbook() unconditionally stamps wb.properties.modified with
    # datetime.now() as its last step before writing, overriding anything set
    # here -- so the desired final value (per modified_policy) is recorded in
    # `_pending_modified_iso` and fixed up via _inject_doc_core_modified AFTER
    # wb.save(), except for update_on_save where openpyxl's own stamp is
    # exactly the desired behavior already.
    dp = copy.deepcopy(data.get("doc_props") or {})
    modified_policy = data.get("modified_policy", "preserve")
    _pending_modified_iso = None
    if modified_policy != "update_on_save" and dp.get("modified"):
        _pending_modified_iso = dp["modified"]
    if dp:
        try:
            from datetime import datetime
            for key, value in dp.items():
                if key in ("created", "modified", "lastPrinted"):
                    setattr(wb.properties, key, datetime.fromisoformat(value))
                else:
                    setattr(wb.properties, key, value)
        except Exception:
            pass

    # Custom document properties (docProps/custom.xml) -- openpyxl round-trips
    # typed values (string/int/float/bool/datetime/link) natively.
    custom_props = data.get("custom_doc_props") or []
    if custom_props:
        try:
            from openpyxl.packaging.custom import (
                StringProperty, IntProperty, FloatProperty, BoolProperty,
                DateTimeProperty, LinkProperty,
            )
            from datetime import datetime as _dt_custom
            _CUSTOM_TYPE_MAP = {
                "StringProperty": StringProperty, "IntProperty": IntProperty,
                "FloatProperty": FloatProperty, "BoolProperty": BoolProperty,
                "DateTimeProperty": DateTimeProperty, "LinkProperty": LinkProperty,
            }
            wb.custom_doc_props.props = []
            for item in custom_props:
                cls = _CUSTOM_TYPE_MAP.get(item.get("type"), StringProperty)
                value = item.get("value")
                if cls is DateTimeProperty and isinstance(value, str):
                    value = _dt_custom.fromisoformat(value)
                wb.custom_doc_props.append(cls(name=item["name"], value=value))
        except Exception:
            pass

    # Calculation properties (calcPr) -- openpyxl's CalcProperties model
    # supports every field the tool exposes natively.
    calc_props = data.get("calculation_properties") or {}
    if calc_props:
        try:
            for key, value in calc_props.items():
                if hasattr(wb.calculation, key):
                    setattr(wb.calculation, key, value)
        except Exception:
            pass

    # Workbook protection (workbookProtection) -- pre-hashed values pass
    # through untouched; plain-text passwords are hashed unless the caller
    # already supplied a hash (already_hashed=True).
    wb_protection = data.get("workbook_protection") or {}
    if wb_protection:
        try:
            already_hashed = bool(wb_protection.get("already_hashed"))
            for key, value in wb_protection.items():
                if key in ("already_hashed", "workbookPassword", "revisionsPassword"):
                    continue
                if hasattr(wb.security, key):
                    setattr(wb.security, key, value)
            if wb_protection.get("workbookPassword") is not None:
                wb.security.set_workbook_password(wb_protection["workbookPassword"], already_hashed=already_hashed)
            if wb_protection.get("revisionsPassword") is not None:
                wb.security.set_revisions_password(wb_protection["revisionsPassword"], already_hashed=already_hashed)
        except Exception:
            pass

    # Workbook properties (workbookPr): codeName + date1904 round-trip
    # natively through openpyxl's object model (wb.code_name / wb.epoch);
    # everything else needs post-save XML injection (_wbpr_extra_attrs,
    # applied after wb.save() below). Cell values in the data model are kept
    # as Python datetime objects (see the per-cell loop above), and openpyxl
    # defers date->serial conversion to save time based on wb.epoch, so
    # simply changing wb.epoch here already achieves "preserve_displayed_dates"
    # (same calendar date, new serial) with no per-cell changes. The opposite
    # policy, "preserve_serial_values" (same serial, calendar date shifts),
    # needs an explicit per-cell re-basing, done further below.
    wb_props = data.get("workbook_properties") or {}
    _WBPR_NON_XML_KEYS = ("codeName", "date1904", "date_system_policy")
    _wbpr_extra_attrs = {k: v for k, v in wb_props.items() if k not in _WBPR_NON_XML_KEYS}
    _date1904_changed = False
    _old_epoch = wb.epoch
    _new_epoch = _old_epoch
    if wb_props.get("codeName"):
        wb.code_name = wb_props["codeName"]
    if "date1904" in wb_props:
        try:
            from openpyxl.utils.datetime import CALENDAR_MAC_1904, CALENDAR_WINDOWS_1900
            baseline_props = data.get("_baseline_workbook_properties") or {}
            baseline_date1904 = bool(baseline_props.get("date1904", False))
            _new_epoch = CALENDAR_MAC_1904 if wb_props["date1904"] else CALENDAR_WINDOWS_1900
            _date1904_changed = bool(wb_props["date1904"]) != baseline_date1904
            wb.epoch = _new_epoch
            if _date1904_changed and wb_props.get("date_system_policy") == "preserve_serial_values":
                from openpyxl.utils.datetime import to_excel, from_excel
                import datetime as _dtmod
                _rebase_epoch = CALENDAR_MAC_1904 if baseline_date1904 else CALENDAR_WINDOWS_1900
                for _ws_rebase in wb.worksheets:
                    for _row_rebase in _ws_rebase.iter_rows():
                        for _c_rebase in _row_rebase:
                            _v_rebase = _c_rebase.value
                            if isinstance(_v_rebase, (_dtmod.datetime, _dtmod.date, _dtmod.time)):
                                _serial = to_excel(_v_rebase, epoch=_rebase_epoch)
                                _c_rebase.value = from_excel(_serial, epoch=_new_epoch)
        except Exception:
            pass

    # Workbook views (bookViews) -- the FULL ordered list, not just index 0.
    wviews = data.get("workbook_views") or ([data.get("wb_view")] if data.get("wb_view") else [])
    if wviews:
        try:
            from openpyxl.workbook.views import BookView
            supported = set(getattr(BookView, "__attrs__", ()))
            new_views = []
            for entry in wviews:
                kwargs = {k: v for k, v in (entry or {}).items() if k in supported and v is not None}
                new_views.append(BookView(**kwargs))
            if new_views:
                wb.views = new_views
        except Exception:
            pass
    primary_view = (wviews[0] if wviews else {}) or {}
    if primary_view.get("activeTab") is not None:
        try:
            # openpyxl's writer takes activeTab from wb.active, not the view
            wb.active = max(0, min(int(primary_view["activeTab"]), len(wb.worksheets) - 1))
        except Exception:
            pass

    creation_plan = _stage_drawing_creations(wb, data)

    # Atomic write: assemble everything in a temp file, replace target on success.
    import os as _os
    out_str = str(output_path)
    tmp_out = out_str + ".~saving.tmp"
    warnings: list[str] = []
    try:
        wb.save(tmp_out)

        # Restore raw theme/indexed/tint fill XML and sheetView attributes after save.
        for w in (
            _inject_raw_fills(tmp_out, data),
            _inject_sheet_view_attrs(tmp_out, data),
            _inject_sheet_format_pr(tmp_out, data),
            _inject_raw_cols(tmp_out, data),
            _inject_data_validations_xml(
                tmp_out,
                {
                    sd["name"]: sd["data_validations_xml"]
                    for sd in data["sheets"]
                    if sd.get("data_validations_xml")
                },
            ),
        ):
            if w:
                warnings.append(w)

        # Inject conditional formatting (XML passthrough, must be after save)
        cf_map = {sd["name"]: sd["cf_xml"] for sd in data["sheets"] if sd.get("cf_xml")}
        dxfs_xml = data.get("dxfs_xml")
        if cf_map or dxfs_xml:
            if dxfs_xml:
                cf_map["__dxfs__"] = dxfs_xml
            w = _inject_cf_xml(tmp_out, cf_map)
            if w:
                warnings.append(w)

        # Inject drawings/charts/images (XML passthrough, must be after save)
        drawing_sheets = {sd["name"]: sd["drawing_data"] for sd in data["sheets"] if sd.get("drawing_data")}
        if drawing_sheets or creation_plan:
            import zipfile as _zf2
            try:
                with _zf2.ZipFile(tmp_out, "r") as _z2:
                    _wb2   = _z2.read("xl/workbook.xml").decode("utf-8")
                    _rels2 = _z2.read("xl/_rels/workbook.xml.rels").decode("utf-8")
                new_sfm = _xlsx_sheet_file_map(_wb2, _rels2)
            except Exception:
                new_sfm = {}
            w = (
                _merge_drawing_packages(tmp_out, drawing_sheets, creation_plan, new_sfm)
                if creation_plan
                else _inject_drawing_data(tmp_out, drawing_sheets, new_sfm)
            )
            if w:
                if creation_plan:
                    raise ValueError(w)
                warnings.append(w)

        w = _inject_sheet_passthrough_relationships(tmp_out, data)
        if w:
            warnings.append(w)

        w = _inject_worksheet_semantics(tmp_out, data)
        if w:
            warnings.append(w)

        w = _inject_comment_vml(tmp_out, data)
        if w:
            warnings.append(w)

        w = _inject_cell_contracts(tmp_out, data)
        if w:
            warnings.append(w)
        w = _inject_xf_contracts(tmp_out, data)
        if w:
            warnings.append(w)

        # Inject workbookPr attributes openpyxl's object model can't set
        # natively (filterPrivacy, saveExternalLinkValues, showObjects,
        # updateLinks, ...), the preserved/explicit document 'modified'
        # timestamp (openpyxl always overwrites it with now() on save), and
        # app.xml (extended) properties (openpyxl has no hook for these at all).
        w = _inject_workbook_pr_extra(tmp_out, _wbpr_extra_attrs)
        if w:
            warnings.append(w)
        if _pending_modified_iso:
            w = _inject_doc_core_modified(tmp_out, _pending_modified_iso)
            if w:
                warnings.append(w)
        w = _inject_app_props(tmp_out, data.get("app_props") or {})
        if w:
            warnings.append(w)

        w = _restore_missing_package_parts(data.get("source"), tmp_out)
        if w:
            warnings.append(w)

        _apply_package_edits(tmp_out, data.get("_package_edits"))
        warnings.extend(validate_xlsx(tmp_out))

        # Windows: a stale GC-held handle or AV scan can briefly lock the
        # target — retry the swap a few times before giving up.
        import gc as _gc
        import time as _time
        for attempt in range(5):
            try:
                _os.replace(tmp_out, out_str)
                break
            except PermissionError:
                if attempt == 4:
                    raise
                _gc.collect()
                _time.sleep(0.2)
    except Exception:
        if _os.path.exists(tmp_out):
            try:
                _os.remove(tmp_out)
            except OSError:
                pass
        raise
    return warnings

