from __future__ import annotations

import json
import sys
from pathlib import Path

import anyio
import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
import save_transaction as ST  # noqa: E402
from cancellable import (  # noqa: E402
    WorkerProcessCancelled,
    WorkerProcessTimeout,
    active_worker_pids,
    run_named_operation,
)


def _write_workbook(path: Path, first: str = "old", second: str = "keep") -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"
    worksheet["A1"] = first
    worksheet["B1"] = second
    workbook.save(path)
    workbook.close()


def _session_key(result: str) -> str:
    return result.split("session_key=", 1)[1].split(" |", 1)[0].strip("'")


def test_successful_transaction_updates_state_only_after_commit(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    original_bytes = source.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(
        session_key,
        "Sheet",
        [{"row_index": 0, "edits": {0: "new"}}],
    )

    original_backup = ST.create_excel_backup
    observed = {}

    def recording_backup(reference_path: str, saved_path: str):
        observed["destination_before_backup"] = Path(reference_path).read_bytes()
        observed["dirty_paths_before_backup"] = list(
            M._sessions[M._resolve_session_key(session_key)].get("_dirty_paths") or []
        )
        return original_backup(reference_path, saved_path)

    monkeypatch.setattr(ST, "create_excel_backup", recording_backup)
    report = json.loads(M.excel_save(
        session_key,
        report_format="json",
        verify_preservation=True,
    ))

    assert observed["destination_before_backup"] == original_bytes
    assert observed["dirty_paths_before_backup"]
    assert report["verification"]["preservation_ok"] is True
    assert report["package_signatures"]["performance"]["before"]["mode"] == "reused_full_inspection"
    assert report["package_signatures"]["performance"]["after"]["mode"] == "reused_full_inspection"
    assert report["performance"]["package_open_count"] == 2
    assert report["performance"]["part_read_count"] > 0
    assert report["verification"]["performance"] == report["performance"]
    assert report["backup"]["sha256"] == ST.sha256_file(Path(report["backup"]["backup_path"]))
    assert report["backup"]["sha256"] == ST.sha256_file(tmp_path / "backups" / Path(report["backup"]["backup_path"]).name)
    assert M._sessions[M._resolve_session_key(session_key)]["_verification_baseline_path"] == str(source.resolve())
    assert M._sessions[M._resolve_session_key(session_key)]["_dirty_features"] == []
    assert M._sessions[M._resolve_session_key(session_key)]["_dirty_paths"] == []
    workbook = openpyxl.load_workbook(source)
    assert workbook.active["A1"].value == "new"
    workbook.close()


def test_save_without_verification_uses_signature_only_reader(tmp_path):
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    session_key = _session_key(M.excel_load(str(source)))

    report = json.loads(M.excel_save(
        session_key,
        report_format="json",
        verify_preservation=False,
    ))

    assert report["verification"]["status"] == "not_run"
    assert report["performance"] is None
    assert report["package_signatures"]["performance"]["before"]["mode"] == "signature_only"
    assert report["package_signatures"]["performance"]["after"]["mode"] == "signature_only"


def test_load_metrics_do_not_disable_exact_copy_fast_path(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "copy.xlsx"
    _write_workbook(source)
    data = M.serialize_excel(str(source))
    session_key = _session_key(M._publish_loaded_session(
        data,
        None,
        load_metrics={"worker_serialization_seconds": 0.01},
    ))

    report = json.loads(M.excel_save_as_copy(
        session_key,
        str(destination),
        report_format="json",
        verify_preservation=True,
    ))

    assert destination.read_bytes() == source.read_bytes()
    assert report["verification"]["preservation_ok"] is True
    assert report["verification"]["change_count"] == 0
    assert report["performance"]["package_open_count"] == 1
    M.excel_close(session_key)


def test_worker_error_removes_staging_and_preserves_destination(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    original_bytes = source.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(
        session_key,
        "Sheet",
        [{"row_index": 0, "edits": {0: "new"}}],
    )
    observed = {}

    def fail_after_staging(session_data, **payload):
        del session_data
        staging = Path(payload["staging_path"])
        staging.write_bytes(b"partial")
        observed["staging"] = staging
        raise RuntimeError("worker failed")

    monkeypatch.setattr(M, "_save_stage_operation", fail_after_staging)
    with pytest.raises(RuntimeError, match="worker failed"):
        M.excel_save(session_key, report_format="json")

    assert source.read_bytes() == original_bytes
    assert not observed["staging"].exists()
    assert M._sessions[M._resolve_session_key(session_key)]["_dirty_paths"]
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS


def test_replace_failure_discards_backup_and_keeps_session_dirty(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    original_bytes = source.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(
        session_key,
        "Sheet",
        [{"row_index": 0, "edits": {0: "new"}}],
    )

    def reject_replace(source_path, destination_path):
        raise PermissionError(f"blocked replace {source_path} -> {destination_path}")

    monkeypatch.setattr(ST.os, "replace", reject_replace)
    with pytest.raises(PermissionError, match="blocked replace"):
        M.excel_save(session_key, report_format="json")

    assert source.read_bytes() == original_bytes
    assert M._sessions[M._resolve_session_key(session_key)].get("_verification_baseline_path") is None
    assert M._sessions[M._resolve_session_key(session_key)]["_dirty_paths"]
    backup_dir = tmp_path / "backups"
    assert not backup_dir.exists() or not list(backup_dir.iterdir())


async def _exercise_busy_session(source: Path, session_key: str, monkeypatch) -> None:
    started = anyio.Event()
    release = anyio.Event()

    async def delayed_worker(operation, payload, **kwargs):
        del operation, kwargs
        Path(payload["staging_path"]).write_bytes(b"staged")
        started.set()
        await release.wait()
        raise RuntimeError("stop after busy assertions")

    monkeypatch.setattr(M, "run_named_operation", delayed_worker)
    outcome = {}

    async def invoke_save():
        try:
            await M._excel_save_tool(session_key, timeout_seconds=30.0)
        except RuntimeError as exc:
            outcome["error"] = exc

    async with anyio.create_task_group() as task_group:
        task_group.start_soon(invoke_save)
        await started.wait()
        status = json.loads(M.excel_get_session_status(session_key))
        assert status["busy"] is True
        assert status["operation"] == "excel_save"
        with pytest.raises(M.ExcelOperationError) as close_error:
            M.excel_close(session_key)
        assert close_error.value.details["code"] == "EXCEL_SESSION_BUSY"
        with pytest.raises(M.ExcelOperationError):
            M.excel_edit_cells(
                session_key,
                "Sheet",
                [{"row_index": 0, "edits": {0: "blocked"}}],
            )
        release.set()

    assert isinstance(outcome.get("error"), RuntimeError)
    assert source.is_file()
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS


def test_busy_session_rejects_close_and_mutation(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    session_key = _session_key(M.excel_load(str(source)))
    anyio.run(_exercise_busy_session, source, session_key, monkeypatch)


async def _invoke_timeout_save(session_key: str, destination: Path, monkeypatch) -> None:
    async def timeout_worker(operation, payload, **kwargs):
        del operation, kwargs
        Path(payload["staging_path"]).write_bytes(b"partial-stage")
        raise WorkerProcessTimeout({
            "code": "HEAVY_OPERATION_TIMEOUT",
            "operation": "excel_save",
            "worker_stopped": True,
            "staging_removed": False,
        })

    monkeypatch.setattr(M, "run_named_operation", timeout_worker)
    with pytest.raises(WorkerProcessTimeout):
        await M._excel_save_tool(
            session_key,
            output_path=str(destination),
            timeout_seconds=0.1,
        )


def test_timeout_preserves_existing_destination_and_removes_stage(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "destination.xlsx"
    _write_workbook(source)
    _write_workbook(destination, first="target")
    original_destination = destination.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))

    anyio.run(_invoke_timeout_save, session_key, destination, monkeypatch)

    assert destination.read_bytes() == original_destination
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS
    assert not list(tmp_path.glob(".*.docloupe-*.xlsx"))


def test_timeout_to_new_path_leaves_no_output(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "new-output.xlsx"
    _write_workbook(source)
    session_key = _session_key(M.excel_load(str(source)))

    anyio.run(_invoke_timeout_save, session_key, destination, monkeypatch)

    assert not destination.exists()
    assert not list(tmp_path.glob(".*.docloupe-*.xlsx"))


async def _run_real_save_stage(data: dict, staging: Path) -> dict:
    return await run_named_operation(
        "save_stage",
        {
            "staging_path": str(staging),
            "verification_reference_path": data["source"],
            "verify_preservation": True,
            "max_differences": 200,
            "requested_paths": [],
            "intentional_edit": False,
        },
        input_data=data,
        failure_cleanup_paths=(staging,),
        timeout_seconds=30.0,
        operation_label="save-stage-input-artifact",
    )


def test_save_stage_worker_uses_validated_input_artifact(tmp_path):
    source = tmp_path / "source.xlsx"
    staging = tmp_path / "staged.xlsx"
    _write_workbook(source)
    data = M.serialize_excel(str(source))

    result = anyio.run(_run_real_save_stage, data, staging)

    assert result["verification"]["preservation_ok"] is True
    assert staging.is_file()
    staging.unlink()


def test_execute_save_stage_preserves_injected_callable_contracts(tmp_path):
    staging = tmp_path / "custom-stage.xlsx"
    calls = {}

    def reconstruct_stub(session_data, path):
        calls["reconstruct"] = (session_data, path)
        Path(path).write_bytes(b"custom-stage")
        return []

    def signature_stub(before_path, after_path, intentional_edit):
        calls["signature"] = (before_path, str(after_path), intentional_edit)
        return {
            "present": False,
            "intentional_edit": intentional_edit,
            "status": "not_present",
            "parts_before": {},
            "parts_after": {},
            "parts_preserved": True,
            "message": "No signatures.",
        }

    def verify_stub(before_path, after_path, max_differences, requested_paths):
        calls["verify"] = (
            before_path,
            after_path,
            max_differences,
            requested_paths,
        )
        return {
            "preservation_ok": True,
            "equivalent": True,
            "change_count": 0,
            "unapproved_difference_count": 0,
            "classification_counts": {},
            "severity_counts": {},
            "changes": [],
            "truncated": False,
            "recommendation": "No differences.",
        }

    result = ST.execute_save_stage(
        {"sheets": [{"rows": []}]},
        staging_path=str(staging),
        verification_reference_path="reference.xlsx",
        verify_preservation=True,
        max_differences=25,
        requested_paths=["worksheets/Sheet/cells/A1/value"],
        intentional_edit=True,
        reconstruct=reconstruct_stub,
        verify=verify_stub,
        signature_report=signature_stub,
    )

    assert calls["signature"] == ("reference.xlsx", str(staging.resolve()), True)
    assert calls["verify"] == (
        "reference.xlsx",
        str(staging.resolve()),
        25,
        ["worksheets/Sheet/cells/A1/value"],
    )
    assert result["verification"]["preservation_ok"] is True
    assert result["inspection_performance"] is None


def test_public_mcp_save_names_are_async_wrappers():
    tools = M.mcp._tool_manager._tools
    assert tools["excel_save"].fn is M._excel_save_tool
    assert tools["excel_save_as_copy"].fn is M._excel_save_as_copy_tool


async def _run_public_async_save(session_key: str) -> str:
    return await M._excel_save_tool(
        session_key,
        report_format="json",
        verify_preservation=True,
        timeout_seconds=30.0,
    )


def test_public_async_save_runs_real_worker_transaction(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(
        session_key,
        "Sheet",
        [{"row_index": 0, "edits": {0: "async-new"}}],
    )

    report = json.loads(anyio.run(_run_public_async_save, session_key))

    assert report["verification"]["preservation_ok"] is True
    workbook = openpyxl.load_workbook(source)
    assert workbook.active["A1"].value == "async-new"
    workbook.close()
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS


async def _cancel_public_save_with_real_worker(
    session_key: str,
    destination: Path,
    monkeypatch,
) -> None:
    real_run_named_operation = run_named_operation

    async def slow_worker(operation, payload, **kwargs):
        del operation
        Path(payload["staging_path"]).write_bytes(b"partial-stage")
        return await real_run_named_operation(
            "_test_sleep",
            {"seconds": 30.0},
            timeout_seconds=20.0,
            operation_label=kwargs.get("operation_label") or "excel_save",
            failure_cleanup_paths=kwargs.get("failure_cleanup_paths") or (),
        )

    monkeypatch.setattr(M, "run_named_operation", slow_worker)

    async def invoke() -> None:
        try:
            await M._excel_save_tool(
                session_key,
                output_path=str(destination),
                timeout_seconds=20.0,
            )
        except WorkerProcessCancelled:
            pass

    async with anyio.create_task_group() as task_group:
        task_group.start_soon(invoke)
        with anyio.fail_after(10.0):
            while not active_worker_pids():
                await anyio.sleep(0.01)
        task_group.cancel_scope.cancel()


def test_host_cancellation_preserves_destination_with_real_worker(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "destination.xlsx"
    _write_workbook(source)
    _write_workbook(destination, first="target-old")
    original_destination = destination.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))

    anyio.run(
        _cancel_public_save_with_real_worker,
        session_key,
        destination,
        monkeypatch,
    )

    assert destination.read_bytes() == original_destination
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS
    assert not list(tmp_path.glob(".*.docloupe-*.xlsx"))


def test_backup_failure_preserves_destination_and_session_state(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    original_bytes = source.read_bytes()
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(
        session_key,
        "Sheet",
        [{"row_index": 0, "edits": {0: "new"}}],
    )

    def fail_backup(reference_path: str, saved_path: str):
        raise OSError(f"backup unavailable for {reference_path} -> {saved_path}")

    monkeypatch.setattr(ST, "create_excel_backup", fail_backup)
    with pytest.raises(OSError, match="backup unavailable"):
        M.excel_save(session_key, report_format="json")

    resolved = M._resolve_session_key(session_key)
    assert source.read_bytes() == original_bytes
    assert M._sessions[resolved]["_dirty_paths"]
    assert M._sessions[resolved].get("_verification_baseline_path") is None
    assert resolved not in M._BUSY_SESSIONS
    assert not list(tmp_path.glob(".*.docloupe-*.xlsx"))


def test_host_cancellation_to_new_path_leaves_no_output(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "new-output.xlsx"
    _write_workbook(source)
    session_key = _session_key(M.excel_load(str(source)))

    anyio.run(
        _cancel_public_save_with_real_worker,
        session_key,
        destination,
        monkeypatch,
    )

    assert not destination.exists()
    assert M._resolve_session_key(session_key) not in M._BUSY_SESSIONS
    assert not list(tmp_path.glob(".*.docloupe-*.xlsx"))
