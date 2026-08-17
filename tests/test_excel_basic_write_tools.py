import json
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


WRITE_PREFIXES = (
    "excel_set_",
    "excel_add_",
    "excel_update_",
    "excel_edit_",
    "excel_insert_",
    "excel_delete_",
    "excel_remove_",
    "excel_clear_",
    "excel_copy_",
    "excel_move_",
    "excel_merge_",
    "excel_fill_",
    "excel_apply_",
    "excel_upsert_",
    "excel_create_",
    "excel_rename_",
    "excel_freeze_",
    "excel_sort_",
    "excel_filter_",
)


def test_basic_fill_clear_freeze_and_dimension_tools_round_trip(tmp_path):
    output = tmp_path / "basic-write-tools.xlsx"
    session_key = json.loads(M.excel_create_workbook(sheet_names=["Data"]))["session_key"]
    try:
        M.excel_edit_cells(session_key, "Data", [{
            "row_index": 0,
            "edits": {0: "template", 1: "seed"},
        }])
        M.excel_set_style(
            session_key,
            "Data",
            0,
            1,
            style={"bold": True, "fill": "FFFFFF00"},
        )
        M.excel_fill_rows(session_key, "Data", template_row=0, after_index=0, count=3)
        M.excel_fill_column(
            session_key,
            "Data",
            col_index=0,
            start_row=0,
            end_row=3,
            sequence_start=0,
            step=2,
        )
        M.excel_fill_column(
            session_key,
            "Data",
            col_index=1,
            start_row=1,
            end_row=3,
            value="kept",
        )
        M.excel_clear_range(
            session_key,
            "Data",
            r1=2,
            c1=1,
            r2=2,
            c2=1,
            clear_values=True,
            clear_styles=False,
        )
        M.excel_clear_range(
            session_key,
            "Data",
            r1=3,
            c1=1,
            r2=3,
            c2=1,
            clear_values=False,
            clear_styles=True,
        )
        M.excel_freeze_panes(session_key, "Data", row=1, col=1)
        M.excel_set_dimension(session_key, "Data", axis="row", index=0, size=18)
        M.excel_set_row_height(session_key, "Data", {"1": 25, "2": None})
        M.excel_set_column_width(session_key, "Data", {"A": 10, "B": 12})
        M.excel_set_column_width(session_key, "Data", {"B": None})
        M.excel_set_dimension(session_key, "Data", axis="col", index=2, size=15)
        M.excel_set_dimension(session_key, "Data", axis="col", index=2, size=None)
        M.excel_save_as_copy(session_key, str(output))
    finally:
        M.excel_close(session_key)

    workbook = openpyxl.load_workbook(output)
    try:
        sheet = workbook["Data"]
        assert [sheet.cell(row=row, column=1).value for row in range(1, 5)] == [0, 2, 4, 6]
        assert sheet["B3"].value is None
        assert sheet["B3"].font.bold is True
        assert sheet["B4"].value == "kept"
        assert sheet["B4"].font.bold is False
        assert sheet.freeze_panes == "B2"
        assert sheet.row_dimensions[1].height == 18
        assert sheet.row_dimensions[2].height == 25
        assert sheet.row_dimensions[3].height is None
        assert sheet.column_dimensions["A"].width == 10
        assert "B" not in sheet.column_dimensions
        assert "C" not in sheet.column_dimensions
    finally:
        workbook.close()


def test_every_registered_write_tool_has_behavioral_test_reference():
    registered = set(M.mcp._tool_manager._tools)
    write_tools = sorted(name for name in registered if name.startswith(WRITE_PREFIXES))
    test_sources = "\n".join(
        path.read_text(encoding="utf-8", errors="ignore")
        for path in sorted((ROOT / "tests").glob("test_*.py"))
    )
    missing = [name for name in write_tools if name not in test_sources]

    assert len(write_tools) == 75
    assert missing == []


def test_major_semantic_groups_register_public_read_and_write_paths():
    registered = set(M.mcp._tool_manager._tools)
    pairs = {
        "workbook": ("excel_get_workbook_semantics", "excel_set_workbook_properties"),
        "worksheet": ("excel_get_sheet_semantics", "excel_set_sheet_properties"),
        "cell": ("excel_get_cell", "excel_edit_cells"),
        "rich_text": ("excel_get_rich_text", "excel_edit_rich_text"),
        "formula": ("excel_get_cell", "excel_set_formula"),
        "style": ("excel_get_cell", "excel_set_cell_style_semantics"),
        "defined_names": ("excel_list_defined_names", "excel_add_defined_name"),
        "tables": ("excel_list_tables", "excel_add_table"),
        "conditional_formatting": (
            "excel_get_conditional_formats",
            "excel_add_conditional_format",
        ),
        "drawings": ("excel_get_shapes", "excel_add_shape"),
        "package": ("excel_read_package_part", "excel_apply_package_transaction"),
    }

    assert {
        group: pair
        for group, pair in pairs.items()
        if pair[0] not in registered or pair[1] not in registered
    } == {}


def test_freeze_panes_synchronizes_first_view_without_resetting_selections(tmp_path):
    output = tmp_path / "freeze-view-sync.xlsx"
    session_key = json.loads(M.excel_create_workbook(sheet_names=["Data"]))["session_key"]
    selections = [{"activeCell": "D4", "sqref": "D4:E5"}]
    try:
        M.excel_set_sheet_views(
            session_key,
            "Data",
            [{"workbookViewId": 0, "topLeftCell": "C3", "selections": selections}],
        )

        M.excel_freeze_panes(session_key, "Data", row=2, col=0)
        row_view = json.loads(M.excel_get_sheet_views(session_key, "Data"))["views"][0]
        assert row_view["pane"] == {
            "ySplit": 2,
            "topLeftCell": "A3",
            "activePane": "bottomLeft",
            "state": "frozen",
        }
        assert row_view["selections"] == selections

        M.excel_freeze_panes(session_key, "Data", row=0, col=2)
        col_view = json.loads(M.excel_get_sheet_views(session_key, "Data"))["views"][0]
        assert col_view["pane"] == {
            "xSplit": 2,
            "topLeftCell": "C1",
            "activePane": "topRight",
            "state": "frozen",
        }
        assert col_view["selections"] == selections

        M.excel_freeze_panes(session_key, "Data", row=2, col=2)
        both_view = json.loads(M.excel_get_sheet_views(session_key, "Data"))["views"][0]
        assert both_view["pane"] == {
            "xSplit": 2,
            "ySplit": 2,
            "topLeftCell": "C3",
            "activePane": "bottomRight",
            "state": "frozen",
        }
        assert both_view["selections"] == selections

        M.excel_freeze_panes(session_key, "Data", row=0, col=0)
        unfrozen_view = json.loads(M.excel_get_sheet_views(session_key, "Data"))["views"][0]
        assert "pane" not in unfrozen_view
        assert unfrozen_view["selections"] == selections

        M.excel_freeze_panes(session_key, "Data", row=2, col=2)
        M.excel_save_as_copy(session_key, str(output))
    finally:
        M.excel_close(session_key)

    workbook = openpyxl.load_workbook(output)
    try:
        sheet = workbook["Data"]
        assert sheet.freeze_panes == "C3"
        assert sheet.sheet_view.selection[0].activeCell == "D4"
        assert str(sheet.sheet_view.selection[0].sqref) == "D4:E5"
    finally:
        workbook.close()
