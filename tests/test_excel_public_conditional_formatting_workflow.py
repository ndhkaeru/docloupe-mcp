import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _new_session():
    return json.loads(M.excel_create_workbook(sheet_names=["S"]))["session_key"]


def _load_key(path: Path) -> str:
    return M.excel_load(str(path)).split("session_key='")[1].split("'")[0]


def _rules_by_sqref(worksheet):
    return {
        str(item.sqref): list(worksheet.conditional_formatting[item])
        for item in worksheet.conditional_formatting
    }


def _inject_cf_extension(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        payloads = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    sheet_path = "xl/worksheets/sheet1.xml"
    xml = payloads[sheet_path].decode("utf-8")
    marker = "</cfRule>"
    extension = '<extLst><ext uri="{DOCLOUPE-CF-TEST}"><x:payload xmlns:x="urn:docloupe:test">keep</x:payload></ext></extLst>'
    xml = xml.replace(marker, extension + marker, 1)
    payloads[sheet_path] = xml.encode("utf-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, payload in payloads.items():
            archive.writestr(name, payload)


def test_public_conditional_formatting_types_and_dxfs_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "conditional-formatting.xlsx"

    M.excel_add_conditional_format(key, "S", "A1:A5", {
        "type": "cellIs",
        "operator": "greaterThan",
        "formula": ["3"],
        "stopIfTrue": True,
        "dxf": {"fill": "FFFF0000"},
    })
    M.excel_add_conditional_format(key, "S", "B1:B5", {
        "type": "formula",
        "formula": ["B1>0"],
        "dxf": {"font": {"bold": True, "color": "FF00AA00"}},
    })
    M.excel_add_conditional_format(key, "S", "C1:C5", {
        "type": "colorScale",
        "colorScale": {
            "cfvo": [{"type": "min"}, {"type": "max"}],
            "colors": ["FFFF0000", "FF00FF00"],
        },
    })
    M.excel_add_conditional_format(key, "S", "D1:D5", {
        "type": "dataBar",
        "dataBar": {
            "cfvo": [{"type": "min"}, {"type": "max"}],
            "color": "FF638EC6",
            "showValue": False,
        },
    })
    M.excel_add_conditional_format(key, "S", "E1:E5", {
        "type": "iconSet",
        "iconSet": {
            "iconSet": "3TrafficLights1",
            "showValue": True,
            "reverse": True,
            "cfvo": [
                {"type": "percent", "val": 0},
                {"type": "percent", "val": 33},
                {"type": "percent", "val": 67},
            ],
        },
    })

    listed = json.loads(M.excel_get_conditional_formats(key, "S"))
    assert listed["count"] == 5
    assert [rule["type"] for rule in listed["rules"]] == [
        "cellIs", "expression", "colorScale", "dataBar", "iconSet",
    ]
    assert [rule["priority"] for rule in listed["rules"]] == [1, 2, 3, 4, 5]
    assert listed["rules"][0]["dxfId"] == 0
    assert listed["rules"][1]["dxfId"] == 1

    M.excel_save(key, str(out))
    assert json.loads(M.excel_validate_workbook(str(out)))["valid"] is True
    workbook = openpyxl.load_workbook(out, data_only=False)
    worksheet = workbook["S"]
    rules = _rules_by_sqref(worksheet)
    assert set(rules) == {"A1:A5", "B1:B5", "C1:C5", "D1:D5", "E1:E5"}
    assert rules["A1:A5"][0].type == "cellIs"
    assert rules["A1:A5"][0].operator == "greaterThan"
    assert rules["A1:A5"][0].formula == ["3"]
    assert rules["A1:A5"][0].stopIfTrue is True
    assert rules["B1:B5"][0].type == "expression"
    assert rules["B1:B5"][0].formula == ["B1>0"]
    assert len(rules["C1:C5"][0].colorScale.cfvo) == 2
    assert rules["D1:D5"][0].dataBar.showValue is False
    assert rules["E1:E5"][0].iconSet.iconSet == "3TrafficLights1"
    assert rules["E1:E5"][0].iconSet.reverse is True
    assert workbook._differential_styles.styles[0].fill.fgColor.rgb == "FFFF0000"
    assert workbook._differential_styles.styles[1].font.bold is True
    assert workbook._differential_styles.styles[1].font.color.rgb == "FF00AA00"
    workbook.close()

    reloaded = _load_key(out)
    listed_after = json.loads(M.excel_get_conditional_formats(reloaded, "S"))
    assert listed_after["count"] == 5
    assert listed_after["rules"][3]["dataBar"]["showValue"] is False
    assert listed_after["rules"][4]["iconSet"]["reverse"] is True


def test_conditional_formatting_patch_priority_move_and_delete(tmp_path):
    key = _new_session()
    out = tmp_path / "conditional-formatting-crud.xlsx"

    first = json.loads(M.excel_add_conditional_format(key, "S", "A1:A5", {
        "type": "expression",
        "formula": ["A1>0"],
        "dxf": {"fill": "FFFFFF00"},
    }))
    second = json.loads(M.excel_add_conditional_format(key, "S", "B1:B5", {
        "type": "cellIs",
        "operator": "lessThan",
        "formula": ["10"],
        "priority": 1,
        "dxf": {"font": {"italic": True}},
    }))
    assert first["rule_id"] == "b0:r0"
    assert second["rule_id"] == "b1:r0"

    rules = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    by_range = {rule["sqref"]: rule for rule in rules}
    assert by_range["B1:B5"]["priority"] == 1
    assert by_range["A1:A5"]["priority"] == 2

    updated = json.loads(M.excel_update_conditional_format(key, "S", by_range["A1:A5"]["rule_id"], {
        "priority": 1,
        "sqref": "C1:C5",
        "formula": ["C1>=5"],
        "stop_if_true": True,
    }))
    assert updated["rule_id"] == "b1:r0"
    rules = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    by_range = {rule["sqref"]: rule for rule in rules}
    assert by_range["C1:C5"]["priority"] == 1
    assert by_range["C1:C5"]["formula"] == ["C1>=5"]
    assert by_range["C1:C5"]["stopIfTrue"] is True
    assert by_range["B1:B5"]["priority"] == 2

    M.excel_delete_conditional_format(key, "S", by_range["B1:B5"]["rule_id"])
    remaining = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    assert len(remaining) == 1
    assert remaining[0]["sqref"] == "C1:C5"
    assert remaining[0]["priority"] == 1

    M.excel_save(key, str(out))
    workbook = openpyxl.load_workbook(out)
    rules = _rules_by_sqref(workbook["S"])
    assert set(rules) == {"C1:C5"}
    assert rules["C1:C5"][0].formula == ["C1>=5"]
    assert rules["C1:C5"][0].stopIfTrue is True
    workbook.close()


def test_conditional_formatting_dxf_dedup_and_copy_on_update(tmp_path):
    key = _new_session()
    out = tmp_path / "conditional-formatting-dxfs.xlsx"
    shared_dxf = {"font": {"bold": True, "color": "FF112233"}}

    M.excel_add_conditional_format(key, "S", "A1:A5", {
        "type": "expression", "formula": ["A1>0"], "dxf": shared_dxf,
    })
    M.excel_add_conditional_format(key, "S", "B1:B5", {
        "type": "expression", "formula": ["B1>0"], "dxf": shared_dxf,
    })
    rules = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    assert [rule["dxfId"] for rule in rules] == [0, 0]
    assert rules[0]["dxf"]["font"]["bold"] is True
    assert rules[0]["dxf"]["font"]["color"]["rgb"] == "FF112233"

    M.excel_update_conditional_format(key, "S", rules[0]["rule_id"], {
        "dxf": {"fill": "FF00FF00"},
    })
    rules = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    by_range = {rule["sqref"]: rule for rule in rules}
    assert by_range["A1:A5"]["dxfId"] == 1
    assert by_range["B1:B5"]["dxfId"] == 0
    assert by_range["A1:A5"]["dxf"]["font"]["bold"] is True
    assert by_range["A1:A5"]["dxf"]["fill"]["fgColor"]["rgb"] == "FF00FF00"
    assert "fill" not in by_range["B1:B5"]["dxf"]

    M.excel_update_conditional_format(key, "S", by_range["A1:A5"]["rule_id"], {
        "dxf": None,
        "clear_nulls": True,
    })
    rules = json.loads(M.excel_get_conditional_formats(key, "S"))["rules"]
    by_range = {rule["sqref"]: rule for rule in rules}
    assert "dxfId" not in by_range["A1:A5"]
    assert by_range["B1:B5"]["dxfId"] == 0

    M.excel_save(key, str(out))
    workbook = openpyxl.load_workbook(out)
    saved = _rules_by_sqref(workbook["S"])
    assert saved["A1:A5"][0].dxfId is None
    assert saved["B1:B5"][0].dxfId == 0
    assert len(workbook._differential_styles.styles) == 2
    workbook.close()


def test_conditional_formatting_patch_preserves_raw_extension(tmp_path):
    source = tmp_path / "conditional-formatting-extension-source.xlsx"
    output = tmp_path / "conditional-formatting-extension-output.xlsx"
    key = _new_session()
    M.excel_add_conditional_format(key, "S", "A1:A5", {
        "type": "expression",
        "formula": ["A1>0"],
        "dxf": {"fill": "FFFF0000"},
    })
    M.excel_save(key, str(source))
    _inject_cf_extension(source)

    loaded = _load_key(source)
    rule = json.loads(M.excel_get_conditional_formats(loaded, "S", include_raw_xml=True))["rules"][0]
    assert "DOCLOUPE-CF-TEST" in rule["raw_xml"]
    M.excel_update_conditional_format(loaded, "S", rule["rule_id"], {
        "stopIfTrue": True,
    })
    M.excel_save(loaded, str(output))

    with zipfile.ZipFile(output, "r") as archive:
        worksheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "DOCLOUPE-CF-TEST" in worksheet_xml
    assert "urn:docloupe:test" in worksheet_xml
    assert ">keep<" in worksheet_xml
    assert 'stopIfTrue="1"' in worksheet_xml
    assert "A1&gt;0" in worksheet_xml or "A1>0" in worksheet_xml


def test_conditional_formatting_invalid_updates_are_atomic():
    key = _new_session()
    M.excel_add_conditional_format(key, "S", "A1:A5", {
        "type": "cellIs",
        "operator": "between",
        "formula": ["1", "5"],
        "dxf": {"fill": "FFFFFF00"},
    })
    before = json.loads(M.excel_get_conditional_formats(key, "S"))
    rule_id = before["rules"][0]["rule_id"]

    with pytest.raises(ValueError, match="requires 2 formula"):
        M.excel_update_conditional_format(key, "S", rule_id, {"formula": ["2"]})
    with pytest.raises(ValueError, match="requires exactly two cfvo"):
        M.excel_add_conditional_format(key, "S", "B1:B5", {
            "type": "dataBar",
            "dataBar": {"cfvo": [{"type": "min"}], "color": "FF0000FF"},
        })
    with pytest.raises(ValueError, match="priority must be between"):
        M.excel_update_conditional_format(key, "S", rule_id, {"priority": 9})

    after = json.loads(M.excel_get_conditional_formats(key, "S"))
    assert after == before
