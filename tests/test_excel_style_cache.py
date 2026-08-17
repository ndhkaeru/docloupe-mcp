import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.colors import Color


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import core  # noqa: E402


def _write_repeated_style_workbook(path: Path, cell_count: int = 128) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Styles"
    accent = NamedStyle(
        name="Accent",
        font=Font(name="Calibri", size=12, bold=True, color=Color(theme=1, tint=0.25)),
        fill=PatternFill(fill_type="solid", fgColor=Color(theme=4, tint=-0.2)),
        border=Border(bottom=Side(style="thin", color=Color(theme=2))),
        alignment=Alignment(horizontal="center", wrap_text=True),
        number_format="0.00",
    )
    workbook.add_named_style(accent)
    for column in range(1, cell_count + 1):
        cell = worksheet.cell(row=1, column=column, value=column)
        if column % 2:
            cell.style = "Accent"
    workbook.save(path)
    workbook.close()


def test_repeated_styles_are_built_once_per_style_key(tmp_path, monkeypatch):
    source = tmp_path / "style-cache.xlsx"
    _write_repeated_style_workbook(source)

    original_builder = core._build_cached_cell_style
    built_keys = []

    def counted_builder(cell, style_id, style_semantics, named_style_names, theme_colors):
        built_keys.append(core._style_cache_key(cell, style_id))
        return original_builder(cell, style_id, style_semantics, named_style_names, theme_colors)

    monkeypatch.setattr(core, "_build_cached_cell_style", counted_builder)
    data = core.serialize_excel(str(source))

    cells = data["sheets"][0]["rows"][0]["cells"]
    assert len(cells) == 128
    assert len(built_keys) == len(set(built_keys)) == 2
    assert cells[0]["alignment"] is cells[2]["alignment"]
    assert cells[0]["xf"] is cells[2]["xf"]
    assert cells[1]["alignment"] is cells[3]["alignment"]
    assert cells[1]["xf"] is cells[3]["xf"]


def test_raw_fill_metadata_stays_per_cell_while_style_semantics_are_shared(tmp_path):
    source = tmp_path / "style-cache-raw-fill.xlsx"
    _write_repeated_style_workbook(source, cell_count=3)

    data = core.serialize_excel(str(source))
    first, _, third = data["sheets"][0]["rows"][0]["cells"]

    assert first["fill"] == third["fill"]
    assert first["fill_color"] is third["fill_color"]
    assert first["_font_raw"] is third["_font_raw"]
    assert first["_fill_raw"] == third["_fill_raw"]
    assert first["_fill_raw"] is not third["_fill_raw"]
