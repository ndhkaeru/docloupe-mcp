import json
import sys
import uuid
import xml.etree.ElementTree as ET
import zipfile
from contextlib import contextmanager
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import core  # noqa: E402
import main as M  # noqa: E402


_BASELINE_KEYS = (
    "_baseline_content_hash",
    "_baseline_style_hash",
    "_baseline_structure_hash",
)


def _write_sparse_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sparse"
    worksheet["A1"] = "start"
    worksheet["D4"] = "merged"
    worksheet.merge_cells("D4:F5")
    worksheet["J10"] = "edge"
    worksheet.row_dimensions[7].height = 31
    worksheet.row_dimensions[7].hidden = True
    worksheet.row_dimensions[7].outlineLevel = 2
    worksheet.column_dimensions["H"].width = 22
    worksheet.column_dimensions["H"].hidden = True
    worksheet.column_dimensions["H"].outlineLevel = 1

    other = workbook.create_sheet("Other")
    other["A1"] = "untouched"
    other["F6"] = "far"
    workbook.save(path)
    workbook.close()
    _inject_explicit_blank(path, "xl/worksheets/sheet1.xml", "B2")


def _inject_explicit_blank(path: Path, part_name: str, coordinate: str) -> None:
    replacement = path.with_name(path.stem + ".explicit-blank.xlsx")
    row_number = int("".join(character for character in coordinate if character.isdigit()))
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w") as target:
        for item in source.infolist():
            raw = source.read(item.filename)
            if item.filename == part_name:
                root = ET.fromstring(raw)
                namespace = root.tag.partition("}")[0].lstrip("{")
                ET.register_namespace("", namespace)
                qname = lambda name: f"{{{namespace}}}{name}"
                sheet_data = root.find(qname("sheetData"))
                rows = list(sheet_data)
                row = next((candidate for candidate in rows if int(candidate.get("r", "0")) == row_number), None)
                if row is None:
                    row = ET.Element(qname("row"), {"r": str(row_number)})
                    insert_at = next(
                        (index for index, candidate in enumerate(rows) if int(candidate.get("r", "0")) > row_number),
                        len(rows),
                    )
                    sheet_data.insert(insert_at, row)
                ET.SubElement(row, qname("c"), {"r": coordinate})
                raw = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(item, raw)
    replacement.replace(path)


def _cell_coordinates(path: Path, part_name: str = "xl/worksheets/sheet1.xml") -> set[str]:
    with zipfile.ZipFile(path, "r") as archive:
        root = ET.fromstring(archive.read(part_name))
    return {
        element.get("r")
        for element in root.iter()
        if element.tag.rsplit("}", 1)[-1] == "c" and element.get("r")
    }


def _part_bytes(path: Path, part_name: str) -> bytes:
    with zipfile.ZipFile(path, "r") as archive:
        return archive.read(part_name)


def _merged_coordinates(range_ref: str) -> set[str]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    return {
        f"{openpyxl.utils.get_column_letter(column)}{row}"
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    }


def _session_key(result: str) -> str:
    return result.split("session_key=", 1)[1].split(" |", 1)[0].strip("'")


@contextmanager
def _installed_session(data: dict):
    session_key = f"implicit-test-{uuid.uuid4().hex}"
    M._sessions[session_key] = data
    try:
        yield session_key
    finally:
        M._sessions.pop(session_key, None)


def test_sparse_serialization_uses_placeholders_and_preserves_grid_semantics(tmp_path):
    source = tmp_path / "sparse.xlsx"
    _write_sparse_fixture(source)

    data = core.serialize_excel(str(source))
    sheet = data["sheets"][0]
    cells = [cell for row in sheet["rows"] for cell in row["cells"]]
    explicit_coordinates = _cell_coordinates(source)
    materialized_coordinates = explicit_coordinates | _merged_coordinates("D4:F5")

    assert len(sheet["rows"]) == 10
    assert {len(row["cells"]) for row in sheet["rows"]} == {10}
    assert sum(bool(cell.get("_implicit")) for cell in cells) == 100 - len(materialized_coordinates)
    assert all(
        cell == {"_implicit": True, "v": None, "merge": {}}
        for cell in cells
        if cell.get("_implicit")
    )
    assert sheet["rows"][1]["cells"][1]["present"] is True
    assert sheet["rows"][1]["cells"][1]["v"] is None
    assert not sheet["rows"][1]["cells"][1].get("_implicit")
    assert sheet["rows"][3]["cells"][3]["merge"]["rowspan"] == 2
    assert sheet["rows"][4]["cells"][5]["merge"] == "slave"
    assert sheet["rows"][6]["h"] == 31
    assert sheet["rows"][6]["hidden"] is True
    assert sheet["rows"][6]["outline"] == 2
    assert sheet["cw"]["H"] == 22
    assert sheet["ch"]["H"] is True
    assert sheet["co"]["H"] == 1


def test_public_reads_expand_implicit_cells_without_leaking_marker(tmp_path):
    source = tmp_path / "public-reads.xlsx"
    _write_sparse_fixture(source)
    data = core.serialize_excel(str(source))

    with _installed_session(data) as session_key:
        cell = json.loads(M.excel_get_cell(session_key, "Sparse", 2, 2))
        column = json.loads(M.excel_get_column(session_key, "Sparse", 2, 2, 4))
        rows = json.loads(M.excel_get_rows(session_key, "Sparse", 2, 3))
        values = json.loads(M.excel_read_range(session_key, "Sparse", "B7:C8"))
        cloned = json.loads(M.excel_clone_rows(session_key, "Sparse", 2))

    assert cell["v"] is None
    assert cell["value"] is None
    assert cell["data_type"] == "n"
    assert cell["present"] is False
    assert cell["font"] == "Calibri"
    assert cell["size"] == 11
    assert cell["named_style"] == "Normal"
    assert cell["xf"]["style_id"] == 0
    assert "_implicit" not in json.dumps(cell)
    assert "_implicit" not in json.dumps(column)
    assert "_implicit" not in json.dumps(rows)
    assert "_implicit" not in json.dumps(cloned)
    assert values["values"] == [[None, None], [None, None]]
    assert cloned[0]["cells"][2]["font"] == "Calibri"


def test_implicit_cells_promote_for_edit_style_fill_and_raw_model_save(tmp_path):
    source = tmp_path / "promote.xlsx"
    output = tmp_path / "promote-output.xlsx"
    _write_sparse_fixture(source)
    data = core.serialize_excel(str(source))
    sheet = data["sheets"][0]

    with _installed_session(data) as session_key:
        M.excel_edit_cells(session_key, "Sparse", [{"row_index": 2, "edits": {2: "edited"}}])
        M.excel_set_style(session_key, "Sparse", 2, 3, style={"bold": True, "fill": "FFFFFF00"})
        M.excel_fill_column(session_key, "Sparse", 1, 6, 7, value="filled")

    raw_model_cell = sheet["rows"][8]["cells"][8]
    raw_model_cell["v"] = "raw-edit"
    raw_model_cell["present"] = True

    for row_index, col_index in ((2, 2), (2, 3), (6, 1), (7, 1)):
        cell = sheet["rows"][row_index]["cells"][col_index]
        assert not cell.get("_implicit")
        assert cell["font"] == "Calibri"
        assert all(key in cell for key in _BASELINE_KEYS)

    assert raw_model_cell.get("_implicit") is True
    core.reconstruct_excel(data, str(output))

    workbook = openpyxl.load_workbook(output)
    worksheet = workbook["Sparse"]
    assert worksheet["C3"].value == "edited"
    assert worksheet["D3"].font.bold is True
    assert worksheet["D3"].fill.fgColor.rgb == "FFFFFF00"
    assert worksheet["B7"].value == "filled"
    assert worksheet["B8"].value == "filled"
    assert worksheet["I9"].value == "raw-edit"
    assert worksheet.row_dimensions[7].height == 31
    assert worksheet.column_dimensions["H"].width == 22
    workbook.close()
    assert "B2" in _cell_coordinates(output)


def test_merge_and_unmerge_across_implicit_cells(tmp_path):
    source = tmp_path / "merge.xlsx"
    merged_output = tmp_path / "merged-output.xlsx"
    unmerged_output = tmp_path / "unmerged-output.xlsx"
    _write_sparse_fixture(source)
    data = core.serialize_excel(str(source))

    with _installed_session(data) as session_key:
        M.excel_merge_cells(session_key, "Sparse", 6, 1, 7, 2)
        core.reconstruct_excel(data, str(merged_output))
        M.excel_merge_cells(session_key, "Sparse", 6, 1, unmerge=True)
        core.reconstruct_excel(data, str(unmerged_output))

    merged = openpyxl.load_workbook(merged_output)
    assert "B7:C8" in {str(item) for item in merged["Sparse"].merged_cells.ranges}
    merged.close()
    unmerged = openpyxl.load_workbook(unmerged_output)
    assert "B7:C8" not in {str(item) for item in unmerged["Sparse"].merged_cells.ranges}
    unmerged.close()


def test_no_edit_reconstruction_remains_byte_identical(tmp_path):
    source = tmp_path / "exact-copy.xlsx"
    output = tmp_path / "exact-copy-output.xlsx"
    _write_sparse_fixture(source)
    original = source.read_bytes()

    data = core.serialize_excel(str(source))
    core.reconstruct_excel(data, str(output))

    assert output.read_bytes() == original


def test_filtered_sheet_save_preserves_unloaded_sheet_part(tmp_path):
    source = tmp_path / "filtered.xlsx"
    output = tmp_path / "filtered-output.xlsx"
    _write_sparse_fixture(source)
    source_hash = core._sha256_file(source)
    original_other_sheet = _part_bytes(source, "xl/worksheets/sheet2.xml")

    session_key = _session_key(M.excel_load(str(source), sheet_name="Sparse"))
    try:
        M.excel_edit_cells(session_key, "Sparse", [{"row_index": 2, "edits": {2: "filtered-edit"}}])
        data = M._sessions[M._resolve_session_key(session_key)]
        core.reconstruct_excel(data, str(output))
    finally:
        M._sessions.pop(M._resolve_session_key(session_key), None)

    workbook = openpyxl.load_workbook(output)
    assert workbook["Sparse"]["C3"].value == "filtered-edit"
    assert workbook["Other"]["A1"].value == "untouched"
    assert workbook["Other"]["F6"].value == "far"
    workbook.close()
    assert _part_bytes(output, "xl/worksheets/sheet2.xml") == original_other_sheet
    assert core._sha256_file(source) == source_hash


def test_content_only_implicit_edit_does_not_build_temporary_workbook(tmp_path, monkeypatch):
    source = tmp_path / "content-only.xlsx"
    output = tmp_path / "content-only-output.xlsx"
    _write_sparse_fixture(source)
    data = core.serialize_excel(str(source))

    with _installed_session(data) as session_key:
        M.excel_edit_cells(session_key, "Sparse", [{"row_index": 2, "edits": {2: "direct patch"}}])

    changes = core._content_only_changes(data)
    assert changes == {"Sparse": ["C3"]}

    def forbidden_reconstruction(*_args, **_kwargs):
        raise AssertionError("content-only save must not reconstruct a temporary workbook")

    monkeypatch.setattr(core, "reconstruct_excel", forbidden_reconstruction)
    warnings = core._reconstruct_content_only(data, str(output), changes)

    assert warnings == []
    workbook = openpyxl.load_workbook(output)
    assert workbook["Sparse"]["C3"].value == "direct patch"
    assert workbook["Other"]["A1"].value == "untouched"
    workbook.close()
    assert not Path(str(output) + ".~content-generated.xlsx").exists()
    assert not Path(str(output) + ".~saving.tmp").exists()


def test_find_cell_xml_stops_at_target_self_closing_cell():
    sheet_xml = (
        '<worksheet><sheetData><row r="4">'
        '<c r="A4" s="16"/><c r="B4" s="16"/><c r="C4" s="16"><v>keep</v></c>'
        '</row></sheetData></worksheet>'
    )
    match, source_cell = core._find_cell_xml(sheet_xml, "A4")

    assert source_cell == '<c r="A4" s="16"/>'
    generated = core._cell_xml_fragment(
        "A4",
        {"v": "edited", "merge": {}, "present": True},
        None,
        include_scalar=True,
    )
    merged = core._merge_generated_cell_content(source_cell, generated)
    patched = sheet_xml[:match.start()] + merged + sheet_xml[match.end():]

    root = ET.fromstring(patched)
    cells = [element for element in root.iter() if element.tag == "c"]
    assert [cell.get("r") for cell in cells] == ["A4", "B4", "C4"]
    assert cells[0].find("is/t").text == "edited"
    assert cells[2].find("v").text == "keep"
