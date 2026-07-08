"""
Regression tests for the MCP tool layer (main.py):

A1. Sheet-filtered load + save merges unloaded sheets back (no data loss).
A2. Macro-enabled OOXML workbooks can be edited while preserving VBA parts.
A3. A failing save never corrupts the existing file (atomic write).
B1. Structural row/column edits shift all coordinate-anchored metadata:
    merges, hyperlinks, comments, data validations, conditional formatting,
    auto filter, freeze panes, hidden columns.
B2. Renaming a sheet updates defined names referencing it.
B3. Overlapping merges are rejected; editing a slave cell is rejected.
C.  Internal hyperlinks, docProps and workbook view survive a round-trip;
    private keys are stripped from read-tool output.
"""
import json
import re
import sys
import types
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
from core import reconstruct_excel, serialize_excel  # noqa: E402


def _make_rich_sheet(path: Path) -> None:
    """One sheet with every kind of coordinate-anchored metadata."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(1, 11):
        for c in range(1, 7):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    ws["B5"].hyperlink = "https://example.com/"
    ws["C6"].comment = Comment("note", "tester")
    dv = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    dv.add("D2:D10")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add(
        "A2:A10",
        CellIsRule(operator="equal", formula=['"x"'],
                   fill=PatternFill("solid", start_color="FFFFC7CE", end_color="FFFFC7CE")))
    ws.merge_cells("A8:B9")
    ws.auto_filter.ref = "A1:F10"
    ws.freeze_panes = "A5"
    ws.column_dimensions["E"].hidden = True
    wb.save(path)


def test_font_color_and_strike_shortcuts_roundtrip(tmp_path):
    src = tmp_path / "src.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "changed"
    wb.save(src)

    key = M.excel_load(str(src)).split("session_key='")[1].split("'")[0]
    M.excel_set_font_color(key, "Sheet", 0, 0, "FF00AA00")
    M.excel_set_strike(key, "Sheet", 0, 0, True)
    M.excel_save(key)

    wb2 = openpyxl.load_workbook(src)
    cell = wb2.active["A1"]
    assert cell.font.color.rgb == "FF00AA00"
    assert cell.font.strike is True


def test_validate_and_diff_package_tools(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src)
    wb.save(out)

    report = json.loads(M.excel_validate_workbook(str(src)))
    assert report["valid"] is True
    diff = json.loads(M.excel_diff_package(str(src), str(out)))
    assert diff["changed"] is False


def test_save_as_copy_does_not_overwrite_source(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "copy.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "old"
    wb.save(src)

    key = M.excel_load(str(src)).split("session_key='")[1].split("'")[0]
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "new"}}])
    M.excel_save_as_copy(key, str(out))

    assert openpyxl.load_workbook(src).active["A1"].value == "old"
    assert openpyxl.load_workbook(out).active["A1"].value == "new"


def test_shape_inventory_and_text_update_from_session_data(tmp_path):
    drawing_xml = '''<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="TextBox 1"/></xdr:nvSpPr><xdr:txBody><a:p><a:r><a:t>Old</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'''
    key = str(tmp_path / "fake.xlsx")
    M._sessions[key] = {
        "source": key,
        "sheets": [{
            "name": "Sheet1",
            "rows": [],
            "drawing_data": {"drawing_file": "xl/drawings/drawing1.xml", "drawing_xml": drawing_xml, "drawing_rels": None, "files": {}},
            "shapes": [{"index": 1, "id": "2", "name": "TextBox 1", "type": "shape", "text": "Old", "relationship_id": None}],
        }],
    }

    shapes = json.loads(M.excel_get_shapes(key, "Sheet1"))
    assert shapes["Sheet1"][0]["text"] == "Old"
    M.excel_update_shape_text(key, "Sheet1", 1, "New & Better")
    sheet = M._sessions[key]["sheets"][0]
    assert sheet["shapes"][0]["text"] == "New & Better"
    assert "New &amp; Better" in sheet["drawing_data"]["drawing_xml"]


def test_shape_style_update_from_session_data(tmp_path):
    drawing_xml = '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="TextBox 1"/></xdr:nvSpPr><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr><xdr:txBody><a:p><a:r><a:rPr/><a:t>Old</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'
    key = str(tmp_path / "fake-style.xlsx")
    M._sessions[key] = {
        "source": key,
        "sheets": [{
            "name": "Sheet1",
            "rows": [],
            "drawing_data": {"drawing_file": "xl/drawings/drawing1.xml", "drawing_xml": drawing_xml, "drawing_rels": None, "files": {}},
            "shapes": [{"index": 1, "id": "2", "name": "TextBox 1", "type": "shape", "text": "Old", "relationship_id": None}],
        }],
    }

    M.excel_set_shape_style(key, "Sheet1", 1, fill_color="FFCCFFFF", outline_color="FFFF0000", outline_width_pt=2, text_color="FF0000FF")

    sheet = M._sessions[key]["sheets"][0]
    xml = sheet["drawing_data"]["drawing_xml"]
    assert re.search(r'<a:srgbClr val="CCFFFF"\s*/>', xml)
    assert re.search(r'<a:ln w="25400">', xml)
    assert re.search(r'<a:srgbClr val="FF0000"\s*/>', xml)
    assert re.search(r'<a:srgbClr val="0000FF"\s*/>', xml)
    assert sheet["shapes"][0]["outline_width_pt"] == 2


def test_shape_fill_does_not_consume_outline_xml(tmp_path):
    drawing_xml = '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="Shape 1"/></xdr:nvSpPr><xdr:spPr><a:noFill/><a:ln><a:solidFill><a:srgbClr val="FF0000"/></a:solidFill></a:ln></xdr:spPr></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'
    key = str(tmp_path / "fake-fill.xlsx")
    M._sessions[key] = {"source": key, "sheets": [{"name": "Sheet1", "rows": [], "drawing_data": {"drawing_file": "xl/drawings/drawing1.xml", "drawing_xml": drawing_xml, "drawing_rels": None, "files": {}}, "shapes": [{"index": 1, "id": "2", "name": "Shape 1", "type": "shape", "text": None, "relationship_id": None}]}]}

    M.excel_set_shape_style(key, "Sheet1", 1, fill_color="00FF00")

    xml = M._sessions[key]["sheets"][0]["drawing_data"]["drawing_xml"]
    assert re.search(r'<a:solidFill><a:srgbClr val="00FF00"\s*/></a:solidFill>', xml)
    assert re.search(r'<a:ln><a:solidFill><a:srgbClr val="FF0000"\s*/></a:solidFill></a:ln>', xml)


def test_shape_outline_width_preserves_existing_outline_color(tmp_path):
    drawing_xml = '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="Shape 1"/></xdr:nvSpPr><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:ln><a:solidFill><a:srgbClr val="FF0000"/></a:solidFill></a:ln></xdr:spPr></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'
    key = str(tmp_path / "fake-line.xlsx")
    M._sessions[key] = {"source": key, "sheets": [{"name": "Sheet1", "rows": [], "drawing_data": {"drawing_file": "xl/drawings/drawing1.xml", "drawing_xml": drawing_xml, "drawing_rels": None, "files": {}}, "shapes": [{"index": 1, "id": "2", "name": "Shape 1", "type": "shape", "text": None, "relationship_id": None}]}]}

    M.excel_set_shape_style(key, "Sheet1", 1, outline_width_pt=3)

    xml = M._sessions[key]["sheets"][0]["drawing_data"]["drawing_xml"]
    assert re.search(r'<a:ln w="38100"><a:solidFill><a:srgbClr val="FF0000"\s*/></a:solidFill></a:ln>', xml)


def test_shape_clear_fill_and_outline_flags_work_through_tool(tmp_path):
    drawing_xml = '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="Shape 1"/></xdr:nvSpPr><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:solidFill><a:srgbClr val="00FF00"/></a:solidFill><a:ln><a:solidFill><a:srgbClr val="FF0000"/></a:solidFill></a:ln></xdr:spPr></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>'
    key = str(tmp_path / "fake-clear.xlsx")
    M._sessions[key] = {"source": key, "sheets": [{"name": "Sheet1", "rows": [], "drawing_data": {"drawing_file": "xl/drawings/drawing1.xml", "drawing_xml": drawing_xml, "drawing_rels": None, "files": {}}, "shapes": [{"index": 1, "id": "2", "name": "Shape 1", "type": "shape", "text": None, "relationship_id": None}]}]}

    M.excel_set_shape_style(key, "Sheet1", 1, clear_fill=True, clear_outline=True)

    xml = M._sessions[key]["sheets"][0]["drawing_data"]["drawing_xml"]
    assert '<a:noFill />' in xml or '<a:noFill/>' in xml
    assert M._sessions[key]["sheets"][0]["shapes"][0]["fill_color"] is None
    assert M._sessions[key]["sheets"][0]["shapes"][0]["outline_color"] is None


def test_underline_true_saves_as_single_underline(tmp_path):
    src = tmp_path / "underline.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "text"
    wb.save(src)
    key = M.excel_load(str(src)).split("session_key='")[1].split("'")[0]

    M.excel_set_style(key, "Sheet", 0, 0, 0, 0, {"underline": True})
    M.excel_save(key)

    assert openpyxl.load_workbook(src).active["A1"].font.underline == "single"


def test_save_refuses_macro_extension_mismatch(tmp_path):
    src = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    wb.save(src)
    key = M.excel_load(str(src)).split("session_key='")[1].split("'")[0]

    with pytest.raises(ValueError, match="Refusing to save non-macro"):
        M.excel_save_as_copy(key, str(tmp_path / "wrong.xlsm"))


# ── A1 ────────────────────────────────────────────────────────────────────────

def test_filtered_load_save_preserves_other_sheets(tmp_path):
    src = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    wb.active["A1"] = "keep1"
    wb.create_sheet("S2")["A1"] = "target"
    wb.create_sheet("S3")["A1"] = "keep3"
    wb.save(src)

    key = str(src.resolve())
    M.excel_load(str(src), sheet_name="S2")
    M.excel_edit_cells(key, "S2", [{"row_index": 0, "edits": {"0": "edited"}}])
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2.sheetnames == ["S1", "S2", "S3"]
    assert wb2["S1"]["A1"].value == "keep1"
    assert wb2["S2"]["A1"].value == "edited"
    assert wb2["S3"]["A1"].value == "keep3"


# ── A2 ────────────────────────────────────────────────────────────────────────

def _add_dummy_vba_project(path: Path) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                xml = raw.decode("utf-8")
                xml = re.sub(
                    r'<Override\b[^>]*\bPartName="/xl/workbook.xml"[^>]*/>',
                    '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"/>',
                    xml,
                    count=1,
                )
                if 'Extension="bin"' not in xml:
                    xml = xml.replace("</Types>", '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>', 1)
                raw = xml.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                xml = raw.decode("utf-8")
                if "vbaProject" not in xml:
                    xml = xml.replace(
                        "</Relationships>",
                        '<Relationship Id="rId999" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/></Relationships>',
                        1,
                    )
                raw = xml.encode("utf-8")
            zout.writestr(item, raw)
        zout.writestr("xl/vbaProject.bin", b"dummy-vba-project")
    tmp.replace(path)


def test_xlsm_load_save_copy_preserves_vba_part(tmp_path):
    xlsm = tmp_path / "macro.xlsm"
    out = tmp_path / "macro-out.xlsm"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "macro container"
    wb.save(xlsm)
    _add_dummy_vba_project(xlsm)

    report = json.loads(M.excel_validate_workbook(str(xlsm)))
    assert report["valid"] is True
    assert report["features"]["vba_project"] == 1

    key = M.excel_load(str(xlsm)).split("session_key='")[1].split("'")[0]
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "kept macro"}}])
    M.excel_save_as_copy(key, str(out))

    out_report = json.loads(M.excel_validate_workbook(str(out)))
    assert out_report["valid"] is True
    assert out_report["features"]["vba_project"] == 1
    with zipfile.ZipFile(out) as zf:
        assert zf.read("xl/vbaProject.bin") == b"dummy-vba-project"
        assert "macroEnabled" in zf.read("[Content_Types].xml").decode("utf-8")

def test_convert_to_markdown_without_session(tmp_path, monkeypatch):
    src = tmp_path / "simple.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score"])
    ws.append(["Ada", 10])
    wb.save(src)

    def fake_convert_excel_to_markdown(data, *, sheet_name=None):
        assert sheet_name is None
        assert data["sheets"][0]["name"] == "Data"
        assert data["sheets"][0]["rows"][1]["cells"][0]["v"] == "Ada"
        return "# Data\n\n| Name | Score |\n| --- | --- |\n| Ada | 10 |"

    monkeypatch.setitem(
        sys.modules,
        "excel_converter",
        types.SimpleNamespace(convert_excel_to_markdown=fake_convert_excel_to_markdown),
    )

    before_sessions = dict(M._sessions)
    result = M.convert_to_markdown(str(src))

    assert result.mimeType == "text/markdown"
    assert "| Ada | 10 |" in result.text
    assert M._sessions == before_sessions

def test_convert_to_markdown_accepts_sheet_range_and_limits(tmp_path, monkeypatch):
    src = tmp_path / "limited.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    wb.save(src)

    seen = {}

    def fake_convert_excel_to_markdown(data, *, sheet_name=None):
        seen["sheets"] = data["sheets"]
        return "ok"

    monkeypatch.setitem(
        sys.modules,
        "excel_converter",
        types.SimpleNamespace(convert_excel_to_markdown=fake_convert_excel_to_markdown),
    )

    result = M.convert_to_markdown(str(src), sheet_name="Data", range_ref="B1:C3", max_rows=2, max_cols=1)
    assert result.text == "ok"
    rows = seen["sheets"][0]["rows"]
    assert len(rows) == 2
    assert len(rows[0]["cells"]) == 1
    assert rows[0]["cells"][0]["v"] == "B"

def test_targeted_range_find_and_summary_tools(tmp_path):
    src = tmp_path / "targeted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score", "Formula"])
    ws.append(["Ada", 10, "=B2*2"])
    ws.append(["Grace", 20, "=B3*2"])
    ws.merge_cells("A4:B4")
    ws["A4"] = "Merged"
    wb.save(src)

    key = str(src.resolve())
    M.excel_load(str(src))
    read = json.loads(M.excel_read_range(key, "Data", range_ref="A2:B3"))
    assert read["values"] == [["Ada", 10], ["Grace", 20]]

    matches = json.loads(M.excel_find_cells(key, "Ada"))
    assert matches["count"] == 1
    assert matches["matches"][0]["row_index"] == 1
    formulas = json.loads(M.excel_find_cells(key, "B3", match_in="formula"))
    assert formulas["matches"][0]["value"] == "=B3*2"
    M.excel_close(key)

    summary = json.loads(M.excel_get_workbook_summary(str(src)))
    assert summary["sheet_count"] == 1
    assert summary["sheets"][0]["formula_count"] == 2
    assert summary["sheets"][0]["merged_ranges"] == 1

def test_excel_table_defined_name_preview_and_markdown_range(tmp_path):
    src = tmp_path / "metadata.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Score"])
    ws.append(["Ada", 10])
    ws.append(["Grace", 20])
    table = Table(displayName="Scores", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    wb.defined_names["ScoreRange"] = DefinedName("ScoreRange", attr_text="Data!$B$2:$B$3")
    wb.save(src)

    info = json.loads(M.excel_get_info(str(src)))
    assert info["sheets"][0]["table_count"] == 1

    key = str(src.resolve())
    M.excel_load(str(src))
    tables = json.loads(M.excel_list_tables(key))
    assert tables["tables"][0]["name"] == "Scores"
    names = json.loads(M.excel_list_defined_names(key))
    assert names["defined_names"][0]["name"] == "ScoreRange"
    md = M.excel_to_markdown_range(key, "Data", range_ref="A1:B2")
    assert "| Name | Score |" in md.text
    M.excel_close(key)

    preview = json.loads(M.excel_get_sheet_preview(str(src), max_rows=2, max_cols=1))
    assert preview["sheets"][0]["rows"] == [["Name"], ["Ada"]]


# ── A3 ────────────────────────────────────────────────────────────────────────

def test_failed_save_leaves_original_intact(tmp_path):
    src = tmp_path / "orig.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "precious"
    wb.save(src)
    original_bytes = src.read_bytes()

    key = M_load(src)
    data = M._get_session(key)
    data["sheets"][0]["name"] = "bad[name]"  # invalid sheet title → save fails
    with pytest.raises(Exception):
        M.excel_save(key)
    M.excel_close(key)

    assert src.read_bytes() == original_bytes, "failed save must not touch the file"


def M_load(path: Path) -> str:
    out = M.excel_load(str(path))
    return str(Path(path).resolve())


# ── B1: row insert ────────────────────────────────────────────────────────────

def test_insert_rows_shifts_anchored_metadata(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    rows = M.excel_clone_rows(key, "Data", 0)
    M.excel_insert_rows(key, "Data", [{"after_index": -1, "rows_json": json.loads(rows)},
                                      {"after_index": -1, "rows_json": json.loads(rows)}])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["B7"].hyperlink is not None            # B5 + 2 rows
    assert ws["C8"].comment is not None              # C6 + 2 rows
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["D4:D12"]
    cf = [str(r.sqref) for r in ws.conditional_formatting]
    assert cf == ["A4:A12"]
    assert "A10:B11" in {str(r) for r in ws.merged_cells.ranges}
    assert str(ws.auto_filter.ref) == "A3:F12"
    assert ws.freeze_panes == "A7"


# ── B1: row delete ────────────────────────────────────────────────────────────

def test_delete_rows_shifts_and_shrinks(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    # delete row 0 (above) and row 2 (inside the DV/CF ranges)
    M.excel_delete_rows(key, "Data", row_indices=[0, 2])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["B3"].hyperlink is not None            # B5 − 2 rows
    assert ws["C4"].comment is not None
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["D1:D8"]                          # D2:D10 minus two rows
    assert "A6:B7" in {str(r) for r in ws.merged_cells.ranges}


def test_delete_merge_origin_row_dissolves_cleanly(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    M.excel_delete_rows(key, "Data", row_indices=[7])   # merge origin row (A8:B9)
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A8:B8" in merged                         # shrunk to surviving row
    # no orphan 2-row merge left behind
    assert "A8:B9" not in merged


# ── B1: column ops ────────────────────────────────────────────────────────────

def test_insert_column_shifts_anchored_metadata(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    M.excel_insert_column(key, "Data", after_col_index=-1)   # prepend column A
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws["C5"].hyperlink is not None            # B5 → C5
    dvs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert dvs == ["E2:E10"]
    assert "B8:C9" in {str(r) for r in ws.merged_cells.ranges}
    assert ws.column_dimensions["F"].hidden          # hidden E → F


def test_delete_column_count_and_hidden_shift(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    msg = M.excel_delete_column(key, "Data", 0)
    assert "from 10 row(s)" in msg
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws = wb["Data"]
    assert ws.column_dimensions["D"].hidden          # hidden E → D
    assert ws["A5"].hyperlink is not None            # B5 → A5


# ── B2 ────────────────────────────────────────────────────────────────────────

def test_rename_sheet_updates_defined_names(tmp_path):
    src = tmp_path / "named.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Old Data"
    ws["A1"] = 1
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="'Old Data'!$A$1")
    wb.save(src)

    key = M_load(src)
    M.excel_rename_sheet(key, "Old Data", "NewData")
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2.defined_names["MyRange"].attr_text == "NewData!$A$1"


# ── B3 ────────────────────────────────────────────────────────────────────────

def test_merge_overlap_rejected(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    with pytest.raises(ValueError, match="overlaps"):
        M.excel_merge_cells(key, "Data", 7, 1, 9, 2)   # overlaps A8:B9
    M.excel_close(key)


def test_edit_slave_cell_rejected(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    with pytest.raises(ValueError, match="slave"):
        M.excel_edit_cells(key, "Data", [{"row_index": 8, "edits": {"0": "x"}}])
    M.excel_close(key)


# ── C ─────────────────────────────────────────────────────────────────────────

def test_internal_hyperlink_roundtrip(tmp_path):
    src = tmp_path / "links.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.create_sheet("Target")
    ws["A1"] = "go"
    from openpyxl.worksheet.hyperlink import Hyperlink
    ws["A1"].hyperlink = Hyperlink(ref="A1", location="Target!A1")
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))
    wb2 = openpyxl.load_workbook(out)
    hl = wb2.active["A1"].hyperlink
    assert hl is not None and hl.location == "Target!A1"


def test_docprops_and_workbook_view_roundtrip(tmp_path):
    src = tmp_path / "props.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.create_sheet("Second")["A1"] = 2
    wb.properties.creator = "Mori Daichi"
    wb.properties.title = "Spec"
    wb.properties.lastModifiedBy = "Reviewer"
    wb.active = 1  # openpyxl persists activeTab from wb.active, not views[0]
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))
    wb2 = openpyxl.load_workbook(out)
    assert wb2.properties.creator == "Mori Daichi"
    assert wb2.properties.title == "Spec"
    assert wb2.properties.lastModifiedBy == "Reviewer"
    assert wb2.views[0].activeTab == 1


def test_read_tools_strip_private_keys(tmp_path):
    src = tmp_path / "rich.xlsx"
    _make_rich_sheet(src)
    key = M_load(src)
    rows = json.loads(M.excel_get_rows(key, "Data", 0, 3))
    cell = json.loads(M.excel_get_cell(key, "Data", 0, 0))
    M.excel_close(key)

    def no_private(obj):
        if isinstance(obj, dict):
            return all(not k.startswith("_") and no_private(v) for k, v in obj.items())
        if isinstance(obj, list):
            return all(no_private(x) for x in obj)
        return True

    assert no_private(rows)
    assert no_private(cell)
