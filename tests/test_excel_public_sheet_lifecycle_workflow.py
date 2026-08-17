"""Public save/reload acceptance coverage for worksheet lifecycle and copy semantics."""

import json
import posixpath
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import SortCondition, SortState
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableFormula
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _load(path: Path) -> str:
    M.excel_load(str(path))
    return str(path.resolve())


def _feature_workbook(
    path: Path,
    sheet_name: str,
    code_name: str,
    fill_color: str,
    table_name: str = "Sales",
) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.sheet_properties.codeName = code_name
    rows = [
        ("Label", "Value"),
        ("Alpha", 2),
        ("Beta", 5),
        ("Gamma", 3),
    ]
    for row in rows:
        sheet.append(row)

    table = Table(displayName=table_name, ref="A1:B4")
    table._initialise_columns()
    table.tableColumns[0].name = "Label"
    table.tableColumns[1].name = "Value"
    table.tableColumns[1].calculatedColumnFormula = TableFormula(attr_text="B2*2")
    table.autoFilter.ref = "A1:B4"
    table.sortState = SortState(
        ref="A2:B4",
        sortCondition=[SortCondition(ref="B2:B4")],
    )
    sheet.add_table(table)

    sheet.conditional_formatting.add(
        "B2:B4",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=fill_color),
        ),
    )
    chart = LineChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=4))
    sheet.add_chart(chart, "D2")
    sheet.defined_names.add(
        DefinedName("LocalInput", attr_text=f"'{sheet_name}'!$B$2:$B$4")
    )
    workbook.active = 0
    workbook.views[0].firstSheet = 0
    workbook.save(path)


def _chart_formulas(path: Path) -> dict[str, list[str]]:
    result = {}
    with zipfile.ZipFile(path, "r") as archive:
        for name in archive.namelist():
            if not name.startswith("xl/charts/chart") or not name.endswith(".xml"):
                continue
            root = ET.fromstring(archive.read(name))
            result[name] = [
                node.text or ""
                for node in root.iter()
                if node.tag.rsplit("}", 1)[-1] == "f"
            ]
    return result


def _sheet_drawing_parts(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in workbook_rels.findall(f"{{{REL_NS}}}Relationship")
        }
        result = {}
        for sheet in workbook_root.findall(f"{{{MAIN_NS}}}sheets/{{{MAIN_NS}}}sheet"):
            relationship_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
            target = rel_targets[relationship_id]
            sheet_part = target.lstrip("/") if target.startswith("/") else posixpath.normpath(
                posixpath.join("xl", target)
            )
            directory, filename = sheet_part.rsplit("/", 1)
            rels_part = f"{directory}/_rels/{filename}.rels"
            if rels_part not in archive.namelist():
                continue
            rels_root = ET.fromstring(archive.read(rels_part))
            for relationship in rels_root.findall(f"{{{REL_NS}}}Relationship"):
                if relationship.attrib.get("Type", "").rstrip("/").endswith("/drawing"):
                    result[sheet.attrib["name"]] = posixpath.normpath(
                        posixpath.join(directory, relationship.attrib["Target"])
                    )
                    break
        return result


def _first_cf_fill(sheet) -> str:
    formatting = list(sheet.conditional_formatting)
    rule = formatting[0].rules[0]
    return rule.dxf.fill.fgColor.rgb


def test_add_move_delete_remaps_views_and_invalidates_deleted_sheet_references(tmp_path):
    path = tmp_path / "lifecycle-delete.xlsx"
    workbook = openpyxl.Workbook()
    sheet_a = workbook.active
    sheet_a.title = "A"
    sheet_b = workbook.create_sheet("B")
    workbook.create_sheet("C")
    sheet_b["A1"] = 1
    sheet_b["A2"] = 2
    sheet_b["A3"] = 3
    sheet_a["A1"] = "=B!A1"
    sheet_a["A2"] = "Go"
    sheet_a["A2"].hyperlink = Hyperlink(ref="A2", location="B!A1")

    validation = DataValidation(type="whole", formula1="=B!$A$1")
    validation.add("C1")
    sheet_a.add_data_validation(validation)
    sheet_a.conditional_formatting.add("A1", FormulaRule(formula=["B!$A$1>0"]))

    sheet_a["D1"] = "Label"
    sheet_a["E1"] = "Calc"
    sheet_a["D2"] = "One"
    sheet_a["E2"] = 1
    sheet_a["D3"] = "Two"
    sheet_a["E3"] = 2
    table = Table(displayName="Refs", ref="D1:E3")
    table._initialise_columns()
    table.tableColumns[0].name = "Label"
    table.tableColumns[1].name = "Calc"
    table.tableColumns[1].calculatedColumnFormula = TableFormula(attr_text="B!A1")
    sheet_a.add_table(table)

    chart = LineChart()
    chart.add_data(Reference(sheet_b, min_col=1, min_row=1, max_row=3))
    sheet_a.add_chart(chart, "G2")
    workbook.defined_names["ToB"] = DefinedName("ToB", attr_text="B!$A$1")
    sheet_b.defined_names.add(DefinedName("LocalB", attr_text="B!$A$1"))
    workbook.active = 1
    workbook.views[0].firstSheet = 1
    workbook.save(path)

    session_key = _load(path)
    try:
        M.excel_add_sheet(session_key, "Intro", position=0)
        M.excel_move_sheet(session_key, "B", position=0)
        M.excel_delete_sheet(session_key, "B")
        M.excel_save(session_key)
    finally:
        M.excel_close(session_key)

    saved = openpyxl.load_workbook(path)
    try:
        assert saved.sheetnames == ["Intro", "A", "C"]
        assert saved.index(saved.active) == 0
        assert saved.views[0].activeTab == 0
        assert saved.views[0].firstSheet == 0
        sheet_a = saved["A"]
        assert sheet_a["A1"].value == "=#REF!A1"
        assert sheet_a["A2"].hyperlink is None
        assert sheet_a.data_validations.dataValidation[0].formula1 == "=#REF!$A$1"
        formatting = list(sheet_a.conditional_formatting)
        assert formatting[0].rules[0].formula == ["#REF!$A$1>0"]
        calculated = sheet_a.tables["Refs"].tableColumns[1].calculatedColumnFormula
        assert calculated.attr_text == "#REF!A1"
        assert saved.defined_names["ToB"].attr_text == "#REF!$A$1"
    finally:
        saved.close()

    formulas = [value for values in _chart_formulas(path).values() for value in values]
    assert any("#REF!$A$1:$A$3" in value for value in formulas)


def test_copy_sheet_preserves_scoped_names_cf_and_uses_unique_object_identity(tmp_path):
    path = tmp_path / "same-workbook-copy.xlsx"
    _feature_workbook(path, "Template", "TemplateCode", "FFFF0000")
    session_key = _load(path)
    try:
        M.excel_copy_sheet(session_key, "Template", "Template Copy", position=1)
        M.excel_save(session_key)
    finally:
        M.excel_close(session_key)

    saved = openpyxl.load_workbook(path)
    try:
        original = saved["Template"]
        copied = saved["Template Copy"]
        assert original.sheet_properties.codeName == "TemplateCode"
        assert copied.sheet_properties.codeName == "TemplateCode_2"
        original_table = next(iter(original.tables.values()))
        copied_table = next(iter(copied.tables.values()))
        assert original_table.name == original_table.displayName == "Sales"
        assert copied_table.name == copied_table.displayName == "Sales_2"
        assert original_table.id != copied_table.id
        assert copied.defined_names["LocalInput"].attr_text == "'Template Copy'!$B$2:$B$4"
        assert _first_cf_fill(original) == "FFFF0000"
        assert _first_cf_fill(copied) == "FFFF0000"
    finally:
        saved.close()

    drawing_parts = _sheet_drawing_parts(path)
    assert drawing_parts["Template"] != drawing_parts["Template Copy"]
    formula_sets = _chart_formulas(path)
    flattened = [formula for values in formula_sets.values() for formula in values]
    assert any("'Template'!" in formula for formula in flattened)
    assert any("'Template Copy'!" in formula for formula in flattened)


def test_copy_sheet_to_merges_dxfs_and_rebases_names_tables_views_and_drawings(tmp_path):
    source_path = tmp_path / "source.xlsx"
    destination_path = tmp_path / "destination.xlsx"
    _feature_workbook(source_path, "Template", "TemplateCode", "FFFF0000")
    _feature_workbook(destination_path, "Existing", "TemplateCode", "FF00FF00")

    source_key = _load(source_path)
    destination_key = _load(destination_path)
    try:
        M.excel_copy_sheet_to(
            source_key,
            "Template",
            destination_key,
            new_name="Imported",
            position=0,
        )
        M.excel_save(destination_key)
    finally:
        M.excel_close(source_key)
        M.excel_close(destination_key)

    saved = openpyxl.load_workbook(destination_path)
    try:
        assert saved.sheetnames == ["Imported", "Existing"]
        assert saved.index(saved.active) == 1
        assert saved.views[0].activeTab == 1
        assert saved.views[0].firstSheet == 1
        imported = saved["Imported"]
        existing = saved["Existing"]
        assert imported.sheet_properties.codeName == "TemplateCode_2"
        assert existing.sheet_properties.codeName == "TemplateCode"
        imported_table = next(iter(imported.tables.values()))
        existing_table = next(iter(existing.tables.values()))
        assert imported_table.name == imported_table.displayName == "Sales_2"
        assert existing_table.name == existing_table.displayName == "Sales"
        assert imported_table.id != existing_table.id
        assert imported.defined_names["LocalInput"].attr_text == "Imported!$B$2:$B$4"
        assert _first_cf_fill(imported) == "FFFF0000"
        assert _first_cf_fill(existing) == "FF00FF00"
    finally:
        saved.close()

    drawing_parts = _sheet_drawing_parts(destination_path)
    assert drawing_parts["Imported"] != drawing_parts["Existing"]
    flattened = [formula for values in _chart_formulas(destination_path).values() for formula in values]
    assert any("Imported!" in formula for formula in flattened)
    assert any("'Existing'!" in formula for formula in flattened)


_CUSTOM_SHEET_REL_TYPE = "urn:docloupe:relationships/sheet-metadata"
_CUSTOM_SHEET_CONTENT_TYPE = "application/vnd.docloupe.sheet-metadata+xml"
_COPY_XF_FLAGS = {
    "applyFont": True,
    "applyFill": True,
    "applyBorder": True,
    "applyAlignment": True,
    "applyProtection": True,
    "xfId": 3,
    "pivotButton": True,
    "quotePrefix": True,
}


def _sheet_part_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        relationship.attrib["Id"]: relationship.attrib["Target"]
        for relationship in workbook_rels.findall(f"{{{REL_NS}}}Relationship")
    }
    result = {}
    for sheet in workbook_root.findall(f"{{{MAIN_NS}}}sheets/{{{MAIN_NS}}}sheet"):
        target = rel_targets[sheet.attrib[f"{{{DOC_REL_NS}}}id"]]
        result[sheet.attrib["name"]] = (
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.normpath(posixpath.join("xl", target))
        )
    return result


def _relationship_part(source_part: str) -> str:
    directory, filename = source_part.rsplit("/", 1)
    return f"{directory}/_rels/{filename}.rels"


def _patch_sheet_passthrough_relationship(
    path: Path,
    sheet_name: str,
    part_name: str,
    payload: str,
) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = {item.filename: archive.read(item.filename) for item in archive.infolist()}
        sheet_part = _sheet_part_map(archive)[sheet_name]

    rels_part = _relationship_part(sheet_part)
    if rels_part in entries:
        rels_root = ET.fromstring(entries[rels_part])
    else:
        rels_root = ET.Element(f"{{{REL_NS}}}Relationships")
    id_mapping = {}
    next_id = 2
    for existing in rels_root:
        old_id = existing.attrib["Id"]
        new_id = f"rId{next_id}"
        next_id += 1
        existing.set("Id", new_id)
        id_mapping[old_id] = new_id
    if id_mapping:
        sheet_xml = entries[sheet_part].decode("utf-8")
        sheet_xml = re.sub(
            r'(\br:id=")([^"]+)(")',
            lambda match: (
                f"{match.group(1)}{id_mapping.get(match.group(2), match.group(2))}{match.group(3)}"
            ),
            sheet_xml,
        )
        entries[sheet_part] = sheet_xml.encode("utf-8")

    relationship = ET.SubElement(rels_root, f"{{{REL_NS}}}Relationship")
    relationship.set("Id", "rId1")
    relationship.set("Type", _CUSTOM_SHEET_REL_TYPE)
    relationship.set("Target", posixpath.relpath(part_name, posixpath.dirname(sheet_part)))
    entries[rels_part] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)
    entries[part_name] = f"<metadata>{payload}</metadata>".encode("utf-8")

    content_types = ET.fromstring(entries["[Content_Types].xml"])
    content_type_namespace = content_types.tag.partition("}")[0].lstrip("{")
    override_tag = f"{{{content_type_namespace}}}Override"
    if not any(node.attrib.get("PartName") == f"/{part_name}" for node in content_types):
        override = ET.SubElement(content_types, override_tag)
        override.set("PartName", f"/{part_name}")
        override.set("ContentType", _CUSTOM_SHEET_CONTENT_TYPE)
    entries["[Content_Types].xml"] = ET.tostring(
        content_types, encoding="utf-8", xml_declaration=True
    )

    replacement = path.with_name(f"{path.stem}.relationship-patch{path.suffix}")
    with zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, raw in entries.items():
            archive.writestr(name, raw)
    replacement.replace(path)


def _custom_sheet_relationships(path: Path) -> dict[str, dict]:
    with zipfile.ZipFile(path, "r") as archive:
        sheet_parts = _sheet_part_map(archive)
        content_types = ET.fromstring(archive.read("[Content_Types].xml"))
        overrides = {
            node.attrib["PartName"].lstrip("/"): node.attrib["ContentType"]
            for node in content_types
            if node.tag.rsplit("}", 1)[-1] == "Override"
        }
        result = {}
        for sheet_name, sheet_part in sheet_parts.items():
            rels_part = _relationship_part(sheet_part)
            if rels_part not in archive.namelist():
                continue
            rels_root = ET.fromstring(archive.read(rels_part))
            for relationship in rels_root.findall(f"{{{REL_NS}}}Relationship"):
                if relationship.attrib.get("Type") != _CUSTOM_SHEET_REL_TYPE:
                    continue
                target = relationship.attrib["Target"]
                target_part = (
                    target.lstrip("/")
                    if target.startswith("/")
                    else posixpath.normpath(posixpath.join(posixpath.dirname(sheet_part), target))
                )
                result[sheet_name] = {
                    "id": relationship.attrib["Id"],
                    "target": target,
                    "part": target_part,
                    "payload": ET.fromstring(archive.read(target_part)).text,
                    "content_type": overrides.get(target_part),
                }
        return result


def _structural_sheet_workbook(path: Path, payload: str) -> None:
    workbook = openpyxl.Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible["A1"] = "Visible sheet"
    template = workbook.create_sheet("Template")
    template.sheet_state = "hidden"
    template.sheet_properties.codeName = "TemplateCode"
    template.sheet_properties.filterMode = True
    template.sheet_properties.published = True
    template.sheet_properties.transitionEntry = True
    template.sheet_properties.transitionEvaluation = True
    template.sheet_properties.outlinePr.summaryBelow = False
    template.sheet_properties.outlinePr.summaryRight = False
    template.sheet_properties.pageSetUpPr.fitToPage = True
    template.sheet_properties.pageSetUpPr.autoPageBreaks = False

    for row in range(1, 10):
        template.cell(row=row, column=1, value=f"Row {row}")
        template.cell(row=row, column=2, value=row)
        template.cell(row=row, column=3, value=row * 2)
    template["A1"].hyperlink = "https://example.com/metadata"
    template.freeze_panes = "B2"
    template.print_area = "A1:C9"
    template.print_title_rows = "$1:$2"
    template.print_title_cols = "$A:$A"
    template.page_setup.orientation = "landscape"
    template.page_setup.paperSize = template.PAPERSIZE_A4
    template.page_setup.fitToWidth = 1
    template.page_setup.fitToHeight = 2
    template.page_setup.firstPageNumber = 3
    template.page_setup.useFirstPageNumber = True
    template.page_setup.horizontalDpi = 300
    template.page_setup.verticalDpi = 300
    template.page_margins.left = 0.25
    template.page_margins.right = 0.35
    template.page_margins.top = 0.45
    template.page_margins.bottom = 0.55
    template.page_margins.header = 0.15
    template.page_margins.footer = 0.2
    template.print_options.horizontalCentered = True
    template.print_options.verticalCentered = True
    template.print_options.headings = True
    template.print_options.gridLines = True
    template.print_options.gridLinesSet = True
    template.oddHeader.center.text = "Confidential &A"
    template.oddFooter.right.text = "Page &P of &N"
    template.HeaderFooter.differentFirst = True
    template.firstHeader.left.text = "First page"
    template.HeaderFooter.differentOddEven = True
    template.evenFooter.center.text = "Even page"
    template.row_breaks.append(Break(id=4, min=0, max=16383, man=True))
    template.col_breaks.append(Break(id=2, min=0, max=1048575, man=True))
    workbook.active = 0
    workbook.save(path)
    workbook.close()
    _patch_sheet_passthrough_relationship(
        path,
        "Template",
        "xl/customParts/sheetMeta1.xml",
        payload,
    )


def _assert_sheet_printing_semantics(sheet) -> None:
    assert sheet.sheet_state == "hidden"
    assert sheet.sheet_properties.filterMode is True
    assert sheet.sheet_properties.published is True
    assert sheet.sheet_properties.transitionEntry is True
    assert sheet.sheet_properties.transitionEvaluation is True
    assert sheet.sheet_properties.outlinePr.summaryBelow is False
    assert sheet.sheet_properties.outlinePr.summaryRight is False
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert sheet.sheet_properties.pageSetUpPr.autoPageBreaks is False
    assert sheet.freeze_panes == "B2"
    assert sheet["A1"].hyperlink.target == "https://example.com/metadata"
    assert "$A$1:$C$9" in str(sheet.print_area)
    assert sheet.print_title_rows == "$1:$2"
    assert sheet.print_title_cols == "$A:$A"
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == str(sheet.PAPERSIZE_A4)
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 2
    assert sheet.page_setup.firstPageNumber == 3
    assert sheet.page_setup.useFirstPageNumber is True
    assert sheet.page_setup.horizontalDpi == 300
    assert sheet.page_setup.verticalDpi == 300
    assert sheet.page_margins.left == 0.25
    assert sheet.page_margins.right == 0.35
    assert sheet.page_margins.top == 0.45
    assert sheet.page_margins.bottom == 0.55
    assert sheet.page_margins.header == 0.15
    assert sheet.page_margins.footer == 0.2
    assert sheet.print_options.horizontalCentered is True
    assert sheet.print_options.verticalCentered is True
    assert sheet.print_options.headings is True
    assert sheet.print_options.gridLines is True
    assert sheet.print_options.gridLinesSet is True
    assert sheet.oddHeader.center.text == "Confidential &A"
    assert sheet.oddFooter.right.text == "Page &P of &N"
    assert sheet.HeaderFooter.differentFirst is True
    assert sheet.firstHeader.left.text == "First page"
    assert sheet.HeaderFooter.differentOddEven is True
    assert sheet.evenFooter.center.text == "Even page"
    assert [page_break.id for page_break in sheet.row_breaks.brk] == [4]
    assert [page_break.id for page_break in sheet.col_breaks.brk] == [2]


def _configure_copy_cell(session_key: str, sheet_name: str) -> None:
    M.excel_edit_rich_text(session_key, sheet_name, "A1", operations=[{
        "op": "replace_runs",
        "runs": [
            {
                "text": "Keep",
                "font": {
                    "bold": True,
                    "color": {"type": "rgb", "rgb": "FFFF0000"},
                },
            },
            {
                "text": "\nStyle",
                "font": {
                    "strike": True,
                    "color": {"type": "rgb", "rgb": "FF0000FF"},
                },
            },
        ],
    }])
    M.excel_set_style(session_key, sheet_name, 0, 0, style={
        "font": {
            "italic": True,
            "color": {"type": "theme", "theme": 4, "tint": -0.2},
        },
        "fill": {"foreground": {"type": "rgb", "rgb": "FFDDEBF7"}},
        "alignment": {
            "horizontal": "right",
            "readingOrder": 2,
            "relativeIndent": 1,
            "justifyLastLine": True,
        },
    })
    M.excel_set_borders(
        session_key,
        sheet_name,
        r1=0,
        c1=0,
        r2=0,
        c2=0,
        style="thin",
        color={"type": "theme", "theme": 5},
    )
    M.excel_set_cell_style_semantics(
        session_key,
        sheet_name,
        "A1",
        xf=dict(_COPY_XF_FLAGS),
    )


def _mutate_copy_source(session_key: str, sheet_name: str) -> None:
    M.excel_edit_rich_text(session_key, sheet_name, "A1", operations=[{
        "op": "style_run",
        "run_index": 0,
        "style": {
            "bold": False,
            "italic": True,
            "color": {"type": "rgb", "rgb": "FF00AA00"},
        },
    }])
    M.excel_set_style(session_key, sheet_name, 0, 0, style={
        "fill": {"foreground": {"type": "rgb", "rgb": "FFFFFF00"}},
        "alignment": {"horizontal": "left"},
    })
    M.excel_set_cell_style_semantics(
        session_key,
        sheet_name,
        "A1",
        xf={"xfId": 1, "pivotButton": False, "quotePrefix": False},
    )


def _assert_copied_cell_semantics(session_key: str, sheet_name: str, row: int, column: int) -> None:
    cell = json.loads(M.excel_get_cell(
        session_key,
        sheet_name,
        row,
        column,
        include_rich_text=True,
        include_semantics=True,
    ))
    runs = cell["rich_text"]["runs"]
    assert [run["text"] for run in runs] == ["Keep", "\nStyle"]
    assert runs[0]["font"]["bold"] is True
    assert runs[0]["font"]["color"]["rgb"] == "FFFF0000"
    assert runs[1]["font"]["strike"] is True
    assert runs[1]["font"]["color"]["rgb"] == "FF0000FF"

    semantics = cell["semantics"]
    assert semantics["font"]["italic"] is True
    assert semantics["font"]["color"]["type"] == "theme"
    assert semantics["font"]["color"]["theme"] == 4
    assert semantics["fill"]["foreground"]["rgb"] == "FFDDEBF7"
    assert semantics["alignment"]["horizontal"] == "right"
    assert semantics["alignment"]["readingOrder"] == 2
    assert semantics["alignment"]["relativeIndent"] == 1
    assert semantics["alignment"]["justifyLastLine"] is True
    assert semantics["border"]["top"]["style"] == "thin"
    assert semantics["border"]["top"]["color_object"]["type"] == "theme"
    assert semantics["border"]["top"]["color_object"]["theme"] == 5
    for field, expected in _COPY_XF_FLAGS.items():
        assert semantics["xf"].get(field) == expected, field


def _assert_named_style(session_key: str, sheet_name: str, row: int, column: int) -> None:
    cell = json.loads(M.excel_get_cell(
        session_key,
        sheet_name,
        row,
        column,
        include_semantics=True,
    ))
    assert cell["semantics"]["named_style"] == "CopyAccent"


def test_copy_row_and_column_deep_copy_rich_text_and_style_semantics(tmp_path):
    path = tmp_path / "copy-cell-semantics.xlsx"
    created = json.loads(M.excel_create_workbook(
        sheet_names=["Rows", "Columns"],
        target_path=str(path),
    ))
    session_key = created["session_key"]
    try:
        M.excel_add_named_style(session_key, "CopyAccent", {"numfmt": "0.00"})
        _configure_copy_cell(session_key, "Rows")
        _configure_copy_cell(session_key, "Columns")
        M.excel_edit_cells(session_key, "Rows", [{"row_index": 0, "edits": {2: "Named"}}])
        M.excel_set_style(
            session_key,
            "Rows",
            0,
            2,
            style={"named_style": "CopyAccent"},
        )

        M.excel_copy_row(session_key, "Rows", row_index=0, after_index=0)
        M.excel_copy_column(session_key, "Columns", col_index=0, after_col_index=0)
        _mutate_copy_source(session_key, "Rows")
        _mutate_copy_source(session_key, "Columns")
        M.excel_set_style(session_key, "Rows", 0, 2, style={"named_style": "Normal"})

        _assert_copied_cell_semantics(session_key, "Rows", 1, 0)
        _assert_copied_cell_semantics(session_key, "Columns", 0, 1)
        _assert_named_style(session_key, "Rows", 1, 2)
        M.excel_save(session_key)
    finally:
        M.excel_close(session_key)

    reloaded_key = _load(path)
    try:
        _assert_copied_cell_semantics(reloaded_key, "Rows", 1, 0)
        _assert_copied_cell_semantics(reloaded_key, "Columns", 0, 1)
        _assert_named_style(reloaded_key, "Rows", 1, 2)
    finally:
        M.excel_close(reloaded_key)


def test_copy_sheet_preserves_hidden_printing_and_rebases_passthrough_relationship(tmp_path):
    path = tmp_path / "copy-sheet-relationship.xlsx"
    _structural_sheet_workbook(path, payload="source-metadata")
    session_key = _load(path)
    try:
        M.excel_copy_sheet(session_key, "Template", "Template Copy", position=2)
        M.excel_save(session_key)
    finally:
        M.excel_close(session_key)

    saved = openpyxl.load_workbook(path)
    try:
        original = saved["Template"]
        copied = saved["Template Copy"]
        _assert_sheet_printing_semantics(original)
        _assert_sheet_printing_semantics(copied)
        assert original.sheet_properties.codeName == "TemplateCode"
        assert copied.sheet_properties.codeName == "TemplateCode_2"
    finally:
        saved.close()

    relationships = _custom_sheet_relationships(path)
    assert relationships["Template"]["payload"] == "source-metadata"
    assert relationships["Template Copy"]["payload"] == "source-metadata"
    assert relationships["Template"]["part"] != relationships["Template Copy"]["part"]
    assert relationships["Template"]["content_type"] == _CUSTOM_SHEET_CONTENT_TYPE
    assert relationships["Template Copy"]["content_type"] == _CUSTOM_SHEET_CONTENT_TYPE


def test_copy_sheet_to_preserves_and_rebases_passthrough_relationship_collision(tmp_path):
    source_path = tmp_path / "copy-sheet-source.xlsx"
    destination_path = tmp_path / "copy-sheet-destination.xlsx"
    _structural_sheet_workbook(source_path, payload="source-metadata")
    _structural_sheet_workbook(destination_path, payload="destination-metadata")

    source_key = _load(source_path)
    destination_key = _load(destination_path)
    try:
        M.excel_copy_sheet_to(
            source_key,
            "Template",
            destination_key,
            new_name="Imported",
            position=2,
        )
        M.excel_save(destination_key)
    finally:
        M.excel_close(source_key)
        M.excel_close(destination_key)

    saved = openpyxl.load_workbook(destination_path)
    try:
        existing = saved["Template"]
        imported = saved["Imported"]
        _assert_sheet_printing_semantics(existing)
        _assert_sheet_printing_semantics(imported)
        assert existing.sheet_properties.codeName == "TemplateCode"
        assert imported.sheet_properties.codeName == "TemplateCode_2"
    finally:
        saved.close()

    relationships = _custom_sheet_relationships(destination_path)
    assert relationships["Template"]["payload"] == "destination-metadata"
    assert relationships["Imported"]["payload"] == "source-metadata"
    assert relationships["Template"]["part"] != relationships["Imported"]["part"]
    assert relationships["Template"]["content_type"] == _CUSTOM_SHEET_CONTENT_TYPE
    assert relationships["Imported"]["content_type"] == _CUSTOM_SHEET_CONTENT_TYPE
