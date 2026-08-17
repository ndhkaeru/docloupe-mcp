"""
Verification tests for PRESERVATION_FIX_CHECKLIST.md bullets extending
excel_set_data_validation, excel_set_borders, and excel_edit_cells. These
exercise the PUBLIC MCP tool functions in servers/excel/main.py exactly as an
agent would call them -- never the internal helpers directly.

Bugs found and fixed here:
  - excel_set_data_validation: adding/patching a validation on a sheet that
    already had ANY data validation on load was silently discarded on save.
    reconstruct_excel unconditionally preferred a byte-for-byte raw-XML
    passthrough of the sheet's ORIGINAL <dataValidations> block (captured at
    load time, to preserve unknown/unsupported attributes) over whatever the
    tool had just built -- the caller's own edit never reached disk, with no
    error at all. Fixed by invalidating that stale passthrough whenever
    excel_set_data_validation actually mutates the validations list.
  - excel_edit_cells: editing an existing rich-text cell's PLAIN scalar value
    (not the typed-payload dict form) silently cleared the rich-text runs
    with no policy and no error -- the exact "silently guessing" behavior the
    typed-payload path already refuses to do. Fixed by requiring the typed
    payload form (with an explicit rich_text_policy) whenever the target
    cell already has rich text.

Genuine structural gap confirmed and documented as GAP (not fixed -- an
openpyxl upstream limitation, not something introduced by any of our public
tools):
  - x14 extended data validations (the genuine "unsupported extension" for
    data validations, living in <extLst>, e.g. ISO-8601 date rules or sqrefs
    openpyxl's own DataValidation model cannot express) are dropped by
    openpyxl's OWN reader at load time already ("Data Validation extension is
    not supported and will be removed"), before any of our code gets a
    chance to preserve them.
"""
import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _load_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _new_session(sheet_names=("S",)):
    created = json.loads(M.excel_create_workbook(sheet_names=list(sheet_names)))
    return created["session_key"]


# ---------------------------------------------------------------------------
# excel_set_data_validation
# ---------------------------------------------------------------------------

def test_data_validation_full_fields_multi_range_sqref_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "dv_full.xlsx"

    M.excel_set_data_validation(key, "S", sqref=["A1:A3", "C1:C3"], validation={
        "type": "list", "formula1": '"A,B,C"', "allowBlank": True,
        "showInputMessage": True, "promptTitle": "Pick one", "prompt": "Choose A, B, or C",
        "showErrorMessage": True, "errorTitle": "Invalid", "error": "Not allowed", "errorStyle": "stop",
        "imeMode": "hiragana",
    })
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    wb2 = openpyxl.load_workbook(str(out))
    dv = wb2["S"].data_validations.dataValidation[0]
    assert str(dv.sqref) == "A1:A3 C1:C3"
    assert dv.type == "list" and dv.formula1 == '"A,B,C"'
    assert dv.promptTitle == "Pick one" and dv.prompt == "Choose A, B, or C"
    assert dv.errorTitle == "Invalid" and dv.error == "Not allowed" and dv.errorStyle == "stop"
    assert dv.imeMode == "hiragana"
    wb2.close()


def test_data_validation_operator_and_numeric_types_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "dv_operator.xlsx"
    M.excel_set_data_validation(key, "S", sqref="B1:B3", validation={
        "type": "whole", "operator": "between", "formula1": "1", "formula2": "10",
    })
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    dv = wb2["S"].data_validations.dataValidation[0]
    assert dv.type == "whole" and dv.operator == "between"
    assert dv.formula1 == "1" and dv.formula2 == "10"
    wb2.close()


def test_data_validation_patch_new_rule_now_persists_alongside_existing_one(tmp_path):
    """Regression test for the silent-discard bug: adding a NEW validation on
    a sheet that already had one (from disk) must actually reach the file."""
    src = tmp_path / "dv_src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    existing = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    existing.add("D2:D10")
    ws.add_data_validation(existing)
    wb.save(src)

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_data_validation(key, "S", sqref="E2:E10", validation={
        "type": "whole", "operator": "greaterThan", "formula1": "0",
    })
    out = tmp_path / "dv_patched.xlsx"
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    wb2 = openpyxl.load_workbook(str(out))
    rules = {str(dv.sqref): dv for dv in wb2["S"].data_validations.dataValidation}
    assert "D2:D10" in rules and rules["D2:D10"].type == "list"
    assert "E2:E10" in rules and rules["E2:E10"].type == "whole"  # previously silently dropped
    wb2.close()


def test_data_validation_x14_extension_survives_unrelated_edit(tmp_path):
    src = tmp_path / "dv_ext_src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    existing = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    existing.add("D2:D10")
    ws.add_data_validation(existing)
    wb.save(src)

    ext_block = (
        '<extLst><ext xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
        'uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}">'
        '<x14:dataValidations xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main" count="1">'
        '<x14:dataValidation type="date" allowBlank="1">'
        "<x14:formula1><xm:f>DATE(2024,1,1)</xm:f></x14:formula1>"
        "<x14:sqref>F2:F10</x14:sqref></x14:dataValidation></x14:dataValidations></ext></extLst>"
    )
    with zipfile.ZipFile(src) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    patched = xml.replace("</worksheet>", ext_block + "</worksheet>")
    tmp_zip = src.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                raw = patched.encode("utf-8")
            zout.writestr(item, raw)
    tmp_zip.replace(src)

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_data_validation(key, "S", sqref="E2:E10", validation={
        "type": "whole", "operator": "greaterThan", "formula1": "0",
    })
    out = tmp_path / "dv_ext_out.xlsx"
    M.excel_save(key, str(out))
    with zipfile.ZipFile(out) as z:
        out_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "x14:dataValidations" in out_xml


# ---------------------------------------------------------------------------
# excel_set_borders
# ---------------------------------------------------------------------------

def test_borders_all_sides_and_color_object_types_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "borders_full.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "x"}}])

    M.excel_set_borders(key, "S", r1=0, c1=0, r2=0, c2=0, border={
        "top": {"style": "thin", "color": {"type": "theme", "theme": 4, "tint": 0.2}},
        "start": {"style": "thick", "color": {"type": "rgb", "rgb": "FF00FF00"}},
        "end": {"style": "dashed", "color": {"type": "indexed", "indexed": 10}},
        "vertical": {"style": "dotted"},
        "horizontal": {"style": "hair"},
        "diagonal": {"style": "thin", "color": {"type": "auto"}},
        "diagonalUp": True, "diagonalDown": True, "outline": False,
    })
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    wb2 = openpyxl.load_workbook(str(out))
    b = wb2["S"]["A1"].border
    assert b.top.style == "thin" and b.top.color.theme == 4 and b.top.color.tint == 0.2
    assert b.start.style == "thick" and b.start.color.rgb == "FF00FF00"
    assert b.end.style == "dashed" and b.end.color.indexed == 10
    assert b.vertical.style == "dotted"
    assert b.horizontal.style == "hair"
    assert b.diagonal.style == "thin" and b.diagonal.color.auto is True
    assert b.diagonalUp is True and b.diagonalDown is True and b.outline is False
    wb2.close()


def test_borders_patching_one_side_preserves_other_sides(tmp_path):
    key = _new_session()
    out = tmp_path / "borders_patch.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "x"}}])

    M.excel_set_borders(key, "S", r1=0, c1=0, r2=0, c2=0,
                         border={"top": {"style": "thin"}, "left": {"style": "thin"}})
    # Patch only bottom -- top and left must survive untouched.
    M.excel_set_borders(key, "S", r1=0, c1=0, r2=0, c2=0, sides=["bottom"], style="medium")

    session = M._get_session(key)
    sheet = next(s for s in session["sheets"] if s["name"] == "S")
    border = sheet["rows"][0]["cells"][0]["border"]
    assert border["top"]["style"] == "thin"
    assert border["left"]["style"] == "thin"
    assert border["bottom"]["style"] == "medium"

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    b = wb2["S"]["A1"].border
    assert b.top.style == "thin"
    assert b.left.style == "thin"
    assert b.bottom.style == "medium"
    wb2.close()


# ---------------------------------------------------------------------------
# excel_edit_cells
# ---------------------------------------------------------------------------

def test_edit_cells_typed_payload_value_data_type_formula_cached_value():
    key = _new_session()

    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {
        0: {"value": "42", "data_type": "s"},  # force text
        1: {"formula": "=1+1", "cached_value": 2, "cache_policy": "clear"},
        2: {"value": True},
    }}])
    row = json.loads(M.excel_get_rows(key, "S", 0, 1, include_formula_cache=True))[0]["cells"]
    assert row[0]["value"] == "42" and row[0]["data_type"] == "s"
    assert row[1]["formula"]["text"].lstrip("=") == "1+1"
    assert row[1]["formula"]["cached_value"] == 2
    assert row[2]["value"] is True


def test_edit_cells_typed_payload_clear_policies():
    key = _new_session()
    M.excel_set_style(key, "S", 0, 0, style={"bold": True})
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "value"}}])

    # clear="content": drops value/formula/rich text but keeps cell style.
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: {"clear": "content"}}}])
    cell = json.loads(M.excel_get_cell(key, "S", 0, 0, include_semantics=True))
    assert cell["value"] in (None, "")
    assert cell["semantics"]["font"]["bold"] is True  # style survives content clear

    # clear="all": drops everything including style.
    M.excel_set_style(key, "S", 0, 0, style={"bold": True})
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "value"}}])
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: {"clear": True}}}])
    cell2 = json.loads(M.excel_get_cell(key, "S", 0, 0, include_semantics=True))
    assert cell2["value"] in (None, "")
    assert not cell2["semantics"]["font"].get("bold")


def test_edit_cells_rich_text_value_requires_explicit_policy_or_errors():
    key = _new_session()
    M.excel_edit_rich_text(key, "S", "A1", operations=[
        {"op": "replace_runs", "runs": [{"text": "Hello", "font": {"bold": True}}]},
    ])

    # A bare scalar edit on a rich-text cell must error, not silently guess.
    with pytest.raises(ValueError):
        M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "Hello"}}])
    still_rich = json.loads(M.excel_get_cell(key, "S", 0, 0, include_rich_text=True))
    assert "rich_text" in still_rich  # the rejected edit did not partially apply

    # preserve_runs_if_text_equal: keeps runs when the new text matches exactly.
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {
        0: {"value": "Hello", "rich_text_policy": "preserve_runs_if_text_equal"},
    }}])
    preserved = json.loads(M.excel_get_cell(key, "S", 0, 0, include_rich_text=True))
    assert preserved["rich_text"]["runs"][0]["font"]["bold"] is True

    # replace_all: explicitly discards the runs.
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {
        0: {"value": "Bye", "rich_text_policy": "replace_all"},
    }}])
    replaced = json.loads(M.excel_get_cell(key, "S", 0, 0, include_rich_text=True))
    assert replaced["value"] == "Bye"
    # The bold run formatting is gone -- any "rich_text" shown here is just
    # _rich_text_model's synthesized single-run view of a now-plain cell.
    assert not any(run.get("font", {}).get("bold") for run in replaced.get("rich_text", {}).get("runs", []))


def test_edit_cells_does_not_create_stray_cells_outside_edited_range(tmp_path):
    key = _new_session()
    out = tmp_path / "sparse.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 49, "edits": {19: "far"}}])
    M.excel_save(key, str(out))

    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    import re
    cells = re.findall(r'<c\b[^>]*r="([A-Z]+\d+)"', xml)
    rows = re.findall(r'<row\b[^>]*r="(\d+)"', xml)
    assert cells == ["T50"]
    assert rows == ["50"]
