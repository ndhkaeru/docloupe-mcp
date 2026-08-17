"""
Verification tests for PRESERVATION_FIX_CHECKLIST.md bullets covering
workbook-level metadata: excel_set_calculation_properties,
excel_set_workbook_properties (incl. date1904 policy), excel_set_document_properties,
excel_set_workbook_protection, excel_get_workbook_views/excel_set_workbook_views.
These exercise the PUBLIC MCP tool functions in servers/excel/main.py exactly as
an agent would call them -- never internal helpers directly.

Bugs found and fixed here (all in servers/excel/core.py's reconstruct_excel /
serialize_excel): every one of these five tool clusters mutated ONLY the
in-session dict. serialize_excel never populated calculation_properties /
workbook_properties / workbook_protection from the source file at load time,
and reconstruct_excel never applied any of them to the openpyxl Workbook
object (or to raw XML) at save time -- a 100% silent discard on every save,
confirmed empirically before the fix (excel_add_image-style stub, but with
no persistence_requires_core_support flag at all to warn the caller).
Fixed by:
  - Reading/writing calcPr via openpyxl's native wb.calculation object.
  - Reading/writing workbookProtection via openpyxl's native wb.security
    object, honoring already_hashed for legacy passwords and passing modern
    hash/salt/spin-count fields straight through.
  - Reading/writing codeName + date1904 via wb.code_name / wb.epoch
    natively, and injecting the workbookPr attributes openpyxl's object
    model has no hook for at all (filterPrivacy, saveExternalLinkValues,
    showObjects, updateLinks) directly into the saved XML post-save.
  - Implementing the date1904 policy for real: preserve_displayed_dates
    falls out for free (cell values are Python datetimes; openpyxl defers
    serial<->datetime conversion to save time based on wb.epoch), while
    preserve_serial_values re-bases every date-valued cell so the raw
    serial number is unchanged and the displayed date shifts instead.
  - Extending doc_props extraction to the full core-property set
    (contentStatus, identifier, language, revision, version, lastPrinted,
    modified) and fixing up the 'modified' timestamp post-save, because
    openpyxl's own save_workbook() unconditionally stamps
    wb.properties.modified = datetime.now() as its very last step,
    silently overriding anything set beforehand -- the preserve/set_explicit
    policies were previously unimplementable no matter what the tool did.
  - Reading/writing typed custom document properties via wb.custom_doc_props
    (openpyxl supports this natively; it just was never wired up).
  - Reading/writing app.xml (extended) properties via direct XML injection,
    since openpyxl's writer always emits a brand new, empty
    ExtendedProperties document with no customization hook at all.
  - Replacing wb.views wholesale with the full ordered workbook_views list
    (previously only views[0] was ever touched, silently dropping any
    additional window/view entries).
"""
import json
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _load_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _new_session(sheet_names=("S",)):
    created = json.loads(M.excel_create_workbook(sheet_names=list(sheet_names)))
    return created["session_key"]


def _patch_core_modified(path, iso_text: str) -> None:
    """Test-fixture helper: openpyxl's own wb.save() unconditionally stamps
    docProps/core.xml's <dcterms:modified> with datetime.now(), so a fixture
    needing an arbitrary historical 'modified' value must patch it directly."""
    import os
    import re

    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        infos = {i.filename: i for i in zin.infolist()}
        parts = {n: zin.read(n) for n in names}
    core = parts["docProps/core.xml"].decode("utf-8")
    core = re.sub(
        r"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
        lambda m: m.group(1) + iso_text + "Z" + m.group(2),
        core,
    )
    parts["docProps/core.xml"] = core.encode("utf-8")
    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(infos[n], parts[n])
    os.replace(tmp, str(path))


# ---------------------------------------------------------------------------
# excel_set_calculation_properties
# ---------------------------------------------------------------------------

def test_calculation_properties_full_field_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "calc.xlsx"
    M.excel_set_calculation_properties(key, {
        "calcMode": "manual", "fullCalcOnLoad": False, "forceFullCalc": True,
        "calcOnSave": False, "iterate": True, "iterateCount": 50,
        "iterateDelta": 0.0005, "refMode": "A1", "fullPrecision": False,
        "concurrentCalc": False, "concurrentManualCount": 2,
    })
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    calc = wb2.calculation
    assert calc.calcMode == "manual"
    assert calc.fullCalcOnLoad is False
    assert calc.forceFullCalc is True
    assert calc.calcOnSave is False
    assert calc.iterate is True
    assert calc.iterateCount == 50
    assert calc.iterateDelta == 0.0005
    assert calc.refMode == "A1"
    assert calc.fullPrecision is False
    assert calc.concurrentCalc is False
    assert calc.concurrentManualCount == 2


def test_calculation_properties_explicit_false_zero_distinct_from_unset(tmp_path):
    """Explicit False/0 must round-trip as False/0, not vanish into 'unset'
    (which openpyxl would otherwise render as None -> attribute omitted)."""
    key = _new_session()
    out = tmp_path / "calc_falsy.xlsx"
    M.excel_set_calculation_properties(key, {
        "fullCalcOnLoad": False, "calcOnSave": False, "iterateCount": 0,
        "iterateDelta": 0.0,
    })
    sem = json.loads(M.excel_get_workbook_semantics(key))
    props = sem["calculation_properties"]
    assert props["fullCalcOnLoad"] is False
    assert props["calcOnSave"] is False
    assert "iterate" not in props or props.get("iterate") is None  # never explicitly set

    M.excel_save(key, str(out))
    z = zipfile.ZipFile(str(out))
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    assert 'fullCalcOnLoad="0"' in wb_xml
    assert 'calcOnSave="0"' in wb_xml


# ---------------------------------------------------------------------------
# excel_set_workbook_properties
# ---------------------------------------------------------------------------

def test_workbook_properties_code_name_and_workbookpr_extras_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "wbpr.xlsx"
    M.excel_set_workbook_properties(key, {
        "codeName": "ThisWorkbook",
        "filterPrivacy": True,
        "saveExternalLinkValues": False,
        "showObjects": "placeholders",
        "updateLinks": "never",
    })
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.code_name == "ThisWorkbook"
    wb_xml = zipfile.ZipFile(str(out)).read("xl/workbook.xml").decode("utf-8")
    assert 'filterPrivacy="1"' in wb_xml
    assert 'saveExternalLinkValues="0"' in wb_xml
    assert 'showObjects="placeholders"' in wb_xml
    assert 'updateLinks="never"' in wb_xml


def test_workbook_properties_date1904_change_requires_explicit_policy():
    key = _new_session()
    # No-op (False -> False, the baseline) must not require a policy.
    M.excel_set_workbook_properties(key, {"date1904": False})
    with pytest.raises(ValueError, match="date_system_policy"):
        M.excel_set_workbook_properties(key, {"date1904": True})


def test_workbook_properties_date1904_preserve_displayed_dates_shifts_serial_not_date(tmp_path):
    """preserve_displayed_dates: the calendar date/time shown to the user
    must NOT shift even though the underlying epoch (and thus raw serial
    number) changes."""
    src = tmp_path / "dates_src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    dt = datetime(2024, 6, 15, 12, 0, 0)
    ws["A1"] = dt
    ws["A1"].number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(src)

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_workbook_properties(key, {"date1904": True}, date_system_policy="preserve_displayed_dates")
    out = tmp_path / "dates_preserved.xlsx"
    M.excel_save(key, str(out))

    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.excel_base_date == datetime(1904, 1, 1)
    assert wb2.active["A1"].value == dt  # unchanged displayed date/time


def test_workbook_properties_date1904_preserve_serial_values_shifts_displayed_date(tmp_path):
    """preserve_serial_values: the raw serial number must stay put, so the
    displayed calendar date shifts by the 1462-day 1900/1904 epoch delta."""
    src = tmp_path / "dates_src2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    dt = datetime(2024, 6, 15, 12, 0, 0)
    ws["A1"] = dt
    ws["A1"].number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(src)

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_workbook_properties(key, {"date1904": True}, date_system_policy="preserve_serial_values")
    out = tmp_path / "dates_shifted.xlsx"
    M.excel_save(key, str(out))

    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.active["A1"].value == dt + timedelta(days=1462)


# ---------------------------------------------------------------------------
# excel_set_document_properties
# ---------------------------------------------------------------------------

def test_document_properties_extended_core_fields_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "docprops.xlsx"
    M.excel_set_document_properties(key, core={
        "contentStatus": "Draft", "identifier": "DOC-1", "language": "en-US",
        "revision": "3", "version": "1.2", "lastPrinted": "2021-05-05T05:05:05",
    })
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.properties.contentStatus == "Draft"
    assert wb2.properties.identifier == "DOC-1"
    assert wb2.properties.language == "en-US"
    assert wb2.properties.revision == "3"
    assert wb2.properties.version == "1.2"
    assert wb2.properties.lastPrinted == datetime(2021, 5, 5, 5, 5, 5)


def test_document_properties_typed_custom_and_app_properties_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "custom.xlsx"
    M.excel_set_document_properties(
        key,
        custom=[
            {"name": "Reviewer", "type": "StringProperty", "value": "Alice"},
            {"name": "RevisionNum", "type": "IntProperty", "value": 7},
            {"name": "Approved", "type": "BoolProperty", "value": True},
        ],
        app={"Company": "Acme Corp", "Manager": "Bob"},
    )
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    values = {p.name: p.value for p in wb2.custom_doc_props}
    assert values["Reviewer"] == "Alice"
    assert values["RevisionNum"] == 7
    assert values["Approved"] is True
    app_xml = zipfile.ZipFile(str(out)).read("docProps/app.xml").decode("utf-8")
    assert "<Company>Acme Corp</Company>" in app_xml
    assert "<Manager>Bob</Manager>" in app_xml


def test_document_properties_modified_preserve_policy_survives_unrelated_edit(tmp_path):
    src = tmp_path / "modsrc.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.save(src)
    _patch_core_modified(src, "2020-01-01T00:00:00")

    key = _load_key(M.excel_load(str(src)))
    sem = json.loads(M.excel_get_workbook_semantics(key))
    assert sem["document_properties"]["modified"] == "2020-01-01T00:00:00"

    # An unrelated edit + save (default policy = preserve) must not bump 'modified'.
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {"1": "x"}}])
    out = tmp_path / "mod_preserved.xlsx"
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.properties.modified == datetime(2020, 1, 1, 0, 0, 0)


def test_document_properties_modified_set_explicit_policy(tmp_path):
    src = tmp_path / "modsrc2.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.save(src)
    _patch_core_modified(src, "2020-01-01T00:00:00")

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_document_properties(key, core={"modified": "2022-07-07T07:07:07"}, modified_policy="set_explicit")
    out = tmp_path / "mod_explicit.xlsx"
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.properties.modified == datetime(2022, 7, 7, 7, 7, 7)


def test_document_properties_modified_update_on_save_policy(tmp_path):
    src = tmp_path / "modsrc3.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.save(src)
    _patch_core_modified(src, "2020-01-01T00:00:00")

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_document_properties(key, modified_policy="update_on_save")
    before = datetime.now(timezone.utc).replace(tzinfo=None)
    out = tmp_path / "mod_update.xlsx"
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert abs((wb2.properties.modified - before).total_seconds()) < 60


def test_document_properties_invalid_modified_policy_rejected():
    key = _new_session()
    with pytest.raises(ValueError, match="modified_policy"):
        M.excel_set_document_properties(key, modified_policy="bogus")


# ---------------------------------------------------------------------------
# excel_set_workbook_protection
# ---------------------------------------------------------------------------

def test_workbook_protection_flags_and_hashed_password_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "protect.xlsx"
    M.excel_set_workbook_protection(key, {
        "lockStructure": True, "lockWindows": True, "lockRevision": False,
        "workbookPassword": "hunter2",
    }, already_hashed=False)
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    sec = wb2.security
    assert sec.lockStructure is True
    assert sec.lockWindows is True
    assert sec.lockRevision is False
    assert sec.workbookPassword  # non-empty -- was hashed on the way in


def test_workbook_protection_already_hashed_value_is_not_rehashed(tmp_path):
    key = _new_session()
    out = tmp_path / "protect_hashed.xlsx"
    # A caller-supplied value marked already_hashed=True must reach the file
    # byte-for-byte, not get run through the hash function a second time.
    M.excel_set_workbook_protection(key, {"workbookPassword": "ABCD"}, already_hashed=True)
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.security.workbookPassword == "ABCD"


def test_workbook_protection_modern_hash_fields_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "protect_modern.xlsx"
    M.excel_set_workbook_protection(key, {
        "workbookAlgorithmName": "SHA-512",
        "workbookHashValue": "YWJjZGVm",
        "workbookSaltValue": "MTIzNDU2",
        "workbookSpinCount": 100000,
    }, already_hashed=True)
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    sec = wb2.security
    assert sec.workbookAlgorithmName == "SHA-512"
    assert sec.workbookHashValue == "YWJjZGVm"
    assert sec.workbookSaltValue == "MTIzNDU2"
    assert sec.workbookSpinCount == 100000


# ---------------------------------------------------------------------------
# excel_get_workbook_views / excel_set_workbook_views
# ---------------------------------------------------------------------------

def test_workbook_views_full_list_insert_update_and_geometry_roundtrip(tmp_path):
    key = _new_session(["A", "B"])
    out = tmp_path / "views.xlsx"
    M.excel_set_workbook_views(key, [
        {"activeTab": 1, "xWindow": 100, "yWindow": 50, "windowWidth": 800,
         "windowHeight": 600, "tabRatio": 500},
        {"activeTab": 0, "xWindow": 900, "yWindow": 50, "windowWidth": 400,
         "windowHeight": 300},
    ], mode="replace")

    views = json.loads(M.excel_get_workbook_views(key))["views"]
    assert len(views) == 2
    assert views[0]["xWindow"] == 100
    assert views[1]["xWindow"] == 900

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert len(wb2.views) == 2, "second workbook view was dropped on save"
    assert wb2.views[0].xWindow == 100
    assert wb2.views[0].windowWidth == 800
    assert wb2.views[1].xWindow == 900
    assert wb2.views[1].windowWidth == 400


def test_workbook_views_patch_mode_updates_one_view_and_preserves_others(tmp_path):
    key = _new_session(["A", "B"])
    out = tmp_path / "views_patch.xlsx"
    M.excel_set_workbook_views(key, [
        {"xWindow": 10, "windowWidth": 111},
        {"xWindow": 20, "windowWidth": 222},
    ], mode="replace")
    # Patch only view index 1's xWindow; windowWidth must survive untouched.
    M.excel_set_workbook_views(key, [{"index": 1, "xWindow": 999}], mode="patch")

    views = json.loads(M.excel_get_workbook_views(key))["views"]
    assert views[0]["xWindow"] == 10
    assert views[1]["xWindow"] == 999
    assert views[1]["windowWidth"] == 222

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.views[0].xWindow == 10
    assert wb2.views[1].xWindow == 999
    assert wb2.views[1].windowWidth == 222
