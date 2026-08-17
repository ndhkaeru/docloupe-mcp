import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.styles.named_styles import NamedStyleList


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

from core import _safe_named_style_name, reconstruct_excel, serialize_excel  # noqa: E402


def _make_named_style_workbook(path: Path, cell_count: int) -> list[str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    wb.add_named_style(NamedStyle(name="AuditNamed"))

    expected = []
    for row_index in range(1, cell_count + 1):
        style_name = "AuditNamed" if row_index % 2 else "Normal"
        ws.cell(row=row_index, column=1, value=row_index).style = style_name
        expected.append(style_name)

    wb.save(path)
    return expected


@pytest.mark.parametrize("cell_count", [2, 128])
def test_serialize_excel_snapshots_named_style_names_once(
    tmp_path, monkeypatch, cell_count
):
    source = tmp_path / f"named-styles-{cell_count}.xlsx"
    expected = _make_named_style_workbook(source, cell_count)

    original_names_getter = NamedStyleList.names.fget
    original_load_workbook = openpyxl.load_workbook
    workbook_loaded = False
    names_accesses = 0

    def counted_names(named_styles):
        nonlocal names_accesses
        if workbook_loaded:
            names_accesses += 1
        return original_names_getter(named_styles)

    def tracked_load_workbook(*args, **kwargs):
        nonlocal workbook_loaded
        workbook = original_load_workbook(*args, **kwargs)
        workbook_loaded = True
        return workbook

    monkeypatch.setattr(NamedStyleList, "names", property(counted_names))
    monkeypatch.setattr(openpyxl, "load_workbook", tracked_load_workbook)

    data = serialize_excel(str(source))

    actual = [row["cells"][0]["named_style"] for row in data["sheets"][0]["rows"]]
    assert actual == expected
    assert names_accesses == 1


@pytest.mark.parametrize(
    ("style", "names", "expected"),
    [
        (None, ("Normal", "AuditNamed"), "Normal"),
        (SimpleNamespace(xfId=1), ("Normal", "AuditNamed"), "AuditNamed"),
        (SimpleNamespace(xfId=99), ("Normal", "AuditNamed"), None),
        (SimpleNamespace(xfId=-1), ("Normal", "AuditNamed"), None),
        (SimpleNamespace(), ("Normal", "AuditNamed"), None),
        (SimpleNamespace(xfId="1"), ("Normal", "AuditNamed"), None),
        (SimpleNamespace(xfId=True), ("Normal", "AuditNamed"), None),
        (SimpleNamespace(xfId=0), (), None),
    ],
)
def test_safe_named_style_name_handles_style_edge_cases(style, names, expected):
    cell = SimpleNamespace(_style=style)

    assert _safe_named_style_name(cell, names) == expected


def test_safe_named_style_name_handles_missing_style_attribute():
    assert _safe_named_style_name(SimpleNamespace(), ("Normal",)) is None


def test_full_reconstruct_preserves_loaded_named_style_definition(tmp_path):
    source = tmp_path / "named-style-source.xlsx"
    target = tmp_path / "named-style-roundtrip.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    named_style = NamedStyle(name="ForegroundOnly")
    named_style.font = Font(bold=True)
    named_style.fill = PatternFill(patternType="solid", fgColor="FF99CC00")
    workbook.add_named_style(named_style)
    worksheet["A1"] = "Styled"
    worksheet["A1"].style = "ForegroundOnly"
    workbook.save(source)
    workbook.close()

    data = serialize_excel(str(source))
    foreground_style = next(
        item for item in data["named_styles"] if item["name"] == "ForegroundOnly"
    )
    assert foreground_style["style"]["fill"]["foreground"]["rgb"] == "FF99CC00"

    data["_dirty"]["workbook"] = ["drawings"]
    reconstruct_excel(data, str(target))

    roundtrip = openpyxl.load_workbook(target)
    try:
        cell = roundtrip.active["A1"]
        assert "ForegroundOnly" in roundtrip.named_styles
        assert cell.style == "ForegroundOnly"
        assert cell.fill.fgColor.rgb == "FF99CC00"
        assert cell._style.xfId == 1
        persisted_style = roundtrip._named_styles[1]
        assert persisted_style.font.bold is True
        assert persisted_style.font.name is None
    finally:
        roundtrip.close()
