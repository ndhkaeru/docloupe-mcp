import json
import re
import shutil
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
import preservation as preservation_module  # noqa: E402
from preservation import (  # noqa: E402
    cleanup_excel_backups,
    create_excel_backup,
    find_latest_excel_backup,
    inspect_workbook_pair,
    package_signature_report,
    verify_xlsx_preservation,
)


def _rewrite_part(path: Path, part_name: str, transform) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as destination:
            for item in source.infolist():
                raw = source.read(item.filename)
                if item.filename == part_name:
                    raw = transform(raw.decode("utf-8")).encode("utf-8")
                destination.writestr(item, raw)
    temporary.replace(path)


def _replace_part_bytes(path: Path, part_name: str, replacement: bytes) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    found = False
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as destination:
            for item in source.infolist():
                raw = source.read(item.filename)
                if item.filename == part_name:
                    raw = replacement
                    found = True
                destination.writestr(item, raw)
    assert found, part_name
    temporary.replace(path)


def _rewrite_part_with_uniform_compression(
    path: Path,
    part_name: str,
    transform,
    compression: int,
) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    found = False
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(temporary, "w", compression=compression) as destination:
            for item in source.infolist():
                raw = source.read(item.filename)
                if item.filename == part_name:
                    raw = transform(raw.decode("utf-8")).encode("utf-8")
                    found = True
                item.compress_type = compression
                destination.writestr(item, raw)
    assert found, part_name
    temporary.replace(path)


def _session_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _replace_named_style_collections(
    path: Path,
    style_count: int,
    changed_style_index: int | None = None,
) -> None:
    def replace_styles(xml: str) -> str:
        style_xfs = []
        cell_styles = []
        for index in range(style_count):
            changed_attribute = ' applyAlignment="1"' if index == changed_style_index else ""
            style_xfs.append(
                '<xf numFmtId="0" fontId="0" fillId="0" borderId="0"'
                f"{changed_attribute}/>"
            )
            cell_styles.append(f'<cellStyle name="Synthetic {index}" xfId="{index}"/>')

        xml, style_xf_replacements = re.subn(
            r"<cellStyleXfs\b[^>]*>.*?</cellStyleXfs>",
            f'<cellStyleXfs count="{style_count}">{"".join(style_xfs)}</cellStyleXfs>',
            xml,
            count=1,
            flags=re.DOTALL,
        )
        xml, cell_style_replacements = re.subn(
            r"<cellStyles\b[^>]*>.*?</cellStyles>",
            f'<cellStyles count="{style_count}">{"".join(cell_styles)}</cellStyles>',
            xml,
            count=1,
            flags=re.DOTALL,
        )
        assert style_xf_replacements == 1
        assert cell_style_replacements == 1
        return xml

    _rewrite_part(path, "xl/styles.xml", replace_styles)


def test_identical_copy_uses_fast_path_without_semantic_or_package_snapshots(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    copied = tmp_path / "copied.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "same"
    workbook.save(source)
    shutil.copy2(source, copied)

    def fail_if_called(*_args, **_kwargs):
        pytest.fail("byte-identical verification must not build semantic or style snapshots")

    monkeypatch.setattr(
        preservation_module,
        "_package_snapshot_from_inspection",
        fail_if_called,
    )
    monkeypatch.setattr(preservation_module, "_style_snapshot", fail_if_called)
    left, right = inspect_workbook_pair(source, copied)

    report = verify_xlsx_preservation(
        str(source),
        str(copied),
        before_inspection=left,
        after_inspection=right,
    )

    assert report["equivalent"] is True
    assert report["preservation_ok"] is True
    assert report["change_count"] == 0
    assert report["part_diff"]["changed"] is False
    assert report["before_sha256"] == report["after_sha256"]
    assert left.package_open_count + right.package_open_count == 1
    assert left.snapshot is None
    assert right.snapshot is None
    assert left.raw_parts_released is True
    assert right.raw_parts_released is True


def test_byte_identical_report_has_exact_contract_schema_and_counters(tmp_path):
    source = tmp_path / "contract-source.xlsx"
    copied = tmp_path / "contract-copy.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "contract"
    workbook.save(source)
    shutil.copy2(source, copied)

    with zipfile.ZipFile(source, "r") as archive:
        expected_part_count = sum(not item.is_dir() for item in archive.infolist())

    report = verify_xlsx_preservation(str(source), str(copied))

    assert set(report) == {
        "schema_version",
        "before_path",
        "after_path",
        "before_sha256",
        "after_sha256",
        "before_size",
        "after_size",
        "equivalent",
        "preservation_ok",
        "recommendation",
        "severity_counts",
        "classification_counts",
        "requested_paths",
        "approved_normalizations",
        "approved_normalization_rules",
        "normalization_evidence_complete",
        "fixture_gap_paths",
        "verifier_gap_paths",
        "unapproved_difference_count",
        "blocking_issue_count",
        "part_diff",
        "change_count",
        "changes",
        "truncated",
    }
    assert report["severity_counts"] == {
        "critical": 0,
        "high": 0,
        "medium": 0,
        "info": 0,
    }
    assert report["classification_counts"] == {
        "REQUESTED": 0,
        "APPROVED_NORMALIZATION": 0,
        "UNAPPROVED_LOSS": 0,
        "FIXTURE_GAP": 0,
        "VERIFIER_GAP": 0,
        "PACKAGE_INVALID": 0,
    }
    assert report["part_diff"] == {
        "before_part_count": expected_part_count,
        "after_part_count": expected_part_count,
        "added": [],
        "removed": [],
        "modified": [],
        "modified_xml": [],
        "modified_binary": [],
        "binary_hash_changes": [],
        "relationship_changes": {"added": {}, "removed": {}, "modified": {}},
        "content_type_changes": {"added": {}, "removed": {}, "modified": {}},
        "changed": False,
    }
    assert report["unapproved_difference_count"] == 0
    assert report["blocking_issue_count"] == 0
    assert report["change_count"] == 0
    assert report["changes"] == []
    assert report["truncated"] is False


def test_byte_identical_part_count_excludes_zip_directory_entries(tmp_path):
    source = tmp_path / "directory-source.xlsx"
    copied = tmp_path / "directory-copy.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "directory"
    workbook.save(source)
    with zipfile.ZipFile(source, "a") as archive:
        archive.writestr("customXml/empty-directory/", b"")
    shutil.copy2(source, copied)

    with zipfile.ZipFile(source, "r") as archive:
        entries = archive.infolist()
    assert any(item.is_dir() for item in entries)
    expected_part_count = sum(not item.is_dir() for item in entries)

    report = verify_xlsx_preservation(str(source), str(copied))

    assert report["part_diff"]["before_part_count"] == expected_part_count
    assert report["part_diff"]["after_part_count"] == expected_part_count


def test_byte_identical_workbook_rejects_malformed_xml_with_part_name(tmp_path):
    source = tmp_path / "malformed-source.xlsx"
    copied = tmp_path / "malformed-copy.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "malformed"
    workbook.save(source)
    _replace_part_bytes(source, "docProps/app.xml", b"<Properties>")
    shutil.copy2(source, copied)

    with pytest.raises(ValueError, match=r"^Invalid XML package part: docProps/app\.xml$"):
        verify_xlsx_preservation(str(source), str(copied))


def test_same_size_different_bytes_builds_semantic_snapshots(tmp_path, monkeypatch):
    source = tmp_path / "same-size-source.xlsx"
    changed = tmp_path / "same-size-changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "left"
    workbook.save(source)
    shutil.copy2(source, changed)

    _rewrite_part_with_uniform_compression(
        source,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml,
        zipfile.ZIP_STORED,
    )

    def replace_same_length_value(xml: str) -> str:
        replaced = xml.replace(">left<", ">east<", 1)
        assert replaced != xml
        return replaced

    _rewrite_part_with_uniform_compression(
        changed,
        "xl/worksheets/sheet1.xml",
        replace_same_length_value,
        zipfile.ZIP_STORED,
    )
    assert source.stat().st_size == changed.stat().st_size
    assert source.read_bytes() != changed.read_bytes()

    original_package_snapshot = preservation_module._package_snapshot_from_inspection
    snapshot_paths = []

    def package_snapshot_spy(inspection):
        snapshot_paths.append(inspection.path)
        return original_package_snapshot(inspection)

    monkeypatch.setattr(
        preservation_module,
        "_package_snapshot_from_inspection",
        package_snapshot_spy,
    )
    left, right = inspect_workbook_pair(source, changed)

    report = verify_xlsx_preservation(
        str(source),
        str(changed),
        before_inspection=left,
        after_inspection=right,
    )

    assert snapshot_paths == [str(source.resolve()), str(changed.resolve())]
    assert report["equivalent"] is False
    assert any(change["category"] == "cell_value" for change in report["changes"])
    assert left.snapshot is not None
    assert right.snapshot is not None
    assert left.raw_parts_released is True
    assert right.raw_parts_released is True


def test_signature_only_report_does_not_build_full_inspection(tmp_path, monkeypatch):
    source = tmp_path / "signature-source.xlsx"
    copied = tmp_path / "signature-copy.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "signature"
    workbook.save(source)
    shutil.copy2(source, copied)

    def fail_if_called(*_args, **_kwargs):
        pytest.fail("signature-only reporting must not build a full workbook inspection")

    monkeypatch.setattr(preservation_module, "inspect_workbook", fail_if_called)
    report = package_signature_report(source, copied, intentional_edit=False)

    assert report["status"] == "not_present"
    assert report["performance"]["before"]["mode"] == "signature_only"
    assert report["performance"]["after"]["mode"] == "signature_only"
    assert report["performance"]["before"]["package_open_count"] == 1
    assert report["performance"]["after"]["package_open_count"] == 1
    assert report["performance"]["before"]["part_read_count"] == 0
    assert report["performance"]["after"]["part_read_count"] == 0


def test_full_inspection_reuses_signatures_relationships_and_content_types(tmp_path, monkeypatch):
    source = tmp_path / "reuse-source.xlsx"
    changed = tmp_path / "reuse-changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "before"
    workbook.save(source)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(">before<", ">after<", 1),
    )

    left, right = inspect_workbook_pair(source, changed)
    assert left.relationship_records
    assert right.relationship_records
    assert left.content_type_records
    assert right.content_type_records

    def fail_if_recomputed(*_args, **_kwargs):
        pytest.fail("cached package inventories must be reused during verification")

    monkeypatch.setattr(preservation_module, "_relationship_records", fail_if_recomputed)
    monkeypatch.setattr(preservation_module, "_content_type_records", fail_if_recomputed)
    signature = package_signature_report(
        source,
        changed,
        intentional_edit=False,
        before_inspection=left,
        after_inspection=right,
    )
    report = verify_xlsx_preservation(
        str(source),
        str(changed),
        before_inspection=left,
        after_inspection=right,
    )

    assert signature["performance"]["before"]["mode"] == "reused_full_inspection"
    assert signature["performance"]["after"]["mode"] == "reused_full_inspection"
    assert signature["performance"]["before"]["package_open_count"] == 0
    assert signature["performance"]["after"]["package_open_count"] == 0
    assert left.package_open_count + right.package_open_count == 2
    assert report["equivalent"] is False
    assert left.snapshot is not None
    assert right.snapshot is not None
    assert left.raw_parts_released is True
    assert right.raw_parts_released is True


def test_large_named_style_graph_has_bounded_conversion_and_preserves_semantics(tmp_path, monkeypatch):
    source = tmp_path / "large-styles-source.xlsx"
    changed = tmp_path / "large-styles-changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "same"
    workbook.save(source)
    shutil.copy2(source, changed)

    style_count = 2_000
    _replace_named_style_collections(source, style_count)
    _replace_named_style_collections(changed, style_count, changed_style_index=style_count - 1)

    original_xml_node = getattr(preservation_module, "_xml_node", None)
    xf_conversion_count = 0
    max_xf_conversions = style_count * 8

    if original_xml_node is not None:
        def counting_xml_node(element):
            nonlocal xf_conversion_count
            if element is not None and element.tag.rsplit("}", 1)[-1] == "xf":
                xf_conversion_count += 1
                if xf_conversion_count > max_xf_conversions:
                    pytest.fail(
                        "style XF conversion count exceeded the bounded-cache regression limit: "
                        f"{xf_conversion_count} > {max_xf_conversions}"
                    )
            return original_xml_node(element)

        monkeypatch.setattr(preservation_module, "_xml_node", counting_xml_node)

    report = verify_xlsx_preservation(str(source), str(changed))

    assert report["equivalent"] is False
    assert report["preservation_ok"] is False
    assert any(
        change["path"] == "styles/named_styles" and change["category"] == "named_styles"
        for change in report["changes"]
    )
    assert not any(change["path"].startswith("worksheets/") for change in report["changes"])
    if original_xml_node is not None:
        assert xf_conversion_count <= max_xf_conversions


def test_rich_text_flattening_is_high_severity(tmp_path):
    source = tmp_path / "rich.xlsx"
    flattened = tmp_path / "flattened.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "HelloWorld"
    workbook.save(source)

    def add_runs(xml: str) -> str:
        replacement = (
            '<is><r><rPr><b/></rPr><t>Hello</t></r>'
            '<r><rPr><i/></rPr><t>World</t></r></is>'
        )
        return re.sub(r"<is>.*?</is>", replacement, xml, count=1, flags=re.DOTALL)

    _rewrite_part(source, "xl/worksheets/sheet1.xml", add_runs)
    shutil.copy2(source, flattened)
    _rewrite_part(
        flattened,
        "xl/worksheets/sheet1.xml",
        lambda xml: re.sub(r"<is>.*?</is>", "<is><t>HelloWorld</t></is>", xml, count=1, flags=re.DOTALL),
    )

    report = verify_xlsx_preservation(str(source), str(flattened))

    rich_changes = [change for change in report["changes"] if change["category"] == "rich_text"]
    assert rich_changes
    assert rich_changes[0]["severity"] == "high"
    assert rich_changes[0]["before"]["storage_form"] == "inlineStr"
    assert rich_changes[0]["before"]["shared_string_index"] is None
    assert rich_changes[0]["before"]["run_count"] == 2
    assert rich_changes[0]["before"]["runs"][0]["text"] == "Hello"
    assert rich_changes[0]["before"]["runs"][0]["start"] == 0
    assert rich_changes[0]["before"]["runs"][0]["end"] == 5
    assert rich_changes[0]["before"]["runs"][0]["properties"] is not None
    assert report["preservation_ok"] is False


def test_cell_presence_report_distinguishes_explicit_empty_from_missing(tmp_path):
    source = tmp_path / "presence.xlsx"
    changed = tmp_path / "presence-missing.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["B2"].font = openpyxl.styles.Font(bold=True)
    workbook.save(source)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: re.sub(r'<c r="B2"[^>]*(?:/>|>.*?</c>)', "", xml, count=1, flags=re.DOTALL),
    )

    report = verify_xlsx_preservation(str(source), str(changed))
    presence = next(change for change in report["changes"] if change["category"] == "cell_presence")

    assert presence["path"] == "worksheets/Sheet/cells/B2"
    assert presence["before"]["present"] is True
    assert presence["before"]["presence_kind"] == "explicit_empty"
    assert presence["after"] is None


def test_hidden_sheet_and_print_settings_are_detected(tmp_path):
    source = tmp_path / "print.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "print"
    sheet.page_setup.orientation = "landscape"
    sheet.print_area = "A1:C5"
    sheet.oddHeader.center.text = "Header"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    workbook.save(source)
    shutil.copy2(source, changed)

    _rewrite_part(
        changed,
        "xl/workbook.xml",
        lambda xml: xml.replace('state="hidden"', 'state="visible"', 1),
    )
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: re.sub(r"<pageSetup\b[^>]*/>", "", xml, count=1),
    )

    report = verify_xlsx_preservation(str(source), str(changed))
    categories = {change["category"] for change in report["changes"]}

    assert "sheet_inventory" in categories
    assert "printing" in categories
    assert report["preservation_ok"] is False


def test_formula_cache_and_calculation_settings_are_detected(tmp_path):
    source = tmp_path / "formula.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "=1+2"
    workbook.save(source)

    _rewrite_part(
        source,
        "xl/worksheets/sheet1.xml",
        lambda xml: re.sub(r"(<f>1\+2</f>)(?:<v\b[^>]*/>|<v>.*?</v>)", r"\1<v>3</v>", xml, count=1),
    )

    def set_calc_mode(xml: str, mode: str) -> str:
        if "calcMode=" in xml:
            return re.sub(r'calcMode="[^"]*"', f'calcMode="{mode}"', xml, count=1)
        return re.sub(r"<calcPr\b", f'<calcPr calcMode="{mode}"', xml, count=1)

    _rewrite_part(source, "xl/workbook.xml", lambda xml: set_calc_mode(xml, "auto"))
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: re.sub(r"(<f>1\+2</f>)(?:<v\b[^>]*/>|<v>.*?</v>)", r"\1<v></v>", xml, count=1),
    )
    _rewrite_part(changed, "xl/workbook.xml", lambda xml: set_calc_mode(xml, "manual"))

    report = verify_xlsx_preservation(str(source), str(changed))
    categories = {change["category"] for change in report["changes"]}

    assert "formula" in categories
    assert "calculation" in categories


def test_table_totals_and_filter_metadata_are_detected(tmp_path):
    source = tmp_path / "table.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Amount"])
    sheet.append(["A", 1])
    sheet.append(["B", 2])
    table = Table(displayName="Sales", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    workbook.save(source)

    def add_totals(xml: str) -> str:
        xml = re.sub(r"<table\b", '<table totalsRowCount="1" totalsRowShown="1"', xml, count=1)
        return re.sub(
            r'(<tableColumn\b[^>]*\bid="2"[^>]*)/>',
            r'\1 totalsRowFunction="sum"><totalsRowFormula>SUBTOTAL(109,Sales[Amount])</totalsRowFormula></tableColumn>',
            xml,
            count=1,
        )

    _rewrite_part(source, "xl/tables/table1.xml", add_totals)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/tables/table1.xml",
        lambda xml: re.sub(r' totalsRowCount="1" totalsRowShown="1"', "", xml, count=1).replace(
            ' totalsRowFunction="sum"><totalsRowFormula>SUBTOTAL(109,Sales[Amount])</totalsRowFormula></tableColumn>',
            "/>",
            1,
        ),
    )

    report = verify_xlsx_preservation(str(source), str(changed))

    assert any(change["category"] == "tables" for change in report["changes"])
    assert report["preservation_ok"] is False


def test_save_creates_backup_and_verify_finds_it(tmp_path, monkeypatch):
    backup_dir = tmp_path / "backups"
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(backup_dir))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "old"
    workbook.save(source)

    key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "new"}}])
    result = M.excel_save(key)

    backup = find_latest_excel_backup(str(source))
    assert backup is not None
    assert backup["backup_path"] in result
    assert openpyxl.load_workbook(backup["backup_path"]).active["A1"].value == "old"
    assert openpyxl.load_workbook(source).active["A1"].value == "new"

    report = json.loads(M.excel_verify_preservation(str(source)))
    assert report["backup"]["backup_path"] == backup["backup_path"]
    assert report["fixture_id"] == "source"
    assert report["fixture_id_source"] == "after_filename"
    assert report["performance"]["package_open_count"] == 2
    assert report["performance"]["part_read_count"] > 0
    assert report["performance"]["metadata_seconds"] >= 0
    assert report["performance"]["signature_seconds"] >= 0
    assert report["performance"]["semantic_verification_seconds"] >= 0
    assert report["performance"]["total_tool_seconds"] >= 0
    assert report["performance"]["peak_memory_bytes"] is not None
    assert any(change["category"] == "cell_value" for change in report["changes"])


def test_expired_backup_and_sidecar_are_deleted(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "old"
    workbook.save(source)

    record = create_excel_backup(str(source), str(source))
    future = datetime.now(timezone.utc) + timedelta(days=3)
    cleanup = cleanup_excel_backups(now=future)

    assert cleanup["removed_backups"] == 1
    assert cleanup["removed_sidecars"] == 1
    assert not Path(record["backup_path"]).exists()
    assert not Path(record["sidecar_path"]).exists()


def test_verifier_classifies_requested_and_approved_changes(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "before-a"
    workbook.active["B1"] = "before-b"
    workbook.save(source)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(">before-a<", ">after-a<", 1).replace(">before-b<", ">after-b<", 1),
    )

    unclassified = verify_xlsx_preservation(str(source), str(changed))
    assert unclassified["preservation_ok"] is False
    assert unclassified["unapproved_difference_count"] >= 2

    report = verify_xlsx_preservation(
        str(source),
        str(changed),
        requested_paths=["worksheets/Sheet/cells/A1/value"],
        approved_normalizations=[{
            "path": "worksheets/Sheet/cells/B1/value",
            "rationale": "Canonical value normalization accepted by the fixture contract.",
            "bidirectional": True,
            "evidence": {"before": "before-b", "after": "after-b"},
        }],
    )

    assert report["preservation_ok"] is True
    assert report["unapproved_difference_count"] == 0
    assert report["classification_counts"]["REQUESTED"] == 1
    assert report["classification_counts"]["APPROVED_NORMALIZATION"] == 1
    assert report["normalization_evidence_complete"] is True
    assert report["before_sha256"] != report["after_sha256"]

    reverse = verify_xlsx_preservation(
        str(changed),
        str(source),
        requested_paths=["worksheets/Sheet/cells/A1/value"],
        approved_normalizations=report["approved_normalization_rules"],
    )
    assert reverse["preservation_ok"] is True
    assert reverse["classification_counts"]["APPROVED_NORMALIZATION"] == 1


def test_verifier_rejects_path_only_normalization_as_verifier_gap(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "before"
    workbook.save(source)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(">before<", ">after<", 1),
    )

    report = verify_xlsx_preservation(
        str(source),
        str(changed),
        approved_normalizations=["worksheets/Sheet/cells/A1/value"],
    )

    assert report["preservation_ok"] is False
    assert report["classification_counts"]["APPROVED_NORMALIZATION"] == 0
    assert report["classification_counts"]["VERIFIER_GAP"] == 1
    assert report["normalization_evidence_complete"] is False
    assert report["approved_normalization_rules"][0]["issues"] == [
        "rationale", "bidirectional", "evidence",
    ]


def test_verifier_classifies_fixture_and_verifier_gaps_as_blocking(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "before-a"
    workbook.active["B1"] = "before-b"
    workbook.save(source)
    shutil.copy2(source, changed)
    _rewrite_part(
        changed,
        "xl/worksheets/sheet1.xml",
        lambda xml: xml.replace(">before-a<", ">after-a<", 1).replace(">before-b<", ">after-b<", 1),
    )

    report = verify_xlsx_preservation(
        str(source),
        str(changed),
        fixture_gap_paths=["worksheets/Sheet/cells/A1/value"],
        verifier_gap_paths=["worksheets/Sheet/cells/B1/value"],
    )

    assert report["classification_counts"]["FIXTURE_GAP"] == 1
    assert report["classification_counts"]["VERIFIER_GAP"] == 1
    assert report["classification_counts"]["PACKAGE_INVALID"] == 0
    assert report["unapproved_difference_count"] == 0
    assert report["blocking_issue_count"] == 2
    assert report["preservation_ok"] is False


def test_package_diff_reports_relationship_and_content_type_changes(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "package"
    workbook.save(source)
    shutil.copy2(source, changed)

    _rewrite_part(
        changed,
        "_rels/.rels",
        lambda xml: xml.replace(
            "</Relationships>",
            '<Relationship Id="rIdAudit" Type="urn:docloupe:test" Target="audit/custom.xml"/></Relationships>',
        ),
    )
    _rewrite_part(
        changed,
        "[Content_Types].xml",
        lambda xml: xml.replace(
            "</Types>",
            '<Override PartName="/audit/custom.xml" ContentType="application/xml"/></Types>',
        ),
    )

    report = verify_xlsx_preservation(str(source), str(changed))
    relationship_changes = report["part_diff"]["relationship_changes"]
    content_type_changes = report["part_diff"]["content_type_changes"]

    assert "_rels/.rels#rIdAudit" in relationship_changes["added"]
    assert "Override:/audit/custom.xml" in content_type_changes["added"]
    assert report["unapproved_difference_count"] >= 2


def test_package_diff_reports_binary_hash_changes(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "binary"
    workbook.save(source)
    with zipfile.ZipFile(source, "a", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("xl/media/audit.bin", b"before-binary")
    shutil.copy2(source, changed)
    _replace_part_bytes(changed, "xl/media/audit.bin", b"after-binary")

    report = verify_xlsx_preservation(str(source), str(changed))

    assert report["part_diff"]["modified_binary"] == ["xl/media/audit.bin"]
    assert report["part_diff"]["binary_hash_changes"][0]["before_sha256"]
    assert report["part_diff"]["binary_hash_changes"][0]["after_sha256"]


def test_report_previews_large_values_without_unbounded_dump(tmp_path):
    source = tmp_path / "source.xlsx"
    changed = tmp_path / "changed.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "a" * 5000
    workbook.save(source)
    shutil.copy2(source, changed)
    changed_workbook = openpyxl.load_workbook(changed)
    changed_workbook.active["A1"] = "b" * 5000
    changed_workbook.save(changed)
    changed_workbook.close()

    report = verify_xlsx_preservation(str(source), str(changed), max_differences=10)
    value_change = next(change for change in report["changes"] if change["category"] == "cell_value")

    assert value_change["before"]["serialized_length"] > 5000
    assert len(value_change["before"]["summary"]) < 1500
    assert value_change["before"]["sha256"]


def test_public_verifier_report_includes_file_health_runtime_and_fixture_id(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_SOURCE_SHA", "source-overlay-sha")
    source = tmp_path / "source.xlsx"
    copied = tmp_path / "copied.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "same"
    workbook.save(source)
    shutil.copy2(source, copied)

    report = json.loads(M.excel_verify_preservation(
        str(copied),
        before_path=str(source),
        fixture_id="fixture-report",
    ))

    assert report["fixture_id"] == "fixture-report"
    assert report["fixture_id_source"] == "provided"
    assert report["backup"] is None
    assert report["files"]["before"]["package_valid"] is True
    assert report["files"]["before"]["loadable"] is True
    assert report["files"]["after"]["package_valid"] is True
    assert report["files"]["after"]["loadable"] is True
    assert report["files"]["before"]["sha256"] == report["before_sha256"]
    assert report["runtime"]["server"] == "excel-tools"
    expected_version = json.loads((ROOT / "server.json").read_text(encoding="utf-8"))["version"]
    assert report["runtime"]["server_version"] == expected_version
    assert report["runtime"]["commit_sha"]
    assert report["runtime"]["source_overlay_sha"] == "source-overlay-sha"
    assert report["runtime"]["library_versions"]["openpyxl"] == "3.1.5"
    assert report["runtime"]["library_versions"]["mcp"] == "1.28.1"
    assert report["runtime"]["python_version"]
    assert set(report["classification_counts"]) == {
        "REQUESTED", "APPROVED_NORMALIZATION", "UNAPPROVED_LOSS",
        "FIXTURE_GAP", "VERIFIER_GAP", "PACKAGE_INVALID",
    }


def test_public_verifier_uses_lightweight_load_probe(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    copied = tmp_path / "copied.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Metadata"
    workbook.active["A1"] = "same"
    workbook.save(source)
    workbook.close()
    shutil.copy2(source, copied)

    def fail_if_serialized(*_args, **_kwargs):
        raise AssertionError("verification metadata must not serialize the full workbook")

    monkeypatch.setattr(M, "serialize_excel", fail_if_serialized)
    report = json.loads(M.excel_verify_preservation(
        str(copied),
        before_path=str(source),
        fixture_id="lightweight-load-probe",
    ))

    assert report["preservation_ok"] is True
    assert report["files"]["before"]["loadable"] is True
    assert report["files"]["after"]["loadable"] is True
    assert report["files"]["before"]["sheet_names"] == ["Metadata"]
    assert report["files"]["after"]["sheet_names"] == ["Metadata"]


def test_public_verifier_reports_invalid_package_without_raising(tmp_path):
    source = tmp_path / "source.xlsx"
    invalid = tmp_path / "invalid.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "source"
    workbook.save(source)
    invalid.write_bytes(b"not-an-ooxml-package")

    report = json.loads(M.excel_verify_preservation(
        str(invalid),
        before_path=str(source),
        fixture_id="invalid-output",
    ))

    assert report["fixture_id"] == "invalid-output"
    assert report["preservation_ok"] is False
    assert report["classification_counts"]["PACKAGE_INVALID"] == 1
    assert report["blocking_issue_count"] == 1
    assert report["files"]["after"]["package_valid"] is False
    assert report["files"]["after"]["loadable"] is False
    assert report["changes"][0]["classification"] == "PACKAGE_INVALID"


def test_backup_retention_boundary_and_hash(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "retention"
    workbook.save(source)

    first = create_excel_backup(str(source), str(source))
    second = create_excel_backup(str(source), str(source))
    assert first["backup_path"] != second["backup_path"]
    assert first["sha256"] == second["sha256"]

    expires_at = datetime.fromisoformat(first["expires_at"].replace("Z", "+00:00"))
    before_boundary = cleanup_excel_backups(now=expires_at - timedelta(microseconds=1))
    assert before_boundary["removed_backups"] == 0
    assert Path(first["backup_path"]).exists()

    at_boundary = cleanup_excel_backups(now=expires_at)
    assert at_boundary["removed_backups"] >= 1
    assert not Path(first["backup_path"]).exists()


def test_multiple_saves_keep_distinct_versioned_backups(tmp_path, monkeypatch):
    backup_dir = tmp_path / "backups"
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(backup_dir))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "version-0"
    workbook.save(source)

    key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "version-1"}}])
    M.excel_save(key)
    first = find_latest_excel_backup(str(source))
    assert first is not None
    assert openpyxl.load_workbook(first["backup_path"]).active["A1"].value == "version-0"

    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "version-2"}}])
    M.excel_save(key)
    second = find_latest_excel_backup(str(source))
    assert second is not None
    assert second["backup_path"] != first["backup_path"]
    assert Path(first["backup_path"]).is_file()
    assert openpyxl.load_workbook(second["backup_path"]).active["A1"].value == "version-1"
    assert openpyxl.load_workbook(source).active["A1"].value == "version-2"

    sidecars = sorted(backup_dir.glob("excel-backup-*.json"))
    records = [json.loads(path.read_text(encoding="utf-8")) for path in sidecars]
    matching = [record for record in records if Path(record["saved_path"]) == source.resolve()]
    assert len(matching) == 2
    assert len({record["backup_path"] for record in matching}) == 2

    report = json.loads(M.excel_verify_preservation(str(source)))
    assert report["backup"]["backup_path"] == second["backup_path"]
    assert any(change["category"] == "cell_value" for change in report["changes"])


def test_restored_backup_validates_loads_and_matches_original_hash(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "before-save"
    workbook.save(source)

    key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {0: "after-save"}}])
    M.excel_save(key)
    backup = find_latest_excel_backup(str(source))
    assert backup is not None
    assert openpyxl.load_workbook(source).active["A1"].value == "after-save"

    shutil.copy2(backup["backup_path"], source)
    assert source.read_bytes() == Path(backup["backup_path"]).read_bytes()

    validation = json.loads(M.excel_validate_workbook(str(source)))
    assert validation["valid"] is True
    restored_key = _session_key(M.excel_load(str(source)))
    restored = json.loads(M.excel_get_cell(restored_key, "Sheet", 0, 0))
    assert restored["value"] == "before-save"
    M.excel_close(restored_key)

    report = json.loads(M.excel_verify_preservation(
        str(source), before_path=backup["backup_path"],
    ))
    assert report["preservation_ok"] is True
    assert report["unapproved_difference_count"] == 0


def test_build_preservation_summary_writes_passing_combined_evidence(tmp_path):
    assert "excel_build_preservation_summary" in M.mcp._tool_manager._tools
    output = tmp_path / "reports" / "final-summary.json"
    coverage_reports = [
        {
            "coverage_id": "exact-87",
            "tested_key_count": 87,
            "source_before_ok_count": 87,
            "fully_reproduced_count": 87,
            "failed": [],
            "fixture_graph_valid": True,
            "path": "exact-87.json",
        },
        {
            "coverage_id": "supplemental-28",
            "case_count": 28,
            "reproduced_count": 28,
            "failed": [],
            "fixture_graph_valid": True,
            "path": "supplemental-28.json",
        },
    ]
    verification_reports = [
        {
            "fixture_id": "public-workflows",
            "before_path": "before.xlsx",
            "after_path": "after.xlsx",
            "preservation_ok": True,
            "classification_counts": {
                "REQUESTED": 2,
                "APPROVED_NORMALIZATION": 1,
                "UNAPPROVED_LOSS": 0,
                "FIXTURE_GAP": 0,
                "VERIFIER_GAP": 0,
                "PACKAGE_INVALID": 0,
            },
            "blocking_issue_count": 0,
            "change_count": 3,
            "changes": [
                {"path": "Book1!A1", "before": "old", "after": "new"},
                {"path": "xl/workbook.xml", "before": {}, "after": {}},
                {"path": "xl/styles.xml", "before": {}, "after": {}},
            ],
            "part_diff": {"added_parts": [], "removed_parts": [], "modified_parts": []},
            "files": {
                "before": {"sha256": "before", "package_valid": True, "loadable": True},
                "after": {"sha256": "after", "package_valid": True, "loadable": True},
            },
        }
    ]
    backup_checks = [
        {"check_id": "backup-before-save", "passed": True},
        {"check_id": "restore", "passed": True},
        {"check_id": "retention-cleanup", "passed": True},
    ]

    payload = M.excel_build_preservation_summary(
        coverage_reports,
        verification_reports,
        backup_checks,
        output_path=str(output),
    )
    report = json.loads(payload)

    assert json.loads(output.read_text(encoding="utf-8")) == report
    assert report["coverage"]["checked_count"] == 115
    assert report["coverage"]["passed_count"] == 115
    assert report["coverage"]["failed_count"] == 0
    assert report["evidence_states"]["source_semantic_present"] == {
        "checked_count": 115,
        "passed_count": 115,
        "failed_count": 0,
        "complete": True,
    }
    assert report["evidence_states"]["legacy_bug_reproduced"]["passed_count"] == 115
    assert report["evidence_states"]["fixed_output_preserved"]["passed_count"] == 1
    assert report["evidence_states"]["fixture_graph_valid"]["passed_count"] == 116
    assert report["verification"]["unapproved_difference_count"] == 0
    assert report["verification"]["invalid_package_count"] == 0
    assert report["backup_retention"]["passed_check_count"] == 3
    assert report["final_gate_passed"] is True
    assert report["failed_gate_reasons"] == {
        "coverage": [],
        "verification": [],
        "backup_retention": [],
    }


def test_build_preservation_summary_distinguishes_reproduction_failure_causes():
    report = json.loads(M.excel_build_preservation_summary(
        [{
            "coverage_id": "reproduction-cases",
            "coverage_mode": "REPRODUCTION",
            "fixture_graph_valid": True,
            "results": [
                {
                    "name": "missing-source-semantic",
                    "source_ok": False,
                    "reproduced": False,
                },
                {
                    "name": "comparator-storage-gap",
                    "source_ok": True,
                    "reproduced": False,
                    "classification": "VERIFIER_GAP",
                },
                {
                    "name": "legacy-loss-reproduced",
                    "source_ok": True,
                    "reproduced": True,
                },
            ],
        }],
        [{
            "fixture_id": "post-fix-output",
            "preservation_ok": True,
            "classification_counts": {},
            "fixture_graph_valid": True,
        }],
        [{"check_id": "backup", "passed": True}],
    ))

    coverage = report["coverage"]["reports"][0]
    assert coverage["failure_classifications"] == {
        "FIXTURE_GAP": ["missing-source-semantic"],
        "VERIFIER_GAP": ["comparator-storage-gap"],
        "NOT_REPRODUCED": [],
    }
    assert coverage["evidence_states"]["source_semantic_present"]["passed_count"] == 2
    assert coverage["evidence_states"]["legacy_bug_reproduced"]["passed_count"] == 1
    assert report["evidence_states"]["fixed_output_preserved"]["complete"] is True
    assert report["evidence_states"]["fixture_graph_valid"]["complete"] is True
    assert report["final_gate_passed"] is False


def test_build_preservation_summary_blocks_all_failure_classes_and_truncates_changes():
    blocking_counts = {
        "REQUESTED": 0,
        "APPROVED_NORMALIZATION": 0,
        "UNAPPROVED_LOSS": 1,
        "FIXTURE_GAP": 1,
        "VERIFIER_GAP": 1,
        "PACKAGE_INVALID": 1,
    }
    changes = [
        {"path": f"part-{index}", "before": index, "after": index + 1}
        for index in range(4)
    ]

    report = json.loads(M.excel_build_preservation_summary(
        [{
            "coverage_id": "exact-87",
            "checked_count": 87,
            "passed_count": 86,
            "failed_items": ["rich-text"],
            "failed_count": 1,
        }],
        [{
            "fixture_id": "broken-fixture",
            "preservation_ok": True,
            "classification_counts": blocking_counts,
            "blocking_issue_count": 0,
            "change_count": 4,
            "changes": changes,
            "part_diff": {"removed_parts": ["xl/styles.xml"]},
        }],
        [{"check_id": "retention-cleanup", "passed": False, "details": {"removed": 0}}],
        max_changes_per_report=2,
    ))

    verification = report["verification"]
    verification_detail = verification["reports"][0]
    assert report["final_gate_passed"] is False
    assert report["failed_gate_reasons"] == {
        "coverage": ["exact-87"],
        "verification": ["broken-fixture"],
        "backup_retention": ["retention-cleanup"],
    }
    assert verification["unapproved_difference_count"] == 1
    assert verification["invalid_package_count"] == 1
    assert verification["fixture_gap_count"] == 1
    assert verification["verifier_gap_count"] == 1
    assert verification_detail["change_count"] == 4
    assert len(verification_detail["changes"]) == 2
    assert verification_detail["changes_truncated"] is True
    assert verification_detail["part_diff"] == {"removed_parts": ["xl/styles.xml"]}


def test_build_preservation_summary_refuses_overwrite_without_explicit_opt_in(tmp_path):
    output = tmp_path / "summary.json"
    output.write_text("existing", encoding="utf-8")
    arguments = (
        [{
            "coverage_id": "exact-87",
            "checked_count": 87,
            "passed_count": 87,
            "fixture_graph_valid": True,
        }],
        [{
            "fixture_id": "clean",
            "preservation_ok": True,
            "classification_counts": {},
            "fixture_graph_valid": True,
        }],
        [{"check_id": "backup", "passed": True}],
    )

    with pytest.raises(FileExistsError, match="Pass overwrite=true"):
        M.excel_build_preservation_summary(*arguments, output_path=str(output))

    report = json.loads(M.excel_build_preservation_summary(
        *arguments,
        output_path=str(output),
        overwrite=True,
    ))
    assert json.loads(output.read_text(encoding="utf-8")) == report
    assert report["final_gate_passed"] is True


@pytest.mark.parametrize(
    ("coverage_reports", "verification_reports", "backup_checks", "message"),
    [
        (
            [{"checked_count": -1, "passed_count": 0}],
            [{"preservation_ok": True, "classification_counts": {}}],
            [{"passed": True}],
            "non-negative integer",
        ),
        (
            [{"checked_count": 1, "passed_count": 1}],
            [{"preservation_ok": True, "classification_counts": {}, "change_count": -1}],
            [{"passed": True}],
            "change_count must be a non-negative integer",
        ),
        (
            [{"checked_count": 1, "passed_count": 1}],
            [{
                "preservation_ok": True,
                "classification_counts": {},
                "blocking_issue_count": "1",
            }],
            [{"passed": True}],
            "blocking_issue_count must be a non-negative integer",
        ),
        (
            [{"checked_count": 1, "passed_count": 1}],
            [{"preservation_ok": True, "classification_counts": {}}],
            [{"passed": 1}],
            "passed must be true or false",
        ),
    ],
)
def test_build_preservation_summary_rejects_invalid_evidence(
    coverage_reports, verification_reports, backup_checks, message,
):
    with pytest.raises(ValueError, match=message):
        M.excel_build_preservation_summary(
            coverage_reports,
            verification_reports,
            backup_checks,
        )
