"""
Regression tests for no-edit round-trip fidelity fixes:

1. Sheets with BOTH a DrawingML drawing (image/chart) and comments/hyperlinks:
   the drawing relationship must be added to the existing sheet rels, the
   comments VML must not be misclassified as a drawing, and <drawing> must
   precede <legacyDrawing> in the worksheet XML.
2. gray125 pattern fill with explicit colors is preserved.
3. fitToPage print setting is preserved.
4. Per-cell protection (locked=False / hidden=True) is preserved.
5. Injected conditionalFormatting blocks appear before pageMargins
   (CT_Worksheet element order).
"""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Protection
from openpyxl.styles.colors import Color

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

from core import inspect_xlsx_package, reconstruct_excel, serialize_excel  # noqa: E402


def _read_part(path: Path, part_name: str) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        return zf.read(part_name).decode("utf-8")


def _assert_rels_complete(path: Path) -> None:
    """Every r:id referenced by a worksheet must exist in its rels part."""
    with zipfile.ZipFile(path, "r") as zf:
        names = set(zf.namelist())
        for part in [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]:
            xml = zf.read(part).decode("utf-8")
            rids = set(re.findall(r'r:id="([^"]+)"', xml))
            rels = part.rsplit("/", 1)[0] + "/_rels/" + part.rsplit("/", 1)[1] + ".rels"
            have = (set(re.findall(r'Id="([^"]+)"', zf.read(rels).decode("utf-8")))
                    if rels in names else set())
            missing = rids - have
            assert not missing, f"{part} references missing rels: {sorted(missing)}"


def _tiny_png(path: Path) -> Path:
    from PIL import Image as PILImage
    PILImage.new("RGB", (4, 4), (255, 0, 0)).save(path)
    return path


def _add_textbox_drawing(path: Path, text: str = "Original Box") -> None:
    anchor_xml = f'''<xdr:twoCellAnchor><xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>4</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>5</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="99" name="TextBox 1"/><xdr:cNvSpPr txBox="1"/></xdr:nvSpPr><xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr><xdr:txBody><a:bodyPr wrap="square"/><a:lstStyle/><a:p><a:r><a:t>{text}</a:t></a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor>'''
    new_drawing_xml = f'''<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">{anchor_xml}</xdr:wsDr>'''
    rels_xml = '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>'''
    with zipfile.ZipFile(path, "r") as zin:
        names = set(zin.namelist())
        existing_rels = "xl/worksheets/_rels/sheet1.xml.rels" in names
        drawing_file = None
        if existing_rels:
            rels = zin.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
            for rel_match in re.finditer(r"<Relationship\b([^>]*)/>", rels):
                attrs = rel_match.group(1)
                type_m = re.search(r'Type="([^"]+)"', attrs)
                target_m = re.search(r'Target="([^"]+)"', attrs)
                if type_m and target_m and type_m.group(1).rstrip("/").endswith("/drawing"):
                    target = target_m.group(1)
                    drawing_file = "xl/" + target[3:] if target.startswith("../") else "xl/drawings/" + target.rsplit("/", 1)[-1]
                    break
    tmp = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if drawing_file and item.filename == drawing_file:
                xml = raw.decode("utf-8")
                if "</xdr:wsDr>" in xml:
                    xml = xml.replace("</xdr:wsDr>", anchor_xml + "</xdr:wsDr>")
                else:
                    root_match = re.search(r"<wsDr\b[^>]*>", xml)
                    if root_match and "xmlns:a=" not in root_match.group(0):
                        xml = xml[:root_match.end() - 1] + ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"' + xml[root_match.end() - 1:]
                    xml = xml.replace("</wsDr>", anchor_xml.replace("xdr:", "") + "</wsDr>")
                raw = xml.encode("utf-8")
            elif not drawing_file and item.filename == "xl/worksheets/sheet1.xml":
                xml = raw.decode("utf-8")
                xml = xml.replace("</worksheet>", '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1"/></worksheet>')
                raw = xml.encode("utf-8")
            elif not drawing_file and item.filename == "[Content_Types].xml":
                xml = raw.decode("utf-8")
                if 'PartName="/xl/drawings/drawing1.xml"' not in xml:
                    xml = xml.replace("</Types>", '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
                raw = xml.encode("utf-8")
            zout.writestr(item, raw)
        if not drawing_file:
            zout.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels_xml.encode("utf-8"))
            zout.writestr("xl/drawings/drawing1.xml", new_drawing_xml.encode("utf-8"))
    tmp.replace(path)


def test_reconstruct_validates_before_replace(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    wb.active["A1"] = "safe"
    wb.save(src)
    out.write_text("keep me", encoding="utf-8")

    data = serialize_excel(str(src))
    data["sheets"][0]["drawing_data"] = {
        "drawing_file": "xl/drawings/drawing1.xml",
        "drawing_xml": "<not xml",
        "drawing_rels": None,
        "files": {},
    }

    with pytest.raises(ValueError, match="failed validation"):
        reconstruct_excel(data, str(out))

    assert out.read_text(encoding="utf-8") == "keep me"


def test_noop_roundtrip_produces_valid_package(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["A1"].font = openpyxl.styles.Font(color="FFFF0000", strike=True)
    wb.save(src)

    warnings = reconstruct_excel(serialize_excel(str(src)), str(out))

    assert not [w for w in warnings if "failed" in w.lower()]
    assert inspect_xlsx_package(str(out))["valid"]
    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["A1"].value == "hello"
    assert wb2.active["A1"].font.strike is True
    assert wb2.active["A1"].font.color.rgb == "FFFF0000"


def test_drawing_with_comments_and_hyperlinks_keeps_rels_intact(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "note"
    ws["A1"].comment = Comment("a comment", "tester")
    ws["B1"] = "link"
    ws["B1"].hyperlink = "https://example.com/"
    from openpyxl.drawing.image import Image
    ws.add_image(Image(str(_tiny_png(tmp_path / "img.png"))), "D2")
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    _assert_rels_complete(out)

    with zipfile.ZipFile(out) as zf:
        names = set(zf.namelist())
        drawings = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        # exactly one DrawingML drawing, and it really is DrawingML (not VML)
        assert len(drawings) == 1
        assert zf.read(drawings[0]).lstrip().startswith(b"<wsDr")
        # the image survived
        assert any(n.startswith("xl/media/") for n in names)

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    draw_pos = sheet_xml.find("<drawing ")
    legacy_pos = sheet_xml.find("<legacyDrawing")
    assert draw_pos != -1
    if legacy_pos != -1:
        assert draw_pos < legacy_pos, "<drawing> must precede <legacyDrawing>"

    # comment and hyperlink survived
    wb2 = openpyxl.load_workbook(out)
    assert wb2.active["A1"].comment is not None
    assert wb2.active["B1"].hyperlink is not None

    # SECOND round-trip: the first output uses ../drawings/… relative rel
    # targets (like Excel-authored files) — the image must survive again.
    out2 = tmp_path / "out2.xlsx"
    reconstruct_excel(serialize_excel(str(out)), str(out2))
    _assert_rels_complete(out2)
    with zipfile.ZipFile(out2) as zf:
        assert any(n.startswith("xl/media/") for n in zf.namelist()), \
            "image lost on second round-trip"
    assert len(openpyxl.load_workbook(out2).active._images) == 1


def test_realistic_chart_image_textbox_corpus_roundtrip(tmp_path):
    src = tmp_path / "corpus.xlsx"
    out = tmp_path / "corpus-out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    for row in range(1, 6):
        ws.cell(row=row, column=1, value=f"Item {row}")
        ws.cell(row=row, column=2, value=row * 10)
    ws["D1"] = "comment"
    ws["D1"].comment = Comment("visible note", "tester")
    from openpyxl.chart import BarChart, Reference
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    chart.title = "Totals"
    ws.add_chart(chart, "F2")
    from openpyxl.drawing.image import Image
    ws.add_image(Image(str(_tiny_png(tmp_path / "img.png"))), "D4")
    wb.save(src)
    _add_textbox_drawing(src)

    data = serialize_excel(str(src))
    assert data["sheets"][0]["shapes"]
    reconstruct_excel(data, str(out))

    report = inspect_xlsx_package(str(out))
    assert report["valid"], report
    _assert_rels_complete(out)
    with zipfile.ZipFile(out) as zf:
        names = set(zf.namelist())
        assert any(n.startswith("xl/charts/") for n in names)
        assert any(n.startswith("xl/media/") for n in names)
        assert any(n.startswith("xl/drawings/") for n in names)
        drawing_xml = "\n".join(zf.read(n).decode("utf-8") for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n))
        assert "Original Box" in drawing_xml

    wb2 = openpyxl.load_workbook(out)
    assert wb2["Dashboard"]["D1"].comment is not None


def test_gray125_fill_with_colors_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "patterned"
    ws["A1"].fill = PatternFill(
        fill_type="gray125",
        fgColor=Color(rgb="FF4472C4"),
        bgColor=Color(rgb="FFFFFFFF"),
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    fill = wb2.active["A1"].fill
    assert fill.fill_type == "gray125"
    assert fill.fgColor.rgb == "FF4472C4"
    assert fill.bgColor.rgb == "FFFFFFFF"


def test_fit_to_page_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    pspr = wb2.active.sheet_properties.pageSetUpPr
    assert pspr is not None and pspr.fitToPage
    assert wb2.active.page_setup.fitToWidth == 1


def test_cell_protection_is_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "locked"
    ws["B1"] = "unlocked"
    ws["B1"].protection = Protection(locked=False)
    ws["C1"] = "hidden formula"
    ws["C1"].protection = Protection(locked=True, hidden=True)
    ws.protection.sheet = True
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    assert ws2.protection.sheet
    assert ws2["A1"].protection.locked is True
    assert ws2["B1"].protection.locked is False
    assert ws2["C1"].protection.hidden is True


def test_mc_ignorable_only_lists_declared_prefixes(tmp_path):
    """Excel refuses files whose mc:Ignorable names undeclared prefixes.

    Real Excel-authored sheets carry mc:Ignorable="x14ac xr xr2 xr3" while the
    reconstructed root only declares the prefixes actually used (x14ac).
    """
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src)

    # Patch source like a real Excel file: extra Ignorable prefixes + x14ac attr
    def patch(xml):
        xml = re.sub(
            r"<worksheet\b[^>]*>",
            ('<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
             'mc:Ignorable="x14ac xr xr2 xr3" '
             'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
             'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
             'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
             'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3">'),
            xml, count=1)
        return re.sub(
            r"<sheetFormatPr\b[^>]*/>",
            '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>',
            xml, count=1)

    tmp = src.with_suffix(".tmp")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                raw = patch(raw.decode("utf-8")).encode("utf-8")
            zout.writestr(item, raw)
    tmp.replace(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    root = re.search(r"<worksheet\b([^>]*)>", sheet_xml).group(1)
    attrs = dict(re.findall(r'([\w:.-]+)="([^"]*)"', root))
    ignorable = attrs.get("mc:Ignorable", "")
    declared = {k[6:] for k in attrs if k.startswith("xmlns:")}
    undeclared = [t for t in ignorable.split() if t not in declared]
    assert not undeclared, f"mc:Ignorable lists undeclared prefixes: {undeclared}"
    # the x14ac extension attr must still be preserved
    assert "x14ac:dyDescent" in sheet_xml
    assert "xmlns:x14ac" in root


def test_prefixed_worksheet_root_attrs_keep_namespace(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src)

    def patch(xml):
        return re.sub(
            r"<worksheet\b[^>]*>",
            ('<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
             'mc:Ignorable="xr" '
             'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
             'xr:uid="{00000000-0001-0000-0300-000000000000}">'),
            xml,
            count=1,
        )

    tmp = src.with_suffix(".tmp")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                raw = patch(raw.decode("utf-8")).encode("utf-8")
            zout.writestr(item, raw)
    tmp.replace(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    report = inspect_xlsx_package(str(out))
    assert report["valid"], report
    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    root = re.search(r"<worksheet\b([^>]*)>", sheet_xml).group(1)
    assert "xmlns:xr" in root
    assert "xr:uid" in root


def test_diagonal_borders_are_preserved(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    from openpyxl.styles import Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "X"
    ws["B2"].border = Border(
        diagonal=Side(style="thin", color=Color(rgb="FFFF0000")),
        diagonalUp=True, diagonalDown=True,
    )
    ws["C3"] = "Y"
    ws["C3"].border = Border(
        diagonal=Side(style="medium"), diagonalDown=True,
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    b2 = wb2.active["B2"].border
    assert b2.diagonal.border_style == "thin"
    assert b2.diagonal.color.rgb == "FFFF0000"
    assert b2.diagonalUp and b2.diagonalDown
    c3 = wb2.active["C3"].border
    assert c3.diagonal.border_style == "medium"
    assert c3.diagonalDown and not c3.diagonalUp


def test_conditional_formatting_injected_in_schema_order(tmp_path):
    src = tmp_path / "src.xlsx"
    out = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=i)
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        "A1:A5",
        CellIsRule(operator="greaterThan", formula=["3"],
                   fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE",
                                    fill_type="solid")),
    )
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    sheet_xml = _read_part(out, "xl/worksheets/sheet1.xml")
    cf_pos = sheet_xml.find("<conditionalFormatting")
    pm_pos = sheet_xml.find("<pageMargins")
    assert cf_pos != -1, "conditional formatting lost"
    assert pm_pos == -1 or cf_pos < pm_pos, \
        "conditionalFormatting must precede pageMargins (CT_Worksheet order)"

    # CT_Stylesheet order: injected dxfs must precede tableStyles/colors
    styles_xml = _read_part(out, "xl/styles.xml")
    dxfs_pos = styles_xml.find("<dxfs")
    assert dxfs_pos != -1, "dxfs section lost"
    for later in ("<tableStyles", "<colors"):
        lp = styles_xml.find(later)
        assert lp == -1 or dxfs_pos < lp, \
            f"dxfs must precede {later} (CT_Stylesheet order)"

    wb2 = openpyxl.load_workbook(out)
    cf_ranges = [str(r.sqref) for r in wb2.active.conditional_formatting]
    assert cf_ranges == ["A1:A5"]
