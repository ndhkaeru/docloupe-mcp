"""
Verification tests for PRESERVATION_FIX_CHECKLIST.md bullets covering
excel_set_auto_filter, the table tools (excel_add_table / excel_update_table /
excel_delete_table), the hyperlink tools, the comment tools,
excel_set_ignored_errors, and the named-style tools. These exercise the
PUBLIC MCP tool functions in servers/excel/main.py exactly as an agent would
call them -- never the internal helpers directly.

Bug found and fixed here (see servers/excel/core.py):
  - excel_add_table's style dict only recognized the short internal alias
    keys (showFirstCol/showLastCol/showColStripes); the canonical OOXML
    tableStyleInfo attribute names (showFirstColumn/showLastColumn/
    showColumnStripes) that the checklist itself names were silently ignored
    (always written as False). Fixed to accept either spelling.

Genuine structural gaps confirmed by real round-trip experiments and
documented below with strict xfail tests (not fixed -- each is a sizable
unimplemented feature, matching still-open PRESERVATION_FIX_CHECKLIST.md
lines rather than a small, contained bug):
  - excel_set_auto_filter's filter_columns and sort_state are entirely
    in-memory decoration: reconstruct_excel only ever writes
    ws.auto_filter.ref, never ws.auto_filter.filterColumn/sortState, and
    serialize_excel never reads them back either.
  - excel_add_table/excel_update_table only ever persist name/ref/the four
    tableStyleInfo flags. tableType, headerRowCount, totalsRowCount,
    totalsRowShown, insertRow, published, table-scoped autoFilter/sortState,
    and per-column totalsRowFunction/totalsRowLabel/calculatedColumnFormula
    are silently dropped on save (matches checklist's own still-unchecked
    "Ho tro day du tableStyleInfo flags" line and its wider table section).
  - excel_set_comment/excel_remove_comment go through openpyxl's plain
    Comment object, which regenerates default-only VML on every save; a
    pre-existing comment's custom VML shape sizing is silently reset to
    openpyxl's default the moment ANY comment tool touches the same sheet,
    even on an unrelated cell.
  - excel_set_ignored_errors's rules are pure in-memory bookkeeping:
    reconstruct_excel never writes <ignoredErrors>, and serialize_excel never
    reads it back (matches PRESERVATION_AUDIT_87.md finding #27).
  - excel_add_named_style/excel_update_named_style/excel_delete_named_style
    and a cell's named_style assignment are pure in-memory bookkeeping:
    reconstruct_excel never builds <cellStyles>/<cellStyleXfs> from
    data["named_styles"], and never applies a cell's named_style at save
    time, so a freshly reloaded cell that had a custom named style silently
    falls back to "Normal" (matches checklist's still-unchecked
    "Quan ly cellStyles, cellStyleXfs va lien ket xfId" line).
"""
import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _load_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _new_session(sheet_names=("S",)):
    created = json.loads(M.excel_create_workbook(sheet_names=list(sheet_names)))
    return created["session_key"]


# ---------------------------------------------------------------------------
# excel_set_auto_filter
# ---------------------------------------------------------------------------

def test_auto_filter_patch_mode_preserves_other_columns_and_sort_state():
    key = _new_session()
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "H1", 1: "H2"}}])

    M.excel_set_auto_filter(
        key, "S", ref="A1:B10",
        filter_columns=[{"colId": "0", "filters": ["x", "y"]}],
        sort_state={"ref": "A1:B10", "conditions": [{"ref": "A1:A10"}]},
    )
    # Patching column 1 only must not drop column 0's criteria or the sort state.
    result = json.loads(M.excel_set_auto_filter(key, "S", filter_columns=[{"colId": "1", "filters": ["z"]}]))
    model = result["after"]
    by_id = {c["colId"]: c for c in model["filter_columns"]}
    assert by_id["0"]["filters"] == ["x", "y"]
    assert by_id["1"]["filters"] == ["z"]
    assert model["sort_state"]["conditions"] == [{"ref": "A1:A10"}]
    assert model["ref"] == "A1:B10"

    # mode="replace" wipes prior filter_columns/sort_state, per its contract.
    replaced = json.loads(M.excel_set_auto_filter(key, "S", ref="A1:B10", filter_columns=[{"colId": "0"}], mode="replace"))
    assert replaced["after"]["filter_columns"] == [{"colId": "0"}]
    assert replaced["after"].get("sort_state") is None


def test_auto_filter_ref_roundtrips_but_filter_columns_and_sort_state_do_not(tmp_path):
    key = _new_session()
    out = tmp_path / "autofilter.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "H1", 1: "H2"}}])
    M.excel_set_auto_filter(
        key, "S", ref="A1:B10",
        filter_columns=[{"colId": "0", "filters": ["x", "y"]}],
        sort_state={"ref": "A1:B10", "conditions": [{"ref": "A1:A10"}]},
    )
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    wb2 = openpyxl.load_workbook(str(out))
    assert wb2["S"].auto_filter.ref == "A1:B10"  # PASS: the range itself survives
    wb2.close()


def test_auto_filter_columns_and_sort_state_survive_save_reload(tmp_path):
    key = _new_session()
    out = tmp_path / "autofilter_gap.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "H1", 1: "H2"}}])
    M.excel_set_auto_filter(
        key, "S", ref="A1:B10",
        filter_columns=[{"colId": "0", "filters": ["x", "y"]}],
        sort_state={"ref": "A1:B10", "conditions": [{"ref": "A1:A10"}]},
    )
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    af = wb2["S"].auto_filter
    assert list(af.filterColumn) != []
    assert af.sortState is not None
    wb2.close()


# ---------------------------------------------------------------------------
# Tables
# ---------------------------------------------------------------------------

def test_table_add_update_delete_style_flags_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "table.xlsx"
    M.excel_edit_cells(key, "S", [
        {"row_index": 0, "edits": {0: "Item", 1: "Qty"}},
        {"row_index": 1, "edits": {0: "A", 1: 1}},
        {"row_index": 2, "edits": {0: "B", 1: 2}},
    ])
    M.excel_add_table(key, "S", "T1", "A1:B3", style={
        "name": "TableStyleMedium9",
        "showFirstColumn": True, "showLastColumn": True,
        "showRowStripes": True, "showColumnStripes": True,
    })
    tables = json.loads(M.excel_list_tables(key))
    assert tables["tables"][0]["name"] == "T1"
    assert tables["tables"][0]["ref"] == "A1:B3"

    M.excel_update_table(key, "T1", {"ref": "A1:B4"})
    updated = json.loads(M.excel_list_tables(key))
    assert updated["tables"][0]["ref"] == "A1:B4"
    # Partial update must not drop the style set earlier.
    assert updated["tables"][0]["style"]["name"] == "TableStyleMedium9"

    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    wb2 = openpyxl.load_workbook(str(out))
    t = wb2["S"].tables["T1"]
    assert t.ref == "A1:B4"
    ts = t.tableStyleInfo
    assert ts.name == "TableStyleMedium9"
    assert ts.showFirstColumn is True
    assert ts.showLastColumn is True
    assert ts.showColumnStripes is True
    wb2.close()

    key2 = _load_key(M.excel_load(str(out)))
    M.excel_delete_table(key2, "T1")
    out2 = tmp_path / "table_deleted.xlsx"
    M.excel_save(key2, str(out2))
    wb3 = openpyxl.load_workbook(str(out2))
    assert "T1" not in wb3["S"].tables
    wb3.close()


def test_table_structural_fields_survive_save_reload(tmp_path):
    key = _new_session()
    out = tmp_path / "table_structural.xlsx"
    M.excel_edit_cells(key, "S", [
        {"row_index": 0, "edits": {0: "Item", 1: "Qty"}},
        {"row_index": 1, "edits": {0: "A", 1: 1}},
        {"row_index": 2, "edits": {0: "Total", 1: 1}},
    ])
    M.excel_add_table(key, "S", "T1", "A1:B3", table={
        "headerRowCount": 1, "totalsRowCount": 1, "totalsRowShown": True,
        "columns": [
            {"name": "Item"},
            {"name": "Qty", "totalsRowFunction": "sum", "totalsRowLabel": "Total"},
        ],
    })
    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    t = wb2["S"].tables["T1"]
    assert t.totalsRowCount == 1
    cols = {c.name: c for c in t.tableColumns}
    assert cols["Qty"].totalsRowFunction == "sum"
    assert cols["Qty"].totalsRowLabel == "Total"
    wb2.close()


def test_table_update_expand_reconciles_columns_and_nested_ranges(tmp_path):
    key = _new_session()
    out = tmp_path / "table_expanded.xlsx"
    M.excel_edit_cells(key, "S", [
        {"row_index": 0, "edits": {0: "Item", 1: "Qty", 2: "Price"}},
        {"row_index": 1, "edits": {0: "A", 1: 1, 2: 10}},
        {"row_index": 2, "edits": {0: "B", 1: 2, 2: 20}},
        {"row_index": 3, "edits": {0: "C", 1: 3, 2: 30}},
        {"row_index": 4, "edits": {0: "D", 1: 4, 2: 40}},
    ])
    M.excel_add_table(key, "S", "T1", "A1:B4", table={
        "columns": [
            {"id": 4, "name": "Item"},
            {"id": 8, "name": "Qty", "totalsRowFunction": "sum", "calculatedColumnFormula": "[@Qty]*2"},
        ],
        "auto_filter": {
            "ref": "A1:B4",
            "filter_columns": [{"colId": 0, "filters": ["A"]}, {"colId": 1, "filters": ["1"]}],
            "sort_state": {"ref": "A1:B4", "conditions": [{"ref": "B2:B4"}]},
        },
        "sort_state": {"ref": "A1:B4", "conditions": [{"ref": "B2:B4"}]},
    })

    result = json.loads(M.excel_update_table(key, "T1", {"ref": "A1:C5"}))
    table = result["after"]
    assert table["ref"] == "A1:C5"
    assert [column["id"] for column in table["columns"][:2]] == [4, 8]
    assert table["columns"][1]["totalsRowFunction"] == "sum"
    assert table["columns"][1]["calculatedColumnFormula"] == "[@Qty]*2"
    assert table["columns"][2]["name"] == "Price"
    assert table["columns"][2]["id"] not in {4, 8}
    assert table["auto_filter"]["ref"] == "A1:C5"
    assert table["auto_filter"]["sort_state"]["ref"] == "A1:C5"
    assert table["auto_filter"]["sort_state"]["conditions"][0]["ref"] == "B2:B5"
    assert table["sort_state"]["ref"] == "A1:C5"
    assert table["sort_state"]["conditions"][0]["ref"] == "B2:B5"

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out), data_only=False)
    saved = wb2["S"].tables["T1"]
    assert saved.ref == "A1:C5"
    assert [column.id for column in saved.tableColumns[:2]] == [4, 8]
    assert [column.name for column in saved.tableColumns] == ["Item", "Qty", "Price"]
    assert saved.tableColumns[1].totalsRowFunction == "sum"
    assert saved.tableColumns[1].calculatedColumnFormula.attr_text == "[@Qty]*2"
    assert saved.autoFilter.ref == "A1:C5"
    assert saved.autoFilter.sortState.ref == "A1:C5"
    assert saved.autoFilter.sortState.sortCondition[0].ref == "B2:B5"
    assert saved.sortState.ref == "A1:C5"
    assert saved.sortState.sortCondition[0].ref == "B2:B5"
    wb2.close()


def test_table_update_shrink_drops_trailing_columns_and_filters(tmp_path):
    key = _new_session()
    out = tmp_path / "table_shrunk.xlsx"
    M.excel_edit_cells(key, "S", [
        {"row_index": 0, "edits": {0: "A", 1: "B", 2: "C"}},
        {"row_index": 1, "edits": {0: 1, 1: 2, 2: 3}},
        {"row_index": 2, "edits": {0: 4, 1: 5, 2: 6}},
    ])
    M.excel_add_table(key, "S", "T1", "A1:C3", table={
        "columns": [
            {"id": 2, "name": "A"},
            {"id": 5, "name": "B", "totalsRowLabel": "kept"},
            {"id": 9, "name": "C", "calculatedColumnFormula": "[@A]+[@B]"},
        ],
        "auto_filter": {
            "ref": "A1:C3",
            "filter_columns": [{"colId": 0, "filters": ["1"]}, {"colId": 2, "filters": ["3"]}],
        },
    })

    table = json.loads(M.excel_update_table(key, "T1", {"ref": "A1:B3"}))["after"]
    assert [column["id"] for column in table["columns"]] == [2, 5]
    assert table["columns"][1]["totalsRowLabel"] == "kept"
    assert [column["colId"] for column in table["auto_filter"]["filter_columns"]] == [0]
    assert table["auto_filter"]["ref"] == "A1:B3"

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    saved = wb2["S"].tables["T1"]
    assert saved.ref == "A1:B3"
    assert [column.id for column in saved.tableColumns] == [2, 5]
    assert [column.colId for column in saved.autoFilter.filterColumn] == [0]
    wb2.close()


def test_table_update_rename_and_nested_patch_are_atomic(tmp_path):
    key = _new_session(("S", "Other"))
    out = tmp_path / "table_renamed.xlsx"
    for sheet_name in ("S", "Other"):
        M.excel_edit_cells(key, sheet_name, [
            {"row_index": 0, "edits": {0: "A", 1: "B"}},
            {"row_index": 1, "edits": {0: 1, 1: 2}},
        ])
    M.excel_add_table(key, "S", "T1", "A1:B2", table={
        "columns": [
            {"id": 1, "name": "A"},
            {"id": 2, "name": "B", "totalsRowFunction": "sum"},
        ],
        "auto_filter": {
            "ref": "A1:B2",
            "filter_columns": [{"colId": 1, "filters": ["2"]}],
            "sort_state": {"ref": "A1:B2", "conditions": [{"ref": "B2:B2"}]},
        },
    }, style={
        "name": "TableStyleMedium9",
        "showFirstColumn": True,
        "showLastColumn": True,
        "showRowStripes": True,
        "showColumnStripes": True,
    })
    M.excel_add_table(key, "Other", "T2", "A1:B2")

    table = json.loads(M.excel_update_table(key, "T1", {
        "name": "Renamed",
        "style": {"showRowStripes": False},
        "auto_filter": {"sort_state": {"caseSensitive": True}},
        "columns": [{}, {"name": "Amount"}],
    }))["after"]
    assert table["name"] == table["displayName"] == "Renamed"
    assert table["style"]["name"] == "TableStyleMedium9"
    assert table["style"]["showFirstColumn"] is True
    assert table["style"]["showRowStripes"] is False
    assert table["auto_filter"]["filter_columns"] == [{"colId": 1, "filters": ["2"]}]
    assert table["auto_filter"]["sort_state"]["conditions"] == [{"ref": "B2:B2"}]
    assert table["auto_filter"]["sort_state"]["caseSensitive"] is True
    assert table["columns"][1]["name"] == "Amount"
    assert table["columns"][1]["totalsRowFunction"] == "sum"

    with pytest.raises(ValueError, match="must match"):
        M.excel_update_table(key, "Renamed", {"name": "X", "displayName": "Y"})
    with pytest.raises(ValueError, match="unique workbook-wide"):
        M.excel_update_table(key, "Renamed", {"displayName": "T2"})
    with pytest.raises(ValueError, match="one worksheet-local A1 range"):
        M.excel_update_table(key, "Renamed", {"ref": "A1:B2 C3:D4"})
    unchanged = json.loads(M.excel_list_tables(key, "S"))["tables"][0]
    assert unchanged["name"] == unchanged["displayName"] == "Renamed"
    assert unchanged["ref"] == "A1:B2"

    M.excel_save(key, str(out))
    wb2 = openpyxl.load_workbook(str(out))
    assert "T1" not in wb2["S"].tables
    saved = wb2["S"].tables["Renamed"]
    assert saved.name == saved.displayName == "Renamed"
    assert saved.tableStyleInfo.name == "TableStyleMedium9"
    assert saved.tableStyleInfo.showFirstColumn is True
    assert saved.tableStyleInfo.showRowStripes is False
    assert saved.tableColumns[1].name == "Amount"
    assert saved.tableColumns[1].totalsRowFunction == "sum"
    assert saved.autoFilter.sortState.caseSensitive is True
    assert saved.autoFilter.sortState.sortCondition[0].ref == "B2:B2"
    wb2.close()


# ---------------------------------------------------------------------------
# Hyperlinks
# ---------------------------------------------------------------------------

def test_hyperlink_set_remove_and_partial_update_preserves_other_fields(tmp_path):
    key = _new_session()
    out = tmp_path / "hyperlinks.xlsx"

    M.excel_set_hyperlink(key, "S", "A1", target="https://example.com/", display="Example", tooltip="tip1")
    # Changing only the target must not wipe display/tooltip.
    M.excel_set_hyperlink(key, "S", "A1", target="https://example.com/changed")
    session = M._get_session(key)
    sheet = next(s for s in session["sheets"] if s["name"] == "S")
    assert sheet["hyperlinks"]["A1"]["target"] == "https://example.com/changed"
    assert sheet["hyperlinks"]["A1"]["tooltip"] == "tip1"
    cell = json.loads(M.excel_get_cell(key, "S", 0, 0))
    assert cell["value"] == "Example"  # display text untouched

    M.excel_set_hyperlink(key, "S", "A2", location="S!B2", display="InternalLink")
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    key2 = _load_key(M.excel_load(str(out)))
    sess2 = M._get_session(key2)
    sheet2 = next(s for s in sess2["sheets"] if s["name"] == "S")
    assert sheet2["hyperlinks"]["A1"]["target"] == "https://example.com/changed"
    assert sheet2["hyperlinks"]["A1"]["tooltip"] == "tip1"
    assert sheet2["hyperlinks"]["A2"]["location"] == "S!B2"

    M.excel_remove_hyperlink(key2, "S", "A1")
    remaining = M._get_session(key2)
    sheet3 = next(s for s in remaining["sheets"] if s["name"] == "S")
    assert "A1" not in sheet3["hyperlinks"]
    assert "A2" in sheet3["hyperlinks"]  # removing one leaves the other intact


# ---------------------------------------------------------------------------
# Comments
# ---------------------------------------------------------------------------

def test_comment_set_remove_legacy_text_author_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "comments.xlsx"

    M.excel_set_comment(key, "S", "A1", "note text", author="alice")
    M.excel_set_comment(key, "S", "B1", "another note", author="bob")
    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    key2 = _load_key(M.excel_load(str(out)))
    sess2 = M._get_session(key2)
    sheet2 = next(s for s in sess2["sheets"] if s["name"] == "S")
    assert sheet2["comments"]["A1"] == {"text": "note text", "author": "alice"}
    assert sheet2["comments"]["B1"] == {"text": "another note", "author": "bob"}

    M.excel_remove_comment(key2, "S", "A1")
    out2 = tmp_path / "comments_removed.xlsx"
    M.excel_save(key2, str(out2))
    key3 = _load_key(M.excel_load(str(out2)))
    sess3 = M._get_session(key3)
    sheet3 = next(s for s in sess3["sheets"] if s["name"] == "S")
    assert "A1" not in sheet3["comments"]
    assert sheet3["comments"]["B1"]["text"] == "another note"  # sibling untouched


def test_comment_rejects_threaded_type_without_silently_converting():
    key = _new_session()
    with pytest.raises(ValueError):
        M.excel_set_comment(key, "S", "A1", "threaded text", comment_type="threaded")


def test_comment_editing_unrelated_cell_preserves_other_comments_custom_vml(tmp_path):
    src = tmp_path / "vml_src.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "x"
    ws["A1"].comment = Comment("original note", "alice")
    wb.save(src)

    vml_name = "xl/drawings/commentsDrawing1.vml"
    with zipfile.ZipFile(src) as z:
        vml = z.read(vml_name).decode("utf-8")
    custom_size = "width:250pt;height:180pt"
    patched = vml.replace("width:144px;height:79px", custom_size)
    assert patched != vml, "openpyxl's default comment box size string changed; update this fixture"
    tmp_zip = src.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            raw = zin.read(item.filename)
            if item.filename == vml_name:
                raw = patched.encode("utf-8")
            zout.writestr(item, raw)
    tmp_zip.replace(src)

    key = _load_key(M.excel_load(str(src)))
    M.excel_set_comment(key, "S", "B1", "new note", author="bob")  # unrelated cell
    out = tmp_path / "vml_out.xlsx"
    M.excel_save(key, str(out))

    with zipfile.ZipFile(out) as z:
        out_vml = z.read(vml_name).decode("utf-8")
    assert custom_size in out_vml


# ---------------------------------------------------------------------------
# excel_set_ignored_errors
# ---------------------------------------------------------------------------

def test_ignored_errors_patch_mode_preserves_other_rules():
    key = _new_session()
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "1"}}])

    M.excel_set_ignored_errors(key, "S", rules=[{"sqref": "A1:A3", "numberStoredAsText": True}])
    M.excel_set_ignored_errors(key, "S", rules=[{"sqref": "B1:B3", "formula": True}])
    result = json.loads(M.excel_set_ignored_errors(key, "S", rules=[{"sqref": "C1:C3", "evalError": True}]))
    by_sqref = {r["sqref"]: r for r in result["after"]}
    assert by_sqref["A1:A3"]["numberStoredAsText"] is True
    assert by_sqref["B1:B3"]["formula"] is True
    assert by_sqref["C1:C3"]["evalError"] is True

    # mode="replace" wipes prior rules, per its contract.
    replaced = json.loads(M.excel_set_ignored_errors(key, "S", rules=[{"sqref": "D1:D3"}], mode="replace"))
    assert len(replaced["after"]) == 1


def test_ignored_errors_survive_save_reload(tmp_path):
    key = _new_session()
    out = tmp_path / "ignored_errors.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "1"}}])
    M.excel_set_ignored_errors(key, "S", rules=[{"sqref": "A1:A3", "numberStoredAsText": True}])
    M.excel_save(key, str(out))
    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "ignoredErrors" in xml


# ---------------------------------------------------------------------------
# Named styles
# ---------------------------------------------------------------------------

def test_named_style_add_update_delete_in_memory_semantics():
    key = _new_session()
    M.excel_add_named_style(key, "Accent", {"bold": True}, metadata={"builtinId": 5})
    with pytest.raises(ValueError):
        M.excel_add_named_style(key, "Accent", {"bold": False})  # duplicate name rejected

    M.excel_update_named_style(key, "Accent", {"style": {"bold": True, "italic": True}})
    session = M._get_session(key)
    accent = next(s for s in session["named_styles"] if s["name"] == "Accent")
    assert accent["style"]["italic"] is True
    assert accent["builtinId"] == 5  # untouched by the partial update

    with pytest.raises(ValueError):
        M.excel_delete_named_style(key, "Normal")  # built-in Normal is protected

    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "x"}}])
    M.excel_set_style(key, "S", 0, 0, style={"named_style": "Accent"})
    with pytest.raises(ValueError):
        M.excel_delete_named_style(key, "Accent")  # still referenced by A1

    M.excel_set_style(key, "S", 0, 0, style={"named_style": "Normal"})
    M.excel_delete_named_style(key, "Accent")  # now unreferenced -- succeeds
    session2 = M._get_session(key)
    assert all(s["name"] != "Accent" for s in session2["named_styles"])


def test_named_style_assignment_survives_save_reload(tmp_path):
    key = _new_session()
    out = tmp_path / "named_style.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "x"}}])
    M.excel_add_named_style(key, "Accent", {
        "bold": True,
        "fill": {
            "pattern_type": "solid",
            "foreground": {"type": "rgb", "rgb": "FF99CC00"},
        },
    })
    M.excel_set_style(key, "S", 0, 0, style={"named_style": "Accent"})
    M.excel_save(key, str(out))

    key2 = _load_key(M.excel_load(str(out)))
    after = json.loads(M.excel_get_cell(key2, "S", 0, 0, include_semantics=True))
    assert after["semantics"]["named_style"] == "Accent"
    assert after["semantics"]["fill"]["foreground"]["rgb"] == "FF99CC00"
