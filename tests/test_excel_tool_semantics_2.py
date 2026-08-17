"""
Verification tests for PRESERVATION_FIX_CHECKLIST.md bullets covering
excel_set_formula and the defined-name tools (excel_add_defined_name /
excel_update_defined_name / excel_delete_defined_name). These exercise the
PUBLIC MCP tool functions in servers/excel/main.py exactly as an agent would
call them -- never the internal helpers directly.

Covers:
  1. excel_set_formula: cached-value state (missing / explicit empty /
     numeric / string / boolean / error) round-trips distinctly through
     save+reload.
  2. excel_set_formula: normal / shared / array / data-table formula types.
  3. excel_set_formula: re-setting a formula clears the old cache by default,
     and only keeps/sets a new cache when the caller explicitly asks.
  4. Defined names: workbook-scope vs worksheet-scope (incl. same name at two
     different worksheet scopes), full metadata fields, built-in name
     classification (_xlnm.Print_Area / _xlnm.Print_Titles /
     _xlnm._FilterDatabase), and partial-update-preserves-metadata.

Bugs found and fixed while writing these tests (see servers/excel/core.py and
servers/excel/main.py for the actual diffs):
  - A formula's string/boolean/error cached value was written to OOXML with
    no t="str"/"b"/"e" attribute (always defaulting to numeric), producing a
    file openpyxl itself could not reload with data_only=True.
  - A reloaded formula cell's cached_value_state always reported "missing"
    (main.py's _formula_model only checked the "cached_value_state" key;
    core.py's raw XML read model uses "cache_state").
  - excel_set_formula's formula_type ("shared"/"array"/"dataTable") had no
    effect on the actual OOXML t="..." attribute unless the caller also
    manually duplicated it into formula_attributes.
  - Defined names: worksheet-scoped names were silently dropped after a
    save+reload (core.py only ever scanned wb.defined_names, but openpyxl's
    own reader routes worksheet-scoped names into each ws.defined_names);
    two different sheets both naming a local range the same name collided in
    a single shared dict; DefinedName metadata (hidden/comment/description/
    help/statusBar/function/vbProcedure/xlm/functionGroupId/shortcutKey/
    publishToServer/workbookParameter) was accepted by the tool but silently
    dropped on save; and _xlnm.Print_Area/_xlnm.Print_Titles were written as
    raw duplicate/malformed defined names that either corrupted the workbook
    (two competing Print_Titles entries) or vanished after reload instead of
    being recognized as built-ins.
"""
import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _load_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _new_session(sheet_names=("S",)):
    created = json.loads(M.excel_create_workbook(sheet_names=list(sheet_names)))
    return created["session_key"]


def _cell_xml(path: Path, sheet_part: str, coord: str) -> str | None:
    import re
    with zipfile.ZipFile(path) as z:
        xml = z.read(sheet_part).decode("utf-8")
    m = re.search(r'<c r="' + re.escape(coord) + r'"[^>]*(?:/>|>.*?</c>)', xml, re.DOTALL)
    return m.group(0) if m else None


# ---------------------------------------------------------------------------
# 1. excel_set_formula: cached-value state distinguishes missing / explicit
#    empty / numeric / string / boolean / error, and each round-trips.
# ---------------------------------------------------------------------------

def test_set_formula_cached_value_states_distinguished_and_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "cache_states.xlsx"

    M.excel_set_formula(key, "S", "A1", "=1+1")  # no cache at all
    M.excel_set_formula(key, "S", "A2", "=1+1", cached_value=None, cached_value_present=True)  # explicit empty
    M.excel_set_formula(key, "S", "A3", "=1+41", cached_value=42, cached_value_present=True)  # numeric (int)
    M.excel_set_formula(key, "S", "A4", "=1+41.5", cached_value=42.5, cached_value_present=True)  # numeric (float)
    M.excel_set_formula(key, "S", "A5", "=A1", cached_value="hello", cached_value_present=True)  # string
    M.excel_set_formula(key, "S", "A6", "=1=1", cached_value=True, cached_value_present=True)  # boolean True
    M.excel_set_formula(key, "S", "A7", "=1=2", cached_value=False, cached_value_present=True)  # boolean False
    M.excel_set_formula(key, "S", "A8", "=1/0", cached_value="#DIV/0!", cached_value_present=True)  # error

    # Session-level view (before any save) already distinguishes all six cases.
    session_state = {
        coord: json.loads(M.excel_get_cell(key, "S", i, 0, include_formula_cache=True))["formula"]
        for i, coord in enumerate(["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8"])
    }
    assert session_state["A1"]["cached_value_state"] == "missing"
    assert "cached_value" not in session_state["A1"] or session_state["A1"].get("cached_value") is None
    assert session_state["A2"]["cached_value_state"] == "empty"
    assert session_state["A2"]["cached_value"] is None
    assert session_state["A3"]["cached_value_state"] == "value"
    assert session_state["A3"]["cached_value"] == 42
    assert session_state["A6"]["cached_value"] is True
    assert session_state["A7"]["cached_value"] is False
    assert session_state["A8"]["cached_value"] == "#DIV/0!"

    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    # The file must be loadable by plain openpyxl in data_only mode -- this
    # is exactly the corruption case the string/boolean/error cache bug
    # produced (ValueError: could not convert string to float).
    wb_check = openpyxl.load_workbook(str(out), data_only=True)
    ws_check = wb_check["S"]
    assert ws_check["A5"].value == "hello"
    assert ws_check["A6"].value is True
    assert ws_check["A7"].value is False
    assert ws_check["A8"].value == "#DIV/0!"
    wb_check.close()

    key2 = _load_key(M.excel_load(str(out)))
    reloaded = {
        coord: json.loads(M.excel_get_cell(key2, "S", i, 0, include_formula_cache=True))["formula"]
        for i, coord in enumerate(["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8"])
    }
    assert reloaded["A1"]["cached_value_state"] == "missing"
    assert reloaded["A2"]["cached_value_state"] == "empty"
    assert reloaded["A2"]["cached_value"] is None
    assert reloaded["A3"]["cached_value_state"] == "value"
    assert float(reloaded["A3"]["cached_value"]) == 42
    assert float(reloaded["A4"]["cached_value"]) == 42.5
    assert reloaded["A5"]["cached_value_state"] == "value"
    assert reloaded["A5"]["cached_value"] == "hello"
    assert reloaded["A8"]["cached_value"] == "#DIV/0!"

    # And the raw XML actually carries the correct t="..." discriminator.
    assert 't="str"' in _cell_xml(out, "xl/worksheets/sheet1.xml", "A5")
    assert 't="b"' in _cell_xml(out, "xl/worksheets/sheet1.xml", "A6")
    assert 't="e"' in _cell_xml(out, "xl/worksheets/sheet1.xml", "A8")
    numeric_xml = _cell_xml(out, "xl/worksheets/sheet1.xml", "A3")
    assert 't="str"' not in numeric_xml and 't="b"' not in numeric_xml and 't="e"' not in numeric_xml


# ---------------------------------------------------------------------------
# 2. excel_set_formula: normal / shared / array / data-table formula types.
# ---------------------------------------------------------------------------

def test_set_formula_supports_normal_shared_array_datatable_types(tmp_path):
    key = _new_session()
    out = tmp_path / "formula_types.xlsx"

    M.excel_set_formula(key, "S", "A1", "=1+1")  # normal (default)
    M.excel_set_formula(key, "S", "A2", "=SUM(B1:B3)", formula_type="array",
                         formula_attributes={"ref": "A2:A2"}, cached_value=0, cached_value_present=True)
    M.excel_set_formula(key, "S", "B1", "=A1", formula_type="shared", formula_attributes={"ref": "B1:B2", "si": "0"})
    M.excel_set_formula(key, "S", "B2", "=A1", formula_type="shared", formula_attributes={"si": "0"})
    # dataTable formula (what-if analysis TABLE()) -- OOXML t="dataTable".
    M.excel_set_formula(key, "S", "C1", "=A1", formula_type="dataTable",
                         formula_attributes={"ref": "C1:C1", "dt2D": "0", "dtr": "0", "r1": "A1"})

    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    a1 = _cell_xml(out, "xl/worksheets/sheet1.xml", "A1")
    a2 = _cell_xml(out, "xl/worksheets/sheet1.xml", "A2")
    b1 = _cell_xml(out, "xl/worksheets/sheet1.xml", "B1")
    b2 = _cell_xml(out, "xl/worksheets/sheet1.xml", "B2")
    c1 = _cell_xml(out, "xl/worksheets/sheet1.xml", "C1")

    assert 't="' not in a1.split("<f")[1].split(">")[0]  # normal: no t= on <f>
    assert 't="array"' in a2
    assert 't="shared"' in b1
    assert 't="shared"' in b2
    assert 't="dataTable"' in c1

    wb2 = openpyxl.load_workbook(str(out))
    ws2 = wb2["S"]
    from openpyxl.worksheet.formula import ArrayFormula
    assert isinstance(ws2["A2"].value, ArrayFormula)
    wb2.close()


# ---------------------------------------------------------------------------
# 3. excel_set_formula: replacing a formula clears the old cache by default;
#    only keeps/sets a new cache when the caller explicitly asks.
# ---------------------------------------------------------------------------

def test_set_formula_clears_cache_by_default_and_preserve_policy_keeps_it():
    key = _new_session()

    M.excel_set_formula(key, "S", "A1", "=1+41", cached_value=42, cached_value_present=True)
    before = json.loads(M.excel_get_cell(key, "S", 0, 0, include_formula_cache=True))
    assert before["formula"]["cached_value_state"] == "value"
    assert before["formula"]["cached_value"] == 42

    # Default behavior (cache_policy="clear"): re-setting the formula text
    # without an explicit new cached_value wipes the stale cache.
    M.excel_set_formula(key, "S", "A1", "=2+41")
    after_default = json.loads(M.excel_get_cell(key, "S", 0, 0, include_formula_cache=True))
    assert after_default["formula"]["text"].lstrip("=") == "2+41"
    assert after_default["formula"]["cached_value_state"] == "missing"
    assert "cached_value" not in after_default["formula"] or after_default["formula"]["cached_value"] is None

    # Explicit cache_policy="preserve": keeps the old cached value when the
    # caller does not pass a new one.
    M.excel_set_formula(key, "S", "A2", "=1+41", cached_value=42, cached_value_present=True)
    M.excel_set_formula(key, "S", "A2", "=2+41", cache_policy="preserve")
    preserved = json.loads(M.excel_get_cell(key, "S", 1, 0, include_formula_cache=True))
    assert preserved["formula"]["text"].lstrip("=") == "2+41"
    assert preserved["formula"]["cached_value_state"] == "value"
    assert preserved["formula"]["cached_value"] == 42

    # Explicit new cached_value always wins, regardless of cache_policy.
    M.excel_set_formula(key, "S", "A3", "=1+41", cached_value=42, cached_value_present=True)
    M.excel_set_formula(key, "S", "A3", "=3+41", cached_value=99, cached_value_present=True)
    replaced = json.loads(M.excel_get_cell(key, "S", 2, 0, include_formula_cache=True))
    assert replaced["formula"]["cached_value"] == 99


# ---------------------------------------------------------------------------
# 4. Defined names: scope, metadata, built-in classification, partial update.
# ---------------------------------------------------------------------------

def test_defined_names_workbook_vs_worksheet_scope_and_same_name_two_scopes(tmp_path):
    key = _new_session(["S1", "S2"])
    out = tmp_path / "defined_names.xlsx"

    M.excel_add_defined_name(key, "GlobalName", "S1!$A$1:$A$2")
    M.excel_add_defined_name(key, "Data", "S1!$A$1", sheet_name="S1")
    M.excel_add_defined_name(key, "Data", "S2!$A$1", sheet_name="S2")
    # local_sheet_id is an equivalent way to address worksheet scope.
    M.excel_add_defined_name(key, "ByLocalId", "S2!$B$1", local_sheet_id=1)

    M.excel_save(key, str(out))
    key2 = _load_key(M.excel_load(str(out)))
    names = {(n["name"], n["sheet_id"]): n for n in json.loads(M.excel_list_defined_names(key2))["defined_names"]}

    assert names[("GlobalName", None)]["value"] == "S1!$A$1:$A$2"
    assert names[("Data", 0)]["value"] == "S1!$A$1"
    assert names[("Data", 1)]["value"] == "S2!$A$1"
    assert names[("ByLocalId", 1)]["value"] == "S2!$B$1"

    wb2 = openpyxl.load_workbook(str(out))
    assert wb2.defined_names["GlobalName"].attr_text == "S1!$A$1:$A$2"
    assert wb2["S1"].defined_names["Data"].attr_text == "S1!$A$1"
    assert wb2["S2"].defined_names["Data"].attr_text == "S2!$A$1"
    wb2.close()


def test_defined_names_full_metadata_fields_roundtrip(tmp_path):
    key = _new_session()
    out = tmp_path / "defined_name_meta.xlsx"

    metadata = {
        "hidden": True,
        "function": True,
        "vbProcedure": False,
        "xlm": False,
        "functionGroupId": 3,
        "shortcutKey": "A",
        "publishToServer": True,
        "workbookParameter": False,
        "comment": "a comment",
        "description": "a description",
        "help": "help text",
        "statusBar": "status text",
    }
    M.excel_add_defined_name(key, "Documented", "S!$A$1", sheet_name="S", metadata=metadata)
    M.excel_save(key, str(out))

    key2 = _load_key(M.excel_load(str(out)))
    names = json.loads(M.excel_list_defined_names(key2))["defined_names"]
    item = next(n for n in names if n["name"] == "Documented")
    for field, expected in metadata.items():
        assert item.get(field) == expected, field


def test_defined_names_builtin_print_area_titles_filterdatabase_classified(tmp_path):
    key = _new_session()
    out = tmp_path / "builtins.xlsx"
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "x"}}])

    M.excel_set_print_area(key, "S", "A1:B2")
    M.excel_set_print_titles(key, "S", repeated_rows="1:1")
    M.excel_set_auto_filter(key, "S", ref="A1:B2")

    # The generic defined-name tools must recognize these as built-ins and
    # refuse to treat them as arbitrary user names (silent data loss / a
    # corrupted duplicate definedName otherwise).
    with pytest.raises(ValueError):
        M.excel_add_defined_name(key, "_xlnm.Print_Area", "A1:C3", sheet_name="S")
    with pytest.raises(ValueError):
        M.excel_update_defined_name(key, "_xlnm.Print_Titles", {"value": "2:2"}, sheet_name="S")
    with pytest.raises(ValueError):
        M.excel_delete_defined_name(key, "_xlnm._FilterDatabase", sheet_name="S")

    M.excel_save(key, str(out))
    report = json.loads(M.excel_validate_workbook(str(out)))
    assert report["valid"] is True

    key2 = _load_key(M.excel_load(str(out)))
    names = {n["name"]: n for n in json.loads(M.excel_list_defined_names(key2))["defined_names"]}
    assert names["_xlnm.Print_Area"]["builtin"] is True
    assert names["_xlnm.Print_Area"]["value"] == "'S'!$A$1:$B$2"
    assert names["_xlnm.Print_Titles"]["builtin"] is True
    assert names["_xlnm._FilterDatabase"]["builtin"] is True

    # No duplicate/competing definedName entries for the same reserved name.
    with zipfile.ZipFile(out) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    assert wb_xml.count("_xlnm.Print_Titles") == 1
    assert wb_xml.count("_xlnm.Print_Area") == 1

    assert json.loads(M.excel_get_sheet_semantics(key2, "S"))["print_area"] == "'S'!$A$1:$B$2"


def test_update_defined_name_partial_update_preserves_other_metadata():
    key = _new_session()
    M.excel_add_defined_name(key, "Documented", "S!$A$1", sheet_name="S", metadata={
        "hidden": True, "comment": "original comment", "description": "original description",
    })
    M.excel_update_defined_name(key, "Documented", {"value": "S!$A$1:$A$5"}, sheet_name="S")

    names = json.loads(M.excel_list_defined_names(key))["defined_names"]
    item = next(n for n in names if n["name"] == "Documented")
    assert item["value"] == "S!$A$1:$A$5"
    # Fields not mentioned in the update call must survive untouched.
    assert item["hidden"] is True
    assert item["comment"] == "original comment"
    assert item["description"] == "original description"
