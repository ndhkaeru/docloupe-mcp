"""
Tests for formula reference shifting on structural edits and for the
quote-prefix input contract.

- Insert/delete rows rewrites formulas on the edited sheet AND on other
  sheets referencing it (sheet-qualified refs), plus defined names.
- References whose entire area is deleted become #REF!.
- String literals inside formulas and refs to OTHER sheets are untouched.
- Literal text starting with "=" survives round-trip as text (dt/qp markers).
- Input contract: leading apostrophe forces text + quotePrefix.
- All six row/column insert, delete, and copy tools shift formulas, defined
  names, filters, tables, validations, conditional formatting, hyperlinks,
  selections, page breaks, and DrawingML anchors.
"""
import json
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import FilterColumn, Filters, SortCondition, SortState
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableFormula
from openpyxl.worksheet.views import Selection
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
from core import reconstruct_excel, serialize_excel  # noqa: E402


def _load(path: Path) -> str:
    M.excel_load(str(path))
    return str(Path(path).resolve())


def _make_formula_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(2, 6):                      # B2:B5 = 1..4
        ws.cell(row=r, column=2, value=r - 1)
    ws["C1"] = "=SUM(B2:B5)"
    ws["C2"] = "=B3*2"
    ws["C3"] = '=IF(B2>0,"B2 is text here",B5)'   # literal must not shift

    other = wb.create_sheet("Other")
    other["A1"] = "=Data!B3"
    other["A2"] = "=SUM('Data'!$B$2:$B$5)"
    other["A3"] = "=Other2!B3"                 # different sheet — untouched
    wb.create_sheet("Other2")["B3"] = 99

    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["TotalRange"] = DefinedName(
        "TotalRange", attr_text="Data!$B$2:$B$5")
    wb.save(path)


def test_insert_rows_shifts_formulas_everywhere(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    rows = M.excel_clone_rows(key, "Data", 0)
    import json
    M.excel_insert_rows(key, "Data", [
        {"after_index": -1, "rows_json": json.loads(rows)},
        {"after_index": -1, "rows_json": json.loads(rows)},
    ])
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws, other = wb["Data"], wb["Other"]
    assert ws["C3"].value == "=SUM(B4:B7)"
    assert ws["C4"].value == "=B5*2"
    assert ws["C5"].value == '=IF(B4>0,"B2 is text here",B7)'  # literal intact
    assert other["A1"].value == "=Data!B5"
    assert other["A2"].value == "=SUM('Data'!$B$4:$B$7)"
    assert other["A3"].value == "=Other2!B3"                   # untouched
    assert wb.defined_names["TotalRange"].attr_text == "Data!$B$4:$B$7"


def test_delete_rows_shrinks_and_ref_errors(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_delete_rows(key, "Data", row_indices=[2])   # row 3 (B3)
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    ws, other = wb["Data"], wb["Other"]
    assert ws["C1"].value == "=SUM(B2:B4)"              # range shrank
    assert ws["C2"].value == "=#REF!*2"                 # single ref deleted
    assert other["A1"].value == "=Data!#REF!"
    assert wb.defined_names["TotalRange"].attr_text == "Data!$B$2:$B$4"


def test_insert_column_shifts_formulas(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_insert_column(key, "Data", after_col_index=-1)   # prepend col A
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    assert wb["Data"]["D1"].value == "=SUM(C2:C5)"
    assert wb["Other"]["A1"].value == "=Data!C3"


def test_rename_sheet_rewrites_cell_formulas(tmp_path):
    src = tmp_path / "f.xlsx"
    _make_formula_workbook(src)
    key = _load(src)
    M.excel_rename_sheet(key, "Data", "Số liệu 2026")
    M.excel_save(key)
    M.excel_close(key)

    wb = openpyxl.load_workbook(src)
    other = wb["Other"]
    assert other["A1"].value == "='Số liệu 2026'!B3"
    assert other["A2"].value == "=SUM('Số liệu 2026'!$B$2:$B$5)"
    assert other["A3"].value == "=Other2!B3"


def test_quote_prefix_input_contract(tmp_path):
    src = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "placeholder"
    wb.save(src)

    key = _load(src)
    M.excel_edit_cells(key, "Sheet", [{"row_index": 0, "edits": {
        "0": "'=not a formula",   # apostrophe → literal text
        "1": "=SUM(1,2)",         # real formula
        "2": "+tăng 5%",          # plus text — stays text automatically
        "3": "'0123",             # apostrophe-protected leading zeros
    }}])
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    ws = wb2.active
    assert ws["A1"].value == "=not a formula"
    assert ws["A1"].data_type == "s"
    assert bool(ws["A1"]._style and ws["A1"]._style.quotePrefix)
    assert ws["B1"].value == "=SUM(1,2)" and ws["B1"].data_type == "f"
    assert ws["C1"].value == "+tăng 5%" and ws["C1"].data_type == "s"
    assert ws["D1"].value == "0123" and ws["D1"].data_type == "s"


def test_text_looking_like_formula_roundtrips_as_text(tmp_path):
    src = tmp_path / "qp.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    c = ws.cell(row=1, column=1, value="=this is plain text")
    c.data_type = "s"
    from openpyxl.styles.cell_style import StyleArray
    c._style = StyleArray()
    c._style.quotePrefix = 1
    ws["A2"] = "=SUM(1,2)"
    wb.save(src)

    reconstruct_excel(serialize_excel(str(src)), str(out))

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.active
    assert ws2["A1"].value == "=this is plain text"
    assert ws2["A1"].data_type == "s", "text must not become a broken formula"
    assert bool(ws2["A1"]._style and ws2["A1"]._style.quotePrefix)
    assert ws2["A2"].data_type == "f"


def test_formula_text_not_shifted_by_structural_edit(tmp_path):
    """A dt='s' literal that LOOKS like a formula must not be rewritten."""
    src = tmp_path / "qp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    c = ws.cell(row=5, column=1, value="=SUM(B2:B5)")    # literal text!
    c.data_type = "s"
    from openpyxl.styles.cell_style import StyleArray
    c._style = StyleArray()
    c._style.quotePrefix = 1
    wb.save(src)

    key = _load(src)
    M.excel_copy_row(key, "Data", 0, -1)   # insert a row above
    M.excel_save(key)
    M.excel_close(key)

    wb2 = openpyxl.load_workbook(src)
    assert wb2["Data"]["A6"].value == "=SUM(B2:B5)"      # content untouched
    assert wb2["Data"]["A6"].data_type == "s"


def test_structural_edits_shift_formula_metadata_and_nested_table_ranges(tmp_path):
    src = tmp_path / "structured.xlsx"
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    other = wb.create_sheet("Other")
    for row in range(1, 7):
        for column in range(1, 6):
            data.cell(row=row, column=column, value=row * 10 + column)
    for row in range(1, 4):
        other.cell(row=row, column=1, value=row)

    local_validation = DataValidation(
        type="whole", operator="between", formula1="=$B$2", formula2="=$B$5",
    )
    local_validation.add("D2:D5")
    data.add_data_validation(local_validation)
    cross_sheet_validation = DataValidation(type="list", formula1="=Data!$B$2:$B$5")
    cross_sheet_validation.add("B1")
    other.add_data_validation(cross_sheet_validation)

    data.conditional_formatting.add("E2:E5", FormulaRule(formula=["$B2>0"]))
    other.conditional_formatting.add("A1:A2", FormulaRule(formula=["Data!$B$2>0"]))
    data["A1"].hyperlink = Hyperlink(ref="A1", location="Data!B2")
    other["A1"].hyperlink = Hyperlink(ref="A1", location="Data!B3")
    data.print_area = ["A1:B3", "D5:E6"]
    data.print_title_rows = "$1:$2"
    data.print_title_cols = "$A:$B"

    table = Table(displayName="Metrics", ref="A1:C6")
    table._initialise_columns()
    for index, column in enumerate(table.tableColumns, 1):
        column.name = f"C{index}"
    table.tableColumns[2].calculatedColumnFormula = TableFormula(attr_text="B2*2")
    table.tableColumns[2].totalsRowFormula = TableFormula(
        array=True, attr_text="SUM(B2:B5)",
    )
    table.autoFilter.ref = "A1:C6"
    table.autoFilter.filterColumn = [
        FilterColumn(colId=1, filters=Filters(filter=["21"])),
    ]
    table.autoFilter.sortState = SortState(
        ref="A2:C6",
        sortCondition=[SortCondition(ref="B2:B6", descending=True)],
    )
    table.sortState = SortState(
        ref="A2:C6",
        sortCondition=[SortCondition(ref="C2:C6")],
    )
    data.add_table(table)
    wb.save(src)

    key = _load(src)
    try:
        M.excel_copy_row(key, "Data", 0, -1)
        M.excel_copy_row(key, "Data", 0, -1)
        M.excel_insert_column(key, "Data", after_col_index=-1)
        M.excel_save(key)
    finally:
        M.excel_close(key)

    saved = openpyxl.load_workbook(src)
    try:
        data = saved["Data"]
        other = saved["Other"]
        assert [
            (str(validation.sqref), validation.formula1, validation.formula2)
            for validation in data.data_validations.dataValidation
        ] == [("E4:E7", "=$C$4", "=$C$7")]
        assert [
            (str(validation.sqref), validation.formula1)
            for validation in other.data_validations.dataValidation
        ] == [("B1", "=Data!$C$4:$C$7")]

        local_cf = list(data.conditional_formatting)
        cross_sheet_cf = list(other.conditional_formatting)
        assert str(local_cf[0].sqref) == "F4:F7"
        assert local_cf[0].rules[0].formula == ["$C4>0"]
        assert str(cross_sheet_cf[0].sqref) == "A1:A2"
        assert cross_sheet_cf[0].rules[0].formula == ["Data!$C$4>0"]

        assert data["B3"].hyperlink.location == "Data!C4"
        assert other["A1"].hyperlink.location == "Data!C5"
        print_area = str(data.print_area)
        assert "$B$3:$C$5" in print_area
        assert "$E$7:$F$8" in print_area
        assert data.print_title_rows == "$3:$4"
        assert data.print_title_cols == "$B:$C"

        saved_table = data.tables["Metrics"]
        assert saved_table.ref == "B3:D8"
        assert saved_table.autoFilter.ref == "B3:D8"
        assert saved_table.autoFilter.sortState.ref == "B4:D8"
        assert [condition.ref for condition in saved_table.autoFilter.sortState.sortCondition] == ["C4:C8"]
        assert [column.colId for column in saved_table.autoFilter.filterColumn] == [1]
        assert saved_table.sortState.ref == "B4:D8"
        assert [condition.ref for condition in saved_table.sortState.sortCondition] == ["D4:D8"]
        calculated = saved_table.tableColumns[2].calculatedColumnFormula
        totals = saved_table.tableColumns[2].totalsRowFormula
        assert calculated.attr_text == "C4*2"
        assert totals.attr_text == "SUM(C4:C7)"
        assert totals.array is True
    finally:
        saved.close()


_STRUCTURAL_SHIFT_CASES = [
    ("insert_rows", "row", 1),
    ("delete_rows", "row", -1),
    ("copy_row", "row", 1),
    ("insert_column", "column", 1),
    ("delete_column", "column", -1),
    ("copy_column", "column", 1),
]


def _make_structural_metadata_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"

    values = [
        ("Category", "Amount", "Double"),
        ("A", 2, 4),
        ("B", 5, 10),
        ("C", 3, 6),
    ]
    for row_offset, row_values in enumerate(values, start=3):
        for column_offset, value in enumerate(row_values, start=3):
            sheet.cell(row=row_offset, column=column_offset, value=value)

    table = Table(displayName="TrackedTable", ref="C3:E6")
    table._initialise_columns()
    table.tableColumns[0].name = "Category"
    table.tableColumns[1].name = "Amount"
    table.tableColumns[2].name = "Double"
    table.autoFilter.ref = "C3:E6"
    table.autoFilter.sortState = SortState(
        ref="C4:E6",
        sortCondition=[SortCondition(ref="D4:D6")],
    )
    table.sortState = SortState(
        ref="C4:E6",
        sortCondition=[SortCondition(ref="E4:E6")],
    )
    sheet.add_table(table)

    sheet.auto_filter.ref = "F3:H6"
    sheet.auto_filter.filterColumn = [
        FilterColumn(colId=0, filters=Filters(filter=["A"])),
    ]
    sheet.auto_filter.sortState = SortState(
        ref="F4:H6",
        sortCondition=[SortCondition(ref="G4:G6")],
    )

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="=$C$3",
        formula2="=$C$5",
    )
    validation.add("I3:I5")
    sheet.add_data_validation(validation)
    sheet.conditional_formatting.add("J3:J5", FormulaRule(formula=["$C3>0"]))
    sheet["K4"] = "Jump"
    sheet["K4"].hyperlink = Hyperlink(ref="K4", location="Data!C3")
    sheet["L8"] = "=SUM(C3:C5)"
    sheet["M5"].comment = Comment("Structural note", "Audit Author")
    sheet.merge_cells("C10:D11")
    workbook.defined_names["TrackedRange"] = DefinedName(
        "TrackedRange", attr_text="Data!$C$3:$E$5"
    )

    sheet.freeze_panes = "C3"
    sheet.sheet_view.topLeftCell = "C3"
    sheet.sheet_view.selection = [
        Selection(pane="bottomRight", activeCell="D4", sqref="D4:E5")
    ]
    sheet.row_breaks.append(Break(id=5, min=0, max=16383, man=True))
    sheet.col_breaks.append(Break(id=4, min=0, max=1048575, man=True))

    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=4, min_row=3, max_row=6),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=3, min_row=4, max_row=6))
    sheet.add_chart(chart, "N4")
    workbook.save(path)
    workbook.close()


def _shifted_position(row: int, column: int, axis: str, delta: int) -> tuple[int, int]:
    if axis == "row":
        row += delta
    else:
        column += delta
    return row, column


def _a1(row: int, column: int, absolute: bool = False) -> str:
    column_letter = openpyxl.utils.get_column_letter(column)
    if absolute:
        return f"${column_letter}${row}"
    return f"{column_letter}{row}"


def _shifted_a1(row: int, column: int, axis: str, delta: int, absolute: bool = False) -> str:
    shifted_row, shifted_column = _shifted_position(row, column, axis, delta)
    return _a1(shifted_row, shifted_column, absolute=absolute)


def _shifted_range(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
    axis: str,
    delta: int,
    absolute: bool = False,
) -> str:
    return ":".join([
        _shifted_a1(start_row, start_column, axis, delta, absolute=absolute),
        _shifted_a1(end_row, end_column, axis, delta, absolute=absolute),
    ])


def _apply_structural_shift_case(session_key: str, operation: str) -> None:
    if operation == "insert_rows":
        rows = json.loads(M.excel_clone_rows(session_key, "Data", 0))
        M.excel_insert_rows(
            session_key,
            "Data",
            [{"after_index": 0, "rows_json": rows}],
        )
    elif operation == "delete_rows":
        M.excel_delete_rows(session_key, "Data", row_indices=[0])
    elif operation == "copy_row":
        M.excel_copy_row(session_key, "Data", row_index=0, after_index=0)
    elif operation == "insert_column":
        M.excel_insert_column(session_key, "Data", after_col_index=0)
    elif operation == "delete_column":
        M.excel_delete_column(session_key, "Data", col_index=0)
    elif operation == "copy_column":
        M.excel_copy_column(session_key, "Data", col_index=0, after_col_index=0)
    else:
        raise AssertionError(f"Unsupported structural shift case: {operation}")


@pytest.mark.parametrize("operation,axis,delta", _STRUCTURAL_SHIFT_CASES)
def test_all_structural_tools_shift_reference_metadata_selections_breaks_and_drawings(
    tmp_path,
    operation,
    axis,
    delta,
):
    path = tmp_path / f"{operation}.xlsx"
    _make_structural_metadata_workbook(path)
    session_key = _load(path)
    try:
        _apply_structural_shift_case(session_key, operation)
        M.excel_save(session_key)
    finally:
        M.excel_close(session_key)

    saved = openpyxl.load_workbook(path)
    try:
        sheet = saved["Data"]
        formula_cell = _shifted_a1(8, 12, axis, delta)
        formula_range = _shifted_range(3, 3, 5, 3, axis, delta)
        assert sheet[formula_cell].value == f"=SUM({formula_range})"

        defined_name = saved.defined_names["TrackedRange"].attr_text.replace("'Data'", "Data")
        assert defined_name == (
            f"Data!{_shifted_range(3, 3, 5, 5, axis, delta, absolute=True)}"
        )

        assert sheet.auto_filter.ref == _shifted_range(3, 6, 6, 8, axis, delta)
        assert sheet.auto_filter.sortState.ref == _shifted_range(4, 6, 6, 8, axis, delta)
        assert [
            condition.ref for condition in sheet.auto_filter.sortState.sortCondition
        ] == [_shifted_range(4, 7, 6, 7, axis, delta)]

        table = sheet.tables["TrackedTable"]
        assert table.ref == _shifted_range(3, 3, 6, 5, axis, delta)
        assert table.autoFilter.ref == _shifted_range(3, 3, 6, 5, axis, delta)
        assert table.autoFilter.sortState.ref == _shifted_range(4, 3, 6, 5, axis, delta)
        assert [
            condition.ref for condition in table.autoFilter.sortState.sortCondition
        ] == [_shifted_range(4, 4, 6, 4, axis, delta)]
        assert table.sortState.ref == _shifted_range(4, 3, 6, 5, axis, delta)
        assert [condition.ref for condition in table.sortState.sortCondition] == [
            _shifted_range(4, 5, 6, 5, axis, delta)
        ]

        validations = sheet.data_validations.dataValidation
        assert len(validations) == 1
        assert str(validations[0].sqref) == _shifted_range(3, 9, 5, 9, axis, delta)
        assert validations[0].formula1 == f"={_shifted_a1(3, 3, axis, delta, absolute=True)}"
        assert validations[0].formula2 == f"={_shifted_a1(5, 3, axis, delta, absolute=True)}"

        conditional_formatting = list(sheet.conditional_formatting)
        assert len(conditional_formatting) == 1
        assert str(conditional_formatting[0].sqref) == _shifted_range(
            3, 10, 5, 10, axis, delta
        )
        shifted_formula_row, shifted_formula_column = _shifted_position(3, 3, axis, delta)
        shifted_formula_column_letter = openpyxl.utils.get_column_letter(shifted_formula_column)
        assert conditional_formatting[0].rules[0].formula == [
            f"${shifted_formula_column_letter}{shifted_formula_row}>0"
        ]

        comment_cell = sheet[_shifted_a1(5, 13, axis, delta)]
        assert comment_cell.comment.text == "Structural note"
        assert comment_cell.comment.author == "Audit Author"
        assert {str(item) for item in sheet.merged_cells.ranges} == {
            _shifted_range(10, 3, 11, 4, axis, delta)
        }

        hyperlink_cell = sheet[_shifted_a1(4, 11, axis, delta)]
        assert hyperlink_cell.hyperlink.location == (
            f"Data!{_shifted_a1(3, 3, axis, delta)}"
        )

        expected_top_left = _shifted_a1(3, 3, axis, delta)
        expected_active = _shifted_a1(4, 4, axis, delta)
        expected_selection = _shifted_range(4, 4, 5, 5, axis, delta)
        assert sheet.freeze_panes == expected_top_left
        assert sheet.sheet_view.topLeftCell == expected_top_left
        assert sheet.sheet_view.pane.topLeftCell == expected_top_left
        assert sheet.sheet_view.selection[0].activeCell == expected_active
        assert str(sheet.sheet_view.selection[0].sqref) == expected_selection

        expected_row_break = 5 + delta if axis == "row" else 5
        expected_column_break = 4 + delta if axis == "column" else 4
        assert [page_break.id for page_break in sheet.row_breaks.brk] == [expected_row_break]
        assert [page_break.id for page_break in sheet.col_breaks.brk] == [expected_column_break]

        chart_anchor = sheet._charts[0].anchor._from
        assert chart_anchor.row == 3 + (delta if axis == "row" else 0)
        assert chart_anchor.col == 13 + (delta if axis == "column" else 0)
    finally:
        saved.close()
