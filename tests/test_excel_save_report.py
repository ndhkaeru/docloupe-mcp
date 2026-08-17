import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402


def _session_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


def _write_workbook(path: Path, first: str = "old", second: str = "keep") -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = first
    sheet["B1"] = second
    workbook.save(path)
    workbook.close()


def _rewrite_zip_part(path: Path, part_name: str, content: bytes) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = [(info, archive.read(info.filename)) for info in archive.infolist()]
    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(temporary, "w") as archive:
        for info, original in entries:
            archive.writestr(info, content if info.filename == part_name else original)
    temporary.replace(path)


def _add_custom_xml_part(path: Path, part_name: str, content: bytes) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        entries = [(info, archive.read(info.filename)) for info in archive.infolist()]
    content_types = next(raw for info, raw in entries if info.filename == "[Content_Types].xml")
    marker = b"</Types>"
    override = (
        f'<Override PartName="/{part_name}" ContentType="application/vnd.docloupe.custom+xml"/>'
    ).encode("utf-8")
    updated_content_types = content_types.replace(marker, override + marker)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(temporary, "w") as archive:
        for info, original in entries:
            archive.writestr(
                info,
                updated_content_types if info.filename == "[Content_Types].xml" else original,
            )
        archive.writestr(part_name, content)
    temporary.replace(path)


def test_json_save_report_includes_backup_paths_and_inline_verification(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    _write_workbook(source)

    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(session_key, "Sheet", [{"row_index": 0, "edits": {0: "new"}}])
    report = json.loads(M.excel_save(
        session_key,
        report_format="json",
        verify_preservation=True,
    ))

    assert report["schema_version"] == 1
    assert report["saved_path"] == str(source.resolve())
    assert Path(report["backup"]["backup_path"]).is_file()
    assert report["backup"]["reference_path"] == str(source.resolve())
    assert report["requested_semantic_paths"] == ["sheets/Sheet/cells/A1"]
    assert report["package_signatures"]["status"] == "not_present"
    assert report["warnings"] == []
    assert report["verification"]["status"] == "completed"
    assert report["verification"]["preservation_ok"] is True
    assert report["verification"]["unapproved_difference_count"] == 0
    assert report["verification"]["classification_counts"]["REQUESTED"] >= 1
    assert "worksheets/Sheet/cells/A1*" in report["verification"]["requested_paths"]
    assert "worksheets/Sheet/cells/A1/value" in report["changed_semantic_paths"]
    assert openpyxl.load_workbook(source).active["A1"].value == "new"


def test_inline_verification_approves_granular_package_edits(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    _write_workbook(source)

    session_key = _session_key(M.excel_load(str(source)))
    M._package_tools["excel_apply_package_transaction"](
        session_key,
        upsert=[
            {
                "path": "customXml/item1.xml",
                "content": "<custom/>",
                "content_type": "application/vnd.docloupe.custom+xml",
            }
        ],
        relationships=[
            {
                "source_part": "customXml/item1.xml",
                "relationships": [
                    {
                        "id": "rId1",
                        "type": "urn:docloupe:test:relationship",
                        "target": "../xl/workbook.xml",
                        "target_mode": "Internal",
                    }
                ],
            }
        ],
        content_types={
            "defaults": [
                {
                    "extension": "foo",
                    "content_type": "application/vnd.docloupe.foo",
                }
            ]
        },
    )
    report = json.loads(M.excel_save_as_copy(
        session_key,
        str(target),
        report_format="json",
        verify_preservation=True,
    ))

    expected_paths = [
        "package/customXml/item1.xml",
        "package/customXml/_rels/item1.xml.rels",
        "package/relationships/customXml/_rels/item1.xml.rels#*",
        "package/content_types/Default:foo",
        "package/content_types/Override:/customXml/item1.xml",
    ]
    assert report["dirty_features"] == ["package"]
    assert report["requested_semantic_paths"] == expected_paths
    assert report["verification"]["requested_paths"] == expected_paths
    assert report["verification"]["preservation_ok"] is True
    assert report["verification"]["unapproved_difference_count"] == 0
    with zipfile.ZipFile(target, "r") as archive:
        assert archive.read("customXml/item1.xml") == b"<custom/>"
        assert "customXml/_rels/item1.xml.rels" in archive.namelist()


def test_dimension_and_freeze_tools_record_inline_verifier_paths(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    _write_workbook(source)

    session_key = _session_key(M.excel_load(str(source)))
    M.excel_set_dimension(session_key, "Sheet", "row", 0, 24)
    M.excel_set_dimension(session_key, "Sheet", "col", 0, 18)
    M.excel_set_row_height(session_key, "Sheet", {"0": 30})
    M.excel_set_column_width(session_key, "Sheet", {"B": 22})
    M.excel_autofit_cols(session_key, "Sheet", [0, 1])
    M.excel_freeze_panes(session_key, "Sheet", 1, 1)
    report = json.loads(M.excel_save_as_copy(
        session_key,
        str(target),
        report_format="json",
        verify_preservation=True,
    ))

    assert report["dirty_features"] == [
        "row_properties",
        "column_properties",
        "sheet_views",
    ]
    assert report["requested_semantic_paths"] == [
        "sheets/Sheet/rows/0",
        "sheets/Sheet/columns",
        "sheets/Sheet/views",
    ]
    assert report["verification"]["requested_paths"] == [
        "worksheets/Sheet/rows/1",
        "worksheets/Sheet/*",
        "worksheets/Sheet/sheet_views",
    ]
    assert report["verification"]["preservation_ok"] is True
    assert report["verification"]["unapproved_difference_count"] == 0


def test_inline_verification_rejects_unrelated_package_corruption(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    _write_workbook(source)
    _add_custom_xml_part(source, "customXml/unrelated.xml", b"<unrelated>keep</unrelated>")
    original_reconstruct = M.reconstruct_excel

    def reconstruct_and_corrupt(data, destination):
        warnings = original_reconstruct(data, destination)
        _rewrite_zip_part(
            Path(destination),
            "customXml/unrelated.xml",
            b"<unrelated>corrupt</unrelated>",
        )
        return warnings

    monkeypatch.setattr(M, "reconstruct_excel", reconstruct_and_corrupt)
    session_key = _session_key(M.excel_load(str(source)))
    M._package_tools["excel_upsert_package_part"](
        session_key,
        "customXml/requested.xml",
        "<requested/>",
        content_type="application/vnd.docloupe.requested+xml",
    )

    with pytest.raises(M.SaveTransactionError) as raised:
        M.excel_save_as_copy(
            session_key,
            str(target),
            report_format="json",
            verify_preservation=True,
        )

    verification = raised.value.details["verification"]
    assert verification["preservation_ok"] is False
    assert verification["unapproved_difference_count"] >= 1
    assert "package/customXml/unrelated.xml" in verification["changed_semantic_paths"]
    assert not target.exists()


def test_inline_verification_reports_unapproved_sibling_loss(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    _write_workbook(source)
    original_reconstruct = M.reconstruct_excel

    def reconstruct_and_corrupt(data, destination):
        warnings = original_reconstruct(data, destination)
        workbook = openpyxl.load_workbook(destination)
        workbook.active["B1"] = "unexpected"
        workbook.save(destination)
        workbook.close()
        return warnings

    monkeypatch.setattr(M, "reconstruct_excel", reconstruct_and_corrupt)
    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(session_key, "Sheet", [{"row_index": 0, "edits": {0: "new"}}])
    original_bytes = source.read_bytes()
    with pytest.raises(M.SaveTransactionError) as raised:
        M.excel_save(
            session_key,
            report_format="json",
            verify_preservation=True,
        )

    verification = raised.value.details["verification"]
    assert verification["status"] == "completed"
    assert verification["preservation_ok"] is False
    assert verification["unapproved_difference_count"] >= 1
    assert any(
        path.startswith("worksheets/Sheet/cells/B1")
        for path in verification["changed_semantic_paths"]
    )
    assert source.read_bytes() == original_bytes
    workbook = openpyxl.load_workbook(source)
    assert workbook.active["A1"].value == "old"
    assert workbook.active["B1"].value == "keep"
    workbook.close()
    assert not (tmp_path / "backups").exists()


def test_save_as_copy_verifies_against_session_source_not_existing_target(tmp_path, monkeypatch):
    monkeypatch.setenv("DOCLOUPE_EXCEL_BACKUP_DIR", str(tmp_path / "backups"))
    source = tmp_path / "source.xlsx"
    target = tmp_path / "target.xlsx"
    _write_workbook(source)
    _write_workbook(target, first="target-old", second="target-old")

    session_key = _session_key(M.excel_load(str(source)))
    M.excel_edit_cells(session_key, "Sheet", [{"row_index": 0, "edits": {0: "new"}}])
    report = json.loads(M.excel_save_as_copy(
        session_key,
        str(target),
        report_format="json",
        verify_preservation=True,
    ))

    assert report["backup"]["reference_path"] == str(target.resolve())
    assert report["verification"]["reference_path"] == str(source.resolve())
    assert report["verification"]["preservation_ok"] is True
    workbook = openpyxl.load_workbook(target)
    assert workbook.active["A1"].value == "new"
    assert workbook.active["B1"].value == "keep"
    workbook.close()


def test_new_workbook_report_skips_inline_verification_without_reference(tmp_path):
    target = tmp_path / "new.xlsx"
    created = json.loads(M.excel_create_workbook(target_path=str(target)))
    session_key = created["session_key"]
    M.excel_edit_cells(session_key, "Sheet1", [{"row_index": 0, "edits": {0: "new"}}])

    report = json.loads(M.excel_save(
        session_key,
        report_format="json",
        verify_preservation=True,
    ))

    assert report["backup"] is None
    assert report["verification"]["status"] == "skipped"
    assert report["verification"]["reference_path"] is None
    assert target.is_file()
