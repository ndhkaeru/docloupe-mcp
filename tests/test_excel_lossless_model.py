import hashlib
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
import pytest
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.colors import Color


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

from core import reconstruct_excel, serialize_excel  # noqa: E402


REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
SHARED_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
CUSTOM_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml"
SHARED_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"


def _cell_xml(xml: str, coord: str) -> str:
    match = re.search(
        rf'<c\b(?=[^>]*\br="{re.escape(coord)}")[^>]*(?:/>|>.*?</c>)',
        xml,
        re.DOTALL,
    )
    assert match, f"missing cell {coord}"
    return match.group(0)


def _replace_cell(xml: str, coord: str, replacement: str) -> str:
    pattern = re.compile(
        rf'<c\b(?=[^>]*\br="{re.escape(coord)}")[^>]*(?:/>|>.*?</c>)',
        re.DOTALL,
    )
    updated, count = pattern.subn(replacement, xml, count=1)
    assert count == 1, f"missing cell {coord}"
    return updated


def _add_relationship(raw: bytes, rel_id: str, rel_type: str, target: str) -> bytes:
    root = ET.fromstring(raw)
    ET.SubElement(
        root,
        f"{{{REL_NS}}}Relationship",
        {"Id": rel_id, "Type": rel_type, "Target": target},
    )
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _add_override(raw: bytes, part_name: str, content_type: str) -> bytes:
    root = ET.fromstring(raw)
    ET.SubElement(
        root,
        f"{{{CT_NS}}}Override",
        {"PartName": part_name, "ContentType": content_type},
    )
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _patch_workbook_xml(raw: bytes) -> bytes:
    xml = raw.decode("utf-8")
    xml = re.sub(
        r"<workbookPr\b[^>]*/>",
        '<workbookPr date1904="1" codeName="LosslessBook" filterPrivacy="1" '
        'saveExternalLinkValues="0" showObjects="none" updateLinks="never"/>',
        xml,
        count=1,
    )
    first_view = re.search(r"<workbookView\b[^>]*/>", xml)
    assert first_view
    second_view = first_view.group(0).replace("/>", ' activeTab="0"/>' if "activeTab=" not in first_view.group(0) else "/>")
    xml = xml.replace(first_view.group(0), first_view.group(0) + second_view, 1)
    xml = xml.replace(
        "<bookViews>",
        '<workbookProtection lockStructure="1" workbookPassword="ABCD"/><bookViews>',
        1,
    )
    return xml.encode("utf-8")


def _make_lossless_fixture(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Lossless"
    sheet["A1"] = "editable"
    sheet["B1"] = "shared placeholder"
    sheet["C1"] = "inline placeholder"
    sheet["A2"] = "=1+2"
    sheet["B2"] = "=2+3"
    sheet["C2"] = "=3+4"
    sheet["D5"].font = Font(bold=True)

    named = NamedStyle(name="AuditNamed")
    named.font = Font(name="Aptos", size=12)
    workbook.add_named_style(named)
    sheet["E1"] = "colors"
    sheet["E1"].style = "AuditNamed"
    sheet["E1"].font = Font(name="Aptos", color=Color(theme=3, tint=-0.25), strike=True)
    sheet["E1"].fill = PatternFill("solid", fgColor=Color(indexed=6))
    sheet["E1"].alignment = Alignment(readingOrder=2, relativeIndent=1, justifyLastLine=True)
    sheet["E1"].border = Border(
        start=Side(style="thin", color=Color(theme=4, tint=0.2)),
        end=Side(style="medium", color=Color(auto=True)),
        vertical=Side(style="dashed", color=Color(indexed=7)),
        diagonal=Side(style="hair", color=Color(rgb="FF112233")),
        diagonalUp=True,
        outline=False,
    )
    workbook.save(path)

    shared_strings = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">'
        '<si>'
        '<r><rPr><rFont val="Aptos"/><b/><strike/><color theme="5" tint="0.25"/>'
        '<sz val="12"/><vertAlign val="superscript"/></rPr><t xml:space="preserve">Shared </t></r>'
        '<r><rPr><i/><u val="double"/><color indexed="5"/></rPr><t>Rich</t></r>'
        '<rPh sb="0" eb="6"><t>ruby</t></rPh>'
        '<phoneticPr fontId="1" type="fullwidthKatakana" alignment="center"/>'
        '</si></sst>'
    ).encode("utf-8")
    custom_xml = b'<?xml version="1.0" encoding="UTF-8"?><audit marker="unknown-part"/>'

    replacement = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            raw = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = raw.decode("utf-8")
                xml = _replace_cell(xml, "B1", '<c r="B1" t="s"><v>0</v></c>')
                xml = _replace_cell(
                    xml,
                    "C1",
                    '<c r="C1" t="inlineStr"><is>'
                    '<r><rPr><color indexed="4"/><outline/></rPr><t>Inline</t></r>'
                    '<r><rPr><color auto="1"/><shadow/></rPr><t xml:space="preserve"> Run</t></r>'
                    '</is></c>',
                )
                xml = _replace_cell(xml, "A2", '<c r="A2"><f t="normal" ca="1">1+2</f></c>')
                xml = _replace_cell(xml, "B2", '<c r="B2"><f t="normal" bx="1">2+3</f><v></v></c>')
                xml = _replace_cell(xml, "C2", '<c r="C2"><f t="normal">3+4</f><v>7</v></c>')
                raw = xml.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                raw = _add_relationship(raw, "rIdSharedLossless", SHARED_REL, "sharedStrings.xml")
            elif item.filename == "_rels/.rels":
                raw = _add_relationship(raw, "rIdCustomLossless", CUSTOM_REL, "customXml/item99.xml")
            elif item.filename == "[Content_Types].xml":
                raw = _add_override(raw, "/xl/sharedStrings.xml", SHARED_CT)
                raw = _add_override(raw, "/customXml/item99.xml", "application/xml")
            elif item.filename == "xl/workbook.xml":
                raw = _patch_workbook_xml(raw)
            target.writestr(item, raw)
        target.writestr("xl/sharedStrings.xml", shared_strings)
        target.writestr("customXml/item99.xml", custom_xml)
    replacement.replace(path)


def _part_bytes(path: Path, part: str) -> bytes:
    with zipfile.ZipFile(path, "r") as archive:
        return archive.read(part)


def test_lossless_session_contract_and_no_edit_exact_copy(tmp_path):
    source = tmp_path / "lossless-source.xlsx"
    output = tmp_path / "lossless-output.xlsx"
    _make_lossless_fixture(source)

    data = serialize_excel(str(source))
    cells = data["sheets"][0]["rows"]
    shared = cells[0]["cells"][1]
    inline = cells[0]["cells"][2]
    formula_missing = cells[1]["cells"][0]
    formula_empty = cells[1]["cells"][1]
    formula_value = cells[1]["cells"][2]
    blank = cells[4]["cells"][3]
    colors = cells[0]["cells"][4]

    assert shared["rich_text"]["storage"] == "shared"
    assert shared["rich_text"]["is_rich_text"] is True
    assert [run["text"] for run in shared["rich_text"]["runs"]] == ["Shared ", "Rich"]
    assert shared["rich_text"]["runs"][0]["start"] == 0
    assert shared["rich_text"]["runs"][1]["end"] == len("Shared Rich")
    assert shared["rich_text"]["runs"][0]["font"]["color"]["type"] == "theme"
    assert shared["rich_text"]["runs"][1]["font"]["color"]["type"] == "indexed"
    assert shared["rich_text"]["phonetic_runs"][0]["text"] == "ruby"
    assert shared["rich_text"]["phonetic_properties"]["type"] == "fullwidthKatakana"
    assert inline["rich_text"]["storage"] == "inline"
    assert inline["rich_text"]["runs"][1]["xml_space"] == "preserve"

    assert formula_missing["formula"] == {
        "text": "1+2",
        "attrs": {"t": "normal", "ca": "1"},
        "cached_value": None,
        "cache_state": "missing",
    }
    assert formula_empty["formula"]["cache_state"] == "empty"
    assert formula_empty["formula"]["attrs"]["bx"] == "1"
    assert formula_value["formula"]["cache_state"] == "value"
    assert formula_value["formula"]["cached_value"] == "7"

    assert blank["present"] is True
    assert blank["v"] is None
    assert colors["font_color"] == {"type": "theme", "theme": 3, "tint": -0.25}
    assert colors["fill_color"] == {"type": "indexed", "indexed": 6}
    assert colors["alignment"]["readingOrder"] == 2
    assert colors["alignment"]["relativeIndent"] == 1
    assert colors["alignment"]["justifyLastLine"] is True
    assert colors["border_semantics"]["start"]["color"]["type"] == "theme"
    assert colors["border_semantics"]["end"]["color"]["type"] == "auto"
    assert colors["border_semantics"]["vertical"]["color"]["type"] == "indexed"
    assert colors["xf"]["definition"]["attrs"].get("xfId") is not None

    workbook_semantics = data["workbook_semantics"]
    assert workbook_semantics["workbook_properties"]["attrs"]["date1904"] == "1"
    assert workbook_semantics["protection"]["attrs"]["lockStructure"] == "1"
    assert len(workbook_semantics["views"]) == 2
    assert "customXml/item99.xml" in data["_lossless"]["package_graph"]["parts"]
    assert "_rels/.rels" in data["_lossless"]["package_graph"]["relationships"]

    reconstruct_excel(data, str(output))

    assert hashlib.sha256(output.read_bytes()).digest() == hashlib.sha256(source.read_bytes()).digest()


def test_content_only_edit_preserves_unrelated_semantics_and_package_graph(tmp_path):
    source = tmp_path / "lossless-source.xlsx"
    output = tmp_path / "content-edit.xlsx"
    _make_lossless_fixture(source)
    data = serialize_excel(str(source))
    data["sheets"][0]["rows"][0]["cells"][0]["v"] = "changed"

    reconstruct_excel(data, str(output))

    loaded = openpyxl.load_workbook(output, rich_text=True)
    assert loaded["Lossless"]["A1"].value == "changed"
    loaded.close()

    exact_parts = [
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/_rels/workbook.xml.rels",
        "xl/workbook.xml",
        "xl/styles.xml",
        "xl/sharedStrings.xml",
        "customXml/item99.xml",
    ]
    for part in exact_parts:
        assert _part_bytes(output, part) == _part_bytes(source, part), part

    source_sheet = _part_bytes(source, "xl/worksheets/sheet1.xml").decode("utf-8")
    output_sheet = _part_bytes(output, "xl/worksheets/sheet1.xml").decode("utf-8")
    for coord in ("B1", "C1", "A2", "B2", "C2", "D5", "E1"):
        assert _cell_xml(output_sheet, coord) == _cell_xml(source_sheet, coord), coord


def test_expert_package_edits_are_applied_transactionally(tmp_path):
    source = tmp_path / "lossless-source.xlsx"
    output = tmp_path / "package-edited.xlsx"
    _make_lossless_fixture(source)
    data = serialize_excel(str(source))
    data["_package_edits"] = {
        "upsert": {"customXml/agent.xml": {"xml": '<agent created="yes"/>'}},
        "relationships": [
            {
                "source": "xl/workbook.xml",
                "id": "rIdAgentPackage",
                "type": CUSTOM_REL,
                "target": "../customXml/agent.xml",
            }
        ],
        "content_types": [
            {"part_name": "/customXml/agent.xml", "content_type": "application/xml"}
        ],
    }

    reconstruct_excel(data, str(output))

    assert _part_bytes(output, "customXml/agent.xml") == b'<agent created="yes"/>'
    rels = ET.fromstring(_part_bytes(output, "xl/_rels/workbook.xml.rels"))
    relation = next(node for node in rels if node.get("Id") == "rIdAgentPackage")
    assert relation.get("Target") == "../customXml/agent.xml"
    content_types = ET.fromstring(_part_bytes(output, "[Content_Types].xml"))
    override = next(node for node in content_types if node.get("PartName") == "/customXml/agent.xml")
    assert override.get("ContentType") == "application/xml"

    failed_output = tmp_path / "failed-package-edit.xlsx"
    broken = serialize_excel(str(source))
    broken["_package_edits"] = {
        "relationships": [{"source": "xl/workbook.xml", "id": "rIdBroken", "type": CUSTOM_REL}]
    }
    with pytest.raises(ValueError):
        reconstruct_excel(broken, str(failed_output))
    assert not failed_output.exists()
