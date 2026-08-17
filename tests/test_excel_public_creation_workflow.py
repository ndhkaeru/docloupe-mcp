"""
Verification tests for the Excel MCP "public tool creation" checklist bullets
(PRESERVATION_FIX_CHECKLIST.md, sections "Muc tieu va Pham vi",
"excel_get_cell/excel_read_range/excel_get_rows", "Rich Text", and
"excel_set_cell_style_semantics"). These exercise the PUBLIC MCP tool
functions in servers/excel/main.py exactly as an agent would call them --
never the internal helpers directly.

Covers:
  1. excel_create_workbook builds a valid, loadable, savable .xlsx with no
     pre-existing template.
  2. A realistic multi-tool workflow (style/borders/formula/defined-name/
     auto-filter/table) built purely from public tools round-trips through
     save + reload.
  3. excel_edit_rich_text partial-cell styling (style_range) touches only the
     targeted Unicode character range, including a Vietnamese multi-byte
     example and a case where a newline is counted in the offsets.
  4. excel_edit_rich_text's other operations (replace_runs, style_run,
     insert_text, delete_range, replace_range, set_phonetic).
  5. excel_get_cell / excel_read_range / excel_get_rows semantic flags
     (include_rich_text, include_formula_cache, include_semantics) return the
     documented shapes and actually gate response size/shape.
  6. Blank-but-styled / hyperlinked / explicit-empty cells stay "present".
  7. Every rich-text run font sub-attribute (name, size, bold, italic,
     underline, strike, color, vertAlign, charset, family, scheme, outline,
     shadow, condense, extend) round-trips through serialize_excel /
     reconstruct_excel and is readable via excel_get_rich_text.
  8. excel_set_cell_style_semantics: dry-run + session-level round trip for
     applyFont/applyFill/applyBorder/applyAlignment/applyProtection/xfId/
     pivotButton/quotePrefix, including persistence through save + reload.
"""
import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "servers" / "excel"))

import main as M  # noqa: E402
from core import reconstruct_excel, serialize_excel  # noqa: E402


def _load_key(load_result: str) -> str:
    return load_result.split("session_key='")[1].split("'")[0]


# ---------------------------------------------------------------------------
# 1. excel_create_workbook -- no pre-existing template required
# ---------------------------------------------------------------------------

def test_create_workbook_without_template_produces_valid_saveable_file(tmp_path):
    out_path = tmp_path / "brand-new.xlsx"
    assert not out_path.exists()

    created = json.loads(M.excel_create_workbook(
        sheet_names=["Cover", "Data"],
        active_sheet="Data",
        document_properties={"title": "Agent-created workbook", "creator": "Docloupe Agent"},
        target_path=str(out_path),
    ))
    key = created["session_key"]
    assert created["sheets"] == ["Cover", "Data"]
    assert created["active_sheet"] == "Data"
    assert created["capabilities"]["requires_explicit_save"] is True

    # Nothing is written until excel_save is called.
    assert not out_path.exists()

    M.excel_edit_cells(key, "Data", [{"row_index": 0, "edits": {0: "hello"}}])
    M.excel_save(key)
    assert out_path.exists()

    report = json.loads(M.excel_validate_workbook(str(out_path)))
    assert report["valid"] is True

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["Cover", "Data"]
    assert wb["Data"]["A1"].value == "hello"
    assert wb.properties.title == "Agent-created workbook"
    assert wb.properties.creator == "Docloupe Agent"
    assert wb.active.title == "Data"
    wb.close()


# ---------------------------------------------------------------------------
# 2. Substantial multi-tool workflow: create -> style/borders/formula/
#    defined-name/auto-filter/table -> save -> reload -> verify semantics.
# ---------------------------------------------------------------------------

def test_workflow_builds_semantics_with_public_tools_and_round_trips(tmp_path):
    out_path = tmp_path / "workflow.xlsx"
    created = json.loads(M.excel_create_workbook(
        sheet_names=["Sales"],
        target_path=str(out_path),
    ))
    key = created["session_key"]

    M.excel_edit_cells(key, "Sales", [
        {"row_index": 0, "edits": {0: "Item", 1: "Qty", 2: "Price", 3: "Total"}},
        {"row_index": 1, "edits": {0: "Widget", 1: 3, 2: 2.5}},
        {"row_index": 2, "edits": {0: "Gadget", 1: 5, 2: 4.0}},
    ])

    # Structured whole-cell style: bold + theme font color + rgb fill + alignment.
    M.excel_set_style(key, "Sales", 0, 0, r2=0, c2=3, style={
        "bold": True,
        "font": {"color": {"type": "theme", "theme": 4, "tint": -0.2}},
        "fill": {"foreground": {"type": "rgb", "rgb": "FFDDEBF7"}},
        "alignment": {"horizontal": "center", "readingOrder": 0},
    })

    # Borders with a theme color across the header row.
    M.excel_set_borders(key, "Sales", r1=0, c1=0, r2=0, c2=3,
                         style="thin", color={"type": "theme", "theme": 5})

    # Formulas with explicit cached values.
    M.excel_set_formula(key, "Sales", "D2", "=B2*C2", cached_value=7.5, cached_value_present=True)
    M.excel_set_formula(key, "Sales", "D3", "=B3*C3", cached_value=20.0, cached_value_present=True)

    # Defined name.
    M.excel_add_defined_name(key, "SalesData", "Sales!$A$1:$D$3")

    # Auto filter.
    M.excel_set_auto_filter(key, "Sales", ref="A1:D3")

    # Table.
    M.excel_add_table(key, "Sales", "SalesTable", "A1:D3", style={"name": "TableStyleMedium9"})

    M.excel_save(key)

    report = json.loads(M.excel_validate_workbook(str(out_path)))
    assert report["valid"] is True

    # ---- reload into a brand-new session and confirm everything reads back ----
    key2 = _load_key(M.excel_load(str(out_path)))

    header_cell = json.loads(M.excel_get_cell(key2, "Sales", 0, 0, include_semantics=True))
    assert header_cell["semantics"]["font"]["bold"] is True
    assert header_cell["semantics"]["font"]["color"]["type"] == "theme"
    assert header_cell["semantics"]["font"]["color"]["theme"] == 4
    assert header_cell["semantics"]["alignment"]["horizontal"] == "center"
    assert header_cell["semantics"]["fill"]["foreground"]["type"] == "rgb"
    assert header_cell["semantics"]["border"]["top"]["color_object"]["type"] == "theme"
    assert header_cell["semantics"]["border"]["top"]["color_object"]["theme"] == 5

    formula_cell = json.loads(M.excel_get_cell(key2, "Sales", 1, 3, include_formula_cache=True))
    # NOTE: excel_set_formula's in-session formula.text includes a leading "="
    # (see _set_formula_cell), but OOXML <f> elements never store the "=", so
    # once a formula round-trips through save+reload, serialize_excel parses
    # it back without the prefix. Both are internally consistent; strip "="
    # here so this assertion focuses on the actual formula content/cache.
    assert formula_cell["formula"]["text"].lstrip("=") == "B2*C2"
    # Cached values always come back as raw XML text (a string) once they
    # have round-tripped through disk, matching serialize_excel's existing
    # convention (see test_excel_lossless_model.py's "7" cached_value case).
    assert float(formula_cell["formula"]["cached_value"]) == 7.5

    names = json.loads(M.excel_list_defined_names(key2))
    assert any(n["name"] == "SalesData" and n["value"] == "Sales!$A$1:$D$3"
               for n in names["defined_names"])

    tables = json.loads(M.excel_list_tables(key2))
    assert any(t["name"] == "SalesTable" for t in tables["tables"])

    wb2 = openpyxl.load_workbook(out_path)
    assert wb2["Sales"].auto_filter.ref == "A1:D3"
    assert "SalesTable" in wb2["Sales"].tables
    wb2.close()


# ---------------------------------------------------------------------------
# 3. excel_edit_rich_text: style_range touches only the targeted characters.
# ---------------------------------------------------------------------------

def test_edit_rich_text_style_range_only_targets_selected_characters_vietnamese():
    created = json.loads(M.excel_create_workbook(sheet_names=["S"]))
    key = created["session_key"]
    text = "Xin chào các bạn"  # "Xin chào các bạn"

    M.excel_edit_rich_text(key, "S", "B2", operations=[
        {"op": "replace_runs", "runs": [{"text": text, "font": {}}]},
    ])
    start = text.index("chào")
    end = start + len("chào")
    M.excel_edit_rich_text(key, "S", "B2", expected_text=text, operations=[
        {"op": "style_range", "start": start, "end": end,
         "style": {"strike": True, "color": {"type": "rgb", "rgb": "FFFF0000"}}},
    ])

    model = json.loads(M.excel_get_rich_text(key, "S", "B2"))
    assert model["text"] == text

    styled_text = "".join(r["text"] for r in model["runs"] if r["font"].get("strike"))
    plain_text = "".join(r["text"] for r in model["runs"] if not r["font"].get("strike"))
    assert styled_text == text[start:end]
    assert plain_text == text[:start] + text[end:]
    # No run outside the targeted range picked up the style.
    offset = 0
    for run in model["runs"]:
        run_start, run_end = offset, offset + len(run["text"])
        if run["font"].get("strike"):
            assert run_start >= start and run_end <= end
        else:
            assert run_start >= end or run_end <= start
        offset = run_end


def test_edit_rich_text_style_range_counts_newline_in_offsets():
    created = json.loads(M.excel_create_workbook(sheet_names=["S"]))
    key = created["session_key"]
    text = "Dòng một\nDòng hai\nDòng ba"  # "Dòng một\nDòng hai\nDòng ba"

    M.excel_edit_rich_text(key, "S", "B3", operations=[
        {"op": "replace_runs", "runs": [{"text": text, "font": {}}]},
    ])

    first_newline = text.index("\n")
    last_newline = text.rindex("\n")
    third_start = last_newline + 1

    # Style only the first line (up to, but excluding, the first newline).
    M.excel_edit_rich_text(key, "S", "B3", expected_text=text, operations=[
        {"op": "style_range", "start": 0, "end": first_newline,
         "style": {"strike": True, "color": {"type": "rgb", "rgb": "FFFF0000"}}},
    ])
    # Style only the third line (starting right after the second newline).
    M.excel_edit_rich_text(key, "S", "B3", operations=[
        {"op": "style_range", "start": third_start, "end": len(text),
         "style": {"bold": True, "underline": "single"}},
    ])

    model = json.loads(M.excel_get_rich_text(key, "S", "B3"))
    assert model["text"] == text

    styled_first = "".join(r["text"] for r in model["runs"] if r["font"].get("strike"))
    styled_third = "".join(r["text"] for r in model["runs"] if r["font"].get("bold"))
    assert styled_first == text[:first_newline]
    assert styled_third == text[third_start:]
    # The newline characters themselves sit exactly at the boundaries used
    # above, proving they were counted as ordinary characters in the offset
    # space (not skipped, not double-counted).
    assert text[first_newline] == "\n"
    assert text[last_newline] == "\n"
    # The (untouched) middle line and both newlines remain plain.
    middle_and_newlines = "".join(
        r["text"] for r in model["runs"]
        if not r["font"].get("strike") and not r["font"].get("bold")
    )
    assert middle_and_newlines == text[first_newline:third_start]


# ---------------------------------------------------------------------------
# 4. excel_edit_rich_text: the other documented operations.
# ---------------------------------------------------------------------------

def test_edit_rich_text_supports_replace_style_insert_delete_replace_and_phonetic():
    created = json.loads(M.excel_create_workbook(sheet_names=["S"]))
    key = created["session_key"]

    # replace_runs
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "replace_runs",
        "runs": [
            {"text": "Hello ", "font": {"bold": True}},
            {"text": "World", "font": {"italic": True}},
        ],
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["text"] == "Hello World"
    assert model["runs"][0]["font"]["bold"] is True
    assert model["runs"][1]["font"]["italic"] is True

    # style_run: patch run 1 only.
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "style_run", "run_index": 1, "style": {"underline": "single"},
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["runs"][1]["font"]["underline"] == "single"
    assert model["runs"][0]["font"].get("underline") is None

    # insert_text
    idx = model["text"].index("World")
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "insert_text", "start": idx, "text": "there, ",
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["text"] == "Hello there, World"

    # delete_range
    start = model["text"].index("there, ")
    end = start + len("there, ")
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "delete_range", "start": start, "end": end,
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["text"] == "Hello World"

    # replace_range
    start = model["text"].index("World")
    end = start + len("World")
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "replace_range", "start": start, "end": end, "text": "Vietnam",
        "style": {"bold": True},
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["text"] == "Hello Vietnam"

    # set_phonetic
    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "set_phonetic",
        "runs": [{"text": "hint", "start": 0, "end": 5}],
        "properties": {"type": "fullwidthKatakana"},
    }])
    model = json.loads(M.excel_get_rich_text(key, "S", "A1"))
    assert model["phonetic_runs"][0]["text"] == "hint"
    assert model["phonetic_properties"]["type"] == "fullwidthKatakana"


# ---------------------------------------------------------------------------
# 5 & 6. excel_get_cell / excel_read_range / excel_get_rows semantic flags.
# ---------------------------------------------------------------------------

def _session_with_rich_formula_and_semantics():
    created = json.loads(M.excel_create_workbook(sheet_names=["S"]))
    key = created["session_key"]

    M.excel_edit_rich_text(key, "S", "A1", operations=[{
        "op": "replace_runs",
        "runs": [
            {"text": "Red", "font": {"color": {"type": "rgb", "rgb": "FFFF0000"}}},
            {"text": "Blue", "font": {"color": {"type": "rgb", "rgb": "FF0000FF"}, "bold": True}},
        ],
    }])

    M.excel_set_formula(key, "S", "B1", "=1+41", cached_value=42, cached_value_present=True,
                         formula_attributes={"ca": "1"})

    M.excel_add_named_style(key, "Accent", {"bold": True})
    M.excel_set_style(key, "S", 2, 0, style={
        "named_style": "Accent",
        "font": {"color": {"type": "theme", "theme": 3, "tint": -0.1}},
        "fill": {"foreground": {"type": "indexed", "indexed": 22}},
        "alignment": {"horizontal": "right", "readingOrder": 2,
                      "relativeIndent": 1, "justifyLastLine": True},
    })
    M.excel_set_borders(key, "S", r1=2, c1=0, r2=2, c2=0,
                         style="thin", color={"type": "indexed", "indexed": 8})
    M.excel_set_cell_style_semantics(key, "S", "A3", xf={"xfId": 7})
    return key


def test_read_tools_return_rich_text_runs_when_flag_set():
    key = _session_with_rich_formula_and_semantics()

    cell = json.loads(M.excel_get_cell(key, "S", 0, 0, include_rich_text=True))
    assert [r["text"] for r in cell["rich_text"]["runs"]] == ["Red", "Blue"]
    assert cell["rich_text"]["runs"][1]["font"]["bold"] is True

    rng = json.loads(M.excel_read_range(key, "S", "A1:A1", include_rich_text=True))
    cell_from_range = rng["values"][0][0]
    assert [r["text"] for r in cell_from_range["rich_text"]["runs"]] == ["Red", "Blue"]

    rows = json.loads(M.excel_get_rows(key, "S", 0, 1, include_rich_text=True))
    cell_from_rows = rows[0]["cells"][0]
    assert [r["text"] for r in cell_from_rows["rich_text"]["runs"]] == ["Red", "Blue"]

    # Without the flag: no rich_text.runs leak into the plain response.
    plain = json.loads(M.excel_get_cell(key, "S", 0, 0))
    assert "rich_text" not in plain
    assert plain["value"] == "RedBlue"


def test_read_tools_separate_formula_text_attrs_and_cached_value():
    key = _session_with_rich_formula_and_semantics()

    with_cache = json.loads(M.excel_get_cell(key, "S", 0, 1, include_formula_cache=True))
    assert with_cache["formula"]["text"] == "=1+41"
    assert with_cache["formula"]["attributes"]["ca"] == "1"
    assert with_cache["formula"]["cached_value"] == 42
    # The cached value is a distinct field, not merged into "value".
    assert with_cache["value"] == "=1+41"
    assert with_cache["value"] != with_cache["formula"]["cached_value"]

    without_cache = json.loads(M.excel_get_cell(key, "S", 0, 1))
    assert without_cache["formula"]["text"] == "=1+41"
    assert "cached_value" not in without_cache["formula"]
    assert "cached_value_state" not in without_cache["formula"]


def test_read_tools_return_raw_color_object_not_resolved_rgb():
    key = _session_with_rich_formula_and_semantics()

    cell = json.loads(M.excel_get_cell(key, "S", 2, 0, include_semantics=True))
    font_color = cell["semantics"]["font"]["color"]
    assert font_color["type"] == "theme"
    assert font_color["theme"] == 3
    assert font_color["tint"] == -0.1

    fill_fg = cell["semantics"]["fill"]["foreground"]
    assert fill_fg["type"] == "indexed"
    assert fill_fg["indexed"] == 22

    border_color = cell["semantics"]["border"]["top"]["color_object"]
    assert border_color["type"] == "indexed"
    assert border_color["indexed"] == 8


def test_read_tools_return_advanced_alignment_border_xf_and_named_style():
    key = _session_with_rich_formula_and_semantics()

    cell = json.loads(M.excel_get_cell(key, "S", 2, 0, include_semantics=True))
    alignment = cell["semantics"]["alignment"]
    assert alignment["readingOrder"] == 2
    assert alignment["relativeIndent"] == 1
    assert alignment["justifyLastLine"] is True
    assert alignment["horizontal"] == "right"

    assert cell["semantics"]["named_style"] == "Accent"
    assert cell["semantics"]["xf"]["xfId"] == 7
    assert cell["semantics"]["border"]["top"]["style"] == "thin"


def test_read_tools_flags_gate_response_shape_and_size():
    key = _session_with_rich_formula_and_semantics()

    full = M.excel_get_cell(key, "S", 0, 1, include_rich_text=True,
                             include_formula_cache=True, include_semantics=True)
    minimal = M.excel_get_cell(key, "S", 0, 1)
    full_obj = json.loads(full)
    minimal_obj = json.loads(minimal)

    assert "semantics" not in minimal_obj
    assert "rich_text" not in minimal_obj
    assert "cached_value" not in minimal_obj.get("formula", {})
    assert "semantics" in full_obj
    assert "cached_value" in full_obj["formula"]
    assert len(minimal) < len(full)


def test_read_tools_preserve_blank_styled_hyperlink_and_explicit_empty_cells():
    created = json.loads(M.excel_create_workbook(sheet_names=["S"]))
    key = created["session_key"]

    # Blank cell with only fill/border style (no value at all).
    M.excel_set_style(key, "S", 5, 0, style={"fill": "FFFFFF00"})
    styled = json.loads(M.excel_get_cell(key, "S", 5, 0, include_semantics=True))
    assert styled["present"] is True
    assert styled["value"] in (None, "")
    assert styled["semantics"]["fill"]["foreground"]["rgb"] == "FFFFFF00"

    # Blank cell with only a hyperlink (no display text).
    M.excel_set_hyperlink(key, "S", "B6", target="https://example.com/")
    linked = json.loads(M.excel_get_cell(key, "S", 5, 1, include_semantics=True))
    assert linked["present"] is True
    assert linked["value"] in (None, "")
    session = M._sessions[key]
    sheet = next(s for s in session["sheets"] if s["name"] == "S")
    assert sheet["hyperlinks"]["B6"]["target"] == "https://example.com/"

    # Explicit-empty formula-cache cell (formula present, cache explicitly empty).
    M.excel_set_formula(key, "S", "C6", "=A1", cached_value=None, cached_value_present=True)
    empty_formula = json.loads(M.excel_get_cell(key, "S", 5, 2, include_formula_cache=True))
    assert empty_formula["present"] is True
    assert empty_formula["formula"]["cached_value_state"] == "empty"
    assert empty_formula["formula"]["cached_value"] is None

    # All three also survive through excel_get_rows.
    rows = json.loads(M.excel_get_rows(key, "S", 5, 6, include_semantics=True, include_formula_cache=True))
    row_cells = rows[0]["cells"]
    assert row_cells[0]["present"] is True
    assert row_cells[1]["present"] is True
    assert row_cells[2]["present"] is True
    assert row_cells[2]["formula"]["cached_value_state"] == "empty"


# ---------------------------------------------------------------------------
# 7. Rich-text run: every font sub-attribute round-trips.
# ---------------------------------------------------------------------------

def _make_full_font_run_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rich"
    ws["A1"] = "placeholder"
    wb.save(path)

    rpr = (
        '<rPr><rFont val="Calibri"/><charset val="1"/><family val="2"/>'
        '<scheme val="minor"/><sz val="14"/><u val="double"/>'
        '<vertAlign val="superscript"/><b/><i/><strike/><outline/><shadow/>'
        '<condense/><extend/><color rgb="FF112233"/></rPr>'
    )
    new_cell = f'<c r="A1" t="inlineStr"><is><r>{rpr}<t>Full</t></r></is></c>'

    replacement = path.with_suffix(".patched.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(replacement, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            raw = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = raw.decode("utf-8")
                xml, count = re.subn(r'<c r="A1"[^>]*(?:/>|>.*?</c>)', new_cell, xml,
                                      count=1, flags=re.DOTALL)
                assert count == 1
                raw = xml.encode("utf-8")
            target.writestr(item, raw)
    replacement.replace(path)


_FONT_SUBATTRS = (
    "name", "size", "bold", "italic", "underline", "strike", "vertAlign",
    "charset", "family", "scheme", "outline", "shadow", "condense", "extend",
)


def test_rich_text_run_all_font_subattributes_round_trip(tmp_path):
    src = tmp_path / "fullfont.xlsx"
    _make_full_font_run_fixture(src)

    data = serialize_excel(str(src))
    run = data["sheets"][0]["rows"][0]["cells"][0]["rich_text"]["runs"][0]
    font = run["font"]
    assert font["name"] == "Calibri"
    assert font["size"] == 14.0
    assert font["bold"] is True
    assert font["italic"] is True
    assert font["underline"] == "double"
    assert font["strike"] is True
    assert font["vertAlign"] == "superscript"
    assert font["charset"] == 1
    assert font["family"] == 2
    assert font["scheme"] == "minor"
    assert font["outline"] is True
    assert font["shadow"] is True
    assert font["condense"] is True
    assert font["extend"] is True
    assert font["color"]["type"] == "rgb"
    assert font["color"]["rgb"] == "FF112233"

    # Force the rich-text feature to actually regenerate (not passthrough) on
    # save, by re-asserting the same run style through the public tool. This
    # exercises the real write path (_run_properties_xml) instead of the
    # byte-identical no-edit fast path.
    key = _load_key(M.excel_load(str(src)))
    M.excel_edit_rich_text(key, "Rich", "A1", operations=[
        {"op": "style_run", "run_index": 0, "style": {}},
    ])
    out = tmp_path / "fullfont-out.xlsx"
    M.excel_save(key, str(out))

    reread = serialize_excel(str(out))
    run2 = reread["sheets"][0]["rows"][0]["cells"][0]["rich_text"]["runs"][0]
    for field in _FONT_SUBATTRS:
        assert run2["font"][field] == font[field], field
    assert run2["font"]["color"]["type"] == "rgb"
    assert run2["font"]["color"]["rgb"] == "FF112233"

    # And via the actual MCP read tool (excel_get_rich_text / _rich_text_model).
    key2 = _load_key(M.excel_load(str(out)))
    rich = json.loads(M.excel_get_rich_text(key2, "Rich", "A1"))
    tool_font = rich["runs"][0]["font"]
    for field in _FONT_SUBATTRS:
        assert tool_font[field] == font[field], field
    assert tool_font["color"]["type"] == "rgb"
    assert tool_font["color"]["rgb"] == "FF112233"


# ---------------------------------------------------------------------------
# 8. excel_set_cell_style_semantics: dry-run, session round trip, and the
#    (currently unwired) persistence gap.
# ---------------------------------------------------------------------------

_XF_FLAGS = {
    "applyFont": True, "applyFill": True, "applyBorder": True,
    "applyAlignment": True, "applyProtection": True,
    "xfId": 3, "pivotButton": True, "quotePrefix": True,
}


def test_set_cell_style_semantics_dry_run_and_session_roundtrip(tmp_path):
    src = tmp_path / "xf.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "value"
    wb.save(src)
    key = _load_key(M.excel_load(str(src)))
    baseline = json.loads(M.excel_get_cell(key, "Sheet", 0, 0, include_semantics=True))
    baseline_xf = baseline["semantics"]["xf"]

    dry = json.loads(M.excel_set_cell_style_semantics(
        key, "Sheet", "A1", xf=dict(_XF_FLAGS), dry_run=True,
    ))
    assert dry["dry_run"] is True
    assert dry["changes"][0]["after"]["xf"]["xfId"] == 3
    # A dry run must not mutate the session: the cell's xf semantics must be
    # exactly what they were before the call (loaded cells already carry
    # style_id/definition metadata, so this is not necessarily {}).
    unaffected = json.loads(M.excel_get_cell(key, "Sheet", 0, 0, include_semantics=True))
    assert unaffected["semantics"]["xf"] == baseline_xf

    applied = json.loads(M.excel_set_cell_style_semantics(
        key, "Sheet", "A1", xf=dict(_XF_FLAGS), dry_run=False,
    ))
    assert applied["dry_run"] is False

    # Session-level round trip: immediately readable via the read tools.
    after = json.loads(M.excel_get_cell(key, "Sheet", 0, 0, include_semantics=True))
    xf = after["semantics"]["xf"]
    for field, expected in _XF_FLAGS.items():
        assert xf[field] == expected, field


def test_set_cell_style_semantics_flags_survive_save_reload(tmp_path):
    src = tmp_path / "xf2.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "value"
    wb.save(src)
    key = _load_key(M.excel_load(str(src)))

    M.excel_set_cell_style_semantics(key, "Sheet", "A1", xf=dict(_XF_FLAGS))
    M.excel_save(key)

    key2 = _load_key(M.excel_load(str(src)))
    reloaded = json.loads(M.excel_get_cell(key2, "Sheet", 0, 0, include_semantics=True))
    xf = reloaded["semantics"]["xf"]
    for field, expected in _XF_FLAGS.items():
        assert xf.get(field) == expected, field


def test_structured_whole_cell_font_all_subattributes_round_trip(tmp_path):
    out = tmp_path / "whole-cell-font.xlsx"
    created = json.loads(M.excel_create_workbook(sheet_names=["S"], target_path=str(out)))
    key = created["session_key"]
    M.excel_edit_cells(key, "S", [{"row_index": 0, "edits": {0: "Styled"}}])
    M.excel_set_style(key, "S", 0, 0, style={
        "font": {
            "name": "Calibri",
            "size": 14,
            "bold": True,
            "italic": True,
            "underline": "doubleAccounting",
            "strike": True,
            "color": {"type": "theme", "theme": 4, "tint": 0.25},
            "vertAlign": "superscript",
            "charset": 1,
            "family": 2,
            "scheme": "minor",
            "outline": True,
            "shadow": True,
            "condense": True,
            "extend": True,
        },
    })
    M.excel_save(key)

    key2 = _load_key(M.excel_load(str(out)))
    cell = json.loads(M.excel_get_cell(key2, "S", 0, 0, include_semantics=True))
    font = cell["semantics"]["font"]
    expected = {
        "name": "Calibri",
        "size": 14.0,
        "bold": True,
        "italic": True,
        "underline": "doubleAccounting",
        "strike": True,
        "vertAlign": "superscript",
        "charset": 1,
        "family": 2.0,
        "scheme": "minor",
        "outline": True,
        "shadow": True,
        "condense": True,
        "extend": True,
    }
    for field, value in expected.items():
        assert font[field] == value, field
    assert font["color"]["type"] == "theme"
    assert font["color"]["theme"] == 4
    assert font["color"]["tint"] == 0.25
